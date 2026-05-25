"""
ניהול חסימות אנטי-ספאם - חלון דיאלוג
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import os
import re
import shutil
import uuid

try:
    from sound_manager import SoundManager
except Exception:
    SoundManager = None

try:
    import pandas as pd
except Exception:
    pd = None

try:
    from database import Database
except Exception:
    Database = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from excel_styling import apply_rtl_and_alternating_colors
except Exception:
    apply_rtl_and_alternating_colors = None


def open_anti_spam_dialog(parent, load_config_func, save_config_func, db=None):
    """פתיחת חלון ניהול חסימות אנטי-ספאם"""
    
    dialog = tk.Toplevel(parent)
    dialog.title("🛡 ניהול חסימות אנטי-ספאם")
    dialog.geometry("950x650")
    dialog.configure(bg='#ecf0f1')
    dialog.transient(parent)
    dialog.grab_set()
    dialog.resizable(True, True)
    
    # טעינת הגדרות
    config = load_config_func()
    enabled = config.get('anti_spam_enabled', True)
    rules_list = config.get('anti_spam_rules', [
        {'type': 'warning', 'count': 10, 'minutes': 1, 'duration': 0,
         'message': 'שים לב! תיקפת {count} פעמים בדקה האחרונה. אם תמשיך, הכרטיס ייחסם.'},
        {'type': 'warning', 'count': 15, 'minutes': 1, 'duration': 0,
         'message': 'אזהרה! זו התראה השנייה. אם תמשיך, הכרטיס ייחסם.'},
        {'type': 'block', 'count': 20, 'minutes': 1, 'duration': 60,
         'message': 'הכרטיס נחסם לשעה עקב ניצול יתר. תוכל לחזור בעוד {time_left}.'},
        {'type': 'block', 'count': 30, 'minutes': 1, 'duration': 1440,
         'message': 'הכרטיס נחסם ל-24 שעות. תוכל לחזור בעוד {time_left}.'}
    ])

    def _get_sounds_root_dir() -> str:
        try:
            cfg = load_config_func() or {}
        except Exception:
            cfg = {}
        try:
            shared = str((cfg or {}).get('shared_folder') or (cfg or {}).get('network_root') or '').strip()
        except Exception:
            shared = ''
        if shared and os.path.isdir(shared):
            return os.path.join(shared, 'sounds')
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        except Exception:
            base_dir = '.'
        return os.path.join(base_dir, 'sounds')

    def _get_spam_sounds_dir() -> str:
        return os.path.join(_get_sounds_root_dir(), 'לספאם')

    def _get_db():
        try:
            if not Database:
                return None
            # השתמש ב-DB שהועבר או צור חדש
            if db:
                return db
            return Database()
        except Exception:
            return None

    def _list_sound_keys_in_spam_folder() -> list:
        base = _get_spam_sounds_dir()
        if not os.path.isdir(base):
            return []
        sounds = {}
        try:
            for root, _, files in os.walk(base):
                for filename in (files or []):
                    if not str(filename).lower().endswith(('.wav', '.mp3', '.ogg')):
                        continue
                    key = os.path.splitext(filename)[0]
                    if not key:
                        continue
                    path = os.path.join(root, filename)
                    prev = sounds.get(key)
                    if not prev:
                        sounds[key] = path
                        continue
                    try:
                        ext = str(os.path.splitext(path)[1] or '').lower()
                    except Exception:
                        ext = ''
                    try:
                        prev_ext = str(os.path.splitext(prev)[1] or '').lower()
                    except Exception:
                        prev_ext = ''
                    priorities = {'.wav': 30, '.mp3': 20, '.ogg': 10}
                    if int(priorities.get(ext, 0)) > int(priorities.get(prev_ext, 0)):
                        sounds[key] = path
        except Exception:
            sounds = {}
        try:
            return sorted(list(sounds.keys()), key=lambda x: str(x))
        except Exception:
            return list(sounds.keys())

    def _import_sound_file_to_spam_folder() -> str:
        try:
            initial_dir = _get_spam_sounds_dir()
        except Exception:
            initial_dir = None
        fp = filedialog.askopenfilename(
            title="בחר קובץ צליל",
            filetypes=[("Sound files", "*.wav;*.mp3;*.ogg"), ("All files", "*.*")],
            initialdir=initial_dir,
            parent=dialog
        )
        if not fp:
            return ''
        try:
            dst_dir = _get_spam_sounds_dir()
            os.makedirs(dst_dir, exist_ok=True)
        except Exception:
            return ''
        try:
            base_name = os.path.splitext(os.path.basename(fp))[0]
            ext = os.path.splitext(os.path.basename(fp))[1]
        except Exception:
            base_name, ext = 'sound', '.wav'
        try:
            safe_base = re.sub(r'[^0-9A-Za-zא-ת _\-]', '', str(base_name)).strip() or 'sound'
        except Exception:
            safe_base = str(base_name).strip() or 'sound'
        dest = os.path.join(dst_dir, safe_base + ext)
        if os.path.abspath(dest) != os.path.abspath(fp) and os.path.exists(dest):
            dest = os.path.join(dst_dir, f"{safe_base}_{uuid.uuid4().hex[:6]}{ext}")
        try:
            if os.path.abspath(dest) != os.path.abspath(fp):
                shutil.copy2(fp, dest)
        except Exception:
            return ''
        try:
            return os.path.splitext(os.path.basename(dest))[0]
        except Exception:
            return ''

    def _preview_sound_key(sound_key: str) -> None:
        try:
            if not SoundManager:
                return
            k = str(sound_key or '').strip()
            if not k:
                return
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
            except Exception:
                base_dir = '.'
            mgr = SoundManager(base_dir, sounds_dir=_get_sounds_root_dir())
            path = mgr.resolve_sound([k])
            if path:
                mgr.play_sound(path, async_play=True)
        except Exception:
            return
    
    # כותרת
    header = tk.Frame(dialog, bg='#e74c3c', height=60)
    header.pack(fill=tk.X)
    header.pack_propagate(False)
    tk.Label(header, text="🛡 ניהול חסימות אנטי-ספאם לכרטיסים", 
             font=('Arial', 14, 'bold'), bg='#e74c3c', fg='white').pack(pady=15)
    
    # הפעלה/כיבוי
    control_frame = tk.Frame(dialog, bg='#ecf0f1')
    control_frame.pack(fill=tk.X, padx=20, pady=10)
    enabled_var = tk.BooleanVar(value=enabled)
    tk.Checkbutton(control_frame, text="הפעל מערכת אנטי-ספאם", variable=enabled_var,
                   font=('Arial', 11, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(anchor='e', padx=10)
    
    # הסבר
    info_frame = tk.Frame(dialog, bg='#ecf0f1')
    info_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
    tk.Label(info_frame, text="💡 הגדר כללים מדורגים. כל כלל בודק תיקופים בחלון זמן ומבצע פעולה.",
             font=('Arial', 9, 'italic'), bg='#ecf0f1', fg='#7f8c8d', 
             anchor='e', justify='right').pack(fill=tk.X, padx=10)

    # ביטול חסימה ידני
    manual_unblock = tk.Frame(dialog, bg='#ecf0f1')
    manual_unblock.pack(fill=tk.X, padx=20, pady=(0, 8))
    tk.Label(manual_unblock, text="ביטול חסימה ידני (למנהל):", font=('Arial', 10, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(side=tk.RIGHT, padx=6)
    unblock_card_var = tk.StringVar(value='')
    tk.Entry(manual_unblock, textvariable=unblock_card_var, font=('Arial', 10), width=18, justify='center').pack(side=tk.RIGHT, padx=6)

    def _manual_unblock():
        card = str(unblock_card_var.get() or '').strip()
        if not card:
            messagebox.showwarning('אזהרה', 'יש להזין מספר כרטיס', parent=dialog)
            return
        db = _get_db()
        if not db:
            messagebox.showerror('שגיאה', 'מסד נתונים לא זמין', parent=dialog)
            return
        try:
            st = db.get_student_by_card(card)
        except Exception:
            st = None
        if not st:
            messagebox.showerror('שגיאה', 'כרטיס לא נמצא במערכת', parent=dialog)
            return
        try:
            sid = int(st.get('id') or 0)
        except Exception:
            sid = 0
        if not sid:
            messagebox.showerror('שגיאה', 'תלמיד לא תקין', parent=dialog)
            return
        nm = f"{str(st.get('first_name') or '').strip()} {str(st.get('last_name') or '').strip()}".strip()
        if not messagebox.askyesno('אישור', f"לבטל חסימה פעילה עבור {nm}?", parent=dialog):
            return
        try:
            ok = bool(db.unblock_card(sid))
        except Exception:
            ok = False
        if ok:
            messagebox.showinfo('בוצע', 'החסימה בוטלה', parent=dialog)
        else:
            messagebox.showerror('שגיאה', 'לא ניתן לבטל חסימה', parent=dialog)

    tk.Button(manual_unblock, text='🔓 בטל חסימה', command=_manual_unblock, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=12, pady=4).pack(side=tk.RIGHT, padx=6)
    
    # טבלה
    table_frame = tk.Frame(dialog, bg='#ecf0f1')
    table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    columns = ('type', 'count', 'minutes', 'duration', 'sound', 'message')
    tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=12)
    tree.heading('type', text='סוג')
    tree.heading('count', text='תיקופים')
    tree.heading('minutes', text='תוך (דקות)')
    tree.heading('duration', text='חסימה (דקות)')
    tree.heading('sound', text='צליל')
    tree.heading('message', text='הודעה')
    tree.column('type', width=100, anchor='center')
    tree.column('count', width=80, anchor='center')
    tree.column('minutes', width=100, anchor='center')
    tree.column('duration', width=120, anchor='center')
    tree.column('sound', width=80, anchor='center')
    tree.column('message', width=360, anchor='e')
    
    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
    
    def refresh_table():
        for item in tree.get_children():
            tree.delete(item)
        for rule in rules_list:
            rule_type = 'התראה' if rule['type'] == 'warning' else 'חסימה'
            duration_text = f"{rule.get('duration', 0)}" if rule['type'] == 'block' else '-'
            sound_key = str(rule.get('sound_key') or '').strip()
            tree.insert('', 'end', values=(
                rule_type, rule['count'], rule['minutes'], duration_text, sound_key, rule.get('message', '')
            ))
    
    def add_or_edit_rule(rule_type=None, edit_index=None):
        """הוספה או עריכת כלל"""
        edit_dialog = tk.Toplevel(dialog)
        edit_dialog.title("עריכת כלל" if edit_index is not None else "הוספת כלל")
        edit_dialog.geometry("720x520")
        try:
            edit_dialog.minsize(700, 500)
        except Exception:
            pass
        edit_dialog.configure(bg='#ecf0f1')
        edit_dialog.transient(dialog)
        edit_dialog.grab_set()
        
        if edit_index is not None:
            rule = rules_list[edit_index].copy()
        else:
            rule = {'type': rule_type or 'warning', 'count': 10, 'minutes': 1, 'duration': 0, 'message': '', 'sound_key': ''}
        
        tk.Label(edit_dialog, text="סוג כלל:", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(20, 5))
        type_var = tk.StringVar(value=rule['type'])
        type_frame = tk.Frame(edit_dialog, bg='#ecf0f1')
        type_frame.pack(anchor='e', padx=20, pady=5)
        tk.Radiobutton(type_frame, text="התראה", variable=type_var, value='warning', bg='#ecf0f1').pack(side=tk.RIGHT, padx=10)
        tk.Radiobutton(type_frame, text="חסימה", variable=type_var, value='block', bg='#ecf0f1').pack(side=tk.RIGHT, padx=10)
        
        tk.Label(edit_dialog, text="מספר תיקופים:", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(10, 5))
        count_var = tk.IntVar(value=rule.get('count', 10))
        tk.Entry(edit_dialog, textvariable=count_var, font=('Arial', 10), width=10, justify='center').pack(anchor='e', padx=20)
        
        tk.Label(edit_dialog, text="תוך כמה דקות:", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(10, 5))
        minutes_var = tk.IntVar(value=rule.get('minutes', 1))
        tk.Entry(edit_dialog, textvariable=minutes_var, font=('Arial', 10), width=10, justify='center').pack(anchor='e', padx=20)
        
        tk.Label(edit_dialog, text="משך חסימה (דקות, רק לחסימה):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(10, 5))
        duration_var = tk.IntVar(value=rule.get('duration', 0))
        duration_entry = tk.Entry(edit_dialog, textvariable=duration_var, font=('Arial', 10), width=10, justify='center')
        duration_entry.pack(anchor='e', padx=20)

        def _sync_duration_state(*_args):
            try:
                is_block = (str(type_var.get() or '').strip().lower() == 'block')
            except Exception:
                is_block = False
            try:
                duration_entry.config(state='normal' if is_block else 'disabled', disabledforeground='#7f8c8d')
            except Exception:
                pass

        try:
            type_var.trace_add('write', _sync_duration_state)
        except Exception:
            pass
        _sync_duration_state()
        
        tk.Label(edit_dialog, text="הודעה:", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(10, 5))
        message_var = tk.StringVar(value=rule.get('message', ''))
        message_entry = tk.Entry(edit_dialog, textvariable=message_var, font=('Arial', 10), width=50, justify='right')
        message_entry.pack(anchor='e', padx=20)

        token_frame = tk.Frame(edit_dialog, bg='#ecf0f1')
        token_frame.pack(anchor='e', padx=20, pady=(6, 0))
        tk.Label(token_frame, text="הוסף קוד:", font=('Arial', 9, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=(0, 6))

        def _insert_token(token: str):
            try:
                message_entry.insert(tk.INSERT, token)
            except Exception:
                try:
                    message_var.set(str(message_var.get() or '') + token)
                except Exception:
                    pass

        tk.Button(
            token_frame,
            text="מס' תיקופים",
            command=lambda: _insert_token('{count}'),
            font=('Arial', 9),
            bg='#bdc3c7',
            fg='black',
            padx=8,
            pady=2
        ).pack(side=tk.RIGHT, padx=4)
        tk.Button(
            token_frame,
            text="זמן נותר",
            command=lambda: _insert_token('{time_left}'),
            font=('Arial', 9),
            bg='#bdc3c7',
            fg='black',
            padx=8,
            pady=2
        ).pack(side=tk.RIGHT, padx=4)

        tk.Label(edit_dialog, text="צליל (מפתח):", font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(anchor='e', padx=20, pady=(10, 5))
        sound_var = tk.StringVar(value=str(rule.get('sound_key') or '').strip())

        sound_row = tk.Frame(edit_dialog, bg='#ecf0f1')
        sound_row.pack(anchor='e', padx=20)

        sound_combo = ttk.Combobox(
            sound_row,
            textvariable=sound_var,
            values=['ללא צליל'] + list(_list_sound_keys_in_spam_folder() or []),
            state='readonly',
            width=22,
            justify='center'
        )
        sound_combo.pack(side=tk.RIGHT)

        def _refresh_sound_values(_event=None):
            try:
                sound_combo.configure(values=['ללא צליל'] + list(_list_sound_keys_in_spam_folder() or []))
            except Exception:
                pass

        sound_combo.bind('<Button-1>', _refresh_sound_values)

        tk.Button(
            sound_row,
            text='▶',
            command=lambda: _preview_sound_key(str(sound_var.get() or '').strip()),
            font=('Arial', 10, 'bold'),
            bg='#bdc3c7',
            fg='black',
            width=3
        ).pack(side=tk.RIGHT, padx=6)

        tk.Button(
            sound_row,
            text='📁',
            command=lambda: (sound_var.set(_import_sound_file_to_spam_folder() or sound_var.get()), _refresh_sound_values()),
            font=('Arial', 10),
            bg='#bdc3c7',
            fg='black',
            width=3
        ).pack(side=tk.RIGHT, padx=6)
        
        def save_rule():
            if type_var.get() == 'block':
                try:
                    dval = int(duration_var.get() or 0)
                except Exception:
                    dval = 0
                if dval < 1:
                    messagebox.showwarning('אזהרה', 'בחסימה חובה להגדיר משך חסימה (בדקות) גדול מ-0', parent=edit_dialog)
                    return
            else:
                try:
                    duration_var.set(0)
                except Exception:
                    pass
            new_rule = {
                'type': type_var.get(),
                'count': count_var.get(),
                'minutes': minutes_var.get(),
                'duration': duration_var.get() if type_var.get() == 'block' else 0,
                'message': message_var.get(),
                'sound_key': '' if str(sound_var.get() or '').strip() == 'ללא צליל' else str(sound_var.get() or '').strip(),
            }
            if edit_index is not None:
                rules_list[edit_index] = new_rule
            else:
                rules_list.append(new_rule)
            refresh_table()
            edit_dialog.destroy()
        
        btn_frame = tk.Frame(edit_dialog, bg='#ecf0f1')
        btn_frame.pack(pady=20)
        tk.Button(btn_frame, text="💾 שמור", command=save_rule, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=15, pady=6).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✖ ביטול", command=edit_dialog.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=15, pady=6).pack(side=tk.LEFT, padx=5)
    
    def add_warning():
        add_or_edit_rule('warning', None)
    
    def add_block():
        add_or_edit_rule('block', None)
    
    def edit_rule():
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור כלל לעריכה")
            return
        index = tree.index(selection[0])
        add_or_edit_rule(None, index)
    
    def delete_rule():
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור כלל למחיקה")
            return
        if messagebox.askyesno("אישור", "האם למחוק את הכלל הנבחר?"):
            index = tree.index(selection[0])
            rules_list.pop(index)
            refresh_table()
    
    def move_up():
        selection = tree.selection()
        if not selection:
            return
        index = tree.index(selection[0])
        if index > 0:
            rules_list[index], rules_list[index-1] = rules_list[index-1], rules_list[index]
            refresh_table()
            tree.selection_set(tree.get_children()[index-1])
    
    def move_down():
        selection = tree.selection()
        if not selection:
            return
        index = tree.index(selection[0])
        if index < len(rules_list) - 1:
            rules_list[index], rules_list[index+1] = rules_list[index+1], rules_list[index]
            refresh_table()
            tree.selection_set(tree.get_children()[index+1])
    
    def save_all():
        import json as _json
        config = load_config_func()
        config['anti_spam_enabled'] = enabled_var.get()
        config['anti_spam_rules'] = rules_list
        if save_config_func(config):
            # סנכרון לטבלת settings ב-DB (לענן)
            _val = ''
            try:
                if db and hasattr(db, 'set_setting'):
                    _val = _json.dumps({'anti_spam_enabled': bool(enabled_var.get()), 'anti_spam_rules': rules_list}, ensure_ascii=False)
                    db.set_setting('anti_spam_config', _val)
            except Exception:
                pass
            # דחיפה ישירה לענן
            if _val:
                try:
                    import threading as _thr, urllib.request as _ureq
                    _cfg2 = load_config_func() or {}
                    _purl = str(_cfg2.get('sync_push_url') or '').strip()
                    _tid = str(_cfg2.get('sync_tenant_id') or '').strip()
                    _akey = str(_cfg2.get('sync_api_key') or _cfg2.get('api_key') or _cfg2.get('sync_key') or '').strip()
                    if _purl and _tid and _akey:
                        _body = _json.dumps({"tenant_id": _tid, "station_id": "admin-settings",
                            "changes": [{"entity_type": "setting", "entity_id": "anti_spam_config",
                                         "action_type": "update",
                                         "payload_json": _json.dumps({"key": "anti_spam_config", "value": _val})}]
                        }).encode('utf-8')
                        def _do_push():
                            try:
                                _req = _ureq.Request(_purl, data=_body, headers={"Content-Type": "application/json", "api-key": _akey})
                                _ureq.urlopen(_req, timeout=6)
                            except Exception:
                                pass
                        _thr.Thread(target=_do_push, daemon=True).start()
                except Exception:
                    pass
            messagebox.showinfo("הצלחה", "ההגדרות נשמרו בהצלחה")
            dialog.destroy()
        else:
            messagebox.showerror("שגיאה", "לא ניתן לשמור את ההגדרות")

    def open_report_dialog():
        rep = tk.Toplevel(dialog)
        rep.title('📊 דו"ח אנטי-ספאם')
        rep.geometry('1050x620')
        rep.configure(bg='#ecf0f1')
        rep.transient(dialog)
        rep.grab_set()
        rep.resizable(True, True)

        top = tk.Frame(rep, bg='#ecf0f1')
        top.pack(fill=tk.X, padx=12, pady=10)

        tk.Label(top, text='ימים אחרונים:', font=('Arial', 10, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
        days_var = tk.IntVar(value=7)
        days_entry = tk.Entry(top, textvariable=days_var, width=6, justify='center')
        days_entry.pack(side=tk.RIGHT)

        status_var = tk.StringVar(value='')
        tk.Label(rep, textvariable=status_var, bg='#ecf0f1', fg='#2c3e50', anchor='e').pack(fill=tk.X, padx=12)

        table_frame2 = tk.Frame(rep, bg='#ecf0f1')
        table_frame2.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        rep_cols = ('time', 'type', 'class', 'name', 'card', 'message')
        rep_tree = ttk.Treeview(table_frame2, columns=rep_cols, show='headings', height=18)
        rep_tree.heading('time', text='זמן')
        rep_tree.heading('type', text='סוג')
        rep_tree.heading('class', text='כיתה')
        rep_tree.heading('name', text='תלמיד')
        rep_tree.heading('card', text='כרטיס')
        rep_tree.heading('message', text='הודעה')

        rep_tree.column('time', width=150, anchor='center')
        rep_tree.column('type', width=90, anchor='center')
        rep_tree.column('class', width=90, anchor='center')
        rep_tree.column('name', width=180, anchor='e')
        rep_tree.column('card', width=140, anchor='center')
        rep_tree.column('message', width=420, anchor='e')

        rep_scroll = ttk.Scrollbar(table_frame2, orient=tk.VERTICAL, command=rep_tree.yview)
        rep_tree.configure(yscroll=rep_scroll.set)
        rep_tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        rep_scroll.pack(side=tk.LEFT, fill=tk.Y)

        btn_row = tk.Frame(rep, bg='#ecf0f1')
        btn_row.pack(fill=tk.X, padx=12, pady=10)

        def _read_days() -> int:
            try:
                d = int(days_var.get() or 7)
            except Exception:
                d = 7
            if d <= 0:
                d = 7
            return d

        def _fetch_rows() -> list:
            db = _get_db()
            if not db:
                return []
            try:
                return db.get_recent_anti_spam_events_report(days=_read_days(), event_types=['warning', 'block'], limit=5000) or []
            except Exception:
                try:
                    return db.get_recent_card_blocks_report(days=_read_days(), limit=5000) or []
                except Exception:
                    return []

        def refresh_report() -> list:
            for it in rep_tree.get_children():
                rep_tree.delete(it)
            rows = _fetch_rows()
            for r in (rows or []):
                t = str(r.get('created_at') or '').replace('T', ' ')
                et = str(r.get('event_type') or '').strip().lower()
                et_txt = 'חסימה' if et == 'block' else 'אזהרה'
                nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                cls = str(r.get('class_name') or '').strip()
                card = str(r.get('card_number') or '').strip()
                msg = str(r.get('message') or r.get('block_reason') or '').strip()
                rep_tree.insert('', 'end', values=(t, et_txt, cls, nm, card, msg))
            status_var.set(f"נמצאו {len(rows or [])} רשומות")
            return rows

        def export_report_to_excel():
            if not pd:
                messagebox.showerror('שגיאה', 'pandas לא מותקן - לא ניתן לייצא לאקסל', parent=rep)
                return
            rows = refresh_report()
            if not rows:
                messagebox.showwarning('אין נתונים', 'אין רשומות לייצוא', parent=rep)
                return
            data = []
            for r in (rows or []):
                et = str(r.get('event_type') or '').strip().lower()
                et_txt = 'חסימה' if et == 'block' else 'אזהרה'
                nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                msg = str(r.get('message') or r.get('block_reason') or '').strip()
                data.append({
                    'זמן': str(r.get('created_at') or '').replace('T', ' '),
                    'סוג': et_txt,
                    'כיתה': str(r.get('class_name') or '').strip(),
                    'תלמיד': nm,
                    'כרטיס': str(r.get('card_number') or '').strip(),
                    'הודעה': msg,
                })

            try:
                do_save = messagebox.askyesno('דו"ח אנטי-ספאם', 'הדו"ח כבר מוצג על המסך. לשמור גם לקובץ Excel?', parent=rep, default='no')
            except Exception:
                try:
                    do_save = messagebox.askyesno('דו"ח אנטי-ספאם', 'הדו"ח כבר מוצג על המסך. לשמור גם לקובץ Excel?', parent=rep)
                except Exception:
                    do_save = False

            if not do_save:
                return

            fp = filedialog.asksaveasfilename(
                title='שמור דו"ח אנטי-ספאם',
                defaultextension='.xlsx',
                filetypes=[('Excel', '*.xlsx')],
                parent=rep
            )
            if not fp:
                return
            try:
                df = pd.DataFrame(data)
                df.to_excel(fp, index=False, engine='openpyxl')
                if load_workbook and apply_rtl_and_alternating_colors:
                    wb = load_workbook(fp)
                    ws = wb.active
                    apply_rtl_and_alternating_colors(ws, has_header=True)
                    wb.save(fp)
                messagebox.showinfo('הצלחה', 'הדו"ח נשמר בהצלחה', parent=rep)
            except Exception as e:
                messagebox.showerror('שגיאה', f'לא ניתן לייצא לאקסל:\n{e}', parent=rep)

        tk.Button(btn_row, text='🔄 רענן', command=refresh_report, font=('Arial', 10), bg='#3498db', fg='white', padx=12, pady=6).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_row, text='📤 ייצוא לאקסל', command=export_report_to_excel, font=('Arial', 10), bg='#9b59b6', fg='white', padx=12, pady=6).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_row, text='✖ סגור', command=rep.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=12, pady=6).pack(side=tk.LEFT, padx=5)

        refresh_report()
    
    buttons_frame = tk.Frame(dialog, bg='#ecf0f1')
    buttons_frame.pack(fill=tk.X, padx=20, pady=10)
    
    tk.Button(buttons_frame, text="➕ הוסף התראה", command=add_warning, font=('Arial', 10), bg='#3498db', fg='white', padx=15, pady=6).pack(side=tk.RIGHT, padx=5)
    tk.Button(buttons_frame, text="➕ הוסף חסימה", command=add_block, font=('Arial', 10), bg='#e67e22', fg='white', padx=15, pady=6).pack(side=tk.RIGHT, padx=5)
    tk.Button(buttons_frame, text="✏ ערוך", command=edit_rule, font=('Arial', 10), bg='#9b59b6', fg='white', padx=15, pady=6).pack(side=tk.RIGHT, padx=5)
    tk.Button(buttons_frame, text="🗑 מחק", command=delete_rule, font=('Arial', 10), bg='#e74c3c', fg='white', padx=15, pady=6).pack(side=tk.RIGHT, padx=5)
    tk.Button(buttons_frame, text="⬆ העלה", command=move_up, font=('Arial', 10), bg='#95a5a6', fg='white', padx=10, pady=6).pack(side=tk.LEFT, padx=5)
    tk.Button(buttons_frame, text="⬇ הורד", command=move_down, font=('Arial', 10), bg='#95a5a6', fg='white', padx=10, pady=6).pack(side=tk.LEFT, padx=5)
    
    bottom_frame = tk.Frame(dialog, bg='#ecf0f1')
    bottom_frame.pack(fill=tk.X, padx=20, pady=15)
    tk.Button(bottom_frame, text="✖ סגור", command=dialog.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', width=12, padx=4, pady=6).pack(side=tk.LEFT, padx=8)
    tk.Button(bottom_frame, text="💾 שמור", command=save_all, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', width=12, padx=4, pady=6).pack(side=tk.LEFT, padx=8)
    tk.Button(bottom_frame, text='📊 דו"ח/ייצוא', command=open_report_dialog, font=('Arial', 10, 'bold'), bg='#9b59b6', fg='white', width=12, padx=4, pady=6).pack(side=tk.RIGHT, padx=8)
    
    refresh_table()
