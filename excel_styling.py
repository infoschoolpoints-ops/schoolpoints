"""
מודול לעיצוב קבצי Excel עם RTL וצבעים מתחלפים
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def apply_rtl_and_alternating_colors(worksheet: Worksheet, has_header: bool = True):
    """
    מעצב גיליון Excel עם:
    1. כיוון מימין לשמאל (RTL)
    2. צבעים מתחלפים בשורות
    3. עיצוב כותרות (אם יש)
    
    Args:
        worksheet: גיליון העבודה של openpyxl
        has_header: האם השורה הראשונה היא כותרת
    """
    # הגדרת כיוון הגיליון מימין לשמאל
    worksheet.sheet_view.rightToLeft = True
    
    # צבעים לשורות מתחלפות
    color1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # לבן
    color2 = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # אפור בהיר
    
    # צבע כותרת
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # כחול
    header_font = Font(bold=True, color="FFFFFF", size=12)  # לבן מודגש
    
    # יישור לימין לעברית
    alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # גבולות לתאים
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # עיבור על כל השורות
    start_row = 1
    for idx, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
        is_header = (idx == 1 and has_header)
        
        for cell in row:
            # הוספת גבולות
            cell.border = thin_border
            
            if is_header:
                # עיצוב כותרת
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_center
            else:
                # עיצוב שורות רגילות
                # צבע מתחלף (מתחילים מהשורה השנייה אם יש כותרת)
                row_num = idx - 1 if has_header else idx
                cell.fill = color1 if row_num % 2 == 1 else color2
                cell.alignment = alignment_right
    
    # התאמת רוחב עמודות אוטומטית
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # מקסימום 50
        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)  # מינימום 12


def add_class_page_breaks_and_freeze(worksheet: Worksheet, has_header: bool = True,
                                     class_col_name: str = 'כיתה'):
    """
    הוספת מעברי עמוד בין כיתות שונות + הקפאת שורת כותרת + כותרת חוזרת בהדפסה.

    Args:
        worksheet: גיליון העבודה של openpyxl
        has_header: האם השורה הראשונה היא כותרת
        class_col_name: שם עמודת הכיתה בכותרת
    """
    from openpyxl.worksheet.pagebreak import Break

    # הקפאת שורת כותרת (freeze panes)
    if has_header:
        worksheet.freeze_panes = 'A2'
        # כותרת חוזרת בכל עמוד בהדפסה
        worksheet.print_title_rows = '1:1'

    # מציאת עמודת כיתה
    class_col_idx = None
    if has_header:
        for cell in worksheet[1]:
            if str(cell.value or '').strip() == class_col_name:
                class_col_idx = cell.column
                break

    if class_col_idx is None:
        return

    # מעבר על השורות וזיהוי מעברי כיתה
    start_row = 2 if has_header else 1
    prev_class = None
    for row_num in range(start_row, worksheet.max_row + 1):
        cell_value = str(worksheet.cell(row=row_num, column=class_col_idx).value or '').strip()
        if prev_class is not None and cell_value != prev_class and cell_value:
            # מעבר עמוד לפני השורה הזו (אחרי השורה הקודמת)
            worksheet.row_breaks.append(Break(id=row_num - 1))
        if cell_value:
            prev_class = cell_value


def apply_rtl_styling_simple(worksheet: Worksheet):
    """
    מעצב גיליון Excel עם כיוון RTL בלבד (ללא צבעים)
    
    Args:
        worksheet: גיליון העבודה של openpyxl
    """
    worksheet.sheet_view.rightToLeft = True
    
    # יישור לימין לעברית
    alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment_right
