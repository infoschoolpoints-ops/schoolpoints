# -*- coding: utf-8 -*-
"""
Generate a printable HTML product catalog for fair students.
Reads the import template Excel and produces a compact, print-friendly HTML.
"""
import openpyxl
import os
import sys
import html as html_mod

# ---- CONFIG ----
TEMPLATE_PATH = r'C:\מיצד\SchoolPoints\גיבוי רובע\יריד\תבנית ייבוא מוצרים.xlsx'
OUTPUT_PATH = r'C:\מיצד\SchoolPoints\גיבוי רובע\יריד\קטלוג_יריד.html'
LOGO_PATH = r'C:\מיצד\לוגו 2.png'  # will be embedded as base64 if exists
FAIR_TITLE = 'קטלוג מוצרים - יריד נקודות'

# Category display order and icons
CAT_ORDER = {
    'חנות - פרס יקר': 1,
    'חנות - פרס בינוני': 2,
    'חנות - פרס זול': 3,
    'דוכני מזון': 4,
    'דוכני פעילות': 5,
    'מכירה סינית': 6,
    'טרמפולינות': 7,
}
CAT_ICONS = {
    'חנות - פרס יקר': '⭐',
    'חנות - פרס בינוני': '🎁',
    'חנות - פרס זול': '🎈',
    'דוכני מזון': '🍿',
    'דוכני פעילות': '🎯',
    'מכירה סינית': '🎟️',
    'טרמפולינות': '🤸',
}

def read_products(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or '').strip())
    products = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        vals = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                vals[headers[i]] = cell.value
        name = str(vals.get('שם תצוגה') or vals.get('שם פנימי') or '').strip()
        if not name:
            continue
        try:
            price = int(vals.get('מחיר נקודות') or 0)
        except:
            price = 0
        try:
            stock = int(vals.get('כמות מלאי') or 0)
        except:
            stock = 0
        try:
            min_pts = int(vals.get('סף נקודות') or 0)
        except:
            min_pts = 0
        try:
            max_stu = int(vals.get('מקס לתלמיד') or 0)
        except:
            max_stu = 0
        cat = str(vals.get('קטגוריה') or '').strip()
        desc = str(vals.get('תיאור') or '').strip()
        img = str(vals.get('נתיב תמונה') or '').strip()
        products.append({
            'name': name,
            'price': price,
            'stock': stock,
            'min_pts': min_pts,
            'max_stu': max_stu,
            'category': cat,
            'desc': desc,
            'img': img,
        })
    return products

def get_logo_base64(path):
    if not os.path.isfile(path):
        return ''
    import base64
    with open(path, 'rb') as f:
        data = base64.b64encode(f.read()).decode('ascii')
    ext = os.path.splitext(path)[1].lower()
    mime = 'image/png' if ext == '.png' else 'image/jpeg'
    return f'data:{mime};base64,{data}'

def generate_html(products, output_path):
    logo_data = get_logo_base64(LOGO_PATH)

    # Group by category
    cats = {}
    for p in products:
        c = p['category'] or 'ללא קטגוריה'
        cats.setdefault(c, []).append(p)
    sorted_cats = sorted(cats.keys(), key=lambda c: CAT_ORDER.get(c, 99))

    h = html_mod.escape  # shortcut

    css = '''
@page {
    size: A4;
    margin: 6mm;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: Arial, 'Segoe UI', Tahoma, sans-serif;
    direction: rtl;
    background: #fff;
    color: #1a1a2e;
    font-size: 9pt;
    line-height: 1.3;
}
.header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 2px solid #2F5496;
    padding: 4mm 2mm;
    margin-bottom: 3mm;
}
.header-right {
    display: flex;
    flex-direction: column;
    gap: 2px;
}
.header-right .title {
    font-size: 16pt;
    font-weight: bold;
    color: #2F5496;
}
.header-right .subtitle {
    font-size: 10pt;
    color: #555;
}
.student-info {
    display: flex;
    gap: 12mm;
    margin-top: 3px;
}
.student-info .field {
    display: flex;
    align-items: center;
    gap: 2mm;
    font-size: 10pt;
}
.student-info .field label {
    font-weight: bold;
    color: #2F5496;
}
.student-info .field .blank {
    display: inline-block;
    border-bottom: 1.5px solid #333;
    min-width: 25mm;
    height: 5mm;
}
.logo {
    width: 22mm;
    height: auto;
    object-fit: contain;
}
.cat-section {
    margin-bottom: 2mm;
    break-inside: avoid;
}
.cat-title {
    background: #2F5496;
    color: #fff;
    padding: 1.5mm 4mm;
    font-size: 10pt;
    font-weight: bold;
    border-radius: 2px;
    margin-bottom: 1.5mm;
    display: flex;
    align-items: center;
    gap: 2mm;
}
.cat-title .icon { font-size: 11pt; }
.products-grid {
    display: grid;
    grid-template-columns: repeat(6, 1fr);
    gap: 1.5mm;
}
.product-card {
    border: 1px solid #ccc;
    border-radius: 3px;
    padding: 1.5mm;
    text-align: center;
    position: relative;
    background: #fafbfd;
    break-inside: avoid;
    display: flex;
    flex-direction: column;
    min-height: 28mm;
}
.product-card.restricted {
    opacity: 0.35;
    background: #f0f0f0;
}
.product-card.restricted::after {
    content: '🔒';
    position: absolute;
    top: 1mm;
    left: 1mm;
    font-size: 10pt;
}
.product-img-wrap {
    width: 100%;
    height: 16mm;
    display: flex;
    align-items: center;
    justify-content: center;
    background: #eef2f7;
    border-radius: 2px;
    margin-bottom: 1mm;
    overflow: hidden;
}
.product-img-wrap img {
    max-width: 100%;
    max-height: 100%;
    object-fit: contain;
}
.product-img-wrap .placeholder {
    font-size: 16pt;
    color: #bbb;
}
.product-name {
    font-size: 7.5pt;
    font-weight: bold;
    color: #1a1a2e;
    margin-bottom: 1mm;
    flex-grow: 1;
    display: flex;
    align-items: center;
    justify-content: center;
    word-break: break-word;
}
.product-bottom {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 1mm;
}
.product-price {
    background: #2F5496;
    color: #fff;
    font-size: 8pt;
    font-weight: bold;
    padding: 0.5mm 2mm;
    border-radius: 2px;
    white-space: nowrap;
}
.product-qty {
    width: 8mm;
    height: 6mm;
    border: 1.5px solid #2F5496;
    border-radius: 2px;
    text-align: center;
    font-size: 9pt;
    color: #2F5496;
    flex-shrink: 0;
}
.product-info {
    font-size: 6pt;
    color: #888;
    margin-top: 0.5mm;
}
.footer {
    margin-top: 3mm;
    padding-top: 2mm;
    border-top: 1px solid #ccc;
    font-size: 7pt;
    color: #888;
    text-align: center;
}
.legend {
    display: flex;
    gap: 6mm;
    justify-content: center;
    margin-top: 2mm;
    font-size: 7.5pt;
}
.legend span { display: flex; align-items: center; gap: 1mm; }
.legend .dot {
    display: inline-block;
    width: 3mm;
    height: 3mm;
    border-radius: 50%;
}

@media print {
    body { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
    .product-card.restricted { opacity: 0.3; }
    .no-print { display: none !important; }
}
'''

    html_parts = []
    html_parts.append(f'''<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{h(FAIR_TITLE)}</title>
<style>{css}</style>
</head>
<body>

<div class="header">
  <div class="header-right">
    <div class="title">{h(FAIR_TITLE)}</div>
    <div class="subtitle">סמן בריבוע כמה יחידות אתה מעוניין לרכוש. מוצרים נעולים 🔒 דורשים יותר נקודות.</div>
    <div class="student-info">
      <div class="field"><label>שם:</label><span class="blank"></span></div>
      <div class="field"><label>כיתה:</label><span class="blank"></span></div>
      <div class="field"><label>נקודות:</label><span class="blank"></span></div>
    </div>
  </div>
  {'<img class="logo" src="' + logo_data + '" alt="לוגו">' if logo_data else ''}
</div>
''')

    # Category colors for variety
    cat_colors = {
        'חנות - פרס יקר': '#8B6914',
        'חנות - פרס בינוני': '#2F5496',
        'חנות - פרס זול': '#548235',
        'דוכני מזון': '#BF4B00',
        'דוכני פעילות': '#7030A0',
        'מכירה סינית': '#C00000',
        'טרמפולינות': '#00B050',
    }

    for cat_name in sorted_cats:
        prods = cats[cat_name]
        icon = CAT_ICONS.get(cat_name, '📦')
        color = cat_colors.get(cat_name, '#2F5496')

        # Use 6 cols for store items, 4 cols for services/food (larger cards)
        if cat_name in ('דוכני פעילות', 'מכירה סינית', 'טרמפולינות'):
            cols = 4
        elif cat_name == 'דוכני מזון':
            cols = 4
        else:
            cols = 6

        html_parts.append(f'''
<div class="cat-section">
  <div class="cat-title" style="background:{color}">
    <span class="icon">{icon}</span> {h(cat_name)}
    <span style="font-size:7pt;font-weight:normal;margin-right:auto;">({len(prods)} פריטים)</span>
  </div>
  <div class="products-grid" style="grid-template-columns: repeat({cols}, 1fr);">
''')

        for p in prods:
            restricted = ' restricted' if p['min_pts'] > 0 else ''
            info_parts = []
            if p['min_pts'] > 0:
                info_parts.append(f"מ-{p['min_pts']} נק'")
            if p['max_stu'] > 0:
                info_parts.append(f"עד {p['max_stu']}")
            info_str = ' | '.join(info_parts)

            # Image - placeholder for now, will support real images later
            if p['img'] and os.path.isfile(p['img']):
                import base64
                with open(p['img'], 'rb') as f:
                    img_data = base64.b64encode(f.read()).decode('ascii')
                ext = os.path.splitext(p['img'])[1].lower()
                mime = 'image/png' if ext == '.png' else 'image/jpeg'
                img_html = f'<img src="data:{mime};base64,{img_data}" alt="">'
            else:
                img_html = f'<span class="placeholder">{icon}</span>'

            html_parts.append(f'''
    <div class="product-card{restricted}">
      <div class="product-img-wrap">{img_html}</div>
      <div class="product-name">{h(p["name"])}</div>
      <div class="product-bottom">
        <div class="product-price">{p["price"]} נק'</div>
        <input type="text" class="product-qty" maxlength="2" title="כמות">
      </div>
      {'<div class="product-info">' + h(info_str) + '</div>' if info_str else ''}
    </div>
''')

        html_parts.append('  </div>\n</div>\n')

    # Legend and footer
    html_parts.append('''
<div class="legend">
  <span><span class="dot" style="background:#2F5496"></span> מחיר בנקודות</span>
  <span>🔒 דורש סף נקודות מינימלי</span>
  <span>☐ רשום כמות רצויה בריבוע</span>
</div>
<div class="footer">
  זהו קטלוג לתכנון בלבד &mdash; הרכישה בפועל תתבצע ביריד עצמו באמצעות כרטיס התלמיד.
</div>

</body>
</html>
''')

    html_content = ''.join(html_parts)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    return output_path

if __name__ == '__main__':
    products = read_products(TEMPLATE_PATH)
    out = generate_html(products, OUTPUT_PATH)
    print(f'Catalog generated: {out}')
    print(f'Products: {len(products)}')
    # Auto-open in browser
    import webbrowser
    webbrowser.open(out)
