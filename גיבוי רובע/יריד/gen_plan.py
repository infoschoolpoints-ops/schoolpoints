# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import sys, os

wb = openpyxl.Workbook()

# ---- STYLES ----
header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_font = Font(name='Arial', bold=True, size=13, color='1F3864')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
normal_font = Font(name='Arial', size=11)
bold_font = Font(name='Arial', bold=True, size=11)
rtl_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
rtl_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
light_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

def style_header(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = rtl_center
        cell.border = thin_border

def style_row(ws, row, max_col, is_alt=False):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = rtl_center
        cell.border = thin_border
        if is_alt:
            cell.fill = light_fill

def add_section_header(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = sub_header_font
    cell.fill = sub_header_fill
    cell.alignment = rtl_center
    cell.border = thin_border
    for c in range(2, max_col+1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = sub_header_fill

# ========================================
# SHEET 1: Summary
# ========================================
ws1 = wb.active
ws1.title = '\u05e1\u05d9\u05db\u05d5\u05dd \u05db\u05dc\u05dc\u05d9'
ws1.sheet_view.rightToLeft = True

r = 1
add_section_header(ws1, r, '\u05e0\u05ea\u05d5\u05e0\u05d9 \u05d1\u05e1\u05d9\u05e1', 3); r += 1
data_rows = [
    ['\u05e1\u05d4"\u05db \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', 234, ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd \u05e8\u05dc\u05d5\u05d5\u05e0\u05d8\u05d9\u05d9\u05dd', 227, '\u05dc\u05dc\u05d0 \u05e7\u05d9\u05e6\u05d5\u05e0\u05d9\u05d9\u05dd \u05de\u05ea\u05d7\u05ea 1,200'],
    ['\u05e1\u05d4"\u05db \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '381,342', ''],
    ['\u05de\u05de\u05d5\u05e6\u05e2 \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', 1680, ''],
    ['\u05de\u05d9\u05e0\u05d9\u05de\u05d5\u05dd \u05e8\u05dc\u05d5\u05d5\u05e0\u05d8\u05d9', 1216, ''],
    ['\u05de\u05e7\u05e1\u05d9\u05de\u05d5\u05dd', 1880, ''],
    ['\u05d7\u05e6\u05d9\u05d5\u05df', 1690, ''],
]
for row_data in data_rows:
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, 3, (r % 2 == 0)); r += 1

r += 1
add_section_header(ws1, r, '\u05d6\u05de\u05df \u05d4\u05d9\u05e8\u05d9\u05d3', 3); r += 1
time_rows = [
    ['\u05d4\u05ea\u05d7\u05dc\u05d4', '12:00', ''],
    ['\u05e1\u05d9\u05d5\u05dd', '16:00', ''],
    ['\u05de\u05e9\u05da', '4 \u05e9\u05e2\u05d5\u05ea = 240 \u05d3\u05e7\u05d5\u05ea', ''],
]
for row_data in time_rows:
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, 3, (r % 2 == 0)); r += 1

r += 1
add_section_header(ws1, r, '\u05d7\u05dc\u05d5\u05e7\u05ea \u05ea\u05e7\u05e6\u05d9\u05d1 \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', 3); r += 1
headers_b = ['\u05ea\u05d7\u05d5\u05dd', '\u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05de\u05d5\u05e6\u05e2 (1,680)', '\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05dc\u05e9 (1,216)']
for c, h in enumerate(headers_b, 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 3); r += 1

budget_rows = [
    ['\u05d7\u05e0\u05d5\u05ea \u05de\u05d5\u05e6\u05e8\u05d9\u05dd (~30%)', 500, 365],
    ['\u05d3\u05d5\u05db\u05e0\u05d9 \u05de\u05d6\u05d5\u05df (~8%)', 120, 120],
    ['\u05de\u05db\u05d9\u05e8\u05d4 \u05e1\u05d9\u05e0\u05d9\u05ea (~15%)', 250, 180],
    ['\u05d3\u05d5\u05db\u05e0\u05d9 \u05e4\u05e2\u05d9\u05dc\u05d5\u05ea (~20%)', 335, 245],
    ['\u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea (~15%)', 250, 180],
    ['\u05e8\u05d6\u05e8\u05d1\u05d4 (~12%)', 225, 126],
    ['\u05e1\u05d4"\u05db', 1680, 1216],
]
for row_data in budget_rows:
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, 3, (r % 2 == 0))
    if row_data[0] == '\u05e1\u05d4"\u05db':
        for c in range(1, 4):
            ws1.cell(row=r, column=c).font = bold_font
            ws1.cell(row=r, column=c).fill = green_fill
    r += 1

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 28

# ========================================
# SHEET 2: Pricing
# ========================================
ws2 = wb.create_sheet('\u05ea\u05de\u05d7\u05d5\u05e8 \u05de\u05e4\u05d5\u05e8\u05d8')
ws2.sheet_view.rightToLeft = True

r = 1
# -- Store --
add_section_header(ws2, r, '\u05d7\u05e0\u05d5\u05ea \u05de\u05d5\u05e6\u05e8\u05d9\u05dd \u2013 \u05de\u05d7\u05d9\u05e8\u05d5\u05df', 4); r += 1
for c, h in enumerate(['\u05e8\u05de\u05d4', '\u05d8\u05d5\u05d5\u05d7 \u05de\u05d7\u05d9\u05e8\u05d9\u05dd (\u05e0\u05e7\u05d5\u05d3\u05d5\u05ea)', '\u05d3\u05d5\u05d2\u05de\u05d0\u05d5\u05ea', '\u05e2\u05dc\u05d5\u05ea \u05e9"\u05d7'], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 4); r += 1

store_items = [
    ['\u05e4\u05e8\u05e1 \u05d9\u05e7\u05e8', '200\u2013350', '\u05d8\u05d9\u05e9\u05d8\u05e8\u05de\u05d9\u05d9\u05dc, \u05e9\u05e2\u05d5\u05df \u05de\u05d7\u05e9\u05d1\u05d5\u05df, \u05d7\u05dc\u05d9\u05e4\u05ea \u05d4\u05e6\u05dc\u05d4, \u05d7\u05d5\u05ea\u05de\u05d5\u05ea', '8\u201330\u20aa'],
    ['\u05e4\u05e8\u05e1 \u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', '80\u2013180', '\u05d8\u05d8\u05e8\u05d9\u05e1, \u05d3\u05d5\u05de\u05d9\u05e0\u05d5, \u05e7\u05e1\u05dd, \u05db\u05d5\u05d1\u05e2 \u05de\u05e8\u05d5\u05e7\u05d0\u05d9, \u05e9\u05dc\u05d2, \u05d6\u05e7\u05df, \u05d6\u05d9\u05e7\u05d5\u05e7\u05d9\u05dd, \u05d7\u05d5\u05d8\u05d9\u05dd', '4\u20136\u20aa'],
    ['\u05e4\u05e8\u05e1 \u05d6\u05d5\u05dc', '20\u201360', '\u05e6\u05d9\u05e0\u05d5\u05e8 \u05e9\u05e8\u05e9\u05d5\u05e8, \u05e4\u05e8\u05d5\u05e4\u05dc\u05d5\u05e8, \u05d8\u05d9\u05e1\u05df, \u05e6\u05e4\u05e6\u05e4\u05d4, \u05e2\u05d8 \u05de\u05e8\u05d2\u05dc\u05d9\u05dd, \u05de\u05e1\u05db\u05d5\u05ea, \u05de\u05d8\u05d5\u05e1, \u05e4\u05d8\u05d9\u05e9\u05d5\u05e0\u05d9\u05dd', '0.8\u20133\u20aa'],
]
for i, item in enumerate(store_items):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, i % 2 == 1); r += 1

r += 1
# -- Food --
add_section_header(ws2, r, '\u05d3\u05d5\u05db\u05e0\u05d9 \u05de\u05d6\u05d5\u05df \u2013 \u05de\u05d7\u05d9\u05e8\u05d5\u05df', 4); r += 1
for c, h in enumerate(['\u05d3\u05d5\u05db\u05df', '\u05de\u05d7\u05d9\u05e8 (\u05e0\u05e7\u05d5\u05d3\u05d5\u05ea)', '\u05d4\u05e2\u05e8\u05d5\u05ea', ''], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 4); r += 1

food = [
    ['\u05d0\u05d9\u05d9\u05e1 \u05e7\u05e4\u05d4', 30, '\u05de\u05e0\u05d4 \u05d0\u05d7\u05ea \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', ''],
    ['\u05e4\u05d5\u05e4\u05e7\u05d5\u05e8\u05df', 30, '\u05de\u05e0\u05d4 \u05d0\u05d7\u05ea \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', ''],
    ['\u05dc\u05d7\u05de\u05e0\u05d9\u05d4 + \u05e0\u05e7\u05e0\u05d9\u05e7', 35, '\u05de\u05e0\u05d4 \u05d0\u05d7\u05ea \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', ''],
    ['\u05d9\u05de\u05d1\u05de\u05d1\u05dd / \u05e9\u05e2\u05e8\u05d5\u05ea \u05e1\u05d1\u05ea\u05d0', 25, '\u05de\u05e0\u05d4 \u05d0\u05d7\u05ea \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', ''],
    ['\u05e1\u05d4"\u05db \u05de\u05d6\u05d5\u05df \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', 120, '', ''],
]
for i, item in enumerate(food):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, i % 2 == 1)
    if i == len(food) - 1:
        for c in range(1, 5):
            ws2.cell(row=r, column=c).font = bold_font
            ws2.cell(row=r, column=c).fill = green_fill
    r += 1

r += 1
# -- Chinese sale --
add_section_header(ws2, r, '\u05de\u05db\u05d9\u05e8\u05d4 \u05e1\u05d9\u05e0\u05d9\u05ea \u2013 \u05d4\u05d2\u05e8\u05dc\u05d4', 4); r += 1
for c, h in enumerate(['\u05e4\u05e8\u05d9\u05d8', '\u05e2\u05e8\u05da', '\u05d4\u05e2\u05e8\u05d5\u05ea', ''], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 4); r += 1

chinese = [
    ['\u05de\u05d7\u05d9\u05e8 \u05db\u05e8\u05d8\u05d9\u05e1 \u05d4\u05d2\u05e8\u05dc\u05d4', '25 \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '', ''],
    ['\u05de\u05e7\u05e1\u05d9\u05de\u05d5\u05dd \u05de\u05d5\u05de\u05dc\u05e5', '~15% \u05de\u05d4\u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '', ''],
    ['', '', '', ''],
    ["\u05e4\u05e8\u05e1 1: \u05de\u05e7\u05e8\u05e0\u05e6'\u05d9\u05e7 + 2 \u05e7\u05dc\u05d8\u05d5\u05ea", '409\u20aa', '\u05e4\u05e8\u05e1 \u05e8\u05d0\u05e9\u05d9', ''],
    ['\u05e4\u05e8\u05e1 2: \u05e7\u05d5\u05e8\u05e7\u05d9\u05e0\u05d8', '150\u20aa', '', ''],
    ['\u05e4\u05e8\u05e1 3: \u05d0\u05d5\u05e7\u05d9 \u05d8\u05d5\u05e7\u05d9', '135\u20aa', '', ''],
    ['\u05e4\u05e8\u05e1 4: \u05de\u05d2\u05e4\u05d5\u05df', '59\u20aa', '', ''],
    ['\u05e4\u05e8\u05e1\u05d9\u05dd 5\u201310', '\u05dc\u05e7\u05d1\u05d5\u05e2', '6 \u05e4\u05e8\u05e1\u05d9\u05dd \u05e0\u05d5\u05e1\u05e4\u05d9\u05dd', ''],
    ['', '', '', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05dc\u05e9 (1,216)', '~7 \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd', '180 \u05e0\u05e7', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05de\u05d5\u05e6\u05e2 (1,680)', '~10 \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd', '250 \u05e0\u05e7', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05d6\u05e7 (1,880)', '~11 \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd', '280 \u05e0\u05e7', ''],
    ['\u05e1\u05d4"\u05db \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd \u05de\u05e9\u05d5\u05e2\u05e8', '~2,100', '227 x ~9.3', ''],
]
for i, item in enumerate(chinese):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, i % 2 == 1); r += 1

r += 1
# -- Activities --
add_section_header(ws2, r, '\u05d3\u05d5\u05db\u05e0\u05d9 \u05e4\u05e2\u05d9\u05dc\u05d5\u05ea - 10 \u05d3\u05d5\u05db\u05e0\u05d9\u05dd', 4); r += 1
for c, h in enumerate(['\u05e4\u05e8\u05d9\u05d8', '\u05e2\u05e8\u05da', '\u05d4\u05e2\u05e8\u05d5\u05ea', ''], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 4); r += 1
activities = [
    ['\u05de\u05d7\u05d9\u05e8 \u05db\u05e0\u05d9\u05e1\u05d4 \u05dc\u05d3\u05d5\u05db\u05df', '25 \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05dc\u05e9 - \u05d3\u05d5\u05db\u05e0\u05d9\u05dd', '~10', '245 \u05e0\u05e7 = \u05db\u05dc 10', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05de\u05d5\u05e6\u05e2 - \u05d3\u05d5\u05db\u05e0\u05d9\u05dd', '~13', '335 \u05e0\u05e7 = \u05db\u05d5\u05dc\u05dd + \u05d7\u05d6\u05e8\u05d5\u05ea', ''],
]
for i, item in enumerate(activities):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, i % 2 == 1); r += 1

r += 1
# -- Trampolines --
add_section_header(ws2, r, '\u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea - 2 \u05de\u05ea\u05e7\u05e0\u05d9\u05dd', 4); r += 1
for c, h in enumerate(['\u05e4\u05e8\u05d9\u05d8', '\u05e2\u05e8\u05da', '\u05d4\u05e2\u05e8\u05d5\u05ea', ''], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 4); r += 1
tramp = [
    ["\u05de\u05d7\u05d9\u05e8 \u05e1\u05d9\u05d1\u05d5\u05d1 (10 \u05d3\u05e7')", '50 \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '', ''],
    ['\u05e7\u05d9\u05d1\u05d5\u05dc\u05ea \u05d1\u05d5-\u05d6\u05de\u05e0\u05d9\u05ea', '10 \u05d9\u05dc\u05d3\u05d9\u05dd x 2 \u05de\u05ea\u05e7\u05e0\u05d9\u05dd = 20', '', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05e9\u05e2\u05d4', '6 x 20 = 120 \u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05e9\u05e2\u05d4', '', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05dc\u05e9 - \u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd', '3\u20134', '180 \u05e0\u05e7', ''],
    ['\u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05de\u05d5\u05e6\u05e2 - \u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd', '5', '250 \u05e0\u05e7', ''],
]
for i, item in enumerate(tramp):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, i % 2 == 1); r += 1

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 15

# ========================================
# SHEET 3: Verification
# ========================================
ws3 = wb.create_sheet('\u05d1\u05d3\u05d9\u05e7\u05ea \u05e7\u05d9\u05e6\u05d5\u05df')
ws3.sheet_view.rightToLeft = True

headers6 = ['\u05e4\u05e2\u05d9\u05dc\u05d5\u05ea', '\u05de\u05d7\u05d9\u05e8 \u05dc\u05d9\u05d7\u05d9\u05d3\u05d4',
            '\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05dc\u05e9 (1,216) \u05e0\u05e7', '\u05db\u05de\u05d5\u05ea',
            '\u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05de\u05d5\u05e6\u05e2 (1,680) \u05e0\u05e7', '\u05db\u05de\u05d5\u05ea',
            '\u05ea\u05dc\u05de\u05d9\u05d3 \u05d7\u05d6\u05e7 (1,880) \u05e0\u05e7', '\u05db\u05de\u05d5\u05ea']
r = 1
for c, h in enumerate(headers6, 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 8); r += 1

verification = [
    ['\u05de\u05d5\u05e6\u05e8 \u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 120, 120, 1, 120, 1, 0, 0],
    ['\u05de\u05d5\u05e6\u05e8 \u05d9\u05e7\u05e8', 300, 0, 0, 300, 1, 300, 1],
    ['\u05de\u05d5\u05e6\u05e8\u05d9\u05dd \u05d6\u05d5\u05dc\u05d9\u05dd', 35, 105, 3, 140, 4, 175, 5],
    ['\u05de\u05d5\u05e6\u05e8\u05d9\u05dd \u05d1\u05d9\u05e0\u05d5\u05e0\u05d9\u05d9\u05dd \u05e0\u05d5\u05e1\u05e4\u05d9\u05dd', 120, 0, 0, 0, 0, 120, 1],
    ['\u05e1\u05d4"\u05db \u05d7\u05e0\u05d5\u05ea', '', 225, '4 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd', 560, '6 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd', 595, '7 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd'],
    ['', '', '', '', '', '', '', ''],
    ['\u05d0\u05d9\u05d9\u05e1 \u05e7\u05e4\u05d4', 30, 30, 1, 30, 1, 30, 1],
    ['\u05e4\u05d5\u05e4\u05e7\u05d5\u05e8\u05df', 30, 30, 1, 30, 1, 30, 1],
    ['\u05dc\u05d7\u05de\u05e0\u05d9\u05d4 + \u05e0\u05e7\u05e0\u05d9\u05e7', 35, 35, 1, 35, 1, 35, 1],
    ['\u05d9\u05de\u05d1\u05de\u05d1\u05dd', 25, 25, 1, 25, 1, 25, 1],
    ['\u05e1\u05d4"\u05db \u05de\u05d6\u05d5\u05df', '', 120, '4 \u05de\u05e0\u05d5\u05ea', 120, '4 \u05de\u05e0\u05d5\u05ea', 120, '4 \u05de\u05e0\u05d5\u05ea'],
    ['', '', '', '', '', '', '', ''],
    ['\u05db\u05e8\u05d8\u05d9\u05e1\u05d9 \u05d4\u05d2\u05e8\u05dc\u05d4', 25, 175, 7, 250, 10, 275, 11],
    ['', '', '', '', '', '', '', ''],
    ['\u05d3\u05d5\u05db\u05e0\u05d9 \u05e4\u05e2\u05d9\u05dc\u05d5\u05ea', 25, 250, 10, 250, 10, 250, 10],
    ['', '', '', '', '', '', '', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1 \u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d4', 50, 150, 3, 250, 5, 250, 5],
    ['', '', '', '', '', '', '', ''],
    ['\u05e1\u05d4"\u05db \u05e0\u05e7\u05d5\u05d3\u05d5\u05ea', '', 920, '', 1450, '', 1490, ''],
    ['\u05e0\u05e9\u05d0\u05e8 \u05dc\u05ea\u05dc\u05de\u05d9\u05d3', '', 296, '', 230, '', 390, ''],
    ['\u05e2\u05d5\u05d3 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd/\u05e4\u05e2\u05d9\u05dc\u05d5\u05d9\u05d5\u05ea', '', '\u05db\u05df!', '', '\u05db\u05df!', '', '\u05db\u05df!', ''],
]

for i, item in enumerate(verification):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, 8, i % 2 == 1)
    name = str(item[0])
    if '\u05e1\u05d4"\u05db' in name:
        for c in range(1, 9):
            ws3.cell(row=r, column=c).font = bold_font
            ws3.cell(row=r, column=c).fill = green_fill
    r += 1

for col in ['A','B','C','D','E','F','G','H']:
    ws3.column_dimensions[col].width = 22

# ========================================
# SHEET 4: Schedule
# ========================================
ws4 = wb.create_sheet('\u05dc\u05d5\u05d7 \u05d6\u05de\u05e0\u05d9\u05dd')
ws4.sheet_view.rightToLeft = True

r = 1
add_section_header(ws4, r, '\u05dc\u05d5\u05d7 \u05d6\u05de\u05e0\u05d9\u05dd \u05de\u05d5\u05de\u05dc\u05e5', 4); r += 1
for c, h in enumerate(['\u05e9\u05e2\u05d4', '\u05d0\u05d9\u05e8\u05d5\u05e2', '\u05e4\u05e8\u05d8\u05d9\u05dd', '\u05d4\u05e2\u05e8\u05d5\u05ea'], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 4); r += 1

schedule = [
    ['12:00\u201312:15', '\u05e4\u05ea\u05d9\u05d7\u05d4 + \u05d4\u05e1\u05d1\u05e8', '\u05db\u05d9\u05e0\u05d5\u05e1 \u05db\u05dc\u05dc\u05d9, \u05d7\u05dc\u05d5\u05e7\u05ea \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd/\u05d4\u05e1\u05d1\u05e8', '15 \u05d3\u05e7'],
    ['12:15\u201314:00', "\u05d9\u05e8\u05d9\u05d3 \u05e4\u05ea\u05d5\u05d7 \u2013 \u05de\u05d7\u05e6\u05d9\u05ea \u05d0'", "\u05db\u05d9\u05ea\u05d5\u05ea \u05d0'-\u05d3' \u05d1\u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea (132 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd)", '\u05db\u05d5\u05dc\u05dd \u05d1\u05e9\u05d0\u05e8 \u05d4\u05d3\u05d5\u05db\u05e0\u05d9\u05dd'],
    ['14:00\u201315:15', "\u05d9\u05e8\u05d9\u05d3 \u05e4\u05ea\u05d5\u05d7 \u2013 \u05de\u05d7\u05e6\u05d9\u05ea \u05d1'", "\u05db\u05d9\u05ea\u05d5\u05ea \u05d4'-\u05d7' \u05d1\u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea (103 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd)", '\u05db\u05d5\u05dc\u05dd \u05d1\u05e9\u05d0\u05e8 \u05d4\u05d3\u05d5\u05db\u05e0\u05d9\u05dd'],
    ['15:15\u201315:30', '\u05e1\u05d2\u05d9\u05e8\u05ea \u05e8\u05db\u05d9\u05e9\u05ea \u05db\u05e8\u05d8\u05d9\u05e1\u05d9\u05dd', '\u05d4\u05d6\u05d3\u05de\u05e0\u05d5\u05ea \u05d0\u05d7\u05e8\u05d5\u05e0\u05d4 \u05dc\u05e7\u05e0\u05d5\u05ea \u05db\u05e8\u05d8\u05d9\u05e1\u05d9 \u05d4\u05d2\u05e8\u05dc\u05d4', ''],
    ['15:30\u201316:00', '\u05d4\u05d2\u05e8\u05dc\u05ea \u05de\u05db\u05d9\u05e8\u05d4 \u05e1\u05d9\u05e0\u05d9\u05ea', '10 \u05e4\u05e8\u05e1\u05d9\u05dd, \u05d4\u05db\u05e8\u05d6\u05d4 \u05d7\u05d2\u05d9\u05d2\u05d9\u05ea', '30 \u05d3\u05e7'],
]

for i, item in enumerate(schedule):
    for c, val in enumerate(item, 1):
        ws4.cell(row=r, column=c, value=val)
    style_row(ws4, r, 4, i % 2 == 1); r += 1

r += 2
add_section_header(ws4, r, '\u05e7\u05d9\u05d1\u05d5\u05dc\u05ea \u05e2\u05de\u05d3\u05d5\u05ea \u05d5\u05ea\u05d5\u05e8\u05d9\u05dd', 4); r += 1
for c, h in enumerate(['\u05e2\u05de\u05d3\u05d4', '\u05e7\u05d9\u05d1\u05d5\u05dc\u05ea/\u05e9\u05e2\u05d4', '\u05d6\u05de\u05df \u05ea\u05d5\u05e8 \u05de\u05de\u05d5\u05e6\u05e2', '\u05e6\u05d5\u05d5\u05d0\u05e8 \u05d1\u05e7\u05d1\u05d5\u05e7?'], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 4); r += 1

capacity = [
    ['\u05d7\u05e0\u05d5\u05ea \u05de\u05d5\u05e6\u05e8\u05d9\u05dd (1)', '~70 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '~5 \u05d3\u05e7 \u05db\u05d5\u05dc\u05dc \u05ea\u05d5\u05e8', '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9 \u2013 \u05e9\u05e7\u05d5\u05dc 2 \u05e7\u05d5\u05e4\u05d5\u05ea'],
    ['\u05d3\u05d5\u05db\u05df \u05de\u05d6\u05d5\u05df (x4)', '~300 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '~5 \u05d3\u05e7 \u05db\u05d5\u05dc\u05dc \u05ea\u05d5\u05e8', '\u05dc\u05d0'],
    ["\u05de\u05db'\u05e1\u05d9\u05e0\u05d9\u05ea \u2013 \u05e8\u05db\u05d9\u05e9\u05d4", '~120 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '~3 \u05d3\u05e7 \u05db\u05d5\u05dc\u05dc \u05ea\u05d5\u05e8', '\u05dc\u05d0'],
    ['\u05d3\u05d5\u05db\u05e0\u05d9 \u05e4\u05e2\u05d9\u05dc\u05d5\u05ea (x10)', '~450 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '~4 \u05d3\u05e7 \u05db\u05d5\u05dc\u05dc \u05ea\u05d5\u05e8', '\u05dc\u05d0'],
    ['\u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea (x2)', '120 \u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05e9\u05e2\u05d4', '~10 \u05d3\u05e7 \u05ea\u05d5\u05e8 + 10 \u05e1\u05d9\u05d1\u05d5\u05d1', '\u05db\u05df \u2013 \u05dc\u05db\u05df \u05de\u05d7\u05d5\u05dc\u05e7 \u05dc\u05de\u05d7\u05e6\u05d9\u05d5\u05ea'],
]
for i, item in enumerate(capacity):
    for c, val in enumerate(item, 1):
        ws4.cell(row=r, column=c, value=val)
    style_row(ws4, r, 4, i % 2 == 1)
    if '\u05db\u05df' in str(item[3]):
        ws4.cell(row=r, column=4).fill = yellow_fill
    r += 1

r += 2
add_section_header(ws4, r, '\u05d7\u05d9\u05e9\u05d5\u05d1 \u05d8\u05e8\u05de\u05e4\u05d5\u05dc\u05d9\u05e0\u05d5\u05ea \u05de\u05e4\u05d5\u05e8\u05d8', 4); r += 1
for c, h in enumerate(['\u05e0\u05ea\u05d5\u05df', '\u05e2\u05e8\u05da', '\u05d7\u05d9\u05e9\u05d5\u05d1', ''], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 4); r += 1

tramp_calc = [
    ['\u05de\u05ea\u05e7\u05e0\u05d9\u05dd', '2', '', ''],
    ['\u05d9\u05dc\u05d3\u05d9\u05dd \u05d1\u05d5-\u05d6\u05de\u05e0\u05d9\u05ea \u05dc\u05de\u05ea\u05e7\u05df', '10', '', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1', '10 \u05d3\u05e7\u05d5\u05ea', '', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05e9\u05e2\u05d4/\u05de\u05ea\u05e7\u05df', '6', '', ''],
    ["\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05e9\u05e2\u05d4 \u05e1\u05d4\"\u05db", '120', '2 x 10 x 6', ''],
    ['', '', '', ''],
    ["\u05de\u05d7\u05e6\u05d9\u05ea \u05d0' (\u05db\u05d9\u05ea\u05d5\u05ea \u05d0'-\u05d3')", '132 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '105 \u05d3\u05e7\u05d5\u05ea (12:15\u201314:00)', ''],
    ["\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd \u05d1\u05de\u05d7\u05e6\u05d9\u05ea \u05d0'", '210', '120 x 1.75 \u05e9\u05e2\u05d5\u05ea', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05ea\u05dc\u05de\u05d9\u05d3', '1.6', '\u05db\u05dc \u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05d5\u05d1\u05d8\u05d7 \u05e1\u05d9\u05d1\u05d5\u05d1 1+', ''],
    ['', '', '', ''],
    ["\u05de\u05d7\u05e6\u05d9\u05ea \u05d1' (\u05db\u05d9\u05ea\u05d5\u05ea \u05d4'-\u05d7')", '103 \u05ea\u05dc\u05de\u05d9\u05d3\u05d9\u05dd', '75 \u05d3\u05e7\u05d5\u05ea (14:00\u201315:15)', ''],
    ["\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd \u05d1\u05de\u05d7\u05e6\u05d9\u05ea \u05d1'", '150', '120 x 1.25 \u05e9\u05e2\u05d5\u05ea', ''],
    ['\u05e1\u05d9\u05d1\u05d5\u05d1\u05d9\u05dd/\u05ea\u05dc\u05de\u05d9\u05d3', '1.5', '\u05db\u05dc \u05ea\u05dc\u05de\u05d9\u05d3 \u05de\u05d5\u05d1\u05d8\u05d7 \u05e1\u05d9\u05d1\u05d5\u05d1 1+', ''],
]
for i, item in enumerate(tramp_calc):
    for c, val in enumerate(item, 1):
        ws4.cell(row=r, column=c, value=val)
    style_row(ws4, r, 4, i % 2 == 1); r += 1

ws4.column_dimensions['A'].width = 38
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 30

# ========================================
# SHEET 5: Product inventory
# ========================================
ws5 = wb.create_sheet('\u05de\u05dc\u05d0\u05d9 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd')
ws5.sheet_view.rightToLeft = True

r = 1
add_section_header(ws5, r, '\u05e4\u05d9\u05e8\u05d5\u05d8 \u05de\u05d5\u05e6\u05e8\u05d9\u05dd \u2013 \u05d7\u05e0\u05d5\u05ea \u05d4\u05d9\u05e8\u05d9\u05d3', 6); r += 1
for c, h in enumerate(['\u05de\u05d5\u05e6\u05e8', '\u05d9\u05d7\u05d9\u05d3\u05d5\u05ea \u05d1\u05de\u05dc\u05d0\u05d9', '\u05e2\u05dc\u05d5\u05ea \u05dc\u05d9\u05d7\u05d9\u05d3\u05d4 \u20aa', '\u05e1\u05d4"\u05db \u05e2\u05dc\u05d5\u05ea \u20aa', '\u05e8\u05de\u05ea \u05de\u05d7\u05d9\u05e8', '\u05de\u05d7\u05d9\u05e8 \u05de\u05d5\u05de\u05dc\u05e5 (\u05e0\u05e7\u05d5\u05d3\u05d5\u05ea)'], 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, 6); r += 1

products_raw = [
    ('\u05d8\u05d9\u05e9\u05d8\u05e8\u05de\u05d9\u05d9\u05dc \u05e9\u05d7\u05d5\u05e8', 5, 30, '\u05d9\u05e7\u05e8', 350),
    ('\u05e9\u05e2\u05d5\u05df \u05de\u05d7\u05e9\u05d1\u05d5\u05df', 48, 8, '\u05d9\u05e7\u05e8', 200),
    ('\u05e0\u05d7\u05e9 \u05e6\u05e4\u05e6\u05e4\u05d4 \u05e4\u05e9\u05d5\u05d8 \u05e7\u05d8\u05df', 1, 23.7, '\u05d9\u05e7\u05e8', 280),
    ('\u05d7\u05d5\u05ea\u05de\u05d5\u05ea', 4, 20, '\u05d9\u05e7\u05e8', 250),
    ('\u05d7\u05dc\u05d9\u05e4\u05ea \u05d4\u05e6\u05dc\u05d4 \u05e2\u05dd \u05db\u05d5\u05d1\u05e2', 3, 15, '\u05d9\u05e7\u05e8', 200),
    ('\u05db\u05d3\u05d5\u05e8 \u05de\u05d1\u05d5\u05da \u05d2\u05d3\u05d5\u05dc', 12, 12, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 160),
    ('\u05db\u05d5\u05d1\u05e2 \u05db\u05d1\u05d0\u05d9', 3, 12, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 160),
    ('\u05de\u05e7\u05dc \u05e1\u05d1\u05d0', 12, 10, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 140),
    ('\u05de\u05e1\u05db\u05d4 \u05d2\u05d5\u05de\u05d9', 10, 10, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 140),
    ('\u05e7\u05e1\u05dd', 6, 7, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 100),
    ('\u05db\u05d5\u05d1\u05e2 \u05e9\u05e2\u05e8\u05d5\u05ea \u05e6\u05d1\u05e2\u05d5\u05e0\u05d9', 8, 6, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 100),
    ('\u05d8\u05d8\u05e8\u05d9\u05e1', 20, 6, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 100),
    ('\u05ea\u05d7\u05e4\u05d5\u05e9\u05ea \u05e0\u05de\u05e8 3 \u05d7\u05dc\u05e7\u05d9\u05dd', 3, 6, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 100),
    ('\u05dc\u05d2\u05d5 \u05d0\u05d5\u05d8\u05d5 \u05de\u05e9\u05d8\u05e8\u05d4', 6, 6, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 100),
    ('\u05db\u05d5\u05d1\u05e2 \u05de\u05e8\u05d5\u05e7\u05d0\u05d9', 10, 5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05d3\u05d5\u05de\u05d9\u05e0\u05d5', 20, 5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05e6\u05e4\u05e6\u05e4\u05d4 (\u05d1\u05d9\u05e7\u05d5\u05e9!)', 30, 5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05e9\u05dc\u05d2', 96, 5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05d7\u05d5\u05d8\u05d9\u05dd', 24, 5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05d6\u05e7\u05df \u05e9\u05d7\u05d5\u05e8', 20, 4.9, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05d6\u05d9\u05e7\u05d5\u05e7\u05d9\u05dd', 20, 4.5, '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9', 80),
    ('\u05e2\u05d9\u05e0\u05d9\u05d9\u05dd \u05de\u05e9\u05e7\u05e4\u05d9\u05d9\u05dd \u05e0\u05d5\u05e4\u05dc\u05d5\u05ea', 7, 4, '\u05d6\u05d5\u05dc', 60),
    ('\u05e9\u05e8\u05d1\u05d9\u05d8 \u05de\u05dc\u05da', 7, 4, '\u05d6\u05d5\u05dc', 60),
    ('\u05d0\u05e8\u05e0\u05e7', 7, 4, '\u05d6\u05d5\u05dc', 60),
    ('\u05e1\u05e4\u05de\u05d9\u05dd', 5, 3, '\u05d6\u05d5\u05dc', 50),
    ('\u05e1\u05d2\u05e8\u05e0\u05d4 \u05de\u05d9\u05dd \u05e9\u05e4\u05e8\u05d9\u05e5', 18, 3, '\u05d6\u05d5\u05dc', 50),
    ('\u05de\u05d8\u05d5\u05e1 \u05d9\u05d3\u05e0\u05d9', 60, 3, '\u05d6\u05d5\u05dc', 50),
    ('\u05e4\u05d8\u05d9\u05e9\u05d5\u05e0\u05d9\u05dd \u05e7\u05d8\u05e0\u05d9\u05dd', 48, 2.5, '\u05d6\u05d5\u05dc', 40),
    ('\u05de\u05e2\u05d5\u05d3\u05d3\u05d5\u05ea', 48, 2.43, '\u05d6\u05d5\u05dc', 40),
    ('\u05e8\u05d5\u05d1\u05d4 \u05de\u05d9\u05dd \u05e7\u05d8\u05df', 36, 2, '\u05d6\u05d5\u05dc', 35),
    ('\u05d1\u05d9\u05dc\u05d9\u05d0\u05e8\u05d3 \u05e7\u05d8\u05e0\u05d8\u05df', 40, 1.8, '\u05d6\u05d5\u05dc', 30),
    ('\u05e6\u05e4\u05e6\u05e4\u05d4', 70, 1.5, '\u05d6\u05d5\u05dc', 25),
    ('\u05d8\u05d9\u05e1\u05df', 50, 1.5, '\u05d6\u05d5\u05dc', 25),
    ('\u05e2\u05d8 \u05de\u05e8\u05d2\u05dc\u05d9\u05dd', 50, 1.5, '\u05d6\u05d5\u05dc', 25),
    ('\u05de\u05d8\u05e8 \u05de\u05d3\u05d9\u05d3\u05d4', 35, 1.2, '\u05d6\u05d5\u05dc', 20),
    ('\u05e6\u05d9\u05e0\u05d5\u05e8 \u05e9\u05e8\u05e9\u05d5\u05e8 \u05e0\u05e4\u05ea\u05d7', 200, 1, '\u05d6\u05d5\u05dc', 20),
    ('\u05d3\u05d5\u05e7\u05d9\u05dd', 20, 1, '\u05d6\u05d5\u05dc', 20),
    ('\u05d0\u05e3 \u05d0\u05d3\u05d5\u05dd', 25, 1, '\u05d6\u05d5\u05dc', 20),
    ('\u05e7\u05e4\u05d9\u05e5 \u05e4\u05dc\u05e1\u05d8\u05d9\u05e7 \u05e4\u05e9\u05d5\u05d8', 24, 0.9, '\u05d6\u05d5\u05dc', 20),
    ('\u05de\u05e1\u05db\u05d5\u05ea \u05d3\u05e7\u05d5\u05ea', 80, 0.9, '\u05d6\u05d5\u05dc', 20),
    ('\u05e4\u05e0\u05e1 \u05d0\u05e6\u05d1\u05e2', 32, 0.88, '\u05d6\u05d5\u05dc', 20),
    ('\u05e4\u05e8\u05d5\u05e4\u05dc\u05d5\u05e8 \u05de\u05e2\u05d5\u05e4\u05e3', 80, 0.8, '\u05d6\u05d5\u05dc', 20),
]

total_units = 0
total_cost = 0
for i, (name, units, cost, level, price) in enumerate(products_raw):
    ws5.cell(row=r, column=1, value=name)
    ws5.cell(row=r, column=2, value=units)
    ws5.cell(row=r, column=3, value=round(cost, 2))
    ws5.cell(row=r, column=4, value=round(units * cost, 1))
    ws5.cell(row=r, column=5, value=level)
    ws5.cell(row=r, column=6, value=price)
    style_row(ws5, r, 6, i % 2 == 1)
    if level == '\u05d9\u05e7\u05e8':
        ws5.cell(row=r, column=5).fill = red_fill
    elif level == '\u05d1\u05d9\u05e0\u05d5\u05e0\u05d9':
        ws5.cell(row=r, column=5).fill = yellow_fill
    else:
        ws5.cell(row=r, column=5).fill = green_fill
    total_units += units
    total_cost += units * cost
    r += 1

# Totals
ws5.cell(row=r, column=1, value='\u05e1\u05d4"\u05db')
ws5.cell(row=r, column=2, value=total_units)
ws5.cell(row=r, column=4, value=round(total_cost, 0))
for c in range(1, 7):
    ws5.cell(row=r, column=c).font = bold_font
    ws5.cell(row=r, column=c).fill = sub_header_fill
    ws5.cell(row=r, column=c).border = thin_border
    ws5.cell(row=r, column=c).alignment = rtl_center
r += 2
ws5.cell(row=r, column=1, value=f'\u05de\u05d5\u05e6\u05e8\u05d9\u05dd \u05dc\u05ea\u05dc\u05de\u05d9\u05d3 (\u05de\u05de\u05d5\u05e6\u05e2): {total_units/227:.1f}')
ws5.cell(row=r, column=1).font = bold_font

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 18
ws5.column_dimensions['E'].width = 15
ws5.column_dimensions['F'].width = 24

out_path = r'C:\מיצד\SchoolPoints\גיבוי רובע\יריד\תכנון_יריד_מפורט_v2.xlsx'
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Total store units: {total_units}')
print(f'Total store cost: {total_cost:.0f}')
print(f'Units per student: {total_units/227:.1f}')
