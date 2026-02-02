"""
עמדת ניהול - עדכון נקודות ושיוך כרטיסים
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser, scrolledtext, simpledialog
from database import Database
from excel_import import ExcelImporter
from PIL import Image, ImageTk, ImageDraw, ImageFont
import os
import json
import shutil
import subprocess
import socket
import sys
import ctypes
import threading
import webbrowser
import urllib.request
import urllib.parse
import base64
import re
from license_manager import LicenseManager
from typing import Optional
from datetime import datetime
import csv
from datetime import date, timedelta
from jewish_calendar import hebrew_date_from_gregorian_str

try:
    from ui_icons import normalize_ui_icons
except Exception:
    def normalize_ui_icons(text):
        return text

try:
    from sound_manager import SoundManager
except Exception:
    SoundManager = None

try:
    from build_flags import ENABLE_PURCHASES
except Exception:
    ENABLE_PURCHASES = True

# בפיתוח (הרצה ישירה מהקוד) נרצה שהפיצ'ר יהיה זמין בלי תלות ב-build_flags.
# בבילדים (PyInstaller) נשאיר את הדגל כפי שהוגדר.
try:
    if not bool(getattr(sys, 'frozen', False)):
        ENABLE_PURCHASES = True
except Exception:
    pass

# סימני BIDI לתיקון כיוון טקסט עברי
RLE = '\u202b'  # Right-to-Left Embedding
PDF = '\u202c'  # Pop Directional Formatting

UNIVERSAL_MASTER_CODE = "05276247440527624744"

APP_VERSION = "1.4.3"

def fix_rtl_text(text):
    """תיקון כיוון טקסט עברי עם סימני קריאה בצד הנכון"""
    if text and text.strip():
        return RLE + text + PDF
    return text

def _strip_asterisk_annotations(text):
    try:
        import re
        if not text:
            return text
        cleaned = re.sub(r'\*[^*]*\*', '', str(text))
        cleaned = re.sub(r'\s{2,}', ' ', cleaned).strip()
        return cleaned
    except Exception:
        return text

def safe_print(text):
    """הדפסה בטוחה שלא נכשלת על אימוג'ים"""
    try:
        print(normalize_ui_icons(text))
    except UnicodeEncodeError:
        # הסר אימוג'ים ונסה שוב
        import re
        text_no_emoji = re.sub(r'[^\u0000-\uFFFF]', '', text)
        try:
            print(text_no_emoji)
        except:
            pass  # אם עדיין יש בעיה, התעלם

def _is_running_as_admin():
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def _restart_as_admin():
    if _is_running_as_admin():
        return False
    try:
        exe = sys.executable
    except Exception:
        return False
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        candidate = os.path.join(base_dir, "admin_station.pyw")
        script = candidate if os.path.exists(candidate) else os.path.abspath(__file__)
    except Exception:
        return False
    params = f'"{script}"'
    try:
        res = ctypes.windll.shell32.ShellExecuteW(None, "runas", exe, params, None, 1)
    except Exception:
        return False
    try:
        code = int(res)
    except Exception:
        code = 0
    return code > 32


def _set_windows_dpi_awareness() -> None:
    try:
        if sys.platform != 'win32':
            return
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
            return
        except Exception:
            pass
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass
    except Exception:
        pass


def _apply_tk_scaling(root: tk.Tk) -> None:
    try:
        try:
            fpix = float(root.winfo_fpixels('1i') or 0.0)
        except Exception:
            fpix = 0.0
        if fpix and fpix > 0:
            scale = fpix / 72.0
        else:
            scale = 96.0 / 72.0
        if scale < 1.0:
            scale = 1.0
        if scale > 4.0:
            scale = 4.0
        root.tk.call('tk', 'scaling', float(scale))
    except Exception:
        pass


class ToggleSwitch(tk.Frame):
    """מתג ON/OFF קומפקטי בסגנון כפתור הזזה."""
    def __init__(self, master, variable=None, on_color="#1e90ff",
                 off_color="#bdc3c7", width=40, height=20, command=None, **kwargs):
        super().__init__(master, width=width, height=height, bg=master.cget('background'), **kwargs)
        self.variable = variable or tk.BooleanVar(value=False)
        self.on_color = on_color
        self.off_color = off_color
        self.command = command

        self.canvas = tk.Canvas(
            self,
            width=width,
            height=height,
            highlightthickness=0,
            bd=0,
            bg=master.cget('background')
        )
        self.canvas.pack()
        self.canvas.bind("<Button-1>", self._on_click)
        self.canvas.configure(cursor="hand2")

        self._width = width
        self._height = height
        self._draw()

    def _draw(self):
        self.canvas.delete("all")
        is_on = bool(self.variable.get())
        bg_color = self.on_color if is_on else self.off_color
        w = self._width
        h = self._height
        r = h // 2

        # רקע מעוגל
        self.canvas.create_oval(0, 0, h, h, outline="", fill=bg_color)
        self.canvas.create_oval(w - h, 0, w, h, outline="", fill=bg_color)
        self.canvas.create_rectangle(r, 0, w - r, h, outline="", fill=bg_color)

        # כפתור לבן זז
        if is_on:
            x0 = w - h
        else:
            x0 = 0
        self.canvas.create_oval(x0 + 2, 2, x0 + h - 2, h - 2, outline="", fill="#ffffff")

    def _on_click(self, event):
        self.variable.set(not self.variable.get())
        self._draw()
        if self.command:
            try:
                self.command()
            except Exception:
                pass

class AdminStation:
    def __init__(self, root, license_manager=None):
        self.root = root
        self.root.title("עמדת ניהול - מערכת ניקוד")
        try:
            # מניעת הבהוב חלונות "ריקים" בתחילת טעינה – נסתיר ויזואלית את ה-root
            # (לא withdraw כדי שחלונות setup/transient יוצגו כראוי)
            self.root.attributes('-alpha', 0.0)
        except Exception:
            pass
        try:
            self.root.update_idletasks()
        except Exception:
            pass
        try:
            self.root.state('zoomed')
        except Exception:
            try:
                sw = int(self.root.winfo_screenwidth() or 1360)
                sh = int(self.root.winfo_screenheight() or 760)
                self.root.geometry(f"{sw}x{sh}+0+0")
            except Exception:
                self.root.geometry("1360x760")
        self.root.configure(bg='#f0f0f0')
        
        # תיקיית בסיס של האפליקציה (תומך ב-UNC)
        try:
            if bool(getattr(sys, 'frozen', False)):
                self.base_dir = os.path.dirname(os.path.abspath(sys.executable))
            else:
                self.base_dir = os.path.dirname(os.path.abspath(__file__))
        except Exception:
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
        
        # טעינת הגדרות והשלמת הגדרות ראשוניות (כולל תיקיית רשת משותפת)
        self.app_config = self.load_app_config()
        if not self.ensure_initial_setup():
            try:
                self.root.after(100, self.root.destroy)
            except Exception:
                pass
            return

        # הפעלת סנכרון רקע אוטומטי (Hybrid/Cloud בלבד)
        self._sync_agent_thread = None
        self._sync_agent_started = False
        try:
            self._maybe_start_sync_agent()
        except Exception:
            pass
        
        # רישוי מערכת – רק לאחר שהוגדרה תיקיית הרשת/הגדרות האפליקציה
        self.license_manager = license_manager or LicenseManager(self.base_dir, "admin")
        
        self.db = Database()

        # השמעת צלילים בעמדת ניהול (בדיקות/תצוגה מקדימה) - אתחול עצל
        self._admin_sound_manager = None
        
        # הרצת המרת נתונים אוטומטית (פעם אחת בלבד)
        try:
            self.db.migrate_comma_separated_data()
        except Exception as e:
            print(f"שגיאה בהמרת נתונים: {e}")
        
        self.importer = ExcelImporter(self.db)
        
        # מורה מחובר (None = טרם התחבר, teacher_dict = מורה רגיל, {'is_admin': 1} = מנהל)
        self.current_teacher = None
        self.teacher_classes_cache = []  # cache של כיתות המורה

        try:
            import time
            self._last_activity_ts = float(time.time())
        except Exception:
            self._last_activity_ts = 0.0
        self._idle_lock_ms = 5 * 60 * 1000
        self._idle_check_job = None
        self._login_active = False
        self._lock_message_shown = False

        self._undo_stack = []
        self._redo_stack = []
        self._undo_max_actions = 100
        self._applying_undo_redo = False
        
        # הסתרה מיידית של חלון השורש כדי למנוע הבהוב/חלון רגעי לפני מסך ההתחברות
        try:
            self.root.withdraw()
        except Exception:
            pass

        # הצגת מסך התחברות
        self.show_login_screen()

    def _maybe_start_sync_agent(self) -> None:
        try:
            if bool(getattr(self, '_sync_agent_started', False)):
                return
        except Exception:
            pass

        try:
            cfg = self.load_app_config() or {}
        except Exception:
            cfg = {}

        try:
            mode = str(cfg.get('deployment_mode') or 'local').strip().lower()
        except Exception:
            mode = 'local'

        if mode not in ('hybrid', 'cloud'):
            return

        try:
            tenant_id = str(cfg.get('sync_tenant_id') or '').strip()
        except Exception:
            tenant_id = ''
        try:
            api_key = str(cfg.get('sync_api_key') or '').strip()
        except Exception:
            api_key = ''
        try:
            push_url = str(cfg.get('sync_push_url') or '').strip()
        except Exception:
            push_url = ''

        if not (tenant_id and api_key and push_url):
            return

        try:
            interval_sec = int(cfg.get('sync_interval_sec') or 60)
        except Exception:
            interval_sec = 60
        interval_sec = max(10, int(interval_sec or 60))

        def _run_sync_loop():
            try:
                import sync_agent
                sync_agent.main_loop(interval_sec=interval_sec)
            except Exception:
                pass

        try:
            self._sync_agent_started = True
        except Exception:
            pass
        try:
            t = threading.Thread(target=_run_sync_loop, daemon=True)
            self._sync_agent_thread = t
            t.start()
        except Exception:
            try:
                self._sync_agent_started = False
            except Exception:
                pass
            return

    def _device_pair_start_and_poll(self, cloud_base_url: str, poll_timeout_sec: int = 240) -> dict:
        try:
            import urllib.request
            import urllib.parse
            import urllib.error
            import time
        except Exception:
            return {'ok': False, 'error': 'missing urllib'}

        base = str(cloud_base_url or '').strip().rstrip('/')
        if not base:
            base = 'https://schoolpoints.co.il'

        start_url = base + '/api/device/pair/start'
        code = ''
        verify_url = ''

        try:
            req = urllib.request.Request(start_url, data=b'', method='POST')
            with urllib.request.urlopen(req, timeout=10) as resp:
                raw = resp.read().decode('utf-8', errors='ignore')
            try:
                data = json.loads(raw or '{}')
            except Exception:
                data = {}
            if isinstance(data, dict) and bool(data.get('ok')):
                code = str(data.get('code') or '').strip().upper()
                verify_url = str(data.get('verify_url') or '').strip()
        except urllib.error.HTTPError as e:
            try:
                body = e.read().decode('utf-8', errors='ignore')
            except Exception:
                body = ''
            msg = f"HTTP {getattr(e, 'code', '')}".strip()
            if body:
                msg = msg + f"\n{body}" if msg else body
            return {'ok': False, 'error': msg or 'http error'}
        except Exception as e:
            return {'ok': False, 'error': str(e)}

        if not (code and verify_url):
            return {'ok': False, 'error': 'start failed'}

        try:
            webbrowser.open(verify_url)
        except Exception:
            pass

        poll_url = base + '/api/device/pair/poll?code=' + urllib.parse.quote(code)
        deadline = time.time() + max(10, int(poll_timeout_sec or 240))
        while time.time() < deadline:
            try:
                with urllib.request.urlopen(poll_url, timeout=10) as resp:
                    raw2 = resp.read().decode('utf-8', errors='ignore')
                try:
                    data2 = json.loads(raw2 or '{}')
                except Exception:
                    data2 = {}
                if isinstance(data2, dict) and bool(data2.get('ok')):
                    st = str(data2.get('status') or '').strip().lower()
                    if st == 'ready':
                        tid = str(data2.get('tenant_id') or '').strip()
                        key = str(data2.get('api_key') or '').strip()
                        purl = str(data2.get('push_url') or '').strip()
                        if tid and key and purl:
                            return {
                                'ok': True,
                                'code': code,
                                'verify_url': verify_url,
                                'tenant_id': tid,
                                'api_key': key,
                                'push_url': purl,
                            }
                    if st == 'consumed':
                        return {'ok': False, 'error': 'consumed'}
            except Exception:
                pass
            time.sleep(2)

        return {'ok': False, 'error': 'timeout', 'code': code, 'verify_url': verify_url}

    def _push_points_action(self, label: str, changes: list):
        """שומר פעולה לשחזור (Undo). changes = [{'student_id':..,'old_points':..,'new_points':..}, ...]"""
        try:
            if getattr(self, '_applying_undo_redo', False):
                return
        except Exception:
            pass

        cleaned = []
        for ch in (changes or []):
            try:
                sid = int(ch.get('student_id'))
                oldp = int(ch.get('old_points'))
                newp = int(ch.get('new_points'))
                if oldp == newp:
                    continue
                cleaned.append({'student_id': sid, 'old_points': oldp, 'new_points': newp})
            except Exception:
                continue

        if not cleaned:
            return

        try:
            self._undo_stack.append({'label': str(label or ''), 'changes': cleaned})
            if len(self._undo_stack) > int(getattr(self, '_undo_max_actions', 100) or 100):
                self._undo_stack = self._undo_stack[-int(getattr(self, '_undo_max_actions', 100) or 100):]
        except Exception:
            return

        try:
            self._redo_stack = []
        except Exception:
            pass

        try:
            self._update_undo_redo_ui_state()
        except Exception:
            pass

    def _apply_points_changes(self, changes: list, direction: str):
        if direction not in ('undo', 'redo'):
            return 0
        applied = 0
        for ch in (changes or []):
            try:
                sid = int(ch.get('student_id'))
                oldp = int(ch.get('old_points'))
                newp = int(ch.get('new_points'))
                target = oldp if direction == 'undo' else newp
                actor_name = self._get_points_actor_name()
                ok = self.db.update_student_points(sid, max(0, int(target)), f"{direction.upper()}", actor_name)
                if ok:
                    applied += 1
            except Exception:
                pass
        return applied

    def undo_last_points_action(self, *_):
        try:
            if not self._undo_stack:
                return
        except Exception:
            return

        action = None
        try:
            action = self._undo_stack.pop()
        except Exception:
            return
        if not action:
            return

        try:
            self._applying_undo_redo = True
        except Exception:
            pass
        try:
            self._apply_points_changes(action.get('changes') or [], 'undo')
        finally:
            try:
                self._applying_undo_redo = False
            except Exception:
                pass

        try:
            self._redo_stack.append(action)
        except Exception:
            pass

        try:
            self.has_changes = True
        except Exception:
            pass
        try:
            self.load_students(keep_selection=True)
        except Exception:
            try:
                self.load_students()
            except Exception:
                pass
        try:
            self.export_to_excel_now()
        except Exception:
            pass
        try:
            self.show_status_message("↩ בוטלה פעולה אחרונה", '#2980b9')
        except Exception:
            pass
        try:
            self._update_undo_redo_ui_state()
        except Exception:
            pass

    def redo_last_points_action(self, *_):
        try:
            if not self._redo_stack:
                return
        except Exception:
            return

        action = None
        try:
            action = self._redo_stack.pop()
        except Exception:
            return
        if not action:
            return

        try:
            self._applying_undo_redo = True
        except Exception:
            pass
        try:
            self._apply_points_changes(action.get('changes') or [], 'redo')
        finally:
            try:
                self._applying_undo_redo = False
            except Exception:
                pass

        try:
            self._undo_stack.append(action)
        except Exception:
            pass

        try:
            self.has_changes = True
        except Exception:
            pass
        try:
            self.load_students(keep_selection=True)
        except Exception:
            try:
                self.load_students()
            except Exception:
                pass
        try:
            self.export_to_excel_now()
        except Exception:
            pass
        try:
            self.show_status_message("↪ הוחזרה פעולה", '#2980b9')
        except Exception:
            pass
        try:
            self._update_undo_redo_ui_state()
        except Exception:
            pass

    def _update_undo_redo_ui_state(self):
        try:
            undo_btn = getattr(self, 'undo_btn', None)
            if undo_btn is not None:
                undo_btn.config(state=(tk.NORMAL if self._undo_stack else tk.DISABLED))
        except Exception:
            pass
        try:
            redo_btn = getattr(self, 'redo_btn', None)
            if redo_btn is not None:
                redo_btn.config(state=(tk.NORMAL if self._redo_stack else tk.DISABLED))
        except Exception:
            pass

    def _touch_activity(self, *_):
        try:
            import time
            self._last_activity_ts = float(time.time())
        except Exception:
            pass

    def _start_idle_monitor(self):
        # עוקב רק אחרי התחברות (לא במסך התחברות)
        try:
            self._touch_activity()
        except Exception:
            pass

        try:
            self.root.bind_all('<Any-KeyPress>', self._touch_activity)
            self.root.bind_all('<Any-Button>', self._touch_activity)
            self.root.bind_all('<Motion>', self._touch_activity)
            self.root.bind_all('<MouseWheel>', self._touch_activity)
        except Exception:
            pass

        try:
            if getattr(self, '_idle_check_job', None):
                self.root.after_cancel(self._idle_check_job)
        except Exception:
            pass

        def _tick():
            try:
                if getattr(self, '_login_active', False):
                    return
                if not getattr(self, 'current_teacher', None):
                    return
                import time
                idle_ms = (float(time.time()) - float(getattr(self, '_last_activity_ts', 0.0) or 0.0)) * 1000.0
                if idle_ms >= float(getattr(self, '_idle_lock_ms', 0) or 0):
                    # בחוסר פעילות נחזור למסך התחברות בלי להציג הודעה, כדי לא לחשוף את הטבלה ברקע
                    self.lock_to_login(show_message=False)
                    return
            except Exception:
                pass
            try:
                self._idle_check_job = self.root.after(1000, _tick)
            except Exception:
                self._idle_check_job = None

        try:
            self._idle_check_job = self.root.after(1000, _tick)
        except Exception:
            self._idle_check_job = None

    def lock_to_login(self, show_message: bool = True):
        try:
            if getattr(self, '_login_active', False):
                return
            if not getattr(self, 'current_teacher', None):
                return
        except Exception:
            return

        try:
            self._login_active = True
        except Exception:
            pass

        # הסתרה מיידית של החלון הראשי כדי שלא תוצג הטבלה/מידע בזמן נעילה
        try:
            self.root.withdraw()
        except Exception:
            pass

        if show_message and not getattr(self, '_lock_message_shown', False):
            try:
                self._lock_message_shown = True
            except Exception:
                pass
            try:
                messagebox.showinfo("נעילה", "התוכנה ננעלה עקב חוסר פעילות.\nיש להתחבר מחדש.")
            except Exception:
                pass

        # עצירת משימות after מחזוריות כדי למנוע הצטברות/רענונים ברקע
        for attr in (
            'auto_refresh_job',
            'auto_sync_job',
            'update_check_job',
            '_idle_check_job'
        ):
            job_id = getattr(self, attr, None)
            if job_id:
                try:
                    self.root.after_cancel(job_id)
                except Exception:
                    pass
                try:
                    setattr(self, attr, None)
                except Exception:
                    pass

        # ניקוי ה-UI הראשי
        try:
            for child in self.root.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
        except Exception:
            pass

        try:
            self.teacher_classes_cache = []
        except Exception:
            pass

        try:
            self.current_teacher = None
        except Exception:
            pass

        try:
            self._lock_message_shown = False
        except Exception:
            pass

        # הצגת מסך התחברות מחדש
        try:
            self.show_login_screen()
        except Exception:
            # fallback: נסה לשחזר חלון
            try:
                self.root.deiconify()
            except Exception:
                pass
    
    def setup_ui(self):
        """בניית ממשק המשתמש"""
        # כותרת עם לוגו וכפתור החלפת משתמש - קטנה יותר
        header = tk.Frame(self.root, bg='#2c3e50', height=50)
        header.pack(fill=tk.X)
        
        # לוגו - קטן יותר
        logo_path = os.path.join(self.base_dir, "דובר שלום לוגו תת.jpg")
        if os.path.exists(logo_path):
            try:
                img = Image.open(logo_path)
                img = img.resize((40, 40), Image.Resampling.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img)
                logo_label = tk.Label(header, image=self.logo_img, bg='#2c3e50')
                logo_label.pack(side=tk.RIGHT, padx=10)
            except:
                pass
        
        # מסגרת לימין (לוגו + כותרת)
        title_container = tk.Frame(header, bg='#2c3e50')
        title_container.pack(side=tk.RIGHT, padx=10, pady=4)

        title_label = tk.Label(
            title_container,
            text="עמדת ניהול - מערכת ניקוד בית ספרית",
            font=('Arial', 14, 'bold'),
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(side=tk.RIGHT, padx=5)

        # כפתור "החלף משתמש" בצד שמאל למעלה
        def _switch_user():
            """החלפת משתמש – חזרה למסך התחברות בלי לסגור את התוכנה."""
            try:
                self.teacher_classes_cache = []
            except Exception:
                pass

            # עצירת משימות after מחזוריות כדי למנוע "הצטברות" חלונות/רענונים ברקע
            for attr in (
                'auto_refresh_job',
                'auto_sync_job',
                'update_check_job'
            ):
                job_id = getattr(self, attr, None)
                if job_id:
                    try:
                        self.root.after_cancel(job_id)
                    except Exception:
                        pass
                    try:
                        setattr(self, attr, None)
                    except Exception:
                        pass

            # ניקוי ה-UI הראשי
            for child in self.root.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

            # איפוס מצב והצגת מסך התחברות מחדש
            self.current_teacher = None
            self.show_login_screen()

        switch_btn = tk.Button(
            header,
            text="🔄 החלף משתמש",
            command=_switch_user,
            font=('Arial', 9),
            bg='#34495e',
            fg='white',
            padx=8,
            pady=3,
            cursor='hand2'
        )
        switch_btn.pack(side=tk.LEFT, padx=10, pady=8)
        
        # שורת משנה בכותרת עבור מצב רישיון (בצד שמאל, מתחת לכותרת)
        sub_header = tk.Frame(header, bg='#2c3e50')
        sub_header.pack(fill=tk.X, side=tk.BOTTOM)

        lm_header = getattr(self, "license_manager", None)

        def get_license_header_text():
            if lm_header is None:
                return "מצב רישיון: (לא זמין)"
            if lm_header.is_licensed:
                name = lm_header.school_name or "לא ידוע"
                used = getattr(lm_header, 'used_stations', 0) or 0
                total = getattr(lm_header, 'max_stations', 0) or 0
                free = max(0, total - used)
                extra = ""
                try:
                    try:
                        _is_term_attr = getattr(lm_header, 'is_term', False)
                        is_term = bool(_is_term_attr() if callable(_is_term_attr) else _is_term_attr)
                    except Exception:
                        is_term = False
                    if getattr(lm_header, 'is_monthly', False):
                        exp = str(getattr(lm_header, 'expiry_date', '') or '').strip()
                        if exp:
                            extra = f" | חודשי עד {exp}"
                        try:
                            left = int(getattr(lm_header, 'monthly_days_left', 0) or 0)
                        except Exception:
                            left = 0
                        if left:
                            extra += f" | נותרו {left} ימים"
                        try:
                            allow_cashier = bool(getattr(lm_header, 'allow_cashier', True))
                        except Exception:
                            allow_cashier = True
                        extra += " | קופה: " + ("כן" if allow_cashier else "לא")
                    elif is_term:
                        exp = str(getattr(lm_header, 'expiry_date', '') or '').strip()
                        act = str(getattr(lm_header, 'activated_at', '') or '').strip()
                        try:
                            td = int(getattr(lm_header, 'term_days', 0) or 0)
                        except Exception:
                            td = 0
                        try:
                            left = int(getattr(lm_header, 'term_days_left', 0) or 0)
                        except Exception:
                            left = 0
                        used_days = max(0, int(td) - int(left)) if td else 0
                        try:
                            allow_cashier = bool(getattr(lm_header, 'allow_cashier', True))
                        except Exception:
                            allow_cashier = True
                        if td:
                            extra = f" | נרשם ל-{td} ימים"
                        else:
                            extra = " | נרשם"
                        if act:
                            extra += f" | הופעל ב-{act}"
                        if exp:
                            extra += f" | עד {exp}"
                        extra += f" | בשימוש {used_days} ימים, נותרו {left}"
                        extra += " | קופה: " + ("כן" if allow_cashier else "לא")
                except Exception:
                    extra = ""
                return (
                    f"מצב רישיון: פעיל – נרשם ל{name} "
                    f"עבור {total} עמדות ({used} בשימוש, {free} פנויות){extra}"
                )
            if lm_header.is_trial:
                return f"מצב רישיון: גרסת ניסיון – נותרו {lm_header.trial_days_left} ימים"
            if lm_header.trial_expired:
                return "מצב רישיון: תקופת הניסיון הסתיימה – ללא רישיון פעיל"
            return "מצב רישיון: לא רשום"

        self.license_header_label = tk.Label(
            sub_header,
            text=get_license_header_text(),
            font=('Arial', 9),
            bg='#2c3e50',
            fg='#ecf0f1',
            anchor='w'
        )
        self.license_header_label.pack(side=tk.LEFT, padx=10, pady=(0, 2))
        
        # פאנל כפתורים עליון - קומפקטי יותר
        button_frame = tk.Frame(self.root, bg='#f0f0f0', pady=5)
        button_frame.pack(fill=tk.X, padx=10)

        button_frame2 = None
        try:
            sw = int(self.root.winfo_screenwidth() or 0)
            sh = int(self.root.winfo_screenheight() or 0)
        except Exception:
            sw, sh = 0, 0
        if bool((sw and sw <= 1100) or (sh and sh <= 820)):
            button_frame2 = tk.Frame(self.root, bg='#f0f0f0', pady=0)
            button_frame2.pack(fill=tk.X, padx=10, pady=(0, 5))
        
        # כפתורים למנהלים בלבד
        is_admin = self.current_teacher and self.current_teacher['is_admin'] == 1
        
        if is_admin:
            # כפתורי ניהול תלמידים: הוספה / עריכה / מחיקה
            manage_frame = tk.Frame(button_frame, bg='#f0f0f0')
            manage_frame.pack(side=tk.RIGHT, padx=(0, 10))

            tk.Button(
                manage_frame,
                text="➕ תלמיד",
                command=self.add_student_dialog,
                font=('Arial', 9),
                bg='#27ae60',
                fg='white',
                padx=6,
                pady=3,
                cursor='hand2'
            ).pack(side=tk.RIGHT, padx=2)

            tk.Button(
                manage_frame,
                text="✏ ערוך",
                command=self.edit_student_dialog,
                font=('Arial', 9),
                bg='#f1c40f',
                fg='black',
                padx=6,
                pady=3,
                cursor='hand2'
            ).pack(side=tk.RIGHT, padx=2)

            tk.Button(
                manage_frame,
                text="🗑 מחק",
                command=self.delete_student_dialog,
                font=('Arial', 9),
                bg='#e74c3c',
                fg='white',
                padx=6,
                pady=3,
                cursor='hand2'
            ).pack(side=tk.RIGHT, padx=2)

            # כפתור ייבוא - קטן יותר (רק מנהלים)
            import_btn = tk.Button(
                button_frame,
                text="📥 ייבוא",
                command=self.import_excel,
                font=('Arial', 9),
                bg='#3498db',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            import_btn.pack(side=tk.RIGHT, padx=3)
            
            # כפתור ייצוא (רק מנהלים)
            export_btn = tk.Button(
                button_frame,
                text="📤 ייצוא",
                command=self.export_excel,
                font=('Arial', 9),
                bg='#2ecc71',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            export_btn.pack(side=tk.RIGHT, padx=3)
        
            # כפתור סינכרון - לכולם! (גיבוי חשוב)
            sync_btn = tk.Button(
                button_frame,
                text="🔄 סינכרון",
                command=self.sync_to_excel,
                font=('Arial', 9),
                bg='#9b59b6',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            sync_btn.pack(side=tk.RIGHT, padx=3)
        
        refresh_btn = tk.Button(
            button_frame,
            text="🔄 רענון",
            command=self.refresh_table,
            font=('Arial', 9),
            bg='#95a5a6',
            fg='white',
            padx=8,
            pady=4,
            cursor='hand2'
        )
        refresh_btn.pack(side=tk.RIGHT, padx=3)
        
        # כפתורים נוספים
        if is_admin:
            # כפתור הגדרות צבעים - רק למנהלים
            color_btn = tk.Button(
                button_frame,
                text="שדרוגים",
                command=self.open_color_settings,
                font=('Arial', 9),
                bg='#8e44ad',
                fg='white',
                padx=10,
                pady=5,
                cursor='hand2'
            )
            color_btn.pack(side=tk.LEFT, padx=5)
            
            # כפתור ניהול הודעות - רק למנהלים (הודעות כלליות)
            messages_btn = tk.Button(
                button_frame2 or button_frame,
                text="💬 הודעות כלליות",
                command=self.open_messages_manager,
                font=('Arial', 9),
                bg='#16a085',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            messages_btn.pack(side=tk.LEFT, padx=3)
            
            # כפתור מצב בונוס מיוחד (מאסטר 2 בעמדה הציבורית) - רק למנהלים
            self.bonus_btn = tk.Button(
                button_frame2 or button_frame,
                text="🎁 בונוס מיוחד",
                command=self.toggle_bonus_mode,
                font=('Arial', 9, 'bold'),
                bg='#f39c12',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            self.bonus_btn.pack(side=tk.LEFT, padx=3)
            
            # כפתור בונוס זמנים - רק למנהלים
            tk.Button(
                button_frame2 or button_frame,
                text="⏰ בונוס זמנים",
                command=self.open_time_bonus_manager,
                font=('Arial', 8, 'bold'),
                bg='#3498db',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=3)

            # כפתור ניהול מורים - רק למנהלים (בין בונוס זמנים להגדרות מערכת)
            tk.Button(
                button_frame2 or button_frame,
                text="👥 ניהול מורים",
                command=self.open_teachers_manager,
                font=('Arial', 9, 'bold'),
                bg='#e67e22',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=3)
            
            # כפתור הגדרות מערכת - רק למנהלים
            tk.Button(
                button_frame2 or button_frame,
                text="⚙ הגדרות מערכת",
                command=self.open_system_settings,
                font=('Arial', 9),
                bg='#34495e',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=3)

            tk.Button(
                button_frame2 or button_frame,
                text="🖥 הגדרות תצוגה",
                command=self.open_display_settings,
                font=('Arial', 9),
                bg='#34495e',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=3)

            # כפתור קניות (רק למנהלים) – בולט יותר
            if bool(ENABLE_PURCHASES) and (self.current_teacher and self.current_teacher.get('is_admin') == 1):
                tk.Button(
                    button_frame2 or button_frame,
                    text="🛒 קניות",
                    command=self.open_purchases_manager,
                    font=('Arial', 9, 'bold'),
                    bg='#8e44ad',
                    fg='white',
                    activebackground='#7d3c98',
                    activeforeground='white',
                    padx=10,
                    pady=4,
                    cursor='hand2'
                ).pack(side=tk.LEFT, padx=3)
        else:
            # מורה (לא מנהל): ייצוא מוגבל לכיתות מורשות + (אופציונלי) כפתור בונוס למורה
            if bool(ENABLE_PURCHASES):
                self.teacher_bonus_btn = tk.Button(
                    button_frame,
                    text="🎁 בונוס",
                    command=self.open_teacher_bonus_dialog,
                    font=('Arial', 9, 'bold'),
                    bg='#f39c12',
                    fg='white',
                    padx=8,
                    pady=4,
                    cursor='hand2'
                )
                self.teacher_bonus_btn.pack(side=tk.LEFT, padx=3)

            export_btn = tk.Button(
                button_frame,
                text="📤 ייצוא",
                command=self.export_excel,
                font=('Arial', 9),
                bg='#2ecc71',
                fg='white',
                padx=8,
                pady=4,
                cursor='hand2'
            )
            export_btn.pack(side=tk.RIGHT, padx=3)
        
        # (הצגת סטטיסטיקות עברה לחלון "הגדרות מערכת")
        
        # כפתור עזרה / הוראות שימוש – ממוקם יחד עם שאר הכפתורים
        tk.Button(
            button_frame2 or button_frame,
            text="❓ הוראות",
            command=self.open_help_text_dialog,
            font=('Arial', 9),
            bg='#2980b9',
            fg='white',
            padx=8,
            pady=4,
            cursor='hand2'
        ).pack(side=tk.LEFT, padx=3)
        
        # תווית סטטוס סינכרון - קטנה יותר, בקצה השמאלי של שורת הכפתורים
        self.sync_label = tk.Label(
            button_frame2 or button_frame,
            text="סטטוס: מוכן",
            font=('Arial', 8),
            bg='#f0f0f0',
            fg='#27ae60'
        )
        self.sync_label.pack(side=tk.LEFT, padx=8)

        self._teacher_stats_frame = None
        self._teacher_stats_classes_label = None
        self._teacher_stats_students_label = None
        self._teacher_stats_points_label = None
        self._teacher_stats_max_allowed_label = None

        if not is_admin:
            self._teacher_stats_frame = tk.Frame(button_frame2 or button_frame, bg='#f0f0f0')
            self._teacher_stats_frame.pack(side=tk.RIGHT, padx=8)

            # Filter button - only show if teacher has more than 1 class
            self._teacher_filter_button = tk.Button(
                self._teacher_stats_frame,
                text='🔍 סנן',
                command=self.open_teacher_class_filter_dialog,
                font=('Arial', 9, 'bold'),
                bg='#34495e',
                fg='white',
                padx=10,
                pady=4,
                cursor='hand2'
            )
            # Will be packed dynamically based on number of classes

            # Order: כללי | גבוה | נמוך | כיתה שלך (left to right, so pack in reverse)
            self._teacher_stats_your_classes_label = tk.Label(
                self._teacher_stats_frame,
                text=fix_rtl_text('הכיתות שלך: ...'),
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#2c3e50',
                anchor='e'
            )
            self._teacher_stats_your_classes_label.pack(side=tk.LEFT, padx=6)

            # Separator before "your classes"
            tk.Label(
                self._teacher_stats_frame,
                text='|',
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#95a5a6'
            ).pack(side=tk.LEFT, padx=2)

            self._teacher_stats_lowest_label = tk.Label(
                self._teacher_stats_frame,
                text=fix_rtl_text('הנמוך ביותר: ...'),
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#2c3e50',
                anchor='e'
            )
            self._teacher_stats_lowest_label.pack(side=tk.LEFT, padx=6)

            # Separator
            tk.Label(
                self._teacher_stats_frame,
                text='|',
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#95a5a6'
            ).pack(side=tk.LEFT, padx=2)

            self._teacher_stats_highest_label = tk.Label(
                self._teacher_stats_frame,
                text=fix_rtl_text('הגבוה ביותר: ...'),
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#2c3e50',
                anchor='e'
            )
            self._teacher_stats_highest_label.pack(side=tk.LEFT, padx=6)

            # Separator
            tk.Label(
                self._teacher_stats_frame,
                text='|',
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#95a5a6'
            ).pack(side=tk.LEFT, padx=2)

            self._teacher_stats_overall_label = tk.Label(
                self._teacher_stats_frame,
                text=fix_rtl_text('ממוצע כללי: ...'),
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#2c3e50',
                anchor='e'
            )
            self._teacher_stats_overall_label.pack(side=tk.LEFT, padx=6)

            # Separator
            tk.Label(
                self._teacher_stats_frame,
                text='|',
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#95a5a6'
            ).pack(side=tk.LEFT, padx=2)

            self._teacher_stats_max_allowed_label = tk.Label(
                self._teacher_stats_frame,
                text=fix_rtl_text('מקסימום אפשרי: ...'),
                font=('Arial', 9),
                bg='#f0f0f0',
                fg='#2c3e50',
                anchor='e'
            )
            self._teacher_stats_max_allowed_label.pack(side=tk.LEFT, padx=6)

        try:
            self.root.bind_all('<Control-z>', self.undo_last_points_action)
            self.root.bind_all('<Control-y>', self.redo_last_points_action)
            self.root.bind_all('<Control-Shift-Z>', self.redo_last_points_action)
        except Exception:
            pass

        try:
            self._update_undo_redo_ui_state()
        except Exception:
            pass
        
        # פאנל חיפוש - קומפקטי
        search_frame = tk.Frame(self.root, bg='#f0f0f0')
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        search_frame.columnconfigure(0, weight=1)
        search_frame.columnconfigure(1, weight=0)
        
        tk.Label(search_frame, text="", bg='#f0f0f0').grid(row=0, column=0, sticky='we')
        tk.Label(
            search_frame,
            text="חיפוש:",
            font=('Arial', 9),
            bg='#f0f0f0'
        ).grid(row=0, column=2, sticky='e', padx=3)
        
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', self.on_search_var_changed)
        
        # שדה חיפוש – נשמר כ-attrib כדי שנוכל להחזיר אליו פוקוס אחרי ניקוי אוטומטי
        self.search_entry = tk.Entry(
            search_frame,
            textvariable=self.search_var,
            font=('Arial', 9),
            width=30,
            justify='right'
        )
        self.search_entry.grid(row=0, column=1, sticky='e', padx=5)
        
        # טבלת תלמידים עם שורות מסומנות
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(table_frame)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        
        # סגנון לטבלה עם שורות מסומנות - קומפקטי
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview",
                       background="#ffffff",
                       foreground="#000000",
                       rowheight=20,
                       fieldbackground="#ffffff",
                       font=('Arial', 9))
        style.configure("Treeview.Heading",
                       font=('Arial', 9, 'bold'),
                       background="#34495e",
                       foreground="#ffffff")
        style.map('Treeview',
                 background=[('selected', '#3498db')])
        
        # Treeview - מימין לשמאל (בלי ID!)
        columns = ('נקודות', 'הודעה פרטית', 'מס\' כרטיס', 'כיתה', 'ת"ז', 'שם פרטי', 'שם משפחה', 'תמונה', 'מס\' סידורי', 'סה"כ תיקופים', 'ממוצע תיקופים ליום')
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show='headings',
            height=20,
            yscrollcommand=scrollbar.set,
            selectmode='browse'
        )
        
        # שמירת ID בנפרד (לא מוצג בטבלה)
        self.student_ids = {}  # מיפוי item_id -> student_id
        
        scrollbar.config(command=self.tree.yview)
        
        # הגדרת עמודות - מימין לשמאל - קומפקטי
        self.tree.column('שם משפחה', width=100, anchor='e')
        self.tree.column('שם פרטי', width=100, anchor='e')

        # הסתרת עמודת ת"ז מהתצוגה (נשארת בנתונים בלבד)
        self.tree.column('ת"ז', width=0, minwidth=0, anchor='center', stretch=False)
        self.tree.heading('ת"ז', text='', anchor='center', command=lambda: None)
        self.tree.column('כיתה', width=60, anchor='center')
        self.tree.column('מס\' כרטיס', width=80, anchor='center')
        self.tree.column('הודעה פרטית', width=180, anchor='e')
        self.tree.column('נקודות', width=60, anchor='center')
        self.tree.column('תמונה', width=40, anchor='center')
        self.tree.column('מס\' סידורי', width=20, anchor='center')
        self.tree.column('סה"כ תיקופים', width=80, anchor='center')
        self.tree.column('ממוצע תיקופים ליום', width=110, anchor='center')

        self.tree['displaycolumns'] = (
            'ממוצע תיקופים ליום',
            'סה"כ תיקופים',
            'נקודות',
            'הודעה פרטית',
            'מס\' כרטיס',
            'כיתה',
            'שם פרטי',
            'שם משפחה',
            'תמונה',
            'מס\' סידורי'
        )
        
        # כותרות
        for col in columns:
            header_text = '#' if col == "מס' סידורי" else col
            self.tree.heading(col, text=header_text, anchor='center',
                              command=lambda c=col: self.sort_by_column(c, False))
        
        # צבעים לשורות לסירוגין
        self.tree.tag_configure('oddrow', background='#ecf0f1')
        self.tree.tag_configure('evenrow', background='#ffffff')
        
        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.RIGHT)
        
        # פאנל עריכה תחתון - קומפקטי
        edit_frame = tk.Frame(self.root, bg='#ecf0f1', pady=5)
        edit_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # שורה ראשונה - סוג עדכון נקודות
        row1 = tk.Frame(edit_frame, bg='#ecf0f1')
        row1.pack(fill=tk.X, pady=2)
        row1.columnconfigure(0, weight=1)
        
        # בחירת סוג עדכון
        tk.Label(
            row1,
            text=fix_rtl_text("סוג עדכון:"),
            font=('Arial', 9, 'bold'),
            bg='#ecf0f1'
        ).grid(row=0, column=4, sticky='e', padx=3)
        self.update_type = tk.StringVar(value="add")  # ברירת מחדל: הוספה
        
        tk.Radiobutton(
            row1,
            text="הוספה (+)",
            variable=self.update_type,
            value="add",
            command=self.on_update_type_changed,
            font=('Arial', 8, 'bold'),
            bg='#ecf0f1',
            fg='#27ae60'
        ).grid(row=0, column=3, sticky='e', padx=3)
        
        tk.Radiobutton(
            row1,
            text="חיסור (-)",
            variable=self.update_type,
            value="subtract",
            command=self.on_update_type_changed,
            font=('Arial', 8),
            bg='#ecf0f1'
        ).grid(row=0, column=2, sticky='e', padx=3)
        
        tk.Radiobutton(
            row1,
            text="מוחלט",
            variable=self.update_type,
            value="set",
            command=self.on_update_type_changed,
            font=('Arial', 8),
            bg='#ecf0f1'
        ).grid(row=0, column=1, sticky='e', padx=3)
        
        # שורה שנייה - נקודות (מימין) + הודעה פרטית (משמאל) + כפתור מהיר
        row2 = tk.Frame(edit_frame, bg='#ecf0f1')
        row2.pack(fill=tk.X, pady=2)

        # נקודות (מימין, מתחת לסוג עדכון)
        tk.Label(
            row2,
            text=fix_rtl_text("נקודות:"),
            font=('Arial', 9),
            bg='#ecf0f1'
        ).pack(side=tk.RIGHT, padx=5)
        
        self.points_entry = tk.Entry(row2, font=('Arial', 9), width=10)
        self.points_entry.pack(side=tk.RIGHT, padx=5)
        self.points_entry.bind('<Return>', self.on_points_entry_return)
        self.points_entry.bind('<Shift-Return>', self.on_points_entry_shift_return)
        self.points_entry.bind('<Tab>', self.on_points_entry_tab)
        self.points_entry.bind('<Down>', lambda e: self._entry_nav_move(1, self.points_entry))
        self.points_entry.bind('<Up>', lambda e: self._entry_nav_move(-1, self.points_entry))
        
        tk.Button(
            row2,
            text="עדכן נקודות",
            command=self.update_points,
            font=('Arial', 9),
            bg='#f39c12',
            fg='white',
            padx=10,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=5)
        
        # כפתור עדכון מהיר - ליד עדכון נקודות!
        self.quick_update_btn = tk.Button(
            row2,
            text="⚡ מהיר",
            command=self.toggle_quick_update,
            font=('Arial', 9, 'bold'),
            bg='#e67e22',
            fg='white',
            padx=15,
            pady=5,
            cursor='hand2'
        )
        self.quick_update_btn.pack(side=tk.RIGHT, padx=5)

        # Undo/Redo לניקוד
        self.undo_btn = tk.Button(
            row2,
            text="↩ בטל",
            command=self.undo_last_points_action,
            font=('Arial', 9, 'bold'),
            bg='#2980b9',
            fg='white',
            padx=10,
            pady=5,
            cursor='hand2',
            state=tk.DISABLED
        )
        self.undo_btn.pack(side=tk.RIGHT, padx=5)

        self.redo_btn = tk.Button(
            row2,
            text="↪ החזר",
            command=self.redo_last_points_action,
            font=('Arial', 9, 'bold'),
            bg='#2980b9',
            fg='white',
            padx=10,
            pady=5,
            cursor='hand2',
            state=tk.DISABLED
        )
        self.redo_btn.pack(side=tk.RIGHT, padx=5)

        tk.Label(row2, text="", bg='#ecf0f1', width=1).pack(side=tk.RIGHT)

        tk.Button(
            row2,
            text="🕒 תיקון תיקוף",
            command=self.open_manual_swipe_override,
            font=('Arial', 9, 'bold'),
            bg='#8e44ad',
            fg='white',
            padx=10,
            pady=5,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=5)

        tk.Label(row2, text="", bg='#ecf0f1', width=1).pack(side=tk.RIGHT)

        # הודעה פרטית (משמאל יותר)
        tk.Label(
            row2,
            text=fix_rtl_text("הודעה פרטית:"),
            font=('Arial', 9),
            bg='#ecf0f1'
        ).pack(side=tk.RIGHT, padx=5)
        
        self.message_entry = tk.Entry(row2, font=('Arial', 9), width=40, justify='right')
        self.message_entry.pack(side=tk.RIGHT, padx=5)
        self.message_entry.bind('<Return>', self.on_message_entry_return)
        self.message_entry.bind('<Shift-Return>', self.on_message_entry_shift_return)
        self.message_entry.bind('<Tab>', self.on_message_entry_tab)
        self.message_entry.bind('<Shift-Tab>', self.on_message_entry_shift_tab)
        self.message_entry.bind('<Down>', lambda e: self._entry_nav_move(1, self.message_entry))
        self.message_entry.bind('<Up>', lambda e: self._entry_nav_move(-1, self.message_entry))
        self._attach_entry_edit_menu(self.message_entry)
        
        tk.Button(
            row2,
            text="עדכן הודעה",
            command=self.update_message,
            font=('Arial', 9),
            bg='#16a085',
            fg='white',
            padx=10,
            pady=3,
            cursor='hand2'
        ).pack(side=tk.RIGHT, padx=10)
        
        # שורה שלישית - כרטיס + נתיב תמונה
        row3 = tk.Frame(edit_frame, bg='#ecf0f1')
        row3.pack(fill=tk.X, pady=2)
        
        card_label = tk.Label(
            row3,
            text=fix_rtl_text("כרטיס:"),
            font=('Arial', 9),
            bg='#ecf0f1'
        )
        card_label.pack(side=tk.RIGHT, padx=5)
        
        self.card_entry = tk.Entry(row3, font=('Arial', 9), width=15)
        self.card_entry.pack(side=tk.RIGHT, padx=5)
        self.card_entry.bind('<Return>', self.on_card_entry_return)
        
        card_btn = tk.Button(
            row3,
            text="עדכן כרטיס",
            command=self.update_card,
            font=('Arial', 9),
            bg='#e74c3c',
            fg='white',
            padx=10,
            pady=3,
            cursor='hand2'
        )
        card_btn.pack(side=tk.RIGHT, padx=5)
        
        photo_label = tk.Label(
            row3,
            text=fix_rtl_text("נתיב תמונה:"),
            font=('Arial', 9),
            bg='#ecf0f1'
        )
        photo_label.pack(side=tk.RIGHT, padx=10)

        self.photo_entry = tk.Entry(row3, font=('Arial', 9), width=20, justify='right')
        self.photo_entry.pack(side=tk.RIGHT, padx=5)
        self.photo_entry.bind('<Return>', self.on_photo_entry_return)

        photo_btn = tk.Button(
            row3,
            text="עדכן תמונה",
            command=self.browse_photo_for_student,
            font=('Arial', 9),
            bg='#2980b9',
            fg='white',
            padx=10,
            pady=3,
            cursor='hand2'
        )
        photo_btn.pack(side=tk.RIGHT, padx=5)

        # הסתרת/נעילת שדות למורים לפי הרשאות
        try:
            if not self._can_current_teacher_edit_student_card():
                try:
                    card_label.pack_forget()
                    self.card_entry.pack_forget()
                    card_btn.pack_forget()
                except Exception:
                    try:
                        self.card_entry.configure(state='disabled')
                        card_btn.configure(state='disabled')
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            if not self._can_current_teacher_edit_student_photo():
                try:
                    photo_label.pack_forget()
                    self.photo_entry.pack_forget()
                    photo_btn.pack_forget()
                except Exception:
                    try:
                        self.photo_entry.configure(state='disabled')
                        photo_btn.configure(state='disabled')
                    except Exception:
                        pass
        except Exception:
            pass
        
        # כפתורי מאסטר - רק למנהלים
        if is_admin:
            # כפתור כרטיס מאסטר
            tk.Button(
                row3,
                text="🔑 מאסטר",
                command=self.update_master_card,
                font=('Arial', 9, 'bold'),
                bg='#8e44ad',
                fg='white',
                padx=10,
                pady=3,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=5)
            
            # כפתור כרטיס מאסטר 2 - ליד מאסטר 1
            tk.Button(
                row3,
                text="🔑 מאסטר 2",
                command=self.update_master_card_2,
                font=('Arial', 9, 'bold'),
                bg='#e67e22',
                fg='white',
                padx=10,
                pady=3,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=5)
        
        # Bind double-click (לבחירה ועריכה מהירה בתא)
        self.tree.bind('<Double-1>', self.on_tree_double_click)

        def _move_selection(delta: int):
            try:
                items = list(self.tree.get_children())
                if not items:
                    return "break"
                sel = self.tree.selection()
                if not sel:
                    if delta > 0:
                        self.tree.selection_set(items[0])
                        self.tree.see(items[0])
                        self.on_student_select(None)
                    else:
                        self.tree.selection_set(items[-1])
                        self.tree.see(items[-1])
                        self.on_student_select(None)
                    return "break"
                cur = sel[0]
                try:
                    idx = items.index(cur)
                except ValueError:
                    idx = 0
                new_idx = max(0, min(len(items) - 1, idx + delta))
                self.tree.selection_set(items[new_idx])
                self.tree.see(items[new_idx])
                self.on_student_select(None)
            except Exception:
                pass
            return "break"

        # ניווט חיצים בטבלת תלמידים
        self.tree.bind('<Down>', lambda e: _move_selection(1))
        self.tree.bind('<Up>', lambda e: _move_selection(-1))

        def _jump_selection(to_end: bool):
            try:
                items = list(self.tree.get_children())
                if not items:
                    return "break"
                target = items[-1] if to_end else items[0]
                self.tree.selection_set(target)
                self.tree.see(target)
                try:
                    self.on_student_select(None)
                except Exception:
                    pass
            except Exception:
                pass
            return "break"

        # ניווט Home/End לראש/סוף הטבלה
        self.tree.bind('<Home>', lambda e: _jump_selection(False))
        self.tree.bind('<End>', lambda e: _jump_selection(True))

    def _attach_entry_edit_menu(self, entry: tk.Entry):
        """הוספת תפריט קליק ימני וקיצורי עריכה לשדה טקסט בודד שורה (Entry)."""
        menu = tk.Menu(entry, tearoff=0)
        menu.add_command(label="גזור", command=lambda: entry.event_generate("<<Cut>>"))
        menu.add_command(label="העתק", command=lambda: entry.event_generate("<<Copy>>"))
        menu.add_command(label="הדבק", command=lambda: entry.event_generate("<<Paste>>"))
        menu.add_separator()
        menu.add_command(label="בחר הכל", command=lambda: (entry.select_range(0, 'end'), entry.icursor('end')))

        def _show_menu(event):
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        # קליק ימני לפתיחת התפריט
        entry.bind("<Button-3>", _show_menu)

        # קיצורי מקשים לעריכה
        entry.bind("<Control-a>", lambda e: (entry.select_range(0, 'end'), entry.icursor('end'), "break"))
        entry.bind("<Control-A>", lambda e: (entry.select_range(0, 'end'), entry.icursor('end'), "break"))
        entry.bind("<Control-c>", lambda e: (entry.event_generate("<<Copy>>"), "break"))
        entry.bind("<Control-C>", lambda e: (entry.event_generate("<<Copy>>"), "break"))
        entry.bind("<Control-v>", lambda e: (entry.event_generate("<<Paste>>"), "break"))
        entry.bind("<Control-V>", lambda e: (entry.event_generate("<<Paste>>"), "break"))
        entry.bind("<Control-x>", lambda e: (entry.event_generate("<<Cut>>"), "break"))
        entry.bind("<Control-X>", lambda e: (entry.event_generate("<<Cut>>"), "break"))

    def _get_selected_student_id(self) -> Optional[int]:
        selection = self.tree.selection()
        if not selection:
            return None
        item_id = selection[0]
        return self.student_ids.get(item_id)

    def add_student_dialog(self):
        """דיאלוג להוספת תלמיד חדש."""
        if not self.ensure_can_modify():
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("הוספת תלמיד")
        dialog.geometry("520x440")
        try:
            dialog.minsize(520, 440)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        def _row(parent, label_text):
            row = tk.Frame(parent, bg='#ecf0f1')
            row.pack(fill=tk.X, pady=4, padx=10)
            tk.Label(row, text=label_text, font=('Arial', 10), bg='#ecf0f1', anchor='e', width=14).pack(side=tk.RIGHT, padx=5)
            entry = tk.Entry(row, font=('Arial', 10), width=24, justify='right')
            entry.pack(side=tk.RIGHT, padx=5)
            return entry

        last_name_entry = _row(dialog, "שם משפחה:")
        first_name_entry = _row(dialog, "שם פרטי:")
        id_entry = _row(dialog, 'ת"ז (לא חובה):')
        class_entry = _row(dialog, "כיתה:")
        card_entry = _row(dialog, "מס' כרטיס:")
        photo_entry = _row(dialog, "מס'/נתיב תמונה:")
        serial_entry = _row(dialog, "מס' סידורי:")
        points_entry = _row(dialog, "נקודות התחלתיות:")
        points_entry.insert(0, "0")

        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=10)

        if not self._can_current_teacher_edit_student_card():
            try:
                card_entry.configure(state='disabled')
            except Exception:
                pass
        if not self._can_current_teacher_edit_student_photo():
            try:
                photo_entry.configure(state='disabled')
            except Exception:
                pass

        def submit():
            last_name = last_name_entry.get().strip()
            first_name = first_name_entry.get().strip()
            if not last_name or not first_name:
                messagebox.showwarning("אזהרה", "יש להזין שם פרטי ושם משפחה")
                return
            id_number = id_entry.get().strip()
            class_name = class_entry.get().strip()
            card = card_entry.get().strip() if self._can_current_teacher_edit_student_card() else ""
            photo = photo_entry.get().strip() if self._can_current_teacher_edit_student_photo() else ""
            serial_val = serial_entry.get().strip()
            points_val = points_entry.get().strip()

            serial_number = None
            if serial_val:
                try:
                    serial_number = int(serial_val)
                except ValueError:
                    messagebox.showwarning("אזהרה", "מס' סידורי חייב להיות מספר שלם")
                    return
            try:
                points = int(points_val) if points_val else 0
            except ValueError:
                messagebox.showwarning("אזהרה", "נקודות חייבות להיות מספר שלם")
                return

            student_id = self.db.add_student(
                last_name=last_name,
                first_name=first_name,
                id_number=id_number,
                class_name=class_name,
                photo_number=photo,
                card_number=card,
                points=points,
                serial_number=serial_number,
            )

            if student_id:
                self.has_changes = True
                self.load_students()
                self.export_to_excel_now()
                self.show_status_message("✓ תלמיד נוסף בהצלחה")
                dialog.destroy()
            else:
                messagebox.showerror("שגיאה", "לא ניתן להוסיף תלמיד חדש")

        tk.Button(
            btn_frame,
            text="שמור",
            command=submit,
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=15,
            pady=5
        ).pack(side=tk.RIGHT, padx=5)

        tk.Button(
            btn_frame,
            text="ביטול",
            command=dialog.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=15,
            pady=5
        ).pack(side=tk.LEFT, padx=5)

    def edit_student_dialog(self):
        """דיאלוג לעריכת פרטי תלמיד קיים."""
        if not self.ensure_can_modify():
            return
        student_id = self._get_selected_student_id()
        if not student_id:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return

        student = self.db.get_student_by_id(student_id)
        if not student:
            messagebox.showerror("שגיאה", "התלמיד לא נמצא במסד הנתונים")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("עריכת תלמיד")
        dialog.geometry("520x420")
        try:
            dialog.minsize(520, 420)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        def _row(parent, label_text, initial=""):
            row = tk.Frame(parent, bg='#ecf0f1')
            row.pack(fill=tk.X, pady=4, padx=10)
            tk.Label(row, text=label_text, font=('Arial', 10), bg='#ecf0f1', anchor='e', width=14).pack(side=tk.RIGHT, padx=5)
            entry = tk.Entry(row, font=('Arial', 10), width=24, justify='right')
            entry.pack(side=tk.RIGHT, padx=5)
            if initial:
                entry.insert(0, initial)
            return entry

        last_name_entry = _row(dialog, "שם משפחה:", student.get('last_name') or "")
        first_name_entry = _row(dialog, "שם פרטי:", student.get('first_name') or "")
        id_entry = _row(dialog, 'ת"ז:', student.get('id_number') or "")
        class_entry = _row(dialog, "כיתה:", student.get('class_name') or "")
        card_entry = _row(dialog, "מס' כרטיס:", student.get('card_number') or "")
        photo_entry = _row(dialog, "מס'/נתיב תמונה:", student.get('photo_number') or "")
        serial_initial = "" if student.get('serial_number') in (None, 0) else str(student.get('serial_number'))
        serial_entry = _row(dialog, "מס' סידורי:", serial_initial)

        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=10)

        if not self._can_current_teacher_edit_student_card():
            try:
                card_entry.configure(state='disabled')
            except Exception:
                pass
        if not self._can_current_teacher_edit_student_photo():
            try:
                photo_entry.configure(state='disabled')
            except Exception:
                pass

        def submit():
            last_name = last_name_entry.get().strip()
            first_name = first_name_entry.get().strip()
            if not last_name or not first_name:
                messagebox.showwarning("אזהרה", "יש להזין שם פרטי ושם משפחה")
                return
            id_number = id_entry.get().strip()
            class_name = class_entry.get().strip()
            card = card_entry.get().strip() if self._can_current_teacher_edit_student_card() else (student.get('card_number') or "")
            photo = photo_entry.get().strip() if self._can_current_teacher_edit_student_photo() else (student.get('photo_number') or "")
            serial_val = serial_entry.get().strip()

            serial_number = None
            if serial_val:
                try:
                    serial_number = int(serial_val)
                except ValueError:
                    messagebox.showwarning("אזהרה", "מס' סידורי חייב להיות מספר שלם")
                    return

            if not self.db.update_student_basic(
                student_id=student_id,
                last_name=last_name,
                first_name=first_name,
                id_number=id_number,
                class_name=class_name,
                card_number=card,
                photo_number=photo,
                serial_number=serial_number,
            ):
                messagebox.showerror("שגיאה", "לא ניתן לעדכן את פרטי התלמיד")
                return

            self.has_changes = True
            self.load_students(keep_selection=True)
            self.export_to_excel_now()
            self.show_status_message("✓ פרטי התלמיד עודכנו בהצלחה")
            dialog.destroy()

        tk.Button(
            btn_frame,
            text="שמור",
            command=submit,
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=15,
            pady=5
        ).pack(side=tk.RIGHT, padx=5)

        tk.Button(
            btn_frame,
            text="ביטול",
            command=dialog.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=15,
            pady=5
        ).pack(side=tk.LEFT, padx=5)

    def delete_student_dialog(self):
        """מחיקת תלמיד נבחר לאחר אישור."""
        if not self.ensure_can_modify():
            return
        student_id = self._get_selected_student_id()
        if not student_id:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return

        student = self.db.get_student_by_id(student_id)
        if not student:
            messagebox.showerror("שגיאה", "התלמיד לא נמצא במסד הנתונים")
            return

        name = f"{student.get('first_name', '')} {student.get('last_name', '')}".strip()
        cls = student.get('class_name') or ""
        msg = f"האם אתה בטוח שברצונך למחוק את התלמיד:\n\n{name} (כיתה {cls})?\n\nהפעולה תמחק גם את היסטוריית הנקודות והתיקופים שלו."
        if not messagebox.askyesno("אישור מחיקה", msg):
            return

        if not self.db.delete_student(student_id):
            messagebox.showerror("שגיאה", "לא ניתן למחוק את התלמיד")
            return

        self.has_changes = True
        self.load_students()
        self.export_to_excel_now()
        self.show_status_message("✓ התלמיד נמחק בהצלחה")
    
    def initial_load(self):
        """טעינה ראשונית - ללא סינכרון Excel (רק קריאה מ-DB)"""
        # שמירת זמן השינוי של Excel (אבל ללא סינכרון)
        if os.path.exists(self.excel_path):
            try:
                self.last_excel_mod_time = os.path.getmtime(self.excel_path)
            except:
                pass
        
        # טעינת תלמידים מייד (ישירות מ-DB - מהיר!)
        self.load_students()
    
    def load_students(self, keep_selection=False):
        """טעינת תלמידים לטבלה - מסונן לפי הרשאות"""
        # אם יש עריכה בתא, בטל אותה לפני רענון הטבלה
        try:
            self.cancel_inline_edit()
        except Exception:
            pass
        # שמירת הבחירה הנוכחית
        selected_student_id = None
        if keep_selection:
            selection = self.tree.selection()
            if selection:
                selected_student_id = self.student_ids.get(selection[0])
        
        # ניקוי הטבלה
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.student_ids.clear()
        
        # טעינת תלמידים - מסונן לפי הרשאות (עם שימוש ב-cache!)
        if self.current_teacher and self.current_teacher['is_admin'] == 0:
            # מורה רגיל - רק תלמידים מהכיתות שלו (משתמש ב-cache!)
            if not self.teacher_classes_cache:
                # אם אין cache עדיין, טען אותו
                self.teacher_classes_cache = self.db.get_teacher_classes(self.current_teacher['id'])
            
            # Check if there's an active class filter
            active_filter = getattr(self, 'teacher_class_filter', None)
            if active_filter:
                # Filter to only selected classes
                classes_to_load = [c for c in self.teacher_classes_cache if c in active_filter]
            else:
                # Load all teacher's classes
                classes_to_load = self.teacher_classes_cache
            
            # טעינה ישירה לפי כיתות (מהיר יותר!)
            conn = self.db.get_connection()
            cursor = conn.cursor()
            if classes_to_load:
                placeholders = ','.join('?' * len(classes_to_load))
                cursor.execute(f'''
                    SELECT * FROM students 
                    WHERE class_name IN ({placeholders})
                    ORDER BY (serial_number IS NULL OR serial_number = 0), serial_number, class_name, last_name, first_name
                ''', classes_to_load)
                rows = cursor.fetchall()
                students = [dict(row) for row in rows]
            else:
                students = []
            conn.close()
        else:
            # מנהל - כל התלמידים
            students = self.db.get_all_students()

        if not students:
            try:
                db_path = getattr(self.db, 'db_path', None)
                teacher_id = self.current_teacher.get('id') if self.current_teacher else None
                safe_print(f"⚠️ אין תלמידים לטעינה. db_path={db_path}, teacher_id={teacher_id}, is_admin={self.current_teacher.get('is_admin') if self.current_teacher else None}")
            except Exception:
                pass
            try:
                self.show_status_message("אין תלמידים לטעינה - בדוק נתיב DB/הרשאות")
            except Exception:
                pass

        # סטטיסטיקות תיקופים עבור כל התלמידים בטבלה
        student_ids_list = [s['id'] for s in students]
        try:
            import time
            now = time.time()
            cache = getattr(self, '_swipe_stats_cache', None)
            cache_ts = float(cache.get('ts', 0.0)) if isinstance(cache, dict) else 0.0
            cache_key = cache.get('key') if isinstance(cache, dict) else None
            cur_key = (len(student_ids_list), student_ids_list[0] if student_ids_list else None, student_ids_list[-1] if student_ids_list else None)
            if isinstance(cache, dict) and cache_key == cur_key and (now - cache_ts) < 60.0:
                swipe_totals = cache.get('swipe_totals') or {}
                total_days = int(cache.get('total_days') or 0)
            else:
                swipe_totals = self.db.get_swipe_totals_for_students(student_ids_list)
                total_days = self.db.get_total_school_days()
                self._swipe_stats_cache = {
                    'ts': now,
                    'key': cur_key,
                    'swipe_totals': swipe_totals,
                    'total_days': total_days,
                }
        except Exception:
            swipe_totals = self.db.get_swipe_totals_for_students(student_ids_list)
            total_days = self.db.get_total_school_days()

        item_to_select = None
        
        for idx, student in enumerate(students):
            # שורות לסירוגין
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            # קיצור הודעה פרטית אם ארוכה
            private_msg = _strip_asterisk_annotations(student.get('private_message', '') or '')
            if len(private_msg) > 30:
                private_msg = private_msg[:27] + '...'
            
            # תיקון RTL להודעה פרטית
            if private_msg:
                private_msg = fix_rtl_text(private_msg)

            # אייקון תמונה
            photo_icon = "📷" if student.get('photo_number') else "✗"

            # מס' סידורי: אם קיים ב-DB השתמש בו, אחרת מספר רץ לפי הסדר
            serial_val = student.get('serial_number')
            serial_display = serial_val if serial_val not in (None, 0) else (idx + 1)

            # סטטיסטיקת תיקופים לתלמיד
            total_swipes = swipe_totals.get(student['id'], 0)
            if total_days > 0 and total_swipes > 0:
                avg_swipes = round(total_swipes / total_days, 2)
            else:
                avg_swipes = 0

            item_id = self.tree.insert('', tk.END, values=(
                student['points'],
                private_msg,
                student['card_number'] if student['card_number'] else '',
                str(student.get('class_name', '') or ''),
                student['id_number'],
                str(student.get('first_name', '') or ''),
                str(student.get('last_name', '') or ''),
                photo_icon,
                serial_display,
                total_swipes,
                avg_swipes
            ), tags=(tag,))
            
            # שמירת מיפוי item_id -> student_id
            self.student_ids[item_id] = student['id']
            
            # שמירת ה-item שצריך לבחור
            if selected_student_id and student['id'] == selected_student_id:
                item_to_select = item_id
        
        # בחירה מחדש של התלמיד
        if item_to_select:
            self.tree.selection_set(item_to_select)
            self.tree.see(item_to_select)

        try:
            self._update_teacher_topbar_stats()
        except Exception:
            pass
    
    def on_search_var_changed(self, *args):
        """תגובה לשינוי בטקסט החיפוש – הפעלה מושהית של החיפוש"""
        try:
            text = self.search_var.get().strip()
        except Exception:
            text = ""

        if not hasattr(self, 'search_typing_job'):
            self.search_typing_job = None
        if not hasattr(self, '_ignore_empty_search_once'):
            self._ignore_empty_search_once = False
        if not hasattr(self, 'search_entry_clear_job'):
            self.search_entry_clear_job = None

        # כל הקלדה חדשה מבטלת ניקוי שדה מתוזמן מהחיפוש הקודם
        if self.search_entry_clear_job is not None:
            try:
                self.root.after_cancel(self.search_entry_clear_job)
            except Exception:
                pass
            self.search_entry_clear_job = None

        # ניקוי שנעשה פנימית אחרי חיפוש – לא להפעיל חיפוש מחדש
        if self._ignore_empty_search_once and text == "":
            self._ignore_empty_search_once = False
            if self.search_typing_job is not None:
                try:
                    self.root.after_cancel(self.search_typing_job)
                except Exception:
                    pass
                self.search_typing_job = None
            return

        # ביטול טיימר הקלדה קודם (אם יש)
        if self.search_typing_job is not None:
            try:
                self.root.after_cancel(self.search_typing_job)
            except Exception:
                pass
            self.search_typing_job = None

        # אם המשתמש מחק הכל ידנית – חזור מיד לרשימה מלאה
        if text == "":
            self.search_students()
            return

        # יש טקסט הקלדה – נפעיל את החיפוש רק אחרי שנייה של שקט
        try:
            self.search_typing_job = self.root.after(1000, self.search_students)
        except Exception:
            self.search_typing_job = None

    def search_students(self):
        """חיפוש תלמידים"""
        search_term = self.search_var.get().strip()

        if not hasattr(self, 'search_clear_job'):
            self.search_clear_job = None
        if not hasattr(self, '_ignore_empty_search_once'):
            self._ignore_empty_search_once = False

        # אם השדה ריק – חזרה לרשימה מלאה, אלא אם כן זה ניקוי פנימי אחרי חיפוש
        if not search_term:
            if self._ignore_empty_search_once:
                # דילוג על רענון – נשאיר את תוצאות החיפוש האחרונות על המסך
                self._ignore_empty_search_once = False
                return

            # אין חיפוש פעיל – בטל טיימר וחזור לרשימה המלאה
            try:
                if self.search_clear_job is not None:
                    self.root.after_cancel(self.search_clear_job)
            except Exception:
                pass
            self.search_clear_job = None
            self.load_students()
            return

        # יש טקסט חיפוש – בטל טיימר קודם אם היה
        if self.search_clear_job is not None:
            try:
                self.root.after_cancel(self.search_clear_job)
            except Exception:
                pass
            self.search_clear_job = None

        # ביטול עריכה בתא לפני רענון התוצאות
        try:
            self.cancel_inline_edit()
        except Exception:
            pass
        
        # ניקוי הטבלה
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.student_ids.clear()
        
        # חיפוש
        students = self.db.search_students(search_term)
        if self.current_teacher and self.current_teacher['is_admin'] == 0:
            if not self.teacher_classes_cache:
                self.teacher_classes_cache = self.db.get_teacher_classes(self.current_teacher['id'])
            if self.teacher_classes_cache:
                allowed_classes = set(self.teacher_classes_cache)
                students = [s for s in students if s.get('class_name') in allowed_classes]
            else:
                students = []

        # סטטיסטיקות תיקופים עבור התלמידים שנמצאו
        student_ids_list = [s['id'] for s in students]
        swipe_totals = self.db.get_swipe_totals_for_students(student_ids_list)
        total_days = self.db.get_total_school_days()

        for idx, student in enumerate(students):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            
            # קיצור הודעה פרטית אם ארוכה
            private_msg = student.get('private_message', '') or ''
            if len(private_msg) > 30:
                private_msg = private_msg[:27] + '...'
            
            # תיקון RTL להודעה פרטית
            if private_msg:
                private_msg = fix_rtl_text(private_msg)

            # אייקון תמונה
            photo_icon = "📷" if student.get('photo_number') else "✗"

            # מס' סידורי: אם קיים ב-DB השתמש בו, אחרת מספר רץ לפי הסדר
            serial_val = student.get('serial_number')
            serial_display = serial_val if serial_val not in (None, 0) else (idx + 1)

            # סטטיסטיקת תיקופים לתלמיד
            total_swipes = swipe_totals.get(student['id'], 0)
            if total_days > 0 and total_swipes > 0:
                avg_swipes = round(total_swipes / total_days, 2)
            else:
                avg_swipes = 0

            item_id = self.tree.insert('', tk.END, values=(
                student['points'],
                private_msg,
                student['card_number'] if student['card_number'] else '',
                _strip_asterisk_annotations(student.get('class_name', '') or ''),
                student['id_number'],
                _strip_asterisk_annotations(student.get('first_name', '') or ''),
                _strip_asterisk_annotations(student.get('last_name', '') or ''),
                photo_icon,
                serial_display,
                total_swipes,
                avg_swipes
            ), tags=(tag,))

            self.student_ids[item_id] = student['id']
        
        # ניקוי שדה החיפוש לאחר השהייה קצרה, כדי שתוכל לראות לרגע את הטקסט שהוקלד
        def _clear_search_entry():
            try:
                self._ignore_empty_search_once = True
                self.search_var.set("")
                if hasattr(self, 'search_entry'):
                    self.search_entry.focus_set()
            except Exception:
                self._ignore_empty_search_once = False
            self.search_entry_clear_job = None

        try:
            if not hasattr(self, 'search_entry_clear_job'):
                self.search_entry_clear_job = None
            # בטל ניקוי קודם אם עדיין לא רץ
            if self.search_entry_clear_job is not None:
                try:
                    self.root.after_cancel(self.search_entry_clear_job)
                except Exception:
                    pass
            # ננקה את השדה שנייה אחרי הצגת התוצאות
            self.search_entry_clear_job = self.root.after(1000, _clear_search_entry)
        except Exception:
            self.search_entry_clear_job = None

        # השארת תוצאות החיפוש על המסך ל-30 שניות נוספות, ואז חזרה לרשימה המלאה
        def _revert_results():
            try:
                # אם בינתיים נקבע טיימר חדש (חיפוש נוסף) – אל נעשה כלום
                if getattr(self, 'search_clear_job', None) != job_id:
                    return
            except Exception:
                return
            self.search_clear_job = None
            self.load_students()

        try:
            job_id = self.root.after(30000, _revert_results)
            self.search_clear_job = job_id
        except Exception:
            self.search_clear_job = None

    def sort_by_column(self, col, reverse=False):
        """מיון הטבלה לפי עמודה (לחיצה חוזרת מחליפה כיוון)"""
        # אחרי מיון המשתמש צריך לראות את הסדר החדש לפני שרענון אוטומטי ידרוס אותו
        try:
            import time
            self._suppress_auto_refresh_until = time.time() + 10.0
        except Exception:
            pass
        # איסוף ערכים לכל שורה
        data = [(self.tree.set(item_id, col), item_id) for item_id in self.tree.get_children('')]
        
        # מיון מספרי לעמודות כמותיות (נקודות, מס' סידורי, תיקופים, ממוצע)
        if col in ('נקודות', 'מס\' סידורי', 'סה"כ תיקופים', 'ממוצע תיקופים ליום'):
            def key_func(t):
                val = t[0]
                try:
                    return float(val)
                except (TypeError, ValueError):
                    return 0.0
        else:
            key_func = lambda t: t[0]
        
        data.sort(key=key_func, reverse=reverse)
        
        # שינוי סדר הרשומות בטבלה לפי המיון
        for index, (_, item_id) in enumerate(data):
            self.tree.move(item_id, '', index)
        
        # בלחיצה הבאה על אותה כותרת נהפוך את כיוון המיון
        self.tree.heading(col, command=lambda c=col: self.sort_by_column(c, not reverse))

    def on_student_select(self, event):
        """טיפול בבחירת תלמיד"""
        # אל תרענן מיד את הטבלה כשיש אינטראקציה עם המשתמש (בחירה/מיון)
        try:
            import time
            self._suppress_auto_refresh_until = time.time() + 10.0
        except Exception:
            pass
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            values = item['values']
            # values: (נקודות, הודעה פרטית, מס' כרטיס, כיתה, ת"ז, שם פרטי, שם משפחה)
            
            # מילוי השדות
            self.card_entry.delete(0, tk.END)
            self.card_entry.insert(0, values[2])  # מס' כרטיס
            
            self.on_update_type_changed(values=values)
            
            # מילוי הודעה פרטית (מהדאטה בייס, לא מהתצוגה המקוצרת)
            student_id = self.student_ids.get(selection[0])
            if student_id:
                student = self.db.get_student_by_id(student_id)
                self.message_entry.delete(0, tk.END)
                if student and student.get('private_message'):
                    self.message_entry.insert(0, _strip_asterisk_annotations(student['private_message']))
                if hasattr(self, 'photo_entry'):
                    self.photo_entry.delete(0, tk.END)
                    if student and student.get('photo_number'):
                        self.photo_entry.insert(0, student['photo_number'])

    def on_update_type_changed(self, values=None):
        if values is None:
            selection = self.tree.selection()
            if not selection:
                return
            item = self.tree.item(selection[0])
            values = item['values']
            if not values:
                return

        update_type = self.update_type.get() if hasattr(self, 'update_type') else "add"
        self.points_entry.delete(0, tk.END)
        if update_type == "set":
            self.points_entry.insert(0, values[0])

    def on_tree_double_click(self, event):
        """עריכה ישירה בתא בטבלה (נקודות / הודעה פרטית)"""
        # זיהוי שורה ועמודה
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item_id:
            return
        # מיפוי מספר עמודה לתעודת העמודה הלוגית (שם העמודה)
        col_id = None
        try:
            # identify_column מחזיר '#N' לפי סדר התצוגה, לכן נשתמש ב-displaycolumns
            idx = int(column.replace('#', '')) - 1
            display_cols = self.tree['displaycolumns'] or self.tree['columns']
            if 0 <= idx < len(display_cols):
                col_id = display_cols[idx]
        except Exception:
            col_id = None

        # בחירת השורה ומילוי השדות התחתונים
        self.tree.selection_set(item_id)
        self.tree.see(item_id)
        self.on_student_select(None)

        # עריכה רק בעמודות "נקודות" ו"הודעה פרטית" (ללא קשר למיקומן על המסך)
        if col_id in ('נקודות', 'הודעה פרטית'):
            self.start_inline_edit(item_id, col_id)

    def start_inline_edit(self, item_id, column):
        """פתיחת שדה עריכה על תא בטבלה"""
        # סגירת עריכה קודמת אם יש
        self.cancel_inline_edit()
        bbox = self.tree.bbox(item_id, column)
        if not bbox:
            return
        x, y, width, height = bbox
        # יצירת שדה עריכה
        justify = 'center' if column == 'נקודות' else 'right'
        editor = tk.Entry(self.tree, font=('Arial', 9), justify=justify)
        editor.place(x=x, y=y, width=width, height=height)
        # תפריט קליק ימני וקיצורי עריכה גם לשדה ההודעה הפרטית בתוך הטבלה
        if column == 'הודעה פרטית':
            self._attach_entry_edit_menu(editor)
        # ערך התחלתי
        values = list(self.tree.item(item_id, 'values'))
        initial = ""
        if column == 'נקודות':
            # עמודת נקודות - הערך נמצא באינדקס 0 ב-values
            initial = str(values[0])
        elif column == 'הודעה פרטית':
            student_id = self.student_ids.get(item_id)
            if student_id:
                student = self.db.get_student_by_id(student_id)
                if student and student.get('private_message'):
                    initial = _strip_asterisk_annotations(student['private_message'])
        editor.insert(0, initial)
        editor.focus_set()
        editor.select_range(0, tk.END)
        # שמירת מצב
        self.inline_editor = editor
        self.inline_edit_item = item_id
        self.inline_edit_column = column
        # קישורי מקלדת
        editor.bind('<Return>', lambda e: self.commit_inline_edit(1, 0))
        editor.bind('<Shift-Return>', lambda e: self.commit_inline_edit(-1, 0))
        editor.bind('<Tab>', lambda e: self.commit_inline_edit(0, 1))
        editor.bind('<Shift-Tab>', lambda e: self.commit_inline_edit(0, -1))
        editor.bind('<ISO_Left_Tab>', lambda e: self.commit_inline_edit(0, -1))
        editor.bind('<Escape>', lambda e: self.cancel_inline_edit())
        editor.bind('<FocusOut>', lambda e: self.cancel_inline_edit())

    def cancel_inline_edit(self):
        """ביטול עריכה ישירה אם קיימת"""
        if hasattr(self, 'inline_editor') and self.inline_editor is not None:
            try:
                self.inline_editor.destroy()
            except Exception:
                pass
        self.inline_editor = None
        self.inline_edit_item = None
        self.inline_edit_column = None

    def _get_points_actor_name(self) -> str:
        try:
            name = str((getattr(self, 'current_teacher', None) or {}).get('name') or '').strip()
        except Exception:
            name = ''
        return name or 'מנהל'

    def commit_inline_edit(self, direction: int = 0, horiz: int = 0):
        """שמירת עריכת תא בטבלה והמשך לשורה הבאה/קודמת"""
        if not hasattr(self, 'inline_editor') or self.inline_editor is None:
            return "break"
        editor = self.inline_editor
        item_id = self.inline_edit_item
        column = self.inline_edit_column
        text = editor.get().strip()
        # סגירת העורך לפני רענון
        self.cancel_inline_edit()
        if not item_id or column not in ('נקודות', 'הודעה פרטית'):
            return "break"
        student_id = self.student_ids.get(item_id)
        if not student_id:
            return "break"
        # עדכון DB + עץ
        if column == 'נקודות':
            actor_name = self._get_points_actor_name()
            # נקודות
            try:
                new_points = int(text) if text else 0
            except ValueError:
                self.show_status_message("✗ יש להזין מספר שלם בלבד", '#e74c3c')
                return "break"

            try:
                if not self._check_max_points_policy(proposed_points=int(new_points), parent=self.root):
                    return "break"
            except Exception:
                pass

            old_points = None
            try:
                old_points = int(self.tree.item(item_id, 'values')[0])
            except Exception:
                old_points = None

            if not self.db.update_student_points(student_id, new_points, "עדכון מוחלט (טבלה)", actor_name):
                self.show_status_message("✗ לא ניתן לעדכן את הנקודות", '#e74c3c')
                return "break"
            values = list(self.tree.item(item_id, 'values'))
            values[0] = new_points
            self.tree.item(item_id, values=tuple(values))

            try:
                if old_points is not None:
                    self._push_points_action("טבלה", [{'student_id': student_id, 'old_points': int(old_points), 'new_points': int(new_points)}])
            except Exception:
                pass
        elif column == 'הודעה פרטית':
            # הודעה פרטית
            if not self.db.update_private_message(student_id, text if text else None):
                self.show_status_message("✗ לא ניתן לעדכן את ההודעה", '#e74c3c')
                return "break"
            display_msg = text or ''
            if len(display_msg) > 30:
                display_msg = display_msg[:27] + '...'
            if display_msg:
                display_msg = fix_rtl_text(display_msg)
            values = list(self.tree.item(item_id, 'values'))
            values[1] = display_msg
            self.tree.item(item_id, values=tuple(values))
        # סימון שינוי וייצוא לאקסל
        self.has_changes = True
        self.export_to_excel_now()
        # רענון השדות התחתונים לפי הבחירה הנוכחית
        self.tree.selection_set(item_id)
        self.tree.see(item_id)
        self.on_student_select(None)
        # מעבר לתא הבא אם נדרש (אנכי / אופקי)
        if direction != 0 or horiz != 0:
            items = list(self.tree.get_children())
            try:
                index = items.index(item_id)
            except ValueError:
                return "break"
            next_item = item_id
            next_column = column
            # תזוזה אנכית (Enter / Shift+Enter)
            if direction != 0:
                new_index = index + direction
                if 0 <= new_index < len(items):
                    next_item = items[new_index]
            # תזוזה אופקית (Tab / Shift+Tab)
            if horiz != 0:
                # סדר אופקי: הודעה פרטית → נקודות → הודעה פרטית בשורה הבאה...
                cols = ['הודעה פרטית', 'נקודות']
                try:
                    col_index = cols.index(column)
                except ValueError:
                    col_index = 0
                new_col_index = col_index + horiz
                if new_col_index < 0:
                    # מעבר לעמודה האחרונה בשורה הקודמת
                    if index > 0:
                        next_item = items[index - 1]
                        new_col_index = len(cols) - 1
                    else:
                        new_col_index = 0
                elif new_col_index >= len(cols):
                    # מעבר לעמודה הראשונה בשורה הבאה
                    if index < len(items) - 1:
                        next_item = items[index + 1]
                        new_col_index = 0
                    else:
                        new_col_index = len(cols) - 1
                next_column = cols[new_col_index]
            # פתיחת עריכה על התא הבא
            if next_item and next_column in ('נקודות', 'הודעה פרטית'):
                self.tree.selection_set(next_item)
                self.tree.see(next_item)
                self.on_student_select(None)
                self.start_inline_edit(next_item, next_column)
        return "break"
    
    def clear_fields(self):
        """ניקוי שדות הקלט"""
        self.card_entry.delete(0, tk.END)
        self.points_entry.delete(0, tk.END)
        self.message_entry.delete(0, tk.END)
        if hasattr(self, 'photo_entry'):
            self.photo_entry.delete(0, tk.END)
    
    def show_status_message(self, message, color='#27ae60', duration=3000):
        """הצגת הודעת סטטוס זמנית"""
        self.sync_label.config(text=message, fg=color)
        self.root.after(duration, lambda: self.sync_label.config(text="סטטוס: מוכן", fg='#27ae60'))
    
    def ensure_can_modify(self) -> bool:
        """בדיקת רישיון לפני ביצוע פעולה המשנה נתונים"""
        lm = getattr(self, "license_manager", None)
        if lm is None:
            return True
        if lm.can_modify_data():
            return True
        msg = lm.get_block_modify_message()
        try:
            messagebox.showerror("רישיון אינו פעיל", msg)
        except Exception:
            pass
        return False

    def _can_current_teacher_edit_student_card(self) -> bool:
        try:
            t = getattr(self, 'current_teacher', None) or {}
            if int(t.get('is_admin', 0) or 0) == 1:
                return True
            return bool(int(t.get('can_edit_student_card', 1) or 0) == 1)
        except Exception:
            return True

    def _can_current_teacher_edit_student_photo(self) -> bool:
        try:
            t = getattr(self, 'current_teacher', None) or {}
            if int(t.get('is_admin', 0) or 0) == 1:
                return True
            return bool(int(t.get('can_edit_student_photo', 1) or 0) == 1)
        except Exception:
            return True
    
    def update_card(self):
        """עדכון מספר כרטיס"""
        if not self.ensure_can_modify():
            return False

        if not self._can_current_teacher_edit_student_card():
            messagebox.showerror("אין הרשאה", "אין הרשאה לשנות מספר כרטיס לתלמיד")
            return False
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return False
        
        item_id = selection[0]
        student_id = self.student_ids.get(item_id)
        
        new_card = self.card_entry.get().strip()
        
        if self.db.update_card_number(student_id, new_card if new_card else None):
            self.has_changes = True
            self.show_status_message("✓ מספר הכרטיס עודכן בהצלחה")
            self.card_entry.delete(0, tk.END)  # ניקוי שדה כרטיס
            self.load_students(keep_selection=True)
            self.export_to_excel_now()  # ייצוא מיידי לאקסל
            return True
        else:
            self.show_status_message("✗ לא ניתן לעדכן את מספר הכרטיס", '#e74c3c')
            return False

    def _get_shared_sounds_dir(self) -> str:
        try:
            cfg = self.load_app_config() or {}
        except Exception:
            cfg = {}
        try:
            shared = str((cfg or {}).get('shared_folder') or (cfg or {}).get('network_root') or '').strip()
        except Exception:
            shared = ''
        if shared and os.path.isdir(shared):
            return os.path.join(shared, 'sounds')
        return os.path.join(self.base_dir, 'sounds')

    def _list_sound_keys_in_folder(self, folder_parts) -> list:
        base = os.path.join(self._get_shared_sounds_dir(), *(folder_parts or []))
        if not os.path.isdir(base):
            return []
        sounds = {}
        try:
            for root, _, files in os.walk(base):
                for filename in (files or []):
                    if not str(filename).lower().endswith(('.wav', '.mp3', '.ogg')):
                        continue
                    key = os.path.splitext(filename)[0]
                    if not key:
                        continue
                    path = os.path.join(root, filename)
                    prev = sounds.get(key)
                    if not prev:
                        sounds[key] = path
                        continue
                    try:
                        ext = str(os.path.splitext(path)[1] or '').lower()
                    except Exception:
                        ext = ''
                    try:
                        prev_ext = str(os.path.splitext(prev)[1] or '').lower()
                    except Exception:
                        prev_ext = ''
                    priorities = {'.wav': 30, '.mp3': 20, '.ogg': 10}
                    if int(priorities.get(ext, 0)) > int(priorities.get(prev_ext, 0)):
                        sounds[key] = path
        except Exception:
            sounds = {}
        try:
            return sorted(list(sounds.keys()), key=lambda x: str(x))
        except Exception:
            return list(sounds.keys())

    def _import_sound_file_to_folder(self, folder_parts) -> str:
        try:
            initial_dir = os.path.join(self._get_shared_sounds_dir(), *(folder_parts or []))
        except Exception:
            initial_dir = self.base_dir
        file_path = filedialog.askopenfilename(
            title="בחר קובץ צליל",
            filetypes=[("Sound files", "*.wav;*.mp3;*.ogg"), ("All files", "*.*")],
            initialdir=initial_dir
        )
        if not file_path:
            return ''
        try:
            target_dir = os.path.join(self._get_shared_sounds_dir(), *(folder_parts or []))
            os.makedirs(target_dir, exist_ok=True)
        except Exception:
            return ''

        try:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            ext = os.path.splitext(os.path.basename(file_path))[1]
        except Exception:
            base_name, ext = 'sound', '.wav'
        try:
            safe_base = re.sub(r'[^0-9A-Za-zא-ת _\-]', '', str(base_name)).strip() or 'sound'
        except Exception:
            safe_base = str(base_name).strip() or 'sound'

        # שם ייחודי כדי לא לדרוס
        dest = os.path.join(target_dir, safe_base + ext)
        if os.path.abspath(dest) == os.path.abspath(file_path):
            return os.path.splitext(os.path.basename(dest))[0]
        if os.path.exists(dest):
            import uuid
            dest = os.path.join(target_dir, f"{safe_base}_{uuid.uuid4().hex[:6]}{ext}")
        try:
            shutil.copy2(file_path, dest)
        except Exception:
            return ''
        try:
            return os.path.splitext(os.path.basename(dest))[0]
        except Exception:
            return ''

    def _admin_play_sound_key(self, sound_key: str) -> None:
        try:
            if not SoundManager:
                return
            k = str(sound_key or '').strip()
            if not k:
                return
            mgr = getattr(self, '_admin_sound_manager', None)
            if mgr is None:
                mgr = SoundManager(self.base_dir, sounds_dir=self._get_shared_sounds_dir())
                self._admin_sound_manager = mgr
            path = mgr.resolve_sound([k])
            if path:
                mgr.play_sound(path, async_play=True)
        except Exception:
            return
    
    def update_message(self):
        """עדכון הודעה פרטית"""
        if not self.ensure_can_modify():
            return False
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return False
        
        item_id = selection[0]
        student_id = self.student_ids.get(item_id)
        
        new_message = self.message_entry.get().strip()
        
        if self.db.update_private_message(student_id, new_message if new_message else None):
            self.has_changes = True
            self.show_status_message("✓ ההודעה הפרטית עודכנה בהצלחה")
            self.load_students(keep_selection=True)
            self.export_to_excel_now()  # ייצוא מיידי לאקסל
            return True
        else:
            self.show_status_message("✗ לא ניתן לעדכן את ההודעה", '#e74c3c')
            return False

    def update_photo(self):
        """עדכון מס' תמונה / קובץ תמונה"""
        if not self.ensure_can_modify():
            return False
        if not self._can_current_teacher_edit_student_photo():
            messagebox.showerror("אין הרשאה", "אין הרשאה לשנות תמונת תלמיד")
            return False
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return False
        
        item_id = selection[0]
        student_id = self.student_ids.get(item_id)
        
        new_photo = self.photo_entry.get().strip()
        
        if self.db.update_photo_number(student_id, new_photo if new_photo else None):
            self.has_changes = True
            self.show_status_message("✓ תמונת התלמיד עודכנה בהצלחה")
            self.load_students(keep_selection=True)
            self.export_to_excel_now()  # ייצוא מיידי לאקסל
            return True
        else:
            self.show_status_message("✗ לא ניתן לעדכן את התמונה", '#e74c3c')
            return False

    def update_points(self):
        """עדכון נקודות"""
        if not self.ensure_can_modify():
            return False
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return False
        
        item_id = selection[0]
        student_id = self.student_ids.get(item_id)
        item = self.tree.item(item_id)
        current_points = item['values'][0]  # נקודות הן בעמודה הראשונה עכשיו
        
        try:
            value = int(self.points_entry.get())
            update_type = self.update_type.get()
            
            success = False
            reason = ""
            old_points = None
            try:
                old_points = int(current_points)
            except Exception:
                old_points = None
            new_points = None

            actor_name = self._get_points_actor_name()

            # חישוב הערך הסופי לצורך אכיפה של "מקסימום נקודות"
            try:
                proposed = None
                if update_type == 'set':
                    proposed = int(value)
                elif update_type == 'add' and old_points is not None:
                    proposed = int(old_points) + int(value)
                elif update_type == 'subtract' and old_points is not None:
                    proposed = int(old_points) - int(value)
                if proposed is not None:
                    if not self._check_max_points_policy(proposed_points=int(proposed), parent=self.root):
                        return False
            except Exception:
                pass
            
            if update_type == "set":
                success = self.db.update_student_points(student_id, value, "עדכון מוחלט", actor_name)
                reason = f"עדכון ל-{value}"
                try:
                    new_points = int(value)
                except Exception:
                    new_points = None
            elif update_type == "add":
                success = self.db.add_points(student_id, value, "הוספת נקודות", actor_name)
                reason = f"הוספת {value}"
                try:
                    if old_points is not None:
                        new_points = int(old_points) + int(value)
                except Exception:
                    new_points = None
            elif update_type == "subtract":
                success = self.db.subtract_points(student_id, value, "חיסור נקודות", actor_name)
                reason = f"חיסור {value}"
                try:
                    if old_points is not None:
                        new_points = int(old_points) - int(value)
                except Exception:
                    new_points = None
            
            if success:
                try:
                    if old_points is not None and new_points is not None:
                        self._push_points_action(
                            "שדה נקודות",
                            [{'student_id': student_id, 'old_points': int(old_points), 'new_points': int(max(0, int(new_points)))}]
                        )
                except Exception:
                    pass
                self.has_changes = True
                self.show_status_message(f"✓ הנקודות עודכנו בהצלחה - {reason}")
                self.points_entry.delete(0, tk.END)  # ניקוי שדה נקודות
                self.load_students(keep_selection=True)
                self.export_to_excel_now()  # ייצוא מיידי לאקסל
                return True
            else:
                self.show_status_message("✗ לא ניתן לעדכן את הנקודות", '#e74c3c')
                return False
        except Exception as e:
            self.show_status_message(f"✗ שגיאה: {e}", '#e74c3c')
            return False

    def open_manual_swipe_override(self):
        if not self.ensure_can_modify():
            return

        try:
            selection = self.tree.selection()
        except Exception:
            selection = ()
        if not selection:
            messagebox.showwarning('אזהרה', 'יש לבחור תלמיד מהרשימה')
            return
        try:
            item_id = selection[0]
            student_id = self.student_ids.get(item_id)
        except Exception:
            student_id = None
        if not student_id:
            messagebox.showwarning('אזהרה', 'יש לבחור תלמיד מהרשימה')
            return

        try:
            student = self.db.get_student_by_id(int(student_id)) or {}
        except Exception:
            student = {}

        dlg = tk.Toplevel(self.root)
        dlg.title('תיקון זמן תיקוף')
        dlg.configure(bg='#ecf0f1')
        dlg.transient(self.root)
        dlg.grab_set()
        try:
            dlg.resizable(False, False)
        except Exception:
            pass

        try:
            ln = str(student.get('last_name') or '').strip()
            fn = str(student.get('first_name') or '').strip()
            full_name = f"{ln} {fn}".strip()
        except Exception:
            full_name = ''

        tk.Label(dlg, text=fix_rtl_text('תיקון זמן תיקוף (ידני)'), font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50', anchor='e').pack(fill=tk.X, padx=14, pady=(12, 6))
        if full_name:
            tk.Label(dlg, text=fix_rtl_text(full_name), font=('Arial', 10), bg='#ecf0f1', fg='#7f8c8d', anchor='e').pack(fill=tk.X, padx=14, pady=(0, 8))

        body = tk.Frame(dlg, bg='#ecf0f1')
        body.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)

        date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        time_var = tk.StringVar(value=datetime.now().strftime('%H:%M'))

        def _open_date_picker():
            picker = tk.Toplevel(dlg)
            picker.title('בחר תאריך')
            picker.transient(dlg)
            picker.grab_set()
            picker.resizable(False, False)

            pf = tk.Frame(picker, padx=12, pady=10)
            pf.pack(fill=tk.BOTH, expand=True)

            tk.Label(pf, text=fix_rtl_text('בחר תאריך'), font=('Arial', 10, 'bold')).pack(anchor='e', pady=(0, 8))

            cur = str(date_var.get() or '').strip()
            yy0, mm0, dd0 = None, None, None
            try:
                parts = cur.split('-')
                if len(parts) == 3:
                    yy0 = int(parts[0]); mm0 = int(parts[1]); dd0 = int(parts[2])
            except Exception:
                yy0, mm0, dd0 = None, None, None

            try:
                now = datetime.now()
                if yy0 is None:
                    yy0 = int(now.year)
                if mm0 is None:
                    mm0 = int(now.month)
                if dd0 is None:
                    dd0 = int(now.day)
            except Exception:
                yy0, mm0, dd0 = 2026, 1, 1

            years = [str(y) for y in range(int(yy0) - 2, int(yy0) + 3)]
            months = [f"{m:02d}" for m in range(1, 13)]
            days = [f"{d:02d}" for d in range(1, 32)]

            row = tk.Frame(pf)
            row.pack(fill=tk.X)

            y_var = tk.StringVar(value=str(yy0))
            m_var = tk.StringVar(value=f"{int(mm0):02d}")
            d_var = tk.StringVar(value=f"{int(dd0):02d}")

            ttk.Combobox(row, textvariable=d_var, values=days, state='readonly', width=4, justify='right').pack(side=tk.RIGHT, padx=4)
            ttk.Combobox(row, textvariable=m_var, values=months, state='readonly', width=4, justify='right').pack(side=tk.RIGHT, padx=4)
            ttk.Combobox(row, textvariable=y_var, values=years, state='readonly', width=6, justify='right').pack(side=tk.RIGHT, padx=4)

            def _ok():
                yy = str(y_var.get() or '').strip()
                mm = str(m_var.get() or '').strip()
                dd = str(d_var.get() or '').strip()
                try:
                    dt = datetime.strptime(f"{yy}-{mm}-{dd}", '%Y-%m-%d')
                    date_var.set(dt.strftime('%Y-%m-%d'))
                except Exception:
                    pass
                picker.destroy()

            def _cancel():
                picker.destroy()

            btns_p = tk.Frame(pf)
            btns_p.pack(fill=tk.X, pady=(10, 0))
            tk.Button(btns_p, text='אישור', width=10, command=_ok).pack(side=tk.LEFT, padx=4)
            tk.Button(btns_p, text='ביטול', width=10, command=_cancel).pack(side=tk.RIGHT, padx=4)

        def _open_time_picker():
            picker = tk.Toplevel(dlg)
            picker.title('בחר שעה')
            picker.transient(dlg)
            picker.grab_set()
            picker.resizable(False, False)

            pf = tk.Frame(picker, padx=12, pady=10)
            pf.pack(fill=tk.BOTH, expand=True)

            tk.Label(pf, text=fix_rtl_text('בחר שעה'), font=('Arial', 10, 'bold')).pack(anchor='e', pady=(0, 8))

            cur = str(time_var.get() or '').strip()
            hh0, mm0 = None, None
            try:
                parts = cur.split(':')
                if len(parts) == 2:
                    hh0 = int(parts[0]); mm0 = int(parts[1])
            except Exception:
                hh0, mm0 = None, None

            try:
                now = datetime.now()
                if hh0 is None:
                    hh0 = int(now.hour)
                if mm0 is None:
                    mm0 = int(now.minute)
            except Exception:
                hh0, mm0 = 8, 0

            hours = [f"{h:02d}" for h in range(0, 24)]
            mins = [f"{m:02d}" for m in range(0, 60, 1)]

            row = tk.Frame(pf)
            row.pack(fill=tk.X)

            h_var = tk.StringVar(value=f"{int(hh0):02d}")
            m_var = tk.StringVar(value=f"{int(mm0):02d}")

            ttk.Combobox(row, textvariable=m_var, values=mins, state='readonly', width=4, justify='right').pack(side=tk.RIGHT, padx=4)
            ttk.Combobox(row, textvariable=h_var, values=hours, state='readonly', width=4, justify='right').pack(side=tk.RIGHT, padx=4)

            def _ok():
                hh = str(h_var.get() or '').strip()
                mm = str(m_var.get() or '').strip()
                try:
                    dt = datetime.strptime(f"{hh}:{mm}", '%H:%M')
                    time_var.set(dt.strftime('%H:%M'))
                except Exception:
                    pass
                picker.destroy()

            def _cancel():
                picker.destroy()

            btns_p = tk.Frame(pf)
            btns_p.pack(fill=tk.X, pady=(10, 0))
            tk.Button(btns_p, text='אישור', width=10, command=_ok).pack(side=tk.LEFT, padx=4)
            tk.Button(btns_p, text='ביטול', width=10, command=_cancel).pack(side=tk.RIGHT, padx=4)

        r1 = tk.Frame(body, bg='#ecf0f1')
        r1.pack(fill=tk.X, pady=6)
        tk.Label(r1, text=fix_rtl_text('תאריך (YYYY-MM-DD):'), bg='#ecf0f1', width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
        date_ent = tk.Entry(r1, textvariable=date_var, font=('Arial', 10), justify='right', width=14)
        date_ent.pack(side=tk.RIGHT, padx=6)
        try:
            date_ent.bind('<Button-1>', lambda _e: _open_date_picker())
        except Exception:
            pass

        r2 = tk.Frame(body, bg='#ecf0f1')
        r2.pack(fill=tk.X, pady=6)
        tk.Label(r2, text=fix_rtl_text('שעה (HH:MM):'), bg='#ecf0f1', width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
        time_ent = tk.Entry(r2, textvariable=time_var, font=('Arial', 10), justify='right', width=14)
        time_ent.pack(side=tk.RIGHT, padx=6)
        try:
            time_ent.bind('<Button-1>', lambda _e: _open_time_picker())
        except Exception:
            pass

        tk.Label(body, text=fix_rtl_text('הפעולה תרשום תיקוף לזמן שנבחר ותחשב בונוס זמנים אם רלוונטי.'), bg='#ecf0f1', fg='#2c3e50', anchor='e', justify='right').pack(fill=tk.X, pady=(8, 0))

        def _apply():
            d = str(date_var.get() or '').strip()
            t = str(time_var.get() or '').strip()
            if not d or not t:
                messagebox.showerror('שגיאה', 'יש להזין תאריך ושעה', parent=dlg)
                return
            try:
                dt_obj = datetime.strptime(f"{d} {t}", '%Y-%m-%d %H:%M')
            except Exception:
                messagebox.showerror('שגיאה', 'תאריך/שעה לא תקינים. פורמט: YYYY-MM-DD ו-HH:MM', parent=dlg)
                return

            swiped_at = dt_obj.strftime('%Y-%m-%d %H:%M:%S')

            # Export expects time_bonus_given.given_at to be UTC (historical behavior).
            try:
                from datetime import timezone
                local_tz = datetime.now().astimezone().tzinfo
                dt_local = dt_obj.replace(tzinfo=local_tz)
                given_at_utc = dt_local.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                given_at_utc = swiped_at
            try:
                card_number = str(student.get('card_number') or '').strip()
            except Exception:
                card_number = ''
            try:
                if not self.db.upsert_first_swipe_for_date(int(student_id), swiped_at=swiped_at, card_number=card_number, station_type='public'):
                    messagebox.showerror('שגיאה', 'לא ניתן לרשום תיקוף', parent=dlg)
                    return
            except Exception as e:
                messagebox.showerror('שגיאה', str(e), parent=dlg)
                return

            try:
                class_name = str(student.get('class_name') or '').strip()
            except Exception:
                class_name = ''

            try:
                time_bonus = self.db.get_active_time_bonus_now(class_name=class_name, now_dt=dt_obj)
            except Exception:
                time_bonus = None

            if time_bonus:
                try:
                    bonus_schedule_id = int(time_bonus.get('id') or 0)
                except Exception:
                    bonus_schedule_id = 0
                bonus_group = ''
                try:
                    bonus_group = str(time_bonus.get('group_name') or time_bonus.get('name') or '').strip()
                except Exception:
                    bonus_group = ''
                try:
                    bonus_points = int(time_bonus.get('bonus_points', 0) or 0)
                except Exception:
                    bonus_points = 0

                existing_bonus = None
                try:
                    if bonus_group:
                        existing_bonus = self.db.get_student_time_bonus_for_group_on_date(int(student_id), bonus_group, d)
                except Exception:
                    existing_bonus = None

                existing_points = 0
                try:
                    if existing_bonus is not None:
                        existing_points = int(existing_bonus.get('bonus_points') or 0)
                except Exception:
                    existing_points = 0

                if bonus_schedule_id:
                    if existing_bonus is None:
                        if bonus_points > 0:
                            try:
                                bonus_name = str(time_bonus.get('name') or '').strip()
                            except Exception:
                                bonus_name = ''
                            try:
                                self.db.add_points(int(student_id), int(bonus_points), f"⏰ בונוס זמנים ({bonus_name}): +{int(bonus_points)}", "תיקון ידני")
                            except Exception:
                                pass
                        try:
                            self.db.record_time_bonus_given_on_date(int(student_id), int(bonus_schedule_id), d, given_at=given_at_utc)
                        except Exception:
                            pass
                    else:
                        if int(bonus_points) > int(existing_points):
                            try:
                                delta = int(bonus_points) - int(existing_points)
                            except Exception:
                                delta = 0
                            if delta > 0:
                                try:
                                    bonus_name = str(time_bonus.get('name') or '').strip()
                                except Exception:
                                    bonus_name = ''
                                try:
                                    self.db.add_points(int(student_id), int(delta), f"⏰ תיקון בונוס זמנים ({bonus_name}): +{int(delta)}", "תיקון ידני")
                                except Exception:
                                    pass
                            try:
                                if bonus_group:
                                    self.db.replace_student_time_bonus_for_group_on_date(int(student_id), str(bonus_group), str(d), int(bonus_schedule_id), given_at=given_at_utc)
                            except Exception:
                                pass
                        else:
                            try:
                                if bonus_group:
                                    self.db.update_time_bonus_given_at_for_group_on_date(int(student_id), str(bonus_group), str(d), given_at=given_at_utc)
                            except Exception:
                                pass



            try:
                self.load_students(keep_selection=True)
                self.export_to_excel_now()
            except Exception:
                pass
            try:
                self.show_status_message('✓ נרשם תיקוף ידני', '#27ae60')
            except Exception:
                pass
            dlg.destroy()

        btns = tk.Frame(body, bg='#ecf0f1')
        btns.pack(fill=tk.X, pady=(12, 0))
        tk.Button(btns, text='אישור', command=_apply, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='ביטול', command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)

    def _select_relative_student(self, step: int):
        selection = self.tree.selection()
        items = list(self.tree.get_children())
        if not items:
            return
        if selection:
            try:
                index = items.index(selection[0])
            except ValueError:
                index = -1
        else:
            index = -1
        new_index = index + step
        if 0 <= new_index < len(items):
            new_item = items[new_index]
            self.tree.selection_set(new_item)
            self.tree.see(new_item)
            self.on_student_select(None)

    def _entry_nav_move(self, step: int, widget: tk.Widget):
        """ניווט תלמידים עם החיצים גם כשהפוקוס בשדות למטה (נקודות/הודעה)."""
        try:
            self._select_relative_student(int(step))
        except Exception:
            pass
        try:
            widget.focus_set()
        except Exception:
            pass
        return "break"

    def select_next_student(self):
        self._select_relative_student(1)

    def on_points_entry_return(self, event):
        try:
            txt = str(self.points_entry.get() or '').strip()
        except Exception:
            txt = ''
        if txt == '':
            self.select_next_student()
            try:
                self.points_entry.focus_set()
            except Exception:
                pass
            return "break"
        if self.update_points():
            self.select_next_student()
            self.points_entry.focus_set()
        return "break"

    def on_message_entry_return(self, event):
        try:
            txt = str(self.message_entry.get() or '').strip()
        except Exception:
            txt = ''
        if txt == '':
            self.select_next_student()
            try:
                self.message_entry.focus_set()
            except Exception:
                pass
            return "break"
        if self.update_message():
            self.select_next_student()
            self.message_entry.focus_set()
        return "break"

    def select_previous_student(self):
        self._select_relative_student(-1)

    def on_points_entry_shift_return(self, event):
        try:
            txt = str(self.points_entry.get() or '').strip()
        except Exception:
            txt = ''
        if txt == '':
            self.select_previous_student()
            try:
                self.points_entry.focus_set()
            except Exception:
                pass
            return "break"
        if self.update_points():
            self.select_previous_student()
            self.points_entry.focus_set()
        return "break"

    def on_message_entry_shift_return(self, event):
        try:
            txt = str(self.message_entry.get() or '').strip()
        except Exception:
            txt = ''
        if txt == '':
            self.select_previous_student()
            try:
                self.message_entry.focus_set()
            except Exception:
                pass
            return "break"
        if self.update_message():
            self.select_previous_student()
            self.message_entry.focus_set()
        return "break"

    def on_points_entry_tab(self, event):
        """Tab משדה נקודות → לתלמיד הבא בשדה הודעה פרטית"""
        if self.update_points():
            self.select_next_student()
            self.message_entry.focus_set()
        return "break"

    def on_message_entry_tab(self, event):
        """Tab משדה הודעה פרטית → לשדה נקודות באותו תלמיד"""
        if self.update_message():
            self.points_entry.focus_set()
        return "break"

    def on_message_entry_shift_tab(self, event):
        """Shift+Tab מהודעה פרטית → חזרה לשדה נקודות באותו תלמיד"""
        self.points_entry.focus_set()
        return "break"

    def on_card_entry_return(self, event):
        if self.update_card():
            self.select_next_student()
            self.card_entry.focus_set()
        return "break"

    def on_photo_entry_return(self, event):
        if self.update_photo():
            self.photo_entry.focus_set()
        return "break"

    def browse_photo_for_student(self):
        """פתיחת סייר קבצים לבחירת תמונת תלמיד ועדכון במסד הנתונים"""
        if not self.ensure_can_modify():
            return False
        if not self._can_current_teacher_edit_student_photo():
            messagebox.showerror("אין הרשאה", "אין הרשאה לשנות תמונת תלמיד")
            return False
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "יש לבחור תלמיד מהרשימה")
            return False

        item_id = selection[0]
        student_id = self.student_ids.get(item_id)
        if not student_id:
            return False

        cfg = self.load_app_config()
        photos_folder = cfg.get('photos_folder', '') or ''

        initial_dir = photos_folder or self.base_dir
        current_val = self.photo_entry.get().strip()
        try:
            if current_val:
                if os.path.isabs(current_val):
                    initial_dir = os.path.dirname(current_val)
                elif photos_folder:
                    initial_dir = os.path.join(photos_folder, os.path.dirname(current_val) or '')
        except Exception:
            pass

        file_path = filedialog.askopenfilename(
            title="בחר תמונה לתלמיד",
            filetypes=[("קבצי תמונה", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("הכל", "*.*")],
            initialdir=initial_dir
        )

        if not file_path:
            return False

        store_value = file_path
        try:
            if photos_folder:
                photos_abs = os.path.abspath(photos_folder)
                file_abs = os.path.abspath(file_path)
                if os.path.commonprefix([photos_abs, file_abs]) == photos_abs:
                    store_value = os.path.relpath(file_abs, photos_abs)
        except Exception:
            store_value = file_path

        self.photo_entry.delete(0, tk.END)
        self.photo_entry.insert(0, store_value)

        return self.update_photo()

    def import_excel(self):
        """ייבוא קובץ Excel"""
        if not self.ensure_can_modify():
            return
        # תיקייה התחלתית - שם שקובץ האקסל נמצא
        excel_dir = os.path.dirname(self.excel_path)
        file_path = filedialog.askopenfilename(
            title="בחר קובץ Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=excel_dir
        )
        
        if not file_path:
            return

        # שאלה האם למחוק נתונים קיימים (כולל איפוס תיקופים והיסטוריית נקודות)
        clear = messagebox.askyesno(
            "ייבוא מ-Excel",
            "האם למחוק את כל התלמידים וההיסטוריה שלהם לפני הייבוא?\n" \
            "הפעולה תאפס גם את כל התיקופים והיסטוריית הנקודות."
        )

        try:
            imported, errors = self.importer.import_from_excel(file_path, clear)

            if errors:
                error_msg = "\n".join(errors[:10])
                messagebox.showwarning(
                    "ייבוא הושלם עם שגיאות",
                    f"יובאו {imported} תלמידים.\n\nשגיאות:\n{error_msg}"
                )
            else:
                messagebox.showinfo(
                    "הצלחה",
                    f"יובאו {imported} תלמידים בהצלחה!"
                )

            self.has_changes = True
            self.load_students()

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייבוא הקובץ:\n{str(e)}")
    
    def ask_export_options(self):
        return self._ask_export_options_impl()
    
    def _ask_export_options_impl(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("אפשרויות ייצוא")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        mode_var = tk.StringVar(value="regular")
        date_var = tk.StringVar(value=date.today().strftime("%d.%m.%Y"))
        bonus_var = tk.StringVar(value="בחר בונוס")
        student_var = tk.StringVar(value="בחר תלמיד")
        teacher_var = tk.StringVar(value="בחר מורה")
        output_type_var = tk.StringVar(value="preview")

        bonuses = []
        try:
            bonuses = self.db.get_all_time_bonuses()
        except Exception as e:
            safe_print(f"שגיאה בשליפת בונוסי זמנים: {e}")
            bonuses = []

        bonus_choices = []
        bonus_id_by_label = {}
        try:
            groups = self.db.get_time_bonus_groups()
        except Exception:
            groups = {}
            for b in bonuses:
                g = (b.get('group_name') or b.get('name') or '').strip() or f"בונוס {b.get('id', '')}".strip()
                groups.setdefault(g, []).append(b)

        for group_name, items in groups.items():
            label = str(group_name or '').strip()
            if not label:
                continue
            bonus_choices.append(label)
            try:
                bonus_id_by_label[label] = int(items[0].get('id'))
            except Exception:
                pass

        combo_values = ["בחר בונוס"] + bonus_choices if bonus_choices else ["אין בונוסים"]

        # מורים (לייצוא פעולות)
        try:
            is_admin_user = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin_user = False

        teachers = []
        try:
            teachers = self.db.get_all_teachers() or []
        except Exception:
            teachers = []
        teacher_choices = []
        teacher_id_by_label = {}

        if is_admin_user:
            teacher_choices.append("כל המורים")
            teacher_id_by_label["כל המורים"] = -1
        for t in (teachers or []):
            try:
                lbl = str(t.get('name') or '').strip()
            except Exception:
                lbl = ''
            if not lbl:
                continue
            teacher_choices.append(lbl)
            try:
                teacher_id_by_label[lbl] = int(t.get('id') or 0)
            except Exception:
                pass

        if not is_admin_user:
            try:
                my_name = str((self.current_teacher or {}).get('name') or '').strip()
            except Exception:
                my_name = ''
            if my_name:
                teacher_var.set(my_name)
                try:
                    if my_name not in teacher_id_by_label:
                        teacher_id_by_label[my_name] = int((self.current_teacher or {}).get('id') or 0)
                except Exception:
                    pass
                teacher_choices = [my_name]

        dialog.configure(bg='#f0f0f0')
        container = tk.Frame(dialog, padx=12, pady=10, bg='#f0f0f0')
        container.pack(fill=tk.BOTH, expand=True)

        title_label = tk.Label(container, text="בחר סוג ייצוא", font=('Arial', 11, 'bold'), bg='#f0f0f0')
        title_label.pack(anchor='e', pady=(0, 6))

        # Output type selection (file or preview)
        output_frame = tk.LabelFrame(container, text="יעד הייצוא", bg='#f0f0f0')
        output_frame.pack(fill=tk.X, pady=(0, 10))

        rb_file = tk.Radiobutton(
            output_frame,
            text="💾 ייצוא לקובץ Excel",
            variable=output_type_var,
            value="file",
            bg='#f0f0f0',
            font=('Arial', 10)
        )
        rb_file.pack(anchor='w', padx=10, pady=2)

        rb_preview = tk.Radiobutton(
            output_frame,
            text="👁 דוח זמני (תצוגה בלבד)",
            variable=output_type_var,
            value="preview",
            bg='#f0f0f0',
            font=('Arial', 10)
        )
        rb_preview.pack(anchor='w', padx=10, pady=2)

        mode_frame = tk.LabelFrame(container, text="מצב ייצוא", bg='#f0f0f0')
        mode_frame.pack(fill=tk.X, pady=(0, 10))

        rb_regular = tk.Radiobutton(
            mode_frame,
            text="ייצוא תלמידים רגיל",
            variable=mode_var,
            value="regular",
            bg='#f0f0f0'
        )
        rb_regular.pack(anchor='w')

        rb_day = tk.Radiobutton(
            mode_frame,
            text="נוכחות ליום מסוים",
            variable=mode_var,
            value="day",
            bg='#f0f0f0'
        )
        rb_day.pack(anchor='w')

        rb_bonus = tk.Radiobutton(
            mode_frame,
            text="נוכחות לפי בונוס נבחר (כל התאריכים)",
            variable=mode_var,
            value="bonus",
            bg='#f0f0f0'
        )
        rb_bonus.pack(anchor='w')

        rb_all = tk.Radiobutton(
            mode_frame,
            text="כל הבונוסים וכל ההיסטוריה (גליון נפרד לכל בונוס)",
            variable=mode_var,
            value="all_bonuses",
            bg='#f0f0f0'
        )
        rb_all.pack(anchor='w')

        rb_points_history = tk.Radiobutton(
            mode_frame,
            text="היסטוריית נקודות לתלמיד",
            variable=mode_var,
            value="points_history",
            bg='#f0f0f0'
        )
        rb_points_history.pack(anchor='w')

        rb_daily_points_summary = tk.Radiobutton(
            mode_frame,
            text='תשקיף יומי נקודות (סיכום יומי לכל תלמיד)',
            variable=mode_var,
            value='daily_points_summary',
            bg='#f0f0f0'
        )
        rb_daily_points_summary.pack(anchor='w')

        rb_teacher_actions = tk.Radiobutton(
            mode_frame,
            text="פעולות לפי מורה",
            variable=mode_var,
            value="teacher_actions",
            bg='#f0f0f0'
        )
        rb_teacher_actions.pack(anchor='w')

        details_frame = tk.LabelFrame(container, text="פרטי ייצוא", bg='#f0f0f0')
        details_frame.pack(fill=tk.X, pady=(0, 8))

        date_frame = tk.Frame(details_frame, bg='#f0f0f0')
        date_frame.pack(fill=tk.X, pady=(4, 4))

        date_label = tk.Label(date_frame, text="תאריך (DD.MM.YYYY):", bg='#f0f0f0')
        date_label.pack(side=tk.RIGHT)

        date_entry = tk.Entry(date_frame, textvariable=date_var, width=12, justify='right')
        date_entry.pack(side=tk.RIGHT, padx=(0, 5))

        bonus_frame = tk.Frame(details_frame, bg='#f0f0f0')

        bonus_label = tk.Label(bonus_frame, text="בונוס זמנים:", bg='#f0f0f0')
        bonus_label.pack(side=tk.RIGHT)

        combo_state = 'readonly' if bonus_choices else 'disabled'
        bonus_combo = ttk.Combobox(
            bonus_frame,
            textvariable=bonus_var,
            values=combo_values,
            state=combo_state,
            width=30,
            justify='right'
        )
        bonus_combo.pack(side=tk.RIGHT, padx=(0, 5))
        bonus_combo.set("בחר בונוס" if bonus_choices else "אין בונוסים")

        # בחירת תלמיד (להיסטוריית נקודות)
        student_frame = tk.Frame(details_frame, bg='#f0f0f0')

        # בחירת מורה (לפעולות לפי מורה)
        teacher_frame = tk.Frame(details_frame, bg='#f0f0f0')
        tk.Label(teacher_frame, text="מורה:", bg='#f0f0f0').pack(side=tk.RIGHT)
        teacher_combo = ttk.Combobox(
            teacher_frame,
            textvariable=teacher_var,
            values=(teacher_choices if (not is_admin_user) else (["בחר מורה"] + teacher_choices)) if teacher_choices else ["אין מורים"],
            state=('disabled' if (not is_admin_user) else ('readonly' if teacher_choices else 'disabled')),
            width=30,
            justify='right'
        )
        teacher_combo.pack(side=tk.RIGHT, padx=(0, 5))
        if is_admin_user:
            teacher_combo.set("בחר מורה" if teacher_choices else "אין מורים")
        else:
            try:
                if teacher_choices:
                    teacher_combo.set(teacher_choices[0])
                else:
                    teacher_combo.set("אין מורים")
            except Exception:
                pass

        # שורה 1: חיפוש לפי מס' סידורי / ID
        student_find_frame = tk.Frame(student_frame, bg='#f0f0f0')
        student_find_frame.pack(fill=tk.X, pady=(0, 4))

        student_find_label = tk.Label(student_find_frame, text="מס' סידורי/ID:", bg='#f0f0f0')
        student_find_label.pack(side=tk.RIGHT)

        student_find_var = tk.StringVar(value="")
        student_find_entry = tk.Entry(student_find_frame, textvariable=student_find_var, width=10, justify='right')
        student_find_entry.pack(side=tk.RIGHT, padx=(0, 5))

        find_btn = tk.Button(student_find_frame, text="אישור")
        find_btn.pack(side=tk.RIGHT, padx=(0, 5))

        # שורה 2: בחירה בשני שלבים - כיתה ואז תלמיד
        student_pick_frame = tk.Frame(student_frame, bg='#f0f0f0')
        student_pick_frame.pack(fill=tk.X)

        class_var = tk.StringVar(value="כל הכיתות")
        class_label = tk.Label(student_pick_frame, text="כיתה:", bg='#f0f0f0')
        class_label.pack(side=tk.RIGHT)

        # רשימת תלמידים – מנהל: כולם; מורה: רק כיתות מורשות
        students_for_export = []
        try:
            is_admin = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False

        allowed_classes = None
        if not is_admin:
            try:
                allowed_classes = (self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id']) or [])
            except Exception:
                allowed_classes = []

        # טעינת תלמידים לפי הרשאות
        try:
            import time
            now = time.time()
            cache = getattr(self, '_export_students_cache', None)
            cache_ts = float(cache.get('ts', 0.0)) if isinstance(cache, dict) else 0.0
            cache_key = cache.get('key') if isinstance(cache, dict) else None
            teacher_id = None
            try:
                teacher_id = int(self.current_teacher.get('id')) if self.current_teacher else None
            except Exception:
                teacher_id = None
            allowed_key = tuple(allowed_classes) if allowed_classes else None
            cur_key = (teacher_id, bool(is_admin), allowed_key)
            if isinstance(cache, dict) and cache_key == cur_key and (now - cache_ts) < 30.0:
                students_for_export = cache.get('students') or []
            else:
                students_for_export = self.db.get_all_students() or []
                self._export_students_cache = {
                    'ts': now,
                    'key': cur_key,
                    'students': students_for_export,
                }
        except Exception:
            try:
                students_for_export = self.db.get_all_students() or []
            except Exception:
                students_for_export = []

        if allowed_classes:
            try:
                allowed_set = set(allowed_classes)
                students_for_export = [s for s in students_for_export if (s.get('class_name') in allowed_set)]
            except Exception:
                pass

        # קומבובוקס כיתה
        try:
            class_choices = ["כל הכיתות"] + sorted({(s.get('class_name') or '').strip() for s in students_for_export if (s.get('class_name') or '').strip()})
        except Exception:
            class_choices = ["כל הכיתות"]

        class_combo = ttk.Combobox(
            student_pick_frame,
            textvariable=class_var,
            values=class_choices,
            state='readonly',
            width=12,
            justify='right'
        )
        class_combo.pack(side=tk.RIGHT, padx=(0, 5))
        class_combo.set("כל הכיתות")

        # קומבובוקס תלמיד
        student_label = tk.Label(student_pick_frame, text="תלמיד:", bg='#f0f0f0')
        student_label.pack(side=tk.RIGHT)

        student_combo = ttk.Combobox(
            student_pick_frame,
            textvariable=student_var,
            values=["בחר תלמיד"],
            state='readonly',
            width=28,
            justify='right'
        )
        student_combo.pack(side=tk.RIGHT, padx=(0, 5))
        student_combo.set("בחר תלמיד")

        student_id_by_label = {}

        def _make_student_label(s: dict) -> str:
            try:
                sid = s.get('id')
            except Exception:
                sid = ''
            try:
                serial = s.get('serial_number')
            except Exception:
                serial = ''
            cls = (s.get('class_name') or '').strip()
            name = f"{(s.get('first_name') or '').strip()} {(s.get('last_name') or '').strip()}".strip()
            parts = []
            if serial:
                parts.append(str(serial))
            elif sid:
                parts.append(str(sid))
            if cls:
                parts.append(cls)
            if name:
                parts.append(name)
            return " - ".join([p for p in parts if p]) or str(sid or '')

        def _refresh_student_choices(*_args):
            try:
                sel_cls = (class_var.get() or '').strip()
            except Exception:
                sel_cls = ''
            filtered = students_for_export
            if sel_cls and sel_cls != "כל הכיתות":
                filtered = [s for s in students_for_export if (s.get('class_name') or '').strip() == sel_cls]

            labels = []
            student_id_by_label.clear()
            for s in filtered:
                lbl = _make_student_label(s)
                if not lbl:
                    continue
                labels.append(lbl)
                try:
                    student_id_by_label[lbl] = int(s.get('id'))
                except Exception:
                    pass

            if labels:
                student_combo.configure(values=["בחר תלמיד"] + labels, state='readonly')
                if student_var.get() not in student_id_by_label:
                    student_combo.set("בחר תלמיד")
            else:
                student_combo.configure(values=["אין תלמידים"], state='disabled')
                student_combo.set("אין תלמידים")

        class_combo.bind('<<ComboboxSelected>>', _refresh_student_choices)
        _refresh_student_choices()

        def _select_student_in_widgets(stu: dict) -> bool:
            if not isinstance(stu, dict) or not stu:
                return False
            try:
                cls = (stu.get('class_name') or '').strip()
            except Exception:
                cls = ''
            try:
                if cls:
                    class_combo.set(cls)
                else:
                    class_combo.set("כל הכיתות")
            except Exception:
                pass
            _refresh_student_choices()
            try:
                target_id = int(stu.get('id') or 0)
            except Exception:
                target_id = 0
            if not target_id:
                return False
            try:
                for lbl, sid in student_id_by_label.items():
                    try:
                        if int(sid) == int(target_id):
                            student_combo.set(lbl)
                            return True
                    except Exception:
                        continue
            except Exception:
                pass
            return False

        def _find_student_by_serial_or_id(val: str):
            raw = str(val or '').strip()
            if not raw:
                return None
            try:
                n = int(raw)
            except Exception:
                return None
            try:
                s = next((x for x in (students_for_export or []) if int(x.get('serial_number') or 0) == n), None)
                if s:
                    return s
            except Exception:
                pass
            try:
                s = next((x for x in (students_for_export or []) if int(x.get('id') or 0) == n), None)
                if s:
                    return s
            except Exception:
                pass
            try:
                return self.db.get_student_by_id(int(n))
            except Exception:
                return None

        def _confirm_student_find(*_args):
            s = _find_student_by_serial_or_id(student_find_var.get())
            if not s:
                messagebox.showerror("שגיאה", "לא נמצא תלמיד לפי המספר שהוזן")
                return
            if not _select_student_in_widgets(s):
                messagebox.showerror("שגיאה", "לא ניתן לבחור את התלמיד ברשימה")
                return

        try:
            find_btn.configure(command=_confirm_student_find)
        except Exception:
            pass
        try:
            student_find_entry.bind('<Return>', _confirm_student_find)
        except Exception:
            pass

        # הצגה/הסתרה לפי מצב
        def _set_visibility(*_args):
            m = mode_var.get()
            try:
                date_frame.pack_forget()
            except Exception:
                pass
            try:
                bonus_frame.pack_forget()
            except Exception:
                pass
            try:
                student_frame.pack_forget()
            except Exception:
                pass
            try:
                teacher_frame.pack_forget()
            except Exception:
                pass

            if m == 'day':
                date_frame.pack(fill=tk.X, pady=(4, 4))
            elif m == 'bonus':
                bonus_frame.pack(fill=tk.X, pady=(4, 4))
            elif m == 'points_history':
                student_frame.pack(fill=tk.X, pady=(4, 4))
            elif m == 'teacher_actions':
                teacher_frame.pack(fill=tk.X, pady=(4, 4))

            # התאמת גובה החלון כדי שלא "יחתוך" את פרטי התלמיד/בונוס
            try:
                dialog.update_idletasks()
                req_w = int(dialog.winfo_reqwidth() or 520)
                req_h = int(dialog.winfo_reqheight() or 260)
                cur_w = int(dialog.winfo_width() or 0)
                cur_h = int(dialog.winfo_height() or 0)
                w = max(520, req_w, cur_w)
                h = max(260, req_h, cur_h)
                dialog.minsize(520, 260)
                dialog.geometry(f'{w}x{h}')
            except Exception:
                pass

        mode_var.trace_add('write', _set_visibility)
        _set_visibility()

        result = {'mode': None, 'date_iso': None, 'bonus_id': None, 'student_id': None, 'teacher_id': None, 'output_type': None}

        def _parse_date_iso():
            try:
                txt = (date_var.get() or '').strip()
            except Exception:
                txt = ''
            if not txt:
                return None
            try:
                dt = datetime.strptime(txt, '%d.%m.%Y').date()
                return dt.isoformat()
            except Exception:
                return None

        def _resolve_bonus_id():
            try:
                lbl = (bonus_var.get() or '').strip()
            except Exception:
                lbl = ''
            if not lbl or lbl in ("בחר בונוס", "אין בונוסים"):
                return None
            try:
                return int(bonus_id_by_label.get(lbl))
            except Exception:
                return None

        def _resolve_student_id():
            raw = (student_find_var.get() or '').strip()
            if raw:
                try:
                    sid = int(raw)
                    st = self.db.get_student_by_id(sid)
                    if st:
                        return int(st.get('id'))
                except Exception:
                    pass
            try:
                lbl = (student_var.get() or '').strip()
            except Exception:
                lbl = ''
            if lbl and lbl in student_id_by_label:
                try:
                    return int(student_id_by_label[lbl])
                except Exception:
                    return None
            return None

        def _resolve_teacher_id() -> int:
            if not is_admin_user:
                try:
                    return int((self.current_teacher or {}).get('id') or 0)
                except Exception:
                    return 0

            try:
                lbl = (teacher_var.get() or '').strip()
            except Exception:
                lbl = ''
            if not lbl or lbl in ("בחר מורה", "אין מורים"):
                return 0
            if lbl == "כל המורים":
                return -1
            try:
                return int(teacher_id_by_label.get(lbl) or 0)
            except Exception:
                return 0

        def submit():
            m = mode_var.get()
            if m == 'day':
                date_iso = _parse_date_iso()
                if not date_iso:
                    messagebox.showerror("שגיאה", "תאריך לא תקין. יש להזין בפורמט DD.MM.YYYY")
                    return
                result['mode'] = m
                result['date_iso'] = date_iso
            elif m == 'bonus':
                bid = _resolve_bonus_id()
                if not bid:
                    messagebox.showerror("שגיאה", "יש לבחור בונוס")
                    return
                result['mode'] = m
                result['bonus_id'] = bid
            elif m == 'points_history':
                sid = _resolve_student_id()
                if not sid:
                    messagebox.showerror("שגיאה", "יש לבחור תלמיד")
                    return
                result['mode'] = m
                result['student_id'] = sid
            elif m == 'teacher_actions':
                tid = _resolve_teacher_id()
                if not tid:
                    messagebox.showerror("שגיאה", "יש לבחור מורה")
                    return
                result['mode'] = m
                result['teacher_id'] = tid
            else:
                result['mode'] = m
            result['output_type'] = output_type_var.get()
            dialog.destroy()

        btns = tk.Frame(container, bg='#f0f0f0')
        btns.pack(fill=tk.X, pady=(8, 0))
        tk.Button(btns, text="ביטול", command=dialog.destroy).pack(side=tk.RIGHT, padx=6)
        tk.Button(btns, text="אישור", command=submit).pack(side=tk.RIGHT)

        try:
            dialog.update_idletasks()
            w = max(520, int(dialog.winfo_reqwidth() or 520))
            h = max(260, int(dialog.winfo_reqheight() or 260))
            sw = int(dialog.winfo_screenwidth() or 1360)
            sh = int(dialog.winfo_screenheight() or 760)
            x = max(0, (sw // 2) - (w // 2))
            y = max(0, (sh // 2) - (h // 2))
            dialog.geometry(f'{w}x{h}+{x}+{y}')
        except Exception:
            pass

        self.root.wait_window(dialog)
        if result['mode'] is None:
            return None, None, None, None, None, None
        return result['mode'], result['date_iso'], result['bonus_id'], result['student_id'], result['teacher_id'], result['output_type']
    
    def _ask_export_options_old2(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("אפשרויות ייצוא")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        mode_var = tk.StringVar(value="regular")
        date_var = tk.StringVar(value=date.today().strftime("%d.%m.%Y"))
        bonus_var = tk.StringVar(value="בחר בונוס")
        student_var = tk.StringVar(value="בחר תלמיד")

        bonuses = []
        try:
            bonuses = self.db.get_all_time_bonuses()
        except Exception as e:
            safe_print(f"שגיאה בשליפת בונוסי זמנים: {e}")
            bonuses = []

        bonus_choices = []
        bonus_id_by_label = {}
        try:
            groups = self.db.get_time_bonus_groups()
        except Exception:
            groups = {}
            for b in bonuses:
                g = (b.get('group_name') or b.get('name') or '').strip() or f"בונוס {b.get('id', '')}".strip()
                groups.setdefault(g, []).append(b)

        for group_name, items in groups.items():
            label = str(group_name or '').strip()
            if not label:
                continue
            bonus_choices.append(label)
            try:
                bonus_id_by_label[label] = int(items[0].get('id'))
            except Exception:
                pass

        combo_values = ["בחר בונוס"] + bonus_choices if bonus_choices else ["אין בונוסים"]

        dialog.configure(bg='#f0f0f0')
        container = tk.Frame(dialog, padx=12, pady=10, bg='#f0f0f0')
        container.pack(fill=tk.BOTH, expand=True)

        title_label = tk.Label(container, text="בחר סוג ייצוא", font=('Arial', 11, 'bold'), bg='#f0f0f0')
        title_label.pack(anchor='e', pady=(0, 6))

        mode_frame = tk.LabelFrame(container, text="מצב ייצוא", bg='#f0f0f0')
        mode_frame.pack(fill=tk.X, pady=(0, 10))

        rb_regular = tk.Radiobutton(
            mode_frame,
            text="ייצוא תלמידים רגיל",
            variable=mode_var,
            value="regular",
            bg='#f0f0f0'
        )
        rb_regular.pack(anchor='w')

        rb_day = tk.Radiobutton(
            mode_frame,
            text="נוכחות ליום מסוים",
            variable=mode_var,
            value="day",
            bg='#f0f0f0'
        )
        rb_day.pack(anchor='w')

        rb_bonus = tk.Radiobutton(
            mode_frame,
            text="נוכחות לפי בונוס נבחר (כל התאריכים)",
            variable=mode_var,
            value="bonus",
            bg='#f0f0f0'
        )
        rb_bonus.pack(anchor='w')

        rb_all = tk.Radiobutton(
            mode_frame,
            text="כל הבונוסים וכל ההיסטוריה (גליון נפרד לכל בונוס)",
            variable=mode_var,
            value="all_bonuses",
            bg='#f0f0f0'
        )
        rb_all.pack(anchor='w')

        rb_points_history = tk.Radiobutton(
            mode_frame,
            text="היסטוריית נקודות לתלמיד",
            variable=mode_var,
            value="points_history",
            bg='#f0f0f0'
        )
        rb_points_history.pack(anchor='w')

        details_frame = tk.LabelFrame(container, text="פרטי ייצוא", bg='#f0f0f0')
        details_frame.pack(fill=tk.X, pady=(0, 8))

        date_frame = tk.Frame(details_frame, bg='#f0f0f0')
        date_frame.pack(fill=tk.X, pady=(4, 4))

        date_label = tk.Label(date_frame, text="תאריך (DD.MM.YYYY):", bg='#f0f0f0')
        date_label.pack(side=tk.RIGHT)

        date_entry = tk.Entry(date_frame, textvariable=date_var, width=12, justify='right')
        date_entry.pack(side=tk.RIGHT, padx=(0, 5))

        bonus_frame = tk.Frame(details_frame, bg='#f0f0f0')

        bonus_label = tk.Label(bonus_frame, text="בונוס זמנים:", bg='#f0f0f0')
        bonus_label.pack(side=tk.RIGHT)

        combo_state = 'readonly' if bonus_choices else 'disabled'
        bonus_combo = ttk.Combobox(
            bonus_frame,
            textvariable=bonus_var,
            values=combo_values,
            state=combo_state,
            width=30,
            justify='right'
        )
        bonus_combo.pack(side=tk.RIGHT, padx=(0, 5))
        bonus_combo.set("בחר בונוס" if bonus_choices else "אין בונוסים")

        # בחירת תלמיד (להיסטוריית נקודות)
        student_frame = tk.Frame(details_frame, bg='#f0f0f0')

        # שורה 1: חיפוש לפי מס' סידורי / ID
        student_find_frame = tk.Frame(student_frame, bg='#f0f0f0')
        student_find_frame.pack(fill=tk.X, pady=(0, 4))

        student_find_label = tk.Label(student_find_frame, text="מס' סידורי/ID:", bg='#f0f0f0')
        student_find_label.pack(side=tk.RIGHT)

        student_find_var = tk.StringVar(value="")
        student_find_entry = tk.Entry(student_find_frame, textvariable=student_find_var, width=10, justify='right')
        student_find_entry.pack(side=tk.RIGHT, padx=(0, 5))

        # שורה 2: בחירה בשני שלבים - כיתה ואז תלמיד
        student_pick_frame = tk.Frame(student_frame, bg='#f0f0f0')
        student_pick_frame.pack(fill=tk.X)

        class_var = tk.StringVar(value="כל הכיתות")
        class_label = tk.Label(student_pick_frame, text="כיתה:", bg='#f0f0f0')
        class_label.pack(side=tk.RIGHT)

        # רשימת תלמידים – מנהל: כולם; מורה: רק כיתות מורשות
        students_for_export = []
        try:
            is_admin = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False

        allowed_classes = None
        if not is_admin:
            try:
                allowed_classes = (self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id']) or [])
            except Exception:
                allowed_classes = []
    
    def export_excel(self):
        """ייצוא לקובץ Excel – תלמידים רגיל או נוכחות (בונוס זמנים) מאותו חלון."""
        if not self.ensure_can_modify():
            return

        # הגבלת מורה – כיתות מורשות (לכל סוגי הייצוא)
        try:
            is_admin = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False

        allowed_classes = None
        if not is_admin:
            try:
                allowed_classes = (self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id']) or [])
            except Exception:
                allowed_classes = []

        mode, target_date_iso, bonus_id, student_id, teacher_id, output_type = self.ask_export_options()
        if mode is None:
            return

        def _safe_filename_part(s: str) -> str:
            try:
                out = str(s or '').strip()
            except Exception:
                out = ''
            if not out:
                return ''
            for ch in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
                out = out.replace(ch, ' ')
            out = ' '.join(out.split())
            return out[:60]

        student_name_for_file = ''
        if mode == "points_history" and student_id:
            try:
                s = self.db.get_student_by_id(int(student_id))
                if s:
                    student_name_for_file = _safe_filename_part(
                        f"{(s.get('first_name') or '').strip()} {(s.get('last_name') or '').strip()}".strip()
                    )
            except Exception:
                student_name_for_file = ''

        bonus_name_for_file = ''
        if mode == "bonus" and bonus_id:
            try:
                bonuses = self.db.get_all_time_bonuses() or []
                b = next((x for x in bonuses if int(x.get('id', 0) or 0) == int(bonus_id)), None)
                if b:
                    bonus_name_for_file = _safe_filename_part((b.get('group_name') or b.get('name') or '').strip())
            except Exception:
                bonus_name_for_file = ''

        teacher_name_for_file = ''
        if mode == 'teacher_actions' and teacher_id:
            try:
                t = self.db.get_teacher_by_id(int(teacher_id))
                if t:
                    teacher_name_for_file = _safe_filename_part((t.get('name') or '').strip())
            except Exception:
                teacher_name_for_file = ''

        excel_dir = os.path.dirname(self.excel_path)
        if mode == "regular":
            initial_name = "טבלה למבצע אשראי.xlsx"
        elif mode == "day":
            # אם יש תאריך – נשתמש בו בשם הקובץ, אחרת שם כללי
            try:
                if target_date_iso:
                    y, m, d = target_date_iso[0:4], target_date_iso[5:7], target_date_iso[8:10]
                    initial_name = f"נוכחות ליום {d}.{m}.{y}.xlsx"
                else:
                    initial_name = "נוכחות ליום מסוים.xlsx"
            except Exception:
                initial_name = "נוכחות ליום מסוים.xlsx"
        elif mode == "bonus":
            base = "נוכחות לפי בונוס"
            if bonus_name_for_file:
                base = f"{base} - {bonus_name_for_file}"
            initial_name = f"{base}.xlsx"
        elif mode == "points_history":
            base = "היסטוריית נקודות"
            if student_name_for_file:
                base = f"{base} - {student_name_for_file}"
            initial_name = f"{base}.xlsx"
        elif mode == "daily_points_summary":
            initial_name = "תשקיף יומי נקודות.xlsx"
        elif mode == "teacher_actions":
            base = "פעולות מורה"
            if teacher_id == -1:
                base = f"{base} - כל המורים"
            elif teacher_name_for_file:
                base = f"{base} - {teacher_name_for_file}"
            initial_name = f"{base}.xlsx"
        else:  # all_bonuses
            initial_name = "נוכחות - כל הבונוסים.xlsx"

        # Handle preview vs file export
        is_preview = (output_type == "preview")
        if is_preview:
            # Preview mode - use temporary file
            import tempfile
            temp_fd, file_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
        else:
            # File mode - ask for save location
            file_path = filedialog.asksaveasfilename(
                title="שמור כקובץ Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=self._get_downloads_dir(),
                initialfile=initial_name
            )
            
            if not file_path:
                return

        if mode == "regular":
            success = self.importer.export_to_excel(file_path, allowed_classes=allowed_classes)
        elif mode == "points_history":
            if not student_id:
                messagebox.showerror("שגיאה", "לא נבחר תלמיד לייצוא היסטוריה.")
                return
            student = None
            try:
                student = self.db.get_student_by_id(int(student_id))
            except Exception:
                student = None
            if not student:
                messagebox.showerror("שגיאה", "תלמיד לא נמצא במסד הנתונים.")
                return

            # הגבלת מורה – בדיקה נוספת ליתר בטיחות
            if not is_admin:
                try:
                    allowed = set(allowed_classes or [])
                    if allowed and (student.get('class_name') not in allowed):
                        messagebox.showerror("שגיאה", "אין הרשאה לייצא היסטוריה עבור תלמיד מכיתה זו.")
                        return
                except Exception:
                    pass

            logs = []
            try:
                logs = self.db.get_points_log_for_student(int(student_id))
            except Exception:
                logs = []
            success = self.importer.export_points_log_excel(file_path, student, logs)
        elif mode == "daily_points_summary":
            success = self.importer.export_daily_points_summary_excel(file_path, allowed_classes=allowed_classes)
        elif mode == "teacher_actions":
            if not teacher_id:
                messagebox.showerror("שגיאה", "לא נבחר מורה לייצוא פעולות.")
                return

            if int(teacher_id) == -1:
                try:
                    import pandas as pd
                except Exception:
                    messagebox.showerror('שגיאה', 'חסר רכיב pandas לייצוא')
                    return

                def _df_from_actions(data_rows):
                    try:
                        df_local = pd.DataFrame(data_rows or [])
                    except Exception:
                        df_local = pd.DataFrame([])

                    preferred_cols_local = [
                        'created_at',
                        'serial_number',
                        'class_name',
                        'first_name',
                        'last_name',
                        'old_points',
                        'new_points',
                        'delta',
                        'action_type',
                        'reason',
                    ]
                    cols_local = [c for c in preferred_cols_local if c in df_local.columns]
                    extra_cols_local = [c for c in df_local.columns if c not in cols_local]
                    df_local = df_local[cols_local + extra_cols_local] if (cols_local or extra_cols_local) else df_local

                    col_map_local = {
                        'created_at': 'תאריך',
                        'serial_number': "מס' סידורי",
                        'class_name': 'כיתה',
                        'first_name': 'שם פרטי',
                        'last_name': 'שם משפחה',
                        'old_points': 'נקודות לפני',
                        'new_points': 'נקודות אחרי',
                        'delta': 'שינוי',
                        'action_type': 'סוג פעולה',
                        'reason': 'סיבה',
                    }
                    try:
                        df_local = df_local.rename(columns={k: v for k, v in col_map_local.items() if k in df_local.columns})
                    except Exception:
                        pass
                    return df_local

                try:
                    teachers_all = self.db.get_all_teachers() or []
                except Exception:
                    teachers_all = []

                def _safe_sheet_name(name: str) -> str:
                    base_name = str(name or '').strip() or 'Sheet'
                    for ch in [':', '\\', '/', '?', '*', '[', ']']:
                        base_name = base_name.replace(ch, ' ')
                    base_name = ' '.join(base_name.split())
                    return base_name[:31] or 'Sheet'

                used_names = set()
                try:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        wrote_any = False
                        for trow in (teachers_all or []):
                            actor_name = ''
                            try:
                                actor_name = str(trow.get('name') or '').strip()
                            except Exception:
                                actor_name = ''
                            if not actor_name:
                                continue
                            try:
                                data_rows = self.db.get_points_log_for_actor(actor_name) or []
                            except Exception:
                                data_rows = []
                            df_one = _df_from_actions(data_rows)
                            sheet = _safe_sheet_name(actor_name)
                            if sheet in used_names:
                                i = 2
                                while True:
                                    candidate = _safe_sheet_name(f"{sheet} {i}")
                                    if candidate not in used_names:
                                        sheet = candidate
                                        break
                                    i += 1
                            used_names.add(sheet)
                            df_one.to_excel(writer, index=False, sheet_name=sheet)
                            wrote_any = True
                        if not wrote_any:
                            _df_from_actions([]).to_excel(writer, index=False, sheet_name='פעולות')

                    try:
                        from openpyxl import load_workbook
                        from excel_styling import apply_rtl_and_alternating_colors
                        wb = load_workbook(file_path)
                        for ws in wb.worksheets:
                            try:
                                apply_rtl_and_alternating_colors(ws, has_header=True)
                            except Exception:
                                pass
                        wb.save(file_path)
                    except Exception:
                        pass

                    success = True
                except Exception as e:
                    messagebox.showerror('שגיאה', f'שגיאה בייצוא פעולות מורה (כל המורים):\n{str(e)}')
                    success = False
            else:

                teacher = None
                try:
                    teacher = self.db.get_teacher_by_id(int(teacher_id))
                except Exception:
                    teacher = None
                if not teacher:
                    messagebox.showerror("שגיאה", "מורה לא נמצא במסד הנתונים.")
                    return

                try:
                    actor_name = str(teacher.get('name') or '').strip()
                except Exception:
                    actor_name = ''
                if not actor_name:
                    messagebox.showerror("שגיאה", "שם מורה לא תקין.")
                    return

                data = []
                try:
                    data = self.db.get_points_log_for_actor(actor_name) or []
                except Exception:
                    data = []

                try:
                    import pandas as pd
                except Exception:
                    messagebox.showerror('שגיאה', 'חסר רכיב pandas לייצוא')
                    return

                try:
                    df = pd.DataFrame(data)
                except Exception:
                    df = pd.DataFrame([])

                # סדר עמודות לייצוא
                preferred_cols = [
                    'created_at',
                    'serial_number',
                    'class_name',
                    'first_name',
                    'last_name',
                    'old_points',
                    'new_points',
                    'delta',
                    'action_type',
                    'reason',
                ]
                cols = [c for c in preferred_cols if c in df.columns]
                extra_cols = [c for c in df.columns if c not in cols]
                df = df[cols + extra_cols] if (cols or extra_cols) else df

                # שמות עמודות בעברית
                col_map = {
                    'created_at': 'תאריך',
                    'serial_number': "מס' סידורי",
                    'class_name': 'כיתה',
                    'first_name': 'שם פרטי',
                    'last_name': 'שם משפחה',
                    'old_points': 'נקודות לפני',
                    'new_points': 'נקודות אחרי',
                    'delta': 'שינוי',
                    'action_type': 'סוג פעולה',
                    'reason': 'סיבה',
                }
                try:
                    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
                except Exception:
                    pass

                try:
                    df.to_excel(file_path, index=False)
                    try:
                        from openpyxl import load_workbook
                        from excel_styling import apply_rtl_and_alternating_colors
                        wb = load_workbook(file_path)
                        ws = wb.active
                        apply_rtl_and_alternating_colors(ws, has_header=True)
                        wb.save(file_path)
                    except Exception:
                        pass
                    success = True
                except Exception as e:
                    messagebox.showerror('שגיאה', f'שגיאה בייצוא פעולות מורה:\n{str(e)}')
                    success = False
        else:
            success = self.importer.export_attendance_excel(
                file_path,
                mode=mode,
                target_date_iso=target_date_iso,
                bonus_id=bonus_id,
                allowed_classes=allowed_classes
            )

        if success and mode == "regular" and not is_preview:
            # ייצוא תלמידים רגיל – עדכון נתיב קובץ האקסל הפעיל ושמירתו בהגדרות היישום
            self.excel_path = file_path
            try:
                cfg = self.load_app_config()
                if isinstance(cfg, dict):
                    cfg['excel_path'] = file_path
                    self.save_app_config(cfg)
            except Exception as e:
                safe_print(f"⚠️ שגיאה בשמירת מיקום קובץ Excel להגדרות: {e}")

        if success:
            if is_preview:
                # Preview mode - show data in window
                try:
                    import pandas as pd
                    df = pd.read_excel(file_path, sheet_name=0)
                    self._show_preview_window(df, initial_name)
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בקריאת הנתונים:\n{str(e)}")
                finally:
                    # Clean up temporary file
                    try:
                        os.remove(file_path)
                    except Exception:
                        pass
            else:
                self._show_export_success_dialog(file_path)
        else:
            messagebox.showerror("שגיאה", "שגיאה בייצוא הקובץ")
            if is_preview:
                # Clean up temporary file on error
                try:
                    os.remove(file_path)
                except Exception:
                    pass

    def _show_preview_window(self, df, title: str):
        """Display DataFrame in a preview window with Treeview"""
        try:
            preview_dialog = tk.Toplevel(self.root)
            preview_dialog.title('תצוגה זמנית - ' + title)
            preview_dialog.geometry('1000x600')
            preview_dialog.configure(bg='#ecf0f1')
            preview_dialog.transient(self.root)
            preview_dialog.resizable(True, True)
            
            header = tk.Frame(preview_dialog, bg='#ecf0f1')
            header.pack(fill=tk.X, padx=10, pady=(10, 6))

            tk.Label(
                header,
                text=fix_rtl_text(f'תצוגה זמנית: {title} ({len(df)} רשומות)'),
                font=('Arial', 14, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50'
            ).pack(side=tk.RIGHT, padx=5)

            def _toggle_maximize():
                try:
                    if getattr(preview_dialog, '_is_maximized', False):
                        preview_dialog.state('normal')
                        preview_dialog._is_maximized = False
                    else:
                        preview_dialog.state('zoomed')
                        preview_dialog._is_maximized = True
                except Exception:
                    try:
                        preview_dialog.state('zoomed')
                        preview_dialog._is_maximized = True
                    except Exception:
                        pass

            tk.Button(
                header,
                text='🗖',
                command=_toggle_maximize,
                font=('Arial', 12, 'bold'),
                bg='#bdc3c7',
                fg='black',
                width=3,
                cursor='hand2'
            ).pack(side=tk.LEFT, padx=4)
            
            tree_frame = tk.Frame(preview_dialog, bg='#ecf0f1')
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            columns = list(df.columns)
            
            # Reverse DataFrame columns for RTL
            df_rtl = df[list(reversed(columns))]
            columns_rtl = list(df_rtl.columns)
            
            tree = ttk.Treeview(tree_frame, columns=columns_rtl, show='headings', height=25)
            
            for col in columns_rtl:
                tree.heading(col, text=str(col), anchor='e')
                tree.column(col, width=120, anchor='e')
            
            # Configure tags for alternating row colors
            tree.tag_configure('oddrow', background='#FFFFFF')
            tree.tag_configure('evenrow', background='#F0F0F0')
            
            for idx, row in df_rtl.iterrows():
                values = [str(val) if val is not None else '' for val in row]
                tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
                tree.insert('', 'end', values=values, tags=(tag,))
            
            scrollbar_y = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            scrollbar_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
            tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
            
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            try:
                preview_dialog.update_idletasks()
                tree.xview_moveto(1.0)
            except Exception:
                pass
            
            tk.Button(
                preview_dialog,
                text='סגור',
                command=preview_dialog.destroy,
                font=('Arial', 12, 'bold'),
                bg='#95a5a6',
                fg='white',
                padx=20,
                pady=8
            ).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בהצגת התצוגה:\n{str(e)}")
    
    def _show_export_success_dialog(self, file_path: str):
        try:
            dlg = tk.Toplevel(self.root)
            dlg.title("הצלחה")
            dlg.transient(self.root)
            dlg.grab_set()
            dlg.resizable(False, False)

            container = tk.Frame(dlg, padx=14, pady=12)
            container.pack(fill=tk.BOTH, expand=True)

            tk.Label(container, text="הקובץ יוצא בהצלחה!", font=('Arial', 12, 'bold')).pack(anchor='e', pady=(0, 6))
            tk.Label(container, text=str(file_path or ''), font=('Arial', 9), fg='#34495e', wraplength=520, justify='right').pack(anchor='e', pady=(0, 10))

            btns = tk.Frame(container)
            btns.pack(fill=tk.X)

            def _open():
                try:
                    if file_path and os.path.exists(file_path):
                        os.startfile(file_path)
                        return
                except Exception:
                    pass
                try:
                    if file_path:
                        webbrowser.open('file:///' + file_path.replace('\\', '/'))
                except Exception:
                    pass

            tk.Button(btns, text="פתח קובץ", width=12, command=_open).pack(side=tk.LEFT, padx=4)
            tk.Button(btns, text="סגור", width=12, command=dlg.destroy).pack(side=tk.RIGHT, padx=4)

            try:
                dlg.update_idletasks()
                w = max(420, int(dlg.winfo_reqwidth() or 420))
                h = max(160, int(dlg.winfo_reqheight() or 160))
                sw = int(dlg.winfo_screenwidth() or 1360)
                sh = int(dlg.winfo_screenheight() or 760)
                x = max(0, (sw // 2) - (w // 2))
                y = max(0, (sh // 2) - (h // 2))
                dlg.geometry(f'{w}x{h}+{x}+{y}')
            except Exception:
                pass
        except Exception:
            try:
                messagebox.showinfo("הצלחה", "הקובץ יוצא בהצלחה!")
            except Exception:
                pass
    
    def sync_with_excel(self):
        """סינכרון עם קובץ Excel - קריאה מהאקסל לעדכון מסד הנתונים"""
        if not self.ensure_can_modify():
            return
        if not os.path.exists(self.excel_path):
            self.sync_label.config(text="סטטוס: קובץ Excel לא נמצא", fg='#e74c3c')
            messagebox.showwarning("שגיאה", "קובץ Excel לא נמצא")
            return
        
        try:
            self.sync_label.config(text="🔄 מסנכרן מאקסל...", fg='#3498db')
            self.root.update()  # עדכון מיידי של התצוגה
            
            # קריאה מהאקסל (עדכון התלמידים)
            updated, errors = self.importer.quick_update_from_excel(self.excel_path)
            self.last_excel_mod_time = os.path.getmtime(self.excel_path)
            
            # רענון הטבלה
            self.load_students()
            
            # הודעה למשתמש
            if updated > 0:
                self.sync_label.config(text=f"✓ סונכרנו {updated} תלמידים מאקסל", fg='#27ae60')
                messagebox.showinfo("הצלחה", f"סונכרנו {updated} תלמידים מהאקסל!")
            else:
                self.sync_label.config(text="✓ אין עדכונים חדשים", fg='#27ae60')
                messagebox.showinfo("סינכרון", "הנתונים מעודכנים - אין שינויים חדשים")
            
            # חזרה להודעה רגילה אחרי 5 שניות
            self.root.after(5000, lambda: self.sync_label.config(text="סטטוס: מוכן", fg='#27ae60'))
            
        except Exception as e:
            self.sync_label.config(text=f"✗ שגיאה בסינכרון", fg='#e74c3c')
            messagebox.showerror("שגיאה", f"שגיאה בסינכרון:\n{e}")
    
    def export_to_excel_now(self):
        """ייצוא לאקסל (רק עמודות G, H, I - לא דורס!).

        כדי למנוע האטות בסביבה מרובת עמדות, הפעולה מתוזמנת (batch) ולא מתבצעת על כל שינוי.
        """
        try:
            self._schedule_excel_export()
        except Exception:
            pass

    def _schedule_excel_export(self, force: bool = False):
        try:
            if not bool(getattr(self, '_excel_auto_export_enabled', True)):
                return
        except Exception:
            return

        try:
            if not self.ensure_can_modify():
                return
        except Exception:
            return

        try:
            if not getattr(self, 'excel_path', None):
                return
        except Exception:
            return

        try:
            self._excel_export_pending = True
        except Exception:
            pass

        try:
            if getattr(self, '_excel_export_job', None) is not None:
                if force:
                    try:
                        self.root.after_cancel(self._excel_export_job)
                    except Exception:
                        pass
                    self._excel_export_job = None
                else:
                    return
        except Exception:
            pass

        delay_sec = float(getattr(self, '_excel_export_interval_sec', 300.0) or 300.0)
        if force:
            delay_sec = 0.05
        delay_ms = int(max(50, int(delay_sec * 1000)))
        try:
            self._excel_export_job = self.root.after(delay_ms, self._export_to_excel_flush)
        except Exception:
            self._excel_export_job = None

    def _export_to_excel_flush(self):
        try:
            self._excel_export_job = None
        except Exception:
            pass

        try:
            if not getattr(self, 'has_changes', False):
                self._excel_export_pending = False
                return
        except Exception:
            return

        # Run export in background to avoid freezing the Tk UI (openpyxl + network Excel can be slow).
        try:
            if bool(getattr(self, '_excel_export_thread_running', False)):
                try:
                    self._schedule_excel_export()
                except Exception:
                    pass
                return
        except Exception:
            pass

        try:
            import time
            now = float(time.time())
            min_gap = float(getattr(self, '_excel_export_min_gap_sec', 3.0) or 3.0)
            last_done = float(getattr(self, '_excel_export_last_done_ts', 0.0) or 0.0)
            if (now - last_done) < min_gap:
                wait_ms = int(max(50, (min_gap - (now - last_done)) * 1000.0))
                try:
                    self._excel_export_job = self.root.after(wait_ms, self._export_to_excel_flush)
                except Exception:
                    self._excel_export_job = None
                return
        except Exception:
            pass

        def _do_export():
            try:
                self.importer.export_columns_only(self.excel_path)

                try:
                    mtime = os.path.getmtime(self.excel_path)
                except Exception:
                    mtime = None

                try:
                    import time
                    last_done = float(time.time())
                except Exception:
                    last_done = None

                def _on_ok():
                    try:
                        if mtime is not None:
                            self.last_excel_mod_time = mtime
                    except Exception:
                        pass
                    try:
                        self.has_changes = False
                    except Exception:
                        pass
                    try:
                        if last_done is not None:
                            self._excel_export_last_done_ts = last_done
                    except Exception:
                        pass
                    try:
                        self._excel_export_pending = False
                    except Exception:
                        pass

                try:
                    self.root.after(0, _on_ok)
                except Exception:
                    _on_ok()
            except Exception:
                # Silent fail like before; just allow future retries.
                def _on_err():
                    try:
                        self._excel_export_pending = False
                    except Exception:
                        pass
                try:
                    self.root.after(0, _on_err)
                except Exception:
                    _on_err()
            finally:
                try:
                    self._excel_export_thread_running = False
                except Exception:
                    pass

        try:
            self._excel_export_thread_running = True
        except Exception:
            pass

        try:
            import threading
            t = threading.Thread(target=_do_export, daemon=True)
            t.start()
        except Exception:
            try:
                self._excel_export_thread_running = False
            except Exception:
                pass
            try:
                _do_export()
            except Exception:
                pass
    
    def sync_to_excel(self):
        """סינכרון מכוון מ-DB אל Excel (רק G,H,I) ללא קריאה מאקסל"""
        if not self.ensure_can_modify():
            return
        if not os.path.exists(self.excel_path):
            self.sync_label.config(text="סטטוס: קובץ Excel לא נמצא", fg='#e74c3c')
            messagebox.showwarning("שגיאה", "קובץ Excel לא נמצא")
            return
        try:
            self.sync_label.config(text="📤 מסנכרן ל-Excel...", fg='#3498db')
            self.root.update()
            self.importer.export_columns_only(self.excel_path)
            self.last_excel_mod_time = os.path.getmtime(self.excel_path)
            self.has_changes = False
            self.sync_label.config(text="✓ סונכרנו נתונים ל-Excel", fg='#27ae60')
            messagebox.showinfo("הצלחה", "הנתונים סונכרנו בהצלחה אל קובץ ה-Excel!")
        except Exception as e:
            self.sync_label.config(text="✗ שגיאה בסינכרון ל-Excel", fg='#e74c3c')
            messagebox.showerror("שגיאה", f"שגיאה בסינכרון ל-Excel:\n{e}")
    
    def sync_from_excel(self):
        """ייבוא שינויים מ-Excel (עדכון נקודות וכרטיסים בלבד)"""
        try:
            import pandas as pd
            # קריאה עם עמודת כרטיס כטקסט!
            df = pd.read_excel(self.excel_path, dtype={'מס\' כרטיס': str})
            
            updates_count = 0
            
            # עבור על כל שורה ועדכן רק אם יש שינויים
            for index, row in df.iterrows():
                last_name = str(row.get('שם משפחה', '')).strip()
                first_name = str(row.get('שם פרטי', '')).strip()
                
                if not last_name or not first_name or last_name == 'nan' or first_name == 'nan':
                    continue
                
                # חיפוש התלמיד במסד הנתונים
                students = self.db.search_students(f"{first_name} {last_name}")
                if students:
                    student = students[0]  # קח את הראשון
                    student_id = student['id']
                    
                    # עדכון כרטיס אם השתנה
                    card_number = str(row.get('מס\' כרטיס', '')).strip() if pd.notna(row.get('מס\' כרטיס')) else ''
                    # הסר apostroph אם יש (מעצם הטקסט באקסל)
                    card_number = card_number.lstrip("'")
                    if card_number and card_number != 'nan' and card_number != '0':
                        current_card = str(student['card_number']) if student['card_number'] else ''
                        
                        if card_number != current_card:
                            safe_print(f"  🔄 עדכון כרטיס: {first_name} {last_name}")
                            safe_print(f"     מ: {current_card} ← ל: {card_number}")
                            self.db.update_card_number(student_id, card_number)
                            updates_count += 1
                    
                    # עדכון נקודות אם השתנו
                    if 'מס\' נקודות' in df.columns and pd.notna(row.get('מס\' נקודות')):
                        try:
                            new_points = int(float(row.get('מס\' נקודות')))
                            if new_points != student['points']:
                                safe_print(f"  🔄 עדכון נקודות: {first_name} {last_name}")
                                safe_print(f"     מ: {student['points']} ← ל: {new_points}")
                                self.db.update_student_points(student_id, new_points, "סינכרון מ-Excel", "מערכת")
                                updates_count += 1
                        except Exception as ex:
                            safe_print(f"  ⚠️ שגיאה בעדכון נקודות: {ex}")
            
            safe_print(f"✅ סה\"כ עדכונים: {updates_count}")
            
        except Exception as e:
            safe_print(f"❌ שגיאה בייבוא מ-Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def start_auto_sync(self):
        """התחלת סינכרון אוטומטי - לכולם (מנהלים ומורים)"""
        # סינכרון עם Excel לכולם - גיבוי חשוב!
        if os.path.exists(self.excel_path):
            try:
                safe_print("🔄 סינכרון אוטומטי - קריאה מאקסל...")
                updated, errors = self.importer.quick_update_from_excel(self.excel_path)
                if updated > 0:
                    safe_print(f"✅ עודכנו {updated} תלמידים מאקסל")
                    self.load_students()  # רענון התצוגה
                elif updated == 0:
                    safe_print("ℹ️ אין עדכונים חדשים מהאקסל")
            except Exception as e:
                safe_print(f"⚠️ שגיאה בסינכרון: {e}")
        
        try:
            self.auto_sync_job = self.root.after(self.sync_interval, self.start_auto_sync)
        except Exception:
            self.auto_sync_job = None
    
    def auto_refresh_table(self):
        """רענון אוטומטי של הטבלה כל 10 שניות (לקליטת עדכונים מהעמדה הציבורית)"""
        try:
            # השהיית רענון אוטומטי לאחר אינטראקציה (מיון/בחירה) כדי לא לדרוס את מה שהמשתמש רואה
            suppressed = False
            try:
                import time
                until_ts = float(getattr(self, '_suppress_auto_refresh_until', 0.0) or 0.0)
                if until_ts and time.time() < until_ts:
                    suppressed = True
            except Exception:
                pass
            inline_editor = getattr(self, 'inline_editor', None)
            search_active = False
            try:
                # אם יש טיימר ניקוי חיפוש פעיל – סימן שיש תוצאות מסוננות על המסך
                if getattr(self, 'search_clear_job', None) is not None:
                    search_active = True
                else:
                    search_active = bool(self.search_var.get().strip())
            except Exception:
                search_active = False
            if (not suppressed) and inline_editor is None and not search_active:
                self.load_students(keep_selection=True)
        except Exception as e:
            safe_print(f"⚠️ שגיאה ברענון אוטומטי: {e}")
        
        # קביעת הרענון הבא
        try:
            self.auto_refresh_job = self.root.after(self.auto_refresh_interval, self.auto_refresh_table)
        except Exception:
            self.auto_refresh_job = None

    def refresh_table(self):
        try:
            if hasattr(self, 'search_clear_job') and self.search_clear_job is not None:
                try:
                    self.root.after_cancel(self.search_clear_job)
                except Exception:
                    pass
                self.search_clear_job = None
        except Exception:
            pass
        try:
            self.search_var.set("")
        except Exception:
            pass
        self.load_students()

        try:
            self._update_teacher_topbar_stats()
        except Exception:
            pass

    def _update_teacher_topbar_stats(self):
        try:
            t = getattr(self, 'current_teacher', None) or {}
            is_admin = bool(int(t.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False
        if is_admin:
            return

        if not getattr(self, '_teacher_stats_frame', None):
            return

        teacher_id = 0
        try:
            teacher_id = int((getattr(self, 'current_teacher', None) or {}).get('id') or 0)
        except Exception:
            teacher_id = 0

        classes = []
        try:
            classes = self.teacher_classes_cache or (self.db.get_teacher_classes(teacher_id) if teacher_id else [])
        except Exception:
            classes = []
        classes = [str(c).strip() for c in (classes or []) if str(c).strip()]

        # Get ALL classes stats for overall/min/max calculation
        try:
            all_classes_rows = self.db.get_all_classes_stats() or []
        except Exception:
            all_classes_rows = []

        # Calculate per-class averages for ALL classes
        all_per_class = []
        for r in (all_classes_rows or []):
            cname = str(r.get('class_name') or '').strip()
            try:
                sc = int(r.get('students_count') or 0)
            except Exception:
                sc = 0
            try:
                tp = int(r.get('total_points') or 0)
            except Exception:
                tp = 0
            avg = (float(tp) / float(sc)) if sc > 0 else 0.0
            all_per_class.append((cname, avg))

        # overall average = average of ALL class averages (unweighted)
        overall_avg = 0.0
        if all_per_class:
            try:
                overall_avg = sum([x[1] for x in all_per_class]) / len(all_per_class)
            except Exception:
                overall_avg = 0.0

        # Find highest and lowest from ALL classes
        best_avg = None
        worst_avg = None
        for cname, avg in all_per_class:
            if best_avg is None or avg > best_avg:
                best_avg = avg
            if worst_avg is None or avg < worst_avg:
                worst_avg = avg

        # Get teacher's subscribed classes stats
        try:
            if teacher_id:
                teacher_rows = self.db.get_teacher_classes_stats(teacher_id) or []
            else:
                teacher_rows = []
        except Exception:
            teacher_rows = []

        # Show only subscribed classes in "הכיתה שלך"
        subscribed = set([str(c).strip() for c in (classes or []) if str(c).strip()])
        your_parts = []
        for r in (teacher_rows or []):
            cname = str(r.get('class_name') or '').strip()
            try:
                sc = int(r.get('students_count') or 0)
            except Exception:
                sc = 0
            try:
                tp = int(r.get('total_points') or 0)
            except Exception:
                tp = 0
            avg = (float(tp) / float(sc)) if sc > 0 else 0.0
            if cname and cname in subscribed:
                your_parts.append(f"{cname}: {int(round(avg))} נק׳")
        
        your_text_full = ' | '.join(your_parts) if your_parts else '(אין כיתות)'
        your_text_truncated = False
        your_text = your_text_full
        
        # Truncate only if too long (estimate ~150 chars for reasonable display)
        if len(your_text) > 150:
            your_text = your_text[:147] + '...'
            your_text_truncated = True

        try:
            if self._teacher_stats_overall_label is not None:
                self._teacher_stats_overall_label.config(text=fix_rtl_text(f"ממוצע כללי: {int(round(overall_avg))} נק׳"))
        except Exception:
            pass

        try:
            if self._teacher_stats_highest_label is not None:
                if best_avg is not None:
                    self._teacher_stats_highest_label.config(text=fix_rtl_text(f"הגבוה ביותר: {int(round(best_avg))} נק׳"))
                else:
                    self._teacher_stats_highest_label.config(text=fix_rtl_text("הגבוה ביותר: -"))
        except Exception:
            pass

        try:
            if self._teacher_stats_lowest_label is not None:
                if worst_avg is not None:
                    self._teacher_stats_lowest_label.config(text=fix_rtl_text(f"הנמוך ביותר: {int(round(worst_avg))} נק׳"))
                else:
                    self._teacher_stats_lowest_label.config(text=fix_rtl_text("הנמוך ביותר: -"))
        except Exception:
            pass

        try:
            if self._teacher_stats_max_allowed_label is not None:
                try:
                    max_allowed_today = int(self.db.compute_max_points_allowed() or 0)
                except Exception:
                    max_allowed_today = 0
                self._teacher_stats_max_allowed_label.config(text=fix_rtl_text(f"מקסימום אפשרי: {int(max_allowed_today)}"))
        except Exception:
            pass

        try:
            if self._teacher_stats_your_classes_label is not None:
                if len(your_parts) == 1:
                    label_text = f"הכיתה שלך - {your_text}"
                else:
                    label_text = f"הכיתות שלך - {your_text}"
                
                self._teacher_stats_your_classes_label.config(text=fix_rtl_text(label_text))
        except Exception:
            pass

        # Show/hide filter button based on number of classes
        try:
            if self._teacher_filter_button is not None:
                if len(classes) > 1:
                    self._teacher_filter_button.pack(side=tk.RIGHT, padx=(0, 8))
                else:
                    self._teacher_filter_button.pack_forget()
        except Exception:
            pass

    def open_teacher_class_filter_dialog(self):
        """פתיחת חלון סינון כיתות למורה"""
        try:
            t = getattr(self, 'current_teacher', None) or {}
            is_admin = bool(int(t.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False
        if is_admin:
            return

        try:
            classes = self.teacher_classes_cache or []
        except Exception:
            classes = []
        classes = [str(c).strip() for c in (classes or []) if str(c).strip()]
        
        if len(classes) <= 1:
            return

        dialog = tk.Toplevel(self.root)
        dialog.title('סינון כיתות')
        dialog.geometry('500x600')
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        tk.Label(
            dialog,
            text=fix_rtl_text('בחר כיתות להצגה'),
            font=('Arial', 14, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(12, 6))

        # Get current filter or default to all classes
        current_filter = getattr(self, 'teacher_class_filter', None)
        if current_filter is None:
            current_filter = set(classes)
        
        # Frame for checkboxes
        cb_frame = tk.Frame(dialog, bg='#ecf0f1')
        cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # Add scrollbar
        canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
        scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        # Store checkbox variables
        checkbox_vars = {}
        
        for cls in sorted(classes):
            var = tk.BooleanVar(value=(cls in current_filter))
            checkbox_vars[cls] = var
            
            # Frame for each checkbox to ensure vertical alignment
            cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
            cb_row.pack(fill=tk.X, padx=20, pady=2)
            
            cb = tk.Checkbutton(
                cb_row,
                text=fix_rtl_text(cls),
                variable=var,
                font=('Arial', 11),
                bg='#ecf0f1',
                anchor='w'
            )
            cb.pack(side=tk.LEFT, fill=tk.X)

        canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)

        # Buttons frame
        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=(0, 12))

        def select_all():
            for var in checkbox_vars.values():
                var.set(True)

        def clear_all():
            for var in checkbox_vars.values():
                var.set(False)

        def apply_filter():
            selected = set([cls for cls, var in checkbox_vars.items() if var.get()])
            if not selected:
                messagebox.showwarning('אזהרה', 'יש לבחור לפחות כיתה אחת')
                return
            
            self.teacher_class_filter = selected
            dialog.destroy()
            self.load_students()

        def clear_filter():
            self.teacher_class_filter = None
            dialog.destroy()
            self.load_students()

        tk.Button(
            btn_frame,
            text='בחר הכל',
            command=select_all,
            font=('Arial', 10),
            bg='#3498db',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT, padx=4)

        tk.Button(
            btn_frame,
            text='נקה הכל',
            command=clear_all,
            font=('Arial', 10),
            bg='#e74c3c',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT, padx=4)

        tk.Button(
            btn_frame,
            text='החל סינון',
            command=apply_filter,
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=16,
            pady=4
        ).pack(side=tk.RIGHT, padx=4)

        tk.Button(
            btn_frame,
            text='הצג הכל',
            command=clear_filter,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT, padx=4)

    def open_teacher_classes_dialog(self):
        try:
            t = getattr(self, 'current_teacher', None) or {}
            is_admin = bool(int(t.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False
        if is_admin:
            return

        try:
            classes = self.teacher_classes_cache or []
        except Exception:
            classes = []
        classes = [str(c).strip() for c in (classes or []) if str(c).strip()]

        teacher_id = 0
        try:
            teacher_id = int((getattr(self, 'current_teacher', None) or {}).get('id') or 0)
        except Exception:
            teacher_id = 0

        dialog = tk.Toplevel(self.root)
        dialog.title('ממוצע הכיתות שלך')
        dialog.geometry('550x620')
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        tk.Label(
            dialog,
            text=fix_rtl_text('ממוצע הכיתות שלך'),
            font=('Arial', 14, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(12, 6))

        frame = tk.Frame(dialog, bg='#ecf0f1')
        frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        lb = tk.Listbox(frame, font=('Arial', 13), justify='right')
        lb.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=lb.yview)
        sb.pack(side=tk.LEFT, fill=tk.Y)
        lb.configure(yscrollcommand=sb.set)

        # Populate with per-class average points (vertical list)
        try:
            cls_rows = self.db.get_teacher_classes_stats(int(teacher_id)) if teacher_id else []
        except Exception:
            cls_rows = []
        if cls_rows:
            for r in (cls_rows or []):
                cname = str(r.get('class_name') or '').strip()
                try:
                    sc = int(r.get('students_count') or 0)
                except Exception:
                    sc = 0
                try:
                    tp = int(r.get('total_points') or 0)
                except Exception:
                    tp = 0
                avg = (float(tp) / float(sc)) if sc > 0 else 0.0
                if cname:
                    lb.insert(tk.END, f"{cname}  {int(round(avg))} נק׳")
        elif classes:
            for c in classes:
                lb.insert(tk.END, f"{str(c)}  0 נק׳")
        else:
            lb.insert(tk.END, '(אין כיתות)')

        tk.Button(
            dialog,
            text='סגור',
            command=dialog.destroy,
            font=('Arial', 11),
            bg='#95a5a6',
            fg='white',
            padx=16,
            pady=6
        ).pack(pady=(0, 12))
    
    def toggle_quick_update(self):
        """הפעלה/ביטול מצב עדכון מהיר"""
        if not self.quick_update_mode:
            opts = self.ask_quick_update_options()
            if not isinstance(opts, dict) or not opts:
                return
            points = opts.get('points')
            mode = opts.get('mode')
            operation = opts.get('operation')
            if points is None or mode is None:
                return
            if operation not in ('add', 'subtract', 'set'):
                operation = 'add'

            # שמירת סוג הפעולה למצב כרטיסים
            try:
                self.quick_update_operation = operation
            except Exception:
                pass

            # מצב 1: לפי כרטיס – מצב מתמשך כמו בעבר
            if mode == 'card':
                if operation == 'set':
                    try:
                        if not messagebox.askyesno("אישור", "בחרת 'מוחלט' (קובע נקודות). האם אתה בטוח?"):
                            return
                    except Exception:
                        pass
                self.quick_update_mode = True
                self.quick_update_points = int(points)
                self.quick_update_btn.config(
                    text=f"🛑 סיים עדכון מהיר ({points} נק')",
                    bg='#c0392b'
                )
                self.sync_label.config(
                    text=f"מצב עדכון מהיר: +{points} נקודות - הצג כרטיסים",
                    fg='#e67e22'
                )
                try:
                    self.open_quick_update_scan_dialog()
                except Exception:
                    pass
                return

            # מצבים 2-4: פעולה חד-פעמית
            self.quick_update_points = int(points)
            # במצב מוחלט נוסיף אזהרה נוספת לפני ביצוע פעולה רחבה
            if operation == 'set':
                try:
                    if not messagebox.askyesno("אישור נוסף", "זהו שינוי מוחלט של נקודות לתלמידים. האם אתה בטוח שברצונך להמשיך?"):
                        return
                except Exception:
                    pass
            try:
                updated = int(self.run_bulk_quick_update(opts) or 0)
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בעדכון מהיר:\n{e}")
                return

            self.has_changes = True
            try:
                self.load_students(keep_selection=True)
            except Exception:
                self.load_students()
            try:
                self.export_to_excel_now()
            except Exception:
                pass
            self.sync_label.config(
                text=f"✓ עודכנו {updated} תלמידים (+{self.quick_update_points} נק')",
                fg='#27ae60'
            )
            try:
                self.root.after(3000, lambda: self.sync_label.config(text="סטטוס: מוכן", fg='#27ae60'))
            except Exception:
                pass
        else:
            # ביטול מצב עדכון מהיר
            self.stop_quick_update()
    
    def ask_quick_update_points(self):
        """שאלת המשתמש כמה נקודות להוסיף בעדכון מהיר"""
        dialog = tk.Toplevel(self.root)
        dialog.title("עדכון מהיר")
        dialog.geometry("650x330")
        try:
            dialog.minsize(650, 330)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        result = {'points': None}
        
        tk.Label(
            dialog,
            text="כמה נקודות להוסיף לכל כרטיס?",
            font=('Arial', 13, 'bold'),
            bg='#ecf0f1'
        ).pack(pady=20)
        
        points_entry = tk.Entry(dialog, font=('Arial', 14), width=10)
        points_entry.pack(pady=10)
        points_entry.focus()
        
        def submit():
            try:
                points = int(points_entry.get())
                result['points'] = points
                dialog.destroy()
            except ValueError:
                messagebox.showerror("שגיאה", "יש להזין מספר שלם")
        
        tk.Button(
            dialog,
            text="התחל עדכון מהיר",
            command=submit,
            font=('Arial', 12),
            bg='#e67e22',
            fg='white',
            padx=20,
            pady=10
        ).pack(pady=10)
        
        points_entry.bind('<Return>', lambda e: submit())
        
        self.root.wait_window(dialog)
        return result['points']

    def ask_quick_update_options(self):
        """בחירת מצב עדכון מהיר והגדרות"""
        dialog = tk.Toplevel(self.root)
        dialog.title("עדכון מהיר")
        dialog.geometry("720x780")
        try:
            dialog.minsize(700, 740)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        result = {
            'points': None,
            'mode': None,
            'operation': 'add',
            'serial_from': None,
            'serial_to': None,
            'class_name': None,
            'class_names': None,
            'student_ids': None,
        }

        is_admin = False
        try:
            is_admin = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False

        tk.Label(dialog, text="עדכון מהיר", font=('Arial', 14, 'bold'), bg='#ecf0f1').pack(pady=(16, 8))

        points_frame = tk.Frame(dialog, bg='#ecf0f1')
        points_frame.pack(fill=tk.X, padx=18, pady=(0, 10))
        tk.Label(points_frame, text=fix_rtl_text("נקודות:"), font=('Arial', 11), bg='#ecf0f1').pack(side=tk.RIGHT)
        points_entry = tk.Entry(points_frame, font=('Arial', 12), width=10, justify='center')
        points_entry.pack(side=tk.RIGHT, padx=10)
        points_entry.focus_set()

        op_var = tk.StringVar(value='add')
        op_frame = tk.LabelFrame(dialog, text=fix_rtl_text("פעולה"), font=('Arial', 11, 'bold'), bg='#ecf0f1')
        op_frame.pack(fill=tk.X, padx=18, pady=(0, 8))
        tk.Radiobutton(op_frame, text=fix_rtl_text("הוספה"), variable=op_var, value='add', bg='#ecf0f1', anchor='e', justify='right').pack(side=tk.RIGHT, padx=10, pady=2)
        tk.Radiobutton(op_frame, text=fix_rtl_text("חיסור"), variable=op_var, value='subtract', bg='#ecf0f1', anchor='e', justify='right').pack(side=tk.RIGHT, padx=10, pady=2)
        tk.Radiobutton(op_frame, text=fix_rtl_text("מוחלט"), variable=op_var, value='set', bg='#ecf0f1', anchor='e', justify='right').pack(side=tk.RIGHT, padx=10, pady=2)

        mode_var = tk.StringVar(value='card')
        modes_frame = tk.LabelFrame(dialog, text=fix_rtl_text("סוג עדכון"), font=('Arial', 11, 'bold'), bg='#ecf0f1')
        modes_frame.pack(fill=tk.X, padx=18, pady=8)

        tk.Radiobutton(modes_frame, text=fix_rtl_text("1) לפי כרטיס (סריקה)"), variable=mode_var, value='card', bg='#ecf0f1', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=2)
        tk.Radiobutton(modes_frame, text=fix_rtl_text("2) טווח מס' סידורי (X עד Y)"), variable=mode_var, value='serial_range', bg='#ecf0f1', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=2)
        tk.Radiobutton(modes_frame, text=fix_rtl_text("3) כל תלמידי כיתה"), variable=mode_var, value='class', bg='#ecf0f1', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=2)
        tk.Radiobutton(modes_frame, text=fix_rtl_text("4) תלמידים נבחרים"), variable=mode_var, value='students', bg='#ecf0f1', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=2)
        if is_admin:
            tk.Radiobutton(modes_frame, text=fix_rtl_text("5) כל תלמידי בית הספר"), variable=mode_var, value='all_school', bg='#ecf0f1', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=2)

        params_frame = tk.Frame(dialog, bg='#ecf0f1')
        params_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=(6, 0))

        serial_frame = tk.LabelFrame(params_frame, text=fix_rtl_text("טווח מס׳ סידורי"), font=('Arial', 10, 'bold'), bg='#ecf0f1')
        serial_frame.pack(fill=tk.X, pady=6)
        tk.Label(serial_frame, text=fix_rtl_text("מ- X:"), font=('Arial', 10), bg='#ecf0f1').pack(side=tk.RIGHT, padx=(10, 4), pady=6)
        serial_from_entry = tk.Entry(serial_frame, font=('Arial', 11), width=8, justify='center')
        serial_from_entry.pack(side=tk.RIGHT, padx=(0, 16), pady=6)
        tk.Label(serial_frame, text=fix_rtl_text("עד Y:"), font=('Arial', 10), bg='#ecf0f1').pack(side=tk.RIGHT, padx=(10, 4), pady=6)
        serial_to_entry = tk.Entry(serial_frame, font=('Arial', 11), width=8, justify='center')
        serial_to_entry.pack(side=tk.RIGHT, padx=(0, 16), pady=6)

        class_frame = tk.LabelFrame(params_frame, text=fix_rtl_text("כיתה"), font=('Arial', 10, 'bold'), bg='#ecf0f1')
        class_frame.pack(fill=tk.X, pady=6)

        class_name_var = tk.StringVar(value='')
        class_entry = None
        class_opt = None

        tk.Label(class_frame, text=fix_rtl_text("כיתה:"), font=('Arial', 10), bg='#ecf0f1').pack(side=tk.RIGHT, padx=(10, 6), pady=6)
        class_entry = tk.Entry(class_frame, font=('Arial', 11), width=28, justify='right', textvariable=class_name_var, state='readonly')
        class_entry.pack(side=tk.RIGHT, padx=(0, 8), pady=6)

        def _get_available_classes_for_quick_update():
            if not is_admin:
                try:
                    cls = self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id'])
                except Exception:
                    cls = []
                return [str(c).strip() for c in (cls or []) if str(c).strip()]
            try:
                cls = self.db.get_all_class_names() or []
            except Exception:
                cls = []
            return [str(c).strip() for c in (cls or []) if str(c).strip()]

        def open_quick_update_class_selector():
            all_classes = _get_available_classes_for_quick_update()
            if not all_classes:
                messagebox.showinfo('מידע', 'אין כיתות במערכת')
                return

            current = [c.strip() for c in str(class_name_var.get() or '').split(',') if c.strip()]
            selected_classes = set(current)
            if not selected_classes:
                selected_classes = set(all_classes)

            selector_dialog = tk.Toplevel(dialog)
            selector_dialog.title('בחירת כיתות')
            selector_dialog.geometry('450x600')
            selector_dialog.configure(bg='#ecf0f1')
            selector_dialog.transient(dialog)
            selector_dialog.grab_set()
            selector_dialog.resizable(True, True)

            tk.Label(
                selector_dialog,
                text=fix_rtl_text('בחר כיתות'),
                font=('Arial', 14, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50'
            ).pack(pady=(12, 6))

            cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
            cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

            canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
            scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
            canvas.configure(yscrollcommand=scrollbar.set)

            def _on_mousewheel(event):
                try:
                    delta = int(event.delta)
                except Exception:
                    delta = 0
                if delta == 0:
                    return
                canvas.yview_scroll(int(-1 * (delta / 120)), 'units')
                return 'break'

            try:
                canvas.bind_all('<MouseWheel>', _on_mousewheel)
            except Exception:
                try:
                    canvas.bind('<MouseWheel>', _on_mousewheel)
                except Exception:
                    pass

            checkbox_vars = {}
            for cls in sorted(all_classes):
                var = tk.BooleanVar(value=(cls in selected_classes))
                checkbox_vars[cls] = var

                cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                cb_row.pack(fill=tk.X, padx=20, pady=2)

                cb = tk.Checkbutton(
                    cb_row,
                    text=fix_rtl_text(cls),
                    variable=var,
                    font=('Arial', 11),
                    bg='#ecf0f1',
                    anchor='w'
                )
                cb.pack(side=tk.LEFT, fill=tk.X)

            canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.LEFT, fill=tk.Y)

            btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
            btn_frame.pack(fill=tk.X, padx=10, pady=(0, 12))

            btn_inner = tk.Frame(btn_frame, bg='#ecf0f1')
            btn_inner.pack(side=tk.RIGHT)

            def select_all():
                for v in checkbox_vars.values():
                    v.set(True)

            def clear_all():
                for v in checkbox_vars.values():
                    v.set(False)

            def apply_selection():
                selected = [cls for cls, v in checkbox_vars.items() if v.get()]
                if not selected:
                    messagebox.showwarning('אזהרה', 'יש לבחור לפחות כיתה אחת')
                    return
                class_name_var.set(', '.join(sorted(selected)))
                selector_dialog.destroy()

            tk.Button(
                btn_inner,
                text='בחר הכל',
                command=select_all,
                font=('Arial', 10),
                bg='#3498db',
                fg='white',
                padx=12,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

            tk.Button(
                btn_inner,
                text='נקה הכל',
                command=clear_all,
                font=('Arial', 10),
                bg='#e74c3c',
                fg='white',
                padx=12,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

            tk.Button(
                btn_inner,
                text='אישור',
                command=apply_selection,
                font=('Arial', 10, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=16,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

        class_opt = tk.Button(
            class_frame,
            text='בחר...',
            command=open_quick_update_class_selector,
            font=('Arial', 10),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=4
        )
        class_opt.pack(side=tk.RIGHT, padx=(0, 16), pady=6)

        # --- Student selector (multi-select) ---
        student_frame = tk.LabelFrame(params_frame, text=fix_rtl_text("תלמידים"), font=('Arial', 10, 'bold'), bg='#ecf0f1')
        student_frame.pack(fill=tk.X, pady=6)

        selected_student_ids = []
        student_summary_var = tk.StringVar(value='')
        tk.Label(student_frame, text=fix_rtl_text("תלמידים שנבחרו:"), font=('Arial', 10), bg='#ecf0f1').pack(side=tk.RIGHT, padx=(10, 6), pady=6)
        student_entry = tk.Entry(student_frame, font=('Arial', 11), width=28, justify='right', textvariable=student_summary_var, state='readonly')
        student_entry.pack(side=tk.RIGHT, padx=(0, 8), pady=6)

        def _get_available_students_for_quick_update():
            # returns list of dict rows from DB
            try:
                all_students = self.db.get_all_students() or []
            except Exception:
                all_students = []

            if is_admin:
                return all_students

            try:
                allowed_classes = self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id'])
            except Exception:
                allowed_classes = []
            allowed_set = set([str(c).strip() for c in (allowed_classes or []) if str(c).strip()])
            if not allowed_set:
                return []
            return [s for s in all_students if str((s.get('class_name') or '')).strip() in allowed_set]

        def open_quick_update_students_selector():
            nonlocal selected_student_ids

            students = _get_available_students_for_quick_update()
            if not students:
                messagebox.showinfo('מידע', 'אין תלמידים לבחירה')
                return

            selector_dialog = tk.Toplevel(dialog)
            selector_dialog.title('בחירת תלמידים')
            selector_dialog.geometry('520x620')
            selector_dialog.configure(bg='#ecf0f1')
            selector_dialog.transient(dialog)
            selector_dialog.grab_set()
            selector_dialog.resizable(True, True)

            tk.Label(
                selector_dialog,
                text=fix_rtl_text('בחר תלמידים'),
                font=('Arial', 14, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50'
            , anchor='e').pack(fill=tk.X, padx=10, pady=(8, 4))

            frame = tk.Frame(selector_dialog, bg='#ecf0f1')
            frame.pack(fill=tk.BOTH, expand=True, padx=(0, 6), pady=6)

            canvas = tk.Canvas(frame, bg='#ecf0f1', highlightthickness=0)
            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor='ne')
            canvas.configure(yscrollcommand=scrollbar.set)

            def _on_canvas_configure(event):
                try:
                    w = int(event.width)
                except Exception:
                    w = 0
                if w > 0:
                    try:
                        canvas.itemconfigure(window_id, width=w)
                        canvas.coords(window_id, w, 0)
                    except Exception:
                        pass

            try:
                canvas.bind('<Configure>', _on_canvas_configure)
            except Exception:
                pass

            def _on_mousewheel(event):
                try:
                    delta = int(event.delta)
                except Exception:
                    delta = 0
                if delta == 0:
                    return
                canvas.yview_scroll(int(-1 * (delta / 120)), 'units')
                return 'break'

            def _bind_mousewheel(_event=None):
                try:
                    selector_dialog.bind_all('<MouseWheel>', _on_mousewheel)
                except Exception:
                    pass

            def _unbind_mousewheel(_event=None):
                try:
                    selector_dialog.unbind_all('<MouseWheel>')
                except Exception:
                    pass

            try:
                canvas.bind('<Enter>', _bind_mousewheel)
                canvas.bind('<Leave>', _unbind_mousewheel)
                scrollable_frame.bind('<Enter>', _bind_mousewheel)
                scrollable_frame.bind('<Leave>', _unbind_mousewheel)
                selector_dialog.bind('<Destroy>', _unbind_mousewheel)
            except Exception:
                pass

            checkbox_vars = {}
            for st in students:
                sid = int(st.get('id') or 0)
                if sid <= 0:
                    continue
                fn = _strip_asterisk_annotations(st.get('first_name', '') or '')
                ln = _strip_asterisk_annotations(st.get('last_name', '') or '')
                cn = str(st.get('class_name') or '').strip()
                sn = st.get('serial_number')
                try:
                    sn_i = int(sn or 0)
                except Exception:
                    sn_i = 0

                label = f"{sn_i} - {ln} {fn} ({cn})" if cn else f"{sn_i} - {ln} {fn}"
                var = tk.BooleanVar(value=(sid in (selected_student_ids or [])))
                checkbox_vars[sid] = var

                cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                cb_row.pack(fill=tk.X, padx=(0, 4), pady=2, anchor='e')

                cb = tk.Checkbutton(
                    cb_row,
                    text='',
                    variable=var,
                    font=('Arial', 11),
                    bg='#ecf0f1'
                )
                cb.pack(side=tk.RIGHT)

                lbl = tk.Label(
                    cb_row,
                    text=fix_rtl_text(label),
                    font=('Arial', 11),
                    bg='#ecf0f1',
                    anchor='e',
                    justify='right'
                )
                lbl.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(0, 6))

            canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.LEFT, fill=tk.Y)

            btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
            btn_frame.pack(fill=tk.X, padx=10, pady=(0, 12))

            btn_inner = tk.Frame(btn_frame, bg='#ecf0f1')
            btn_inner.pack(side=tk.RIGHT)

            def select_all():
                for v in checkbox_vars.values():
                    v.set(True)

            def clear_all():
                for v in checkbox_vars.values():
                    v.set(False)

            def apply_selection():
                nonlocal selected_student_ids
                selected_student_ids = [int(sid) for sid, v in checkbox_vars.items() if v.get()]
                if not selected_student_ids:
                    messagebox.showwarning('אזהרה', 'יש לבחור לפחות תלמיד אחד')
                    return
                student_summary_var.set(f"נבחרו {len(selected_student_ids)} תלמידים")
                selector_dialog.destroy()

            tk.Button(
                btn_inner,
                text='בחר הכל',
                command=select_all,
                font=('Arial', 10),
                bg='#3498db',
                fg='white',
                padx=12,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

            tk.Button(
                btn_inner,
                text='נקה הכל',
                command=clear_all,
                font=('Arial', 10),
                bg='#e74c3c',
                fg='white',
                padx=12,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

            tk.Button(
                btn_inner,
                text='אישור',
                command=apply_selection,
                font=('Arial', 10, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=16,
                pady=4
            ).pack(side=tk.RIGHT, padx=4)

        student_opt = tk.Button(
            student_frame,
            text='בחר...',
            command=open_quick_update_students_selector,
            font=('Arial', 10),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=4
        )
        student_opt.pack(side=tk.RIGHT, padx=(0, 16), pady=6)

        classes = []
        if not is_admin:
            try:
                classes = self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id'])
            except Exception:
                classes = []
            classes = [str(c).strip() for c in (classes or []) if str(c).strip()]
            if classes:
                class_name_var.set(classes[0])
                try:
                    tk.Label(class_frame, text=fix_rtl_text("כיתות מותרות: " + ', '.join(classes)), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d', anchor='e', justify='right').pack(fill=tk.X, padx=10, pady=(0, 6))
                except Exception:
                    pass

        def _sync_params_state(*_):
            m = mode_var.get()
            try:
                serial_state = 'normal' if m == 'serial_range' else 'disabled'
                class_state = 'normal' if m == 'class' else 'disabled'
                students_state = 'normal' if m == 'students' else 'disabled'
                serial_from_entry.config(state=serial_state)
                serial_to_entry.config(state=serial_state)
                if class_entry is not None:
                    class_entry.config(state=class_state)
                if class_opt is not None:
                    class_opt.config(state=class_state)
                try:
                    student_entry.config(state=students_state)
                except Exception:
                    pass
                try:
                    student_opt.config(state=students_state)
                except Exception:
                    pass
            except Exception:
                pass

        try:
            mode_var.trace_add('write', _sync_params_state)
        except Exception:
            pass
        _sync_params_state()

        btns = tk.Frame(dialog, bg='#ecf0f1')
        btns.pack(pady=14)

        result = {"mode": None, "date": None, "bonus_id": None, "student_id": None}

        def submit():
            try:
                pts = int(str(points_entry.get() or '').strip())
            except Exception:
                messagebox.showerror("שגיאה", "יש להזין מספר שלם של נקודות")
                return

            op = op_var.get()
            if op not in ('add', 'subtract', 'set'):
                op = 'add'
            # כללים בסיסיים
            if op in ('add', 'subtract') and pts <= 0:
                messagebox.showerror("שגיאה", "בהוספה/חיסור יש להזין מספר גדול מ-0")
                return
            if op == 'set' and pts < 0:
                messagebox.showerror("שגיאה", "במצב מוחלט לא ניתן להזין מספר שלילי")
                return

            m = mode_var.get()
            if m == 'serial_range':
                try:
                    x = int(str(serial_from_entry.get() or '').strip())
                    y = int(str(serial_to_entry.get() or '').strip())
                except Exception:
                    messagebox.showerror("שגיאה", "יש להזין טווח מס׳ סידורי תקין (X ו-Y)")
                    return
                if x <= 0 or y <= 0 or x > y:
                    messagebox.showerror("שגיאה", "טווח מס׳ סידורי לא תקין")
                    return
                result['serial_from'] = x
                result['serial_to'] = y
            elif m == 'class':
                raw = str(class_name_var.get() or '').strip()
                parts = [p.strip() for p in raw.split(',') if p.strip()]
                if not parts:
                    messagebox.showerror("שגיאה", "יש לבחור כיתה")
                    return
                result['class_names'] = parts
                result['class_name'] = parts[0]
            elif m == 'students':
                if not selected_student_ids:
                    messagebox.showerror("שגיאה", "יש לבחור תלמידים")
                    return
                result['student_ids'] = list(selected_student_ids)
            elif m == 'all_school':
                if not is_admin:
                    messagebox.showerror("שגיאה", "אפשרות זו זמינה רק למנהל")
                    return

            # אזהרה ראשונה במצב מוחלט
            if op == 'set':
                try:
                    if not messagebox.askyesno("אישור", "בחרת 'מוחלט' (קובע נקודות). האם אתה בטוח?"):
                        return
                except Exception:
                    pass

            result['points'] = pts
            result['mode'] = m
            result['operation'] = op
            dialog.destroy()

        tk.Button(btns, text="ביטול", command=dialog.destroy, font=('Arial', 11), bg='#95a5a6', fg='white', padx=14, pady=8).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text="בצע", command=submit, font=('Arial', 11, 'bold'), bg='#e67e22', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=8)

        points_entry.bind('<Return>', lambda e: submit())
        self.root.wait_window(dialog)
        if result.get('points') is None or result.get('mode') is None:
            return None
        return result

    def run_bulk_quick_update(self, opts: dict) -> int:
        """הרצת עדכון מהיר חד-פעמי לפי הגדרות"""
        mode = opts.get('mode')
        points = int(opts.get('points') or 0)
        operation = opts.get('operation')
        if operation not in ('add', 'subtract', 'set'):
            operation = 'add'

        is_admin = False
        try:
            is_admin = bool(self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1)
        except Exception:
            is_admin = False

        if mode == 'all_school' and not is_admin:
            raise ValueError("אפשרות זו זמינה רק למנהל")

        allowed_classes = None
        if not is_admin:
            try:
                allowed_classes = self.teacher_classes_cache or self.db.get_teacher_classes(self.current_teacher['id'])
            except Exception:
                allowed_classes = []
            allowed_classes = [str(c).strip() for c in (allowed_classes or []) if str(c).strip()]

        # איסוף רשימת תלמידים לעדכון
        conn = self.db.get_connection()
        cursor = conn.cursor()
        students = []
        try:
            if mode == 'serial_range':
                x = int(opts.get('serial_from') or 0)
                y = int(opts.get('serial_to') or 0)
                if x <= 0 or y <= 0 or x > y:
                    raise ValueError("טווח מס׳ סידורי לא תקין")
                if allowed_classes is not None:
                    if not allowed_classes:
                        students = []
                    else:
                        placeholders = ','.join('?' * len(allowed_classes))
                        cursor.execute(
                            f"SELECT id FROM students WHERE serial_number IS NOT NULL AND serial_number >= ? AND serial_number <= ? AND class_name IN ({placeholders})",
                            (x, y, *allowed_classes)
                        )
                        students = [int(r['id']) for r in cursor.fetchall()]
                else:
                    cursor.execute(
                        "SELECT id FROM students WHERE serial_number IS NOT NULL AND serial_number >= ? AND serial_number <= ?",
                        (x, y)
                    )
                    students = [int(r['id']) for r in cursor.fetchall()]

            elif mode == 'class':
                raw_classes = opts.get('class_names')
                if isinstance(raw_classes, (list, tuple)):
                    cls_list = [str(c).strip() for c in raw_classes if str(c).strip()]
                else:
                    raw = str(opts.get('class_name') or '').strip()
                    cls_list = [p.strip() for p in raw.split(',') if p.strip()]
                if not cls_list:
                    raise ValueError("יש לבחור כיתה")
                if allowed_classes is not None:
                    allowed_set = set(allowed_classes or [])
                    bad = [c for c in cls_list if c not in allowed_set]
                    if bad:
                        raise ValueError("אין הרשאה לעדכן כיתה זו")
                placeholders = ','.join('?' * len(cls_list))
                cursor.execute(f"SELECT id FROM students WHERE class_name IN ({placeholders})", tuple(cls_list))
                students = [int(r['id']) for r in cursor.fetchall()]

            elif mode == 'students':
                raw_ids = opts.get('student_ids')
                if not isinstance(raw_ids, (list, tuple)):
                    raise ValueError("יש לבחור תלמידים")
                try:
                    sid_list = [int(x) for x in raw_ids if int(x) > 0]
                except Exception:
                    sid_list = []
                if not sid_list:
                    raise ValueError("יש לבחור תלמידים")

                if allowed_classes is not None:
                    allowed_set = set(allowed_classes or [])
                    placeholders = ','.join('?' * len(sid_list))
                    cursor.execute(
                        f"SELECT id FROM students WHERE id IN ({placeholders}) AND class_name IN ({','.join('?' * len(allowed_set))})",
                        tuple(sid_list) + tuple(allowed_set)
                    )
                    students = [int(r['id']) for r in cursor.fetchall()]
                    if len(students) != len(set(sid_list)):
                        raise ValueError("אין הרשאה לעדכן אחד או יותר מהתלמידים שנבחרו")
                else:
                    placeholders = ','.join('?' * len(sid_list))
                    cursor.execute(f"SELECT id FROM students WHERE id IN ({placeholders})", tuple(sid_list))
                    students = [int(r['id']) for r in cursor.fetchall()]
                    if len(students) != len(set(sid_list)):
                        raise ValueError("אחד או יותר מהתלמידים שנבחרו לא נמצאו")

            elif mode == 'all_school':
                cursor.execute("SELECT id FROM students")
                students = [int(r['id']) for r in cursor.fetchall()]

            else:
                raise ValueError("מצב עדכון לא נתמך")
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # ביצוע עדכון נקודות לכל תלמיד
        updated = 0
        changes = []
        actor_name = self._get_points_actor_name()
        for sid in students:
            try:
                old_points = None
                try:
                    st = self.db.get_student_by_id(int(sid))
                    if st is not None and st.get('points') is not None:
                        old_points = int(st.get('points') or 0)
                except Exception:
                    old_points = None

                ok = False
                if operation == 'add':
                    ok = self.db.add_points(int(sid), points, f"עדכון מהיר +{points}", actor_name)
                    try:
                        if old_points is not None:
                            changes.append({'student_id': int(sid), 'old_points': int(old_points), 'new_points': int(old_points) + int(points)})
                    except Exception:
                        pass
                elif operation == 'subtract':
                    ok = self.db.subtract_points(int(sid), abs(points), f"עדכון מהיר -{abs(points)}", actor_name)
                    try:
                        if old_points is not None:
                            changes.append({'student_id': int(sid), 'old_points': int(old_points), 'new_points': int(old_points) - int(abs(points))})
                    except Exception:
                        pass
                elif operation == 'set':
                    ok = self.db.update_student_points(int(sid), max(0, int(points)), f"עדכון מהיר = {max(0, int(points))}", actor_name)
                    try:
                        if old_points is not None:
                            changes.append({'student_id': int(sid), 'old_points': int(old_points), 'new_points': int(max(0, int(points)))})
                    except Exception:
                        pass
                if ok:
                    updated += 1
            except Exception:
                pass

        try:
            if changes:
                op_label = "עדכון מהיר"
                try:
                    if operation == 'add':
                        op_label = f"עדכון מהיר (+{points})"
                    elif operation == 'subtract':
                        op_label = f"עדכון מהיר (-{abs(points)})"
                    elif operation == 'set':
                        op_label = f"עדכון מהיר (= {max(0, int(points))})"
                except Exception:
                    pass
                self._push_points_action(op_label, changes)
        except Exception:
            pass
        return updated
    
    def quick_update_on_enter(self, event):
        """טיפול בהקשת Enter במצב עדכון מהיר"""
        # הדבקת הטקסט שהוקלד אם יש
        try:
            # מחשב מיקוד למקרה שהיה שדה אחר במיקוד
            focused = self.root.focus_get()
            if hasattr(focused, 'get') and hasattr(focused, 'delete'):
                card_number = focused.get().strip()
                if card_number:
                    self.process_quick_update_card(card_number)
                    focused.delete(0, tk.END)
        except:
            pass
    
    def process_quick_update_card(self, card_number):
        """עדכון נקודות לכרטיס במצב עדכון מהיר"""
        student = self.db.get_student_by_card(card_number)
        
        if student:
            old_points = None
            try:
                old_points = int(student.get('points') or 0)
            except Exception:
                old_points = None

            op = getattr(self, 'quick_update_operation', 'add')
            if op not in ('add', 'subtract', 'set'):
                op = 'add'

            actor_name = self._get_points_actor_name()
            success = False
            if op == 'add':
                success = self.db.add_points(
                    student['id'],
                    int(self.quick_update_points),
                    f"עדכון מהיר +{int(self.quick_update_points)}",
                    actor_name
                )
            elif op == 'subtract':
                success = self.db.subtract_points(
                    student['id'],
                    abs(int(self.quick_update_points)),
                    f"עדכון מהיר -{abs(int(self.quick_update_points))}",
                    actor_name
                )
            elif op == 'set':
                target = max(0, int(self.quick_update_points))
                success = self.db.update_student_points(
                    student['id'],
                    target,
                    f"עדכון מהיר = {target}",
                    actor_name
                )

            try:
                if success and old_points is not None:
                    if op == 'add':
                        new_points = int(old_points) + int(self.quick_update_points)
                    elif op == 'subtract':
                        new_points = int(old_points) - abs(int(self.quick_update_points))
                    else:
                        new_points = max(0, int(self.quick_update_points))
                    self._push_points_action(
                        "עדכון מהיר (כרטיס)",
                        [{'student_id': int(student['id']), 'old_points': int(old_points), 'new_points': int(max(0, int(new_points)))}]
                    )
            except Exception:
                pass
            
            if success:
                self.has_changes = True
                # הצג הודעה קצרה
                try:
                    if op == 'add':
                        msg = f"✓ {_strip_asterisk_annotations(student.get('first_name','') or '')} {_strip_asterisk_annotations(student.get('last_name','') or '')}: +{int(self.quick_update_points)} נקודות"
                    elif op == 'subtract':
                        msg = f"✓ {_strip_asterisk_annotations(student.get('first_name','') or '')} {_strip_asterisk_annotations(student.get('last_name','') or '')}: -{abs(int(self.quick_update_points))} נקודות"
                    else:
                        msg = f"✓ {_strip_asterisk_annotations(student.get('first_name','') or '')} {_strip_asterisk_annotations(student.get('last_name','') or '')}: = {max(0, int(self.quick_update_points))} נקודות"
                except Exception:
                    msg = "✓ עודכן"
                self.sync_label.config(text=msg, fg='#27ae60')
                # רענן את הטבלה
                self.load_students()
                # ייצוא מיידי לאקסל
                self.export_to_excel_now()
                
                # הצג הודעה זמנית
                def _back_to_status():
                    try:
                        if op == 'add':
                            txt = f"מצב עדכון מהיר: +{int(self.quick_update_points)} נקודות - הצג כרטיסים"
                        elif op == 'subtract':
                            txt = f"מצב עדכון מהיר: -{abs(int(self.quick_update_points))} נקודות - הצג כרטיסים"
                        else:
                            txt = f"מצב עדכון מהיר: = {max(0, int(self.quick_update_points))} נקודות - הצג כרטיסים"
                        self.sync_label.config(text=txt, fg='#e67e22')
                    except Exception:
                        pass
                self.root.after(2000, _back_to_status)
        else:
            # כרטיס לא נמצא
            self.sync_label.config(
                text=f"✗ כרטיס {card_number} לא נמצא במערכת",
                fg='#e74c3c'
            )
            # חזרה להודעה הרגילה אחרי 3 שניות
            def _back_to_status_not_found():
                try:
                    op = getattr(self, 'quick_update_operation', 'add')
                    if op == 'subtract':
                        txt = f"מצב עדכון מהיר: -{abs(int(self.quick_update_points))} נקודות - הצג כרטיסים"
                    elif op == 'set':
                        txt = f"מצב עדכון מהיר: = {max(0, int(self.quick_update_points))} נקודות - הצג כרטיסים"
                    else:
                        txt = f"מצב עדכון מהיר: +{int(self.quick_update_points)} נקודות - הצג כרטיסים"
                    self.sync_label.config(text=txt, fg='#e67e22')
                except Exception:
                    pass
            self.root.after(3000, _back_to_status_not_found)
    
    def stop_quick_update(self):
        """ביטול מצב עדכון מהיר"""
        self.quick_update_mode = False
        self.quick_update_points = 0
        try:
            dlg = getattr(self, 'quick_update_scan_dialog', None)
            if dlg is not None:
                try:
                    dlg.destroy()
                except Exception:
                    pass
            self.quick_update_scan_dialog = None
            self.quick_update_scan_entry = None
        except Exception:
            pass
        self.quick_update_btn.config(
            text="⚡ עדכון מהיר",
            bg='#e67e22'
        )
        self.sync_label.config(
            text="סטטוס: מוכן",
            fg='#27ae60'
        )
        # הסרת מאזין מקלדת
        self.root.unbind('<Return>')

    def open_quick_update_scan_dialog(self):
        dlg = getattr(self, 'quick_update_scan_dialog', None)
        try:
            if dlg is not None and dlg.winfo_exists():
                try:
                    dlg.deiconify()
                    dlg.lift()
                except Exception:
                    pass
                try:
                    ent = getattr(self, 'quick_update_scan_entry', None)
                    if ent is not None:
                        ent.focus_set()
                except Exception:
                    pass
                return
        except Exception:
            pass

        dlg = tk.Toplevel(self.root)
        self.quick_update_scan_dialog = dlg
        dlg.title("עדכון מהיר - סריקה")
        dlg.geometry("520x250")
        try:
            dlg.minsize(500, 230)
        except Exception:
            pass
        dlg.configure(bg='#ecf0f1')
        dlg.transient(self.root)
        try:
            dlg.grab_set()
        except Exception:
            pass

        op = getattr(self, 'quick_update_operation', 'add')
        try:
            pts = int(getattr(self, 'quick_update_points', 0) or 0)
        except Exception:
            pts = 0
        if op == 'subtract':
            op_txt = f"-{abs(int(pts))}"
        elif op == 'set':
            op_txt = f"= {max(0, int(pts))}"
        else:
            op_txt = f"+{int(pts)}"

        tk.Label(dlg, text=fix_rtl_text(f"סרוק כרטיס כדי לבצע {op_txt} נק'"), font=('Arial', 13, 'bold'), bg='#ecf0f1').pack(pady=(18, 10))

        ent = tk.Entry(dlg, font=('Arial', 16), width=20, justify='center')
        self.quick_update_scan_entry = ent
        ent.pack(pady=(0, 12))
        ent.focus_set()

        btns = tk.Frame(dlg, bg='#ecf0f1')
        btns.pack(pady=(0, 12))

        def _do_scan():
            try:
                card_number = str(ent.get() or '').strip()
            except Exception:
                card_number = ''
            if not card_number:
                return
            try:
                self.process_quick_update_card(card_number)
            finally:
                try:
                    ent.delete(0, tk.END)
                    ent.focus_set()
                except Exception:
                    pass

        def _close():
            try:
                self.stop_quick_update()
            except Exception:
                try:
                    dlg.destroy()
                except Exception:
                    pass

        tk.Button(btns, text="סיים", command=_close, font=('Arial', 12, 'bold'), bg='#c0392b', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text="בצע", command=_do_scan, font=('Arial', 12), bg='#e67e22', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=8)

        ent.bind('<Return>', lambda e: _do_scan())
        try:
            dlg.protocol("WM_DELETE_WINDOW", _close)
        except Exception:
            pass
    
    def update_master_card(self):
        """עדכון כרטיס מאסטר - קוד יציאה לעמדה הציבורית"""
        if not self.ensure_can_modify():
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("כרטיס מאסטר")
        dialog.geometry("650x330")
        try:
            dialog.minsize(650, 330)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        tk.Label(
            dialog,
            text="הזן מספר כרטיס מאסטר:",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1'
        ).pack(pady=20)

        card_entry = tk.Entry(dialog, font=('Arial', 14), width=15)
        card_entry.pack(pady=10)
        card_entry.focus()

        # נסיון למלא את המספר הקיים מקובץ master_card.txt
        try:
            # קודם כל ננסה לקרוא מתיקיית הרשת המשותפת (אם הוגדרה)
            try:
                cfg = self.load_app_config()
                if isinstance(cfg, dict):
                    shared_folder = cfg.get('shared_folder') or cfg.get('network_root')
                    if shared_folder and os.path.isdir(shared_folder):
                        master_file = os.path.join(shared_folder, 'master_card.txt')
                        if os.path.exists(master_file):
                            with open(master_file, 'r', encoding='utf-8') as f:
                                current = f.read().strip()
                                if current:
                                    card_entry.insert(0, current)
            except Exception:
                pass

            # אם לא נמצא בקובץ משותף – ננסה בקובץ הנתונים המקומי (ליד config.json החי)
            if not card_entry.get():
                config_file = self._get_config_file_path()
                data_dir = os.path.dirname(config_file) or self.base_dir
                master_file = os.path.join(data_dir, 'master_card.txt')
                if os.path.exists(master_file):
                    with open(master_file, 'r', encoding='utf-8') as f:
                        current = f.read().strip()
                        if current:
                            card_entry.insert(0, current)
        except Exception:
            pass

        def submit():
            card_number = card_entry.get().strip()
            if not card_number:
                messagebox.showwarning("אזהרה", "יש להזין מספר כרטיס")
                return

            # שמירת הכרטיס המאסטר לקובץ
            try:
                # יעד ברירת מחדל: תיקיית הנתונים של config.json (ProgramData/LocalAppData)
                config_file = self._get_config_file_path()
                data_dir = os.path.dirname(config_file) or self.base_dir
                try:
                    os.makedirs(data_dir, exist_ok=True)
                except Exception:
                    pass

                targets = [os.path.join(data_dir, 'master_card.txt')]

                # אם מוגדרת תיקיית רשת משותפת – נשמור גם שם כדי שכל העמדות ישתמשו באותו כרטיס מאסטר
                try:
                    cfg = self.load_app_config()
                    if isinstance(cfg, dict):
                        shared_folder = cfg.get('shared_folder') or cfg.get('network_root')
                        if shared_folder:
                            try:
                                os.makedirs(shared_folder, exist_ok=True)
                            except Exception:
                                pass
                            targets.insert(0, os.path.join(shared_folder, 'master_card.txt'))
                except Exception:
                    pass

                write_error = None
                wrote_any = False
                for path in targets:
                    try:
                        with open(path, 'w', encoding='utf-8') as f:
                            f.write(card_number)
                        wrote_any = True
                    except Exception as e2:
                        write_error = e2

                if not wrote_any:
                    raise write_error or Exception("לא ניתן לשמור קובץ כרטיס מאסטר")
                
                messagebox.showinfo(
                    "הצלחה", 
                    f"כרטיס מאסטר עודכן!\n\n"
                    f"מספר כרטיס: {card_number}\n\n"
                    f"כרטיס זה ישמש כקוד יציאה בעמדה הציבורית."
                )
                
                dialog.destroy()
            
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לשמור כרטיס מאסטר:\n{e}")

        tk.Button(
            dialog,
            text="שמור",
            command=submit,
            font=('Arial', 12),
            bg='#8e44ad',
            fg='white',
            padx=20,
            pady=10
        ).pack(pady=10)

        card_entry.bind('<Return>', lambda e: submit())

        self.root.wait_window(dialog)
    
    def open_color_settings(self):
        """פתיחת חלון הגדרות צבעים"""
        if not self.ensure_can_modify():
            return
        try:
            import color_editor

            # פתיחת עורך הצבעים והמטבעות כחלון משנה בתוך אותה ריצה (מודאלי וצמוד לעמדת הניהול)
            dialog = tk.Toplevel(self.root)
            dialog.title("הגדרות מטבעות וצבעים - מערכת ניקוד")
            dialog.transient(self.root)
            dialog.grab_set()

            color_editor.ColorEditor(dialog)

            # המתנה עד סגירת חלון הצבעים לפני חזרה לעמדת הניהול
            self.root.wait_window(dialog)

        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לפתוח עורך צבעים:\n{e}")
    
    def open_messages_manager(self):
        """פתיחת חלון ניהול הודעות"""
        if not self.ensure_can_modify():
            return
        try:
            import messages_manager

            # פתיחת מנהל ההודעות כחלון משנה בתוך אותה ריצה
            dialog = tk.Toplevel(self.root)
            dialog.transient(self.root)
            dialog.grab_set()
            messages_manager.MessagesManager(dialog)

        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לפתוח מנהל הודעות:\n{e}")
    
    def open_teachers_manager(self):
        """פתיחת חלון ניהול מורים (רק למנהלים)"""
        if not self.ensure_can_modify():
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("👥 ניהול מורים והרשאות")
        dialog.geometry("1000x700")
        try:
            dialog.minsize(1000, 700)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        # כותרת
        tk.Label(
            dialog,
            text="👥 ניהול מורים והרשאות",
            font=('Arial', 16, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=15)
        
        # מסגרת רשימה
        list_frame = tk.Frame(dialog, bg='#ecf0f1')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # רשימת מורים (מימין לשמאל: שם, כרטיס, תפקיד, כיתות, מגבלות, שימושים היום)
        columns = (
            'bonus_points_today',
            'bonus_runs_used_today',
            'bonus_runs_limit',
            'bonus_points_limit',
            'classes',
            'role',
            'card',
            'name'
        )
        style = ttk.Style(dialog)
        style.configure("Teachers.Treeview", font=('Arial', 10))
        # הגדלת ה-padding האנכי לכותרות כדי לאפשר הצגת שתי שורות טקסט (שוליים גדולים יותר)
        style.configure("Teachers.Treeview.Heading", font=('Arial', 10, 'bold'), padding=(6, 22))
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15, style="Teachers.Treeview")

        tree.heading('name', text='שם', anchor='e')
        tree.heading('card', text="מס׳ כרטיס / סיסמה", anchor='e')
        tree.heading('role', text='תפקיד', anchor='e')
        tree.heading('classes', text='כיתות', anchor='e')
        tree.heading('bonus_points_limit', text='מקס׳ נק׳\nבונוס לתלמיד', anchor='e')
        tree.heading('bonus_runs_limit', text='מקס׳ הפעלות\nבונוס ליום', anchor='e')
        tree.heading('bonus_runs_used_today', text='הפעלות\nבונוס היום', anchor='e')
        tree.heading('bonus_points_today', text='סה"כ נק׳\nשהמורה חילק היום', anchor='e')

        tree.column('name', width=130, anchor='e')
        tree.column('card', width=130, anchor='e')
        tree.column('role', width=70, anchor='e')
        tree.column('classes', width=260, anchor='e')
        tree.column('bonus_points_limit', width=90, anchor='e')
        tree.column('bonus_runs_limit', width=90, anchor='e')
        tree.column('bonus_runs_used_today', width=90, anchor='e')
        tree.column('bonus_points_today', width=90, anchor='e')

        tree.tag_configure('evenrow', background='#ffffff')
        tree.tag_configure('oddrow', background='#f4f6f7')
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        def load_teachers():
            """טעינת מורים"""
            tree.delete(*tree.get_children())
            teachers = self.db.get_all_teachers()
            today_iso = date.today().isoformat()
            for index, teacher in enumerate(teachers):
                role = "מנהל" if teacher['is_admin'] else "מורה"
                classes = self.db.get_teacher_classes(teacher['id'])
                classes_str = ", ".join(classes) if classes else "(אין כיתות)"
                max_points = teacher.get('bonus_max_points_per_student')
                max_runs = teacher.get('bonus_max_total_runs')
                max_points_str = "" if max_points is None else str(max_points)
                max_runs_str = "" if max_runs is None else str(max_runs)

                # חישוב הפעלות בונוס היום
                runs_used = teacher.get('bonus_runs_used') or 0
                runs_reset_date = teacher.get('bonus_runs_reset_date')
                try:
                    runs_used_int = int(runs_used)
                except Exception:
                    runs_used_int = 0
                if not runs_reset_date or runs_reset_date != today_iso:
                    runs_used_int = 0

                # פורמט הפעלות בונוס היום (2/3 כאשר קיימת מגבלה יומית)
                try:
                    max_runs_int = int(max_runs) if max_runs is not None else None
                except Exception:
                    max_runs_int = None
                if max_runs_int is not None and max_runs_int > 0:
                    runs_used_str = f"{runs_used_int}/{max_runs_int}"
                else:
                    runs_used_str = str(runs_used_int)

                # חישוב נקודות בונוס היום
                points_used = teacher.get('bonus_points_used') or 0
                points_reset_date = teacher.get('bonus_points_reset_date')
                try:
                    points_used_int = int(points_used)
                except Exception:
                    points_used_int = 0
                if not points_reset_date or points_reset_date != today_iso:
                    points_used_int = 0
                points_used_str = str(points_used_int)

                # הסתרת מספרי כרטיסים בתצוגה – יופיעו רק בעריכה
                has_any_card = bool(
                    (teacher.get('card_number') or '').strip()
                    or (teacher.get('card_number2') or '').strip()
                    or (teacher.get('card_number3') or '').strip()
                )
                masked_card = "••••••" if has_any_card else "(ללא כרטיס)"

                tag_name = 'evenrow' if index % 2 == 0 else 'oddrow'
                tree.insert('', tk.END, values=(
                    points_used_str,
                    runs_used_str,
                    max_runs_str,
                    max_points_str,
                    classes_str,
                    role,
                    masked_card,
                    teacher['name']
                ), tags=(str(teacher['id']), tag_name))

        editable_columns = {'name', 'classes', 'bonus_points_limit', 'bonus_runs_limit'}

        def start_inline_edit(event):
            region = tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            row_id = tree.identify_row(event.y)
            col_id = tree.identify_column(event.x)
            if not row_id or not col_id:
                return
            try:
                col_index = int(col_id[1:]) - 1
            except Exception:
                return
            if col_index < 0 or col_index >= len(columns):
                return
            column_key = columns[col_index]
            if column_key not in editable_columns:
                return
            item = tree.item(row_id)
            teacher_id = int(item['tags'][0])
            old_value = tree.set(row_id, column_key)
            bbox = tree.bbox(row_id, col_id)
            if not bbox:
                return
            x, y, width, height = bbox
            if width <= 0 or height <= 0:
                return
            editor = tk.Entry(tree, font=('Arial', 10), justify='right')
            editor.insert(0, old_value)
            editor.select_range(0, tk.END)
            editor.place(x=x, y=y, width=width, height=height)
            editor.focus_set()

            def commit_edit(event=None):
                new_value = editor.get().strip()
                editor.destroy()
                if new_value == old_value:
                    return
                try:
                    if column_key == 'name':
                        if not new_value:
                            messagebox.showwarning("שגיאה", "שם מורה לא יכול להיות ריק.")
                            return
                        self.db.update_teacher(teacher_id, name=new_value)
                    elif column_key == 'card':
                        if not new_value:
                            messagebox.showwarning("שגיאה", "יש להזין מספר כרטיס.")
                            return
                        self.db.update_teacher(teacher_id, card_number=new_value)
                    elif column_key == 'classes':
                        classes_text = new_value
                        old_classes = self.db.get_teacher_classes(teacher_id)
                        for c in old_classes:
                            self.db.remove_teacher_class(teacher_id, c)
                        if classes_text:
                            classes = [c.strip() for c in classes_text.split(',') if c.strip()]
                            for class_name in classes:
                                self.db.add_teacher_class(teacher_id, class_name)
                    elif column_key in ('bonus_points_limit', 'bonus_runs_limit'):
                        teacher_row = self.db.get_teacher_by_id(teacher_id)
                        current_max_points = teacher_row.get('bonus_max_points_per_student')
                        current_max_runs = teacher_row.get('bonus_max_total_runs')
                        if column_key == 'bonus_points_limit':
                            if not new_value:
                                new_points = None
                            else:
                                try:
                                    v = int(new_value)
                                except ValueError:
                                    messagebox.showerror("שגיאה", "יש להזין מספר שלם עבור מקסימום נקודות בונוס.")
                                    return
                                new_points = v if v > 0 else None
                            new_runs = current_max_runs
                        else:
                            if not new_value:
                                new_runs = None
                            else:
                                try:
                                    v = int(new_value)
                                except ValueError:
                                    messagebox.showerror("שגיאה", "יש להזין מספר שלם עבור מקסימום הפעלות בונוס.")
                                    return
                                new_runs = v if v > 0 else None
                            new_points = current_max_points
                        self.db.set_teacher_bonus_limits(teacher_id, new_points, new_runs)
                    load_teachers()
                except Exception as e:
                    messagebox.showerror("שגיאה", f"עדכון נכשל:\n{e}")

            def cancel_edit(event=None):
                editor.destroy()

            editor.bind("<Return>", commit_edit)
            editor.bind("<Escape>", cancel_edit)
            editor.bind("<FocusOut>", commit_edit)

        tree.bind("<Double-1>", start_inline_edit)
        
        def add_teacher():
            """הוספת מורה חדש"""
            add_dialog = tk.Toplevel(dialog)
            add_dialog.title("הוספת מורה")
            add_dialog.geometry("700x760")
            try:
                add_dialog.minsize(680, 720)
            except Exception:
                pass
            add_dialog.configure(bg='#ecf0f1')
            add_dialog.resizable(True, True)
            add_dialog.transient(dialog)
            add_dialog.grab_set()
            
            tk.Label(add_dialog, text="שם המורה:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            name_entry = tk.Entry(add_dialog, font=('Arial', 12), width=30, justify='right')
            name_entry.pack(pady=5)
            
            tk.Label(add_dialog, text="מספר כרטיס 1:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry1 = tk.Entry(add_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry1.pack(pady=5)

            tk.Label(add_dialog, text="מספר כרטיס 2:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry2 = tk.Entry(add_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry2.pack(pady=5)

            tk.Label(add_dialog, text="מספר כרטיס 3:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry3 = tk.Entry(add_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry3.pack(pady=5)
            
            is_admin_var = tk.BooleanVar()
            tk.Checkbutton(
                add_dialog,
                text="✓ מנהל (הרשאות מלאות)",
                variable=is_admin_var,
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1'
            ).pack(pady=10)

            can_edit_card_var = tk.BooleanVar(value=True)
            can_edit_photo_var = tk.BooleanVar(value=True)
            tk.Checkbutton(
                add_dialog,
                text="✓ מורה יכול לשנות מס' כרטיס לתלמיד",
                variable=can_edit_card_var,
                font=('Arial', 10),
                bg='#ecf0f1'
            ).pack(pady=(0, 4))
            tk.Checkbutton(
                add_dialog,
                text="✓ מורה יכול לשנות תמונת תלמיד",
                variable=can_edit_photo_var,
                font=('Arial', 10),
                bg='#ecf0f1'
            ).pack(pady=(0, 10))
            
            # Classes selection with button
            classes_frame = tk.Frame(add_dialog, bg='#ecf0f1')
            classes_frame.pack(pady=5)
            
            tk.Label(classes_frame, text="כיתות:", font=('Arial', 11), bg='#ecf0f1').pack(side=tk.TOP, pady=5)
            
            classes_display_frame = tk.Frame(classes_frame, bg='#ecf0f1')
            classes_display_frame.pack(side=tk.TOP)
            
            classes_entry = tk.Entry(classes_display_frame, font=('Arial', 12), width=30, justify='right', state='readonly')
            classes_entry.pack(side=tk.LEFT, padx=5)
            
            selected_classes = []
            
            def open_class_selector():
                # Get all unique classes from students
                all_classes = set()
                try:
                    students = self.db.get_all_students()
                    for s in students:
                        cn = (s.get('class_name') or '').strip()
                        if cn:
                            all_classes.add(cn)
                except Exception:
                    pass
                
                if not all_classes:
                    messagebox.showinfo('מידע', 'אין כיתות במערכת')
                    return
                
                selector_dialog = tk.Toplevel(add_dialog)
                selector_dialog.title('בחירת כיתות')
                selector_dialog.geometry('500x600')
                selector_dialog.configure(bg='#ecf0f1')
                selector_dialog.transient(add_dialog)
                selector_dialog.grab_set()
                selector_dialog.resizable(True, True)
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('בחר כיתות'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(12, 6))
                
                cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                checkbox_vars = {}
                
                for cls in sorted(all_classes):
                    var = tk.BooleanVar(value=(cls in selected_classes))
                    checkbox_vars[cls] = var
                    
                    cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                    cb_row.pack(fill=tk.X, padx=20, pady=2)
                    
                    cb = tk.Checkbutton(
                        cb_row,
                        text=fix_rtl_text(cls),
                        variable=var,
                        font=('Arial', 11),
                        bg='#ecf0f1',
                        anchor='w'
                    )
                    cb.pack(side=tk.LEFT, fill=tk.X)
                
                canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                
                btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=(0, 12))
                
                def select_all():
                    for var in checkbox_vars.values():
                        var.set(True)
                
                def clear_all():
                    for var in checkbox_vars.values():
                        var.set(False)
                
                def apply_selection():
                    selected_classes.clear()
                    selected_classes.extend([cls for cls, var in checkbox_vars.items() if var.get()])
                    classes_entry.config(state='normal')
                    classes_entry.delete(0, tk.END)
                    classes_entry.insert(0, ', '.join(selected_classes) if selected_classes else '(אין כיתות)')
                    classes_entry.config(state='readonly')
                    selector_dialog.destroy()
                
                tk.Button(
                    btn_frame,
                    text='בחר הכל',
                    command=select_all,
                    font=('Arial', 10),
                    bg='#3498db',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='נקה הכל',
                    command=clear_all,
                    font=('Arial', 10),
                    bg='#e74c3c',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='אישור',
                    command=apply_selection,
                    font=('Arial', 10, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=16,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
            
            tk.Button(
                classes_display_frame,
                text='בחר...',
                command=open_class_selector,
                font=('Arial', 10),
                bg='#3498db',
                fg='white',
                padx=8,
                pady=2
            ).pack(side=tk.LEFT)

            tk.Label(add_dialog, text="מקסימום נקודות לבונוס מורה (לכל תלמיד בסבב, ריק = ללא הגבלה):", font=('Arial', 10), bg='#ecf0f1', justify='right').pack(pady=5)
            max_points_entry = tk.Entry(add_dialog, font=('Arial', 11), width=20, justify='center')
            max_points_entry.pack(pady=2)

            tk.Label(add_dialog, text="מקסימום מספר הפעלות בונוס למורה זה (ריק = ללא הגבלה):", font=('Arial', 10), bg='#ecf0f1', justify='right').pack(pady=5)
            max_runs_entry = tk.Entry(add_dialog, font=('Arial', 11), width=20, justify='center')
            max_runs_entry.pack(pady=2)
            
            def save():
                name = name_entry.get().strip()
                card1 = card_entry1.get().strip()
                card2 = card_entry2.get().strip()
                card3 = card_entry3.get().strip()
                is_admin = is_admin_var.get()
                max_points_text = max_points_entry.get().strip()
                max_runs_text = max_runs_entry.get().strip()
                
                if not name:
                    messagebox.showwarning("שגיאה", "יש להזין שם מורה")
                    return
                
                if not (card1 or card2 or card3):
                    messagebox.showwarning("שגיאה", "יש להזין לפחות מספר כרטיס אחד")
                    return
                
                # הוספת מורה
                teacher_id = self.db.add_teacher(
                    name,
                    card1,
                    is_admin,
                    card_number2=card2,
                    card_number3=card3,
                    can_edit_student_card=bool(can_edit_card_var.get()),
                    can_edit_student_photo=bool(can_edit_photo_var.get()),
                )
                if teacher_id:
                    max_points_val = None
                    max_runs_val = None
                    if max_points_text:
                        try:
                            v = int(max_points_text)
                            if v > 0:
                                max_points_val = v
                        except ValueError:
                            pass
                    if max_runs_text:
                        try:
                            v = int(max_runs_text)
                            if v > 0:
                                max_runs_val = v
                        except ValueError:
                            pass
                    if max_points_val is not None or max_runs_val is not None:
                        self.db.set_teacher_bonus_limits(teacher_id, max_points_val, max_runs_val)
                    # הוספת כיתות
                    if not is_admin and selected_classes:
                        for class_name in selected_classes:
                            self.db.add_teacher_class(teacher_id, class_name)
                    
                    messagebox.showinfo("הצלחה", f"מורה '{name}' נוסף בהצלחה!")
                    load_teachers()
                    add_dialog.destroy()
                else:
                    messagebox.showerror("שגיאה", "לא ניתן להוסיף מורה (אולי הכרטיס כבר קיים)")
            
            tk.Button(
                add_dialog,
                text="💾 שמור",
                command=save,
                font=('Arial', 12, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=30,
                pady=10
            ).pack(pady=15)
        
        def edit_teacher():
            """עריכת מורה"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור מורה")
                return
            
            item = tree.item(selected[0])
            teacher_id = int(item['tags'][0])
            teacher = self.db.get_teacher_by_id(teacher_id)
            
            edit_dialog = tk.Toplevel(dialog)
            edit_dialog.title("עריכת מורה")
            edit_dialog.geometry("700x760")
            try:
                edit_dialog.minsize(680, 720)
            except Exception:
                pass
            edit_dialog.configure(bg='#ecf0f1')
            edit_dialog.resizable(True, True)
            edit_dialog.transient(dialog)
            edit_dialog.grab_set()
            
            tk.Label(edit_dialog, text="שם המורה:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            name_entry = tk.Entry(edit_dialog, font=('Arial', 12), width=30, justify='right')
            name_entry.insert(0, teacher['name'])
            name_entry.pack(pady=5)
            
            tk.Label(edit_dialog, text="מספר כרטיס 1:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry1 = tk.Entry(edit_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry1.insert(0, teacher.get('card_number') or "")
            card_entry1.pack(pady=5)

            tk.Label(edit_dialog, text="מספר כרטיס 2:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry2 = tk.Entry(edit_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry2.insert(0, teacher.get('card_number2') or "")
            card_entry2.pack(pady=5)

            tk.Label(edit_dialog, text="מספר כרטיס 3:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            card_entry3 = tk.Entry(edit_dialog, font=('Arial', 12), width=30, justify='right')
            card_entry3.insert(0, teacher.get('card_number3') or "")
            card_entry3.pack(pady=5)
            
            is_admin_var = tk.BooleanVar(value=teacher['is_admin'] == 1)
            tk.Checkbutton(
                edit_dialog,
                text="✓ מנהל (הרשאות מלאות)",
                variable=is_admin_var,
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1'
            ).pack(pady=10)

            can_edit_card_var = tk.BooleanVar(value=int(teacher.get('can_edit_student_card', 1) or 0) == 1)
            can_edit_photo_var = tk.BooleanVar(value=int(teacher.get('can_edit_student_photo', 1) or 0) == 1)
            tk.Checkbutton(
                edit_dialog,
                text="✓ מורה יכול לשנות מס' כרטיס לתלמיד",
                variable=can_edit_card_var,
                font=('Arial', 10),
                bg='#ecf0f1'
            ).pack(pady=(0, 4))
            tk.Checkbutton(
                edit_dialog,
                text="✓ מורה יכול לשנות תמונת תלמיד",
                variable=can_edit_photo_var,
                font=('Arial', 10),
                bg='#ecf0f1'
            ).pack(pady=(0, 10))
            
            tk.Label(edit_dialog, text="כיתות:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
            
            classes_frame = tk.Frame(edit_dialog, bg='#ecf0f1')
            classes_frame.pack(pady=5)
            
            classes_entry = tk.Entry(classes_frame, font=('Arial', 12), width=22, justify='right', state='readonly')
            classes_entry.pack(side=tk.LEFT, padx=5)
            
            current_classes = self.db.get_teacher_classes(teacher_id)
            selected_classes_list = list(current_classes)
            if selected_classes_list:
                classes_entry.config(state='normal')
                classes_entry.insert(0, ", ".join(selected_classes_list))
                classes_entry.config(state='readonly')
            
            def open_teacher_classes_selector():
                all_classes = set()
                try:
                    students = self.db.get_all_students()
                    for s in students:
                        cn = (s.get('class_name') or '').strip()
                        if cn:
                            all_classes.add(cn)
                except Exception:
                    pass
                
                if not all_classes:
                    messagebox.showinfo('מידע', 'אין כיתות במערכת')
                    return
                
                selector_dialog = tk.Toplevel(edit_dialog)
                selector_dialog.title('בחירת כיתות')
                selector_dialog.geometry('500x600')
                selector_dialog.configure(bg='#ecf0f1')
                selector_dialog.transient(edit_dialog)
                selector_dialog.grab_set()
                selector_dialog.resizable(True, True)
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('בחר כיתות למורה'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(12, 6))
                
                cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                checkbox_vars = {}
                
                for cls in sorted(all_classes):
                    var = tk.BooleanVar(value=(cls in selected_classes_list))
                    checkbox_vars[cls] = var
                    
                    cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                    cb_row.pack(fill=tk.X, padx=20, pady=2)
                    
                    cb = tk.Checkbutton(
                        cb_row,
                        text=fix_rtl_text(cls),
                        variable=var,
                        font=('Arial', 11),
                        bg='#ecf0f1',
                        anchor='w'
                    )
                    cb.pack(side=tk.LEFT, fill=tk.X)
                
                canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                
                btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=(0, 12))
                
                def select_all():
                    for var in checkbox_vars.values():
                        var.set(True)
                
                def clear_all():
                    for var in checkbox_vars.values():
                        var.set(False)
                
                def apply_selection():
                    selected = [cls for cls, var in checkbox_vars.items() if var.get()]
                    selected_classes_list.clear()
                    selected_classes_list.extend(selected)
                    classes_entry.config(state='normal')
                    classes_entry.delete(0, tk.END)
                    classes_entry.insert(0, ', '.join(selected) if selected else '')
                    classes_entry.config(state='readonly')
                    selector_dialog.destroy()
                
                tk.Button(
                    btn_frame,
                    text='בחר הכל',
                    command=select_all,
                    font=('Arial', 10),
                    bg='#3498db',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='נקה הכל',
                    command=clear_all,
                    font=('Arial', 10),
                    bg='#e74c3c',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='אישור',
                    command=apply_selection,
                    font=('Arial', 10, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=16,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
            
            tk.Button(
                classes_frame,
                text='בחר...',
                command=open_teacher_classes_selector,
                font=('Arial', 10, 'bold'),
                bg='#3498db',
                fg='white',
                padx=12,
                pady=4
            ).pack(side=tk.LEFT, padx=5)

            tk.Label(edit_dialog, text="מקסימום נקודות לבונוס מורה (לכל תלמיד בסבב, ריק = ללא הגבלה):", font=('Arial', 10), bg='#ecf0f1', justify='right').pack(pady=5)
            max_points_entry = tk.Entry(edit_dialog, font=('Arial', 11), width=20, justify='center')
            current_max_points = teacher.get('bonus_max_points_per_student')
            if current_max_points is not None and current_max_points > 0:
                try:
                    max_points_entry.insert(0, str(int(current_max_points)))
                except Exception:
                    pass
            max_points_entry.pack(pady=2)

            tk.Label(edit_dialog, text="מקסימום מספר הפעלות בונוס למורה זה (ריק = ללא הגבלה):", font=('Arial', 10), bg='#ecf0f1', justify='right').pack(pady=5)
            max_runs_entry = tk.Entry(edit_dialog, font=('Arial', 11), width=20, justify='center')
            current_max_runs = teacher.get('bonus_max_total_runs')
            if current_max_runs is not None and current_max_runs > 0:
                try:
                    max_runs_entry.insert(0, str(int(current_max_runs)))
                except Exception:
                    pass
            max_runs_entry.pack(pady=2)
            
            def save():
                name = name_entry.get().strip()
                card1 = card_entry1.get().strip()
                card2 = card_entry2.get().strip()
                card3 = card_entry3.get().strip()
                is_admin = is_admin_var.get()
                max_points_text = max_points_entry.get().strip()
                max_runs_text = max_runs_entry.get().strip()
                
                if not name:
                    messagebox.showwarning("שגיאה", "יש למלא שם")
                    return
                if not (card1 or card2 or card3):
                    messagebox.showwarning("שגיאה", "יש להזין לפחות מספר כרטיס אחד")
                    return
                
                # עדכון מורה
                success = self.db.update_teacher(
                    teacher_id,
                    name=name,
                    card_number=card1,
                    card_number2=card2,
                    card_number3=card3,
                    is_admin=is_admin,
                    can_edit_student_card=bool(can_edit_card_var.get()),
                    can_edit_student_photo=bool(can_edit_photo_var.get()),
                )
                if success:
                    max_points_val = None
                    max_runs_val = None
                    if max_points_text:
                        try:
                            v = int(max_points_text)
                            if v > 0:
                                max_points_val = v
                        except ValueError:
                            pass
                    if max_runs_text:
                        try:
                            v = int(max_runs_text)
                            if v > 0:
                                max_runs_val = v
                        except ValueError:
                            pass
                    if max_points_val is not None or max_runs_val is not None:
                        self.db.set_teacher_bonus_limits(teacher_id, max_points_val, max_runs_val)
                    else:
                        # אם שני השדות ריקים – נבטל הגבלות (ללא הגבלה)
                        self.db.set_teacher_bonus_limits(teacher_id, None, None)
                    # עדכון כיתות - מחק ישנות והוסף חדשות
                    old_classes = self.db.get_teacher_classes(teacher_id)
                    for old_class in old_classes:
                        self.db.remove_teacher_class(teacher_id, old_class)
                    
                    if not is_admin and selected_classes_list:
                        for class_name in selected_classes_list:
                            self.db.add_teacher_class(teacher_id, class_name)
                    
                    messagebox.showinfo("הצלחה", "המורה עודכן בהצלחה!")
                    load_teachers()
                    edit_dialog.destroy()
                else:
                    messagebox.showerror("שגיאה", "לא ניתן לעדכן מורה")
            
            tk.Button(
                edit_dialog,
                text="💾 שמור",
                command=save,
                font=('Arial', 12, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=30,
                pady=10
            ).pack(pady=15)
        
        def delete_teacher():
            """מחיקת מורה"""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור מורה")
                return
            
            item = tree.item(selected[0])
            teacher_id = int(item['tags'][0])
            # אחרי שינוי סדר העמודות, שם המורה הוא הערך האחרון, אך נעדיף למשוך מהמסד
            try:
                teacher_row = self.db.get_teacher_by_id(teacher_id)
                teacher_name = teacher_row['name'] if teacher_row and teacher_row.get('name') else item['values'][-1]
            except Exception:
                teacher_name = item['values'][-1]
            
            if messagebox.askyesno("אישור מחיקה", f"האם למחוק את המורה '{teacher_name}'?"):
                if self.db.delete_teacher(teacher_id):
                    messagebox.showinfo("הצלחה", "המורה נמחק בהצלחה!")
                    load_teachers()
                else:
                    messagebox.showerror("שגיאה", "לא ניתן למחוק מורה")
        
        # כפתורים
        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="➕ הוסף מורה",
            command=add_teacher,
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="✏️ ערוך",
            command=edit_teacher,
            font=('Arial', 11),
            bg='#3498db',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="🗑️ מחק",
            command=delete_teacher,
            font=('Arial', 11),
            bg='#e74c3c',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="✖ סגור",
            command=dialog.destroy,
            font=('Arial', 11),
            bg='#95a5a6',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        # טעינה ראשונית
        load_teachers()
    
    def toggle_statistics(self):
        """הפעלה/כיבוי הצגת סטטיסטיקות"""
        try:
            # שמירת ההגדרה
            show = '1' if self.show_statistics_var.get() else '0'
            self.db.set_setting('show_statistics', show)
            
            # הודעה למשתמש
            status = "מופעל" if show == '1' else "מכובה"
            messagebox.showinfo(
                "הצגת סטטיסטיקות",
                f"הצגת סטטיסטיקות {status}\n\n"
                "השינוי ייכנס לתוקף מיד בעמדה הציבורית."
            )
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לעדכן הגדרה:\n{e}")
    
    def _get_config_file_path(self) -> str:
        """החזרת נתיב קובץ ההגדרות ה"חי" מחוץ ל-Program Files במידת האפשר.

        העדיפות: PROGRAMDATA / LOCALAPPDATA / APPDATA, ורק כמוצא אחרון – תיקיית האפליקציה.
        """
        for env_name in ("PROGRAMDATA", "LOCALAPPDATA", "APPDATA"):
            root = os.environ.get(env_name)
            if not root:
                continue
            try:
                if os.path.isdir(root) and os.access(root, os.W_OK):
                    cfg_dir = os.path.join(root, "SchoolPoints")
                    try:
                        os.makedirs(cfg_dir, exist_ok=True)
                    except Exception:
                        pass
                    return os.path.join(cfg_dir, "config.json")
            except Exception:
                continue

        # מוצא אחרון – עשוי להיות לקריאה בלבד בהתקנה, אבל בסביבת פיתוח זה תקין
        return os.path.join(self.base_dir, 'config.json')

    @staticmethod
    def _get_config_file_path_static() -> str:
        for env_name in ("PROGRAMDATA", "LOCALAPPDATA", "APPDATA"):
            root = os.environ.get(env_name)
            if not root:
                continue
            try:
                if os.path.isdir(root) and os.access(root, os.W_OK):
                    cfg_dir = os.path.join(root, "SchoolPoints")
                    try:
                        os.makedirs(cfg_dir, exist_ok=True)
                    except Exception:
                        pass
                    return os.path.join(cfg_dir, "config.json")
            except Exception:
                continue

        base_dir = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base_dir, 'config.json')

    @staticmethod
    def load_app_config_static():
        try:
            live_config = AdminStation._get_config_file_path_static()
            base_dir = os.path.dirname(os.path.abspath(__file__))
            base_config = os.path.join(base_dir, 'config.json')

            if not os.path.exists(live_config) and os.path.exists(base_config):
                try:
                    shutil.copy2(base_config, live_config)
                except Exception:
                    pass

            local_cfg = {}
            if os.path.exists(live_config):
                with open(live_config, 'r', encoding='utf-8') as f:
                    local_cfg = json.load(f)

            shared_folder = None
            if isinstance(local_cfg, dict):
                shared_folder = local_cfg.get('shared_folder') or local_cfg.get('network_root')

            if shared_folder and os.path.isdir(shared_folder):
                shared_config_path = os.path.join(shared_folder, 'config.json')
                shared_cfg = None

                if os.path.exists(shared_config_path):
                    try:
                        with open(shared_config_path, 'r', encoding='utf-8') as f:
                            shared_cfg = json.load(f)
                    except Exception:
                        shared_cfg = None

                if not isinstance(shared_cfg, dict):
                    shared_cfg = dict(local_cfg) if isinstance(local_cfg, dict) else {}
                    try:
                        os.makedirs(shared_folder, exist_ok=True)
                    except Exception:
                        pass
                    try:
                        with open(shared_config_path, 'w', encoding='utf-8') as f:
                            json.dump(shared_cfg, f, ensure_ascii=False, indent=4)
                    except Exception:
                        pass

                return shared_cfg

            if local_cfg:
                return local_cfg

            if os.path.exists(base_config):
                with open(base_config, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    @staticmethod
    def save_app_config_static(config: dict) -> bool:
        try:
            config_file = AdminStation._get_config_file_path_static()
            cfg_dir = os.path.dirname(config_file) or '.'
            try:
                os.makedirs(cfg_dir, exist_ok=True)
            except Exception:
                pass

            local_cfg = dict(config) if isinstance(config, dict) else {}
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(local_cfg, f, ensure_ascii=False, indent=4)

            shared_folder = None
            if isinstance(local_cfg, dict):
                shared_folder = local_cfg.get('shared_folder') or local_cfg.get('network_root')

            if shared_folder and os.path.isdir(shared_folder):
                shared_config_path = os.path.join(shared_folder, 'config.json')
                try:
                    os.makedirs(shared_folder, exist_ok=True)
                except Exception:
                    pass
                try:
                    with open(shared_config_path, 'w', encoding='utf-8') as f:
                        json.dump(local_cfg, f, ensure_ascii=False, indent=4)
                except Exception:
                    pass

            return True
        except Exception:
            return False

    def load_app_config(self):
        """טעינת הגדרות יישום מקובץ הגדרות חיצוני, עם נפילה חזרה לקובץ המובנה."""
        try:
            live_config = self._get_config_file_path()
            base_config = os.path.join(self.base_dir, 'config.json')

            # אם עדיין אין קובץ "חי" אבל יש קובץ ברירת-מחדל מותקן – ננסה להעתיקו
            if not os.path.exists(live_config) and os.path.exists(base_config):
                try:
                    shutil.copy2(base_config, live_config)
                except Exception:
                    # אם לא הצליח – נמשיך לפחות לקרוא ממנו
                    pass

            # קריאת ההגדרות המקומיות (לשימוש גם כאינדיקציה לתיקיית הרשת)
            local_cfg = {}
            if os.path.exists(live_config):
                with open(live_config, 'r', encoding='utf-8') as f:
                    local_cfg = json.load(f)

            # אם מוגדרת תיקיית רשת משותפת – ננסה להשתמש בקובץ config.json משותף בתיקייה זו
            shared_folder = None
            if isinstance(local_cfg, dict):
                shared_folder = local_cfg.get('shared_folder') or local_cfg.get('network_root')

            if shared_folder and os.path.isdir(shared_folder):
                shared_config_path = os.path.join(shared_folder, 'config.json')
                shared_cfg = None

                # אם כבר קיים קובץ הגדרות משותף – נקרא ממנו
                if os.path.exists(shared_config_path):
                    try:
                        with open(shared_config_path, 'r', encoding='utf-8') as f:
                            shared_cfg = json.load(f)
                    except Exception:
                        shared_cfg = None

                # אם אין קובץ משותף תקין – ניצור אחד על בסיס ההגדרות המקומיות
                if not isinstance(shared_cfg, dict):
                    shared_cfg = dict(local_cfg) if isinstance(local_cfg, dict) else {}
                    try:
                        os.makedirs(shared_folder, exist_ok=True)
                    except Exception:
                        pass
                    try:
                        with open(shared_config_path, 'w', encoding='utf-8') as f:
                            json.dump(shared_cfg, f, ensure_ascii=False, indent=4)
                    except Exception:
                        pass

                return shared_cfg

            # ללא תיקיית רשת – החזר את ההגדרות המקומיות או קובץ ברירת המחדל
            if local_cfg:
                return local_cfg

            if os.path.exists(base_config):
                with open(base_config, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            safe_print(f"שגיאה בטעינת config.json: {e}")
        return {}

    def save_app_config(self, config: dict) -> bool:
        """שמירת הגדרות יישום לקובץ הגדרות חיצוני במיקום כתיב."""
        try:
            # קודם כל – שמירה לקובץ המקומי (שמשמש גם לאיתור תיקיית הרשת)
            config_file = self._get_config_file_path()
            cfg_dir = os.path.dirname(config_file) or '.'
            try:
                os.makedirs(cfg_dir, exist_ok=True)
            except Exception:
                pass

            local_cfg = dict(config) if isinstance(config, dict) else {}
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(local_cfg, f, ensure_ascii=False, indent=4)

            # אם מוגדרת תיקיית רשת משותפת – נשמור גם קובץ config.json משותף בתיקייה זו
            shared_folder = None
            if isinstance(local_cfg, dict):
                shared_folder = local_cfg.get('shared_folder') or local_cfg.get('network_root')

            if shared_folder and os.path.isdir(shared_folder):
                shared_config_path = os.path.join(shared_folder, 'config.json')
                try:
                    os.makedirs(shared_folder, exist_ok=True)
                except Exception:
                    pass
                try:
                    with open(shared_config_path, 'w', encoding='utf-8') as f:
                        json.dump(local_cfg, f, ensure_ascii=False, indent=4)
                except Exception:
                    # כשל בשמירת הקובץ המשותף לא צריך להפיל את האפליקציה
                    pass

            return True
        except Exception as e:
            safe_print(f"שגיאה בשמירת config.json: {e}")
            messagebox.showerror("שגיאה", f"לא ניתן לשמור קובץ הגדרות:\n{e}")
            return False

    def _seed_shared_sounds_folder(self, shared_folder: str) -> None:
        try:
            folder = str(shared_folder or '').strip()
            if not folder:
                return
            if not os.path.isdir(folder):
                return
            src_root = os.path.join(self.base_dir, 'sounds')
            dst_root = os.path.join(folder, 'sounds')
            try:
                os.makedirs(dst_root, exist_ok=True)
            except Exception:
                pass
            if not os.path.isdir(src_root):
                return
            for root, _, files in os.walk(src_root):
                rel = os.path.relpath(root, src_root)
                if rel == '.':
                    rel = ''
                dst_dir = os.path.join(dst_root, rel)
                try:
                    os.makedirs(dst_dir, exist_ok=True)
                except Exception:
                    pass
                for fn in (files or []):
                    try:
                        if not str(fn).lower().endswith(('.wav', '.mp3', '.ogg')):
                            continue
                        sp = os.path.join(root, fn)
                        dp = os.path.join(dst_dir, fn)
                        if os.path.exists(dp):
                            continue
                        try:
                            shutil.copy2(sp, dp)
                        except Exception:
                            pass
                    except Exception:
                        continue
        except Exception:
            return

    def ensure_initial_setup(self):
        try:
            cfg = self.load_app_config() or {}
        except Exception:
            cfg = {}

        # איתור נתיב ה-DB הקיים לפני שינוי ההגדרות (לצורך העתקה לתיקיית הרשת)
        current_db_path = None
        try:
            _db_for_path = Database()
            current_db_path = getattr(_db_for_path, 'db_path', None)
        except Exception:
            current_db_path = None

        shared = None
        if isinstance(cfg, dict):
            shared = cfg.get('shared_folder') or cfg.get('network_root')

        previously_configured = bool(shared)
        if shared and not os.path.isdir(shared):
            try:
                messagebox.showwarning(
                    "תיקייה משותפת לא זמינה",
                    "נראה שהמחשב כבר הוגדר בעבר לתיקיית נתונים משותפת, אך כרגע אין גישה אליה.\n\n"
                    "בדרך כלל זו בעיית רשת/חיבור למחשב המרכזי.\n"
                    "מומלץ לבדוק חיבור לרשת, שם מחשב/נתיב, והרשאות – ולא ליצור תיקייה משותפת חדשה.\n\n"
                    "שינוי התיקייה המשותפת עלול ליצור פיצול נתונים (\"מערכת חדשה\") ולשבש את העבודה."
                )
            except Exception:
                pass

        if not (shared and os.path.isdir(shared)):
            if not self._open_admin_shared_folder_dialog(
                cfg,
                first_run=not previously_configured,
                current_db_path=current_db_path,
                previous_shared_folder=shared if previously_configured else None,
                previously_configured=previously_configured,
            ):
                return False
            try:
                cfg = self.load_app_config() or cfg
            except Exception:
                pass
        try:
            shared = None
            if isinstance(cfg, dict):
                shared = cfg.get('shared_folder') or cfg.get('network_root')
            if shared and os.path.isdir(shared):
                self._seed_shared_sounds_folder(shared)
        except Exception:
            pass
        self.app_config = cfg
        temp_db = Database()
        teachers = temp_db.get_all_teachers()
        if not teachers:
            if not self._open_initial_admin_dialog(temp_db):
                return False
        return True

    def _open_admin_shared_folder_dialog(self, cfg, first_run, current_db_path=None, previous_shared_folder=None, previously_configured=False):
        dialog = tk.Toplevel(self.root)
        dialog.title("הגדרת עמדת ניהול - תיקיית רשת משותפת")
        dialog.geometry("820x320")
        try:
            dialog.minsize(800, 300)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        title_text = "ברוכים הבאים לעמדת הניהול" if first_run else "חיבור לתיקיית נתונים משותפת"
        tk.Label(
            dialog,
            text=title_text,
            font=('Arial', 16, 'bold'),
            bg='#ecf0f1'
        ).pack(pady=(15, 5))
        tk.Label(
            dialog,
            text=(
                "בחר או הדבק נתיב תיקיית הנתונים המשותפת של המערכת.\n"
                "אם אינך בטוח – אפשר פשוט ללחוץ על 'צור תיקיית שיתוף אוטומטית'."
            ),
            font=('Arial', 11),
            bg='#ecf0f1',
            justify='right'
        ).pack(pady=(0, 10))
        if previously_configured and previous_shared_folder:
            try:
                tk.Label(
                    dialog,
                    text=(
                        "שימו לב: המחשב כבר הוגדר בעבר. שינוי תיקיית הנתונים עלול ליצור פיצול נתונים.\n"
                        f"תיקייה קודמת: {previous_shared_folder}"
                    ),
                    font=('Arial', 9),
                    bg='#ecf0f1',
                    fg='#c0392b',
                    justify='right'
                ).pack(pady=(0, 8))
            except Exception:
                pass
        frame = tk.Frame(dialog, bg='#ecf0f1')
        frame.pack(fill=tk.X, padx=20, pady=5)
        shared_var = tk.StringVar(value=cfg.get('shared_folder') or cfg.get('network_root') or "")
        entry = tk.Entry(frame, textvariable=shared_var, font=('Arial', 11), width=40)
        entry.pack(side=tk.RIGHT, padx=5, pady=5, fill=tk.X, expand=True)

        def browse():
            folder = filedialog.askdirectory(title="בחר תיקיית רשת משותפת")
            if folder:
                shared_var.set(folder)

        def auto_create_share():
            folder = ""
            try:
                folder = shared_var.get().strip()
            except Exception:
                folder = ""
            if not folder:
                try:
                    system_drive = os.environ.get("SystemDrive") or "C:"
                except Exception:
                    system_drive = "C:"
                if not system_drive.endswith("\\") and not system_drive.endswith("/"):
                    system_drive = system_drive + "\\"
                try:
                    folder = os.path.join(system_drive, "SchoolPointsShare")
                except Exception:
                    folder = system_drive + "SchoolPointsShare"
            if folder.startswith("\\\\"):
                shared_var.set(folder)
                try:
                    messagebox.showinfo("תיקיית רשת קיימת", "נבחרה כבר תיקיית רשת משותפת (נתיב UNC).")
                except Exception:
                    pass
                return
            try:
                os.makedirs(folder, exist_ok=True)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן ליצור תיקייה משותפת:\n{e}")
                return

            share_name = "SchoolPoints$"
            try:
                computer = os.environ.get("COMPUTERNAME") or socket.gethostname() or ""
            except Exception:
                computer = ""
            unc_path = f"\\\\{computer}\\{share_name}" if computer else None

            def _run_share(cmd_args):
                try:
                    completed = subprocess.run(cmd_args, capture_output=True, text=True, shell=False)
                    if completed.returncode != 0:
                        return completed.stderr or completed.stdout or str(completed.returncode)
                    return None
                except Exception as e:
                    return str(e)

            base_cmd = ["net", "share", f"{share_name}={folder}"]
            cmd_args = base_cmd + ["/GRANT:Everyone,CHANGE"]
            error_text = _run_share(cmd_args)

            if error_text:
                fallback_error = _run_share(base_cmd)
                if not fallback_error:
                    error_text = None
                    cmd_args = base_cmd
                else:
                    error_text = fallback_error

            if error_text:
                msg_lines = [
                    "לא ניתן ליצור את השיתוף אוטומטית.",
                    "",
                    "בדרך כלל יש להפעיל את התוכנה או את חלון הפקודה עם הרשאות מנהל.",
                    "ניתן להריץ ידנית את הפקודה הבאה בחלון 'שורת הפקודה' כמנהל:",
                    "",
                    " ".join(cmd_args),
                    "",
                    f"שגיאה:\n{error_text}",
                ]
                msg = "\n".join(msg_lines)
                try:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(" ".join(cmd_args))
                except Exception:
                    pass
                try:
                    messagebox.showwarning("שיתוף לא נוצר", msg)
                except Exception:
                    pass
            else:
                if unc_path:
                    shared_var.set(unc_path)
                else:
                    shared_var.set(folder)
                try:
                    info_msg = "תיקיית השיתוף נוצרה בהצלחה."
                    if unc_path:
                        info_msg += f"\n\nנתיב השיתוף:\n{unc_path}"
                    messagebox.showinfo("שיתוף נוצר", info_msg)
                except Exception:
                    pass

        tk.Button(
            frame,
            text="בחר...",
            command=browse,
            font=('Arial', 10),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=4
        ).pack(side=tk.LEFT, padx=5)
        if first_run:
            tk.Button(
                frame,
                text="צור תיקיית שיתוף אוטומטית",
                command=auto_create_share,
                font=('Arial', 9),
                bg='#2ecc71',
                fg='white',
                padx=8,
                pady=4
            ).pack(side=tk.LEFT, padx=5)
        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=20)

        def open_initial_setup_guide():
            html_path = None
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                candidate = os.path.join(base_dir, "הוראות התקנה ראשוניות.html")
                if os.path.exists(candidate):
                    html_path = candidate
            except Exception:
                html_path = None
            if not html_path:
                try:
                    messagebox.showinfo("הוראות הגדרה", "קובץ 'הוראות התקנה ראשוניות.html' לא נמצא בתיקיית התוכנה.")
                except Exception:
                    pass
                return
            try:
                os.startfile(html_path)
            except Exception as e:
                try:
                    messagebox.showerror("שגיאה", f"לא ניתן לפתוח את קובץ ההוראות:\n{e}")
                except Exception:
                    pass

        def restart_as_admin():
            if _is_running_as_admin():
                try:
                    messagebox.showinfo("הרצה כמנהל", "היישום כבר פועל עם הרשאות מנהל.")
                except Exception:
                    pass
                return
            ok = _restart_as_admin()
            if not ok:
                try:
                    messagebox.showerror("שגיאה", "לא ניתן להפעיל מחדש עם הרשאות מנהל.\nניתן לנסות ללחוץ על הקיצור עם 'הפעל כמנהל'.")
                except Exception:
                    pass
                return
            try:
                dialog.destroy()
            except Exception:
                pass
            try:
                self.root.after(100, self.root.destroy)
            except Exception:
                try:
                    self.root.destroy()
                except Exception:
                    pass

        if first_run:
            tk.Button(
                btn_frame,
                text="הוראות הגדרה ראשוניות",
                command=open_initial_setup_guide,
                font=('Arial', 9),
                bg='#2980b9',
                fg='white',
                padx=10,
                pady=6
            ).pack(side=tk.LEFT, padx=10)
            tk.Button(
                btn_frame,
                text="הפעל מחדש כמנהל",
                command=restart_as_admin,
                font=('Arial', 10),
                bg='#e67e22',
                fg='white',
                padx=16,
                pady=6
            ).pack(side=tk.LEFT, padx=10)

        result = {'ok': False}

        def save_and_close():
            folder = shared_var.get().strip()
            if not folder:
                messagebox.showerror("שגיאה", "יש לבחור תיקיית רשת משותפת.")
                return

            if previously_configured and previous_shared_folder:
                try:
                    old_s = str(previous_shared_folder or '').strip()
                    new_s = str(folder or '').strip()
                except Exception:
                    old_s = str(previous_shared_folder or '')
                    new_s = str(folder or '')
                if old_s and new_s and old_s != new_s:
                    try:
                        ok = messagebox.askyesno(
                            "אישור שינוי תיקייה משותפת",
                            "המחשב כבר היה מחובר בעבר לתיקיית נתונים משותפת.\n\n"
                            "שינוי התיקייה המשותפת עלול ליצור פיצול נתונים (\"מערכת חדשה\"),\n"
                            "ולגרום לכך שעמדות אחרות ימשיכו לעבוד על נתונים ישנים.\n\n"
                            "האם לבצע שינוי בכל זאת?"
                        )
                    except Exception:
                        ok = False
                    if not ok:
                        return
            try:
                os.makedirs(folder, exist_ok=True)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן להשתמש בתיקייה:\n{e}")
                return

            # אם כבר קיימת תיקיית רשת עם config.json – נטען אותה ונעדיף את ההגדרות המשותפות
            base_cfg = dict(cfg) if isinstance(cfg, dict) else {}
            shared_cfg_path = os.path.join(folder, 'config.json')
            merged_cfg = None
            if os.path.exists(shared_cfg_path):
                try:
                    with open(shared_cfg_path, 'r', encoding='utf-8') as f:
                        shared_cfg = json.load(f)
                    if isinstance(shared_cfg, dict):
                        merged_cfg = dict(shared_cfg)
                        for k, v in base_cfg.items():
                            merged_cfg.setdefault(k, v)
                except Exception:
                    merged_cfg = None
            if merged_cfg is None:
                merged_cfg = base_cfg

            # טיפול במסד נתונים קיים – העתקה ראשונית לתיקיית הרשת אם צריך
            shared_db_path = os.path.join(folder, 'school_points.db')
            try:
                if os.path.exists(shared_db_path):
                    # אם כבר קיים DB בתיקיה המשותפת – רק נבדוק שניתן לעבוד איתו
                    from database import Database as _DBTest
                    _ = _DBTest(db_path=shared_db_path)
                elif current_db_path and os.path.exists(current_db_path) and os.path.abspath(current_db_path) != os.path.abspath(shared_db_path):
                    # אין DB ברשת אבל יש DB קיים מקומי – נעתיק אותו לתיקיית הרשת
                    try:
                        shutil.copy2(current_db_path, shared_db_path)
                    except Exception as copy_err:
                        messagebox.showerror(
                            "שגיאה",
                            f"לא ניתן להעתיק את מסד הנתונים הקיים לתיקיית הרשת:\n"
                            f"מ־{current_db_path}\nאל {shared_db_path}\n\n{copy_err}"
                        )
                        return
                else:
                    # אין DB קיים – ניצור DB חדש בתיקיית הרשת
                    from database import Database as _DBTest
                    _ = _DBTest(db_path=shared_db_path)
            except Exception as e:
                messagebox.showerror(
                    "שגיאה",
                    f"לא ניתן לעבוד עם מסד נתונים בתיקיית הרשת:\n{shared_db_path}\n\n{e}"
                )
                return

            try:
                self._seed_shared_sounds_folder(folder)
            except Exception:
                pass

            merged_cfg['shared_folder'] = folder
            if 'network_root' in merged_cfg:
                merged_cfg.pop('network_root', None)
            if not merged_cfg.get('excel_path'):
                merged_cfg['excel_path'] = os.path.join(folder, "טבלה למבצע אשראי.xlsx")
            if not self.save_app_config(merged_cfg):
                return
            self.app_config = merged_cfg
            result['ok'] = True
            dialog.destroy()

        def cancel():
            dialog.destroy()

        tk.Button(
            btn_frame,
            text="💾 שמור והמשך",
            command=save_and_close,
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=25,
            pady=8
        ).pack(side=tk.RIGHT, padx=10)
        tk.Button(
            btn_frame,
            text="ביטול",
            command=cancel,
            font=('Arial', 11),
            bg='#95a5a6',
            fg='white',
            padx=20,
            pady=8
        ).pack(side=tk.LEFT, padx=10)
        self.root.wait_window(dialog)
        return result['ok']

    def _open_initial_admin_dialog(self, db):
        teachers = db.get_all_teachers()
        if teachers:
            return True
        dialog = tk.Toplevel(self.root)
        dialog.title("יצירת מנהל ראשוני")
        dialog.geometry("600x350")
        try:
            dialog.minsize(600, 350)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        tk.Label(dialog, text="יצירת מנהל ראשוני", font=('Arial', 14, 'bold'), bg='#ecf0f1').pack(pady=15)
        tk.Label(dialog, text="שם המנהל:", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
        name_entry = tk.Entry(dialog, font=('Arial', 12), width=25, justify='right')
        name_entry.pack(pady=5)
        tk.Label(dialog, text="העבר כרטיס מאסטר (או הזן מספר כרטיס):", font=('Arial', 11), bg='#ecf0f1').pack(pady=5)
        card_entry = tk.Entry(dialog, font=('Arial', 12), width=25, justify='right', show='•')
        card_entry.pack(pady=5)
        tk.Label(dialog, text="המספר מוסתר (●●●●) לצורך פרטיות.", font=('Arial', 9), fg='#7f8c8d', bg='#ecf0f1').pack(pady=(0, 5))
        name_entry.focus_set()
        card_entry.focus_set()
        result = {'ok': False}

        def create_admin():
            name = name_entry.get().strip()
            card = card_entry.get().strip()
            if not (name and card):
                messagebox.showwarning("שגיאה", "יש למלא את כל השדות")
                return
            teacher_id = db.add_teacher(name, card, is_admin=True)
            if teacher_id:
                messagebox.showinfo("הצלחה", "מנהל נוצר בהצלחה.\nכעת ניתן להתחבר עם הכרטיס.")
                result['ok'] = True
                dialog.destroy()
            else:
                messagebox.showerror("שגיאה", "לא ניתן ליצור מנהל")

        tk.Button(
            dialog,
            text="צור מנהל",
            command=create_admin,
            font=('Arial', 12, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=20,
            pady=10
        ).pack(pady=15)
        self.root.wait_window(dialog)
        return result['ok']

    def open_help_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("הוראות שימוש - מערכת הניקוד")
        win.geometry("1000x760")
        try:
            win.minsize(880, 640)
        except Exception:
            pass
        win.configure(bg='#ecf0f1')
        win.resizable(True, True)
        win.transient(self.root)
        win.grab_set()

        container = tk.Frame(win, bg='#ecf0f1')
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        canvas = tk.Canvas(container, bg='#ecf0f1', highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='#ecf0f1')

        scroll_frame_id = canvas.create_window((0, 0), window=scroll_frame, anchor='nw')

        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            canvas.itemconfig(scroll_frame_id, width=event.width)

        scroll_frame.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)

        title = tk.Label(
            scroll_frame,
            text="📘 הוראות שימוש במערכת הניקוד",
            font=('Arial', 16, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        title.pack(fill=tk.X, pady=(0, 10))

        section1 = tk.Label(
            scroll_frame,
            text="1️⃣  רישום מערכת ורישיון",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section1.pack(fill=tk.X, pady=(5, 2))

        text1 = (
            "• רישום המערכת מתבצע רק בעמדת הניהול, על ידי מנהל.\n"
            "• תחילה יש להגדיר תיקיית רשת משותפת בחלון 'הגדרות מערכת' כדי שכל העמדות ישתמשו באותו קובץ רישיון ומסד נתונים.\n"
            "• לאחר מכן יש ללחוץ על 'רישום מערכת' בחלון ההגדרות. ברישיון פעיל יוצג סיכום בלבד, ובלחיצה על 'החלפת רישיון' ניתן להזין קוד חדש.\n"
            "• בגרסת ניסיון המערכת פועלת 7 ימים ומאפשרת עד 2 עמדות. לאחר מכן עמדת הניהול הופכת לקריאה בלבד והעמדה הציבורית נחסמת.\n"
            "• לאחר הפעלת רישיון (בסיסי / מורחב / ללא הגבלה) מספר העמדות המרבי נקבע לפי סוג הרישיון. כל מחשב שבו מופעלת עמדה נספר אוטומטית כל עוד לא חורגים מהמכסה."
        )
        tk.Label(
            scroll_frame,
            text=text1,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section2 = tk.Label(
            scroll_frame,
            text="2️⃣  הכנת נתונים וייבוא מאקסל",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section2.pack(fill=tk.X, pady=(5, 2))

        text2 = (
            "• יש להכין קובץ אקסל עם רשימת תלמידים לפי המבנה שנקבע למערכת (עמודות מזהה, שם פרטי, שם משפחה, כיתה, מספר כרטיס ועוד).\n"
            "• בלחיצה על כפתור '📥 ייבוא' בעמדת הניהול ניתן לבחור את קובץ האקסל, לסקור את הנתונים ולייבא אותם למערכת.\n"
            "• לאחר הייבוא הנתונים נשמרים במסד הנתונים הפנימי, והעמדה הציבורית מציגה את הנתונים מתוך המסד בלבד.\n"
            "• מומלץ לבצע ייצוא או סינכרון תקופתי לקובץ אקסל גיבוי, באמצעות כפתורי '📤 ייצוא' ו-'🔄 סינכרון'."
        )
        tk.Label(
            scroll_frame,
            text=text2,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section3 = tk.Label(
            scroll_frame,
            text="3️⃣  הודעות, צבעים וניקוד",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section3.pack(fill=tk.X, pady=(5, 2))

        text3 = (
            "• הודעות כלליות (📢) מנוהלות בחלון 'הודעות כלליות':\n"
            "  – 'הצג תמיד' – ההודעה מופיעה בעמדה הציבורית גם בלי שסורק כרטיס.\n"
            "  – 'עם כרטיס' – ההודעה מופיעה יחד עם נתוני התלמיד בכל סריקה.\n"
            "• הודעות לפי נקודות (🎯) מאפשרות להגדיר טווחי נקודות (0–10, 11–30, 31–100...) ולכל טווח טקסט מתאים. העמדה הציבורית בוחרת את ההודעה לפי מספר הנקודות של התלמיד.\n"
            "• הודעות אישיות (👤) לכל תלמיד מופיעות כשדה 'הודעה פרטית' בטבלת הניהול, ונראות בעמדה הציבורית רק לתלמיד שסורק את כרטיסו.\n"
            "• צבעי הניקוד בעמדה הציבורית נקבעים לפי טווחי נקודות בקובץ צבעים. בחלון '🎨 צבעים' ניתן להגדיר לכל טווח נקודות צבע טקסט נפרד (לדוגמה: אדום לנמוך, כתום לביניים, ירוק לגבוה)."
        )
        tk.Label(
            scroll_frame,
            text=text3,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section4 = tk.Label(
            scroll_frame,
            text="4️⃣  מצב בונוס, בונוס מיוחד ובונוס זמנים",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section4.pack(fill=tk.X, pady=(5, 2))

        text4 = (
            "• '🎁 בונוס מיוחד' מוגדר בעמדת הניהול באמצעות כפתור 'בונוס מיוחד' (מנהל בלבד):\n"
            "  – המנהל קובע כמה נקודות נוספות יקבל כל תלמיד בזמן שהבונוס המיוחד פעיל.\n"
            "  – בעמדה הציבורית מפעילים/מכבים את הבונוס המיוחד באמצעות כרטיס 'מאסטר 2' שהוגדר מראש.\n"
            "  – כל תלמיד יכול לקבל את נקודות הבונוס המיוחד פעם אחת בלבד בכל ריצת בונוס מיוחד.\n"
            "• '🎁 בונוס מורה' מאפשר לכל מורה להגדיר לעצמו כמה נקודות יקבל כל תלמיד בסבב בונוס שלו:\n"
            "  – המנהל יכול להגדיר בחלון 'הגדרות מערכת' מקסימום נקודות לבונוס מורה (לכל תלמיד בסבב).\n"
            "  – כל מורה מגדיר בעמדת הניהול שלו את כמות הנקודות לבונוס, בכפתור 'בונוס'.\n"
            "  – בעמדה הציבורית מפעילים/מכבים את בונוס המורה בעזרת כרטיס המורה; כל תלמיד מקבל את הבונוס פעם אחת בלבד בכל סבב.\n"
            "• '⏰ בונוס זמנים' מנוהל בחלון 'בונוס זמנים':\n"
            "  – מגדירים שם, שעת התחלה וסיום, וכמות נקודות.\n"
            "  – בזמן שבונוס זמנים פעיל, תלמיד שסורק כרטיס מקבל אוטומטית את נקודות הבונוס אם עדיין לא קיבל באותו יום, ומוצגת לו הודעת אישור מתאימה."
        )
        tk.Label(
            scroll_frame,
            text=text4,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section5 = tk.Label(
            scroll_frame,
            text="5️⃣  כרטיסי מאסטר והעמדה הציבורית",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section5.pack(fill=tk.X, pady=(5, 2))

        text5 = (
            "• כרטיס 'מאסטר' (🔑) מוגדר בעמדת הניהול ומשמש כקוד יציאה מהעמדה הציבורית: סריקתו פותחת חלון יציאה למפעיל.\n"
            "• כרטיס 'מאסטר 2' משמש להפעלה ולסיום של מצב 'בונוס מיוחד' בעמדה הציבורית.\n"
            "  – כאשר בונוס מיוחד פעיל, הודעת מצב מוצגת על המסך הציבורי ותלמידים מקבלים נקודות נוספות בעת סריקה.\n"
            "  – כאשר מסיימים את הבונוס המיוחד, מוצגת בעמדה הציבורית הודעה מסכמת עם מספר התלמידים שקיבלו את הבונוס המיוחד.\n"
            "• אם אין רישיון תקף ותקופת הניסיון הסתיימה – העמדה הציבורית לא תיפתח כלל, גם לא באמצעות כרטיסי מאסטר, ותוצג הודעה שמסבירה שיש להפעיל רישיון בעמדת הניהול."
        )
        tk.Label(
            scroll_frame,
            text=text5,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section6 = tk.Label(
            scroll_frame,
            text="6️⃣  רקעים, מצגת וערכת צבעים",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section6.pack(fill=tk.X, pady=(5, 2))

        text6 = (
            "• בחלון 'הגדרות מערכת' ניתן לבחור מצב רקע לעמדה הציבורית:\n"
            "  – ללא רקע מיוחד.\n"
            "  – תמונת רקע אחת.\n"
            "  – מצגת תמונות מתיקייה (תמונה אחת מתחלפת או מונטאז' ריבועים).\n"
            "  – צבע רקע אחיד שנבחר מטבלת צבעים.\n"
            "• ניתן גם לבחור פריסת רקע (מילוי, התאמה, מתיחה, אריח, מרכז) כך שהתמונה תתאים לגודל המסך בצורה נאה.\n"
            "• ערכת הצבעים הכללית של המסך הציבורי (רקע פנלים וצבעי טקסט בסיסיים) נבחרת מתוך רשימת ערכות מוכנות, כדי להבטיח ניגודיות וקריאות טובות לתלמידים."
        )
        tk.Label(
            scroll_frame,
            text=text6,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 8))

        section7 = tk.Label(
            scroll_frame,
            text="7️⃣  סיכום וקווים מנחים",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        )
        section7.pack(fill=tk.X, pady=(5, 2))

        text7 = (
            "• מומלץ להתחיל מהגדרת תיקיית הרשת והרישיון, ואז לייבא תלמידים מאקסל.\n"
            "• חשוב לעבוד תמיד מאותה תיקיית רשת כדי שכל העמדות ישתפו את אותו מסד נתונים ורישיון.\n"
            "• לפני שינוי רישיון מומלץ לוודא כמה עמדות יופעלו בפועל (עמדת ניהול + עמדות ציבוריות).\n"
            "• במידה ומופיעה שגיאה הקשורה לרישיון, לבונוס או לעמדות – מומלץ לצלם מסך ולפנות למנהל המערכת או לתמיכה."
        )
        tk.Label(
            scroll_frame,
            text=text7,
            font=('Arial', 10),
            bg='#ecf0f1',
            fg='#2c3e50',
            anchor='e',
            justify='right'
        ).pack(fill=tk.X, pady=(0, 10))

        btn_close = tk.Button(
            scroll_frame,
            text="✖ סגור",
            command=win.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=16,
            pady=6
        )
        btn_close.pack(pady=(0, 5))

        try:
            win.after(100, lambda: canvas.yview_moveto(0.0))
        except Exception:
            pass

    def open_help_text_dialog(self):
        # נסיון ראשון: לפתוח דף הוראות HTML חיצוני
        is_admin = bool(self.current_teacher and self.current_teacher['is_admin'] == 1)

        def try_build_embedded_guide():
            try:
                src_html = os.path.join(self.base_dir, "guide_user.html")
                if not os.path.exists(src_html):
                    return ""

                try:
                    config_file = self._get_config_file_path()
                    data_dir = os.path.dirname(config_file) or self.base_dir
                except Exception:
                    data_dir = self.base_dir

                out_html = os.path.join(data_dir, "guide_user_embedded.html")
                try:
                    if os.path.exists(out_html):
                        if os.path.getmtime(out_html) >= os.path.getmtime(src_html):
                            return out_html
                except Exception:
                    pass

                with open(src_html, 'r', encoding='utf-8') as f:
                    html = f.read()

                def repl(m):
                    src = m.group(1)
                    if not src:
                        return m.group(0)
                    if src.startswith('data:'):
                        return m.group(0)
                    if not (src.startswith('תמונות/להוראות/') or src.startswith('תמונות\\להוראות\\')):
                        return m.group(0)

                    rel = src.replace('/', os.sep).replace('\\', os.sep)
                    img_path = os.path.join(self.base_dir, rel)
                    if not os.path.exists(img_path):
                        return m.group(0)

                    ext = os.path.splitext(img_path)[1].lower()
                    mime = 'image/png'
                    if ext == '.jpg' or ext == '.jpeg':
                        mime = 'image/jpeg'
                    elif ext == '.gif':
                        mime = 'image/gif'
                    elif ext == '.svg':
                        mime = 'image/svg+xml'

                    try:
                        if mime == 'image/svg+xml':
                            with open(img_path, 'r', encoding='utf-8') as rf:
                                raw = rf.read().encode('utf-8')
                        else:
                            with open(img_path, 'rb') as rf:
                                raw = rf.read()
                        b64 = base64.b64encode(raw).decode('ascii')
                        return f'src="data:{mime};base64,{b64}"'
                    except Exception:
                        return m.group(0)

                html2 = re.sub(r'src\s*=\s*"([^"]+)"', repl, html)
                try:
                    os.makedirs(os.path.dirname(out_html), exist_ok=True)
                except Exception:
                    pass
                with open(out_html, 'w', encoding='utf-8') as f:
                    f.write(html2)

                try:
                    equip_src = os.path.join(self.base_dir, "equipment_required.html")
                    equip_dst = os.path.join(data_dir, "equipment_required.html")
                    if os.path.exists(equip_src):
                        if (not os.path.exists(equip_dst)) or (os.path.getmtime(equip_src) > os.path.getmtime(equip_dst)):
                            try:
                                shutil.copy2(equip_src, equip_dst)
                            except Exception:
                                pass
                    equip_dir_src = os.path.join(self.base_dir, "equipment_required_files")
                    equip_dir_dst = os.path.join(data_dir, "equipment_required_files")
                    if os.path.isdir(equip_dir_src):
                        try:
                            shutil.copytree(equip_dir_src, equip_dir_dst, dirs_exist_ok=True)
                        except Exception:
                            try:
                                if not os.path.exists(equip_dir_dst):
                                    os.makedirs(equip_dir_dst, exist_ok=True)
                                for root, _dirs, files in os.walk(equip_dir_src):
                                    rel = os.path.relpath(root, equip_dir_src)
                                    dst_root = equip_dir_dst if rel == '.' else os.path.join(equip_dir_dst, rel)
                                    os.makedirs(dst_root, exist_ok=True)
                                    for fn in files:
                                        try:
                                            shutil.copy2(os.path.join(root, fn), os.path.join(dst_root, fn))
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                except Exception:
                    pass
                return out_html
            except Exception:
                return ""

        embedded_path = try_build_embedded_guide()

        html_candidates = []
        if embedded_path:
            html_candidates.append(embedded_path)
        html_candidates += [
            os.path.join(self.base_dir, "guide_user.html"),
            os.path.join(self.base_dir, "guide_index.html"),
            os.path.join(self.base_dir, "guide_admin.html" if is_admin else "guide_teacher.html"),
        ]
        html_path = ""
        for nm in html_candidates:
            try:
                p = nm
                if p and os.path.exists(p):
                    html_path = p
                    break
            except Exception:
                continue
        try:
            if html_path and os.path.exists(html_path):
                try:
                    os.startfile(html_path)
                    return
                except Exception as e:
                    messagebox.showerror("שגיאה", f"לא ניתן לפתוח קובץ ההוראות:\n{html_path}\n\n{e}")
                    return
        except Exception:
            # אם משהו נכשל בניסיון לפתוח HTML – נמשיך לגרסת ההוראות הפנימית
            pass

        win = tk.Toplevel(self.root)
        win.title("הוראות שימוש - מערכת הניקוד")
        win.geometry("1060x840")
        try:
            win.minsize(920, 680)
        except Exception:
            pass
        win.configure(bg='#ecf0f1')
        win.resizable(True, True)
        win.transient(self.root)
        win.grab_set()

        # מסגרת גלילה עם קנבס ותת-מסגרת (תוויות בלבד – ללא בחירת טקסט)
        container = tk.Frame(win, bg='#ecf0f1')
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        canvas = tk.Canvas(container, bg='#ecf0f1', highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='#ecf0f1')

        # חשוב: עוגן 'nw' (פינה עליונה-שמאלית) כדי שהתוכן יוצג בתוך הקנבס ולא "יברח" שמאלה
        scroll_frame_id = canvas.create_window((0, 0), window=scroll_frame, anchor='nw')

        def on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_canvas_configure(event):
            # התאמת רוחב המסגרת הפנימית כך שהפריסה תישאר יציבה גם בשינוי גודל חלון
            canvas.itemconfig(scroll_frame_id, width=event.width)

        scroll_frame.bind("<Configure>", on_configure)
        canvas.bind("<Configure>", on_canvas_configure)

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)

        def _on_mousewheel(event):
            if event.delta:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _on_mousewheel_linux(event):
            if event.num == 4:
                canvas.yview_scroll(-3, "units")
            elif event.num == 5:
                canvas.yview_scroll(3, "units")

        def _bind_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_mousewheel_linux)
            canvas.bind_all("<Button-5>", _on_mousewheel_linux)

        def _unbind_from_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _bind_to_mousewheel)
        canvas.bind("<Leave>", _unbind_from_mousewheel)
        scroll_frame.bind("<Enter>", _bind_to_mousewheel)
        scroll_frame.bind("<Leave>", _unbind_from_mousewheel)

        # האם המשתמש הנוכחי הוא מנהל (בשביל ההוראות הפנימיות במקרה שאין PDF חיצוני)
        is_admin = bool(self.current_teacher and self.current_teacher['is_admin'] == 1)

        def add_main_title(text):
            lbl = tk.Label(
                scroll_frame,
                text=text,
                font=('Arial', 16, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50',
                anchor='e',
                justify='right'
            )
            lbl.pack(fill=tk.X, pady=(0, 10))

        def add_section_title(text, color="#2980b9", icon='◆'):
            frame = tk.Frame(scroll_frame, bg=color)
            frame.pack(fill=tk.X, pady=(10, 4))
            icon_lbl = tk.Label(
                frame,
                text=icon,
                font=('Arial', 12, 'bold'),
                bg=color,
                fg='white',
                width=2,
                anchor='e'
            )
            icon_lbl.pack(side=tk.RIGHT, padx=(0, 4))
            lbl = tk.Label(
                frame,
                text=text,
                font=('Arial', 12, 'bold'),
                bg=color,
                fg='white',
                anchor='e',
                justify='right'
            )
            lbl.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(0, 8), pady=4)

        def add_bullet(text, icon='•', icon_color='#2c3e50'):
            """שורת תבליט – האייקון ב-Label נפרד כך שתמיד יופיע בקצה הימני."""
            row = tk.Frame(scroll_frame, bg='#ecf0f1')
            row.pack(fill=tk.X, pady=1)
            bullet_lbl = tk.Label(
                row,
                text=icon,
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1',
                fg=icon_color,
                width=2,
                anchor='e'
            )
            bullet_lbl.pack(side=tk.RIGHT, padx=(0, 2))
            text_lbl = tk.Label(
                row,
                text=text,
                font=('Arial', 11),
                bg='#ecf0f1',
                fg='#2c3e50',
                anchor='e',
                justify='right',
                wraplength=780
            )
            text_lbl.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        def add_numbered(number_text, text, color='#8e44ad'):
            """שורה ממוספרת – המספר בצד ימין כ"אייקון" נפרד."""
            row = tk.Frame(scroll_frame, bg='#ecf0f1')
            row.pack(fill=tk.X, pady=(8, 2))
            num_lbl = tk.Label(
                row,
                text=number_text,
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1',
                fg=color,
                width=4,
                anchor='e'
            )
            num_lbl.pack(side=tk.RIGHT, padx=(0, 4))
            text_lbl = tk.Label(
                row,
                text=text,
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50',
                anchor='e',
                justify='right',
                wraplength=780
            )
            text_lbl.pack(side=tk.RIGHT, fill=tk.X, expand=True)

        # כותרת כללית
        if is_admin:
            add_main_title("מדריך שימוש – עמדת ניהול ומורה")
        else:
            add_main_title("מדריך שימוש – למורה")

        # === חלק א' – מנהל מערכת (מוצג רק למנהלים) ===
        if is_admin:
            add_section_title("חלק א – הוראות לעמדת ניהול (מנהל מערכת)", color='#2980b9', icon='◆')

            # 1. הגדרה ראשונית – תיקיית רשת וכרטיס מנהל
            add_numbered("1", "הגדרה ראשונית – תיקיית רשת וכרטיס מנהל", color='#e67e22')
            add_bullet("בהפעלה הראשונה של עמדת הניהול מופיע אשף הגדרה שמחייב לבחור תיקיית רשת משותפת ולהגדיר כרטיס מנהל ראשי (כרטיס כניסה לעמדת הניהול).", icon='■', icon_color='#e67e22')
            add_bullet("תיקיית הרשת המשותפת תשמש כמיקום המרכזי למסד הנתונים ולקובץ האקסל, וכל העמדות (ניהול וציבורית) צריכות להתחבר אליה.", icon='■', icon_color='#e67e22')
            add_bullet("בהמשך ניתן לשנות את תיקיית הרשת בחלון 'הגדרות מערכת'; בעת מעבר לתיקייה חדשה המערכת תנסה להעתיק אליה את מסד הנתונים הקיים כדי שלא יאבדו נתונים.", icon='■', icon_color='#e67e22')
            add_bullet("כרטיס המנהל הראשי שנקבע באשף ההפעלה הראשונית משמש להזדהות בעמדת הניהול; מומלץ להשתמש בכרטיס הנמצא באופן קבוע אצל מנהל המערכת.", icon='■', icon_color='#e67e22')

            # 2. מבנה קובץ Excel
            add_numbered("2", "מבנה מומלץ לקובץ Excel לייבוא תלמידים", color='#16a085')
            add_bullet("שורה ראשונה בקובץ צריכה להיות שורת כותרות (שמות העמודות).", icon='•', icon_color='#16a085')
            add_bullet("עמודות חובה ליבוא: מס' סידורי, שם משפחה, שם פרטי, כיתה.", icon='•', icon_color='#16a085')
            add_bullet("עמודות נוספות שאפשר (לא חובה) להוסיף: נתיב תמונה, מס' כרטיס, מס' נקודות, הודעה פרטית.", icon='•', icon_color='#16a085')
            add_bullet("אם בוחרים להגדיר באקסל עמודת מס' כרטיס, מומלץ שהיא תהיה בפורמט טקסט כדי לשמור על אפסים מובילים ומספרים ארוכים.", icon='•', icon_color='#16a085')
            add_bullet("אם בוחרים להגדיר באקסל עמודת מס' נקודות – זו צריכה להיות מספר שלם; אם התא ריק, המערכת תתחיל מ-0 נקודות.", icon='•', icon_color='#16a085')
            add_bullet("עמודת הודעה פרטית (אם קיימת) היא טקסט חופשי שמוצג לתלמיד בעמדה הציבורית כאשר הוא סורק את הכרטיס.", icon='•', icon_color='#16a085')
            add_bullet("לייבוא מלא של תלמידים משתמשים בכפתור 'ייבוא'; לסינכרון מהמערכת אל קובץ האקסל (ייצוא מהיר של מס' כרטיס, נקודות והודעה פרטית בלבד) משתמשים בכפתור 'סינכרון'.", icon='•', icon_color='#16a085')
            add_bullet("מומלץ, ולעיתים נוח יותר, לעדכן מס' כרטיס ונקודות מתוך מסך הניהול אחרי הייבוא הראשוני של רשימת התלמידים.", icon='•', icon_color='#16a085')

            # 3. כפתורי המסך הראשי
            add_numbered("3", "כפתורי המסך הראשי בעמדת הניהול", color='#8e44ad')
            add_bullet("כפתור 'ייבוא' – טעינת רשימת תלמידים מקובץ אקסל אל המערכת.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'ייצוא' – כתיבת הנתונים מהמערכת חזרה לאקסל לצורך גיבוי או דוחות.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'סינכרון' – עדכון מהיר בין מסד הנתונים לקובץ האקסל בעמודות הנקודות וההודעות בלבד.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'רענון' – ריענון רשימת התלמידים מהמסד, כולל עדכונים מהעמדה הציבורית.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'מטבעות וצבעים' – פתיחת עורך צבעי הניקוד ומטבעות/יהלומים המבוססים על נקודות בעמדה הציבורית.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'הודעות כלליות' – פתיחת מנהל ההודעות הסטטיות, הודעות לפי נקודות והודעות אישיות (כולל כרטיסייה 'חדשות' לניהול הטקסטים של הטיקר בתחתית העמדה הציבורית).", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'בונוס מיוחד' – הגדרת מצב בונוס מיוחד וכמות הנקודות שכל תלמיד יקבל בזמן שריצת הבונוס המיוחד פעילה (מופעל בעמדה הציבורית עם כרטיס מאסטר 2).", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'בונוס זמנים' – ניהול לוחות זמנים של בונוסים אוטומטיים לפי שעות.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'ניהול מורים' – הגדרת מורים, כיתות והרשאות גישה (כרטיס מורה, האם המורה מנהל, אילו כיתות יוכל לראות ומה מותר לו לעדכן).", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'הגדרות מערכת' – רישיון, תיקיית רשת, לוגו, שם מבצע, רקעים, מצגת וסטטיסטיקות.", icon='▶', icon_color='#8e44ad')
            add_bullet("כפתור 'הגדרות תצוגה' – כיוון מסך (אורך/רוחב), ערכת צבעים, מצב רקע (תמונה אחת, מצגת, צבע אחיד), פריסת רקע וסגנון הפנלים בעמדה הציבורית.", icon='▶', icon_color='#8e44ad')
            add_bullet("מעל טבלת התלמידים מופיעים כפתורי '➕ תלמיד', '✏ ערוך' ו'🗑 מחק' המאפשרים להוסיף תלמיד חדש, לערוך פרטי תלמיד קיים או למחוק תלמיד (כולל היסטוריית הנקודות והתיקופים שלו).", icon='▶', icon_color='#8e44ad')

            # 4. הודעות, צבעים ובונוסים
            add_numbered("4", "הודעות, צבעים ובונוסים", color='#c0392b')
            add_bullet("הודעות כלליות – הודעות שמופיעות לכל התלמידים; ניתן להגדיר הצגה תמידית או רק עם כרטיס.", icon='◆', icon_color='#c0392b')
            add_bullet("הודעות לפי נקודות – הודעות שונות לטווחי ניקוד שונים; העמדה הציבורית בוחרת אוטומטית את ההודעה לפי מספר הנקודות של התלמיד.", icon='◆', icon_color='#c0392b')
            add_bullet("הודעות אישיות – הודעות פרטיות לתלמיד, שמופיעות רק כשהוא סורק את הכרטיס.", icon='◆', icon_color='#c0392b')
            add_bullet("צבעי ניקוד – בחלון 'מטבעות וצבעים' מגדירים לכל טווח נקודות צבע טקסט שונה בעמדה הציבורית.", icon='◆', icon_color='#c0392b')
            add_bullet("באותו חלון ניתן להגדיר גם מטבעות ויהלומים מבוססי נקודות; בעמדה הציבורית הם מוצגים לפי הערכים שהוגדרו, כך שהערכים הגבוהים יותר מופיעים בצד ימין והרצף מסודר מהגדול לקטן.", icon='◆', icon_color='#c0392b')
            add_bullet("המטבעות והיהלומים נועדו למנוע מצב שתלמידים מגיעים למספרי נקודות של אלפים רבים שכבר קשה להרגיש בהם; במקום מספר ארוך הם רואים כמה 'אוצר' צברו ברמות ברורות (יהלומים ומטבעות).", icon='◆', icon_color='#c0392b')
            add_bullet("המלצה פדגוגית: לא להפעיל בבת אחת את כל סוגי הבונוסים, המצגת, החדשות ושאר האפשרויות, אלא להוסיף אותם בהדרגה במהלך השנה – כך תמיד נשאר 'קלף' חדש להפתיע בו את התלמידים ולשמור על עניין ומוטיבציה.", icon='◆', icon_color='#c0392b')
            add_bullet("מצב 'בונוס מיוחד' – המנהל קובע כמה נקודות נוספות יקבל כל תלמיד בזמן שהבונוס המיוחד פעיל; הפעלה וסיום נעשים בעמדה הציבורית בעזרת כרטיס מאסטר 2, וכל תלמיד מקבל את הבונוס פעם אחת בכל ריצה.", icon='◆', icon_color='#c0392b')
            add_bullet("'בונוס זמנים' – בונוסים אוטומטיים לפי שעות; תלמיד יכול לקבל כל בונוס זמן פעם אחת בלבד ביום.", icon='◆', icon_color='#c0392b')

            # 5. כרטיסי מאסטר והעמדה הציבורית
            add_numbered("5", "כרטיסי מאסטר והעמדה הציבורית", color='#34495e')
            add_bullet("כרטיס מאסטר 1 – קוד יציאה מהעמדה הציבורית; סריקתו פותחת חלון יציאה למפעיל.", icon='■', icon_color='#34495e')
            add_bullet("כרטיס מאסטר 2 – הפעלה וסיום של מצב הבונוס הרגיל בעמדה הציבורית.", icon='■', icon_color='#34495e')
            add_bullet("העמדה הציבורית מציגה נקודות, הודעות כלליות, הודעות לפי נקודות והודעות אישיות לכל תלמיד שסורק כרטיס.", icon='■', icon_color='#34495e')
            add_bullet("הלוגו, שם המבצע, תיקיית תמונות תלמידים, רקעי המסך, האם להציג סטטיסטיקות והאם להציג תמונת תלמיד – כולם נקבעים בחלון 'הגדרות מערכת'.", icon='■', icon_color='#34495e')
            add_bullet("במצב 'מצגת תמונות' אפשר לבחור אם להציג תמונה אחת מתחלפת או מונטאז' ריבועים (סטטי או מתחלף), וכן לקבוע את מספר העמודות במונטאז'; העמדה מחשבת אוטומטית כמה שורות מלאות ניתן להציג כך שכל התמונות יוצגו בשלמותן מעל רצועת החדשות.", icon='■', icon_color='#34495e')
            add_bullet("שדה 'פריסת רקע' קובע איך התמונה/הריבועים ממלאים את המסך: 'מילוי' (cover) למילוי מלא גם במחיר חיתוך קל בשוליים, 'התאמה' (contain) לשמירה מלאה על התמונה עם שוליים במידת הצורך, ו'מתח' (stretch) למתיחה מדויקת למסך.", icon='■', icon_color='#34495e')
            add_bullet("בסיום תקופת ניסיון ללא רישיון פעיל, העמדה הציבורית לא תיפתח ותופיע הודעה שיש להפעיל רישיון.", icon='■', icon_color='#34495e')

            # 6. רישום מערכת ורישיון (לביצוע לאחר ההגדרה הראשונית)
            add_numbered("6", "רישום מערכת ורישיון", color='#e67e22')
            add_bullet("רישום המערכת (הכנסת הרישיון) מתבצע רק בעמדת הניהול על ידי מנהל.", icon='■', icon_color='#e67e22')
            try:
                row = tk.Frame(scroll_frame, bg='#ecf0f1')
                row.pack(fill=tk.X, pady=1)
                bullet_lbl = tk.Label(
                    row,
                    text='■',
                    font=('Arial', 11, 'bold'),
                    bg='#ecf0f1',
                    fg='#e67e22',
                    width=2,
                    anchor='e'
                )
                bullet_lbl.pack(side=tk.RIGHT, padx=(0, 2))

                prefix_lbl = tk.Label(
                    row,
                    text="לביצוע רישום יש להיכנס ",
                    font=('Arial', 11),
                    bg='#ecf0f1',
                    fg='#2c3e50',
                    anchor='e',
                    justify='right'
                )
                prefix_lbl.pack(side=tk.RIGHT)

                link_lbl = tk.Label(
                    row,
                    text="כאן",
                    font=('Arial', 11, 'bold'),
                    bg='#ecf0f1',
                    fg='#2980b9',
                    cursor='hand2',
                    anchor='e'
                )
                link_lbl.pack(side=tk.RIGHT)
                link_lbl.bind('<Button-1>', lambda _e: webbrowser.open('https://forms.gle/hE6DAmyHabDxKnf48'))
            except Exception:
                pass
            add_bullet("בחלון 'רישום מערכת' מוצגים שם המוסד וקוד המערכת של המחשב המרכזי; לאחר שתעבירו אלינו את שם המוסד וקוד המערכת, תקבלו מאיתנו קוד הפעלה מותאם למחשב זה, אותו יש להזין בחלון זה.", icon='■', icon_color='#e67e22')
            add_bullet("קוד ההפעלה ניתן לשימוש רק על אותו מחשב שבו הופק קוד המערכת; מומלץ שמחשב זה יהיה המחשב המרכזי המחובר לתיקיית הרשת המשותפת, ולא מחשב אישי זמני.", icon='■', icon_color='#e67e22')
            add_bullet("בגרסת ניסיון המערכת פועלת 7 ימים ומאפשרת עד שתי עמדות;\nלאחר מכן עמדת הניהול עוברת למצב צפייה בלבד והעמדה הציבורית נחסמת.", icon='■', icon_color='#e67e22')
            add_bullet("ברישיון פעיל מספר העמדות המרבי נקבע לפי סוג הרישיון (עד 2 / עד 5 / ללא הגבלה),\nוכל מחשב שבו מופעלת עמדה נספר אוטומטית עד למכסה.", icon='■', icon_color='#e67e22')

        # === חלק ב' – הוראות למורה (מוצג לכל המשתמשים) ===
        add_section_title("חלק ב – הוראות למורה", color='#27ae60', icon='◆')

        # 1. כניסה וסקירת מסך
        add_numbered("1", "כניסה למערכת כמורה", color='#27ae60')
        add_bullet("כדי להזדהות בעמדת הניהול/עמדת המורה מעבירים את כרטיס המורה מול קורא הכרטיסים; אם אין קורא כרטיסים במחשב, ניתן להזדהות באמצעות סיסמה שהוגדרה מראש במקום כרטיס.", icon='•', icon_color='#27ae60')
        add_bullet("לאחר הכניסה מוצגת רשימת תלמידים רק של הכיתות שהוקצו למורה זה.", icon='•', icon_color='#27ae60')

        # 2. עבודה עם רשימת התלמידים
        add_numbered("2", "עבודה עם רשימת התלמידים", color='#27ae60')
        add_bullet("ניתן להשתמש בתיבת החיפוש העליונה כדי לסנן תלמידים לפי שם או כיתה.", icon='•', icon_color='#27ae60')
        add_bullet("בעמודת הנקודות ניתן לעדכן נקודות לתלמידים: להוסיף נקודות חיוביות או להפחית לפי הצורך.", icon='•', icon_color='#27ae60')
        add_bullet("בעמודת 'הודעה פרטית' ניתן לכתוב הערה אישית לתלמיד; ההודעה תופיע לו בעמדה הציבורית בעת סריקת הכרטיס.", icon='•', icon_color='#27ae60')

        # 3. מה רואה התלמיד בעמדה הציבורית
        add_numbered("3", "מה רואה התלמיד בעמדה הציבורית", color='#27ae60')
        add_bullet("כאשר תלמיד סורק כרטיס, הוא רואה את מספר הנקודות שלו בצבע התואם לטווח הניקוד שהוגדר.", icon='•', icon_color='#27ae60')
        add_bullet("אם קיימת הודעה כללית פעילה, היא תופיע לכל התלמידים בהתאם להגדרה.", icon='•', icon_color='#27ae60')
        add_bullet("הודעה לפי נקודות תופיע לפי מספר הנקודות המצטבר של התלמיד.", icon='•', icon_color='#27ae60')
        add_bullet("אם למורה הוגדרה הודעה אישית לתלמיד, היא תופיע רק כשהתלמיד סורק את הכרטיס.", icon='•', icon_color='#27ae60')

        # 4. בונוס למורה וסדר עבודה מומלץ
        add_numbered("4", "בונוס למורה וסדר עבודה מומלץ", color='#27ae60')
        add_bullet("המורה יכול להגדיר לעצמו 'בונוס מורה' לתלמידים בעזרת כפתור '🎁 בונוס מורה' בעמדת הניהול – שם קובעים כמה נקודות יקבל כל תלמיד בסבב הבונוס הבא (בכפוף למגבלת המקסימום שקבע המנהל).", icon='•', icon_color='#27ae60')
        add_bullet("להפעלת הבונוס בפועל בעמדה הציבורית – המורה ניגש לעמדה, מציג את כרטיס המורה שלו (או מזין סיסמת מורה במקום כרטיס), ופותח בכך מצב 'בונוס מורה'. כל תלמיד מכיתות המורה שסורק כרטיס בזמן שהמצב פעיל מקבל את הבונוס פעם אחת בלבד בסבב.", icon='•', icon_color='#27ae60')
        add_bullet("תלמידים מכיתות אחרות (שאינם משויכים למורה זה בעמדת הניהול) אינם מקבלים את בונוס המורה גם אם סרקו כרטיס בזמן שהמצב פעיל.", icon='•', icon_color='#27ae60')
        add_bullet("בסיום הסבב, המורה מציג שוב את כרטיס המורה (או מזין את סיסמת המורה) בעמדה הציבורית כדי לסיים את מצב 'בונוס מורה'.", icon='•', icon_color='#27ae60')
        add_bullet("במהלך יום הלימודים לעדכן נקודות לתלמידים בהתאם להתנהגות ועמידה ביעדים, ולהוסיף בונוס במידת הצורך.", icon='•', icon_color='#27ae60')
        add_bullet("להשתמש בהודעות האישיות כדי לחזק תלמידים, להציב מטרות או להעביר מסר קצר.", icon='•', icon_color='#27ae60')
        add_bullet("לעודד תלמידים לסרוק את הכרטיס בעמדה הציבורית כדי לראות את ההתקדמות שלהם, את ההודעות ואת השפעת הבונוסים.", icon='•', icon_color='#27ae60')
        add_bullet("במידת הצורך לפנות למנהל המערכת לגבי בעיות רישיון, רקעים, צבעים או הגדרות בונוסים מערכתיים.", icon='•', icon_color='#27ae60')

        # כפתור סגירה
        close_frame = tk.Frame(scroll_frame, bg='#ecf0f1')
        close_frame.pack(fill=tk.X, pady=(12, 4))
        btn_close = tk.Button(
            close_frame,
            text="✖ סגור",
            command=win.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=16,
            pady=4
        )
        btn_close.pack(pady=(0, 4))

    def _normalize_quiet_time(self, value: str) -> str:
        try:
            s = str(value or '').strip().replace('.', ':')
        except Exception:
            return ''
        if not s:
            return ''
        parts = s.split(':')
        if len(parts) != 2:
            return s
        try:
            hh = int(parts[0])
            mm = int(parts[1])
            if hh < 0 or hh > 23 or mm < 0 or mm > 59:
                return s
            return f"{hh:02d}:{mm:02d}"
        except Exception:
            return s

    def _quiet_time_to_minutes(self, value: str):
        try:
            s = self._normalize_quiet_time(value)
            if not s:
                return None
            parts = s.split(':')
            if len(parts) != 2:
                return None
            hh = int(parts[0])
            mm = int(parts[1])
            if hh < 0 or hh > 23 or mm < 0 or mm > 59:
                return None
            return hh * 60 + mm
        except Exception:
            return None

    def _load_quiet_mode_ranges(self, cfg: dict) -> list:
        c = cfg if isinstance(cfg, dict) else {}
        raw = c.get('quiet_mode_ranges')
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                raw = None
        if isinstance(raw, dict):
            raw = [raw]
        ranges = list(raw) if isinstance(raw, list) else []

        if not ranges:
            enabled_raw = str(c.get('quiet_mode_enabled', '0')).strip().lower()
            enabled = enabled_raw in ('1', 'true', 'yes', 'on')
            if enabled:
                ranges = [{
                    'start': c.get('quiet_mode_start', ''),
                    'end': c.get('quiet_mode_end', ''),
                    'mode': 'low',
                    'volume': c.get('quiet_mode_volume', 20)
                }]

        cleaned = []
        for r in ranges:
            if not isinstance(r, dict):
                continue
            start = self._normalize_quiet_time(r.get('start') or r.get('start_time') or '')
            end = self._normalize_quiet_time(r.get('end') or r.get('end_time') or '')
            if not start or not end:
                continue
            mode_raw = str(r.get('mode') or '').strip().lower()
            mode = 'silent' if mode_raw in ('silent', 'quiet', 'mute', 'שקט') else 'low'
            try:
                vol_int = int(float(str(r.get('volume', 20)).strip()))
            except Exception:
                vol_int = 20
            vol_int = max(0, min(100, vol_int))
            cleaned.append({'start': start, 'end': end, 'mode': mode, 'volume': vol_int})
        return cleaned

    def _quiet_ranges_summary(self, ranges: list) -> str:
        try:
            count = len(ranges) if ranges else 0
        except Exception:
            count = 0
        return "ללא טווחים" if count == 0 else f"{count} טווחים"

    def _open_quiet_mode_manager_dialog(self, parent, ranges: list) -> list:
        ranges = list(ranges) if isinstance(ranges, list) else []
        result = list(ranges)

        try:
            return self._open_quiet_mode_manager_dialog_form(parent, ranges)
        except Exception:
            pass

        dlg = tk.Toplevel(parent)
        dlg.title("🌙 ניהול מצב שקט")
        dlg.configure(bg='#ecf0f1')
        dlg.transient(parent)
        dlg.grab_set()
        dlg.resizable(True, True)
        try:
            dlg.minsize(620, 380)
        except Exception:
            pass

        tk.Label(
            dlg,
            text=fix_rtl_text("טווחי שעות למצב שקט"),
            font=('Arial', 13, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(12, 4))

        tk.Label(
            dlg,
            text=fix_rtl_text("פורמט: 08:30-10:15, שקט, 0 | 13:00-14:00, ווליום נמוך, 25"),
            font=('Arial', 9),
            bg='#ecf0f1',
            fg='#7f8c8d'
        ).pack(pady=(0, 8))

        text_box = tk.Text(dlg, height=10, font=('Arial', 10), wrap='none')
        text_box.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))

        def _format_line(entry: dict) -> str:
            start = self._normalize_quiet_time(entry.get('start') or '')
            end = self._normalize_quiet_time(entry.get('end') or '')
            mode = str(entry.get('mode') or 'low').strip().lower()
            mode_label = 'שקט' if mode in ('silent', 'quiet', 'mute', 'שקט') else 'ווליום נמוך'
            try:
                vol = int(entry.get('volume', 20) or 20)
            except Exception:
                vol = 20
            if mode_label == 'שקט':
                vol = 0
            return f"{start}-{end}, {mode_label}, {vol}"

        try:
            lines = [_format_line(r) for r in ranges if isinstance(r, dict)]
        except Exception:
            lines = []
        if not lines:
            lines = ["08:00-09:00, שקט, 0"]
        text_box.insert('1.0', "\n".join(lines))

        def _add_default_line():
            text_box.insert(tk.END, "\n08:00-09:00, שקט, 0")

        btn_frame = tk.Frame(dlg, bg='#ecf0f1')
        btn_frame.pack(fill=tk.X, padx=16, pady=(0, 12))

        def _parse_lines():
            parsed = []
            invalid = 0
            raw_text = text_box.get('1.0', tk.END)
            for line in raw_text.splitlines():
                raw = str(line or '').strip()
                if not raw:
                    continue
                parts = [p.strip() for p in re.split(r"[,|]", raw) if p.strip()]
                if not parts:
                    continue
                time_part = parts[0]
                if '-' in time_part:
                    start_raw, end_raw = time_part.split('-', 1)
                elif 'עד' in time_part:
                    start_raw, end_raw = time_part.split('עד', 1)
                else:
                    invalid += 1
                    continue
                start = self._normalize_quiet_time(start_raw.strip())
                end = self._normalize_quiet_time(end_raw.strip())
                if not start or not end:
                    invalid += 1
                    continue
                mode_text = parts[1] if len(parts) > 1 else ''
                mode_low = str(mode_text or '').strip().lower()
                mode = 'silent' if any(k in mode_low for k in ('silent', 'quiet', 'mute', 'שקט')) else 'low'
                vol = 0 if mode == 'silent' else 20
                if len(parts) > 2:
                    try:
                        vol = int(float(parts[2]))
                    except Exception:
                        vol = 0 if mode == 'silent' else 20
                vol = max(0, min(100, int(vol)))
                if mode == 'silent':
                    vol = 0
                parsed.append({'start': start, 'end': end, 'mode': mode, 'volume': vol})
            return parsed, invalid

        def _on_save():
            nonlocal result
            parsed, invalid = _parse_lines()
            if not parsed and invalid > 0:
                messagebox.showwarning("שגיאה", "לא נמצאו טווחים תקינים. בדוק את הפורמט.")
                return
            result = parsed
            dlg.destroy()

        def _on_cancel():
            dlg.destroy()

        tk.Button(
            btn_frame,
            text="➕ הוסף טווח",
            command=_add_default_line,
            font=('Arial', 9, 'bold'),
            bg='#3498db',
            fg='white',
            padx=8,
            pady=4
        ).pack(side=tk.LEFT)

        tk.Button(
            btn_frame,
            text="שמירה",
            command=_on_save,
            font=('Arial', 9, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT, padx=(0, 6))

        tk.Button(
            btn_frame,
            text="ביטול",
            command=_on_cancel,
            font=('Arial', 9, 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT)

        dlg.wait_window()
        return result

    def _open_quiet_mode_manager_dialog_form(self, parent, ranges: list) -> list:
        ranges = list(ranges) if isinstance(ranges, list) else []
        result = list(ranges)

        dlg = tk.Toplevel(parent)
        dlg.title("🌙 ניהול מצב שקט")
        dlg.configure(bg='#ecf0f1')
        dlg.transient(parent)
        dlg.grab_set()
        dlg.resizable(True, True)
        try:
            dlg.minsize(720, 420)
        except Exception:
            pass

        tk.Label(
            dlg,
            text=fix_rtl_text("טווחי שעות למצב שקט"),
            font=('Arial', 14, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(12, 6))

        tk.Label(
            dlg,
            text=fix_rtl_text("בחר טווחי שעות והגדר אם זה שקט מלא או ווליום נמוך."),
            font=('Arial', 9),
            bg='#ecf0f1',
            fg='#7f8c8d'
        ).pack(pady=(0, 8))

        header = tk.Frame(dlg, bg='#ecf0f1')
        header.pack(fill=tk.X, padx=16)
        tk.Label(header, text=fix_rtl_text("משעה"), font=('Arial', 10, 'bold'), bg='#ecf0f1', width=8).pack(side=tk.RIGHT, padx=4)
        tk.Label(header, text=fix_rtl_text("עד שעה"), font=('Arial', 10, 'bold'), bg='#ecf0f1', width=8).pack(side=tk.RIGHT, padx=4)
        tk.Label(header, text=fix_rtl_text("מצב"), font=('Arial', 10, 'bold'), bg='#ecf0f1', width=12).pack(side=tk.RIGHT, padx=4)
        tk.Label(header, text=fix_rtl_text("ווליום"), font=('Arial', 10, 'bold'), bg='#ecf0f1', width=10).pack(side=tk.RIGHT, padx=4)
        tk.Label(header, text="", font=('Arial', 10, 'bold'), bg='#ecf0f1', width=6).pack(side=tk.RIGHT, padx=4)

        rows_container = tk.Frame(dlg, bg='#ecf0f1')
        rows_container.pack(fill=tk.BOTH, expand=True, padx=16, pady=(6, 8))

        row_widgets = []

        def _sync_volume_state(mode_var, volume_var, volume_spin):
            try:
                mode_val = str(mode_var.get() or '').strip()
            except Exception:
                mode_val = ''
            if mode_val == 'שקט':
                try:
                    volume_var.set(0)
                except Exception:
                    pass
                try:
                    volume_spin.config(state='disabled')
                except Exception:
                    pass
            else:
                try:
                    volume_spin.config(state='normal')
                except Exception:
                    pass

        def _remove_row(info: dict):
            try:
                info['frame'].destroy()
            except Exception:
                pass
            try:
                row_widgets.remove(info)
            except Exception:
                pass

        def _add_row(initial=None):
            initial = initial or {}
            row = tk.Frame(rows_container, bg='#ecf0f1')
            row.pack(fill=tk.X, pady=4)

            start_var = tk.StringVar(value=self._normalize_quiet_time(initial.get('start') or '08:00'))
            end_var = tk.StringVar(value=self._normalize_quiet_time(initial.get('end') or '09:00'))
            mode_raw = str(initial.get('mode') or 'silent').strip().lower()
            mode_var = tk.StringVar(value='שקט' if mode_raw in ('silent', 'quiet', 'mute', 'שקט') else 'ווליום נמוך')
            try:
                vol_default = int(initial.get('volume', 20) or 20)
            except Exception:
                vol_default = 20
            if mode_var.get() == 'שקט':
                vol_default = 0
            volume_var = tk.IntVar(value=max(0, min(100, int(vol_default))))

            tk.Entry(row, textvariable=start_var, font=('Arial', 10), width=8, justify='right').pack(side=tk.RIGHT, padx=4)
            tk.Entry(row, textvariable=end_var, font=('Arial', 10), width=8, justify='right').pack(side=tk.RIGHT, padx=4)

            mode_combo = ttk.Combobox(
                row,
                values=['שקט', 'ווליום נמוך'],
                state='readonly',
                width=12,
                textvariable=mode_var,
                justify='center'
            )
            mode_combo.pack(side=tk.RIGHT, padx=4)

            volume_spin = tk.Spinbox(
                row,
                from_=0,
                to=100,
                textvariable=volume_var,
                width=6,
                justify='center',
                font=('Arial', 10)
            )
            volume_spin.pack(side=tk.RIGHT, padx=4)

            row_info = {
                'frame': row,
                'start_var': start_var,
                'end_var': end_var,
                'mode_var': mode_var,
                'volume_var': volume_var,
                'volume_spin': volume_spin
            }
            remove_btn = tk.Button(
                row,
                text='✖',
                command=lambda info=row_info: _remove_row(info),
                font=('Arial', 9, 'bold'),
                bg='#e74c3c',
                fg='white',
                padx=6,
                pady=2
            )
            remove_btn.pack(side=tk.RIGHT, padx=4)

            mode_combo.bind('<<ComboboxSelected>>', lambda _evt: _sync_volume_state(mode_var, volume_var, volume_spin))
            _sync_volume_state(mode_var, volume_var, volume_spin)

            row_widgets.append(row_info)
            return row_info

        for entry in ranges:
            if isinstance(entry, dict):
                _add_row(entry)

        if not row_widgets:
            _add_row({'start': '08:00', 'end': '09:00', 'mode': 'silent', 'volume': 0})

        def _parse_rows():
            parsed = []
            invalid = 0
            for info in list(row_widgets):
                start = self._normalize_quiet_time(info['start_var'].get())
                end = self._normalize_quiet_time(info['end_var'].get())
                if not start or not end:
                    invalid += 1
                    continue
                if self._quiet_time_to_minutes(start) is None or self._quiet_time_to_minutes(end) is None:
                    invalid += 1
                    continue
                mode_label = str(info['mode_var'].get() or '').strip()
                mode = 'silent' if mode_label == 'שקט' else 'low'
                try:
                    vol = int(info['volume_var'].get())
                except Exception:
                    vol = 20
                vol = max(0, min(100, int(vol)))
                if mode == 'silent':
                    vol = 0
                parsed.append({'start': start, 'end': end, 'mode': mode, 'volume': vol})
            return parsed, invalid

        def _on_save():
            nonlocal result
            parsed, invalid = _parse_rows()
            if not parsed and invalid > 0:
                messagebox.showwarning("שגיאה", "יש שורות לא תקינות. בדוק שעות בפורמט HH:MM.")
                return
            result = parsed
            dlg.destroy()

        def _on_cancel():
            dlg.destroy()

        btn_frame = tk.Frame(dlg, bg='#ecf0f1')
        btn_frame.pack(fill=tk.X, padx=16, pady=(4, 12))

        tk.Button(
            btn_frame,
            text="➕ הוסף טווח",
            command=lambda: _add_row({'start': '08:00', 'end': '09:00', 'mode': 'silent', 'volume': 0}),
            font=('Arial', 9, 'bold'),
            bg='#3498db',
            fg='white',
            padx=8,
            pady=4
        ).pack(side=tk.LEFT)

        tk.Button(
            btn_frame,
            text="שמירה",
            command=_on_save,
            font=('Arial', 9, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT, padx=(0, 6))

        tk.Button(
            btn_frame,
            text="ביטול",
            command=_on_cancel,
            font=('Arial', 9, 'bold'),
            bg='#95a5a6',
            fg='white',
            padx=12,
            pady=4
        ).pack(side=tk.RIGHT)

        dlg.wait_window()
        return result

    def open_system_settings(self):
        """חלון הגדרות מערכת (תיקיית רשת, לוגו, שם מבצע)"""
        # רק למנהלים
        if not (self.current_teacher and self.current_teacher['is_admin'] == 1):
            messagebox.showwarning("אזהרה", "רק מנהלים יכולים לשנות הגדרות מערכת")
            return
        
        config = self.load_app_config()
        
        dialog = tk.Toplevel(self.root)
        dialog.title("⚙ הגדרות מערכת")
        # רוחב מינימלי קבוע, גובה משתנה לפי מספר השדות
        dialog.configure(bg='#ecf0f1')
        dialog.minsize(700, 470)
        dialog.geometry("760x520")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        # מסגרת מרכזית לכל שורות השדות (מעל כפתורי שמור/סגור)
        content_frame = tk.Frame(dialog, bg='#ecf0f1')
        # ממלא לרוחב עם שוליים סימטריים כך שכל השורות ישתמשו באותו רוחב
        content_frame.pack(fill=tk.X, padx=20, pady=8)
        
        # שדות
        shared_folder_var = tk.StringVar(value=config.get('shared_folder') or config.get('network_root') or "")
        logo_path_var = tk.StringVar(value=config.get('logo_path', ""))
        photos_folder_var = tk.StringVar(value=config.get('photos_folder', ""))
        campaign_var = tk.StringVar(value=config.get('campaign_name', "אשראיכם"))
        default_printer_var = tk.StringVar(value=str(config.get('default_printer') or '').strip())
        mode_var = tk.StringVar(value=str(config.get('deployment_mode') or 'local'))
        tenant_id_var = tk.StringVar(value=str(config.get('sync_tenant_id') or ''))
        tenant_name_var = tk.StringVar(value=str(config.get('sync_tenant_name') or ''))

        show_default_printer = False

        # אחידות גדלים לשדות, תוויות וכפתורים
        FIELD_WIDTH = 34
        BUTTON_WIDTH = 14
        LABEL_WIDTH = 32

        def _open_initial_settings_dialog():
            dialog2 = tk.Toplevel(self.root)
            dialog2.title('הגדרות ראשוניות')
            dialog2.configure(bg='#ecf0f1')
            dialog2.transient(self.root)
            dialog2.grab_set()
            nonlocal license_status_label
            try:
                dialog2.geometry('780x460')
                dialog2.minsize(740, 420)
            except Exception:
                pass

            tk.Label(
                dialog2,
                text=fix_rtl_text('זהירות: שינוי הגדרות ראשוניות משפיע על כל העמדות.'),
                font=('Arial', 12, 'bold'),
                bg='#ecf0f1',
                fg='#c0392b'
            ).pack(pady=(14, 6))

            frame2 = tk.Frame(dialog2, bg='#ecf0f1')
            frame2.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))

            def _lock_entry(entry_widget, value: str) -> None:
                if value:
                    try:
                        entry_widget.configure(state='readonly', readonlybackground='#e5e5e5')
                    except Exception:
                        pass
                else:
                    try:
                        entry_widget.configure(state='normal')
                    except Exception:
                        pass

            # שורה - תיקיית רשת משותפת
            row1i = tk.Frame(frame2, bg='#ecf0f1')
            row1i.pack(fill=tk.X, pady=3)
            row1i.columnconfigure(1, weight=1)
            shared_folder_label_i = tk.Label(row1i, text=fix_rtl_text("תיקיית רשת משותפת:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH)
            shared_folder_label_i.grid(row=0, column=2, sticky='e', padx=3)
            folder_entry_i = tk.Entry(row1i, textvariable=shared_folder_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
            folder_entry_i.grid(row=0, column=1, sticky='e', padx=3)

            def browse_folder_i():
                initial = shared_folder_var.get().strip() or os.path.dirname(self.base_dir)
                path = filedialog.askdirectory(title="בחר תיקיית רשת משותפת", initialdir=initial)
                if path:
                    shared_folder_var.set(path)
                    _lock_entry(folder_entry_i, path)

            browse_folder_btn_i = tk.Button(row1i, text="בחר תיקייה", command=browse_folder_i, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4)
            browse_folder_btn_i.grid(row=0, column=0, sticky='e', padx=3)
            _lock_entry(folder_entry_i, shared_folder_var.get().strip())

            # שורה - מצב עבודה
            row1bi = tk.Frame(frame2, bg='#ecf0f1')
            row1bi.pack(fill=tk.X, pady=3)
            row1bi.columnconfigure(1, weight=1)
            tk.Label(row1bi, text=fix_rtl_text("מצב עבודה:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            mode_frame_i = tk.Frame(row1bi, bg='#ecf0f1')
            mode_frame_i.grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row1bi, text="", bg='#ecf0f1').grid(row=0, column=0, padx=5)
            tk.Radiobutton(mode_frame_i, text=fix_rtl_text('רשת ביתית'), variable=mode_var, value='local', bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
            tk.Radiobutton(mode_frame_i, text=fix_rtl_text('היברידי'), variable=mode_var, value='hybrid', bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
            tk.Radiobutton(mode_frame_i, text=fix_rtl_text('מקוון בלבד'), variable=mode_var, value='cloud', bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)

            # שורה - שם מוסד (אנגלית)
            row1ci = tk.Frame(frame2, bg='#ecf0f1')
            row1ci.pack_forget()
            row1ci.columnconfigure(1, weight=1)
            tk.Label(row1ci, text=fix_rtl_text("שם מוסד (אנגלית):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            tk.Entry(row1ci, textvariable=tenant_name_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right').grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row1ci, text=fix_rtl_text("(מופיע באתר)"), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d').grid(row=0, column=0, sticky='e', padx=6)

            # שורה - קוד מוסד לסנכרון
            row1di = tk.Frame(frame2, bg='#ecf0f1')
            row1di.pack_forget()
            row1di.columnconfigure(1, weight=1)
            tk.Label(row1di, text=fix_rtl_text("קוד מוסד (Tenant ID):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            tenant_id_entry_i = tk.Entry(row1di, textvariable=tenant_id_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
            tenant_id_entry_i.grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row1di, text=fix_rtl_text("(אותיות/מספרים)"), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d').grid(row=0, column=0, sticky='e', padx=6)

            # שורה - כתובת סנכרון (Push URL)
            try:
                push_url_var = tk.StringVar(value=str(config.get('sync_push_url') or '').strip())
            except Exception:
                push_url_var = tk.StringVar(value='')
            row1ei = tk.Frame(frame2, bg='#ecf0f1')
            row1ei.pack_forget()
            row1ei.columnconfigure(1, weight=1)
            tk.Label(row1ei, text=fix_rtl_text("כתובת סנכרון (Push URL):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            push_url_entry_i = tk.Entry(row1ei, textvariable=push_url_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
            push_url_entry_i.grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row1ei, text=fix_rtl_text("(/sync/push)"), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d').grid(row=0, column=0, sticky='e', padx=6)

            # שורה - מפתח API
            try:
                api_key_var = tk.StringVar(value=str(config.get('sync_api_key') or config.get('api_key') or config.get('sync_key') or '').strip())
            except Exception:
                api_key_var = tk.StringVar(value='')
            row1fi = tk.Frame(frame2, bg='#ecf0f1')
            row1fi.pack_forget()
            row1fi.columnconfigure(1, weight=1)
            tk.Label(row1fi, text=fix_rtl_text("מפתח API (לסנכרון):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            api_key_entry_i = tk.Entry(row1fi, textvariable=api_key_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right', show='*')
            api_key_entry_i.grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row1fi, text=fix_rtl_text("(סודי)"), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d').grid(row=0, column=0, sticky='e', padx=6)

            try:
                tenant_id_entry_i.configure(state='readonly', readonlybackground='#e5e5e5')
            except Exception:
                pass
            try:
                push_url_entry_i.configure(state='readonly', readonlybackground='#e5e5e5')
            except Exception:
                pass
            try:
                api_key_entry_i.configure(state='readonly', readonlybackground='#e5e5e5')
            except Exception:
                pass

            # שורה - לוגו לעמדה הציבורית
            row2i = tk.Frame(frame2, bg='#ecf0f1')
            row2i.pack(fill=tk.X, pady=3)
            row2i.columnconfigure(1, weight=1)
            tk.Label(row2i, text=fix_rtl_text("לוגו לעמדה הציבורית:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            logo_entry_i = tk.Entry(row2i, textvariable=logo_path_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
            logo_entry_i.grid(row=0, column=1, sticky='e', padx=5)

            def _lock_logo_entry_i() -> None:
                try:
                    value = str(logo_path_var.get() or '').strip()
                except Exception:
                    value = ''
                _lock_entry(logo_entry_i, value)

            def browse_logo_i():
                try:
                    current = str(logo_path_var.get() or '').strip()
                except Exception:
                    current = ''
                if current:
                    try:
                        if not messagebox.askyesno('שינוי לוגו', 'לוגו כבר הוגדר.\nלשנות אותו?', parent=dialog2):
                            return
                    except Exception:
                        pass
                initial = os.path.dirname(current) if current else self.base_dir
                path = filedialog.askopenfilename(
                    title="בחר קובץ לוגו",
                    filetypes=[("קבצי תמונה", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("הכל", "*.*")],
                    initialdir=initial
                )
                if path:
                    logo_path_var.set(path)
                    _lock_logo_entry_i()

            tk.Button(row2i, text="בחר קובץ", command=browse_logo_i, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).grid(row=0, column=0, sticky='e', padx=5)
            _lock_logo_entry_i()

            # שורה - תיקיית תמונות תלמידים
            row2bi = tk.Frame(frame2, bg='#ecf0f1')
            row2bi.pack(fill=tk.X, pady=3)
            row2bi.columnconfigure(1, weight=1)
            tk.Label(row2bi, text=fix_rtl_text("תיקיית תמונות תלמידים:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            photos_entry_i = tk.Entry(row2bi, textvariable=photos_folder_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
            photos_entry_i.grid(row=0, column=1, sticky='e', padx=5)

            def _lock_photos_entry_i() -> None:
                try:
                    value = str(photos_folder_var.get() or '').strip()
                except Exception:
                    value = ''
                _lock_entry(photos_entry_i, value)

            def browse_photos_folder_i():
                try:
                    current = str(photos_folder_var.get() or '').strip()
                except Exception:
                    current = ''
                if current:
                    try:
                        if not messagebox.askyesno('שינוי תיקייה', 'תיקיית תמונות כבר הוגדרה.\nלשנות אותה?', parent=dialog2):
                            return
                    except Exception:
                        pass
                initial = current or self.base_dir
                path = filedialog.askdirectory(title="בחר תיקיית תמונות תלמידים", initialdir=initial)
                if path:
                    photos_folder_var.set(path)
                    _lock_photos_entry_i()

            tk.Button(row2bi, text="בחר תיקייה", command=browse_photos_folder_i, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).grid(row=0, column=0, sticky='e', padx=5)
            _lock_photos_entry_i()

            # שורה - רישום מערכת
            row_license = tk.Frame(frame2, bg='#ecf0f1')
            row_license.pack(fill=tk.X, pady=3)
            row_license.columnconfigure(1, weight=1)
            tk.Label(
                row_license,
                text=fix_rtl_text("רישום מערכת:"),
                font=('Arial', 10, 'bold'),
                bg='#ecf0f1',
                anchor='e',
                width=LABEL_WIDTH
            ).grid(row=0, column=2, sticky='e', padx=5)
            tk.Button(
                row_license,
                text="🔑 רישום מערכת",
                command=open_license_dialog,
                font=('Arial', 9, 'bold'),
                bg='#16a085',
                fg='white',
                padx=14,
                pady=5
            ).grid(row=0, column=1, sticky='e', padx=5)
            tk.Label(row_license, text="", bg='#ecf0f1').grid(row=0, column=0, padx=5)

            license_status_label = tk.Label(
                frame2,
                text=get_license_status_text(),
                font=('Arial', 9),
                bg='#ecf0f1',
                fg='#2c3e50',
                anchor='e',
                justify='right',
                wraplength=740
            )
            license_status_label.pack(fill=tk.X, padx=12, pady=(0, 6))

            def _open_cloud_connect():
                try:
                    url = 'https://schoolpoints.co.il/web/signin'
                    webbrowser.open(url)
                except Exception as e:
                    try:
                        messagebox.showerror('שגיאה', str(e), parent=dialog2)
                    except Exception:
                        pass

            def _get_cloud_base_url() -> str:
                try:
                    cfg0 = self.load_app_config() or {}
                except Exception:
                    cfg0 = {}
                try:
                    base = str(cfg0.get('cloud_base_url') or '').strip()
                except Exception:
                    base = ''
                if not base:
                    base = 'https://schoolpoints.co.il'
                return str(base).strip().rstrip('/')

            def _is_cloud_connected() -> bool:
                try:
                    cfg0 = self.load_app_config() or {}
                except Exception:
                    cfg0 = {}
                try:
                    tid0 = str(cfg0.get('sync_tenant_id') or '').strip()
                except Exception:
                    tid0 = ''
                try:
                    key0 = str(cfg0.get('sync_api_key') or cfg0.get('api_key') or cfg0.get('sync_key') or '').strip()
                except Exception:
                    key0 = ''
                try:
                    push0 = str(cfg0.get('sync_push_url') or '').strip()
                except Exception:
                    push0 = ''
                return bool(tid0 and (key0 or push0))

            def _verify_cloud_connection(tenant_id: str, api_key: str, push_url: str) -> tuple[bool, str]:
                try:
                    import urllib.request
                    import urllib.parse
                    import urllib.error
                except Exception:
                    return False, 'חסרה תמיכה ברשת (urllib)'

                tid = str(tenant_id or '').strip()
                key = str(api_key or '').strip()
                purl = str(push_url or '').strip()
                if not tid:
                    return False, 'חסר Tenant ID'
                if not key:
                    return False, 'חסר API Key'
                if not purl:
                    return False, 'חסרה כתובת Push URL'

                base = purl
                if base.endswith('/sync/push'):
                    base = base[:-len('/sync/push')]
                status_url = base.rstrip('/') + '/sync/status'
                q = 'tenant_id=' + urllib.parse.quote(tid)
                url = status_url + ('&' if '?' in status_url else '?') + q
                req = urllib.request.Request(
                    url,
                    headers={
                        'Accept': 'application/json',
                        'api-key': key,
                    }
                )
                try:
                    with urllib.request.urlopen(req, timeout=8) as resp:
                        raw = resp.read().decode('utf-8', errors='ignore')
                    # success is any 2xx with JSON containing ok
                    try:
                        import json
                        data = json.loads(raw or '{}')
                    except Exception:
                        data = {}
                    if isinstance(data, dict) and bool(data.get('ok')):
                        return True, '✓ אושר מול השרת'
                    return False, 'השרת לא אישר התחברות'
                except urllib.error.HTTPError as e:
                    try:
                        body = e.read().decode('utf-8', errors='ignore')
                    except Exception:
                        body = ''
                    msg = f"HTTP {getattr(e, 'code', '')}".strip()
                    if body:
                        msg = msg + f"\n{body}" if msg else body
                    return False, msg or 'שגיאת התחברות'
                except Exception as e:
                    return False, str(e)

            connect_row = tk.Frame(frame2, bg='#ecf0f1')
            connect_row.pack(fill=tk.X, pady=(6, 0))
            connect_row.columnconfigure(1, weight=1)
            tk.Label(connect_row, text=fix_rtl_text("חיבור לחשבון ענן:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=5)
            pairing_in_progress = {'on': False}
            connect_btn = tk.Button(
                connect_row,
                text='🌐 התחבר',
                command=_open_cloud_connect,
                font=('Arial', 9, 'bold'),
                bg='#2980b9',
                fg='white',
                padx=14,
                pady=5
            )
            connect_btn.grid(row=0, column=1, sticky='e', padx=5)

            disconnect_btn = tk.Button(
                connect_row,
                text='התנתק',
                font=('Arial', 9, 'bold'),
                bg='#95a5a6',
                fg='white',
                padx=14,
                pady=5
            )
            disconnect_btn.grid(row=0, column=0, sticky='e', padx=6)

            cloud_state_lbl = tk.Label(connect_row, text=fix_rtl_text(''), font=('Arial', 9), bg='#ecf0f1', fg='#7f8c8d', anchor='e', justify='right')
            cloud_state_lbl.grid(row=1, column=1, columnspan=2, sticky='e', padx=5, pady=(2, 0))

            def _do_disconnect():
                try:
                    if not messagebox.askyesno('התנתקות', 'לנתק את החיבור לענן ולמחוק פרטי סנכרון מהמחשב?', parent=dialog2):
                        return
                except Exception:
                    pass
                try:
                    cfg1 = self.load_app_config() or {}
                except Exception:
                    cfg1 = {}
                try:
                    cfg1.pop('sync_api_key', None)
                    cfg1.pop('api_key', None)
                    cfg1.pop('sync_key', None)
                    cfg1.pop('sync_push_url', None)
                    cfg1.pop('sync_pull_url', None)
                    cfg1.pop('sync_snapshot_url', None)
                    cfg1.pop('sync_station_id', None)
                except Exception:
                    pass
                try:
                    self.save_app_config(cfg1)
                except Exception:
                    pass
                try:
                    _refresh_cloud_ui()
                except Exception:
                    pass

                try:
                    api_key_var.set('')
                except Exception:
                    pass
                try:
                    push_url_var.set('')
                except Exception:
                    pass

            try:
                disconnect_btn.configure(command=_do_disconnect)
            except Exception:
                pass

            def _apply_mode_ui(*_args):
                mode = str(mode_var.get() or 'local').strip().lower()
                allow_cloud = mode in ('hybrid', 'cloud')
                try:
                    # When connected: always show disabled 'מחובר'
                    if allow_cloud and _is_cloud_connected():
                        connect_btn.configure(state='disabled')
                    else:
                        connect_btn.configure(state=('normal' if allow_cloud else 'disabled'))
                except Exception:
                    pass
                # במצב מקוון בלבד: אין תיקיית רשת משותפת
                disable_shared = mode == 'cloud'
                try:
                    folder_entry_i.configure(state=('disabled' if disable_shared else 'normal'))
                except Exception:
                    pass
                try:
                    browse_folder_btn_i.configure(state=('disabled' if disable_shared else 'normal'))
                except Exception:
                    pass
                try:
                    shared_folder_label_i.configure(fg=('#95a5a6' if disable_shared else '#2c3e50'))
                except Exception:
                    pass

                try:
                    _refresh_cloud_ui()
                except Exception:
                    pass

            def _refresh_cloud_ui():
                try:
                    connected = bool(_is_cloud_connected())
                except Exception:
                    connected = False
                try:
                    allow_cloud = str(mode_var.get() or 'local').strip().lower() in ('hybrid', 'cloud')
                except Exception:
                    allow_cloud = False

                try:
                    if bool(pairing_in_progress.get('on')) and allow_cloud and (not connected):
                        try:
                            connect_btn.configure(text=fix_rtl_text('ממתין...'), state='disabled')
                        except Exception:
                            pass
                        try:
                            disconnect_btn.configure(state='disabled')
                        except Exception:
                            pass
                        try:
                            cloud_state_lbl.configure(text=fix_rtl_text('ממתין לאישור בדפדפן...'))
                        except Exception:
                            pass
                        return
                except Exception:
                    pass

                if (not allow_cloud):
                    try:
                        connect_btn.configure(text='🌐 התחבר')
                    except Exception:
                        pass
                    try:
                        disconnect_btn.configure(state='disabled')
                    except Exception:
                        pass
                    try:
                        cloud_state_lbl.configure(text=fix_rtl_text('מצב מקומי - חיבור ענן כבוי'))
                    except Exception:
                        pass
                    return

                if connected:
                    try:
                        connect_btn.configure(text=fix_rtl_text('מחובר'), state='disabled')
                    except Exception:
                        pass
                    try:
                        disconnect_btn.configure(state='normal')
                    except Exception:
                        pass
                    try:
                        cloud_state_lbl.configure(text=fix_rtl_text('מחובר לענן'))
                    except Exception:
                        pass
                else:
                    try:
                        connect_btn.configure(text='🌐 התחבר')
                    except Exception:
                        pass
                    try:
                        disconnect_btn.configure(state='disabled')
                    except Exception:
                        pass
                    try:
                        cloud_state_lbl.configure(text=fix_rtl_text('(פותח דפדפן)'))
                    except Exception:
                        pass

            def _connect_verified():
                # Open browser for signup/login, then verify credentials against server.
                # Important: do NOT call _open_cloud_connect here because it may be wrapped later.
                _ = None
                try:
                    import threading
                    base = _get_cloud_base_url()

                    try:
                        pairing_in_progress['on'] = True
                    except Exception:
                        pass
                    try:
                        _refresh_cloud_ui()
                    except Exception:
                        pass

                    def _pair_worker():
                        res = self._device_pair_start_and_poll(base, poll_timeout_sec=600) or {}

                        def _on_done():
                            try:
                                pairing_in_progress['on'] = False
                            except Exception:
                                pass
                            if bool(res.get('ok')):
                                cfg2 = self.load_app_config() or {}
                                cfg2['sync_tenant_id'] = str(res.get('tenant_id') or '').strip()
                                cfg2['sync_api_key'] = str(res.get('api_key') or '').strip()
                                cfg2['sync_push_url'] = str(res.get('push_url') or '').strip()
                                try:
                                    self.save_app_config(cfg2)
                                except Exception:
                                    pass
                                try:
                                    tenant_id_var.set(str(cfg2.get('sync_tenant_id') or ''))
                                except Exception:
                                    pass
                                try:
                                    api_key_var.set(str(cfg2.get('sync_api_key') or ''))
                                except Exception:
                                    pass
                                try:
                                    push_url_var.set(str(cfg2.get('sync_push_url') or ''))
                                except Exception:
                                    pass
                                try:
                                    self._maybe_start_sync_agent()
                                except Exception:
                                    pass
                                try:
                                    cloud_state_lbl.configure(text=fix_rtl_text('מחובר לענן'))
                                except Exception:
                                    pass
                                try:
                                    messagebox.showinfo('חיבור לענן', 'העמדה התחברה לענן בהצלחה.', parent=dialog2)
                                except Exception:
                                    pass
                            else:
                                try:
                                    cloud_state_lbl.configure(text=fix_rtl_text('לא מחובר'))
                                except Exception:
                                    pass
                                try:
                                    err = str(res.get('error') or 'שגיאת חיבור')
                                except Exception:
                                    err = 'שגיאת חיבור'
                                try:
                                    messagebox.showerror('חיבור לענן', err, parent=dialog2)
                                except Exception:
                                    pass
                            try:
                                _refresh_cloud_ui()
                            except Exception:
                                pass

                        try:
                            dialog2.after(0, _on_done)
                        except Exception:
                            _on_done()

                    try:
                        cloud_state_lbl.configure(text=fix_rtl_text('חיבור אוטומטי... (פותח דפדפן)'))
                    except Exception:
                        pass
                    threading.Thread(target=_pair_worker, daemon=True).start()
                    return
                except Exception:
                    pass
                try:
                    _orig_open_cloud()
                except Exception:
                    pass

                tid = str(tenant_id_var.get() or '').strip()
                key = str(api_key_var.get() or '').strip()
                purl = str(push_url_var.get() or '').strip()

                def _set_busy(on: bool):
                    try:
                        if on:
                            connect_btn.configure(state='disabled')
                            disconnect_btn.configure(state='disabled')
                        else:
                            pass
                    except Exception:
                        pass

                try:
                    cloud_state_lbl.configure(text=fix_rtl_text('בודק מול השרת...'))
                except Exception:
                    pass
                _set_busy(True)

                def _worker():
                    ok, msg = _verify_cloud_connection(tid, key, purl)

                    def _on_done():
                        try:
                            if ok:
                                cfg2 = self.load_app_config() or {}
                                cfg2['sync_tenant_id'] = tid
                                cfg2['sync_api_key'] = key
                                cfg2['sync_push_url'] = purl
                                try:
                                    self.save_app_config(cfg2)
                                except Exception:
                                    pass
                                try:
                                    cloud_state_lbl.configure(text=fix_rtl_text('מחובר לענן'))
                                except Exception:
                                    pass
                            else:
                                try:
                                    cloud_state_lbl.configure(text=fix_rtl_text('לא מחובר'))
                                except Exception:
                                    pass
                                try:
                                    messagebox.showerror('התחברות לענן', str(msg or 'שגיאת התחברות'), parent=dialog2)
                                except Exception:
                                    pass
                        finally:
                            try:
                                _refresh_cloud_ui()
                            except Exception:
                                pass
                    try:
                        dialog2.after(0, _on_done)
                    except Exception:
                        _on_done()

                try:
                    import threading
                    threading.Thread(target=_worker, daemon=True).start()
                except Exception:
                    _worker()

            try:
                _refresh_cloud_ui()
            except Exception:
                pass

            try:
                def _on_cloud_focus(_e=None):
                    try:
                        _refresh_cloud_ui()
                    except Exception:
                        pass
                dialog2.bind('<FocusIn>', _on_cloud_focus)
            except Exception:
                pass

            try:
                connect_btn.configure(command=_connect_verified)
            except Exception:
                pass

            try:
                mode_var.trace_add('write', _apply_mode_ui)
            except Exception:
                try:
                    mode_var.trace('w', lambda *_a: _apply_mode_ui())
                except Exception:
                    pass
            _apply_mode_ui()

            init_btn_frame = tk.Frame(dialog2, bg='#ecf0f1')
            init_btn_frame.pack(pady=(10, 14))
            tk.Button(
                init_btn_frame,
                text="✖ סגור",
                command=dialog2.destroy,
                font=('Arial', 10),
                bg='#95a5a6',
                fg='white',
                width=BUTTON_WIDTH,
                padx=4,
                pady=6
            ).pack()

        init_btn_row = tk.Frame(content_frame, bg='#ecf0f1')
        init_btn_row.pack(fill=tk.X, pady=3)
        init_btn_row.columnconfigure(0, weight=1)
        init_btn_row.columnconfigure(1, weight=1)
        init_btn_row.columnconfigure(2, weight=1)
        tk.Button(
            init_btn_row,
            text="פתח הגדרות ראשוניות",
            command=_open_initial_settings_dialog,
            font=('Arial', 10, 'bold'),
            bg='#8e44ad',
            fg='white',
            width=BUTTON_WIDTH + 10,
            padx=6,
            pady=4
        ).grid(row=0, column=1, sticky='n', padx=3)
        
        # שורה 3 - שם המבצע
        row3 = tk.Frame(content_frame, bg='#ecf0f1')
        row3.pack(fill=tk.X, pady=6)
        row3.columnconfigure(1, weight=1)
        tk.Label(row3, text=fix_rtl_text("שם המבצע (עמדה ציבורית):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=3)
        campaign_entry = tk.Entry(row3, textvariable=campaign_var, font=('Arial', 10), width=max(16, FIELD_WIDTH - 16), justify='right')
        campaign_entry.grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(row3, text="", bg='#ecf0f1').grid(row=0, column=0)

        if show_default_printer:
            # שורה 4 - מדפסת ברירת מחדל (לקופה/קבלות)
            row4 = tk.Frame(content_frame, bg='#ecf0f1')
            row4.pack(fill=tk.X, pady=6)
            row4.columnconfigure(1, weight=1)
            tk.Label(row4, text=fix_rtl_text("מדפסת ברירת מחדל (קבלות):"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).grid(row=0, column=2, sticky='e', padx=3)

            printers = []
            try:
                import win32print  # type: ignore
                flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
                for p in (win32print.EnumPrinters(flags) or []):
                    try:
                        nm = str(p[2] or '').strip()
                    except Exception:
                        nm = ''
                    if nm:
                        printers.append(nm)
            except Exception:
                printers = []
            printers = [p for p in printers if p]
            # de-dup keep order
            seen = set()
            dedup = []
            for p in printers:
                if p not in seen:
                    seen.add(p)
                    dedup.append(p)
            printers = dedup

            if printers:
                printer_cb = ttk.Combobox(row4, textvariable=default_printer_var, values=printers, state='readonly', width=FIELD_WIDTH, justify='right')
                printer_cb.grid(row=0, column=1, sticky='e', padx=3)
            else:
                printer_entry = tk.Entry(row4, textvariable=default_printer_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
                printer_entry.grid(row=0, column=1, sticky='e', padx=3)
            tk.Label(row4, text="", bg='#ecf0f1').grid(row=0, column=0)

        # סטטוס רישיון + כפתור רישום מערכת
        lm = getattr(self, "license_manager", None)

        def get_license_status_text():
            if lm is None:
                return "מצב רישיון: (לא זמין)"
            if lm.is_licensed:
                name = lm.school_name or "לא ידוע"
                used = getattr(lm, 'used_stations', 0) or 0
                total = getattr(lm, 'max_stations', 0) or 0
                free = max(0, total - used)
                extra = ""
                try:
                    try:
                        _is_term_attr = getattr(lm, 'is_term', False)
                        is_term = bool(_is_term_attr() if callable(_is_term_attr) else _is_term_attr)
                    except Exception:
                        is_term = False
                    if getattr(lm, 'is_monthly', False):
                        exp = str(getattr(lm, 'expiry_date', '') or '').strip()
                        if exp:
                            extra = f" | חודשי עד {exp}"
                        try:
                            left = int(getattr(lm, 'monthly_days_left', 0) or 0)
                        except Exception:
                            left = 0
                        if left:
                            extra += f" | נותרו {left} ימים"
                        try:
                            allow_cashier = bool(getattr(lm, 'allow_cashier', True))
                        except Exception:
                            allow_cashier = True
                        extra += " | קופה: " + ("כן" if allow_cashier else "לא")
                    elif is_term:
                        exp = str(getattr(lm, 'expiry_date', '') or '').strip()
                        act = str(getattr(lm, 'activated_at', '') or '').strip()
                        try:
                            td = int(getattr(lm, 'term_days', 0) or 0)
                        except Exception:
                            td = 0
                        try:
                            left = int(getattr(lm, 'term_days_left', 0) or 0)
                        except Exception:
                            left = 0
                        used_days = max(0, int(td) - int(left)) if td else 0
                        try:
                            allow_cashier = bool(getattr(lm, 'allow_cashier', True))
                        except Exception:
                            allow_cashier = True
                        if td:
                            extra = f" | נרשם ל-{td} ימים"
                        else:
                            extra = " | נרשם"
                        if act:
                            extra += f" | הופעל ב-{act}"
                        if exp:
                            extra += f" | עד {exp}"
                        extra += f" | בשימוש {used_days} ימים, נותרו {left}"
                        extra += " | קופה: " + ("כן" if allow_cashier else "לא")
                except Exception:
                    extra = ""
                return (
                    f"מצב רישיון: פעיל – נרשם ל{name} "
                    f"עבור {total} עמדות ({used} בשימוש, {free} פנויות){extra}"
                )
            if lm.is_trial:
                return f"מצב רישיון: גרסת ניסיון – נותרו {lm.trial_days_left} ימים"
            try:
                if getattr(lm, 'is_monthly', False) and bool(getattr(lm, 'monthly_expired', False)):
                    exp = str(getattr(lm, 'expiry_date', '') or '').strip()
                    suffix = f" (עד {exp})" if exp else ""
                    return f"מצב רישיון: רישיון חודשי פג תוקף{suffix}"
            except Exception:
                pass
            try:
                _is_term_attr2 = getattr(lm, 'is_term', False)
                is_term2 = bool(_is_term_attr2() if callable(_is_term_attr2) else _is_term_attr2)
            except Exception:
                is_term2 = False
            try:
                if is_term2 and bool(getattr(lm, 'term_expired', False)):
                    exp = str(getattr(lm, 'expiry_date', '') or '').strip()
                    suffix = f" (עד {exp})" if exp else ""
                    return f"מצב רישיון: רישיון פג תוקף{suffix}"
            except Exception:
                pass
            if lm.trial_expired:
                return "מצב רישיון: תקופת הניסיון הסתיימה – ללא רישיון פעיל"
            return "מצב רישיון: לא רשום"

        license_status_label = None

        # הגדרת השמעת צלילים בעמדה הציבורית
        sounds_enabled_value = config.get('sounds_enabled', '1')
        sounds_enabled_var = tk.BooleanVar(value=sounds_enabled_value == '1')
        
        sounds_frame = tk.Frame(content_frame, bg='#ecf0f1')
        sounds_frame.pack(fill=tk.X, pady=(0, 5))
        sounds_frame.columnconfigure(1, weight=1)
        
        tk.Label(
            sounds_frame,
            text=fix_rtl_text("השמעת צלילים בעמדה הציבורית:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).grid(row=0, column=2, sticky='e', padx=3)

        tk.Checkbutton(
            sounds_frame,
            variable=sounds_enabled_var,
            onvalue=True,
            offvalue=False,
            bg='#ecf0f1',
            activebackground='#ecf0f1'
        ).grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(sounds_frame, text="", bg='#ecf0f1').grid(row=0, column=0)
        
        # הגדרת עוצמת שמע (0-100)
        volume_value = int(config.get('sound_volume', 80))
        volume_var = tk.IntVar(value=volume_value)
        
        volume_frame = tk.Frame(content_frame, bg='#ecf0f1')
        volume_frame.pack(fill=tk.X, pady=(0, 5))
        volume_frame.columnconfigure(1, weight=1)
        
        tk.Label(
            volume_frame,
            text=fix_rtl_text("עוצמת צלילים (0-100):"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).grid(row=0, column=2, sticky='e', padx=3)
        
        volume_scale = tk.Scale(
            volume_frame,
            from_=0,
            to=100,
            orient=tk.HORIZONTAL,
            variable=volume_var,
            length=200,
            bg='#ecf0f1'
        )
        volume_scale.grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(volume_frame, text="", bg='#ecf0f1').grid(row=0, column=0)

        quiet_mode_ranges = self._load_quiet_mode_ranges(config)
        quiet_button_var = tk.StringVar(value=f"🌙 מצב שקט ({self._quiet_ranges_summary(quiet_mode_ranges)})")

        def open_quiet_mode_manager():
            nonlocal quiet_mode_ranges
            quiet_mode_ranges = self._open_quiet_mode_manager_dialog(dialog, quiet_mode_ranges)
            quiet_button_var.set(f"🌙 מצב שקט ({self._quiet_ranges_summary(quiet_mode_ranges)})")

        quiet_frame = tk.Frame(content_frame, bg='#ecf0f1')
        quiet_frame.pack(fill=tk.X, pady=(0, 8))
        quiet_frame.columnconfigure(1, weight=1)

        tk.Label(
            quiet_frame,
            text=fix_rtl_text("מצב שקט:"),
            font=('Arial', 10, 'bold'),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).grid(row=0, column=2, sticky='e', padx=3)

        tk.Button(
            quiet_frame,
            textvariable=quiet_button_var,
            command=open_quiet_mode_manager,
            font=('Arial', 9, 'bold'),
            bg='#34495e',
            fg='white',
            padx=12,
            pady=4
        ).grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(quiet_frame, text="", bg='#ecf0f1').grid(row=0, column=0)

        # מקסימום נקודות (תקרה דינמית)
        max_points_frame = tk.Frame(content_frame, bg='#ecf0f1')
        max_points_frame.pack(fill=tk.X, pady=(10, 5))
        max_points_frame.columnconfigure(1, weight=1)
        tk.Label(
            max_points_frame,
            text=fix_rtl_text("מקסימום נקודות:"),
            font=('Arial', 10, 'bold'),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).grid(row=0, column=2, sticky='e', padx=3)
        tk.Button(
            max_points_frame,
            text="📈 מקסימום נקודות",
            command=self.open_max_points_settings,
            font=('Arial', 10, 'bold'),
            bg='#8e44ad',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=4
        ).grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(max_points_frame, text="", bg='#ecf0f1').grid(row=0, column=0)

        def open_license_dialog(force_edit=False):
            if lm is None:
                messagebox.showerror("שגיאה", "מודול הרישוי אינו זמין.")
                return

            # מצב ברירת מחדל: אם כבר יש רישיון פעיל, הצג חלון תצוגה בלבד
            if lm.is_licensed and not force_edit:
                lic_win = tk.Toplevel(dialog)
                lic_win.title("פרטי רישיון מערכת")
                lic_win.geometry("720x260")
                try:
                    lic_win.minsize(720, 260)
                except Exception:
                    pass
                lic_win.configure(bg='#ecf0f1')
                lic_win.transient(dialog)
                lic_win.grab_set()

                tk.Label(
                    lic_win,
                    text="פרטי רישיון מערכת",
                    font=('Arial', 11, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(10, 5))

                # סיכום מצב הרישיון (כמו בשורת הסטטוס)
                status_text = get_license_status_text()
                tk.Label(
                    lic_win,
                    text=status_text,
                    font=('Arial', 10),
                    bg='#ecf0f1',
                    fg='#34495e',
                    anchor='e',
                    justify='right',
                    wraplength=680
                ).pack(fill=tk.X, padx=20, pady=(5, 5))

                # האם המחשב הנוכחי רשום ברישיון
                try:
                    machines = set(str(m) for m in (lm.machines or []))
                    current_id = str(getattr(lm, "machine_id", ""))
                    if current_id and current_id in machines and not lm.over_limit:
                        comp_text = "מחשב זה רשום כרגע לרישיון הפעיל ונספר כאחת העמדות המורשות."
                    else:
                        comp_text = "מחשב זה אינו רשום כרגע לרישיון הפעיל (ייתכן חריגה ממספר העמדות המותר)."
                except Exception:
                    comp_text = "לא ניתן לקבוע את מצב רישום המחשב הנוכחי."

                tk.Label(
                    lic_win,
                    text=comp_text,
                    font=('Arial', 10),
                    bg='#ecf0f1',
                    fg='#2c3e50',
                    anchor='e',
                    justify='right'
                ).pack(fill=tk.X, padx=20, pady=(0, 10))

                btn_bar = tk.Frame(lic_win, bg='#ecf0f1')
                btn_bar.pack(pady=10)

                # כפתור מעבר למצב החלפת רישיון (פותח שוב את הדיאלוג במצב עריכה)
                tk.Button(
                    btn_bar,
                    text="🔁 החלפת רישיון...",
                    command=lambda: (lic_win.destroy(), open_license_dialog(True)),
                    font=('Arial', 10, 'bold'),
                    bg='#e67e22',
                    fg='white',
                    padx=18,
                    pady=6
                ).pack(side=tk.LEFT, padx=8)

                tk.Button(
                    btn_bar,
                    text="✖ סגור",
                    command=lic_win.destroy,
                    font=('Arial', 10),
                    bg='#95a5a6',
                    fg='white',
                    padx=16,
                    pady=6
                ).pack(side=tk.LEFT, padx=8)

                return

            # מצב רישום / החלפת רישיון – טופס מלא
            lic_win = tk.Toplevel(dialog)
            lic_win.title("רישום / החלפת רישיון")
            lic_win.geometry("720x330")
            try:
                lic_win.minsize(720, 330)
            except Exception:
                pass
            lic_win.configure(bg='#ecf0f1')
            lic_win.transient(dialog)
            lic_win.grab_set()

            tk.Label(
                lic_win,
                text="רישום / החלפת רישיון – הזן שם מוסד, העתק קוד מערכת והזן קוד הפעלה שקיבלת",
                font=('Arial', 11, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50',
                wraplength=680,
                justify='right'
            ).pack(pady=10)

            form = tk.Frame(lic_win, bg='#ecf0f1')
            form.pack(fill=tk.X, padx=20, pady=10)

            tk.Label(form, text="שם מוסד:", font=('Arial', 10), bg='#ecf0f1', anchor='e', width=14).pack(side=tk.RIGHT, padx=5)
            school_var = tk.StringVar(value=lm.school_name or "")
            school_entry = tk.Entry(form, textvariable=school_var, font=('Arial', 10), width=32, justify='right')
            school_entry.pack(side=tk.RIGHT, padx=5)

            def paste_school_from_clipboard():
                try:
                    text = lic_win.clipboard_get()
                    if text:
                        school_var.set(text.strip())
                except Exception:
                    pass

            tk.Button(
                form,
                text="📋 הדבק",
                command=paste_school_from_clipboard,
                font=('Arial', 9),
                bg='#bdc3c7',
                fg='#2c3e50',
                padx=6,
                pady=2
            ).pack(side=tk.LEFT, padx=5)

            # שורת קוד מערכת – לקריאה בלבד, להעתקה אליך לצורך יצירת קוד הפעלה
            form_sys = tk.Frame(lic_win, bg='#ecf0f1')
            form_sys.pack(fill=tk.X, padx=20, pady=5)

            tk.Label(
                form_sys,
                text="קוד מערכת:",
                font=('Arial', 10),
                bg='#ecf0f1',
                anchor='e',
                width=14
            ).pack(side=tk.RIGHT, padx=5)

            system_code_var = tk.StringVar(value=getattr(lm, "system_code", ""))
            system_entry = tk.Entry(
                form_sys,
                textvariable=system_code_var,
                font=('Consolas', 10),
                width=32,
                justify='left',
                state='readonly'
            )
            system_entry.pack(side=tk.RIGHT, padx=5)

            def copy_system_code_to_clipboard():
                try:
                    code = system_code_var.get().strip()
                    if code:
                        lic_win.clipboard_clear()
                        lic_win.clipboard_append(code)
                except Exception:
                    pass

            tk.Button(
                form_sys,
                text="📋 העתק",
                command=copy_system_code_to_clipboard,
                font=('Arial', 9),
                bg='#bdc3c7',
                fg='#2c3e50',
                padx=6,
                pady=2
            ).pack(side=tk.LEFT, padx=5)

            form2 = tk.Frame(lic_win, bg='#ecf0f1')
            form2.pack(fill=tk.X, padx=20, pady=5)

            tk.Label(form2, text="קוד הפעלה:", font=('Arial', 10), bg='#ecf0f1', anchor='e', width=14).pack(side=tk.RIGHT, padx=5)
            key_var = tk.StringVar()
            key_entry = tk.Entry(form2, textvariable=key_var, font=('Arial', 10), width=32, justify='right')
            key_entry.pack(side=tk.RIGHT, padx=5)

            def paste_key_from_clipboard():
                try:
                    text = lic_win.clipboard_get()
                    if text:
                        key_var.set(text.strip())
                except Exception:
                    pass

            tk.Button(
                form2,
                text="📋 הדבק",
                command=paste_key_from_clipboard,
                font=('Arial', 9),
                bg='#bdc3c7',
                fg='#2c3e50',
                padx=6,
                pady=2
            ).pack(side=tk.LEFT, padx=5)

            def do_activate():
                school = school_var.get().strip()
                key = key_var.get().strip()
                success, msg = lm.activate(school, key, None)
                if success:
                    # עדכון סטטוס רישיון בחלון ההגדרות
                    if license_status_label is not None:
                        license_status_label.config(text=get_license_status_text())
                    # עדכון סטטוס רישיון בכותרת הראשית (אם קיים)
                    if hasattr(self, "license_header_label"):
                        try:
                            try:
                                self.license_header_label.config(text=get_license_header_text())
                            except Exception:
                                self.license_header_label.config(text=get_license_status_text())
                        except Exception:
                            pass
                    messagebox.showinfo("רישום מערכת", msg)
                    lic_win.destroy()
                else:
                    messagebox.showerror("רישום מערכת", msg)

            btn_bar = tk.Frame(lic_win, bg='#ecf0f1')
            btn_bar.pack(pady=15)

            tk.Button(
                btn_bar,
                text="🔑 הפעל רישיון",
                command=do_activate,
                font=('Arial', 10, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=20,
                pady=6
            ).pack(side=tk.LEFT, padx=8)

            tk.Button(
                btn_bar,
                text="✖ סגור",
                command=lic_win.destroy,
                font=('Arial', 10),
                bg='#95a5a6',
                fg='white',
                padx=16,
                pady=6
            ).pack(side=tk.LEFT, padx=8)

            school_entry.focus_set()

        # (הגדרות תצוגה בעמדה הציבורית עברו לחלון "הגדרות תצוגה")

        # לוח חופשות / חסימות לעמדה הציבורית
        try:
            closures_frame = tk.Frame(content_frame, bg='#ecf0f1')
            closures_frame.pack(fill=tk.X, pady=(8, 0))
            closures_frame.columnconfigure(1, weight=1)
            tk.Label(
                closures_frame,
                text=fix_rtl_text("חופשות/שבת (עמדה ציבורית):"),
                font=('Arial', 10, 'bold'),
                bg='#ecf0f1',
                anchor='e',
                width=LABEL_WIDTH
            ).grid(row=0, column=2, sticky='e', padx=3)
            tk.Button(
                closures_frame,
                text="🗓 ניהול לוח חופשות",
                command=self.open_public_closures_manager,
                font=('Arial', 10, 'bold'),
                bg='#34495e',
                fg='white',
                width=BUTTON_WIDTH,
                padx=4,
                pady=4
            ).grid(row=0, column=1, sticky='e', padx=3)
            tk.Label(closures_frame, text="", bg='#ecf0f1').grid(row=0, column=0)
        except Exception:
            pass

        # הגדרות אנטי-ספאם לכרטיסים - כפתור לפתיחת חלון ניהול
        spam_frame = tk.Frame(content_frame, bg='#ecf0f1')
        spam_frame.pack(fill=tk.X, pady=(15, 5))
        spam_frame.columnconfigure(1, weight=1)
        
        tk.Label(
            spam_frame,
            text=fix_rtl_text("ניהול חסימות אנטי-ספאם:"),
            font=('Arial', 10, 'bold'),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).grid(row=0, column=2, sticky='e', padx=3)
        
        tk.Button(
            spam_frame,
            text="🛡 ניהול חסימות",
            command=self.open_anti_spam_manager,
            font=('Arial', 10, 'bold'),
            bg='#e74c3c',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=4
        ).grid(row=0, column=1, sticky='e', padx=3)
        tk.Label(spam_frame, text="", bg='#ecf0f1').grid(row=0, column=0)
        
        # כפתורי שמירה/סגירה - תמיד בתחתית, ממורכזים
        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=20)
        
        def save_settings():
            cfg = self.load_app_config()
            folder = shared_folder_var.get().strip()

            # טיפול בתיקיית רשת משותפת (DB + Excel) בצורה אוטומטית
            if folder:
                folder = os.path.abspath(folder)
                try:
                    os.makedirs(folder, exist_ok=True)
                except Exception as e:
                    messagebox.showerror("שגיאה", f"לא ניתן ליצור/להשתמש בתיקייה המשותפת:\n{folder}\n\n{e}")
                    return

                # נתיב DB נוכחי (לפני שינוי ההגדרות) לצורך העתקה ראשונית לתיקיית הרשת
                current_db_path = getattr(self.db, 'db_path', None)

                # נסה להשתמש ב-DB קיים בתיקייה המשותפת או להעתיק את הקיים מהמחשב הנוכחי
                shared_db_path = os.path.join(folder, 'school_points.db')
                try:
                    if os.path.exists(shared_db_path):
                        # אם כבר קיים DB בתיקיה המשותפת – רק נבדוק שניתן לעבוד איתו
                        from database import Database as _DBTest
                        _ = _DBTest(db_path=shared_db_path)
                    elif current_db_path and os.path.exists(current_db_path) and os.path.abspath(current_db_path) != os.path.abspath(shared_db_path):
                        # אין DB ברשת אבל יש DB קיים מקומי – נעתיק אותו לתיקיית הרשת
                        try:
                            shutil.copy2(current_db_path, shared_db_path)
                        except Exception as copy_err:
                            messagebox.showerror(
                                "שגיאה",
                                f"לא ניתן להעתיק את מסד הנתונים הקיים לתיקייה המשותפת:\n"
                                f"מ־{current_db_path}\nאל {shared_db_path}\n\n{copy_err}"
                            )
                            return
                    else:
                        # אין DB קיים – ניצור DB חדש בתיקיית הרשת
                        from database import Database as _DBTest
                        _ = _DBTest(db_path=shared_db_path)
                except Exception as e:
                    messagebox.showerror(
                        "שגיאה",
                        f"לא ניתן לעבוד עם מסד נתונים בתיקייה המשותפת:\n{shared_db_path}\n\n{e}"
                    )
                    return

                # נסה ליצור קובץ אקסל שבלונה בתיקייה המשותפת (אם אינו קיים)
                shared_excel_path = os.path.join(folder, "טבלה למבצע אשראי.xlsx")
                if not os.path.exists(shared_excel_path):
                    try:
                        self.importer.export_to_excel(shared_excel_path)
                    except Exception as e:
                        messagebox.showerror(
                            "שגיאה",
                            f"לא ניתן ליצור קובץ Excel בתיקייה המשותפת:\n{shared_excel_path}\n\n{e}"
                        )
                        return

                # שמירת ההגדרות רק לאחר שהכל הצליח
                cfg['shared_folder'] = folder
                cfg.pop('network_root', None)
                cfg['excel_path'] = shared_excel_path

                try:
                    self._seed_shared_sounds_folder(folder)
                except Exception:
                    pass
            else:
                # ללא תיקיית רשת – מחיקה מההגדרות (עבודה לוקאלית לפי ברירת מחדל)
                cfg.pop('shared_folder', None)
                cfg.pop('network_root', None)
                cfg.pop('excel_path', None)
            
            mode = str(mode_var.get() or '').strip() or 'local'
            cfg['deployment_mode'] = mode

            tenant_name = str(tenant_name_var.get() or '').strip()
            if tenant_name:
                cfg['sync_tenant_name'] = tenant_name
            else:
                cfg.pop('sync_tenant_name', None)

            tenant_id = str(tenant_id_var.get() or '').strip()
            if tenant_id:
                cfg['sync_tenant_id'] = tenant_id
            else:
                cfg.pop('sync_tenant_id', None)

            logo = logo_path_var.get().strip()
            if logo:
                cfg['logo_path'] = logo
            else:
                cfg.pop('logo_path', None)
            
            name = campaign_var.get().strip() or "אשראיכם"
            cfg['campaign_name'] = name

            # שמירת הגדרות צלילים (לעמדה הציבורית)
            try:
                cfg['sounds_enabled'] = '1' if bool(sounds_enabled_var.get()) else '0'
            except Exception:
                cfg['sounds_enabled'] = '1'
            try:
                cfg['sound_volume'] = str(int(volume_var.get()))
            except Exception:
                cfg['sound_volume'] = str(cfg.get('sound_volume', '80'))

            # שמירת מצב שקט מתקדם (רשימת טווחים)
            try:
                cfg['quiet_mode_ranges'] = json.dumps(quiet_mode_ranges, ensure_ascii=False)
            except Exception:
                try:
                    cfg['quiet_mode_ranges'] = quiet_mode_ranges
                except Exception:
                    cfg['quiet_mode_ranges'] = []

            # תיקיית תמונות תלמידים
            photos_folder = photos_folder_var.get().strip()
            if photos_folder:
                cfg['photos_folder'] = photos_folder
            else:
                cfg.pop('photos_folder', None)

            if show_default_printer:
                # מדפסת ברירת מחדל (קבלות)
                dp = default_printer_var.get().strip()
                if dp:
                    cfg['default_printer'] = dp
                else:
                    cfg.pop('default_printer', None)

            if self.save_app_config(cfg):
                messagebox.showinfo(
                    "הצלחה",
                    "ההגדרות נשמרו בהצלחה.\n\n"
                    "שינויים בתיקיית הרשת, לוגו, שם המבצע ותיקיית תמונות התלמידים ייכנסו לתוקף בפתיחה הבאה של העמדות."
                )
                dialog.destroy()
        
        tk.Button(
            btn_frame,
            text="✖ סגור",
            command=dialog.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=6
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame,
            text="💾 שמור",
            command=save_settings,
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=6
        ).pack(side=tk.LEFT, padx=8)

    def open_display_settings(self):
        """הגדרות תצוגה לעמדה הציבורית (ערכת צבעים, רקעים וסגנון פנלים)."""
        # רק למנהלים
        if not (self.current_teacher and self.current_teacher['is_admin'] == 1):
            messagebox.showwarning("אזהרה", "רק מנהלים יכולים לשנות הגדרות תצוגה")
            return

        try:
            self._open_display_settings_dialog_original()
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))
        return

    def open_max_points_settings(self):
        """חלון הגדרות 'מקסימום נקודות' (תקרה דינמית)."""
        from datetime import date, datetime, timedelta

        dlg = tk.Toplevel(self.root)
        dlg.title('📈 מקסימום נקודות')
        dlg.configure(bg='#ecf0f1')
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.resizable(True, True)
        try:
            dlg.minsize(760, 420)
        except Exception:
            pass

        try:
            cfg = self.db.get_max_points_config() or {}
        except Exception:
            cfg = {}

        def _iso_to_dmy(s: str) -> str:
            try:
                s = str(s or '').strip()
                if not s:
                    return ''
                d = date.fromisoformat(s)
                return d.strftime('%d-%m-%Y')
            except Exception:
                return str(s or '').strip()

        def _dmy_to_iso(s: str) -> str:
            try:
                s = str(s or '').strip()
                if not s:
                    return ''
                d = datetime.strptime(s, '%d-%m-%Y').date()
                return d.isoformat()
            except Exception:
                return str(s or '').strip()

        start_date_var = tk.StringVar(value=_iso_to_dmy(str(cfg.get('start_date') or date.today().isoformat())))
        daily_var = tk.StringVar(value=str(int(cfg.get('daily_points', 0) or 0)))
        weekly_var = tk.StringVar(value=str(int(cfg.get('weekly_points', 0) or 0)))
        daily_details_var = tk.StringVar(value=str(cfg.get('daily_details') or ''))
        weekly_details_var = tk.StringVar(value=str(cfg.get('weekly_details') or ''))
        warn_within_var = tk.StringVar(value=str(int(cfg.get('warn_within_points', 0) or 0)))
        policy_code = str(cfg.get('policy') or 'none').strip().lower()
        free_additions = list((cfg.get('free_additions') or []))

        daily_by_weekday = dict(cfg.get('daily_points_by_weekday') or {}) if isinstance(cfg.get('daily_points_by_weekday'), dict) else {}

        tk.Label(dlg, text=fix_rtl_text('הגדרות מקסימום נקודות'), font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(12, 8))

        body = tk.Frame(dlg, bg='#ecf0f1')
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=10)

        def _row(label: str):
            r = tk.Frame(body, bg='#ecf0f1')
            r.pack(fill=tk.X, pady=6)
            tk.Label(r, text=fix_rtl_text(label), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
            return r

        def _open_date_picker_for_dmy(var: tk.StringVar):
            picker = tk.Toplevel(dlg)
            picker.title("בחר תאריך")
            picker.transient(dlg)
            picker.grab_set()
            picker.resizable(False, False)

            pf = tk.Frame(picker, padx=10, pady=10)
            pf.pack(fill=tk.BOTH, expand=True)
            tk.Label(pf, text=fix_rtl_text('בחר תאריך'), font=('Arial', 10, 'bold')).pack(pady=(0, 6))

            try:
                cur = str(var.get() or '').strip()
            except Exception:
                cur = ''
            today = date.today()
            d0, m0, y0 = today.day, today.month, today.year
            try:
                if cur and '-' in cur:
                    d0, m0, y0 = [int(x) for x in cur.split('-', 2)]
            except Exception:
                d0, m0, y0 = today.day, today.month, today.year

            rowp = tk.Frame(pf)
            rowp.pack(pady=(0, 6))
            yv = tk.StringVar(value=str(y0))
            mv = tk.StringVar(value=str(m0))
            dv = tk.StringVar(value=str(d0))
            tk.Entry(rowp, textvariable=dv, width=4, justify='center').pack(side=tk.LEFT)
            tk.Label(rowp, text='-').pack(side=tk.LEFT)
            tk.Entry(rowp, textvariable=mv, width=4, justify='center').pack(side=tk.LEFT)
            tk.Label(rowp, text='-').pack(side=tk.LEFT)
            tk.Entry(rowp, textvariable=yv, width=6, justify='center').pack(side=tk.LEFT)

            def _ok_date():
                try:
                    dd = int(dv.get())
                    mm = int(mv.get())
                    yy = int(yv.get())
                    dsel = date(yy, mm, dd)
                except Exception:
                    messagebox.showwarning('שגיאה', 'תאריך לא תקין', parent=picker)
                    return
                var.set(dsel.strftime('%d-%m-%Y'))
                picker.destroy()

            tk.Button(pf, text=fix_rtl_text('אישור'), command=_ok_date, width=10).pack(side=tk.LEFT, padx=5)
            tk.Button(pf, text=fix_rtl_text('ביטול'), command=picker.destroy, width=10).pack(side=tk.RIGHT, padx=5)
            picker.wait_window()

        def _open_details_editor(title: str, var: tk.StringVar):
            d2 = tk.Toplevel(dlg)
            d2.title(title)
            d2.configure(bg='#ecf0f1')
            d2.transient(dlg)
            d2.grab_set()
            d2.resizable(True, True)
            try:
                d2.minsize(520, 260)
            except Exception:
                pass

            tk.Label(d2, text=fix_rtl_text(title), font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(pady=(12, 8))
            txt = tk.Text(d2, wrap='word', font=('Arial', 10))
            txt.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
            try:
                txt.insert('1.0', str(var.get() or ''))
            except Exception:
                pass

            btns2 = tk.Frame(d2, bg='#ecf0f1')
            btns2.pack(pady=10)

            def _save_close():
                try:
                    var.set((txt.get('1.0', 'end') or '').strip())
                except Exception:
                    pass
                d2.destroy()

            tk.Button(btns2, text=fix_rtl_text('✖ סגור'), command=d2.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
            tk.Button(btns2, text=fix_rtl_text('💾 שמור'), command=_save_close, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)

        def _open_free_additions_dialog():
            try:
                self._open_free_additions_dialog_max_points(parent=dlg, free_additions=free_additions, date_picker=_open_date_picker_for_dmy, on_changed=_refresh_calc)
            except Exception as e:
                messagebox.showerror('שגיאה', str(e), parent=dlg)

        def _open_daily_by_weekday_editor():
            d2 = tk.Toplevel(dlg)
            d2.title(fix_rtl_text('נקודות יומיות לפי יום בשבוע'))
            try:
                d2.configure(bg='#ecf0f1')
            except Exception:
                pass
            d2.transient(dlg)
            d2.grab_set()
            d2.resizable(False, False)

            tk.Label(d2, text=fix_rtl_text('נקודות יומיות לפי יום בשבוע'), font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(pady=(12, 8))

            body2 = tk.Frame(d2, bg='#ecf0f1')
            body2.pack(fill=tk.BOTH, expand=True, padx=14, pady=10)

            day_defs = [('א', 6), ('ב', 0), ('ג', 1), ('ד', 2), ('ה', 3), ('ו', 4), ('שבת', 5)]
            vars_by_wd = {}
            for heb, wd in day_defs:
                r = tk.Frame(body2, bg='#ecf0f1')
                r.pack(fill=tk.X, pady=4)
                tk.Label(r, text=fix_rtl_text(f'{heb}:'), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=6, anchor='e').pack(side=tk.RIGHT, padx=6)
                v = tk.StringVar(value=str(int((daily_by_weekday or {}).get(str(wd), 0) or 0)))
                tk.Entry(r, textvariable=v, font=('Arial', 10), width=10, justify='right').pack(side=tk.RIGHT, padx=6)
                vars_by_wd[int(wd)] = v

            btns2 = tk.Frame(d2, bg='#ecf0f1')
            btns2.pack(pady=12)

            def _save_close():
                out = {}
                for wd, v in vars_by_wd.items():
                    try:
                        out[str(int(wd))] = int(str(v.get() or '0').strip() or 0)
                    except Exception:
                        out[str(int(wd))] = 0
                try:
                    daily_by_weekday.clear()
                    daily_by_weekday.update(out)
                except Exception:
                    pass
                try:
                    _refresh_calc()
                except Exception:
                    pass
                d2.destroy()

            tk.Button(btns2, text=fix_rtl_text('✖ סגור'), command=d2.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
            tk.Button(btns2, text=fix_rtl_text('💾 שמור'), command=_save_close, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)

        r0 = _row('תאריך התחלה:')
        start_ent = tk.Entry(r0, textvariable=start_date_var, font=('Consolas', 10), width=14, justify='right')
        start_ent.pack(side=tk.RIGHT, padx=6)
        try:
            start_ent.bind('<Button-1>', lambda _e: _open_date_picker_for_dmy(start_date_var))
        except Exception:
            pass
        tk.Label(r0, text=fix_rtl_text('(DD-MM-YYYY)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

        r1 = _row('נקודות יומיות:')
        tk.Button(
            r1,
            text=fix_rtl_text('📅 עריכה'),
            command=_open_daily_by_weekday_editor,
            font=('Arial', 10, 'bold'),
            bg='#2980b9',
            fg='white',
            padx=16,
            pady=4
        ).pack(side=tk.RIGHT, padx=6)
        tk.Button(
            r1,
            text=fix_rtl_text('פירוט'),
            command=lambda: _open_details_editor('פירוט נקודות יומיות', daily_details_var),
            font=('Arial', 9),
            bg='#bdc3c7',
            fg='#2c3e50',
            padx=10,
            pady=2
        ).pack(side=tk.RIGHT, padx=6)

        r2 = _row('נקודות שבועיות:')
        tk.Entry(r2, textvariable=weekly_var, font=('Arial', 10), width=10, justify='right').pack(side=tk.RIGHT, padx=6)
        tk.Button(
            r2,
            text=fix_rtl_text('פירוט'),
            command=lambda: _open_details_editor('פירוט נקודות שבועיות', weekly_details_var),
            font=('Arial', 9),
            bg='#bdc3c7',
            fg='#2c3e50',
            padx=10,
            pady=2
        ).pack(side=tk.RIGHT, padx=6)

        r2b = _row('ניקוד נוסף חופשי:')
        tk.Button(
            r2b,
            text=fix_rtl_text('📋 פירוט / עריכה'),
            command=lambda: _open_free_additions_dialog(),
            font=('Arial', 10, 'bold'),
            bg='#8e44ad',
            fg='white',
            padx=16,
            pady=4
        ).pack(side=tk.RIGHT, padx=6)

        r3 = _row('אזהרה אם נשארו:')
        tk.Entry(r3, textvariable=warn_within_var, font=('Arial', 10), width=10, justify='right').pack(side=tk.RIGHT, padx=6)
        tk.Label(r3, text=fix_rtl_text('נקודות למקסימום'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

        r4 = _row('מדיניות חריגה:')
        pol_map = {
            'לא לעשות דבר': 'none',
            'להציג אזהרה': 'warn',
            'לחסום חריגה': 'block',
        }
        pol_rev = {v: k for k, v in pol_map.items()}
        policy_var = tk.StringVar(value=pol_rev.get(policy_code, 'לא לעשות דבר'))
        ttk.Combobox(r4, textvariable=policy_var, values=list(pol_map.keys()), state='readonly', width=18, justify='right').pack(side=tk.RIGHT, padx=6)

        calc_var = tk.StringVar(value='')
        tk.Label(body, textvariable=calc_var, bg='#ecf0f1', fg='#2c3e50', anchor='e').pack(fill=tk.X, pady=(10, 0))

        def _refresh_calc():
            # calculate using DB helper by building a tmp cfg and calling compute_max_points_allowed
            try:
                sd_dmy = str(start_date_var.get() or '').strip()
                sd = _dmy_to_iso(sd_dmy) if sd_dmy else date.today().isoformat()
                wv = int(str(weekly_var.get() or '0').strip() or 0)

                # נקודות יומיות נקבעות רק לפי יום בשבוע
                dv = 0

                cfg_tmp = {
                    'start_date': sd,
                    'daily_points': int(dv),
                    'daily_points_by_weekday': dict(daily_by_weekday or {}),
                    'weekly_points': int(wv),
                    'free_additions': list(free_additions or []),
                }
                total = int(self.db.compute_max_points_allowed_from_cfg(cfg_tmp, for_date=date.today().isoformat()))
                calc_var.set(fix_rtl_text(f"מקסימום נקודות נכון להיום: {max(0, int(total))}"))
            except Exception:
                calc_var.set('')

        def _save():
            sd_dmy = str(start_date_var.get() or '').strip()
            sd = _dmy_to_iso(sd_dmy)
            if not sd:
                messagebox.showerror('שגיאה', 'חובה להזין תאריך התחלה', parent=dlg)
                return
            try:
                datetime.strptime(sd, '%Y-%m-%d')
            except Exception:
                messagebox.showerror('שגיאה', 'תאריך לא תקין (DD-MM-YYYY)', parent=dlg)
                return
            try:
                wv = int(str(weekly_var.get() or '0').strip() or 0)
                ww = int(str(warn_within_var.get() or '0').strip() or 0)
            except Exception:
                messagebox.showerror('שגיאה', 'נקודות/סף אזהרה חייבים להיות מספרים', parent=dlg)
                return
            dv = 0
            cfg2 = {
                'start_date': sd,
                'daily_points': int(dv),
                'daily_points_by_weekday': dict(daily_by_weekday or {}),
                'weekly_points': int(wv),
                'daily_details': str(daily_details_var.get() or ''),
                'weekly_details': str(weekly_details_var.get() or ''),
                'warn_within_points': int(ww),
                'policy': pol_map.get(str(policy_var.get() or '').strip(), 'none'),
                'free_additions': list(free_additions or []),
            }
            try:
                if not self.db.set_max_points_config(cfg2):
                    messagebox.showerror('שגיאה', 'לא ניתן לשמור את ההגדרות', parent=dlg)
                    return
            except Exception as e:
                messagebox.showerror('שגיאה', str(e), parent=dlg)
                return
            messagebox.showinfo('הצלחה', 'ההגדרות נשמרו', parent=dlg)
            _refresh_calc()
            try:
                dlg.destroy()
            except Exception:
                pass

        # live calc on changes
        for v in (start_date_var, weekly_var, warn_within_var, policy_var):
            try:
                v.trace_add('write', lambda *_: _refresh_calc())
            except Exception:
                pass
        _refresh_calc()

        btns = tk.Frame(dlg, bg='#ecf0f1')
        btns.pack(pady=12)
        tk.Button(btns, text='✖ סגור', command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text='💾 שמור', command=_save, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)

    def _check_max_points_policy(self, *, proposed_points: int, parent=None) -> bool:
        """אכיפה/אזהרה לפי הגדרות מקסימום נקודות. מחזיר האם לאפשר את העדכון."""
        try:
            ev = self.db.evaluate_points_against_max(proposed_points=int(proposed_points))
        except Exception:
            return True

        try:
            policy = str(ev.get('policy') or 'none').strip().lower()
        except Exception:
            policy = 'none'

        status = str(ev.get('status') or 'ok').strip().lower()
        max_allowed = int(ev.get('max_allowed', 0) or 0)
        proposed = int(ev.get('proposed_points', proposed_points) or 0)

        if policy == 'none':
            return True

        if policy == 'block' and status == 'exceed':
            try:
                messagebox.showwarning(
                    'חריגה ממקסימום נקודות',
                    fix_rtl_text(f"לא ניתן לעדכן נקודות ל-{proposed} כי המקסימום המותר כיום הוא {max_allowed}."),
                    parent=parent or self.root
                )
            except Exception:
                pass
            return False

        if policy == 'warn' and status in ('near', 'exceed'):
            try:
                msg = f"המקסימום המותר כיום הוא {max_allowed}.\nהערך המבוקש הוא {proposed}.\n\nלהמשיך בכל זאת?"
                return bool(messagebox.askyesno('אזהרה: קרוב/מעל מקסימום נקודות', fix_rtl_text(msg), parent=parent or self.root))
            except Exception:
                return True

        return True

    def _open_free_additions_dialog_max_points(self, *, parent, free_additions: list, date_picker=None, on_changed=None):
        from datetime import date, datetime

        dlg = tk.Toplevel(parent)
        dlg.title(fix_rtl_text('ניקוד נוסף חופשי'))
        try:
            dlg.configure(bg='#ecf0f1')
        except Exception:
            pass
        dlg.transient(parent)
        dlg.grab_set()
        dlg.resizable(True, True)
        try:
            dlg.minsize(760, 420)
        except Exception:
            pass

        tk.Label(dlg, text=fix_rtl_text('ניקוד נוסף חופשי'), font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(12, 8))

        table = tk.Frame(dlg, bg='#ecf0f1')
        table.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        cols = ('date', 'points', 'note')
        tree = ttk.Treeview(table, columns=cols, show='headings', height=12)
        tree.heading('date', text=fix_rtl_text('תאריך'))
        tree.heading('points', text=fix_rtl_text('נקודות'))
        tree.heading('note', text=fix_rtl_text('הערה'))
        tree.column('date', width=130, anchor='center')
        tree.column('points', width=90, anchor='center')
        tree.column('note', width=430, anchor='e')
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        sb = ttk.Scrollbar(table, orient='vertical', command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        try:
            tree.configure(yscrollcommand=sb.set)
        except Exception:
            pass

        def iso_to_dmy(s: str) -> str:
            try:
                return date.fromisoformat(str(s)).strftime('%d-%m-%Y')
            except Exception:
                return str(s or '').strip()

        def dmy_to_iso(s: str) -> str:
            try:
                return datetime.strptime(str(s or '').strip(), '%d-%m-%Y').date().isoformat()
            except Exception:
                return ''

        def normalize_inplace():
            rows = []
            try:
                rows = list(free_additions or [])
            except Exception:
                rows = []
            norm = []
            for r in rows:
                if not isinstance(r, dict):
                    continue
                ds = str(r.get('date') or '').strip()
                if not ds:
                    continue
                try:
                    datetime.strptime(ds, '%Y-%m-%d')
                except Exception:
                    continue
                try:
                    pts = int(r.get('points', 0) or 0)
                except Exception:
                    pts = 0
                try:
                    note = str(r.get('note') or '').strip()
                except Exception:
                    note = ''
                norm.append({'date': ds, 'points': int(pts), 'note': note})
            norm.sort(key=lambda x: str(x.get('date') or ''))
            try:
                free_additions[:] = norm
            except Exception:
                pass

        def refresh_tree():
            normalize_inplace()
            try:
                for iid in tree.get_children():
                    tree.delete(iid)
            except Exception:
                pass
            for idx, r in enumerate(list(free_additions or [])):
                ds = str(r.get('date') or '').strip()
                pts = int(r.get('points', 0) or 0)
                note = str(r.get('note') or '').strip()
                tree.insert('', 'end', iid=str(idx), values=(iso_to_dmy(ds), str(pts), fix_rtl_text(note) if note else ''))
            try:
                if callable(on_changed):
                    on_changed()
            except Exception:
                pass

        def sel_index() -> int:
            try:
                sel = tree.selection() or ()
                if not sel:
                    return -1
                return int(sel[0])
            except Exception:
                return -1

        def open_editor(initial: dict = None):
            init = dict(initial or {})
            ed = tk.Toplevel(dlg)
            ed.title(fix_rtl_text('חריגה'))
            try:
                ed.configure(bg='#ecf0f1')
            except Exception:
                pass
            ed.transient(dlg)
            ed.grab_set()
            ed.resizable(False, False)

            frm = tk.Frame(ed, bg='#ecf0f1')
            frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

            date_var = tk.StringVar(value=iso_to_dmy(str(init.get('date') or '')) or date.today().strftime('%d-%m-%Y'))
            points_var = tk.StringVar(value=str(int(init.get('points', 0) or 0)))
            note_var = tk.StringVar(value=str(init.get('note') or ''))

            r0 = tk.Frame(frm, bg='#ecf0f1')
            r0.pack(fill=tk.X, pady=6)
            tk.Label(r0, text=fix_rtl_text('תאריך:'), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=12, anchor='e').pack(side=tk.RIGHT, padx=6)
            ent_date = tk.Entry(r0, textvariable=date_var, font=('Consolas', 10), width=14, justify='right')
            ent_date.pack(side=tk.RIGHT, padx=6)
            try:
                if callable(date_picker):
                    ent_date.bind('<Button-1>', lambda _e: date_picker(date_var))
            except Exception:
                pass
            tk.Label(r0, text=fix_rtl_text('(DD-MM-YYYY)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r1 = tk.Frame(frm, bg='#ecf0f1')
            r1.pack(fill=tk.X, pady=6)
            tk.Label(r1, text=fix_rtl_text('נקודות:'), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=12, anchor='e').pack(side=tk.RIGHT, padx=6)
            tk.Entry(r1, textvariable=points_var, font=('Arial', 10), width=10, justify='right').pack(side=tk.RIGHT, padx=6)

            r2 = tk.Frame(frm, bg='#ecf0f1')
            r2.pack(fill=tk.X, pady=6)
            tk.Label(r2, text=fix_rtl_text('הערה:'), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=12, anchor='e').pack(side=tk.RIGHT, padx=6)
            tk.Entry(r2, textvariable=note_var, font=('Arial', 10), width=42, justify='right').pack(side=tk.RIGHT, padx=6)

            out = {'data': None}

            def ok():
                ds = dmy_to_iso(str(date_var.get() or '').strip())
                if not ds:
                    messagebox.showerror('שגיאה', 'תאריך לא תקין (DD-MM-YYYY)', parent=ed)
                    return
                try:
                    pts = int(str(points_var.get() or '0').strip() or 0)
                except Exception:
                    messagebox.showerror('שגיאה', 'נקודות חייב להיות מספר', parent=ed)
                    return
                try:
                    note = str(note_var.get() or '').strip()
                except Exception:
                    note = ''
                out['data'] = {'date': ds, 'points': int(pts), 'note': note}
                ed.destroy()

            btns = tk.Frame(frm, bg='#ecf0f1')
            btns.pack(pady=(10, 0))
            tk.Button(btns, text=fix_rtl_text('✖ ביטול'), command=ed.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
            tk.Button(btns, text=fix_rtl_text('✓ אישור'), command=ok, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)

            ed.wait_window()
            return out.get('data')

        def add_row():
            data = open_editor(None)
            if not data:
                return
            try:
                free_additions.append(data)
            except Exception:
                pass
            refresh_tree()

        def edit_row():
            idx = sel_index()
            if idx < 0 or idx >= len(free_additions or []):
                return
            cur = None
            try:
                cur = dict(free_additions[idx])
            except Exception:
                cur = None
            data = open_editor(cur)
            if not data:
                return
            try:
                free_additions[idx] = data
            except Exception:
                pass
            refresh_tree()

        def del_row():
            idx = sel_index()
            if idx < 0 or idx >= len(free_additions or []):
                return
            try:
                if not messagebox.askyesno('אישור', fix_rtl_text('למחוק את השורה שנבחרה?'), parent=dlg):
                    return
            except Exception:
                pass
            try:
                del free_additions[idx]
            except Exception:
                pass
            refresh_tree()

        try:
            tree.bind('<Double-1>', lambda _e: edit_row())
        except Exception:
            pass

        btns = tk.Frame(dlg, bg='#ecf0f1')
        btns.pack(pady=10)
        tk.Button(btns, text=fix_rtl_text('➕ הוספה'), command=add_row, font=('Arial', 10, 'bold'), bg='#2980b9', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text=fix_rtl_text('✏ עריכה'), command=edit_row, font=('Arial', 10, 'bold'), bg='#8e44ad', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text=fix_rtl_text('🗑 מחיקה'), command=del_row, font=('Arial', 10, 'bold'), bg='#c0392b', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)
        tk.Button(btns, text=fix_rtl_text('✖ סגור'), command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=8)

        refresh_tree()

    def _open_display_settings_dialog_original(self):
        config = self.load_app_config()

        dialog = tk.Toplevel(self.root)
        dialog.title("🖥 הגדרות תצוגה")
        dialog.configure(bg='#ecf0f1')
        dialog.minsize(750, 550)
        dialog.geometry("850x600")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        content_frame = tk.Frame(dialog, bg='#ecf0f1')
        content_frame.pack(fill=tk.X, padx=40, pady=10)

        # ערכת צבעים לעמדה הציבורית – שני מצבים בלבד (כהה / בהיר)
        theme_map = {
            "רקע כהה, כיתוב בהיר": "dark",
            "רקע בהיר, כיתוב כהה": "light",
        }

        # מצב רקע לעמדה הציבורית
        # "ברירת מחדל (דף רקע אוטומטי)" ישתמש בתבנית הגרפית המובנית (template1)
        background_mode_map = {
            "ברירת מחדל (דף רקע אוטומטי)": "default",
            "תמונת רקע אחת": "image",
            "מצגת תמונות (תיקייה)": "slideshow",
            "צבע רקע אחיד (טבלת צבעים)": "color",
        }

        # מצב פריסת תמונת הרקע (בדומה לשולחן עבודה)
        background_layout_map = {
            "מילוי": "cover",      # ממלא את המסך (חיתוך בשוליים)
            "התאמה": "contain",    # מתאים למסך עם שוליים אם צריך
            "מתח": "stretch",      # מתיחה מדויקת לגודל המסך
            "אריח": "tile",        # חזרה על התמונה
            "מרכז": "center",      # במרכז המסך
            "פרישה": "cover",      # שם נוסף ל"מילוי"
        }

        # כיוון מסך בעמדה הציבורית
        orientation_map = {
            "מסך לרוחב (אופקי)": "landscape",
            "מסך לאורך (אנכי)": "portrait",
        }

        raw_theme = config.get('theme', 'dark')
        current_bg_mode_code = config.get('background_mode', 'default') or 'default'
        current_theme_code = raw_theme

        # תאימות לאחור: אם theme נשמר כ-background_image / slideshow
        if raw_theme in ('background_image', 'slideshow') and current_bg_mode_code in ('none', '', None):
            # בעבר ערך זה סימל "רקע תמונה" / "מצגת" – היום נתייחס לכך כאל ערכת צבעים כהה
            current_theme_code = 'dark'
            current_bg_mode_code = 'image' if raw_theme == 'background_image' else 'slideshow'

        # תאימות לאחור: ערכת "ניגודיות גבוהה" מתורגמת לערכת כהה
        if current_theme_code in ('high_contrast', 'background_image', 'slideshow', 'none'):
            current_theme_code = 'dark'

        reverse_theme_map = {v: k for k, v in theme_map.items()}
        theme_var = tk.StringVar(value=reverse_theme_map.get(current_theme_code, "רקע כהה, כיתוב בהיר"))

        # תאימות לאחור: מצב 'none' הישן יתנהג כעת כברירת מחדל (דף רקע אוטומטי)
        if current_bg_mode_code in ('none', '', None):
            current_bg_mode_code = 'default'

        reverse_bg_mode_map = {v: k for k, v in background_mode_map.items()}
        bg_mode_var = tk.StringVar(value=reverse_bg_mode_map.get(current_bg_mode_code, "ברירת מחדל (דף רקע אוטומטי)"))

        current_bg_layout_code = config.get('background_layout', 'cover')
        reverse_bg_layout_map = {v: k for k, v in background_layout_map.items()}
        bg_layout_var = tk.StringVar(value=reverse_bg_layout_map.get(current_bg_layout_code, "מילוי"))

        background_image_var = tk.StringVar(value=config.get('background_image_path', ""))
        background_folder_var = tk.StringVar(value=config.get('background_folder', ""))
        background_color_var = tk.StringVar(value=config.get('background_color', "#000000"))

        slide_interval_var = tk.StringVar(value=str(config.get('background_interval_sec', 15)))

        slideshow_mode_map = {
            "תמונה אחת מתחלפת": "single",
            "מונטאז' ריבועים סטטי": "grid_static",
            "מונטאז' ריבועים מתחלפים": "grid_dynamic",
        }
        current_slideshow_mode = config.get('slideshow_display_mode', 'single')
        reverse_slideshow_mode_map = {v: k for k, v in slideshow_mode_map.items()}
        slideshow_mode_var = tk.StringVar(value=reverse_slideshow_mode_map.get(current_slideshow_mode, "תמונה אחת מתחלפת"))

        slideshow_cols_var = tk.StringVar(value=str(config.get('slideshow_grid_cols', 4)))

        panel_style_map = {
            "פנלים רגילים (מלאים)": "solid",
            "פנלים שקופים חלקית": "floating",
        }
        current_panel_style = config.get('panel_style', 'floating')
        reverse_panel_style_map = {v: k for k, v in panel_style_map.items()}
        panel_style_var = tk.StringVar(value=reverse_panel_style_map.get(current_panel_style, "פנלים רגילים (מלאים)"))

        # כיוון מסך נוכחי
        current_orientation_code = config.get('screen_orientation', 'landscape')
        reverse_orientation_map = {v: k for k, v in orientation_map.items()}
        orientation_var = tk.StringVar(value=reverse_orientation_map.get(current_orientation_code, "מסך לרוחב (אופקי)"))

        FIELD_WIDTH = 42
        BUTTON_WIDTH = 14
        LABEL_WIDTH = 40

        # שורה 1 - כיוון מסך
        row1 = tk.Frame(content_frame, bg='#ecf0f1')
        row1.pack(fill=tk.X, pady=6)
        tk.Label(row1, text=fix_rtl_text("כיוון מסך בעמדה הציבורית:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        orientation_choices = list(orientation_map.keys())
        orientation_menu = tk.OptionMenu(row1, orientation_var, *orientation_choices)
        orientation_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        orientation_menu.pack(side=tk.RIGHT, padx=5)

        # שורה 2 - ערכת צבעים לעמדה הציבורית
        row2 = tk.Frame(content_frame, bg='#ecf0f1')
        row2.pack(fill=tk.X, pady=6)
        tk.Label(row2, text=fix_rtl_text("ערכת צבעים לעמדה הציבורית:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        theme_choices = list(theme_map.keys())
        theme_menu = tk.OptionMenu(row2, theme_var, *theme_choices)
        theme_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        theme_menu.pack(side=tk.RIGHT, padx=5)

        # שורה 3 - מצב רקע
        row3b = tk.Frame(content_frame, bg='#ecf0f1')
        row3b.pack(fill=tk.X, pady=6)
        tk.Label(row3b, text=fix_rtl_text("מצב רקע:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_mode_choices = list(background_mode_map.keys())
        bg_mode_menu = tk.OptionMenu(row3b, bg_mode_var, *bg_mode_choices)
        bg_mode_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        bg_mode_menu.pack(side=tk.RIGHT, padx=5)

        # מסגרת לשורות הרקע הדינמיות
        bg_rows_frame = tk.Frame(content_frame, bg='#ecf0f1')

        # פריסת רקע
        row4 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row4, text=fix_rtl_text("פריסת רקע:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_layout_choices = list(background_layout_map.keys())
        bg_layout_menu = tk.OptionMenu(row4, bg_layout_var, *bg_layout_choices)
        bg_layout_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        bg_layout_menu.pack(side=tk.RIGHT, padx=5)

        # תמונת רקע בודדת
        row5 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row5, text=fix_rtl_text("תמונת רקע (למצב 'תמונת רקע אחת'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_image_entry = tk.Entry(row5, textvariable=background_image_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_image_entry.pack(side=tk.RIGHT, padx=5)

        def browse_bg_image():
            initial = os.path.dirname(background_image_var.get()) if background_image_var.get() else self.base_dir
            path = filedialog.askopenfilename(
                title="בחר תמונת רקע",
                filetypes=[("קבצי תמונה", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("הכל", "*.*")],
                initialdir=initial
            )
            if path:
                background_image_var.set(path)

        tk.Button(row5, text="בחר קובץ", command=browse_bg_image, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # תיקיית מצגת
        row6 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row6, text=fix_rtl_text("תיקיית תמונות (למצב 'מצגת תמונות'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_folder_entry = tk.Entry(row6, textvariable=background_folder_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_folder_entry.pack(side=tk.RIGHT, padx=5)

        def browse_bg_folder():
            initial = background_folder_var.get().strip() or self.base_dir
            path = filedialog.askdirectory(title="בחר תיקיית תמונות למצגת", initialdir=initial)
            if path:
                background_folder_var.set(path)

        tk.Button(row6, text="בחר תיקייה", command=browse_bg_folder, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # צבע רקע אחיד
        row6b = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row6b, text=fix_rtl_text("צבע רקע (למצב 'צבע אחיד'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_color_entry = tk.Entry(row6b, textvariable=background_color_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_color_entry.pack(side=tk.RIGHT, padx=5)

        def choose_bg_color():
            initial = background_color_var.get().strip() or "#000000"
            try:
                color = colorchooser.askcolor(color=initial, title="בחר צבע רקע אחיד")
            except Exception:
                color = None

            hex_color = None
            if color:
                if len(color) > 1 and color[1]:
                    hex_color = color[1]
                elif color[0] and isinstance(color[0], (tuple, list)):
                    try:
                        r, g, b = map(int, color[0])
                        hex_color = f"#{r:02x}{g:02x}{b:02x}"
                    except Exception:
                        hex_color = None

            if hex_color:
                background_color_var.set(hex_color)

        tk.Button(row6b, text="בחר צבע", command=choose_bg_color, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # זמן מעבר במצגת
        row7 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7, text=fix_rtl_text("זמן מעבר במצגת (שניות):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        tk.Entry(row7, textvariable=slide_interval_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right').pack(side=tk.RIGHT, padx=5)

        # מצב תצוגת מצגת
        row7b = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7b, text=fix_rtl_text("מצב תצוגת מצגת:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        slideshow_mode_choices = list(slideshow_mode_map.keys())
        slideshow_mode_menu = tk.OptionMenu(row7b, slideshow_mode_var, *slideshow_mode_choices)
        slideshow_mode_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        slideshow_mode_menu.pack(side=tk.RIGHT, padx=5)

        # מספר עמודות במונטאז'
        row7c = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7c, text=fix_rtl_text("מספר עמודות למונטאז' (1-10):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        tk.Entry(row7c, textvariable=slideshow_cols_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right').pack(side=tk.RIGHT, padx=5)

        # סגנון פנלים
        row8 = tk.Frame(content_frame, bg='#ecf0f1')
        row8.pack(fill=tk.X, pady=10)
        tk.Label(row8, text=fix_rtl_text("סגנון פנלים מעל הרקע:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        panel_style_choices = list(panel_style_map.keys())
        panel_style_menu = tk.OptionMenu(row8, panel_style_var, *panel_style_choices)
        panel_style_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        panel_style_menu.pack(side=tk.RIGHT, padx=5)

        # הגדרת הצגת סטטיסטיקות בעמדה הציבורית (ברירת מחדל: כבוי)
        show_stats_value = self.db.get_setting('show_statistics', '0')
        show_stats_var = tk.BooleanVar(value=show_stats_value == '1')

        stats_frame = tk.Frame(content_frame, bg='#ecf0f1')
        stats_frame.pack(fill=tk.X, pady=(5, 5))

        tk.Label(
            stats_frame,
            text=fix_rtl_text("הצגת סטטיסטיקות בעמדה הציבורית:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).pack(side=tk.RIGHT, padx=5)
        ToggleSwitch(
            stats_frame,
            variable=show_stats_var
        ).pack(side=tk.RIGHT, padx=5)

        # הגדרת הצגת תמונת תלמיד בעמדה הציבורית (ברירת מחדל: כבוי)
        show_photo_value = self.db.get_setting('show_student_photo', '0')
        show_photo_var = tk.BooleanVar(value=show_photo_value == '1')

        photo_frame = tk.Frame(content_frame, bg='#ecf0f1')
        photo_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(
            photo_frame,
            text=fix_rtl_text("הצגת תמונת תלמיד בעמדה הציבורית:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).pack(side=tk.RIGHT, padx=5)
        ToggleSwitch(
            photo_frame,
            variable=show_photo_var
        ).pack(side=tk.RIGHT, padx=5)


        # מהירות טיקר חדשות
        # שינוי סקאלה: המהירות שהייתה עד היום "רגיל" תיחשב כ"איטי" (normal).
        # "רגיל" יהיה מהיר יותר (fast), ו"מהיר" יהיה הכי מהיר (very_fast).
        ticker_speed_map = {
            "איטי": "normal",
            "רגיל": "fast",
            "מהיר": "very_fast",
        }
        current_ticker_speed = config.get('news_ticker_speed', 'normal')
        reverse_ticker_speed_map = {v: k for k, v in ticker_speed_map.items()}
        ticker_speed_var = tk.StringVar(value=reverse_ticker_speed_map.get(current_ticker_speed, "איטי"))
        
        ticker_frame = tk.Frame(content_frame, bg='#ecf0f1')
        ticker_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(
            ticker_frame,
            text=fix_rtl_text("מהירות טיקר חדשות:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).pack(side=tk.RIGHT, padx=5)
        
        ticker_speed_choices = list(ticker_speed_map.keys())
        ticker_speed_menu = tk.OptionMenu(ticker_frame, ticker_speed_var, *ticker_speed_choices)
        ticker_speed_menu.config(font=('Arial', 10), width=20)
        ticker_speed_menu.pack(side=tk.RIGHT, padx=5)

        def update_background_rows(*args):
            mode_code = background_mode_map.get(bg_mode_var.get(), 'none')

            bg_rows_frame.pack_forget()
            for row in (row4, row5, row6, row6b, row7, row7b, row7c):
                row.pack_forget()

            if mode_code == 'image':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row5.pack(fill=tk.X, pady=6)
                row4.pack(fill=tk.X, pady=6)
            elif mode_code == 'slideshow':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row6.pack(fill=tk.X, pady=6)
                row4.pack(fill=tk.X, pady=6)
                row7.pack(fill=tk.X, pady=6)
                row7b.pack(fill=tk.X, pady=6)
                row7c.pack(fill=tk.X, pady=6)
            elif mode_code == 'color':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row6b.pack(fill=tk.X, pady=6)

            try:
                dialog.update_idletasks()
                dialog.geometry("")
            except Exception:
                pass

        update_background_rows()
        bg_mode_var.trace_add('write', lambda *args: update_background_rows())

        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=20)

        def save_settings():
            cfg = self.load_app_config() or {}

            # כיוון מסך
            selected_orientation_label = orientation_var.get()
            cfg['screen_orientation'] = orientation_map.get(selected_orientation_label, 'landscape')

            # ערכת צבעים לעמדה הציבורית
            selected_theme_label = theme_var.get()
            cfg['theme'] = theme_map.get(selected_theme_label, 'dark')

            # מצב רקע
            selected_bg_label = bg_mode_var.get()
            bg_mode_code = background_mode_map.get(selected_bg_label, 'default')
            cfg['background_mode'] = bg_mode_code

            # קביעת דף הרקע הגרפי בהתאם למצב הרקע
            # ברירת מחדל – שימוש בתבנית הגרפית template1 עם דפי רקע אוטומטיים
            cfg['background_template'] = 'template1'

            # תמונת רקע
            bg_image = background_image_var.get().strip()
            if bg_image:
                cfg['background_image_path'] = bg_image
            else:
                cfg.pop('background_image_path', None)

            # תיקיית מצגת
            bg_folder = background_folder_var.get().strip()
            if bg_folder:
                cfg['background_folder'] = bg_folder
            else:
                cfg.pop('background_folder', None)

            # צבע רקע אחיד
            bg_color = background_color_var.get().strip()
            if bg_color:
                cfg['background_color'] = bg_color
            else:
                cfg.pop('background_color', None)

            # זמן מעבר במצגת
            interval_str = slide_interval_var.get().strip()
            try:
                interval_val = int(interval_str)
                if interval_val < 3:
                    interval_val = 3
                if interval_val > 600:
                    interval_val = 600
                cfg['background_interval_sec'] = interval_val
            except ValueError:
                cfg.pop('background_interval_sec', None)

            # מצב תצוגת מצגת
            selected_slideshow_label = slideshow_mode_var.get()
            cfg['slideshow_display_mode'] = slideshow_mode_map.get(selected_slideshow_label, 'single')

            # מספר עמודות במונטאז'
            cols_str = slideshow_cols_var.get().strip()
            try:
                cols_val = int(cols_str)
                if cols_val < 1:
                    cols_val = 1
                if cols_val > 10:
                    cols_val = 10
                cfg['slideshow_grid_cols'] = cols_val
            except ValueError:
                cfg.pop('slideshow_grid_cols', None)

            # פריסת רקע
            selected_layout_label = bg_layout_var.get()
            cfg['background_layout'] = background_layout_map.get(selected_layout_label, 'cover')

            # סגנון פנלים
            selected_panel_label = panel_style_var.get()
            cfg['panel_style'] = panel_style_map.get(selected_panel_label, 'solid')

            # מהירות טיקר חדשות
            selected_ticker_speed_label = ticker_speed_var.get()
            cfg['news_ticker_speed'] = ticker_speed_map.get(selected_ticker_speed_label, 'normal')

            # שמירת הגדרת הצגת סטטיסטיקות בעמדה הציבורית
            try:
                show = '1' if show_stats_var.get() else '0'
                self.db.set_setting('show_statistics', show)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לעדכן הגדרת סטטיסטיקות:\n{e}")
                return

            # שמירת הגדרת הצגת תמונת תלמיד בעמדה הציבורית
            try:
                show_photo = '1' if show_photo_var.get() else '0'
                self.db.set_setting('show_student_photo', show_photo)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לעדכן הגדרת תמונת תלמיד:\n{e}")
                return

            if self.save_app_config(cfg):
                # שליחת הפעלה מחדש אוטומטית לכל העמדות הציבוריות
                restart_cfg = dict(cfg) if isinstance(cfg, dict) else {}
                if isinstance(restart_cfg, dict):
                    old_token = restart_cfg.get('restart_public_stations_token')
                    try:
                        token_val = int(old_token)
                        token_val += 1
                    except Exception:
                        token_val = 1
                    restart_cfg['restart_public_stations_token'] = str(token_val)
                    self.save_app_config(restart_cfg)

                messagebox.showinfo(
                    "הפעלה מחדש נשלחה",
                    "הגדרות התצוגה נשמרו בהצלחה.\n\n"
                    "המערכת שלחה הוראה להפעיל מחדש את כל העמדות הציבוריות.\n"
                    "שינויים בלוגו וברקע הגרפי יופיעו בתוך מספר שניות."
                )
                dialog.destroy()
            else:
                messagebox.showerror("שגיאה", "לא ניתן לשמור את ההגדרות")

        def send_restart_only():
            try:
                if not messagebox.askyesno(
                    "אזהרה",
                    "פעולה זו תשלח הוראה להפעלה מחדש של כל העמדות הציבוריות.\n\nלהמשיך?",
                ):
                    return
            except Exception:
                pass

            try:
                cfg2 = self.load_app_config() or {}
            except Exception:
                cfg2 = {}
            restart_cfg = dict(cfg2) if isinstance(cfg2, dict) else {}
            try:
                old_token = restart_cfg.get('restart_public_stations_token')
                try:
                    token_val = int(old_token)
                    token_val += 1
                except Exception:
                    token_val = 1
                restart_cfg['restart_public_stations_token'] = str(token_val)
                if self.save_app_config(restart_cfg):
                    messagebox.showinfo("נשלח", "נשלחה הוראת הפעלה מחדש לעמדות הציבוריות.")
                else:
                    messagebox.showerror("שגיאה", "לא ניתן לשלוח הוראת הפעלה מחדש")
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))

        tk.Button(
            btn_frame,
            text="🔴 הפעלה מחדש לעמדות",
            command=send_restart_only,
            font=('Arial', 11, 'bold'),
            bg='#e74c3c',
            fg='white',
            width=18,
            padx=10,
            pady=6
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="שמור",
            command=save_settings,
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            width=12,
            padx=10,
            pady=6
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            btn_frame,
            text="סגור",
            command=dialog.destroy,
            font=('Arial', 11),
            bg='#95a5a6',
            fg='white',
            width=12,
            padx=10,
            pady=6
        ).pack(side=tk.LEFT, padx=5)

    def _print_or_open_pdf_for_receipts(self, pdf_path: str):
        pdf_path = str(pdf_path or '').strip()
        if not pdf_path or not os.path.exists(pdf_path):
            return
        printer = ''
        try:
            cfg = self.load_app_config() or {}
        except Exception:
            cfg = {}
        try:
            printer = str((cfg or {}).get('default_printer') or '').strip()
        except Exception:
            printer = ''
        if printer:
            try:
                os.startfile(pdf_path, 'printto', printer)
                return
            except Exception:
                pass
        try:
            os.startfile(pdf_path)
        except Exception:
            pass

    def _generate_receipt_pdf_from_snapshot(self, snapshot: dict) -> str:
        snap = snapshot or {}
        data = snap.get('data') or {}
        if not isinstance(data, dict):
            data = {}

        try:
            student_id = int(data.get('student_id') or 0)
        except Exception:
            student_id = 0
        try:
            st = self.db.get_student_by_id(int(student_id or 0))
        except Exception:
            st = None
        st = st or {}
        student_name = f"{str(st.get('first_name') or '').strip()} {str(st.get('last_name') or '').strip()}".strip()
        cls = str(st.get('class_name') or '').strip()
        created_at = str(snap.get('created_at') or '').strip()
        dt_txt = created_at[:16] if created_at else datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Add Hebrew date
        heb_date = ''
        try:
            from jewish_calendar import hebrew_date_from_gregorian_str
            greg_date = dt_txt[:10] if len(dt_txt) >= 10 else datetime.now().strftime('%Y-%m-%d')
            heb_date = hebrew_date_from_gregorian_str(greg_date)
        except Exception:
            heb_date = ''

        lines = ['חשבונית', dt_txt]
        if heb_date:
            lines.append(heb_date)
        lines.append(f"{student_name} | {cls}".strip(' |'))
        lines.append('--------------------------')
        total = 0

        for it in (data.get('items') or []):
            try:
                pid = int(it.get('product_id') or 0)
            except Exception:
                pid = 0
            try:
                qty = int(it.get('qty') or 0)
            except Exception:
                qty = 0
            if not pid or qty <= 0:
                continue
            label = f"מוצר לא ידוע (#{pid})".strip()
            try:
                prows = self.db.get_all_products(active_only=False) or []
                for p in prows:
                    if int(p.get('id') or 0) == pid:
                        label = str(p.get('display_name') or '').strip() or str(p.get('name') or '').strip() or label
                        break
            except Exception:
                pass
            try:
                price_each = int(it.get('price_points') or 0)
            except Exception:
                price_each = 0
            if price_each <= 0:
                try:
                    conn = self.db.get_connection()
                    cur = conn.cursor()
                    cur.execute('SELECT points_each FROM purchases_log WHERE id = ?', (int(it.get('purchase_id') or 0),))
                    row = cur.fetchone()
                    conn.close()
                    if row:
                        price_each = int(row['points_each'] or 0)
                except Exception:
                    pass
            lines.append(f"{qty}x {label} - {price_each} נק")
            total += int(price_each) * int(qty)

        for sr in (data.get('scheduled_reservations') or []):
            try:
                sid = int(sr.get('service_id') or 0)
            except Exception:
                sid = 0
            sd = str(sr.get('service_date') or '').strip()
            stt = str(sr.get('slot_start_time') or '').strip()
            if sid:
                lines.append(f"אתגר {sid} {sd} {stt}".strip())

        lines.append('--------------------------')
        lines.append(f"סה\"כ: {int(total)} נק")

        font = ImageFont.load_default()
        width = 620
        pad = 22
        line_h = 28
        height = max(220, pad * 2 + line_h * (len(lines) + 2))
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        y = pad
        for ln in lines:
            try:
                draw.text((pad, y), fix_rtl_text(_strip_asterisk_annotations(str(ln))), font=font, fill='black')
            except Exception:
                try:
                    draw.text((pad, y), str(ln), font=font, fill='black')
                except Exception:
                    pass
            y += line_h

        out_dir = os.path.join(os.environ.get('PROGRAMDATA', r'C:\ProgramData'), 'SchoolPoints', 'receipts')
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception:
            out_dir = self.base_dir
        fn = f"receipt_reprint_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        out_path = os.path.join(out_dir, fn)
        try:
            img.save(out_path, 'PDF', resolution=200.0)
        except Exception:
            return ''
        return out_path

    def open_purchases_manager(self):
        if not bool(ENABLE_PURCHASES):
            messagebox.showwarning("לא זמין", "מסך קניות אינו זמין בגרסה זו")
            return
        if not (self.current_teacher and self.current_teacher.get('is_admin') == 1):
            messagebox.showwarning("אין הרשאה", "גישה רק למנהלים")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("🛒 ניהול קניות")
        dialog.configure(bg='#ecf0f1')
        dialog.geometry("1200x750")
        try:
            dialog.minsize(1150, 700)
        except Exception:
            pass
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        tk.Label(
            dialog,
            text=fix_rtl_text("ניהול קופה"),
            font=('Arial', 14, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(14, 8))

        nb = ttk.Notebook(dialog)
        nb.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 10))

        products_tab = tk.Frame(nb, bg='#ecf0f1')
        categories_tab = tk.Frame(nb, bg='#ecf0f1')
        settings_tab = tk.Frame(nb, bg='#ecf0f1')
        nb.add(products_tab, text='מוצרים')
        nb.add(categories_tab, text='קטגוריות')
        nb.add(settings_tab, text='הגדרות קופה')

        list_frame = tk.Frame(products_tab, bg='#ecf0f1')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ('type', 'active', 'name', 'display_name', 'category', 'price', 'stock', 'image')
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=16)
        tree.heading('type', text='סוג')
        tree.heading('active', text='פעיל')
        tree.heading('name', text='שם פנימי')
        tree.heading('display_name', text='שם תצוגה')
        tree.heading('category', text='קטגוריה')
        tree.heading('price', text='מחיר')
        tree.heading('stock', text='מלאי')
        tree.heading('image', text='תמונה')

        tree.column('type', width=80, anchor='center')
        tree.column('active', width=60, anchor='center')
        tree.column('name', width=200, anchor='e')
        tree.column('display_name', width=220, anchor='e')
        tree.column('category', width=140, anchor='e')
        tree.column('price', width=80, anchor='center')
        tree.column('stock', width=90, anchor='center')
        tree.column('image', width=220, anchor='e')

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=sb.set)

        def _get_selected_product_id() -> int:
            sel = tree.selection()
            if not sel:
                return 0
            tags = tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('pid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _load_rows(select_pid: int = 0):
            try:
                rows = self.db.get_all_products(active_only=False) or []
            except Exception:
                rows = []
            try:
                tree.delete(*tree.get_children())
            except Exception:
                pass

            to_select = None
            for r in rows:
                try:
                    pid = int(r.get('id', 0) or 0)
                except Exception:
                    pid = 0
                if not pid:
                    continue
                active_txt = 'כן' if int(r.get('is_active', 1) or 0) == 1 else 'לא'
                name = str(r.get('name') or '').strip()
                display_name = str(r.get('display_name') or '').strip()
                cat_name = str(r.get('category_name') or '').strip()
                price = str(int(r.get('price_points', 0) or 0))
                stock_val = r.get('stock_qty', None)
                stock = '∞' if stock_val is None else str(stock_val)
                img = str(r.get('image_path') or '').strip()
                img_display = '🖼' if img else ''

                kind = '🛍 מוצר'
                try:
                    ss = self.db.get_scheduled_service_by_product(int(pid))
                    if ss and int(ss.get('is_active', 1) or 0) == 1:
                        kind = '⏱ אתגר'
                except Exception:
                    kind = '🛍 מוצר'

                iid = tree.insert('', 'end', values=(kind, active_txt, name, display_name, cat_name, price, stock, img_display), tags=(f"pid:{pid}",))
                if select_pid and pid == int(select_pid):
                    to_select = iid
            if to_select:
                try:
                    tree.selection_set(to_select)
                    tree.see(to_select)
                except Exception:
                    pass

        def _open_product_dialog(existing: dict = None) -> dict:
            existing = existing or {}
            dlg = tk.Toplevel(dialog)
            dlg.title("עריכת מוצר" if existing.get('id') else "הוספת מוצר")
            dlg.configure(bg='#ecf0f1')
            dlg.transient(dialog)
            dlg.grab_set()

            result = {'ok': False}

            name_var = tk.StringVar(value=str(existing.get('name') or ''))
            display_name_var = tk.StringVar(value=str(existing.get('display_name') or ''))
            image_var = tk.StringVar(value=str(existing.get('image_path') or ''))
            price_var = tk.StringVar(value=str(int(existing.get('price_points', 0) or 0)))
            stock_var = tk.StringVar(value='' if existing.get('stock_qty', None) is None else str(existing.get('stock_qty')))
            active_var = tk.IntVar(value=1 if int(existing.get('is_active', 1) or 0) == 1 else 0)

            allowed_classes_var = tk.StringVar(value=str(existing.get('allowed_classes') or ''))
            min_points_var = tk.StringVar(value=str(int(existing.get('min_points_required', 0) or 0)))
            mps_var = tk.StringVar(value='' if existing.get('max_per_student', None) is None else str(existing.get('max_per_student')))
            mpc_var = tk.StringVar(value='' if existing.get('max_per_class', None) is None else str(existing.get('max_per_class')))
            po_min_var = tk.StringVar(value='' if existing.get('price_override_min_points', None) is None else str(existing.get('price_override_min_points')))
            po_price_var = tk.StringVar(value='' if existing.get('price_override_discount_pct', None) is None else str(existing.get('price_override_discount_pct')))

            # Categories
            try:
                cats = self.db.get_product_categories(active_only=False) or []
            except Exception:
                cats = []
            cats = list(cats)
            cats_names = [str(c.get('name') or '').strip() for c in cats if str(c.get('name') or '').strip()]
            cat_values = ['ללא קטגוריה'] + cats_names
            cats_by_name = {str(c.get('name') or '').strip(): int(c.get('id') or 0) for c in cats}
            cur_cat_name = str(existing.get('category_name') or '').strip()
            if not cur_cat_name:
                try:
                    cur_cat_id = existing.get('category_id', None)
                    if cur_cat_id is not None:
                        cur_cat_id = int(cur_cat_id)
                        for c in cats:
                            if int(c.get('id') or 0) == cur_cat_id:
                                cur_cat_name = str(c.get('name') or '').strip()
                                break
                except Exception:
                    cur_cat_name = ''
            cat_var = tk.StringVar(value=(cur_cat_name if cur_cat_name in cat_values else 'ללא קטגוריה'))

            body = tk.Frame(dlg, bg='#ecf0f1')
            body.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

            def _row(label: str):
                r = tk.Frame(body, bg='#ecf0f1')
                r.pack(fill=tk.X, pady=6)
                tk.Label(r, text=fix_rtl_text(label), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=14, anchor='e').pack(side=tk.RIGHT, padx=6)
                return r

            r1 = _row('שם מוצר:')
            tk.Entry(r1, textvariable=name_var, font=('Arial', 11), justify='right', width=36).pack(side=tk.RIGHT, padx=6)

            r1b = _row('שם תצוגה:')
            tk.Entry(r1b, textvariable=display_name_var, font=('Arial', 11), justify='right', width=36).pack(side=tk.RIGHT, padx=6)

            rcat = _row('קטגוריה:')
            ttk.Combobox(rcat, textvariable=cat_var, values=cat_values, state='readonly', width=34, justify='right').pack(side=tk.RIGHT, padx=6)

            rimg = _row('תמונה:')
            tk.Entry(rimg, textvariable=image_var, font=('Arial', 10), justify='right', width=42).pack(side=tk.RIGHT, padx=6)

            def _browse_img():
                try:
                    fp = filedialog.askopenfilename(
                        title='בחר תמונת מוצר',
                        filetypes=[('תמונות', '*.png;*.jpg;*.jpeg;*.bmp;*.gif'), ('הכל', '*.*')]
                    )
                except Exception:
                    fp = ''
                if fp:
                    image_var.set(fp)

            tk.Button(rimg, text='בחר…', command=_browse_img, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)

            r2 = _row('מחיר נקודות:')
            tk.Entry(r2, textvariable=price_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r2, text=fix_rtl_text('(ניתן לרשום 0 עבור זיכוי)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r3 = _row('מלאי:')
            tk.Entry(r3, textvariable=stock_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r3, text=fix_rtl_text("(ריק = ללא הגבלת מלאי)"), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r4 = _row('דגלים:')
            tk.Checkbutton(r4, text=fix_rtl_text('פעיל'), variable=active_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)

            rules_box = tk.LabelFrame(body, text=fix_rtl_text('זמינות/הגבלות (אופציונלי)'), bg='#ecf0f1', font=('Arial', 10, 'bold'))
            rules_box.pack(fill=tk.X, pady=(10, 0), anchor='e')

            def _rrow(label: str):
                r = tk.Frame(rules_box, bg='#ecf0f1')
                r.pack(fill=tk.X, pady=5, padx=10)
                tk.Label(r, text=fix_rtl_text(label), bg='#ecf0f1', font=('Arial', 9, 'bold'), width=16, anchor='e').pack(side=tk.RIGHT, padx=6)
                return r

            rcls = _rrow('כיתות מורשות:')
            allowed_classes_entry = tk.Entry(rcls, textvariable=allowed_classes_var, font=('Arial', 10), justify='right', width=28, state='readonly')
            allowed_classes_entry.pack(side=tk.RIGHT, padx=6)
            
            def open_product_classes_selector():
                all_classes = set()
                try:
                    students = self.db.get_all_students()
                    for s in students:
                        cn = (s.get('class_name') or '').strip()
                        if cn:
                            all_classes.add(cn)
                except Exception:
                    pass
                
                if not all_classes:
                    messagebox.showinfo('מידע', 'אין כיתות במערכת')
                    return
                
                selector_dialog = tk.Toplevel(dlg)
                selector_dialog.title('בחירת כיתות')
                selector_dialog.geometry('500x600')
                selector_dialog.configure(bg='#ecf0f1')
                selector_dialog.transient(dlg)
                selector_dialog.grab_set()
                selector_dialog.resizable(True, True)
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('בחר כיתות מורשות'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(12, 6))
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('(ריק = כל הכיתות)'),
                    font=('Arial', 9),
                    bg='#ecf0f1',
                    fg='#7f8c8d'
                ).pack(pady=(0, 6))
                
                cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                checkbox_vars = {}
                current_classes = [c.strip() for c in str(allowed_classes_var.get() or '').split(',') if c.strip()]
                
                for cls in sorted(all_classes):
                    var = tk.BooleanVar(value=(cls in current_classes))
                    checkbox_vars[cls] = var
                    
                    cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                    cb_row.pack(fill=tk.X, padx=20, pady=2)
                    
                    cb = tk.Checkbutton(
                        cb_row,
                        text=fix_rtl_text(cls),
                        variable=var,
                        font=('Arial', 11),
                        bg='#ecf0f1',
                        anchor='w'
                    )
                    cb.pack(side=tk.LEFT, fill=tk.X)
                
                canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                
                btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=(0, 12))
                
                def select_all():
                    for var in checkbox_vars.values():
                        var.set(True)
                
                def clear_all():
                    for var in checkbox_vars.values():
                        var.set(False)
                
                def apply_selection():
                    selected = [cls for cls, var in checkbox_vars.items() if var.get()]
                    allowed_classes_var.set(', '.join(selected) if selected else '')
                    selector_dialog.destroy()
                
                tk.Button(
                    btn_frame,
                    text='בחר הכל',
                    command=select_all,
                    font=('Arial', 10),
                    bg='#3498db',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='נקה הכל',
                    command=clear_all,
                    font=('Arial', 10),
                    bg='#e74c3c',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='אישור',
                    command=apply_selection,
                    font=('Arial', 10, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=16,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
            
            tk.Button(rcls, text='בחר...', command=open_product_classes_selector, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)
            tk.Label(rcls, text=fix_rtl_text('(ריק = כל הכיתות)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rmin = _rrow('סף נקודות:')
            tk.Entry(rmin, textvariable=min_points_var, font=('Arial', 10), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(rmin, text=fix_rtl_text('(0 = ללא סף)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rmps = _rrow('מקס׳ לתלמיד:')
            tk.Entry(rmps, textvariable=mps_var, font=('Arial', 10), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(rmps, text=fix_rtl_text('(ריק = ללא הגבלה)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rmpc = _rrow('מקס׳ לכיתה:')
            tk.Entry(rmpc, textvariable=mpc_var, font=('Arial', 10), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(rmpc, text=fix_rtl_text('(ריק = ללא הגבלה)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rpo = _rrow('מחיר מותנה:')
            tk.Label(rpo, text=fix_rtl_text('מעל'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=4)
            tk.Entry(rpo, textvariable=po_min_var, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)
            tk.Label(rpo, text=fix_rtl_text("נק׳ → הנחה %"), bg='#ecf0f1').pack(side=tk.RIGHT, padx=4)
            tk.Entry(rpo, textvariable=po_price_var, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)
            tk.Label(rpo, text=fix_rtl_text("% (ריק = לא פעיל)"), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            # וריאציות בתוך מסך הוספת מוצר
            variants_box = tk.LabelFrame(body, text=fix_rtl_text('וריאציות (אופציונלי)'), bg='#ecf0f1', font=('Arial', 10, 'bold'))
            variants_box.pack(fill=tk.X, pady=(10, 0), anchor='e')

            variants_rows_frame = tk.Frame(variants_box, bg='#ecf0f1')
            variants_rows_frame.pack(fill=tk.X, padx=10, pady=(8, 6))

            hdr = tk.Frame(variants_rows_frame, bg='#ecf0f1')
            hdr.pack(fill=tk.X, pady=(0, 4))
            tk.Label(hdr, text=fix_rtl_text('שם וריאציה'), bg='#ecf0f1', font=('Arial', 9, 'bold'), width=18, anchor='e').pack(side=tk.RIGHT, padx=4)
            tk.Label(hdr, text=fix_rtl_text('מחיר'), bg='#ecf0f1', font=('Arial', 9, 'bold'), width=8, anchor='e').pack(side=tk.RIGHT, padx=4)
            tk.Label(hdr, text=fix_rtl_text('מלאי'), bg='#ecf0f1', font=('Arial', 9, 'bold'), width=8, anchor='e').pack(side=tk.RIGHT, padx=4)

            variants_entries = []  # list of {'frame','name_var','price_var','stock_var'}

            def _add_variant_row(v: dict = None):
                v = v or {}
                rowf = tk.Frame(variants_rows_frame, bg='#ecf0f1')
                rowf.pack(fill=tk.X, pady=2)

                namev = tk.StringVar(value=str(v.get('variant_name') or ''))
                pricev = tk.StringVar(value=str(int(v.get('price_points', 0) or 0)))
                stock0 = v.get('stock_qty', None)
                stockv = tk.StringVar(value='' if stock0 is None else str(stock0))

                tk.Entry(rowf, textvariable=namev, font=('Arial', 10), justify='right', width=22).pack(side=tk.RIGHT, padx=4)
                tk.Entry(rowf, textvariable=pricev, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)
                tk.Entry(rowf, textvariable=stockv, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)

                def _remove():
                    try:
                        rowf.destroy()
                    except Exception:
                        pass
                    try:
                        variants_entries[:] = [x for x in variants_entries if x.get('frame') is not rowf]
                    except Exception:
                        pass

                tk.Button(rowf, text='✖', command=_remove, font=('Arial', 9, 'bold'), bg='#95a5a6', fg='white', padx=8, pady=2).pack(side=tk.RIGHT, padx=(0, 4))

                variants_entries.append({'frame': rowf, 'name_var': namev, 'price_var': pricev, 'stock_var': stockv})

            # טען וריאציות קיימות בעריכה
            try:
                if existing.get('id'):
                    _existing_vars = self.db.get_product_variants(int(existing.get('id') or 0), active_only=False) or []
                    for _v in _existing_vars:
                        _add_variant_row(_v)
            except Exception:
                pass

            vb = tk.Frame(variants_box, bg='#ecf0f1')
            vb.pack(fill=tk.X, padx=10, pady=(0, 8), anchor='e')
            tk.Button(vb, text='➕ הוסף וריאציה', command=lambda: _add_variant_row({}), font=('Arial', 9, 'bold'), bg='#27ae60', fg='white', padx=12, pady=4).pack(side=tk.RIGHT)

            btns = tk.Frame(body, bg='#ecf0f1')
            btns.pack(pady=(10, 0), anchor='e')

            def _save():
                name = name_var.get().strip()
                if not name:
                    messagebox.showerror('שגיאה', 'יש להזין שם מוצר')
                    return
                try:
                    if not str(display_name_var.get() or '').strip():
                        display_name_var.set(str(name))
                except Exception:
                    pass
                try:
                    price = int(str(price_var.get() or '0').strip())
                except Exception:
                    price = 0
                sraw = str(stock_var.get() or '').strip()
                if not sraw:
                    stock = None
                else:
                    try:
                        stock = int(sraw)
                    except Exception:
                        messagebox.showerror('שגיאה', 'מלאי חייב להיות מספר או ריק')
                        return

                variants_out = []
                try:
                    for x in list(variants_entries):
                        vn = str((x.get('name_var').get() if x.get('name_var') else '') or '').strip()
                        if not vn:
                            continue
                        try:
                            vp = int(str((x.get('price_var').get() if x.get('price_var') else '0') or '0').strip())
                        except Exception:
                            messagebox.showerror('שגיאה', 'מחיר וריאציה חייב להיות מספר')
                            return
                        sraw2 = str((x.get('stock_var').get() if x.get('stock_var') else '') or '').strip()
                        if not sraw2:
                            vstock = None
                        else:
                            try:
                                vstock = int(sraw2)
                            except Exception:
                                messagebox.showerror('שגיאה', 'מלאי וריאציה חייב להיות מספר או ריק')
                                return
                        variants_out.append({'variant_name': vn, 'price_points': int(vp), 'stock_qty': vstock})
                except Exception:
                    variants_out = []

                result.update({
                    'ok': True,
                    'name': name,
                    'display_name': display_name_var.get().strip(),
                    'image_path': image_var.get().strip(),
                    'category_id': (cats_by_name.get(str(cat_var.get() or '').strip()) if str(cat_var.get() or '').strip() != 'ללא קטגוריה' else None),
                    'price_points': int(price),
                    'stock_qty': stock,
                    'deduct_points': 1,
                    'is_active': 1 if int(active_var.get() or 0) == 1 else 0,
                    'allowed_classes': str(allowed_classes_var.get() or '').strip(),
                    'min_points_required': (int(str(min_points_var.get() or '0').strip() or '0') if str(min_points_var.get() or '').strip() else 0),
                    'max_per_student': (int(str(mps_var.get()).strip()) if str(mps_var.get() or '').strip() else None),
                    'max_per_class': (int(str(mpc_var.get()).strip()) if str(mpc_var.get() or '').strip() else None),
                    'price_override_min_points': (int(str(po_min_var.get()).strip()) if str(po_min_var.get() or '').strip() else None),
                    'price_override_points': None,
                    'price_override_discount_pct': None,
                    'variants': variants_out,
                })

                try:
                    _pct_raw = str(po_price_var.get() or '').strip()
                    if _pct_raw:
                        _pct = int(_pct_raw)
                        if _pct < 0:
                            _pct = 0
                        if _pct > 100:
                            _pct = 100
                        result['price_override_discount_pct'] = int(_pct)
                except Exception:
                    pass
                dlg.destroy()

            tk.Button(btns, text='💾 שמור', command=_save, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text='✖ ביטול', command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)

            dlg.wait_window()
            return result

        def _get_row_by_pid(pid: int) -> dict:
            try:
                rows = self.db.get_all_products(active_only=False) or []
            except Exception:
                rows = []
            for r in rows:
                try:
                    if int(r.get('id', 0) or 0) == int(pid or 0):
                        return r
                except Exception:
                    continue
            return {}

        def add_product():
            data = _open_product_dialog(existing=None)
            if not data.get('ok'):
                return 0
            try:
                pid = self.db.add_product(
                    name=data['name'],
                    display_name=data.get('display_name', ''),
                    image_path=data.get('image_path', ''),
                    category_id=data.get('category_id', None),
                    price_points=data['price_points'],
                    stock_qty=data['stock_qty'],
                    deduct_points=data['deduct_points'],
                    allowed_classes=data.get('allowed_classes', ''),
                    min_points_required=data.get('min_points_required', 0),
                    max_per_student=data.get('max_per_student', None),
                    max_per_class=data.get('max_per_class', None),
                    price_override_min_points=data.get('price_override_min_points', None),
                    price_override_points=None,
                    price_override_discount_pct=data.get('price_override_discount_pct', None),
                    is_active=data['is_active'],
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return 0
            try:
                _apply_variants(pid, data.get('variants') or [])
            except Exception:
                pass
            _load_rows(select_pid=pid)
            return int(pid or 0)

        def _apply_variants(pid: int, variants: list):
            try:
                pid = int(pid or 0)
            except Exception:
                pid = 0
            if not pid:
                return
            # delete all existing
            try:
                old = self.db.get_product_variants(pid, active_only=False) or []
            except Exception:
                old = []
            for r in old:
                try:
                    vid = int(r.get('id') or 0)
                except Exception:
                    vid = 0
                if vid:
                    try:
                        self.db.delete_product_variant(int(vid))
                    except Exception:
                        pass
            # insert new
            for i, v in enumerate(list(variants or [])):
                vn = str(v.get('variant_name') or '').strip()
                if not vn:
                    continue
                try:
                    vp = int(v.get('price_points', 0) or 0)
                except Exception:
                    vp = 0
                vstock = v.get('stock_qty', None)
                if vstock is not None:
                    try:
                        vstock = int(vstock)
                    except Exception:
                        vstock = None
                try:
                    self.db.add_product_variant(
                        product_id=int(pid),
                        variant_name=vn,
                        display_name='',
                        price_points=int(vp),
                        stock_qty=vstock,
                        deduct_points=1,
                        is_active=1,
                        sort_order=int(i),
                    )
                except Exception:
                    pass

        def _open_scheduled_service_dialog(product_id: int):
            pid = int(product_id or 0)
            dlg = tk.Toplevel(dialog)
            dlg.title('⏱ הגדרת אתגר מתוזמן')
            dlg.configure(bg='#ecf0f1')
            dlg.transient(dialog)
            dlg.grab_set()

            prod = None
            if pid:
                try:
                    prod = self.db.get_product_by_id(pid)
                except Exception:
                    prod = None
            title_name = ''
            try:
                title_name = str((prod or {}).get('display_name') or '').strip() or str((prod or {}).get('name') or '').strip()
            except Exception:
                title_name = ''
            if not title_name:
                title_name = "אתגר חדש" if not pid else f"מוצר לא ידוע (#{pid})"

            tk.Label(dlg, text=fix_rtl_text(f"אתגר מתוזמן למוצר: {title_name}"), font=('Arial', 12, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(12, 8))

            existing = None
            try:
                if pid:
                    existing = self.db.get_scheduled_service_by_product(pid)
            except Exception:
                existing = None

            enabled_var = tk.IntVar(value=1 if (existing is None or int((existing or {}).get('is_active', 1) or 0) == 1) else 0)
            dur_var = tk.StringVar(value=str(int((existing or {}).get('duration_minutes', 10) or 10)))
            cap_var = tk.StringVar(value=str(int((existing or {}).get('capacity_per_slot', 1) or 1)))
            start_var = tk.StringVar(value=str((existing or {}).get('start_time') or '16:00'))
            end_var = tk.StringVar(value=str((existing or {}).get('end_time') or '19:00'))
            auto_var = tk.IntVar(value=1 if int((existing or {}).get('allow_auto_time', 1) or 0) == 1 else 0)
            mps_var = tk.StringVar(value='' if (existing is None or (existing.get('max_per_student', None) is None)) else str(existing.get('max_per_student')))
            mpc_var = tk.StringVar(value='' if (existing is None or (existing.get('max_per_class', None) is None)) else str(existing.get('max_per_class')))

            qp_mode_var = tk.StringVar(value=str((existing or {}).get('queue_priority_mode') or 'none'))
            qp_custom_var = tk.StringVar(value=str((existing or {}).get('queue_priority_custom') or ''))
            allowed_classes_var = tk.StringVar(value=str((existing or {}).get('allowed_classes') or ''))
            min_points_var = tk.StringVar(value=str(int((existing or {}).get('min_points_required', 0) or 0)))

            dates_existing = []
            try:
                if existing:
                    dates_existing = self.db.get_scheduled_service_dates(int(existing.get('id') or 0), active_only=True) or []
            except Exception:
                dates_existing = []
            dates_var = tk.StringVar(value='\n'.join([d for d in dates_existing if str(d or '').strip()]))

            body = tk.Frame(dlg, bg='#ecf0f1')
            body.pack(fill=tk.BOTH, expand=True, padx=16, pady=10)

            def _row(label: str):
                r = tk.Frame(body, bg='#ecf0f1')
                r.pack(fill=tk.X, pady=5)
                tk.Label(r, text=fix_rtl_text(label), bg='#ecf0f1', font=('Arial', 10, 'bold'), width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
                return r

            r0 = _row('פעיל:')
            tk.Checkbutton(r0, text=fix_rtl_text('אתגר מתוזמן פעיל'), variable=enabled_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)

            # Product fields (for both new and existing challenges)
            new_prod_name_var = tk.StringVar(value='')
            new_prod_display_var = tk.StringVar(value='')
            new_prod_price_var = tk.StringVar(value='0')
            new_prod_image_var = tk.StringVar(value='')
            new_prod_category_var = tk.StringVar(value='ללא קטגוריה')

            display_state = {'touched': False}

            def _mark_display_touched(*_a):
                display_state['touched'] = True

            def _sync_display_from_name(*_a):
                if display_state.get('touched'):
                    return
                try:
                    nm = str(new_prod_name_var.get() or '').strip()
                except Exception:
                    nm = ''
                if not nm:
                    return
                try:
                    cur = str(new_prod_display_var.get() or '').strip()
                except Exception:
                    cur = ''
                if not cur or cur == nm:
                    try:
                        new_prod_display_var.set(nm)
                    except Exception:
                        pass

            try:
                new_prod_display_var.trace_add('write', _mark_display_touched)
            except Exception:
                pass
            try:
                new_prod_name_var.trace_add('write', _sync_display_from_name)
            except Exception:
                pass
            
            # Conditional pricing fields
            po_min_var = tk.StringVar(value='')
            po_price_var = tk.StringVar(value='')
            
            # Initialize conditional pricing for existing product
            if pid and prod:
                try:
                    po_min = prod.get('price_override_min_points', None)
                    if po_min is not None:
                        po_min_var.set(str(po_min))
                except Exception:
                    pass
                try:
                    po_pct = prod.get('price_override_discount_pct', None)
                    if po_pct is not None:
                        po_price_var.set(str(po_pct))
                except Exception:
                    pass

            # Load categories for challenge (both new and edit)
            cats = []
            cats_by_name = {}
            try:
                cats = self.db.get_product_categories(active_only=False) or []
            except Exception:
                cats = []
            cats = list(cats)
            cats_names = [str(c.get('name') or '').strip() for c in cats if str(c.get('name') or '').strip()]
            cat_values = ['ללא קטגוריה'] + cats_names
            cats_by_name = {str(c.get('name') or '').strip(): int(c.get('id') or 0) for c in cats}
            
            # Initialize category for existing product
            if pid and prod:
                try:
                    cat_id = int(prod.get('category_id') or 0)
                    if cat_id:
                        for c in cats:
                            if int(c.get('id') or 0) == cat_id:
                                new_prod_category_var.set(str(c.get('name') or '').strip())
                                break
                    else:
                        new_prod_category_var.set('ללא קטגוריה')
                except Exception:
                    new_prod_category_var.set('ללא קטגוריה')
            else:
                new_prod_category_var.set('ללא קטגוריה')
            
            # Initialize image for existing product
            if pid and prod:
                try:
                    img_path = str(prod.get('image_path') or '').strip()
                    if img_path:
                        new_prod_image_var.set(img_path)
                except Exception:
                    pass

            # Initialize price for existing product
            if pid and prod:
                try:
                    new_prod_price_var.set(str(int(prod.get('price_points', 0) or 0)))
                except Exception:
                    pass

            # Initialize name/display for existing product
            if pid and prod:
                try:
                    new_prod_name_var.set(str(prod.get('name') or '').strip())
                except Exception:
                    pass
                try:
                    new_prod_display_var.set(str(prod.get('display_name') or '').strip())
                except Exception:
                    pass

            # Default display name to name (if empty)
            try:
                if not str(new_prod_display_var.get() or '').strip():
                    new_prod_display_var.set(str(new_prod_name_var.get() or '').strip())
            except Exception:
                pass

            if not pid:
                rp0 = _row('פרטי מוצר:')
                tk.Label(rp0, text=fix_rtl_text('שם:'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
                tk.Entry(rp0, textvariable=new_prod_name_var, font=('Arial', 11), justify='right', width=18).pack(side=tk.RIGHT, padx=6)
                tk.Label(rp0, text=fix_rtl_text('שם תצוגה:'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
                tk.Entry(rp0, textvariable=new_prod_display_var, font=('Arial', 11), justify='right', width=18).pack(side=tk.RIGHT, padx=6)

                rp1 = _row('מחיר נקודות:')
                tk.Entry(rp1, textvariable=new_prod_price_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)

                rpcat = _row('קטגוריה:')
                ttk.Combobox(rpcat, textvariable=new_prod_category_var, values=cat_values, state='readonly', width=34, justify='right').pack(side=tk.RIGHT, padx=6)

                rpimg = _row('תמונה:')
                tk.Entry(rpimg, textvariable=new_prod_image_var, font=('Arial', 10), justify='right', width=32).pack(side=tk.RIGHT, padx=6)

                def _browse_challenge_img():
                    try:
                        fp = filedialog.askopenfilename(
                            title='בחר תמונת אתגר',
                            filetypes=[('תמונות', '*.png;*.jpg;*.jpeg;*.bmp;*.gif'), ('הכל', '*.*')]
                        )
                    except Exception:
                        fp = ''
                    if fp:
                        new_prod_image_var.set(fp)

                tk.Button(rpimg, text='בחר…', command=_browse_challenge_img, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)
            else:
                # For existing challenges - show editable name/display + category + image
                rpname_edit = _row('שם:')
                tk.Entry(rpname_edit, textvariable=new_prod_name_var, font=('Arial', 11), justify='right', width=34).pack(side=tk.RIGHT, padx=6)

                rpdisp_edit = _row('שם תצוגה:')
                tk.Entry(rpdisp_edit, textvariable=new_prod_display_var, font=('Arial', 11), justify='right', width=34).pack(side=tk.RIGHT, padx=6)

                rpprice_edit = _row('מחיר נקודות:')
                tk.Entry(rpprice_edit, textvariable=new_prod_price_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)

                rpcat_edit = _row('קטגוריה:')
                ttk.Combobox(rpcat_edit, textvariable=new_prod_category_var, values=cat_values, state='readonly', width=34, justify='right').pack(side=tk.RIGHT, padx=6)

                rpimg_edit = _row('תמונה:')
                tk.Entry(rpimg_edit, textvariable=new_prod_image_var, font=('Arial', 10), justify='right', width=32).pack(side=tk.RIGHT, padx=6)

                def _browse_challenge_img_edit():
                    try:
                        fp = filedialog.askopenfilename(
                            title='בחר תמונת אתגר',
                            filetypes=[('תמונות', '*.png;*.jpg;*.jpeg;*.bmp;*.gif'), ('הכל', '*.*')]
                        )
                    except Exception:
                        fp = ''
                    if fp:
                        new_prod_image_var.set(fp)

                tk.Button(rpimg_edit, text='בחר…', command=_browse_challenge_img_edit, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)

            r1 = _row('משך (דקות):')
            tk.Entry(r1, textvariable=dur_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)

            r2 = _row('קיבולת לסלוט:')
            tk.Entry(r2, textvariable=cap_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)

            r3 = _row('שעת התחלה:')
            tk.Entry(r3, textvariable=start_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r3, text=fix_rtl_text('(HH:MM)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r4 = _row('שעת סיום:')
            tk.Entry(r4, textvariable=end_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r4, text=fix_rtl_text('(HH:MM)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r5 = _row('שיבוץ אוטומטי:')
            tk.Checkbutton(r5, text=fix_rtl_text('ברירת מחדל לשיבוץ אוטומטי'), variable=auto_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)

            r6 = _row('מקסימום לתלמיד:')
            tk.Entry(r6, textvariable=mps_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r6, text=fix_rtl_text('(ריק = ללא הגבלה)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r7 = _row('מקסימום לכיתה:')
            tk.Entry(r7, textvariable=mpc_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(r7, text=fix_rtl_text('(ריק = ללא הגבלה)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r8 = _row('קדימות בתור:')
            qp_map = {
                'ללא': 'none',
                'כיתה בסדר עולה': 'class_asc',
                'כיתה בסדר יורד': 'class_desc',
                'מותאם אישית': 'custom',
            }
            qp_rev = {v: k for k, v in qp_map.items()}
            try:
                qp_mode_var.set(qp_rev.get(str(qp_mode_var.get() or '').strip(), 'ללא'))
            except Exception:
                qp_mode_var.set('ללא')
            ttk.Combobox(
                r8,
                textvariable=qp_mode_var,
                values=list(qp_map.keys()),
                state='readonly',
                width=18,
                justify='right'
            ).pack(side=tk.RIGHT, padx=6)
            tk.Label(r8, text=fix_rtl_text('(ברירת מחדל: ללא)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            r9 = _row('רשימת כיתות:')
            qp_custom_entry = tk.Entry(r9, textvariable=qp_custom_var, font=('Arial', 11), justify='right', width=20, state='readonly')
            qp_custom_entry.pack(side=tk.RIGHT, padx=6)
            
            def open_qp_classes_selector():
                all_classes = set()
                try:
                    students = self.db.get_all_students()
                    for s in students:
                        cn = (s.get('class_name') or '').strip()
                        if cn:
                            all_classes.add(cn)
                except Exception:
                    pass
                
                if not all_classes:
                    messagebox.showinfo('מידע', 'אין כיתות במערכת')
                    return
                
                selector_dialog = tk.Toplevel(dlg)
                selector_dialog.title('בחירת כיתות')
                selector_dialog.geometry('500x600')
                selector_dialog.configure(bg='#ecf0f1')
                selector_dialog.transient(dlg)
                selector_dialog.grab_set()
                selector_dialog.resizable(True, True)
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('בחר כיתות לקדימות'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(12, 6))
                
                cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                checkbox_vars = {}
                current_classes = [c.strip() for c in str(qp_custom_var.get() or '').split(',') if c.strip()]
                
                for cls in sorted(all_classes):
                    var = tk.BooleanVar(value=(cls in current_classes))
                    checkbox_vars[cls] = var
                    
                    cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                    cb_row.pack(fill=tk.X, padx=20, pady=2)
                    
                    cb = tk.Checkbutton(
                        cb_row,
                        text=fix_rtl_text(cls),
                        variable=var,
                        font=('Arial', 11),
                        bg='#ecf0f1',
                        anchor='w'
                    )
                    cb.pack(side=tk.LEFT, fill=tk.X)
                
                canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                
                btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=(0, 12))
                
                def select_all():
                    for var in checkbox_vars.values():
                        var.set(True)
                
                def clear_all():
                    for var in checkbox_vars.values():
                        var.set(False)
                
                def apply_selection():
                    selected = [cls for cls, var in checkbox_vars.items() if var.get()]
                    qp_custom_var.set(', '.join(selected) if selected else '')
                    selector_dialog.destroy()
                
                tk.Button(
                    btn_frame,
                    text='בחר הכל',
                    command=select_all,
                    font=('Arial', 10),
                    bg='#3498db',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='נקה הכל',
                    command=clear_all,
                    font=('Arial', 10),
                    bg='#e74c3c',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='אישור',
                    command=apply_selection,
                    font=('Arial', 10, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=16,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
            
            tk.Button(r9, text='בחר...', command=open_qp_classes_selector, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)

            # Availability criteria section
            avail_box = tk.LabelFrame(body, text=fix_rtl_text('קריטריוני זמינות (אופציונלי)'), bg='#ecf0f1', font=('Arial', 10, 'bold'))
            avail_box.pack(fill=tk.X, pady=(10, 0), anchor='e')

            def _arow(label: str):
                r = tk.Frame(avail_box, bg='#ecf0f1')
                r.pack(fill=tk.X, pady=5, padx=10)
                tk.Label(r, text=fix_rtl_text(label), bg='#ecf0f1', font=('Arial', 9, 'bold'), width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
                return r

            racls = _arow('כיתות מורשות:')
            allowed_classes_entry = tk.Entry(racls, textvariable=allowed_classes_var, font=('Arial', 10), justify='right', width=24, state='readonly')
            allowed_classes_entry.pack(side=tk.RIGHT, padx=6)
            
            def open_allowed_classes_selector():
                all_classes = set()
                try:
                    students = self.db.get_all_students()
                    for s in students:
                        cn = (s.get('class_name') or '').strip()
                        if cn:
                            all_classes.add(cn)
                except Exception:
                    pass
                
                if not all_classes:
                    messagebox.showinfo('מידע', 'אין כיתות במערכת')
                    return
                
                selector_dialog = tk.Toplevel(dlg)
                selector_dialog.title('בחירת כיתות מורשות')
                selector_dialog.geometry('400x500')
                selector_dialog.configure(bg='#ecf0f1')
                selector_dialog.transient(dlg)
                selector_dialog.grab_set()
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('בחר כיתות מורשות'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=(12, 6))
                
                tk.Label(
                    selector_dialog,
                    text=fix_rtl_text('(ריק = כל הכיתות)'),
                    font=('Arial', 9),
                    bg='#ecf0f1',
                    fg='#7f8c8d'
                ).pack(pady=(0, 6))
                
                cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                checkbox_vars = {}
                current_classes = [c.strip() for c in str(allowed_classes_var.get() or '').split(',') if c.strip()]
                
                for cls in sorted(all_classes):
                    var = tk.BooleanVar(value=(cls in current_classes))
                    checkbox_vars[cls] = var
                    
                    cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                    cb_row.pack(fill=tk.X, padx=20, pady=2)
                    
                    cb = tk.Checkbutton(
                        cb_row,
                        text=fix_rtl_text(cls),
                        variable=var,
                        font=('Arial', 11),
                        bg='#ecf0f1',
                        anchor='w'
                    )
                    cb.pack(side=tk.LEFT, fill=tk.X)
                
                canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                
                btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=(0, 12))
                
                def select_all():
                    for var in checkbox_vars.values():
                        var.set(True)
                
                def clear_all():
                    for var in checkbox_vars.values():
                        var.set(False)
                
                def apply_selection():
                    selected = [cls for cls, var in checkbox_vars.items() if var.get()]
                    allowed_classes_var.set(', '.join(selected) if selected else '')
                    selector_dialog.destroy()
                
                tk.Button(
                    btn_frame,
                    text='בחר הכל',
                    command=select_all,
                    font=('Arial', 10),
                    bg='#3498db',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='נקה הכל',
                    command=clear_all,
                    font=('Arial', 10),
                    bg='#e74c3c',
                    fg='white',
                    padx=12,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
                
                tk.Button(
                    btn_frame,
                    text='אישור',
                    command=apply_selection,
                    font=('Arial', 10, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=16,
                    pady=4
                ).pack(side=tk.RIGHT, padx=4)
            
            tk.Button(racls, text='בחר...', command=open_allowed_classes_selector, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)

            ramin = _arow('סף נקודות:')
            tk.Entry(ramin, textvariable=min_points_var, font=('Arial', 10), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
            tk.Label(ramin, text=fix_rtl_text('(0 = ללא סף)'), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rapo = _arow('מחיר מותנה:')
            tk.Label(rapo, text=fix_rtl_text('מעל'), bg='#ecf0f1', font=('Arial', 9)).pack(side=tk.RIGHT, padx=4)
            tk.Entry(rapo, textvariable=po_min_var, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)
            tk.Label(rapo, text=fix_rtl_text("נק׳ → הנחה %"), bg='#ecf0f1', font=('Arial', 9)).pack(side=tk.RIGHT, padx=4)
            tk.Entry(rapo, textvariable=po_price_var, font=('Arial', 10), justify='right', width=8).pack(side=tk.RIGHT, padx=4)
            tk.Label(rapo, text=fix_rtl_text("% (ריק = לא פעיל)"), bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).pack(side=tk.RIGHT, padx=6)

            rd = tk.Frame(body, bg='#ecf0f1')
            rd.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
            tk.Label(rd, text=fix_rtl_text('תאריכים:'), bg='#ecf0f1', font=('Arial', 10, 'bold'), anchor='e').pack(anchor='e')

            date_mode_var = tk.StringVar(value='single')
            start_date_var = tk.StringVar(value='')
            end_date_var = tk.StringVar(value='')

            # init from existing list
            _dates_init = [str(d or '').strip() for d in (dates_existing or []) if str(d or '').strip()]
            _today = date.today()
            _tomorrow = _today + timedelta(days=1)
            if len(_dates_init) == 1:
                date_mode_var.set('single')
                start_date_var.set(_dates_init[0])
                end_date_var.set(_dates_init[0])
            elif len(_dates_init) >= 2:
                date_mode_var.set('range')
                try:
                    start_date_var.set(min(_dates_init))
                    end_date_var.set(max(_dates_init))
                except Exception:
                    start_date_var.set(_dates_init[0])
                    end_date_var.set(_dates_init[-1])
            else:
                date_mode_var.set('single')
                start_date_var.set(_today.isoformat())
                end_date_var.set(_today.isoformat())

            dates_frame = tk.Frame(rd, bg='#ecf0f1')
            dates_frame.pack(fill=tk.X, expand=False, pady=6)

            rm = tk.Frame(dates_frame, bg='#ecf0f1')
            rm.pack(fill=tk.X, pady=(0, 6), anchor='e')
            tk.Radiobutton(rm, text=fix_rtl_text('תאריך בודד'), variable=date_mode_var, value='single', bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)
            tk.Radiobutton(rm, text=fix_rtl_text('טווח תאריכים'), variable=date_mode_var, value='range', bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)

            def _open_date_picker_for(var: tk.StringVar):
                # Reuse simple day-picker (DD/MM/YYYY) but store in YYYY-MM-DD
                picker = tk.Toplevel(dlg)
                picker.title("בחר תאריך")
                picker.transient(dlg)
                picker.grab_set()
                picker.resizable(False, False)

                pf = tk.Frame(picker, padx=10, pady=10)
                pf.pack(fill=tk.BOTH, expand=True)
                tk.Label(pf, text="בחר תאריך", font=('Arial', 10, 'bold')).pack(pady=(0, 6))

                try:
                    cur = str(var.get() or '').strip()
                except Exception:
                    cur = ''
                d0 = _today.day
                m0 = _today.month
                y0 = _today.year
                try:
                    if cur and '-' in cur:
                        y0, m0, d0 = [int(x) for x in cur.split('-', 2)]
                except Exception:
                    d0, m0, y0 = _today.day, _today.month, _today.year

                rowp = tk.Frame(pf)
                rowp.pack(pady=(0, 6))
                yv = tk.StringVar(value=str(y0))
                mv = tk.StringVar(value=str(m0))
                dv = tk.StringVar(value=str(d0))
                tk.Entry(rowp, textvariable=dv, width=4, justify='center').pack(side=tk.LEFT)
                tk.Label(rowp, text='/').pack(side=tk.LEFT)
                tk.Entry(rowp, textvariable=mv, width=4, justify='center').pack(side=tk.LEFT)
                tk.Label(rowp, text='/').pack(side=tk.LEFT)
                tk.Entry(rowp, textvariable=yv, width=6, justify='center').pack(side=tk.LEFT)

                def _ok_date():
                    try:
                        dd = int(dv.get())
                        mm = int(mv.get())
                        yy = int(yv.get())
                        dsel = date(yy, mm, dd)
                    except Exception:
                        messagebox.showwarning('שגיאה', 'תאריך לא תקין')
                        return
                    var.set(dsel.isoformat())
                    picker.destroy()

                tk.Button(pf, text='אישור', command=_ok_date, width=10).pack(side=tk.LEFT, padx=5)
                tk.Button(pf, text='ביטול', command=picker.destroy, width=10).pack(side=tk.RIGHT, padx=5)
                picker.wait_window()

            rdates = tk.Frame(dates_frame, bg='#ecf0f1')
            rdates.pack(fill=tk.X, anchor='e')
            tk.Label(rdates, text=fix_rtl_text('מתאריך:'), bg='#ecf0f1', width=10, anchor='e').pack(side=tk.RIGHT, padx=4)
            start_ent = tk.Entry(rdates, textvariable=start_date_var, font=('Arial', 11), justify='right', width=12)
            start_ent.pack(side=tk.RIGHT, padx=4)
            tk.Label(rdates, text=fix_rtl_text('עד תאריך:'), bg='#ecf0f1', width=10, anchor='e').pack(side=tk.RIGHT, padx=4)
            end_ent = tk.Entry(rdates, textvariable=end_date_var, font=('Arial', 11), justify='right', width=12)
            end_ent.pack(side=tk.RIGHT, padx=4)

            try:
                start_ent.bind('<Button-1>', lambda _e: _open_date_picker_for(start_date_var))
                end_ent.bind('<Button-1>', lambda _e: _open_date_picker_for(end_date_var))
            except Exception:
                pass

            def _build_dates_from_ui() -> list:
                mode = str(date_mode_var.get() or 'single')
                s0 = str(start_date_var.get() or '').strip()
                e0 = str(end_date_var.get() or '').strip()
                if not s0:
                    s0 = _today.isoformat()
                if not e0:
                    e0 = s0
                if mode == 'single':
                    return [s0]
                # range
                try:
                    sd = date.fromisoformat(s0)
                    ed = date.fromisoformat(e0)
                except Exception:
                    return [s0]
                if ed < sd:
                    sd, ed = ed, sd
                out = []
                dcur = sd
                while dcur <= ed:
                    out.append(dcur.isoformat())
                    dcur = dcur + timedelta(days=1)
                return out

            def _on_mode_change(*_a):
                m = str(date_mode_var.get() or 'single')
                if m == 'single':
                    try:
                        end_date_var.set(start_date_var.get())
                    except Exception:
                        pass

            def _fill_defaults_for_range():
                try:
                    start_date_var.set(_today.isoformat())
                    end_date_var.set((_today + timedelta(days=1)).isoformat())
                except Exception:
                    pass

            try:
                date_mode_var.trace_add('write', lambda *_a: (_fill_defaults_for_range() if date_mode_var.get() == 'range' and not end_date_var.get().strip() else None, _on_mode_change()))
            except Exception:
                pass
            try:
                start_date_var.trace_add('write', lambda *_a: _on_mode_change())
            except Exception:
                pass

            btns2 = tk.Frame(dlg, bg='#ecf0f1')
            btns2.pack(pady=(6, 12))

            def _save():
                nonlocal pid
                
                # Update category, image, and conditional pricing for existing product
                if pid:
                    try:
                        # update_product דורש שדות חובה (למשל name), לכן נטען את המוצר הקיים
                        try:
                            cur_prod = self.db.get_product_by_id(int(pid)) or {}
                        except Exception:
                            cur_prod = {}

                        img_path = str(new_prod_image_var.get() or '').strip()
                        cat_id = None
                        try:
                            cat_name = str(new_prod_category_var.get() or '').strip()
                            if cat_name and cat_name != 'ללא קטגוריה':
                                cat_id = cats_by_name.get(cat_name, None)
                        except Exception:
                            cat_id = None
                        
                        # Parse conditional pricing (percent discount)
                        po_min = None
                        po_pct = None
                        try:
                            po_min_str = str(po_min_var.get() or '').strip()
                            if po_min_str:
                                po_min = int(po_min_str)
                        except Exception:
                            pass
                        try:
                            po_pct_str = str(po_price_var.get() or '').strip()
                            if po_pct_str:
                                po_pct = int(po_pct_str)
                        except Exception:
                            pass
                        if po_pct is not None:
                            if po_pct < 0:
                                po_pct = 0
                            if po_pct > 100:
                                po_pct = 100
                        
                        # Update product with edited fields
                        cur_name = str(new_prod_name_var.get() or '').strip() or str(cur_prod.get('name') or '').strip()
                        if not cur_name:
                            cur_name = f"product_{int(pid)}"
                        cur_display_name = str(new_prod_display_var.get() or '').strip()
                        if not cur_display_name:
                            cur_display_name = str(cur_prod.get('display_name') or '').strip()
                        try:
                            cur_price_points = int(str(new_prod_price_var.get() or '').strip() or 0)
                        except Exception:
                            cur_price_points = 0
                        if cur_price_points < 0:
                            cur_price_points = 0
                        cur_stock = cur_prod.get('stock_qty', None)
                        try:
                            cur_deduct = int(cur_prod.get('deduct_points', 1) or 1)
                        except Exception:
                            cur_deduct = 1
                        try:
                            cur_active = int(enabled_var.get() or 0)
                        except Exception:
                            cur_active = 0
                        cur_allowed = str(cur_prod.get('allowed_classes') or '').strip()
                        try:
                            cur_min_pts = int(cur_prod.get('min_points_required', 0) or 0)
                        except Exception:
                            cur_min_pts = 0
                        cur_mps = cur_prod.get('max_per_student', None)
                        cur_mpc = cur_prod.get('max_per_class', None)

                        ok_prod = self.db.update_product(
                            product_id=int(pid),
                            name=cur_name,
                            display_name=cur_display_name,
                            image_path=img_path,
                            category_id=cat_id,
                            price_points=int(cur_price_points),
                            stock_qty=cur_stock,
                            deduct_points=int(cur_deduct),
                            allowed_classes=cur_allowed,
                            min_points_required=int(cur_min_pts),
                            max_per_student=cur_mps,
                            max_per_class=cur_mpc,
                            price_override_min_points=po_min,
                            price_override_points=None,
                            price_override_discount_pct=po_pct,
                            is_active=int(cur_active),
                        )
                        if not ok_prod:
                            messagebox.showwarning('אזהרה', 'לא נשמרו שינויים במוצר (קטגוריה/תמונה). נסה שוב.', parent=dlg)
                    except Exception as e:
                        messagebox.showerror('שגיאה', f"שגיאה בעדכון קטגוריה/תמונה: {e}", parent=dlg)
                
                if not pid:
                    name = str(new_prod_name_var.get() or '').strip()
                    disp = str(new_prod_display_var.get() or '').strip()
                    if not disp:
                        disp = name
                    price_raw = str(new_prod_price_var.get() or '0').strip()
                    try:
                        price_pts = int(price_raw)
                    except Exception:
                        messagebox.showerror('שגיאה', 'מחיר נקודות חייב להיות מספר')
                        return
                    if price_pts < 0:
                        messagebox.showerror('שגיאה', 'מחיר נקודות חייב להיות 0 או יותר')
                        return
                    if not name:
                        name = disp
                    if not name:
                        messagebox.showerror('שגיאה', 'חובה להזין שם לאתגר')
                        return
                    img_path = str(new_prod_image_var.get() or '').strip()
                    cat_id = None
                    try:
                        cat_name = str(new_prod_category_var.get() or '').strip()
                        if cat_name and cat_name != 'ללא קטגוריה':
                            cat_id = cats_by_name.get(cat_name, None)
                    except Exception:
                        cat_id = None
                    
                    # Parse conditional pricing (percent discount)
                    po_min = None
                    po_pct = None
                    try:
                        po_min_str = str(po_min_var.get() or '').strip()
                        if po_min_str:
                            po_min = int(po_min_str)
                    except Exception:
                        pass
                    try:
                        po_pct_str = str(po_price_var.get() or '').strip()
                        if po_pct_str:
                            po_pct = int(po_pct_str)
                    except Exception:
                        pass
                    if po_pct is not None:
                        if po_pct < 0:
                            po_pct = 0
                        if po_pct > 100:
                            po_pct = 100
                    
                    try:
                        pid = self.db.add_product(
                            name=name,
                            display_name=disp,
                            image_path=img_path,
                            category_id=cat_id,
                            price_points=int(price_pts),
                            stock_qty=None,
                            deduct_points=1,
                            is_active=1 if int(enabled_var.get() or 0) == 1 else 0,
                            price_override_min_points=po_min,
                            price_override_points=None,
                            price_override_discount_pct=po_pct,
                        )
                    except Exception as e:
                        messagebox.showerror('שגיאה', str(e))
                        return
                    if not pid:
                        messagebox.showerror('שגיאה', 'לא ניתן ליצור מוצר עבור האתגר')
                        return

                try:
                    dur = int(str(dur_var.get() or '0').strip())
                except Exception:
                    messagebox.showerror('שגיאה', 'משך חייב להיות מספר')
                    return
                try:
                    cap = int(str(cap_var.get() or '0').strip())
                except Exception:
                    messagebox.showerror('שגיאה', 'קיבולת חייבת להיות מספר')
                    return
                if dur <= 0 or cap <= 0:
                    messagebox.showerror('שגיאה', 'משך/קיבולת חייבים להיות גדולים מ-0')
                    return
                st = str(start_var.get() or '').strip()
                et = str(end_var.get() or '').strip()
                if not st or not et:
                    messagebox.showerror('שגיאה', 'חובה להזין שעות התחלה/סיום')
                    return

                mps_raw = str(mps_var.get() or '').strip()
                mpc_raw = str(mpc_var.get() or '').strip()
                mps = None
                mpc = None
                if mps_raw:
                    try:
                        mps = int(mps_raw)
                    except Exception:
                        messagebox.showerror('שגיאה', 'מקסימום לתלמיד חייב להיות מספר או ריק')
                        return
                if mpc_raw:
                    try:
                        mpc = int(mpc_raw)
                    except Exception:
                        messagebox.showerror('שגיאה', 'מקסימום לכיתה חייב להיות מספר או ריק')
                        return

                dates_list = _build_dates_from_ui()
                if not dates_list:
                    messagebox.showerror('שגיאה', 'חובה להזין לפחות תאריך אחד')
                    return

                is_active = 1 if int(enabled_var.get() or 0) == 1 else 0
                allow_auto = 1 if int(auto_var.get() or 0) == 1 else 0

                allowed_cls = str(allowed_classes_var.get() or '').strip()
                min_pts = 0
                try:
                    min_pts = int(str(min_points_var.get() or '0').strip())
                except Exception:
                    min_pts = 0

                try:
                    sid = self.db.upsert_scheduled_service(
                        product_id=int(pid),
                        duration_minutes=int(dur),
                        capacity_per_slot=int(cap),
                        start_time=st,
                        end_time=et,
                        allow_auto_time=int(allow_auto),
                        max_per_student=mps,
                        max_per_class=mpc,
                        queue_priority_mode=str(qp_map.get(str(qp_mode_var.get() or '').strip(), 'class_asc')).strip(),
                        queue_priority_custom=str(qp_custom_var.get() or '').strip(),
                        allowed_classes=allowed_cls,
                        min_points_required=min_pts,
                        is_active=int(is_active),
                    )
                    self.db.set_scheduled_service_dates(service_id=int(sid), dates_greg=dates_list)
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return

                messagebox.showinfo('נשמר', 'האתגר המתוזמן נשמר')
                try:
                    _load_rows(select_pid=int(pid))
                except Exception:
                    pass
                dlg.destroy()

            def _disable():
                if not existing:
                    dlg.destroy()
                    return
                if not messagebox.askyesno('ביטול', 'לבטל את האתגר המתוזמן עבור מוצר זה?'):
                    return
                try:
                    self.db.delete_scheduled_service_by_product(int(pid))
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return
                messagebox.showinfo('בוצע', 'בוטל אתגר מתוזמן למוצר')
                try:
                    _load_rows(select_pid=int(pid))
                except Exception:
                    pass
                dlg.destroy()

            def _export_queue_by_time():
                if not pid:
                    messagebox.showwarning('אזהרה', 'כדי לייצא תור, יש לשמור קודם את האתגר')
                    return
                try:
                    svc = self.db.get_scheduled_service_by_product(int(pid))
                except Exception:
                    svc = None
                if not svc:
                    messagebox.showwarning('אזהרה', 'כדי לייצא תור, יש לשמור קודם את האתגר')
                    return
                try:
                    service_id = int(svc.get('id') or 0)
                except Exception:
                    service_id = 0
                if not service_id:
                    messagebox.showwarning('אזהרה', 'כדי לייצא תור, יש לשמור קודם את האתגר')
                    return

                dates_list = _build_dates_from_ui()
                dates_list = [str(d or '').strip() for d in (dates_list or []) if str(d or '').strip()]
                if not dates_list:
                    messagebox.showwarning('אזהרה', 'אין תאריכים לייצוא')
                    return

                chosen_date = dates_list[0]
                if len(dates_list) > 1:
                    pick = tk.Toplevel(dlg)
                    pick.title('בחר תאריך לייצוא')
                    pick.transient(dlg)
                    pick.grab_set()
                    pick.resizable(False, False)
                    out = {'val': ''}
                    box = tk.Frame(pick, padx=12, pady=10)
                    box.pack(fill=tk.BOTH, expand=True)
                    tk.Label(box, text=fix_rtl_text('בחר תאריך:'), font=('Arial', 10, 'bold')).pack(anchor='e', pady=(0, 6))
                    v = tk.StringVar(value=dates_list[0])
                    cb = ttk.Combobox(box, textvariable=v, values=dates_list, state='readonly', width=18, justify='right')
                    cb.pack(anchor='e')

                    def _ok():
                        out['val'] = str(v.get() or '').strip()
                        pick.destroy()

                    def _cancel():
                        out['val'] = ''
                        pick.destroy()

                    btnrow = tk.Frame(box)
                    btnrow.pack(fill=tk.X, pady=(10, 0))
                    tk.Button(btnrow, text='אישור', width=10, command=_ok).pack(side=tk.LEFT, padx=4)
                    tk.Button(btnrow, text='ביטול', width=10, command=_cancel).pack(side=tk.RIGHT, padx=4)
                    pick.wait_window()
                    chosen_date = str(out.get('val') or '').strip()

                if not chosen_date:
                    return

                try:
                    rows = self.db.get_scheduled_queue_export(service_id=int(service_id), service_date=str(chosen_date)) or []
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return
                if not rows:
                    messagebox.showwarning('אין נתונים', 'אין שיבוצים ליום שנבחר')
                    return

                try:
                    import pandas as pd
                except Exception:
                    messagebox.showerror('שגיאה', 'חסר רכיב pandas לייצוא')
                    return

                # Ask user: preview (default) or save to file
                try:
                    choice_dialog = tk.Toplevel(dlg)
                    choice_dialog.title('ייצוא תור לפי שעות')
                    choice_dialog.geometry('520x240')
                    try:
                        choice_dialog.minsize(520, 220)
                    except Exception:
                        pass
                    choice_dialog.configure(bg='#ecf0f1')
                    choice_dialog.transient(dlg)
                    choice_dialog.grab_set()

                    tk.Label(
                        choice_dialog,
                        text=fix_rtl_text('בחר אופציית ייצוא:'),
                        font=('Arial', 13, 'bold'),
                        bg='#ecf0f1',
                        fg='#2c3e50'
                    ).pack(pady=18)

                    user_choice = {'action': 'preview'}

                    def choose_preview():
                        user_choice['action'] = 'preview'
                        choice_dialog.destroy()

                    def choose_save():
                        user_choice['action'] = 'save'
                        choice_dialog.destroy()

                    btn_frame = tk.Frame(choice_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=12)

                    btn_preview = tk.Button(
                        btn_frame,
                        text='👁 דוח זמני',
                        command=choose_preview,
                        font=('Arial', 12, 'bold'),
                        bg='#3498db',
                        fg='white',
                        padx=18,
                        pady=8
                    )
                    btn_preview.pack(side=tk.LEFT, padx=10)

                    tk.Button(
                        btn_frame,
                        text='💾 שמירה לקובץ',
                        command=choose_save,
                        font=('Arial', 12, 'bold'),
                        bg='#27ae60',
                        fg='white',
                        padx=18,
                        pady=8
                    ).pack(side=tk.LEFT, padx=10)

                    try:
                        btn_preview.focus_set()
                        choice_dialog.bind('<Return>', lambda _e: choose_preview())
                    except Exception:
                        pass

                    try:
                        choice_dialog.wait_window()
                    except Exception:
                        pass
                except Exception:
                    user_choice = {'action': 'preview'}

                data = []
                for r in rows:
                    slot = str(r.get('slot_start_time') or '').strip()
                    durm = str(r.get('duration_minutes') or '').strip()
                    sn = r.get('serial_number')
                    cname = str(r.get('class_name') or '').strip()
                    nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                    data.append({'שעה': slot, 'משך': durm, "מס' סידורי": sn, 'כיתה': cname, 'תלמיד': nm})

                try:
                    df = pd.DataFrame(data, columns=['שעה', 'משך', "מס' סידורי", 'כיתה', 'תלמיד'])

                    if user_choice.get('action') == 'preview':
                        try:
                            self._show_preview_window(df, f'תור לפי שעות - {title_name}')
                        except Exception as e:
                            messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                        return

                    try:
                        fp = filedialog.asksaveasfilename(
                            title='שמור תור לפי שעות',
                            defaultextension='.xlsx',
                            initialdir=self._get_downloads_dir(),
                            filetypes=[('Excel', '*.xlsx')],
                            initialfile=f"תור לפי שעות - {title_name}.xlsx"
                        )
                    except Exception:
                        fp = ''
                    if not fp:
                        return

                    df.to_excel(fp, index=False, engine='openpyxl')
                    try:
                        from openpyxl import load_workbook
                        from excel_styling import apply_rtl_and_alternating_colors
                        wb = load_workbook(fp)
                        ws = wb.active
                        apply_rtl_and_alternating_colors(ws, has_header=True)
                        wb.save(fp)
                    except Exception:
                        pass
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return

                messagebox.showinfo('נשמר', f'נשמר קובץ:\n{fp}')

            # Buttons at bottom of dialog
            tk.Button(btns2, text='💾 שמור', command=_save, font=('Arial', 11, 'bold'), bg='#27ae60', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=6)
            if existing:
                tk.Button(btns2, text='🗑 ביטול אתגר', command=_disable, font=('Arial', 11), bg='#e74c3c', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=6)
            tk.Button(btns2, text='✖ סגור', command=dlg.destroy, font=('Arial', 11), bg='#95a5a6', fg='white', padx=18, pady=8).pack(side=tk.LEFT, padx=6)

        def _open_purchases_by_category_dialog():
            dlg = tk.Toplevel(dialog)
            dlg.title('📊 סיכום רכישות לפי קטגוריה')
            dlg.configure(bg='#ecf0f1')
            dlg.transient(dialog)
            dlg.grab_set()
            try:
                dlg.minsize(720, 520)
            except Exception:
                pass

            tk.Label(dlg, text=fix_rtl_text('סיכום רכישות לפי קטגוריה'), font=('Arial', 13, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(12, 8))

            filters = tk.Frame(dlg, bg='#ecf0f1')
            filters.pack(fill=tk.X, padx=12, pady=(0, 8))

            from_var = tk.StringVar(value='')
            to_var = tk.StringVar(value='')
            include_ref_var = tk.IntVar(value=0)

            tk.Label(filters, text=fix_rtl_text('מתאריך (YYYY-MM-DD):'), bg='#ecf0f1', width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
            tk.Entry(filters, textvariable=from_var, font=('Arial', 10), justify='right', width=14).pack(side=tk.RIGHT, padx=6)
            tk.Label(filters, text=fix_rtl_text('עד תאריך (YYYY-MM-DD):'), bg='#ecf0f1', width=18, anchor='e').pack(side=tk.RIGHT, padx=6)
            tk.Entry(filters, textvariable=to_var, font=('Arial', 10), justify='right', width=14).pack(side=tk.RIGHT, padx=6)
            tk.Checkbutton(filters, text=fix_rtl_text('כולל עסקאות שבוטלו'), variable=include_ref_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)

            cols = ('category', 'rows', 'qty', 'points')
            tree2 = ttk.Treeview(dlg, columns=cols, show='headings', height=14)
            tree2.heading('category', text='קטגוריה')
            tree2.heading('rows', text='שורות')
            tree2.heading('qty', text='כמות')
            tree2.heading('points', text='סה"כ נק')
            tree2.column('category', width=260, anchor='e')
            tree2.column('rows', width=90, anchor='center')
            tree2.column('qty', width=90, anchor='center')
            tree2.column('points', width=120, anchor='center')
            tree2.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 10), side=tk.LEFT)
            sb2 = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=tree2.yview)
            sb2.pack(side=tk.RIGHT, fill=tk.Y, pady=(0, 10))
            tree2.configure(yscrollcommand=sb2.set)

            totals_var = tk.StringVar(value='')
            tk.Label(dlg, textvariable=totals_var, bg='#ecf0f1', fg='#2c3e50', anchor='e').pack(fill=tk.X, padx=12, pady=(0, 10))

            def _reload():
                try:
                    rows = self.db.get_purchases_summary_by_category(
                        from_date=str(from_var.get() or '').strip(),
                        to_date=str(to_var.get() or '').strip(),
                        include_refunded=bool(int(include_ref_var.get() or 0)),
                    ) or []
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e), parent=dlg)
                    return
                try:
                    tree2.delete(*tree2.get_children())
                except Exception:
                    pass
                sum_rows = 0
                sum_qty = 0
                sum_pts = 0
                for r in rows:
                    cat = str(r.get('category_name') or '').strip()
                    try:
                        rc = int(r.get('rows_count') or 0)
                    except Exception:
                        rc = 0
                    try:
                        tq = int(r.get('total_qty') or 0)
                    except Exception:
                        tq = 0
                    try:
                        tp = int(r.get('total_points') or 0)
                    except Exception:
                        tp = 0
                    sum_rows += rc
                    sum_qty += tq
                    sum_pts += tp
                    tree2.insert('', 'end', values=(cat, rc, tq, tp))
                totals_var.set(f"סה\"כ: {sum_rows} שורות | {sum_qty} פריטים | {sum_pts} נק")

            btns2 = tk.Frame(dlg, bg='#ecf0f1')
            btns2.pack(fill=tk.X, padx=12, pady=(0, 12))
            tk.Button(btns2, text='רענן', command=_reload, font=('Arial', 10, 'bold'), bg='#3498db', fg='white', padx=16, pady=6).pack(side=tk.LEFT)
            tk.Button(btns2, text='סגור', command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=(10, 0))
            _reload()

            try:
                dlg.wait_window()
            except Exception:
                pass

        def edit_product():
            pid = _get_selected_product_id()
            if not pid:
                messagebox.showwarning('אזהרה', 'בחר מוצר לעריכה')
                return

            # Decide editor by actual item type
            try:
                ss = self.db.get_scheduled_service_by_product(int(pid))
            except Exception:
                ss = None
            if ss and int((ss or {}).get('is_active', 1) or 0) == 1:
                _open_scheduled_service_dialog(int(pid))
                return

            existing = _get_row_by_pid(pid)
            data = _open_product_dialog(existing=existing)
            if not data.get('ok'):
                return
            try:
                ok = self.db.update_product(
                    pid,
                    name=data['name'],
                    display_name=data.get('display_name', ''),
                    image_path=data.get('image_path', ''),
                    category_id=data.get('category_id', None),
                    price_points=data['price_points'],
                    stock_qty=data['stock_qty'],
                    deduct_points=data['deduct_points'],
                    allowed_classes=data.get('allowed_classes', ''),
                    min_points_required=data.get('min_points_required', 0),
                    max_per_student=data.get('max_per_student', None),
                    max_per_class=data.get('max_per_class', None),
                    price_override_min_points=data.get('price_override_min_points', None),
                    price_override_points=None,
                    price_override_discount_pct=data.get('price_override_discount_pct', None),
                    is_active=data['is_active'],
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            if ok:
                try:
                    _apply_variants(pid, data.get('variants') or [])
                except Exception:
                    pass
                _load_rows(select_pid=pid)

        def delete_product():
            pid = _get_selected_product_id()
            if not pid:
                messagebox.showwarning('אזהרה', 'בחר מוצר למחיקה')
                return
            if not messagebox.askyesno('מחיקה', 'למחוק את המוצר שנבחר?'):
                return
            try:
                self.db.delete_product(pid)
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_rows()

        def toggle_active():
            pid = _get_selected_product_id()
            if not pid:
                messagebox.showwarning('אזהרה', 'בחר מוצר')
                return
            r = _get_row_by_pid(pid)
            if not r:
                return
            try:
                new_active = 0 if int(r.get('is_active', 1) or 0) == 1 else 1
            except Exception:
                new_active = 1
            try:
                self.db.update_product(
                    pid,
                    name=str(r.get('name') or '').strip(),
                    display_name=str(r.get('display_name') or '').strip(),
                    image_path=str(r.get('image_path') or '').strip(),
                    category_id=r.get('category_id', None),
                    price_points=int(r.get('price_points', 0) or 0),
                    stock_qty=r.get('stock_qty', None),
                    deduct_points=int(r.get('deduct_points', 1) or 0),
                    is_active=int(new_active),
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_rows(select_pid=pid)

        btns = tk.Frame(products_tab, bg='#ecf0f1')
        btns.pack(pady=(0, 12))

        def _open_scheduled_for_selected_product():
            pid = _get_selected_product_id()
            if not pid:
                _open_scheduled_service_dialog(0)
                return
            _open_scheduled_service_dialog(int(pid))

        add_menu_btn = tk.Menubutton(btns, text='➕ הוסף', font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6, relief=tk.RAISED)
        add_menu = tk.Menu(add_menu_btn, tearoff=0)
        add_menu.add_command(label=fix_rtl_text('מוצר'), command=add_product)
        add_menu.add_command(label=fix_rtl_text('אתגר מתוזמן'), command=lambda: _open_scheduled_service_dialog(0))
        add_menu_btn.configure(menu=add_menu)
        add_menu_btn.pack(side=tk.LEFT, padx=6)

        tk.Button(btns, text='✏ ערוך', command=edit_product, font=('Arial', 10, 'bold'), bg='#3498db', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)

        def _move_selected(direction: int):
            pid = _get_selected_product_id()
            if not pid:
                messagebox.showwarning('אזהרה', 'בחר מוצר כדי לשנות סדר')
                return
            try:
                ok = self.db.move_product_sort_order(product_id=int(pid), direction=int(direction))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            if ok:
                _load_rows(select_pid=pid)

        def _export_purchases_log():
            try:
                rows = self.db.get_purchases_log_export(limit=5000) or []
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            if not rows:
                messagebox.showwarning('אין נתונים', 'אין רכישות לייצוא')
                return

            # Ask user: preview or save to file
            choice_dialog = tk.Toplevel(dialog)
            choice_dialog.title('ייצוא רכישות')
            choice_dialog.geometry('600x300')
            try:
                choice_dialog.minsize(600, 300)
            except Exception:
                pass
            choice_dialog.configure(bg='#ecf0f1')
            choice_dialog.transient(dialog)
            choice_dialog.grab_set()
            choice_dialog.resizable(True, True)
            
            tk.Label(
                choice_dialog,
                text=fix_rtl_text('בחר אופציית ייצוא:'),
                font=('Arial', 14, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50'
            ).pack(pady=20)
            
            user_choice = {'action': None}
            
            def choose_preview():
                user_choice['action'] = 'preview'
                choice_dialog.destroy()
            
            def choose_save():
                user_choice['action'] = 'save'
                choice_dialog.destroy()
            
            btn_frame = tk.Frame(choice_dialog, bg='#ecf0f1')
            btn_frame.pack(pady=20)
            
            tk.Button(
                btn_frame,
                text='👁 תצוגה זמנית',
                command=choose_preview,
                font=('Arial', 12, 'bold'),
                bg='#3498db',
                fg='white',
                padx=20,
                pady=10
            ).pack(side=tk.LEFT, padx=10)
            
            tk.Button(
                btn_frame,
                text='💾 שמירה לקובץ',
                command=choose_save,
                font=('Arial', 12, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=20,
                pady=10
            ).pack(side=tk.LEFT, padx=10)
            
            choice_dialog.wait_window()
            
            if not user_choice['action']:
                return

            data = []
            for r in rows:
                dt = str(r.get('created_at') or '').strip()
                d = dt[0:10] if len(dt) >= 10 else dt
                t = dt[11:16] if len(dt) >= 16 else ''
                nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                cname = str(r.get('class_name') or '').strip()
                prod_name = str(r.get('product_display_name') or '').strip() or str(r.get('product_name') or '').strip()
                var_name = str(r.get('variant_display_name') or '').strip() or str(r.get('variant_name') or '').strip()
                item_name = prod_name
                if var_name:
                    item_name = f"{item_name} - {var_name}".strip(' -')
                refunded = 'כן' if int(r.get('is_refunded', 0) or 0) == 1 else 'לא'

                data.append({
                    'תאריך': d,
                    'שעה': t,
                    "מס' סידורי": r.get('serial_number'),
                    'כיתה': cname,
                    'תלמיד': nm,
                    'מוצר': item_name,
                    'כמות': r.get('qty'),
                    'נקודות ליחידה': r.get('points_each'),
                    'סה"כ נקודות': r.get('total_points'),
                    'תחנה': str(r.get('station_type') or '').strip(),
                    'בוטל': refunded,
                })

            if user_choice['action'] == 'preview':
                try:
                    import pandas as pd
                except Exception:
                    messagebox.showerror('שגיאה', 'חסר רכיב pandas לייצוא')
                    return

                try:
                    df = pd.DataFrame(
                        data,
                        columns=['תאריך', 'שעה', "מס' סידורי", 'כיתה', 'תלמיד', 'מוצר', 'כמות', 'נקודות ליחידה', 'סה"כ נקודות', 'תחנה', 'בוטל']
                    )
                    self._show_preview_window(df, 'רכישות')
                except Exception as e:
                    messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                return

            else:
                # Save to file
                try:
                    import pandas as pd
                except Exception:
                    messagebox.showerror('שגיאה', 'חסר רכיב pandas לייצוא')
                    return
                
                try:
                    fp = filedialog.asksaveasfilename(
                        title='שמור ייצוא רכישות',
                        defaultextension='.xlsx',
                        filetypes=[('Excel', '*.xlsx')],
                        initialdir=self._get_downloads_dir(),
                        initialfile='רשימת_רכישות.xlsx'
                    )
                except Exception:
                    fp = ''
                if not fp:
                    return
                
                try:
                    df = pd.DataFrame(
                        data,
                        columns=['תאריך', 'שעה', "מס' סידורי", 'כיתה', 'תלמיד', 'מוצר', 'כמות', 'נקודות ליחידה', 'סה"כ נקודות', 'תחנה', 'בוטל']
                    )
                    df.to_excel(fp, index=False, engine='openpyxl')
                    try:
                        from openpyxl import load_workbook
                        from excel_styling import apply_rtl_and_alternating_colors
                        wb = load_workbook(fp)
                        ws = wb.active
                        apply_rtl_and_alternating_colors(ws, has_header=True)
                        wb.save(fp)
                    except Exception:
                        pass
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return

                try:
                    if messagebox.askyesno('הייצוא הסתיים', 'לפתוח את הקובץ עכשיו?'):
                        try:
                            os.startfile(fp)
                        except Exception:
                            pass
                except Exception:
                    pass

        tk.Button(btns, text='⬆ מעלה', command=lambda: _move_selected(-1), font=('Arial', 10, 'bold'), bg='#7f8c8d', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='⬇ מטה', command=lambda: _move_selected(1), font=('Arial', 10, 'bold'), bg='#7f8c8d', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='📤 ייצוא רכישות', command=_export_purchases_log, font=('Arial', 10, 'bold'), bg='#2980b9', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✅ פעיל/כבוי', command=toggle_active, font=('Arial', 10, 'bold'), bg='#16a085', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🗑 מחק', command=delete_product, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)

        # Categories tab (list only; CRUD buttons added separately)
        cat_list_frame = tk.Frame(categories_tab, bg='#ecf0f1')
        cat_list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        cat_cols = ('active', 'name')
        cat_tree = ttk.Treeview(cat_list_frame, columns=cat_cols, show='headings', height=16)
        cat_tree.heading('active', text='פעיל')
        cat_tree.heading('name', text='שם קטגוריה')
        cat_tree.column('active', width=70, anchor='center')
        cat_tree.column('name', width=360, anchor='e')
        cat_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cat_sb = ttk.Scrollbar(cat_list_frame, orient=tk.VERTICAL, command=cat_tree.yview)
        cat_sb.pack(side=tk.RIGHT, fill=tk.Y)
        cat_tree.configure(yscrollcommand=cat_sb.set)

        def _load_categories(select_cid: int = 0):
            try:
                rows = self.db.get_product_categories(active_only=False) or []
            except Exception:
                rows = []
            try:
                cat_tree.delete(*cat_tree.get_children())
            except Exception:
                pass
            to_select = None
            for r in (rows or []):
                try:
                    cid = int(r.get('id') or 0)
                except Exception:
                    cid = 0
                if not cid:
                    continue
                active_txt = 'כן' if int(r.get('is_active', 1) or 0) == 1 else 'לא'
                name = str(r.get('name') or '').strip()
                iid = cat_tree.insert('', 'end', values=(active_txt, name), tags=(f"cid:{cid}",))
                if select_cid and cid == int(select_cid):
                    to_select = iid
            if to_select:
                try:
                    cat_tree.selection_set(to_select)
                    cat_tree.see(to_select)
                except Exception:
                    pass

        try:
            _load_categories()
        except Exception:
            pass

        def _selected_cid() -> int:
            sel = cat_tree.selection()
            if not sel:
                return 0
            tags = cat_tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('cid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _get_cat_row(cid: int) -> dict:
            try:
                rows = self.db.get_product_categories(active_only=False) or []
            except Exception:
                rows = []
            for r in rows:
                try:
                    if int(r.get('id') or 0) == int(cid or 0):
                        return dict(r)
                except Exception:
                    continue
            return {}

        def add_category():
            nm = simpledialog.askstring('קטגוריה', 'שם קטגוריה:', parent=dialog)
            if nm is None:
                return
            nm = str(nm or '').strip()
            if not nm:
                messagebox.showwarning('אזהרה', 'חובה להזין שם קטגוריה')
                return
            is_active = 1 if messagebox.askyesno('קטגוריה', 'להפעיל את הקטגוריה?') else 0
            try:
                cid = self.db.add_product_category(name=nm, sort_order=0, is_active=int(is_active))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_categories(select_cid=int(cid or 0))

        def edit_category():
            cid = _selected_cid()
            if not cid:
                messagebox.showwarning('אזהרה', 'בחר קטגוריה לעריכה')
                return
            row = _get_cat_row(cid)
            nm = simpledialog.askstring('קטגוריה', 'שם קטגוריה:', initialvalue=str(row.get('name') or ''), parent=dialog)
            if nm is None:
                return
            nm = str(nm or '').strip()
            if not nm:
                messagebox.showwarning('אזהרה', 'חובה להזין שם קטגוריה')
                return
            cur_active = 1 if int(row.get('is_active', 1) or 0) == 1 else 0
            is_active = 1 if messagebox.askyesno('קטגוריה', 'להפעיל את הקטגוריה?', default=('yes' if cur_active else 'no')) else 0
            try:
                self.db.update_product_category(int(cid), name=nm, sort_order=int(row.get('sort_order', 0) or 0), is_active=int(is_active))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_categories(select_cid=int(cid))

        def toggle_category_active():
            cid = _selected_cid()
            if not cid:
                messagebox.showwarning('אזהרה', 'בחר קטגוריה')
                return
            row = _get_cat_row(cid)
            if not row:
                return
            new_active = 0 if int(row.get('is_active', 1) or 0) == 1 else 1
            try:
                self.db.update_product_category(int(cid), name=str(row.get('name') or ''), sort_order=int(row.get('sort_order', 0) or 0), is_active=int(new_active))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_categories(select_cid=int(cid))

        def delete_category():
            cid = _selected_cid()
            if not cid:
                messagebox.showwarning('אזהרה', 'בחר קטגוריה למחיקה')
                return
            if not messagebox.askyesno('מחיקה', 'למחוק את הקטגוריה? (מוצרים בקטגוריה יחזרו ל"ללא קטגוריה")'):
                return
            try:
                self.db.delete_product_category(int(cid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_categories()

        cat_btns = tk.Frame(categories_tab, bg='#ecf0f1')
        cat_btns.pack(pady=(0, 12))
        tk.Button(cat_btns, text='➕ הוסף', command=add_category, bg='#27ae60', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(cat_btns, text='✏ ערוך', command=edit_category, bg='#3498db', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(cat_btns, text='✅ פעיל/כבוי', command=toggle_category_active, bg='#16a085', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(cat_btns, text='🗑 מחק', command=delete_category, bg='#e74c3c', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)

        cashier_save_cb = {'fn': None}

        footer = tk.Frame(dialog, bg='#ecf0f1')
        footer.pack(fill=tk.X, pady=(0, 14))
        tk.Button(
            footer,
            text='💾 שמור',
            command=lambda: (cashier_save_cb.get('fn') or (lambda: None))(),
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=18,
            pady=6
        ).pack(side=tk.LEFT, padx=8)
        tk.Button(footer, text='✖ סגור', command=dialog.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=18, pady=6).pack(side=tk.LEFT, padx=8)

        settings_container = tk.Frame(settings_tab, bg='#ecf0f1')
        settings_container.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        # RTL style לרדיובאטון (נקודה מימין לטקסט)
        try:
            _rtl_rb_style = ttk.Style(dialog)
            _rtl_rb_style.layout(
                'Rtl.TRadiobutton',
                [
                    ('Radiobutton.padding', {
                        'sticky': 'nswe',
                        'children': [
                            ('Radiobutton.indicator', {'side': 'right', 'sticky': ''}),
                            ('Radiobutton.label', {'side': 'right', 'sticky': 'nswe'}),
                        ]
                    })
                ]
            )
            _rtl_rb_style.configure('Rtl.TRadiobutton', background='#ecf0f1', font=('Arial', 10))
        except Exception:
            pass

        mode_var = tk.StringVar(value='teacher')
        try:
            mode_var.set(self.db.get_cashier_mode())
        except Exception:
            mode_var.set('teacher')

        idle_var = tk.StringVar(value='300')
        try:
            idle_var.set(str(self.db.get_cashier_idle_timeout_sec()))
        except Exception:
            idle_var.set('300')

        bw_logo_var = tk.StringVar(value='')
        try:
            bw_logo_var.set(self.db.get_cashier_bw_logo_path())
        except Exception:
            bw_logo_var.set('')

        confirm_mode_var = tk.StringVar(value='always')
        try:
            confirm_mode_var.set(self.db.get_cashier_payment_confirm_mode())
        except Exception:
            try:
                confirm_mode_var.set('always' if self.db.get_cashier_require_rescan_confirm() else 'never')
            except Exception:
                confirm_mode_var.set('always')

        confirm_threshold_var = tk.StringVar(value='0')
        try:
            confirm_threshold_var.set(str(self.db.get_cashier_payment_confirm_threshold()))
        except Exception:
            confirm_threshold_var.set('0')

        print_items_var = tk.IntVar(value=0)
        try:
            print_items_var.set(1 if self.db.get_cashier_print_item_receipts() else 0)
        except Exception:
            print_items_var.set(0)

        footer_text_var = tk.StringVar(value='')
        try:
            footer_text_var.set(self.db.get_cashier_receipt_footer_text())
        except Exception:
            footer_text_var.set('תודה ולהתראות')

        ef = tk.LabelFrame(settings_container, text='ייצוא', bg='#ecf0f1', font=('Arial', 11, 'bold'))
        ef.pack(fill=tk.X, pady=(0, 10), anchor='e')

        # Output type selection for catalog export
        catalog_output_var = tk.StringVar(value='preview')
        output_choice_frame = tk.Frame(ef, bg='#ecf0f1')
        output_choice_frame.pack(anchor='e', padx=10, pady=(8, 4))
        
        tk.Radiobutton(
            output_choice_frame,
            text='💾 ייצוא לקובץ Excel',
            variable=catalog_output_var,
            value='file',
            bg='#ecf0f1',
            font=('Arial', 9)
        ).pack(side=tk.RIGHT, padx=8)
        
        tk.Radiobutton(
            output_choice_frame,
            text='👁 דוח זמני',
            variable=catalog_output_var,
            value='preview',
            bg='#ecf0f1',
            font=('Arial', 9)
        ).pack(side=tk.RIGHT, padx=8)

        def _export_catalog_csv(*, mode: str):
            try:
                rows = self.db.get_cashier_catalog_export() or []
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return

            m = str(mode or '').strip().lower()
            if m == 'products':
                rows = [r for r in rows if not int(r.get('challenge_id') or 0)]
                default_name = 'קטלוג_מוצרים.xlsx'
                title = 'קטלוג מוצרים'
            elif m == 'challenges':
                rows = [r for r in rows if int(r.get('challenge_id') or 0)]
                default_name = 'קטלוג_אתגרים.xlsx'
                title = 'קטלוג אתגרים'
            else:
                default_name = 'קטלוג_מלא.xlsx'
                title = 'קטלוג מלא'

            if not rows:
                messagebox.showwarning('אין נתונים', 'אין פריטים לייצוא')
                return

            is_preview = (catalog_output_var.get() == 'preview')
            
            if is_preview:
                # Preview mode - use temporary file
                import tempfile
                temp_fd, fp = tempfile.mkstemp(suffix='.xlsx')
                os.close(temp_fd)
            else:
                # File mode - ask for save location
                try:
                    fp = filedialog.asksaveasfilename(
                        title='שמור ייצוא קטלוג',
                        defaultextension='.xlsx',
                        initialfile=default_name,
                        initialdir=self._get_downloads_dir(),
                        filetypes=[('Excel', '*.xlsx')]
                    )
                except Exception:
                    fp = ''
                if not fp:
                    return

            try:
                import pandas as pd
                df = pd.DataFrame(rows)
                
                # Rename columns to Hebrew
                hebrew_columns = {
                    'product_id': 'מזהה מוצר',
                    'product_is_active': 'פעיל',
                    'product_name': 'שם פנימי',
                    'product_display_name': 'שם תצוגה',
                    'product_price_points': 'מחיר (נקודות)',
                    'product_stock_qty': 'מלאי',
                    'product_image_path': 'נתיב תמונה',
                    'category_name': 'קטגוריה',
                    'product_allowed_classes': 'כיתות מורשות',
                    'product_min_points_required': 'מינימום נקודות',
                    'product_max_per_student': 'מקסימום לתלמיד',
                    'product_max_per_class': 'מקסימום לכיתה',
                    'product_price_override_min_points': 'מינימום נקודות למחיר מיוחד',
                    'product_price_override_discount_pct': 'אחוז הנחה למחיר מיוחד',
                    'challenge_id': 'מזהה אתגר',
                    'challenge_is_active': 'אתגר פעיל',
                    'challenge_duration_minutes': 'משך (דקות)',
                    'challenge_capacity_per_slot': 'קיבולת למשבצת',
                    'challenge_start_time': 'שעת התחלה',
                    'challenge_end_time': 'שעת סיום',
                    'challenge_allow_auto_time': 'בחירת זמן אוטומטית',
                    'challenge_max_per_student': 'מקסימום לתלמיד (אתגר)',
                    'challenge_max_per_class': 'מקסימום לכיתה (אתגר)',
                    'challenge_queue_priority_mode': 'מצב עדיפות',
                    'challenge_queue_priority_custom': 'עדיפות מותאמת',
                    'challenge_allowed_classes': 'כיתות מורשות (אתגר)',
                    'challenge_min_points_required': 'מינימום נקודות (אתגר)'
                }
                df = df.rename(columns=hebrew_columns)
                
                df.to_excel(fp, index=False, engine='openpyxl')
                
                # Apply RTL and alternating colors styling
                try:
                    from openpyxl import load_workbook
                    from excel_styling import apply_rtl_and_alternating_colors
                    wb = load_workbook(fp)
                    ws = wb.active
                    apply_rtl_and_alternating_colors(ws, has_header=True)
                    wb.save(fp)
                except Exception:
                    pass
                
                if is_preview:
                    # Show preview window
                    try:
                        df_display = pd.read_excel(fp, sheet_name=0)
                        self._show_preview_window(df_display, title)
                    except Exception as e:
                        messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                    finally:
                        # Clean up temp file
                        try:
                            os.remove(fp)
                        except Exception:
                            pass
                else:
                    messagebox.showinfo('ייצוא', f'הייצוא נשמר בהצלחה:\n{fp}')
            except Exception as e:
                messagebox.showerror('שגיאה', f'שגיאה בייצוא:\n{str(e)}')
                if is_preview:
                    try:
                        os.remove(fp)
                    except Exception:
                        pass

        def _export_product_purchases():
            """ייצוא רכישות מוצרים - מי קנה מה וכמה"""
            try:
                # Get all products
                products = self.db.get_all_products() or []
                products = [p for p in products if not int(p.get('challenge_id') or 0)]
                if not products:
                    messagebox.showwarning('אין נתונים', 'אין מוצרים במערכת')
                    return
                
                # Let user select products
                selection_dialog = tk.Toplevel(dialog)
                selection_dialog.title('בחר מוצרים לייצוא')
                selection_dialog.geometry('500x600')
                try:
                    selection_dialog.minsize(500, 560)
                except Exception:
                    pass
                selection_dialog.configure(bg='#ecf0f1')
                selection_dialog.transient(dialog)
                selection_dialog.grab_set()
                selection_dialog.resizable(True, True)
                
                tk.Label(
                    selection_dialog,
                    text=fix_rtl_text('בחר מוצרים לייצוא:'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=10)
                
                # Listbox with checkboxes
                list_frame = tk.Frame(selection_dialog, bg='#ecf0f1')
                list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                
                scrollbar = tk.Scrollbar(list_frame)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set, font=('Arial', 10), height=20)
                listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.config(command=listbox.yview)
                
                product_map = {}
                for p in products:
                    display_name = str(p.get('display_name') or p.get('name') or '').strip()
                    if display_name:
                        listbox.insert(tk.END, display_name)
                        product_map[display_name] = p
                
                def select_all():
                    listbox.select_set(0, tk.END)
                
                def deselect_all():
                    listbox.select_clear(0, tk.END)
                
                btn_frame = tk.Frame(selection_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=5)
                tk.Button(btn_frame, text='בחר הכל', command=select_all, font=('Arial', 9), bg='#95a5a6', fg='white', padx=10, pady=4).pack(side=tk.LEFT, padx=5)
                tk.Button(btn_frame, text='בטל הכל', command=deselect_all, font=('Arial', 9), bg='#95a5a6', fg='white', padx=10, pady=4).pack(side=tk.LEFT, padx=5)
                
                result = {'selected': []}
                
                def confirm_selection():
                    selected_indices = listbox.curselection()
                    if not selected_indices:
                        messagebox.showwarning('אזהרה', 'לא נבחרו מוצרים')
                        return
                    result['selected'] = [listbox.get(i) for i in selected_indices]
                    selection_dialog.destroy()
                
                action_frame = tk.Frame(selection_dialog, bg='#ecf0f1')
                action_frame.pack(pady=10)
                tk.Button(
                    action_frame,
                    text='ייצא',
                    command=confirm_selection,
                    font=('Arial', 12, 'bold'),
                    bg='#27ae60',
                    fg='white',
                    padx=20,
                    pady=8
                ).pack(side=tk.LEFT, padx=5)
                tk.Button(
                    action_frame,
                    text='ביטול',
                    command=selection_dialog.destroy,
                    font=('Arial', 12),
                    bg='#95a5a6',
                    fg='white',
                    padx=20,
                    pady=8
                ).pack(side=tk.LEFT, padx=5)
                
                selection_dialog.wait_window()
                
                if not result['selected']:
                    return
                
                # Get purchase data for selected products
                selected_product_ids = [product_map[name]['id'] for name in result['selected']]
                
                data = []
                for pid in selected_product_ids:
                    product = next((p for p in products if p['id'] == pid), None)
                    if not product:
                        continue
                    
                    product_name = str(product.get('display_name') or product.get('name') or '').strip()
                    
                    # Get purchases for this product
                    try:
                        purchases = self.db.get_purchases_by_product(pid) or []
                    except Exception:
                        purchases = []
                    
                    # Group by student
                    student_purchases = {}
                    for purchase in purchases:
                        if int(purchase.get('is_refunded', 0) or 0) == 1:
                            continue
                        sid = purchase.get('student_id')
                        qty = int(purchase.get('qty', 0) or 0)
                        if sid in student_purchases:
                            student_purchases[sid]['qty'] += qty
                        else:
                            student_purchases[sid] = {
                                'student_name': f"{purchase.get('first_name', '')} {purchase.get('last_name', '')}".strip(),
                                'class_name': purchase.get('class_name', ''),
                                'qty': qty
                            }
                    
                    total_qty = sum(s['qty'] for s in student_purchases.values())
                    
                    for sid, info in student_purchases.items():
                        data.append({
                            'מוצר': product_name,
                            'תלמיד': info['student_name'],
                            'כיתה': info['class_name'],
                            'כמות': info['qty'],
                            'סה"כ מוצר': total_qty
                        })
                
                if not data:
                    messagebox.showwarning('אין נתונים', 'אין רכישות למוצרים שנבחרו')
                    return
                
                # Export to Excel
                is_preview = (catalog_output_var.get() == 'preview')
                
                if is_preview:
                    import tempfile
                    temp_fd, fp = tempfile.mkstemp(suffix='.xlsx')
                    os.close(temp_fd)
                else:
                    try:
                        fp = filedialog.asksaveasfilename(
                            title='שמור ייצוא רכישות מוצרים',
                            defaultextension='.xlsx',
                            initialfile='רכישות_מוצרים.xlsx',
                            initialdir=self._get_downloads_dir(),
                            filetypes=[('Excel', '*.xlsx')]
                        )
                    except Exception:
                        fp = ''
                    if not fp:
                        return
                
                try:
                    import pandas as pd
                    df = pd.DataFrame(data, columns=['מוצר', 'תלמיד', 'כיתה', 'כמות', 'סה"כ מוצר'])
                    df.to_excel(fp, index=False, engine='openpyxl')
                    
                    # Apply styling
                    try:
                        from openpyxl import load_workbook
                        from excel_styling import apply_rtl_and_alternating_colors
                        wb = load_workbook(fp)
                        ws = wb.active
                        apply_rtl_and_alternating_colors(ws, has_header=True)
                        wb.save(fp)
                    except Exception:
                        pass
                    
                    if is_preview:
                        try:
                            df_display = pd.read_excel(fp, sheet_name=0)
                            self._show_preview_window(df_display, 'רכישות מוצרים')
                        except Exception as e:
                            messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                        finally:
                            try:
                                os.remove(fp)
                            except Exception:
                                pass
                    else:
                        messagebox.showinfo('ייצוא', f'הייצוא נשמר בהצלחה:\n{fp}')
                        try:
                            if messagebox.askyesno('הייצוא הסתיים', 'לפתוח את הקובץ עכשיו?'):
                                try:
                                    os.startfile(fp)
                                except Exception:
                                    pass
                        except Exception:
                            pass
                except Exception as e:
                    messagebox.showerror('שגיאה', f'שגיאה בייצוא:\n{str(e)}')
                    if is_preview:
                        try:
                            os.remove(fp)
                        except Exception:
                            pass
                    
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))

        def _export_challenge_schedules():
            """ייצוא תורנות אתגרים - לפי אתגר או כל האתגרים"""
            try:
                # Get all scheduled services (challenges) directly from scheduled_services
                services = []
                try:
                    services = self.db.get_all_scheduled_services(active_only=True) or []
                except Exception:
                    services = []

                for svc in services:
                    try:
                        svc['product_name'] = str(svc.get('product_display_name') or svc.get('product_name') or '').strip()
                    except Exception:
                        svc['product_name'] = ''
                
                if not services:
                    messagebox.showwarning('אין נתונים', 'אין אתגרים מתוזמנים במערכת')
                    return
                
                # Let user choose: single challenge or all challenges
                choice_dialog = tk.Toplevel(dialog)
                choice_dialog.title('בחר אופציית ייצוא')
                choice_dialog.geometry('500x400')
                try:
                    choice_dialog.minsize(500, 360)
                except Exception:
                    pass
                choice_dialog.configure(bg='#ecf0f1')
                choice_dialog.transient(dialog)
                choice_dialog.grab_set()
                choice_dialog.resizable(True, True)
                
                tk.Label(
                    choice_dialog,
                    text=fix_rtl_text('בחר אתגר או כל האתגרים:'),
                    font=('Arial', 14, 'bold'),
                    bg='#ecf0f1',
                    fg='#2c3e50'
                ).pack(pady=10)
                
                result = {'mode': None, 'service_id': None}
                
                def export_all():
                    result['mode'] = 'all'
                    choice_dialog.destroy()
                
                def export_single():
                    # Show service selection
                    if not services:
                        return
                    
                    select_dialog = tk.Toplevel(choice_dialog)
                    select_dialog.title('בחר אתגר')
                    select_dialog.geometry('400x500')
                    try:
                        select_dialog.minsize(400, 460)
                    except Exception:
                        pass
                    select_dialog.configure(bg='#ecf0f1')
                    select_dialog.transient(choice_dialog)
                    select_dialog.grab_set()
                    select_dialog.resizable(True, True)
                    
                    tk.Label(
                        select_dialog,
                        text=fix_rtl_text('בחר אתגר:'),
                        font=('Arial', 12, 'bold'),
                        bg='#ecf0f1'
                    ).pack(pady=10)
                    
                    list_frame = tk.Frame(select_dialog, bg='#ecf0f1')
                    list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                    
                    scrollbar = tk.Scrollbar(list_frame)
                    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                    
                    listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('Arial', 10))
                    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    scrollbar.config(command=listbox.yview)
                    
                    service_map = {}
                    for svc in services:
                        name = svc.get('product_name', '')
                        if name:
                            listbox.insert(tk.END, name)
                            service_map[name] = svc['id']
                    
                    def confirm():
                        selection = listbox.curselection()
                        if not selection:
                            messagebox.showwarning('אזהרה', 'לא נבחר אתגר')
                            return
                        selected_name = listbox.get(selection[0])
                        result['mode'] = 'single'
                        result['service_id'] = service_map[selected_name]
                        select_dialog.destroy()
                        choice_dialog.destroy()
                    
                    btn_frame = tk.Frame(select_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=10)
                    tk.Button(btn_frame, text='אישור', command=confirm, font=('Arial', 11, 'bold'), bg='#27ae60', fg='white', padx=20, pady=8).pack(side=tk.LEFT, padx=5)
                    tk.Button(btn_frame, text='ביטול', command=select_dialog.destroy, font=('Arial', 11), bg='#95a5a6', fg='white', padx=20, pady=8).pack(side=tk.LEFT, padx=5)
                
                btn_frame = tk.Frame(choice_dialog, bg='#ecf0f1')
                btn_frame.pack(pady=20)
                
                tk.Button(
                    btn_frame,
                    text='ייצוא אתגר בודד',
                    command=export_single,
                    font=('Arial', 12, 'bold'),
                    bg='#3498db',
                    fg='white',
                    padx=20,
                    pady=10
                ).pack(pady=10)
                
                tk.Button(
                    btn_frame,
                    text='ייצוא כל האתגרים (לשוניות נפרדות)',
                    command=export_all,
                    font=('Arial', 12, 'bold'),
                    bg='#16a085',
                    fg='white',
                    padx=20,
                    pady=10
                ).pack(pady=10)
                
                tk.Button(
                    btn_frame,
                    text='ביטול',
                    command=choice_dialog.destroy,
                    font=('Arial', 11),
                    bg='#95a5a6',
                    fg='white',
                    padx=20,
                    pady=8
                ).pack(pady=10)
                
                choice_dialog.wait_window()
                
                if not result['mode']:
                    return
                
                # Export based on selection
                is_preview = (catalog_output_var.get() == 'preview')

                def _extract_date_list(items):
                    out = []
                    for it in (items or []):
                        if isinstance(it, dict):
                            v = it.get('service_date')
                        else:
                            v = it
                        v = str(v or '').strip()
                        if v:
                            out.append(v)
                    return out

                def _sanitize_sheet_name(name: str, fallback: str) -> str:
                    n = str(name or '').strip() or str(fallback or '').strip() or 'תורנות'
                    for ch in (':', '\\', '/', '?', '*', '[', ']'):
                        n = n.replace(ch, '_')
                    n = n[:31].strip()
                    return n or 'תורנות'
                
                if result['mode'] == 'single':
                    # Export single challenge
                    service_id = result['service_id']
                    service = next((s for s in services if s['id'] == service_id), None)
                    if not service:
                        return
                    
                    service_name = service.get('product_name', 'אתגר')
                    
                    # Get dates for this service
                    try:
                        dates = self.db.get_scheduled_service_dates(service_id) or []
                    except Exception:
                        dates = []
                    
                    if not dates:
                        messagebox.showwarning('אין נתונים', 'אין תאריכים מוגדרים לאתגר זה')
                        return
                    
                    # Let user choose date
                    date_list = _extract_date_list(dates)
                    if not date_list:
                        messagebox.showwarning('אין נתונים', 'אין תאריכים מוגדרים')
                        return
                    
                    chosen_date = date_list[0]
                    if len(date_list) > 1:
                        # Show date picker
                        date_dialog = tk.Toplevel(dialog)
                        date_dialog.title('בחר תאריך')
                        date_dialog.geometry('300x200')
                        try:
                            date_dialog.minsize(300, 200)
                        except Exception:
                            pass
                        date_dialog.configure(bg='#ecf0f1')
                        date_dialog.transient(dialog)
                        date_dialog.grab_set()
                        
                        tk.Label(date_dialog, text=fix_rtl_text('בחר תאריך:'), font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(pady=10)
                        
                        date_var = tk.StringVar(value=date_list[0])
                        combo = ttk.Combobox(date_dialog, textvariable=date_var, values=date_list, state='readonly', width=20, justify='right')
                        combo.pack(pady=10)
                        
                        date_result = {'date': None}
                        
                        def confirm_date():
                            date_result['date'] = date_var.get()
                            date_dialog.destroy()
                        
                        tk.Button(date_dialog, text='אישור', command=confirm_date, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=20, pady=6).pack(pady=10)
                        
                        date_dialog.wait_window()
                        
                        if date_result['date']:
                            chosen_date = date_result['date']
                        else:
                            return
                    
                    # Get queue data
                    try:
                        rows = self.db.get_scheduled_queue_export(service_id=service_id, service_date=chosen_date) or []
                    except Exception as e:
                        messagebox.showerror('שגיאה', str(e))
                        return
                    
                    if not rows:
                        messagebox.showwarning('אין נתונים', 'אין שיבוצים ליום שנבחר')
                        return
                    
                    # Format data
                    data = []
                    for r in rows:
                        slot = str(r.get('slot_start_time') or '').strip()
                        end_time = str(r.get('slot_end_time') or '').strip()
                        time_range = f"{slot}-{end_time}" if end_time else slot
                        nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                        cname = str(r.get('class_name') or '').strip()
                        data.append({
                            'שעות': time_range,
                            'תלמיד': nm,
                            'כיתה': cname
                        })
                    
                    # Group by time slot
                    grouped_data = {}
                    for item in data:
                        time_slot = item['שעות']
                        if time_slot not in grouped_data:
                            grouped_data[time_slot] = []
                        grouped_data[time_slot].append(f"{item['תלמיד']} ({item['כיתה']})")
                    
                    # Create final data
                    final_data = []
                    for time_slot in sorted(grouped_data.keys()):
                        students = ' | '.join(grouped_data[time_slot])
                        final_data.append({
                            'שעות': time_slot,
                            'תלמידים': students
                        })
                    
                    if is_preview:
                        import tempfile
                        temp_fd, fp = tempfile.mkstemp(suffix='.xlsx')
                        os.close(temp_fd)
                    else:
                        try:
                            fp = filedialog.asksaveasfilename(
                                title='שמור ייצוא תורנות',
                                defaultextension='.xlsx',
                                initialfile=f'תורנות_{service_name}.xlsx',
                                initialdir=self._get_downloads_dir(),
                                filetypes=[('Excel', '*.xlsx')]
                            )
                        except Exception:
                            fp = ''
                        if not fp:
                            return
                    
                    try:
                        import pandas as pd
                        df = pd.DataFrame(final_data, columns=['שעות', 'תלמידים'])
                        df.to_excel(fp, index=False, engine='openpyxl')
                        
                        # Apply styling
                        try:
                            from openpyxl import load_workbook
                            from excel_styling import apply_rtl_and_alternating_colors
                            wb = load_workbook(fp)
                            ws = wb.active
                            apply_rtl_and_alternating_colors(ws, has_header=True)
                            wb.save(fp)
                        except Exception:
                            pass
                        
                        if is_preview:
                            try:
                                df_display = pd.read_excel(fp, sheet_name=0)
                                self._show_preview_window(df_display, f'תורנות {service_name}')
                            except Exception as e:
                                messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                            finally:
                                try:
                                    os.remove(fp)
                                except Exception:
                                    pass
                        else:
                            messagebox.showinfo('ייצוא', f'הייצוא נשמר בהצלחה:\n{fp}')
                    except Exception as e:
                        messagebox.showerror('שגיאה', f'שגיאה בייצוא:\n{str(e)}')
                        if is_preview:
                            try:
                                os.remove(fp)
                            except Exception:
                                pass
                
                else:  # Export all challenges
                    if is_preview:
                        import tempfile
                        temp_fd, fp = tempfile.mkstemp(suffix='.xlsx')
                        os.close(temp_fd)
                    else:
                        try:
                            fp = filedialog.asksaveasfilename(
                                title='שמור ייצוא תורנות',
                                defaultextension='.xlsx',
                                initialfile='תורנות_כל_האתגרים.xlsx',
                                initialdir=self._get_downloads_dir(),
                                filetypes=[('Excel', '*.xlsx')]
                            )
                        except Exception:
                            fp = ''
                        if not fp:
                            return
                    
                    try:
                        import pandas as pd
                        
                        with pd.ExcelWriter(fp, engine='openpyxl') as writer:
                            wrote_any = False
                            
                            for service in services:
                                try:
                                    service_id = int(service.get('id') or 0)
                                except Exception:
                                    service_id = 0
                                if not service_id:
                                    continue
                                service_name = str(service.get('product_name') or '').strip() or 'אתגר'

                                try:
                                    dates = self.db.get_scheduled_service_dates(service_id) or []
                                except Exception:
                                    dates = []
                                if not dates:
                                    continue

                                date_list = _extract_date_list(dates)
                                if not date_list:
                                    continue
                                chosen_date = date_list[0]

                                try:
                                    rows = self.db.get_scheduled_queue_export(service_id=service_id, service_date=chosen_date) or []
                                except Exception:
                                    continue
                                if not rows:
                                    continue

                                data = []
                                for r in rows:
                                    slot = str(r.get('slot_start_time') or '').strip()
                                    end_time = str(r.get('slot_end_time') or '').strip()
                                    time_range = f"{slot}-{end_time}" if end_time else slot
                                    nm = f"{str(r.get('first_name') or '').strip()} {str(r.get('last_name') or '').strip()}".strip()
                                    cname = str(r.get('class_name') or '').strip()
                                    data.append({
                                        'שעות': time_range,
                                        'תלמיד': nm,
                                        'כיתה': cname
                                    })

                                grouped_data = {}
                                for item in data:
                                    time_slot = item['שעות']
                                    if time_slot not in grouped_data:
                                        grouped_data[time_slot] = []
                                    grouped_data[time_slot].append(f"{item['תלמיד']} ({item['כיתה']})")

                                final_data = []
                                for time_slot in sorted(grouped_data.keys()):
                                    students = ' | '.join(grouped_data[time_slot])
                                    final_data.append({
                                        'שעות': time_slot,
                                        'תלמידים': students
                                    })

                                df = pd.DataFrame(final_data, columns=['שעות', 'תלמידים'])
                                sheet_name = _sanitize_sheet_name(service_name, f"אתגר_{service_id}")
                                try:
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                except Exception:
                                    sheet_name = _sanitize_sheet_name(f"אתגר_{service_id}", 'תורנות')
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                wrote_any = True
                            
                            if not wrote_any:
                                # Create empty sheet
                                pd.DataFrame([], columns=['שעות', 'תלמידים']).to_excel(writer, sheet_name='תורנות', index=False)
                        
                        # Apply styling to all sheets
                        try:
                            from openpyxl import load_workbook
                            from excel_styling import apply_rtl_and_alternating_colors
                            wb = load_workbook(fp)
                            for ws in wb.worksheets:
                                apply_rtl_and_alternating_colors(ws, has_header=True)
                            wb.save(fp)
                        except Exception:
                            pass
                        
                        if is_preview:
                            try:
                                df_display = pd.read_excel(fp, sheet_name=0)
                                self._show_preview_window(df_display, 'תורנות אתגרים')
                            except Exception as e:
                                messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                            finally:
                                try:
                                    os.remove(fp)
                                except Exception:
                                    pass
                        else:
                            messagebox.showinfo('ייצוא', f'הייצוא נשמר בהצלחה:\n{fp}')
                    except Exception as e:
                        messagebox.showerror('שגיאה', f'שגיאה בייצוא:\n{str(e)}')
                        if is_preview:
                            try:
                                os.remove(fp)
                            except Exception:
                                pass
                    
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))

        ebtns = tk.Frame(ef, bg='#ecf0f1')
        ebtns.pack(anchor='e', padx=10, pady=8)
        tk.Button(ebtns, text='📤 ייצוא רכישות מוצרים', command=_export_product_purchases, bg='#3498db', fg='white', padx=14, pady=6).pack(side=tk.RIGHT, padx=6)
        tk.Button(ebtns, text='📤 ייצוא תורנות אתגרים', command=_export_challenge_schedules, bg='#16a085', fg='white', padx=14, pady=6).pack(side=tk.RIGHT, padx=6)
        tk.Button(ebtns, text='📤 ייצוא קטלוג', command=lambda: _export_catalog_csv(mode='all'), bg='#2c3e50', fg='white', padx=14, pady=6).pack(side=tk.RIGHT, padx=6)

        mf = tk.LabelFrame(settings_container, text='מצב קופה', bg='#ecf0f1', font=('Arial', 11, 'bold'))
        mf.pack(fill=tk.X, pady=(0, 10), anchor='e')

        mode_row = tk.Frame(mf, bg='#ecf0f1')
        mode_row.pack(anchor='e', padx=10, pady=6)
        ttk.Radiobutton(mode_row, text=fix_rtl_text('ניהול ע"י מורה'), variable=mode_var, value='teacher', style='Rtl.TRadiobutton').pack(side=tk.RIGHT, padx=10, pady=2)
        ttk.Radiobutton(mode_row, text=fix_rtl_text('ניהול ע"י תלמיד אחראי'), variable=mode_var, value='responsible_student', style='Rtl.TRadiobutton').pack(side=tk.RIGHT, padx=10, pady=2)
        ttk.Radiobutton(mode_row, text=fix_rtl_text('שירות עצמי (פתיחה ע"י תלמיד)'), variable=mode_var, value='self_service', style='Rtl.TRadiobutton').pack(side=tk.RIGHT, padx=10, pady=2)

        sf = tk.LabelFrame(settings_container, text='פרטים', bg='#ecf0f1', font=('Arial', 11, 'bold'))
        sf.pack(fill=tk.X, pady=(0, 10), anchor='e')

        r_idle = tk.Frame(sf, bg='#ecf0f1')
        r_idle.pack(fill=tk.X, pady=6)
        tk.Label(r_idle, text=fix_rtl_text('נעילה אוטומטית (שניות):'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
        tk.Entry(r_idle, textvariable=idle_var, font=('Arial', 11), justify='right', width=10).pack(side=tk.RIGHT, padx=6)
        tk.Label(r_idle, text=fix_rtl_text('(300 = 5 דקות)'), bg='#ecf0f1', fg='#7f8c8d').pack(side=tk.RIGHT, padx=6)

        r_logo = tk.Frame(sf, bg='#ecf0f1')
        r_logo.pack(fill=tk.X, pady=6)
        tk.Label(r_logo, text=fix_rtl_text('לוגו שחור-לבן להדפסה:'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
        tk.Entry(r_logo, textvariable=bw_logo_var, font=('Arial', 10), justify='right', width=54).pack(side=tk.RIGHT, padx=6)

        def _browse_bw_logo():
            try:
                fp = filedialog.askopenfilename(
                    title='בחר לוגו שחור-לבן',
                    filetypes=[('תמונות', '*.png;*.jpg;*.jpeg;*.bmp;*.gif'), ('הכל', '*.*')]
                )
            except Exception:
                fp = ''
            if fp:
                bw_logo_var.set(fp)

        tk.Button(r_logo, text='בחר…', command=_browse_bw_logo, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=10, pady=3).pack(side=tk.RIGHT, padx=6)

        r_confirm = tk.Frame(sf, bg='#ecf0f1')
        r_confirm.pack(fill=tk.X, pady=6)
        tk.Label(r_confirm, text=fix_rtl_text('אישור תשלום:'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)

        r_confirm_inner = tk.Frame(r_confirm, bg='#ecf0f1')
        r_confirm_inner.pack(side=tk.RIGHT, padx=6)

        confirm_row = tk.Frame(r_confirm_inner, bg='#ecf0f1')
        confirm_row.pack(anchor='e')

        # סדר RTL: מימין לשמאל -> תמיד, אף פעם לא, רק מעל סף נקודות + שדה
        ttk.Radiobutton(confirm_row, text=fix_rtl_text('תמיד בסריקה חוזרת'), variable=confirm_mode_var, value='always', style='Rtl.TRadiobutton').pack(side=tk.RIGHT, padx=(0, 12))
        ttk.Radiobutton(confirm_row, text=fix_rtl_text('אף פעם לא'), variable=confirm_mode_var, value='never', style='Rtl.TRadiobutton').pack(side=tk.RIGHT, padx=(0, 12))

        thr = tk.Frame(confirm_row, bg='#ecf0f1')
        thr.pack(side=tk.RIGHT, padx=(0, 12))
        ttk.Radiobutton(thr, text=fix_rtl_text('רק מעל סף נקודות:'), variable=confirm_mode_var, value='threshold', style='Rtl.TRadiobutton').pack(side=tk.RIGHT)
        tk.Entry(thr, textvariable=confirm_threshold_var, font=('Arial', 11), justify='right', width=8).pack(side=tk.RIGHT, padx=(6, 0))

        r_print_items = tk.Frame(sf, bg='#ecf0f1')
        r_print_items.pack(fill=tk.X, pady=6)
        tk.Label(r_print_items, text=fix_rtl_text('הדפסות:'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
        tk.Checkbutton(r_print_items, text=fix_rtl_text('להדפיס שובר נוסף לכל מוצר/אתגר (למפעיל הדוכן)'), variable=print_items_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=6)

        r_footer = tk.Frame(sf, bg='#ecf0f1')
        r_footer.pack(fill=tk.X, pady=6)
        tk.Label(r_footer, text=fix_rtl_text('ברכת סיום בקבלה:'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
        tk.Entry(r_footer, textvariable=footer_text_var, font=('Arial', 10), justify='right', width=54).pack(side=tk.RIGHT, padx=6)

        responsibles_cb = {'fn': None}

        r_resp = tk.Frame(sf, bg='#ecf0f1')
        r_resp.pack(fill=tk.X, pady=6)
        tk.Label(r_resp, text=fix_rtl_text('תלמידים אחראיים:'), bg='#ecf0f1', width=22, anchor='e').pack(side=tk.RIGHT, padx=6)
        tk.Button(
            r_resp,
            text='פתח ניהול…',
            command=lambda: (responsibles_cb.get('fn') or (lambda: None))(),
            font=('Arial', 9, 'bold'),
            bg='#3498db',
            fg='white',
            padx=10,
            pady=3
        ).pack(side=tk.RIGHT, padx=6)

        rf = tk.LabelFrame(settings_container, text='תלמידים אחראיים', bg='#ecf0f1', font=('Arial', 11, 'bold'))
        # הועבר לניהול בחלון נפרד כדי למנוע גלילה שמסתירה את כפתורי השמירה
        rf.pack_forget()

        rlist = tk.Frame(rf, bg='#ecf0f1')
        rlist.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        rcols = ('serial', 'name', 'class')
        rtree = ttk.Treeview(rlist, columns=rcols, show='headings', height=8)
        rtree.heading('serial', text='מס"ד')
        rtree.heading('name', text='שם')
        rtree.heading('class', text='כיתה')
        rtree.column('serial', width=70, anchor='center')
        rtree.column('name', width=260, anchor='e')
        rtree.column('class', width=120, anchor='center')
        rtree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rsb = ttk.Scrollbar(rlist, orient=tk.VERTICAL, command=rtree.yview)
        rsb.pack(side=tk.RIGHT, fill=tk.Y)
        rtree.configure(yscrollcommand=rsb.set)

        def _load_responsibles():
            try:
                rows = self.db.get_cashier_responsibles() or []
            except Exception:
                rows = []
            try:
                rtree.delete(*rtree.get_children())
            except Exception:
                pass
            for s in rows:
                try:
                    sid = int(s.get('id') or 0)
                except Exception:
                    sid = 0
                if not sid:
                    continue
                sn = s.get('serial_number', '')
                name = f"{(s.get('first_name') or '').strip()} {(s.get('last_name') or '').strip()}".strip()
                cls = str(s.get('class_name') or '').strip()
                rtree.insert('', 'end', values=(sn, name, cls), tags=(f"sid:{sid}",))

        def _get_selected_responsible_id() -> int:
            sel = rtree.selection()
            if not sel:
                return 0
            tags = rtree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('sid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _pick_student_dialog() -> int:
            dlg = tk.Toplevel(dialog)
            dlg.title('בחר תלמיד')
            dlg.configure(bg='#ecf0f1')
            dlg.transient(dialog)
            dlg.grab_set()
            result = {'sid': 0}

            search_var = tk.StringVar(value='')
            tk.Label(dlg, text=fix_rtl_text('חיפוש:'), bg='#ecf0f1').pack(anchor='e', padx=10, pady=(10, 0))
            ent = tk.Entry(dlg, textvariable=search_var, font=('Arial', 11), justify='right')
            ent.pack(fill=tk.X, padx=10, pady=6)

            frame = tk.Frame(dlg, bg='#ecf0f1')
            frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            cols = ('serial', 'name', 'class')
            st = ttk.Treeview(frame, columns=cols, show='headings', height=10)
            st.heading('serial', text='מס"ד')
            st.heading('name', text='שם')
            st.heading('class', text='כיתה')
            st.column('serial', width=70, anchor='center')
            st.column('name', width=280, anchor='e')
            st.column('class', width=120, anchor='center')
            st.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            ssb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=st.yview)
            ssb.pack(side=tk.RIGHT, fill=tk.Y)
            st.configure(yscrollcommand=ssb.set)

            try:
                all_students = self.db.get_all_students() or []
            except Exception:
                all_students = []

            def _refresh():
                q = search_var.get().strip().lower()
                try:
                    st.delete(*st.get_children())
                except Exception:
                    pass
                for s in all_students:
                    name = f"{(s.get('first_name') or '').strip()} {(s.get('last_name') or '').strip()}".strip()
                    cls = str(s.get('class_name') or '').strip()
                    sn = str(s.get('serial_number') or '').strip()
                    if q and (q not in name.lower()) and (q not in cls.lower()) and (q not in sn.lower()):
                        continue
                    try:
                        sid = int(s.get('id') or 0)
                    except Exception:
                        sid = 0
                    if not sid:
                        continue
                    st.insert('', 'end', values=(sn, name, cls), tags=(f"sid:{sid}",))

            def _ok():
                sel = st.selection()
                if not sel:
                    return
                tags = st.item(sel[0]).get('tags') or ()
                for t in tags:
                    if isinstance(t, str) and t.startswith('sid:'):
                        try:
                            result['sid'] = int(t.split(':', 1)[1])
                        except Exception:
                            result['sid'] = 0
                dlg.destroy()

            def _cancel():
                dlg.destroy()

            b = tk.Frame(dlg, bg='#ecf0f1')
            b.pack(fill=tk.X, padx=10, pady=(0, 10))
            tk.Button(b, text='אישור', command=_ok, width=10).pack(side=tk.LEFT, padx=5)
            tk.Button(b, text='ביטול', command=_cancel, width=10).pack(side=tk.RIGHT, padx=5)

            try:
                ent.bind('<KeyRelease>', lambda _e: _refresh())
            except Exception:
                pass
            try:
                st.bind('<Double-1>', lambda _e: _ok())
            except Exception:
                pass

            _refresh()
            try:
                dlg.minsize(560, 420)
            except Exception:
                pass
            dlg.wait_window()
            return int(result.get('sid') or 0)

        def add_responsible():
            sid = _pick_student_dialog()
            if not sid:
                return
            try:
                self.db.add_cashier_responsible(int(sid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_responsibles()

        def remove_responsible():
            sid = _get_selected_responsible_id()
            if not sid:
                messagebox.showwarning('אזהרה', 'בחר תלמיד להסרה')
                return
            try:
                self.db.remove_cashier_responsible(int(sid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load_responsibles()

        def _open_responsibles_manager():
            dlg = tk.Toplevel(dialog)
            dlg.title('תלמידים אחראיים')
            dlg.configure(bg='#ecf0f1')
            dlg.transient(dialog)
            dlg.grab_set()
            dlg.resizable(True, True)

            rlist = tk.Frame(dlg, bg='#ecf0f1')
            rlist.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

            cols = ('serial', 'name', 'class')
            tree = ttk.Treeview(rlist, columns=cols, show='headings', height=12)
            tree.heading('serial', text='מס"ד')
            tree.heading('name', text='שם')
            tree.heading('class', text='כיתה')
            tree.column('serial', width=70, anchor='center')
            tree.column('name', width=320, anchor='e')
            tree.column('class', width=140, anchor='center')
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb = ttk.Scrollbar(rlist, orient=tk.VERTICAL, command=tree.yview)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            tree.configure(yscrollcommand=sb.set)

            def _reload():
                try:
                    rows = self.db.get_cashier_responsibles() or []
                except Exception:
                    rows = []
                try:
                    tree.delete(*tree.get_children())
                except Exception:
                    pass
                for s in rows:
                    try:
                        sid = int(s.get('id') or 0)
                    except Exception:
                        sid = 0
                    if not sid:
                        continue
                    sn = s.get('serial_number', '')
                    name = f"{(s.get('first_name') or '').strip()} {(s.get('last_name') or '').strip()}".strip()
                    cls = str(s.get('class_name') or '').strip()
                    tree.insert('', 'end', values=(sn, name, cls), tags=(f"sid:{sid}",))

            def _selected_sid() -> int:
                sel = tree.selection()
                if not sel:
                    return 0
                tags = tree.item(sel[0]).get('tags') or ()
                for t in tags:
                    if isinstance(t, str) and t.startswith('sid:'):
                        try:
                            return int(t.split(':', 1)[1])
                        except Exception:
                            return 0
                return 0

            def _add():
                sid = _pick_student_dialog()
                if not sid:
                    return
                try:
                    self.db.add_cashier_responsible(int(sid))
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return
                _reload()

            def _remove():
                sid = _selected_sid()
                if not sid:
                    messagebox.showwarning('אזהרה', 'בחר תלמיד להסרה')
                    return
                try:
                    self.db.remove_cashier_responsible(int(sid))
                except Exception as e:
                    messagebox.showerror('שגיאה', str(e))
                    return
                _reload()

            btns = tk.Frame(dlg, bg='#ecf0f1')
            btns.pack(fill=tk.X, padx=12, pady=(0, 12))
            tk.Button(btns, text='➕ הוסף אחראי', command=_add, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=14, pady=5).pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text='🗑 הסר אחראי', command=_remove, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=14, pady=5).pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text='✖ סגור', command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=14, pady=5).pack(side=tk.RIGHT, padx=6)

            try:
                dlg.minsize(620, 520)
            except Exception:
                pass
            _reload()
            dlg.wait_window()

        responsibles_cb['fn'] = _open_responsibles_manager

        rbtns = tk.Frame(rf, bg='#ecf0f1')
        rbtns.pack(pady=(0, 10))
        tk.Button(rbtns, text='➕ הוסף אחראי', command=add_responsible, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=14, pady=5).pack(side=tk.LEFT, padx=6)
        tk.Button(rbtns, text='🗑 הסר אחראי', command=remove_responsible, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=14, pady=5).pack(side=tk.LEFT, padx=6)

        def _save_settings():
            try:
                self.db.set_cashier_mode(mode_var.get())
            except Exception:
                pass
            try:
                self.db.set_cashier_idle_timeout_sec(int(str(idle_var.get() or '300').strip()))
            except Exception:
                pass
            try:
                self.db.set_cashier_bw_logo_path(bw_logo_var.get().strip())
            except Exception:
                pass
            try:
                self.db.set_cashier_payment_confirm_mode(confirm_mode_var.get())
            except Exception:
                pass
            try:
                self.db.set_cashier_payment_confirm_threshold(int(str(confirm_threshold_var.get() or '0').strip() or '0'))
            except Exception:
                pass
            try:
                self.db.set_cashier_print_item_receipts(bool(int(print_items_var.get() or 0)))
            except Exception:
                pass
            try:
                self.db.set_cashier_receipt_footer_text(footer_text_var.get().strip())
            except Exception:
                pass
            messagebox.showinfo('נשמר', 'הגדרות קופה נשמרו')

        cashier_save_cb['fn'] = _save_settings

        sbtns = tk.Frame(settings_tab, bg='#ecf0f1')
        sbtns.pack(side=tk.BOTTOM, fill=tk.X, padx=12, pady=(0, 10))
        tk.Button(sbtns, text='💾 שמור הגדרות', command=_save_settings, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=18, pady=6).pack(side=tk.LEFT, padx=6)

        try:
            settings_container.pack_forget()
        except Exception:
            pass
        settings_container.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        _load_responsibles()

        _load_rows()

        try:
            tree.bind('<Double-1>', lambda _e: edit_product())
        except Exception:
            pass

        return

    def _open_public_closure_edit_dialog(self, parent, existing_row: dict = None) -> int:
        existing_row = existing_row or {}
        row_id = int(existing_row.get('id', 0) or 0)

        default_portrait = r"C:\מיצד\SchoolPoints\תמונות\shabat1.png"
        default_landscape = r"C:\מיצד\SchoolPoints\תמונות\shabat2.png"

        dlg = tk.Toplevel(parent)
        dlg.title("עריכת חסימה" if row_id else "הוספת חסימה")
        dlg.configure(bg='#ecf0f1')
        dlg.transient(parent)
        dlg.grab_set()

        result = {'id': 0}

        title_var = tk.StringVar(value=str(existing_row.get('title') or ''))
        subtitle_var = tk.StringVar(value=str(existing_row.get('subtitle') or ''))
        enabled_var = tk.IntVar(value=int(existing_row.get('enabled', 1) or 0))
        weekly_var = tk.IntVar(value=int(existing_row.get('repeat_weekly', 0) or 0))

        start_at_var = tk.StringVar(value=str(existing_row.get('start_at') or ''))
        end_at_var = tk.StringVar(value=str(existing_row.get('end_at') or ''))
        sd_var = tk.StringVar(value=str(existing_row.get('weekly_start_day') or ''))
        st_var = tk.StringVar(value=str(existing_row.get('weekly_start_time') or ''))
        ed_var = tk.StringVar(value=str(existing_row.get('weekly_end_day') or ''))
        et_var = tk.StringVar(value=str(existing_row.get('weekly_end_time') or ''))

        portrait_var = tk.StringVar(value=str(existing_row.get('image_path_portrait') or ''))
        landscape_var = tk.StringVar(value=str(existing_row.get('image_path_landscape') or ''))

        body = tk.Frame(dlg, bg='#ecf0f1')
        body.pack(fill=tk.BOTH, expand=True, padx=18, pady=16)

        body.grid_columnconfigure(0, weight=0)
        body.grid_columnconfigure(1, weight=1)
        body.grid_columnconfigure(2, weight=0)

        row_idx = 0

        try:
            cfg = self.load_app_config() or {}
        except Exception:
            cfg = {}
        orientation_code = str(cfg.get('screen_orientation', 'landscape') or 'landscape')
        orientation_label = 'רוחב' if orientation_code == 'landscape' else 'אורך'
        hint_text = fix_rtl_text(f"שים לב: כיוון המסך בעמדה הציבורית מוגדר כרגע {orientation_label}. מומלץ לבחור תמונה מתאימה.")
        tk.Label(body, text=hint_text, bg='#ecf0f1', fg='#7f8c8d', font=('Arial', 9)).grid(row=row_idx, column=0, columnspan=3, sticky='e', pady=(0, 10))
        row_idx += 1

        tk.Label(body, text=fix_rtl_text("כותרת:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        title_entry = tk.Entry(body, textvariable=title_var, font=('Arial', 11), justify='right')
        title_entry.grid(row=row_idx, column=1, sticky='ew', pady=4)
        row_idx += 1

        tk.Label(body, text=fix_rtl_text("טקסט משנה:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        subtitle_entry = tk.Entry(body, textvariable=subtitle_var, font=('Arial', 11), justify='right')
        subtitle_entry.grid(row=row_idx, column=1, sticky='ew', pady=4)
        row_idx += 1

        flags = tk.Frame(body, bg='#ecf0f1')
        flags.grid(row=row_idx, column=0, columnspan=3, sticky='e', pady=(6, 10))
        tk.Checkbutton(flags, text=fix_rtl_text("פעיל"), variable=enabled_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=10)
        tk.Checkbutton(flags, text=fix_rtl_text("חסימה שבועית"), variable=weekly_var, bg='#ecf0f1').pack(side=tk.RIGHT, padx=10)
        row_idx += 1

        # one-time window
        tk.Label(body, text=fix_rtl_text("תאריך/שעה התחלה:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        start_at_entry = tk.Entry(body, textvariable=start_at_var, font=('Arial', 11), justify='right')
        start_at_entry.grid(row=row_idx, column=1, sticky='ew', pady=4)
        row_idx += 1

        tk.Label(body, text=fix_rtl_text("תאריך/שעה סיום:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        end_at_entry = tk.Entry(body, textvariable=end_at_var, font=('Arial', 11), justify='right')
        end_at_entry.grid(row=row_idx, column=1, sticky='ew', pady=4)
        row_idx += 1

        # weekly window
        tk.Label(body, text=fix_rtl_text("שבועי – התחלה (יום/שעה):"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        weekly_start_frame = tk.Frame(body, bg='#ecf0f1')
        weekly_start_frame.grid(row=row_idx, column=1, sticky='e', pady=4)
        sd_entry = tk.Entry(weekly_start_frame, textvariable=sd_var, font=('Arial', 11), width=4, justify='right')
        sd_entry.pack(side=tk.RIGHT, padx=(0, 6))
        st_entry = tk.Entry(weekly_start_frame, textvariable=st_var, font=('Arial', 11), width=8, justify='right')
        st_entry.pack(side=tk.RIGHT)
        row_idx += 1

        tk.Label(body, text=fix_rtl_text("שבועי – סיום (יום/שעה):"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=4)
        weekly_end_frame = tk.Frame(body, bg='#ecf0f1')
        weekly_end_frame.grid(row=row_idx, column=1, sticky='e', pady=4)
        ed_entry = tk.Entry(weekly_end_frame, textvariable=ed_var, font=('Arial', 11), width=4, justify='right')
        ed_entry.pack(side=tk.RIGHT, padx=(0, 6))
        et_entry = tk.Entry(weekly_end_frame, textvariable=et_var, font=('Arial', 11), width=8, justify='right')
        et_entry.pack(side=tk.RIGHT)
        row_idx += 1

        def _browse_image(target_var: tk.StringVar, title: str):
            try:
                initial = os.path.dirname(target_var.get()) if target_var.get() else self.base_dir
            except Exception:
                initial = self.base_dir
            path = filedialog.askopenfilename(
                title=title,
                filetypes=[("קבצי תמונה", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("הכל", "*.*")],
                initialdir=initial
            )
            if path:
                target_var.set(path)

        tk.Label(body, text=fix_rtl_text("תמונה לאורך:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=6)
        p_frame = tk.Frame(body, bg='#ecf0f1')
        p_frame.grid(row=row_idx, column=1, sticky='ew', pady=6)
        p_frame.grid_columnconfigure(0, weight=1)
        p_entry = tk.Entry(p_frame, textvariable=portrait_var, font=('Arial', 10), justify='right')
        p_entry.grid(row=0, column=0, sticky='ew')
        tk.Button(p_frame, text="בחר...", command=lambda: _browse_image(portrait_var, "בחר תמונה לאורך"), font=('Arial', 9), bg='#3498db', fg='white', padx=10, pady=3).grid(row=0, column=1, padx=(8, 0))
        row_idx += 1

        tk.Label(body, text=fix_rtl_text("תמונה לרוחב:"), bg='#ecf0f1', font=('Arial', 10, 'bold')).grid(row=row_idx, column=2, sticky='e', padx=(8, 0), pady=6)
        l_frame = tk.Frame(body, bg='#ecf0f1')
        l_frame.grid(row=row_idx, column=1, sticky='ew', pady=6)
        l_frame.grid_columnconfigure(0, weight=1)
        l_entry = tk.Entry(l_frame, textvariable=landscape_var, font=('Arial', 10), justify='right')
        l_entry.grid(row=0, column=0, sticky='ew')
        tk.Button(l_frame, text="בחר...", command=lambda: _browse_image(landscape_var, "בחר תמונה לרוחב"), font=('Arial', 9), bg='#3498db', fg='white', padx=10, pady=3).grid(row=0, column=1, padx=(8, 0))
        row_idx += 1

        def _sync_mode(*_):
            is_weekly = 1 if int(weekly_var.get() or 0) == 1 else 0
            try:
                start_at_entry.config(state='disabled' if is_weekly else 'normal')
                end_at_entry.config(state='disabled' if is_weekly else 'normal')
                sd_entry.config(state='normal' if is_weekly else 'disabled')
                st_entry.config(state='normal' if is_weekly else 'disabled')
                ed_entry.config(state='normal' if is_weekly else 'disabled')
                et_entry.config(state='normal' if is_weekly else 'disabled')
            except Exception:
                pass

        try:
            weekly_var.trace_add('write', _sync_mode)
        except Exception:
            pass
        _sync_mode()

        def save_it():
            title = title_var.get().strip()
            if not title:
                messagebox.showerror("שגיאה", "יש להזין כותרת")
                return

            enabled = 1 if int(enabled_var.get() or 0) == 1 else 0
            repeat_weekly = 1 if int(weekly_var.get() or 0) == 1 else 0

            start_at = start_at_var.get().strip() or '2000-01-01 00:00:00'
            end_at = end_at_var.get().strip() or '2000-01-01 00:00:00'

            portrait = portrait_var.get().strip() or default_portrait
            landscape = landscape_var.get().strip() or default_landscape

            try:
                if row_id:
                    ok = self.db.update_public_closure(
                        row_id,
                        title=title,
                        subtitle=subtitle_var.get().strip() or None,
                        start_at=start_at,
                        end_at=end_at,
                        enabled=enabled,
                        repeat_weekly=repeat_weekly,
                        weekly_start_day=sd_var.get().strip() or None,
                        weekly_start_time=st_var.get().strip() or None,
                        weekly_end_day=ed_var.get().strip() or None,
                        weekly_end_time=et_var.get().strip() or None,
                        image_path_portrait=portrait,
                        image_path_landscape=landscape,
                    )
                    if ok:
                        result['id'] = row_id
                        dlg.destroy()
                    else:
                        messagebox.showerror("שגיאה", "לא ניתן לעדכן")
                else:
                    new_id = self.db.add_public_closure(
                        title=title,
                        subtitle=subtitle_var.get().strip() or None,
                        start_at=start_at,
                        end_at=end_at,
                        enabled=enabled,
                        repeat_weekly=repeat_weekly,
                        weekly_start_day=sd_var.get().strip() or None,
                        weekly_start_time=st_var.get().strip() or None,
                        weekly_end_day=ed_var.get().strip() or None,
                        weekly_end_time=et_var.get().strip() or None,
                        image_path_portrait=portrait,
                        image_path_landscape=landscape,
                    )
                    result['id'] = int(new_id or 0)
                    dlg.destroy()
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))

        btns = tk.Frame(body, bg='#ecf0f1')
        btns.grid(row=row_idx, column=0, columnspan=3, sticky='e', pady=(12, 0))
        tk.Button(btns, text="💾 שמור", command=save_it, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=18, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="✖ ביטול", command=dlg.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=18, pady=6).pack(side=tk.LEFT, padx=6)

        dlg.wait_window()
        return int(result['id'] or 0)

    def open_anti_spam_manager(self):
        """פתיחת חלון ניהול חסימות אנטי-ספאם"""
        if not (self.current_teacher and self.current_teacher.get('is_admin') == 1):
            messagebox.showwarning("אין הרשאה", "גישה רק למנהלים")
            return
        
        try:
            from anti_spam_dialog import open_anti_spam_dialog
            open_anti_spam_dialog(self.root, self.load_app_config, self.save_app_config)
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לפתוח חלון ניהול חסימות:\n{e}")

    def open_public_closures_manager(self):
        if not (self.current_teacher and self.current_teacher.get('is_admin') == 1):
            messagebox.showwarning("אין הרשאה", "גישה רק למנהלים")
            return

        # יצירת ברירת מחדל לשבת (אם לא קיימת)
        try:
            rows0 = self.db.get_all_public_closures() or []
        except Exception:
            rows0 = []
        try:
            has_shabbat = any(
                str(x.get('title') or '').strip() == 'שבת' and int(x.get('repeat_weekly', 0) or 0) == 1
                for x in (rows0 or [])
            )
        except Exception:
            has_shabbat = False
        if not has_shabbat:
            try:
                self.db.add_public_closure(
                    title='שבת',
                    subtitle='חסימה שבועית',
                    start_at='2000-01-01 00:00:00',
                    end_at='2000-01-01 00:00:00',
                    enabled=1,
                    repeat_weekly=1,
                    weekly_start_day='ו',
                    weekly_start_time='13:15',
                    weekly_end_day='ש',
                    weekly_end_time='23:30',
                    image_path_portrait=r"C:\מיצד\SchoolPoints\תמונות\shabat1.png",
                    image_path_landscape=r"C:\מיצד\SchoolPoints\תמונות\shabat2.png",
                )
            except Exception:
                pass

        dialog = tk.Toplevel(self.root)
        dialog.title("🗓 לוח חופשות / חסימות (עמדה ציבורית)")
        dialog.configure(bg='#ecf0f1')
        try:
            dialog.minsize(1060, 560)
        except Exception:
            pass
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(
            dialog,
            text=fix_rtl_text("לוח חופשות/חסימות – חוסם את העמדה הציבורית בלבד"),
            font=('Arial', 14, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=(14, 8))

        list_frame = tk.Frame(dialog, bg='#ecf0f1')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=18, pady=10)

        columns = ('enabled', 'kind', 'img', 'title', 'window')
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15, selectmode='extended')
        tree.heading('enabled', text='פעיל')
        tree.heading('kind', text='סוג')
        tree.heading('img', text='תמונה')
        tree.heading('title', text='כותרת')
        tree.heading('window', text='טווח')

        tree.column('enabled', width=60, anchor='center')
        tree.column('kind', width=90, anchor='e')
        tree.column('img', width=70, anchor='center')
        tree.column('title', width=260, anchor='e')
        tree.column('window', width=420, anchor='e')

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=sb.set)

        def _format_window(row: dict) -> str:
            try:
                if int(row.get('repeat_weekly', 0) or 0) == 1:
                    sd = str(row.get('weekly_start_day') or '').strip()
                    st = str(row.get('weekly_start_time') or '').strip()
                    ed = str(row.get('weekly_end_day') or '').strip()
                    et = str(row.get('weekly_end_time') or '').strip()
                    return f"{sd} {st} עד {ed} {et}".strip()
                s = str(row.get('start_at') or '').replace('T', ' ').strip()
                e = str(row.get('end_at') or '').replace('T', ' ').strip()
                return f"{s} עד {e}".strip()
            except Exception:
                return ''

        def _get_selected_id() -> int:
            sel = tree.selection()
            if not sel:
                return 0
            tags = tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('row:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _get_selected_ids() -> list:
            ids = []
            sel = tree.selection() or ()
            for iid in sel:
                try:
                    tags = tree.item(iid).get('tags') or ()
                except Exception:
                    tags = ()
                for t in tags:
                    if isinstance(t, str) and t.startswith('row:'):
                        try:
                            rid = int(t.split(':', 1)[1])
                        except Exception:
                            rid = 0
                        if rid and rid not in ids:
                            ids.append(rid)
            return ids

        def _get_row_by_id(row_id: int) -> dict:
            try:
                rows = self.db.get_all_public_closures() or []
            except Exception:
                rows = []
            for r in rows:
                try:
                    if int(r.get('id', 0) or 0) == int(row_id):
                        return r
                except Exception:
                    continue
            return {}

        def load_rows(select_id: int = None):
            try:
                rows = self.db.get_all_public_closures() or []
            except Exception:
                rows = []
            tree.delete(*tree.get_children())
            item_to_select = None
            for r in rows:
                rid = int(r.get('id', 0) or 0)
                enabled_txt = '✓' if int(r.get('enabled', 0) or 0) == 1 else '✗'
                kind = 'שבועי' if int(r.get('repeat_weekly', 0) or 0) == 1 else 'חד-פעמי'
                try:
                    has_img = bool(str(r.get('image_path_portrait') or '').strip() or str(r.get('image_path_landscape') or '').strip())
                except Exception:
                    has_img = False
                img_txt = '✓' if has_img else '✗'
                title = str(r.get('title') or '').strip()
                win = _format_window(r)
                iid = tree.insert('', tk.END, values=(enabled_txt, kind, img_txt, title, win), tags=(f"row:{rid}",))
                if select_id and rid == int(select_id):
                    item_to_select = iid
            if item_to_select:
                try:
                    tree.selection_set(item_to_select)
                    tree.see(item_to_select)
                except Exception:
                    pass

        def add_row():
            new_id = self._open_public_closure_edit_dialog(dialog)
            if new_id:
                load_rows(select_id=new_id)

        def edit_row():
            rid = _get_selected_id()
            if not rid:
                messagebox.showwarning("אזהרה", "בחר חסימה לעריכה")
                return
            r = _get_row_by_id(rid)
            out_id = self._open_public_closure_edit_dialog(dialog, existing_row=r)
            if out_id:
                load_rows(select_id=out_id)

        def toggle_enabled():
            rid = _get_selected_id()
            if not rid:
                messagebox.showwarning("אזהרה", "בחר חסימה")
                return
            r = _get_row_by_id(rid)
            if not r:
                return
            new_enabled = 0 if int(r.get('enabled', 0) or 0) == 1 else 1
            try:
                ok = self.db.update_public_closure(
                    rid,
                    title=str(r.get('title') or ''),
                    subtitle=r.get('subtitle'),
                    start_at=str(r.get('start_at') or ''),
                    end_at=str(r.get('end_at') or ''),
                    enabled=new_enabled,
                    repeat_weekly=int(r.get('repeat_weekly', 0) or 0),
                    weekly_start_day=r.get('weekly_start_day'),
                    weekly_start_time=r.get('weekly_start_time'),
                    weekly_end_day=r.get('weekly_end_day'),
                    weekly_end_time=r.get('weekly_end_time'),
                    image_path_portrait=r.get('image_path_portrait'),
                    image_path_landscape=r.get('image_path_landscape'),
                )
            except Exception:
                ok = False
            if ok:
                load_rows(select_id=rid)

        def delete_row():
            ids = _get_selected_ids()
            if not ids:
                messagebox.showwarning("אזהרה", "בחר חסימה למחיקה")
                return
            msg = "למחוק את החסימה שנבחרה?" if len(ids) == 1 else f"למחוק {len(ids)} חסימות שנבחרו?"
            if not messagebox.askyesno("מחיקה", msg):
                return

            deleted_any = False
            for rid in ids:
                try:
                    self.db.delete_public_closure(int(rid))
                    deleted_any = True
                except Exception:
                    pass
            if deleted_any:
                load_rows()

        def _on_delete_key(event=None):
            try:
                delete_row()
            except Exception:
                pass
            return "break"

        def seed_holidays():
            try:
                created = int(self.db.seed_public_holidays_template(days_ahead=450, israel=True) or 0)
                if created > 0:
                    messagebox.showinfo("תבנית חגים", f"נוצרה תבנית חגים (מושבתת) עם {created} רשומות.")
                else:
                    messagebox.showinfo("תבנית חגים", "לא נוצרו רשומות חדשות (ייתכן שכבר קיימת תבנית).")
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))
            load_rows()

        btns = tk.Frame(dialog, bg='#ecf0f1')
        btns.pack(pady=(0, 16))
        tk.Button(btns, text="➕ הוסף", command=add_row, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="✏ ערוך", command=edit_row, font=('Arial', 10, 'bold'), bg='#3498db', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="✅ פעיל/כבוי", command=toggle_enabled, font=('Arial', 10, 'bold'), bg='#16a085', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="🗑 מחק", command=delete_row, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="📅 צור תבנית חגים", command=seed_holidays, font=('Arial', 10, 'bold'), bg='#8e44ad', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="✖ סגור", command=dialog.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=16, pady=6).pack(side=tk.LEFT, padx=6)

        load_rows()

        try:
            tree.focus_set()
        except Exception:
            pass

        try:
            tree.bind('<Delete>', _on_delete_key)
        except Exception:
            pass

        try:
            dialog.bind('<Delete>', _on_delete_key)
        except Exception:
            pass

        return

        config = self.load_app_config()

        dialog = tk.Toplevel(self.root)
        dialog.title("🖥 הגדרות תצוגה")
        dialog.configure(bg='#ecf0f1')
        dialog.minsize(720, 0)
        dialog.transient(self.root)
        dialog.grab_set()

        content_frame = tk.Frame(dialog, bg='#ecf0f1')
        content_frame.pack(fill=tk.X, padx=40, pady=10)

        # ערכת צבעים לעמדה הציבורית – שני מצבים בלבד (כהה / בהיר)
        theme_map = {
            "רקע כהה, כיתוב בהיר": "dark",
            "רקע בהיר, כיתוב כהה": "light",
        }

        # מצב רקע לעמדה הציבורית
        # "ברירת מחדל (דף רקע אוטומטי)" ישתמש בתבנית הגרפית המובנית (template1)
        background_mode_map = {
            "ברירת מחדל (דף רקע אוטומטי)": "default",
            "תמונת רקע אחת": "image",
            "מצגת תמונות (תיקייה)": "slideshow",
            "צבע רקע אחיד (טבלת צבעים)": "color",
        }

        # מצב פריסת תמונת הרקע (בדומה לשולחן עבודה)
        background_layout_map = {
            "מילוי": "cover",      # ממלא את המסך (חיתוך בשוליים)
            "התאמה": "contain",    # מתאים למסך עם שוליים אם צריך
            "מתח": "stretch",      # מתיחה מדויקת לגודל המסך
            "אריח": "tile",        # חזרה על התמונה
            "מרכז": "center",      # במרכז המסך
            "פרישה": "cover",      # שם נוסף ל"מילוי"
        }

        # כיוון מסך בעמדה הציבורית
        orientation_map = {
            "מסך לרוחב (אופקי)": "landscape",
            "מסך לאורך (אנכי)": "portrait",
        }

        raw_theme = config.get('theme', 'dark')
        current_bg_mode_code = config.get('background_mode', 'default') or 'default'
        current_theme_code = raw_theme

        # תאימות לאחור: אם theme נשמר כ-background_image / slideshow
        if raw_theme in ('background_image', 'slideshow') and current_bg_mode_code in ('none', '', None):
            # בעבר ערך זה סימל "רקע תמונה" / "מצגת" – היום נתייחס לכך כאל ערכת צבעים כהה
            current_theme_code = 'dark'
            current_bg_mode_code = 'image' if raw_theme == 'background_image' else 'slideshow'

        # תאימות לאחור: ערכת "ניגודיות גבוהה" מתורגמת לערכת כהה
        if current_theme_code in ('high_contrast', 'background_image', 'slideshow', 'none'):
            current_theme_code = 'dark'

        reverse_theme_map = {v: k for k, v in theme_map.items()}
        theme_var = tk.StringVar(value=reverse_theme_map.get(current_theme_code, "רקע כהה, כיתוב בהיר"))

        # תאימות לאחור: מצב 'none' הישן יתנהג כעת כברירת מחדל (דף רקע אוטומטי)
        if current_bg_mode_code in ('none', '', None):
            current_bg_mode_code = 'default'

        reverse_bg_mode_map = {v: k for k, v in background_mode_map.items()}
        bg_mode_var = tk.StringVar(value=reverse_bg_mode_map.get(current_bg_mode_code, "ברירת מחדל (דף רקע אוטומטי)"))

        current_bg_layout_code = config.get('background_layout', 'cover')
        reverse_bg_layout_map = {v: k for k, v in background_layout_map.items()}
        bg_layout_var = tk.StringVar(value=reverse_bg_layout_map.get(current_bg_layout_code, "מילוי"))

        background_image_var = tk.StringVar(value=config.get('background_image_path', ""))
        background_folder_var = tk.StringVar(value=config.get('background_folder', ""))
        background_color_var = tk.StringVar(value=config.get('background_color', "#000000"))

        slide_interval_var = tk.StringVar(value=str(config.get('background_interval_sec', 15)))

        slideshow_mode_map = {
            "תמונה אחת מתחלפת": "single",
            "מונטאז' ריבועים סטטי": "grid_static",
            "מונטאז' ריבועים מתחלפים": "grid_dynamic",
        }
        current_slideshow_mode = config.get('slideshow_display_mode', 'single')
        reverse_slideshow_mode_map = {v: k for k, v in slideshow_mode_map.items()}
        slideshow_mode_var = tk.StringVar(value=reverse_slideshow_mode_map.get(current_slideshow_mode, "תמונה אחת מתחלפת"))

        slideshow_cols_var = tk.StringVar(value=str(config.get('slideshow_grid_cols', 4)))

        panel_style_map = {
            "פנלים רגילים (מלאים)": "solid",
            "פנלים שקופים חלקית": "floating",
        }
        current_panel_style = config.get('panel_style', 'floating')
        reverse_panel_style_map = {v: k for k, v in panel_style_map.items()}
        panel_style_var = tk.StringVar(value=reverse_panel_style_map.get(current_panel_style, "פנלים רגילים (מלאים)"))

        # כיוון מסך נוכחי
        current_orientation_code = config.get('screen_orientation', 'landscape')
        reverse_orientation_map = {v: k for k, v in orientation_map.items()}
        orientation_var = tk.StringVar(value=reverse_orientation_map.get(current_orientation_code, "מסך לרוחב (אופקי)"))

        FIELD_WIDTH = 42
        BUTTON_WIDTH = 14
        LABEL_WIDTH = 40

        # שורה 1 - כיוון מסך
        row1 = tk.Frame(content_frame, bg='#ecf0f1')
        row1.pack(fill=tk.X, pady=6)
        tk.Label(row1, text=fix_rtl_text("כיוון מסך בעמדה הציבורית:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        orientation_choices = list(orientation_map.keys())
        orientation_menu = tk.OptionMenu(row1, orientation_var, *orientation_choices)
        orientation_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        orientation_menu.pack(side=tk.RIGHT, padx=5)

        # שורה 2 - ערכת צבעים לעמדה הציבורית
        row2 = tk.Frame(content_frame, bg='#ecf0f1')
        row2.pack(fill=tk.X, pady=6)
        tk.Label(row2, text=fix_rtl_text("ערכת צבעים לעמדה הציבורית:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        theme_choices = list(theme_map.keys())
        theme_menu = tk.OptionMenu(row2, theme_var, *theme_choices)
        theme_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        theme_menu.pack(side=tk.RIGHT, padx=5)

        # שורה 3 - מצב רקע
        row3b = tk.Frame(content_frame, bg='#ecf0f1')
        row3b.pack(fill=tk.X, pady=6)
        tk.Label(row3b, text=fix_rtl_text("מצב רקע:"), font=('Arial', 10, 'bold'), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_mode_choices = list(background_mode_map.keys())
        bg_mode_menu = tk.OptionMenu(row3b, bg_mode_var, *bg_mode_choices)
        bg_mode_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        bg_mode_menu.pack(side=tk.RIGHT, padx=5)

        # מסגרת לשורות הרקע הדינמיות
        bg_rows_frame = tk.Frame(content_frame, bg='#ecf0f1')

        # פריסת רקע
        row4 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row4, text=fix_rtl_text("פריסת רקע:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_layout_choices = list(background_layout_map.keys())
        bg_layout_menu = tk.OptionMenu(row4, bg_layout_var, *bg_layout_choices)
        bg_layout_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        bg_layout_menu.pack(side=tk.RIGHT, padx=5)

        # תמונת רקע בודדת
        row5 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row5, text=fix_rtl_text("תמונת רקע (למצב 'תמונת רקע אחת'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_image_entry = tk.Entry(row5, textvariable=background_image_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_image_entry.pack(side=tk.RIGHT, padx=5)

        def browse_bg_image():
            initial = os.path.dirname(background_image_var.get()) if background_image_var.get() else self.base_dir
            path = filedialog.askopenfilename(
                title="בחר תמונת רקע",
                filetypes=[("קבצי תמונה", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("הכל", "*.*")],
                initialdir=initial
            )
            if path:
                background_image_var.set(path)

        tk.Button(row5, text="בחר קובץ", command=browse_bg_image, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # תיקיית מצגת
        row6 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row6, text=fix_rtl_text("תיקיית תמונות (למצב 'מצגת תמונות'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_folder_entry = tk.Entry(row6, textvariable=background_folder_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_folder_entry.pack(side=tk.RIGHT, padx=5)

        def browse_bg_folder():
            initial = background_folder_var.get().strip() or self.base_dir
            path = filedialog.askdirectory(title="בחר תיקיית תמונות למצגת", initialdir=initial)
            if path:
                background_folder_var.set(path)

        tk.Button(row6, text="בחר תיקייה", command=browse_bg_folder, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # צבע רקע אחיד
        row6b = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row6b, text=fix_rtl_text("צבע רקע (למצב 'צבע אחיד'):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        bg_color_entry = tk.Entry(row6b, textvariable=background_color_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right')
        bg_color_entry.pack(side=tk.RIGHT, padx=5)

        def choose_bg_color():
            initial = background_color_var.get().strip() or "#000000"
            try:
                color = colorchooser.askcolor(color=initial, title="בחר צבע רקע אחיד")
            except Exception:
                color = None

            hex_color = None
            if color:
                if len(color) > 1 and color[1]:
                    hex_color = color[1]
                elif color[0] and isinstance(color[0], (tuple, list)):
                    try:
                        r, g, b = map(int, color[0])
                        hex_color = f"#{r:02x}{g:02x}{b:02x}"
                    except Exception:
                        hex_color = None

            if hex_color:
                background_color_var.set(hex_color)

        tk.Button(row6b, text="בחר צבע", command=choose_bg_color, font=('Arial', 10), bg='#3498db', fg='white', width=BUTTON_WIDTH, padx=4, pady=4).pack(side=tk.RIGHT, padx=5)

        # זמן מעבר במצגת
        row7 = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7, text=fix_rtl_text("זמן מעבר במצגת (שניות):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        tk.Entry(row7, textvariable=slide_interval_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right').pack(side=tk.RIGHT, padx=5)

        # מצב תצוגת מצגת
        row7b = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7b, text=fix_rtl_text("מצב תצוגת מצגת:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        slideshow_mode_choices = list(slideshow_mode_map.keys())
        slideshow_mode_menu = tk.OptionMenu(row7b, slideshow_mode_var, *slideshow_mode_choices)
        slideshow_mode_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        slideshow_mode_menu.pack(side=tk.RIGHT, padx=5)

        # מספר עמודות במונטאז'
        row7c = tk.Frame(bg_rows_frame, bg='#ecf0f1')
        tk.Label(row7c, text=fix_rtl_text("מספר עמודות למונטאז' (1-10):"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        tk.Entry(row7c, textvariable=slideshow_cols_var, font=('Arial', 10), width=FIELD_WIDTH, justify='right').pack(side=tk.RIGHT, padx=5)

        # סגנון פנלים
        row8 = tk.Frame(content_frame, bg='#ecf0f1')
        row8.pack(fill=tk.X, pady=10)
        tk.Label(row8, text=fix_rtl_text("סגנון פנלים מעל הרקע:"), font=('Arial', 10), bg='#ecf0f1', anchor='e', width=LABEL_WIDTH).pack(side=tk.RIGHT, padx=5)
        panel_style_choices = list(panel_style_map.keys())
        panel_style_menu = tk.OptionMenu(row8, panel_style_var, *panel_style_choices)
        panel_style_menu.config(font=('Arial', 10), width=FIELD_WIDTH)
        panel_style_menu.pack(side=tk.RIGHT, padx=5)

        # הגדרת הצגת סטטיסטיקות בעמדה הציבורית (ברירת מחדל: כבוי)
        show_stats_value = self.db.get_setting('show_statistics', '0')
        show_stats_var = tk.BooleanVar(value=show_stats_value == '1')

        stats_frame = tk.Frame(content_frame, bg='#ecf0f1')
        stats_frame.pack(fill=tk.X, pady=(5, 5))

        tk.Label(
            stats_frame,
            text=fix_rtl_text("הצגת סטטיסטיקות בעמדה הציבורית:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).pack(side=tk.RIGHT, padx=5)
        ToggleSwitch(
            stats_frame,
            variable=show_stats_var
        ).pack(side=tk.RIGHT, padx=5)

        # הגדרת הצגת תמונת תלמיד בעמדה הציבורית (ברירת מחדל: כבוי)
        show_photo_value = self.db.get_setting('show_student_photo', '0')
        show_photo_var = tk.BooleanVar(value=show_photo_value == '1')

        photo_frame = tk.Frame(content_frame, bg='#ecf0f1')
        photo_frame.pack(fill=tk.X, pady=(0, 5))

        tk.Label(
            photo_frame,
            text=fix_rtl_text("הצגת תמונת תלמיד בעמדה הציבורית:"),
            font=('Arial', 10),
            bg='#ecf0f1',
            anchor='e',
            width=LABEL_WIDTH
        ).pack(side=tk.RIGHT, padx=5)
        ToggleSwitch(
            photo_frame,
            variable=show_photo_var
        ).pack(side=tk.RIGHT, padx=5)

        def update_background_rows(*args):
            mode_code = background_mode_map.get(bg_mode_var.get(), 'none')

            bg_rows_frame.pack_forget()
            for row in (row4, row5, row6, row6b, row7, row7b, row7c):
                row.pack_forget()

            if mode_code == 'image':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row5.pack(fill=tk.X, pady=6)
                row4.pack(fill=tk.X, pady=6)
            elif mode_code == 'slideshow':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row6.pack(fill=tk.X, pady=6)
                row4.pack(fill=tk.X, pady=6)
                row7.pack(fill=tk.X, pady=6)
                row7b.pack(fill=tk.X, pady=6)
                row7c.pack(fill=tk.X, pady=6)
            elif mode_code == 'color':
                if not bg_rows_frame.winfo_ismapped():
                    bg_rows_frame.pack(fill=tk.X, pady=0)
                row6b.pack(fill=tk.X, pady=6)

            try:
                dialog.update_idletasks()
                dialog.geometry("")
            except Exception:
                pass

        update_background_rows()
        bg_mode_var.trace_add('write', lambda *args: update_background_rows())

        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=20)

        def save_settings():
            cfg = self.load_app_config() or {}

            # כיוון מסך
            selected_orientation_label = orientation_var.get()
            cfg['screen_orientation'] = orientation_map.get(selected_orientation_label, 'landscape')

            # ערכת צבעים לעמדה הציבורית
            selected_theme_label = theme_var.get()
            cfg['theme'] = theme_map.get(selected_theme_label, 'dark')

            # מצב רקע
            selected_bg_label = bg_mode_var.get()
            bg_mode_code = background_mode_map.get(selected_bg_label, 'default')
            cfg['background_mode'] = bg_mode_code

            # קביעת דף הרקע הגרפי בהתאם למצב הרקע
            # ברירת מחדל – שימוש בתבנית הגרפית template1 עם דפי רקע אוטומטיים
            cfg['background_template'] = 'template1'

            # תמונת רקע
            bg_image = background_image_var.get().strip()
            if bg_image:
                cfg['background_image_path'] = bg_image
            else:
                cfg.pop('background_image_path', None)

            # תיקיית מצגת
            bg_folder = background_folder_var.get().strip()
            if bg_folder:
                cfg['background_folder'] = bg_folder
            else:
                cfg.pop('background_folder', None)

            # צבע רקע אחיד
            bg_color = background_color_var.get().strip()
            if bg_color:
                cfg['background_color'] = bg_color
            else:
                cfg.pop('background_color', None)

            # זמן מעבר במצגת
            interval_str = slide_interval_var.get().strip()
            try:
                interval_val = int(interval_str)
                if interval_val < 3:
                    interval_val = 3
                if interval_val > 600:
                    interval_val = 600
                cfg['background_interval_sec'] = interval_val
            except ValueError:
                cfg.pop('background_interval_sec', None)

            # מצב תצוגת מצגת
            selected_slideshow_label = slideshow_mode_var.get()
            cfg['slideshow_display_mode'] = slideshow_mode_map.get(selected_slideshow_label, 'single')

            # מספר עמודות במונטאז'
            cols_str = slideshow_cols_var.get().strip()
            try:
                cols_val = int(cols_str)
                if cols_val < 1:
                    cols_val = 1
                if cols_val > 10:
                    cols_val = 10
                cfg['slideshow_grid_cols'] = cols_val
            except ValueError:
                cfg.pop('slideshow_grid_cols', None)

            # פריסת רקע
            selected_layout_label = bg_layout_var.get()
            cfg['background_layout'] = background_layout_map.get(selected_layout_label, 'cover')

            # סגנון פנלים
            selected_panel_label = panel_style_var.get()
            cfg['panel_style'] = panel_style_map.get(selected_panel_label, 'solid')

            # שמירת הגדרת הצגת סטטיסטיקות בעמדה הציבורית
            try:
                show = '1' if show_stats_var.get() else '0'
                self.db.set_setting('show_statistics', show)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לעדכן הגדרת סטטיסטיקות:\n{e}")
                return

            # שמירת הגדרת הצגת תמונת תלמיד בעמדה הציבורית
            try:
                show_photo = '1' if show_photo_var.get() else '0'
                self.db.set_setting('show_student_photo', show_photo)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לעדכן הגדרת תמונת תלמיד:\n{e}")
                return

            if self.save_app_config(cfg):
                # שליחת הפעלה מחדש אוטומטית לכל העמדות הציבוריות
                try:
                    restart_cfg = self.load_app_config() or {}
                except Exception:
                    restart_cfg = {}
                if isinstance(restart_cfg, dict):
                    old_token = restart_cfg.get('restart_public_stations_token')
                    try:
                        token_val = int(old_token)
                        token_val += 1
                    except Exception:
                        token_val = 1
                    restart_cfg['restart_public_stations_token'] = str(token_val)
                    self.save_app_config(restart_cfg)

                messagebox.showinfo(
                    "הפעלה מחדש נשלחה",
                    "הגדרות התצוגה נשמרו בהצלחה.\n\n"
                    "המערכת שלחה הוראה להפעיל מחדש את כל העמדות הציבוריות.\n"
                    "שינויים בלוגו וברקע הגרפי יופיעו בתוך מספר שניות."
                )
                dialog.destroy()

        tk.Button(
            btn_frame,
            text="✖ סגור",
            command=dialog.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=6
        ).pack(side=tk.LEFT, padx=8)

        tk.Button(
            btn_frame,
            text="💾 שמור",
            command=save_settings,
            font=('Arial', 10, 'bold'),
            bg='#27ae60',
            fg='white',
            width=BUTTON_WIDTH,
            padx=4,
            pady=6
        ).pack(side=tk.LEFT, padx=8)

    def open_teacher_bonus_dialog(self):
        """פתיחת חלון הגדרת בונוס מורה (למורים רגילים בלבד)"""
        # בונוס מורה זמין רק למורה מחובר שאינו מנהל
        if not self.current_teacher or self.current_teacher.get('is_admin') == 1:
            messagebox.showwarning("אזהרה", "בונוס מורה זמין רק למורים רגילים.")
            return

        if not self.ensure_can_modify():
            return

        teacher_id = self.current_teacher['id']
        try:
            current_bonus = self.db.get_teacher_bonus(teacher_id)
        except Exception:
            current_bonus = 0

        # אם כבר מוגדר בונוס חיובי למורה – לחיצה על הכפתור מבטלת אותו (toggle off)
        if current_bonus and current_bonus > 0:
            try:
                self.db.set_teacher_bonus(teacher_id, 0)
            except Exception as e:
                messagebox.showerror("שגיאה", f"לא ניתן לבטל בונוס למורה:\n{e}")
                return

            messagebox.showinfo("הצלחה", "בונוס המורה בוטל (0 נקודות).")
            if hasattr(self, 'teacher_bonus_btn'):
                self.teacher_bonus_btn.config(text="🎁 בונוס", bg='#f39c12')
            return

        teacher_limit_val = None
        try:
            teacher_row = self.db.get_teacher_by_id(teacher_id)
        except Exception:
            teacher_row = self.current_teacher or {}
        if teacher_row:
            try:
                raw_limit = teacher_row.get('bonus_max_points_per_student')
                if raw_limit is not None:
                    teacher_limit_val = int(raw_limit)
            except Exception:
                teacher_limit_val = teacher_limit_val

        effective_max = None
        if teacher_limit_val is not None and teacher_limit_val > 0:
            effective_max = teacher_limit_val

        dialog = tk.Toplevel(self.root)
        dialog.title("הגדרת בונוס למורה")
        dialog.geometry("460x300")
        try:
            dialog.minsize(460, 300)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()

        lines = ["כמה נקודות בונוס יקבל כל תלמיד בסבב? "]
        if teacher_limit_val is not None and teacher_limit_val > 0:
            lines.append(f"(מקסימום למורה זה: {teacher_limit_val} נקודות)")
        else:
            lines.append("(אין מגבלה מוגדרת למורה זה.)")
        label_text = "\n".join(lines)

        tk.Label(
            dialog,
            text=label_text,
            font=('Arial', 11, 'bold'),
            bg='#ecf0f1',
            justify='right'
        ).pack(pady=15, padx=10)

        entry = tk.Entry(dialog, font=('Arial', 14), width=10, justify='center')
        if current_bonus and current_bonus > 0:
            try:
                entry.insert(0, str(int(current_bonus)))
            except Exception:
                pass
        entry.pack(pady=5)
        entry.focus()

        result = {'points': None}

        def submit():
            value = entry.get().strip()
            if not value:
                pts = 0
            else:
                try:
                    pts = int(value)
                except ValueError:
                    messagebox.showerror("שגיאה", "יש להזין מספר שלם.")
                    return
                if pts < 0:
                    pts = 0
                if effective_max is not None and pts > effective_max:
                    messagebox.showerror(
                        "שגיאה",
                        f"לא ניתן להגדיר יותר מ-{effective_max} נקודות לבונוס מורה (לכל תלמיד בסבב)."
                    )
                    return

            result['points'] = pts
            dialog.destroy()

        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=10)

        tk.Button(
            btn_frame,
            text="💾 שמור",
            command=submit,
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=20,
            pady=6
        ).pack(side=tk.RIGHT, padx=5)

        tk.Button(
            btn_frame,
            text="✖ בטל",
            command=dialog.destroy,
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            padx=16,
            pady=6
        ).pack(side=tk.RIGHT, padx=5)

        entry.bind('<Return>', lambda e: submit())

        self.root.wait_window(dialog)

        if result['points'] is None:
            return

        try:
            self.db.set_teacher_bonus(teacher_id, result['points'])
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לעדכן בונוס למורה:\n{e}")
            return

        if result['points'] > 0:
            messagebox.showinfo("הצלחה", f"בונוס המורה עודכן ל-{result['points']} נקודות לכל תלמיד בסבב.")
        else:
            messagebox.showinfo("הצלחה", "בונוס המורה בוטל (0 נקודות).")

        # עדכון כיתוב כפתור בונוס למורה (אם קיים)
        if hasattr(self, 'teacher_bonus_btn'):
            if result['points'] > 0:
                self.teacher_bonus_btn.config(text=f"🎁 בונוס ({result['points']} נק')", bg='#f39c12')
            else:
                self.teacher_bonus_btn.config(text="🎁 בונוס", bg='#f39c12')

    def load_bonus_settings(self):
        """טעינת הגדרות בונוס (משותף לכל העמדות כאשר מוגדרת תיקיית רשת)."""
        try:
            # ברירת מחדל: קובץ ליד קוד התוכנה
            local_bonus_file = os.path.join(self.base_dir, 'bonus_settings.json')
            bonus_file = local_bonus_file
            # אם מוגדרת תיקיית רשת משותפת – נעדיף קובץ משותף שם
            try:
                cfg = self.load_app_config()
                if isinstance(cfg, dict):
                    shared_folder = cfg.get('shared_folder') or cfg.get('network_root')
                    if shared_folder and os.path.isdir(shared_folder):
                        bonus_file = os.path.join(shared_folder, 'bonus_settings.json')
                        # אם אין עדיין קובץ משותף אבל יש מקומי – נעתיק
                        if (not os.path.exists(bonus_file)) and os.path.exists(local_bonus_file):
                            try:
                                shutil.copy2(local_bonus_file, bonus_file)
                            except Exception:
                                pass
            except Exception:
                pass

            if os.path.exists(bonus_file):
                with open(bonus_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception:
            pass
        return {"bonus_active": False, "bonus_points": 0, "bonus_running": False, "students_got_bonus": [], "master_card_2": "8888"}
    
    def save_bonus_settings(self, settings):
        """שמירת הגדרות בונוס (משותף לכל העמדות כאשר מוגדרת תיקיית רשת)."""
        try:
            # ברירת מחדל: קובץ ליד קוד התוכנה
            bonus_file = os.path.join(self.base_dir, 'bonus_settings.json')
            # אם מוגדרת תיקיית רשת משותפת – נשמור שם כדי לסנכרן בין כל העמדות
            try:
                cfg = self.load_app_config()
                if isinstance(cfg, dict):
                    shared_folder = cfg.get('shared_folder') or cfg.get('network_root')
                    if shared_folder and os.path.isdir(shared_folder):
                        try:
                            os.makedirs(shared_folder, exist_ok=True)
                        except Exception:
                            pass
                        bonus_file = os.path.join(shared_folder, 'bonus_settings.json')
            except Exception:
                pass

            with open(bonus_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"שגיאה בשמירת בונוס: {e}")
    
    def update_bonus_button(self):
        """עדכון מראה כפתור הבונוס לפי הסטטוס"""
        # רק אם זה מנהל ויש כפתור בונוס
        if not hasattr(self, 'bonus_btn'):
            return
        
        settings = self.load_bonus_settings()
        if settings.get('bonus_active', False):
            points = settings.get('bonus_points', 0)
            # טקסט מרובה שורות כדי לשמור על רוחב כפתור דומה למצב הרגיל
            # ובכך לא לדחוק את כפתור "הוראות" מהסרגל.
            self.bonus_btn.config(
                text=f"🎁 בונוס מיוחד\n({points} נק' פעיל)",
                bg='#e74c3c'
            )
        else:
            self.bonus_btn.config(text="🎁 בונוס מיוחד", bg='#f39c12')
    
    def toggle_bonus_mode(self):
        """הפעלה/כיבוי מצב בונוס מיוחד (מאסטר 2 בעמדה הציבורית)"""
        if not self.ensure_can_modify():
            return
        settings = self.load_bonus_settings()
        
        if not settings['bonus_active']:
            # הפעלת מצב בונוס מיוחד - שאל כמה נקודות
            points = self.ask_bonus_points()
            if points is not None:
                settings['bonus_active'] = True
                settings['bonus_points'] = points
                settings['bonus_running'] = False
                settings['students_got_bonus'] = []
                self.save_bonus_settings(settings)
                
                self.update_bonus_button()
                self.show_status_message(f"✓ מצב בונוס מיוחד הופעל! {points} נקודות - הצג כרטיס מאסטר 2 בעמדה הציבורית להפעלה/סיום")
        else:
            # כיבוי מצב בונוס מיוחד
            settings['bonus_active'] = False
            settings['bonus_running'] = False
            settings['students_got_bonus'] = []
            self.save_bonus_settings(settings)
            
            self.update_bonus_button()
            self.show_status_message("✓ מצב בונוס מיוחד כובה")
    
    def ask_bonus_points(self):
        """שאלת משתמש כמה נקודות לבונוס המיוחד (מאסטר 2)"""
        dialog = tk.Toplevel(self.root)
        dialog.title("מצב בונוס מיוחד")
        dialog.geometry("560x340")
        try:
            dialog.minsize(560, 340)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        
        result = {'points': None}
        
        tk.Label(
            dialog,
            text="כמה נקודות 'בונוס מיוחד' יקבל כל תלמיד בזמן שהבונוס פעיל?\n(הבונוס מופעל/מכובה בעמדה הציבורית בעזרת כרטיס מאסטר 2)",
            font=('Arial', 13, 'bold'),
            bg='#ecf0f1',
            wraplength=430,
            justify='right'
        ).pack(pady=20)
        
        points_entry = tk.Entry(dialog, font=('Arial', 14), width=10)
        points_entry.pack(pady=10)
        points_entry.focus()
        
        def submit():
            try:
                points = int(points_entry.get())
                result['points'] = points
                dialog.destroy()
            except ValueError:
                messagebox.showerror("שגיאה", "יש להזין מספר שלם")
        
        tk.Button(
            dialog,
            text="הפעל בונוס מיוחד",
            command=submit,
            font=('Arial', 12),
            bg='#f39c12',
            fg='white',
            padx=20,
            pady=10
        ).pack(pady=10)
        
        points_entry.bind('<Return>', lambda e: submit())
        
        self.root.wait_window(dialog)
        return result['points']
    
    def update_master_card_2(self):
        """עדכון כרטיס מאסטר 2 להפעלת הבונוס המיוחד בעמדה הציבורית"""
        dialog = tk.Toplevel(self.root)
        dialog.title("כרטיס מאסטר 2")
        dialog.geometry("460x260")
        try:
            dialog.minsize(460, 260)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(
            dialog,
            text="הזן מספר כרטיס מאסטר 2 (בונוס מיוחד):",
            font=('Arial', 12, 'bold'),
            bg='#ecf0f1'
        ).pack(pady=20)
        
        card_entry = tk.Entry(dialog, font=('Arial', 14), width=15)
        card_entry.pack(pady=10)
        card_entry.focus()
        
        settings = self.load_bonus_settings()
        card_entry.insert(0, settings.get('master_card_2', '8888'))
        
        def submit():
            card_number = card_entry.get().strip()
            if card_number:
                settings = self.load_bonus_settings()
                settings['master_card_2'] = card_number
                self.save_bonus_settings(settings)
                
                messagebox.showinfo(
                    "הצלחה",
                    f"כרטיס מאסטר 2 עודכן!\n\n"
                    f"מספר: {card_number}\n\n"
                    f"כרטיס זה ישמש להפעלת/סיום מצב 'בונוס מיוחד' בעמדה הציבורית."
                )
                dialog.destroy()
        
        tk.Button(
            dialog,
            text="שמור",
            command=submit,
            font=('Arial', 12),
            bg='#e67e22',
            fg='white',
            padx=20,
            pady=10
        ).pack(pady=10)
        
        card_entry.bind('<Return>', lambda e: submit())
        
        self.root.wait_window(dialog)

    def open_product_variants_manager(self, product_id: int):
        if not bool(ENABLE_PURCHASES):
            messagebox.showwarning("לא זמין", "מסך קניות אינו זמין בגרסה זו")
            return
        if not self.ensure_can_modify():
            return
        try:
            pid = int(product_id or 0)
        except Exception:
            pid = 0
        if not pid:
            messagebox.showwarning('אזהרה', 'בחר מוצר לניהול וריאציות')
            return

        product_name = f"מוצר לא ידוע (#{pid})"
        try:
            p = None
            try:
                p = self.db.get_product_by_id(pid)
            except Exception:
                p = None
            if not p:
                try:
                    allp = self.db.get_all_products(active_only=False) or []
                    p = next((x for x in allp if int(x.get('id', 0) or 0) == pid), None)
                except Exception:
                    p = None
            if p:
                product_name = (str(p.get('display_name') or '').strip() or str(p.get('name') or '').strip() or product_name)
        except Exception:
            pass

        dialog = tk.Toplevel(self.root)
        dialog.title(f"📦 וריאציות: {product_name}")
        dialog.configure(bg='#ecf0f1')
        try:
            dialog.minsize(820, 500)
        except Exception:
            pass
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text=fix_rtl_text(f"ניהול וריאציות למוצר: {product_name}"), font=('Arial', 13, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(14, 8))

        list_frame = tk.Frame(dialog, bg='#ecf0f1')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        cols = ('active', 'sort', 'variant_name', 'display_name', 'price', 'stock')
        tree = ttk.Treeview(list_frame, columns=cols, show='headings', height=14)
        tree.heading('active', text='פעיל')
        tree.heading('sort', text='סדר')
        tree.heading('variant_name', text='שם וריאציה')
        tree.heading('display_name', text='שם תצוגה')
        tree.heading('price', text='מחיר')
        tree.heading('stock', text='מלאי')

        tree.column('active', width=60, anchor='center')
        tree.column('sort', width=70, anchor='center')
        tree.column('variant_name', width=170, anchor='e')
        tree.column('display_name', width=220, anchor='e')
        tree.column('price', width=90, anchor='center')
        tree.column('stock', width=90, anchor='center')

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=sb.set)

        def _load(select_vid: int = None):
            try:
                rows = self.db.get_product_variants(pid, active_only=False) or []
            except Exception:
                rows = []
            try:
                tree.delete(*tree.get_children())
            except Exception:
                pass
            item_to_select = None
            for r in rows:
                try:
                    vid = int(r.get('id') or 0)
                except Exception:
                    vid = 0
                if not vid:
                    continue
                active_txt = '✓' if int(r.get('is_active', 1) or 0) == 1 else '✗'
                stock_val = r.get('stock_qty', None)
                stock_txt = '' if stock_val is None else str(stock_val)
                it = tree.insert('', tk.END, values=(
                    active_txt,
                    str(r.get('sort_order', 0) or 0),
                    str(r.get('variant_name') or ''),
                    str(r.get('display_name') or ''),
                    str(r.get('price_points', 0) or 0),
                    stock_txt,
                ), tags=(f"vid:{vid}",))
                if select_vid and vid == int(select_vid):
                    item_to_select = it
            if item_to_select:
                try:
                    tree.selection_set(item_to_select)
                    tree.see(item_to_select)
                except Exception:
                    pass

        def _selected_vid() -> int:
            sel = tree.selection()
            if not sel:
                return 0
            tags = tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('vid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _get_row(vid: int) -> dict:
            try:
                rows = self.db.get_product_variants(pid, active_only=False) or []
            except Exception:
                rows = []
            for r in rows:
                try:
                    if int(r.get('id', 0) or 0) == int(vid or 0):
                        return r
                except Exception:
                    continue
            return {}

        def _ask_stock(initial):
            s0 = '' if initial is None else str(initial)
            raw = simpledialog.askstring('וריאציה', 'מלאי (ריק=ללא):', initialvalue=s0, parent=dialog)
            if raw is None:
                return None, False
            raw = str(raw or '').strip()
            if raw == '':
                return None, True
            try:
                return int(raw), True
            except Exception:
                messagebox.showwarning('אזהרה', 'מלאי חייב להיות מספר או ריק')
                return None, False

        def add_variant():
            vn = simpledialog.askstring('וריאציה', 'שם וריאציה (למשל קטן/בינוני/גדול):', parent=dialog)
            if vn is None:
                return
            vn = str(vn or '').strip()
            if not vn:
                messagebox.showwarning('אזהרה', 'חובה להזין שם וריאציה')
                return
            dn = simpledialog.askstring('וריאציה', 'שם תצוגה (אופציונלי):', parent=dialog)
            if dn is None:
                dn = ''
            price = simpledialog.askinteger('וריאציה', 'מחיר בנקודות:', parent=dialog, minvalue=0)
            if price is None:
                price = 0
            stock_qty, ok = _ask_stock(None)
            if not ok:
                return
            sort_order = simpledialog.askinteger('וריאציה', 'סדר (מספר):', parent=dialog, minvalue=0)
            if sort_order is None:
                sort_order = 0
            is_active = 1 if messagebox.askyesno('וריאציה', 'להפעיל את הווריאציה?') else 0
            try:
                vid = self.db.add_product_variant(
                    product_id=pid,
                    variant_name=vn,
                    display_name=str(dn or '').strip(),
                    price_points=int(price or 0),
                    stock_qty=stock_qty,
                    deduct_points=1,
                    is_active=int(is_active),
                    sort_order=int(sort_order or 0),
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_vid=vid)

        def edit_variant():
            vid = _selected_vid()
            if not vid:
                messagebox.showwarning('אזהרה', 'בחר וריאציה לעריכה')
                return
            r = _get_row(vid)
            vn = simpledialog.askstring('עריכה', 'שם וריאציה:', initialvalue=str(r.get('variant_name') or ''), parent=dialog)
            if vn is None:
                return
            vn = str(vn or '').strip()
            if not vn:
                messagebox.showwarning('אזהרה', 'חובה להזין שם וריאציה')
                return
            dn = simpledialog.askstring('עריכה', 'שם תצוגה (אופציונלי):', initialvalue=str(r.get('display_name') or ''), parent=dialog)
            if dn is None:
                dn = ''
            price0 = 0
            try:
                price0 = int(r.get('price_points', 0) or 0)
            except Exception:
                price0 = 0
            price = simpledialog.askinteger('עריכה', 'מחיר בנקודות:', initialvalue=price0, parent=dialog, minvalue=0)
            if price is None:
                price = price0
            stock_qty, ok = _ask_stock(r.get('stock_qty', None))
            if not ok:
                return
            sort0 = 0
            try:
                sort0 = int(r.get('sort_order', 0) or 0)
            except Exception:
                sort0 = 0
            sort_order = simpledialog.askinteger('עריכה', 'סדר (מספר):', initialvalue=sort0, parent=dialog, minvalue=0)
            if sort_order is None:
                sort_order = sort0
            is_active = 1 if messagebox.askyesno('עריכה', 'להפעיל את הווריאציה?') else 0
            try:
                self.db.update_product_variant(
                    int(vid),
                    variant_name=vn,
                    display_name=str(dn or '').strip(),
                    price_points=int(price or 0),
                    stock_qty=stock_qty,
                    deduct_points=1,
                    is_active=int(is_active),
                    sort_order=int(sort_order or 0),
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_vid=vid)

        def toggle_variant():
            vid = _selected_vid()
            if not vid:
                messagebox.showwarning('אזהרה', 'בחר וריאציה')
                return
            r = _get_row(vid)
            if not r:
                return
            try:
                new_active = 0 if int(r.get('is_active', 1) or 0) == 1 else 1
            except Exception:
                new_active = 1
            try:
                self.db.update_product_variant(
                    int(vid),
                    variant_name=str(r.get('variant_name') or '').strip(),
                    display_name=str(r.get('display_name') or '').strip(),
                    price_points=int(r.get('price_points', 0) or 0),
                    stock_qty=r.get('stock_qty', None),
                    deduct_points=1,
                    is_active=int(new_active),
                    sort_order=int(r.get('sort_order', 0) or 0),
                )
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_vid=vid)

        def delete_variant():
            vid = _selected_vid()
            if not vid:
                messagebox.showwarning('אזהרה', 'בחר וריאציה למחיקה')
                return
            if not messagebox.askyesno('מחיקה', 'למחוק את הווריאציה שנבחרה?'):
                return
            try:
                self.db.delete_product_variant(int(vid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load()

        btns = tk.Frame(dialog, bg='#ecf0f1')
        btns.pack(pady=(0, 14))
        tk.Button(btns, text='➕ הוסף', command=add_variant, font=('Arial', 10, 'bold'), bg='#27ae60', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✏ ערוך', command=edit_variant, font=('Arial', 10, 'bold'), bg='#3498db', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✅ פעיל/כבוי', command=toggle_variant, font=('Arial', 10, 'bold'), bg='#16a085', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🗑 מחק', command=delete_variant, font=('Arial', 10, 'bold'), bg='#e74c3c', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✖ סגור', command=dialog.destroy, font=('Arial', 10), bg='#95a5a6', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)

        try:
            tree.bind('<Double-1>', lambda _e: edit_variant())
        except Exception:
            pass
        _load()
        self.root.wait_window(dialog)

    def open_activity_schedules_manager(self, activity_id: int):
        if not self.ensure_can_modify():
            return
        try:
            aid = int(activity_id or 0)
        except Exception:
            aid = 0
        if not aid:
            return

        dialog = tk.Toplevel(self.root)
        dialog.title('🗓 לוחות זמנים לאתגר')
        dialog.configure(bg='#ecf0f1')
        try:
            dialog.geometry('1000x560')
        except Exception:
            pass
        try:
            dialog.minsize(980, 480)
        except Exception:
            pass
        dialog.transient(self.root)
        dialog.grab_set()
        try:
            dialog.resizable(True, True)
        except Exception:
            pass

        tk.Label(dialog, text=fix_rtl_text('לוחות זמנים'), font=('Arial', 13, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(14, 8))

        frame = tk.Frame(dialog, bg='#ecf0f1')
        frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        cols = ('active', 'time', 'days', 'date', 'general', 'classes')
        tree = ttk.Treeview(frame, columns=cols, show='headings', height=12)
        tree.heading('active', text='פעיל')
        tree.heading('time', text='שעות')
        tree.heading('days', text='ימים')
        tree.heading('date', text='תאריכים')
        tree.heading('general', text='כללי')
        tree.heading('classes', text='כיתות')
        tree.column('active', width=60, anchor='center')
        tree.column('time', width=200, anchor='center')
        tree.column('days', width=120, anchor='center')
        tree.column('date', width=220, anchor='center')
        tree.column('general', width=70, anchor='center')
        tree.column('classes', width=260, anchor='e')
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=sb.set)

        def _selected_sid() -> int:
            sel = tree.selection()
            if not sel:
                return 0
            tags = tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('sid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _load(select_id: int = None):
            try:
                rows = self.db.get_activity_schedules(int(aid), active_only=False) or []
            except Exception:
                rows = []
            try:
                tree.delete(*tree.get_children())
            except Exception:
                pass
            it_sel = None
            for r in rows:
                sid = int(r.get('id', 0) or 0)
                if not sid:
                    continue
                active_txt = '✓' if int(r.get('is_active', 1) or 0) == 1 else '✗'
                gen_txt = '✓' if int(r.get('is_general', 1) or 0) == 1 else '✗'
                t1 = str(r.get('start_time') or '')
                t2 = str(r.get('end_time') or '')
                time_txt = f"{t1} - {t2}".strip(' -')
                days_txt = str(r.get('days_of_week') or '')
                d1 = str(r.get('start_date') or '')
                d2 = str(r.get('end_date') or '')
                date_txt = f"{d1} עד {d2}".strip()
                classes_txt = str(r.get('classes') or '')
                it = tree.insert('', tk.END, values=(active_txt, time_txt, days_txt, date_txt, gen_txt, classes_txt), tags=(f"sid:{sid}",))
                if select_id and sid == int(select_id):
                    it_sel = it
            if it_sel:
                try:
                    tree.selection_set(it_sel)
                    tree.see(it_sel)
                except Exception:
                    pass

        def _get_row(sid: int) -> dict:
            try:
                rows = self.db.get_activity_schedules(int(aid), active_only=False) or []
            except Exception:
                rows = []
            return next((x for x in rows if int(x.get('id', 0) or 0) == int(sid)), {})

        def _ask_schedule(existing: dict = None) -> dict:
            existing = existing or {}
            st = simpledialog.askstring('לו"ז', 'שעת התחלה (HH:MM או ריק):', initialvalue=str(existing.get('start_time') or ''), parent=dialog)
            if st is None:
                return {'ok': False}
            et = simpledialog.askstring('לו"ז', 'שעת סיום (HH:MM או ריק):', initialvalue=str(existing.get('end_time') or ''), parent=dialog)
            if et is None:
                return {'ok': False}
            days = simpledialog.askstring('לו"ז', 'ימים (א,ב,ג... או ריק=כל):', initialvalue=str(existing.get('days_of_week') or ''), parent=dialog)
            if days is None:
                days = str(existing.get('days_of_week') or '')
            sd = simpledialog.askstring('לו"ז', 'תאריך התחלה (YYYY-MM-DD או ריק):', initialvalue=str(existing.get('start_date') or ''), parent=dialog)
            if sd is None:
                sd = str(existing.get('start_date') or '')
            ed = simpledialog.askstring('לו"ז', 'תאריך סיום (YYYY-MM-DD או ריק):', initialvalue=str(existing.get('end_date') or ''), parent=dialog)
            if ed is None:
                ed = str(existing.get('end_date') or '')
            is_general = 1 if messagebox.askyesno('לו"ז', 'כללי לכל הכיתות?') else 0
            classes = ''
            if not is_general:
                classes = simpledialog.askstring('לו"ז', 'כיתות (מופרד בפסיק):', initialvalue=str(existing.get('classes') or ''), parent=dialog)
                if classes is None:
                    classes = str(existing.get('classes') or '')
                if not str(classes or '').strip():
                    messagebox.showwarning('אזהרה', 'כשכללי כבוי חובה להזין כיתות')
                    return {'ok': False}
            is_active = 1 if messagebox.askyesno('לו"ז', 'להפעיל את שורת הלו"ז?') else 0
            return {
                'ok': True,
                'start_time': str(st or '').strip(),
                'end_time': str(et or '').strip(),
                'days_of_week': str(days or '').strip(),
                'start_date': str(sd or '').strip(),
                'end_date': str(ed or '').strip(),
                'is_general': int(is_general),
                'classes': str(classes or '').strip(),
                'is_active': int(is_active),
            }

        def add_row():
            data = _ask_schedule()
            if not data.get('ok'):
                return
            try:
                sid = self.db.add_activity_schedule(activity_id=int(aid), **data)
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_id=sid)

        def edit_row():
            sid = _selected_sid()
            if not sid:
                messagebox.showwarning('אזהרה', 'בחר שורה לעריכה')
                return
            row = _get_row(sid)
            data = _ask_schedule(existing=row)
            if not data.get('ok'):
                return
            try:
                self.db.update_activity_schedule(int(sid), **data)
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_id=sid)

        def delete_row():
            sid = _selected_sid()
            if not sid:
                messagebox.showwarning('אזהרה', 'בחר שורה למחיקה')
                return
            if not messagebox.askyesno('מחיקה', 'למחוק את שורת הלו"ז?'):
                return
            try:
                self.db.delete_activity_schedule(int(sid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load()

        btns = tk.Frame(dialog, bg='#ecf0f1')
        btns.pack(pady=(0, 14))
        tk.Button(btns, text='➕ הוסף', command=add_row, bg='#27ae60', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✏ ערוך', command=edit_row, bg='#3498db', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🗑 מחק', command=delete_row, bg='#e74c3c', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✖ סגור', command=dialog.destroy, bg='#95a5a6', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)

        try:
            tree.bind('<Double-1>', lambda _e: edit_row())
        except Exception:
            pass
        _load()
        self.root.wait_window(dialog)

    def export_activity_cards_excel(self):
        if not self.ensure_can_modify():
            return
        if not (self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1):
            messagebox.showwarning("אין הרשאה", "גישה רק למנהלים")
            return
        rows = []
        source = ''
        try:
            # אתגרים מתוזמנים (קופה)
            rows = self.db.get_all_scheduled_services() or []
            source = 'scheduled_services'
        except Exception as e:
            print(f"שגיאה בקריאת אתגרים (scheduled_services): {e}")
            rows = []
        if not rows:
            # תאימות לאחור: אתגרים בטבלה הישנה (activities)
            try:
                rows = self.db.get_all_activities() or []
                source = 'activities'
            except Exception as e:
                print(f"שגיאה בקריאת אתגרים (activities): {e}")
                rows = []
        if not rows:
            messagebox.showwarning('אין נתונים', 'אין אתגרים במערכת. נא להוסיף אתגרים')
            return

        # Ask user: preview (default) or save to file
        try:
            choice_dialog = tk.Toplevel(self.root)
            choice_dialog.title('ייצוא כרטיסי אתגרים')
            choice_dialog.geometry('520x240')
            try:
                choice_dialog.minsize(520, 220)
            except Exception:
                pass
            choice_dialog.configure(bg='#ecf0f1')
            choice_dialog.transient(self.root)
            choice_dialog.grab_set()

            tk.Label(
                choice_dialog,
                text=fix_rtl_text('בחר אופציית ייצוא:'),
                font=('Arial', 13, 'bold'),
                bg='#ecf0f1',
                fg='#2c3e50'
            ).pack(pady=18)

            user_choice = {'action': 'preview'}

            def choose_preview():
                user_choice['action'] = 'preview'
                choice_dialog.destroy()

            def choose_save():
                user_choice['action'] = 'save'
                choice_dialog.destroy()

            btn_frame = tk.Frame(choice_dialog, bg='#ecf0f1')
            btn_frame.pack(pady=12)

            tk.Button(
                btn_frame,
                text='👁 דוח זמני',
                command=choose_preview,
                font=('Arial', 12, 'bold'),
                bg='#3498db',
                fg='white',
                padx=18,
                pady=8
            ).pack(side=tk.LEFT, padx=10)

            tk.Button(
                btn_frame,
                text='💾 שמירה לקובץ',
                command=choose_save,
                font=('Arial', 12, 'bold'),
                bg='#27ae60',
                fg='white',
                padx=18,
                pady=8
            ).pack(side=tk.LEFT, padx=10)

            try:
                choice_dialog.wait_window()
            except Exception:
                pass
        except Exception:
            user_choice = {'action': 'preview'}

        fp = ''
        if user_choice.get('action') == 'save':
            try:
                default_name = 'כרטיסי אתגרים.xlsx'
                fp = filedialog.asksaveasfilename(
                    title='שמירת קובץ להדפסה',
                    defaultextension='.xlsx',
                    initialdir=self._get_downloads_dir(),
                    initialfile=default_name,
                    filetypes=[('Excel', '*.xlsx')]
                )
            except Exception:
                fp = ''
            if not fp:
                return

        try:
            data = []
            if source == 'activities':
                for r in rows:
                    if int(r.get('is_active', 1) or 0) != 1:
                        continue
                    data.append({
                        'שם אתגר': str(r.get('name') or ''),
                        'נקודות': int(r.get('points', 0) or 0),
                        'משך (דקות)': 0,
                        'קיבולת': 0,
                    })
            else:
                for r in rows:
                    if int(r.get('is_active', 1) or 0) != 1:
                        continue
                    # scheduled_services מחוברים ל-products
                    product_name = str(r.get('product_display_name') or r.get('product_name') or r.get('name') or '')
                    product_points = int(r.get('product_price_points') or r.get('points', 0) or 0)
                    data.append({
                        'שם אתגר': product_name,
                        'נקודות': product_points,
                        'משך (דקות)': int(r.get('duration_minutes', 0) or 0),
                        'קיבולת': int(r.get('capacity_per_slot', 0) or 0),
                    })
            if not data:
                messagebox.showwarning('אין נתונים', 'אין אתגרים פעילים לייצוא')
                return
            df = pd.DataFrame(data, columns=['שם אתגר', 'נקודות', 'משך (דקות)', 'קיבולת'])
            if user_choice.get('action') == 'preview':
                try:
                    self._show_preview_window(df, 'כרטיסי אתגרים')
                except Exception as e:
                    messagebox.showerror('שגיאה', f'שגיאה בהצגת תצוגה:\n{str(e)}')
                return

            df.to_excel(fp, index=False, engine='openpyxl')
            
            # Apply RTL and alternating colors styling
            try:
                from openpyxl import load_workbook
                from excel_styling import apply_rtl_and_alternating_colors
                wb = load_workbook(fp)
                ws = wb.active
                apply_rtl_and_alternating_colors(ws, has_header=True)
                wb.save(fp)
            except Exception:
                pass
            
            messagebox.showinfo('נשמר', f'נשמר קובץ:\n{fp}')
            try:
                if messagebox.askyesno('הייצוא הסתיים', 'לפתוח את הקובץ עכשיו?'):
                    try:
                        os.startfile(fp)
                    except Exception:
                        pass
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def open_activities_manager(self):
        if not self.ensure_can_modify():
            return
        if not (self.current_teacher and int(self.current_teacher.get('is_admin', 0) or 0) == 1):
            messagebox.showwarning("אין הרשאה", "גישה רק למנהלים")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("🏆 ניהול אתגרים")
        dialog.configure(bg='#ecf0f1')
        dialog.geometry("1000x620")
        try:
            dialog.minsize(950, 580)
        except Exception:
            pass
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)

        tk.Label(dialog, text=fix_rtl_text('ניהול אתגרים'), font=('Arial', 14, 'bold'), bg='#ecf0f1', fg='#2c3e50').pack(pady=(14, 8))

        frame = tk.Frame(dialog, bg='#ecf0f1')
        frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        cols = ('active', 'name', 'points', 'print_code')
        tree = ttk.Treeview(frame, columns=cols, show='headings', height=16)
        tree.heading('active', text='פעיל')
        tree.heading('name', text='שם אתגר')
        tree.heading('points', text='נקודות')
        tree.heading('print_code', text='קוד הדפסה')
        tree.column('active', width=60, anchor='center')
        tree.column('name', width=420, anchor='e')
        tree.column('points', width=90, anchor='center')
        tree.column('print_code', width=180, anchor='center')
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=sb.set)

        def _selected_aid() -> int:
            sel = tree.selection()
            if not sel:
                return 0
            tags = tree.item(sel[0]).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('aid:'):
                    try:
                        return int(t.split(':', 1)[1])
                    except Exception:
                        return 0
            return 0

        def _load(select_id: int = None):
            try:
                rows = self.db.get_all_activities() or []
            except Exception:
                rows = []
            try:
                tree.delete(*tree.get_children())
            except Exception:
                pass
            it_sel = None
            for r in rows:
                aid = int(r.get('id', 0) or 0)
                if not aid:
                    continue
                active_txt = '✓' if int(r.get('is_active', 1) or 0) == 1 else '✗'
                it = tree.insert('', tk.END, values=(
                    active_txt,
                    str(r.get('name') or ''),
                    str(r.get('points', 0) or 0),
                    str(r.get('print_code') or ''),
                ), tags=(f"aid:{aid}",))
                if select_id and aid == int(select_id):
                    it_sel = it
            if it_sel:
                try:
                    tree.selection_set(it_sel)
                    tree.see(it_sel)
                except Exception:
                    pass

        def _ask_fields(existing: dict = None) -> dict:
            existing = existing or {}
            name = simpledialog.askstring('אתגר', 'שם אתגר:', initialvalue=str(existing.get('name') or ''), parent=dialog)
            if name is None:
                return {'ok': False}
            name = str(name or '').strip()
            if not name:
                messagebox.showwarning('אזהרה', 'חובה להזין שם אתגר')
                return {'ok': False}
            pts0 = 0
            try:
                pts0 = int(existing.get('points', 0) or 0)
            except Exception:
                pts0 = 0
            pts = simpledialog.askinteger('אתגר', 'נקודות:', initialvalue=pts0, parent=dialog, minvalue=0)
            if pts is None:
                pts = pts0
            code = simpledialog.askstring('אתגר', 'קוד הדפסה (אופציונלי):', initialvalue=str(existing.get('print_code') or ''), parent=dialog)
            if code is None:
                code = str(existing.get('print_code') or '')
            is_active = 1 if messagebox.askyesno('אתגר', 'להפעיל את האתגר?') else 0
            return {'ok': True, 'name': name, 'points': int(pts or 0), 'print_code': str(code or '').strip(), 'is_active': int(is_active)}

        def add_activity():
            data = _ask_fields()
            if not data.get('ok'):
                return
            try:
                aid = self.db.add_activity(name=data['name'], points=data['points'], print_code=data['print_code'], is_active=data['is_active'])
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_id=aid)

        def edit_activity():
            aid = _selected_aid()
            if not aid:
                messagebox.showwarning('אזהרה', 'בחר אתגר לעריכה')
                return
            try:
                rows = self.db.get_all_activities() or []
            except Exception:
                rows = []
            r = next((x for x in rows if int(x.get('id', 0) or 0) == int(aid)), {})
            data = _ask_fields(existing=r)
            if not data.get('ok'):
                return
            try:
                self.db.update_activity(int(aid), name=data['name'], points=data['points'], print_code=data['print_code'], is_active=data['is_active'])
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load(select_id=aid)

        def delete_activity():
            aid = _selected_aid()
            if not aid:
                messagebox.showwarning('אזהרה', 'בחר אתגר למחיקה')
                return
            if not messagebox.askyesno('מחיקה', 'למחוק את האתגר (כולל לוחות זמנים)?'):
                return
            try:
                self.db.delete_activity(int(aid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
                return
            _load()

        def manage_schedules():
            aid = _selected_aid()
            if not aid:
                messagebox.showwarning('אזהרה', 'בחר אתגר לניהול לוחות זמנים')
                return
            try:
                self.open_activity_schedules_manager(int(aid))
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))

        def export_cards():
            try:
                self.export_activity_cards_excel()
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))

        btns = tk.Frame(dialog, bg='#ecf0f1')
        btns.pack(pady=(0, 14))
        tk.Button(btns, text='➕ הוסף', command=add_activity, bg='#27ae60', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✏ ערוך', command=edit_activity, bg='#3498db', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🗓 לוחות זמנים', command=manage_schedules, bg='#8e44ad', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🖨 ייצוא להדפסה', command=export_cards, bg='#34495e', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='🗑 מחק', command=delete_activity, bg='#e74c3c', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text='✖ סגור', command=dialog.destroy, bg='#95a5a6', fg='white', padx=14, pady=6).pack(side=tk.LEFT, padx=6)

        try:
            tree.bind('<Double-1>', lambda _e: edit_activity())
        except Exception:
            pass
        _load()
        self.root.wait_window(dialog)
    
    def open_time_bonus_manager(self):
        """פתיחת חלון ניהול בונוס זמנים"""
        if not self.ensure_can_modify():
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("⏰ ניהול בונוס זמנים")
        dialog.geometry("860x620")
        try:
            dialog.minsize(820, 560)
        except Exception:
            pass
        dialog.configure(bg='#ecf0f1')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        # כותרת
        tk.Label(
            dialog,
            text="⏰ ניהול בונוס זמנים אוטומטי",
            font=('Arial', 16, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=15)
        
        # מסגרת רשימה
        list_frame = tk.Frame(dialog, bg='#ecf0f1')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # תצוגה שטוחה (RTL): נהפוך סדר עמודות כך ש"שם" יופיע מימין
        columns = ('shown', 'days', 'classes', 'general', 'active', 'points', 'end', 'start', 'name')
        tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12, selectmode='extended')

        tree.heading('name', text='שם')
        tree.heading('start', text='שעת התחלה')
        tree.heading('end', text='שעת סיום')
        tree.heading('points', text='נקודות')
        tree.heading('active', text='פעיל')
        tree.heading('general', text='כללי')
        tree.heading('classes', text='כיתות')
        tree.heading('days', text='ימים')
        tree.heading('shown', text='מוצג')

        tree.column('name', width=300, anchor='e')
        tree.column('start', width=120, anchor='e')
        tree.column('end', width=120, anchor='e')
        tree.column('points', width=90, anchor='e')
        tree.column('active', width=60, anchor='center')
        tree.column('general', width=60, anchor='center')
        tree.column('classes', width=130, anchor='e')
        tree.column('days', width=90, anchor='e')
        tree.column('shown', width=60, anchor='center')

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)

        try:
            tree.tag_configure('group_sep', background='#dfe6e9')
        except Exception:
            pass

        def _apply_rtl_entry_behavior(entry: tk.Entry):
            try:
                entry.config(justify='right')
            except Exception:
                pass

            def _force_cursor_end(_event=None):
                try:
                    entry.icursor(tk.END)
                except Exception:
                    pass

            def _on_key_release(_event=None):
                _force_cursor_end()

            def _on_focus_in(_event=None):
                _force_cursor_end()

            try:
                entry.bind('<KeyRelease>', _on_key_release, add=True)
                entry.bind('<FocusIn>', _on_focus_in, add=True)
            except Exception:
                pass
        
        def load_bonuses():
            """טעינת בונוסים בתצוגה שטוחה"""
            tree.delete(*tree.get_children())
            try:
                groups = self.db.get_time_bonus_groups()
            except Exception:
                bonuses = self.db.get_all_time_bonuses()
                groups = {}
                for b in bonuses:
                    g = (b.get('group_name') or b.get('name') or '').strip() or f"בונוס {b.get('id', '')}".strip()
                    groups.setdefault(g, []).append(b)

            group_index = 0
            for group_name, rows in groups.items():
                if group_index > 0:
                    tree.insert(
                        '',
                        tk.END,
                        values=('', '', '', '', '', '', '', '────────'),
                        tags=('group_sep',)
                    )
                group_index += 1
                for idx, r in enumerate(rows):
                    rid = int(r['id'])
                    row_active = "✓" if int(r.get('is_active', 0) or 0) == 1 else "✗"
                    row_general = "✓" if int(r.get('is_general', 1) or 0) == 1 else "✗"
                    row_shown = "✓" if int(r.get('is_shown_public', 1) or 0) == 1 else "✗"
                    row_classes = ''
                    try:
                        row_classes = str(r.get('classes') or '').strip()
                    except Exception:
                        row_classes = ''
                    row_days = ''
                    try:
                        row_days = str(r.get('days_of_week') or '').strip()
                    except Exception:
                        row_days = ''
                    if not row_days:
                        row_days = 'כל'
                    tree.insert(
                        '',
                        tk.END,
                        values=(
                            row_shown,
                            row_days,
                            row_classes,
                            row_general,
                            row_active,
                            str(r.get('bonus_points') or ''),
                            r.get('end_time') or '',
                            r.get('start_time') or '',
                            str(group_name) if idx == 0 else ''
                        ),
                        tags=(f"row:{rid}", f"group:{group_name}")
                    )

        def _parse_classes_list(classes_text: str) -> list:
            if not classes_text:
                return []
            s = str(classes_text).strip()
            if not s:
                return []
            s = s.replace(';', ',').replace('\u05f3', ',').replace('\u05f4', ',')
            parts = [p.strip() for p in s.split(',')]
            return [p for p in parts if p]

        def _parse_days_list(days_text: str) -> set:
            s = str(days_text or '').strip()
            if (not s) or s == 'כל':
                # Empty means "all days" (matches UI behavior).
                return {'א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ש'}
            s = s.replace(';', ',').replace('\u05f3', ',').replace('\u05f4', ',')
            parts = [p.strip() for p in s.split(',') if p.strip()]
            out = set()
            for p in parts:
                # Support either single Hebrew letter (א/ב/ג...) or common names (א׳, וכו')
                p0 = str(p).strip()
                if not p0:
                    continue
                # Normalize quotes-like marks
                p0 = p0.replace("'", '').replace('"', '').replace('׳', '').replace('""', '').strip()
                if not p0 or p0 == 'כל':
                    return {'א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ש'}
                if p0:
                    out.add(p0[0])
            if not out:
                return {'א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ש'}
            return out

        def _ranges_overlap(a_start: str, a_end: str, b_start: str, b_end: str) -> bool:
            try:
                # חפיפה כוללת גבול: אם שעה זהה לסיום/התחלה נחשבת התנגשות.
                # זה תואם ללוגיקת מסד הנתונים שמחשיבה end_time ככולל.
                return (a_start <= b_end) and (a_end >= b_start)
            except Exception:
                return False

        def _validate_time_bonus_conflicts(group_name: str, new_row: dict, existing_rows: list, exclude_row_id: int = None):
            """ולידציה: אין חפיפת שעות לפי כללי כללי/כיתות באותה קבוצה."""
            n_start = new_row.get('start_time') or ''
            n_end = new_row.get('end_time') or ''
            n_general = int(new_row.get('is_general', 0) or 0) == 1
            n_classes = set(_parse_classes_list(new_row.get('classes') or ''))
            n_days = _parse_days_list(new_row.get('days_of_week') or '')
            if (not n_general) and (not n_classes):
                raise ValueError("כש'כללי' כבוי חובה לבחור כיתות")

            for r in existing_rows:
                try:
                    if exclude_row_id is not None and int(r.get('id', 0) or 0) == int(exclude_row_id):
                        continue
                except Exception:
                    pass

                g = (r.get('group_name') or r.get('name') or '').strip()
                if g != group_name:
                    continue
                if int(r.get('is_active', 0) or 0) != 1:
                    continue

                # No conflict if the weekday sets do not intersect
                r_days = _parse_days_list(r.get('days_of_week') or '')
                try:
                    if not (n_days.intersection(r_days)):
                        continue
                except Exception:
                    pass

                r_start = str(r.get('start_time') or '')
                r_end = str(r.get('end_time') or '')
                if not _ranges_overlap(n_start, n_end, r_start, r_end):
                    continue

                r_general = int(r.get('is_general', 1) or 0) == 1
                if n_general or r_general:
                    raise ValueError("אין אפשרות לחפיפה בשעות כאשר קיימת שורה 'כללי' באותה קבוצה")

                r_classes = set(_parse_classes_list(r.get('classes') or ''))
                if n_classes.intersection(r_classes):
                    raise ValueError("אין אפשרות לחפיפה בשעות עבור אותן כיתות באותה קבוצה")
        
        def add_group():
            """הוספת בונוס חדש: חלון רוחבי עם אפשרות להוסיף שורות (קריטריונים)"""
            import re

            add_dialog = tk.Toplevel(dialog)
            add_dialog.title("הוספת בונוס")
            add_dialog.geometry("1x1")
            add_dialog.configure(bg='#ecf0f1')
            add_dialog.transient(dialog)
            add_dialog.grab_set()

            container = tk.Frame(add_dialog, bg='#ecf0f1')
            container.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)

            # grid: col2 = ימין, col0 = שמאל
            container.grid_columnconfigure(0, weight=0)
            container.grid_columnconfigure(1, weight=1)
            container.grid_columnconfigure(2, weight=0)

            tk.Label(container, text=":בונוס", font=('Arial', 12, 'bold'), bg='#ecf0f1').grid(row=0, column=2, sticky='e', padx=(8, 0), pady=(0, 10))
            group_entry = tk.Entry(container, font=('Arial', 13), width=28, justify='right')
            group_entry.grid(row=0, column=1, sticky='e', pady=(0, 10))
            group_entry.insert(0, "בוקר טוב")

            # טבלת קריטריונים: מימין לשמאל תצוגה = משעה | עד שעה | נקודות בונוס | כללי | כיתות | ימים | מוצג | מחק
            # ב-grid: col8=משעה (ימין), col7=עד שעה, col5=נקודות, col4=כללי, col3=כיתות, col2=ימים, col1=מוצג, col0=מחק (שמאל)
            grid_frame = tk.Frame(container, bg='#ecf0f1')
            grid_frame.grid(row=1, column=0, columnspan=3, sticky='nsew')

            for c in range(9):
                grid_frame.grid_columnconfigure(c, weight=0)

            tk.Label(grid_frame, text=":משעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=8, sticky='e', padx=8)
            tk.Label(grid_frame, text=":עד שעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=7, sticky='e', padx=8)
            tk.Label(grid_frame, text=":נקודות בונוס", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=5, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כללי", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=4, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כיתות", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=3, sticky='e', padx=8)
            tk.Label(grid_frame, text=":ימים", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=2, sticky='e', padx=8)
            tk.Label(grid_frame, text=":מוצג", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=1, sticky='e', padx=8)
            tk.Label(grid_frame, text=":מחק", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=0, sticky='e', padx=8)

            row_widgets = []

            def _normalize_time_str(s: str) -> str:
                try:
                    t = str(s or '').strip()
                    parts = t.split(':')
                    if len(parts) != 2:
                        return t
                    hh = int(parts[0])
                    mm = int(parts[1])
                    if hh < 0 or hh > 23 or mm < 0 or mm > 59:
                        return t
                    return f"{hh:02d}:{mm:02d}"
                except Exception:
                    return str(s or '').strip()

            def _add_minutes(t: str, minutes: int) -> str:
                try:
                    parts = str(t or '').strip().split(':')
                    if len(parts) != 2:
                        return ''
                    hh = int(parts[0])
                    mm = int(parts[1])
                    if hh < 0 or hh > 23 or mm < 0 or mm > 59:
                        return ''
                    total = hh * 60 + mm + int(minutes)
                    if total < 0 or total > (23 * 60 + 59):
                        return ''
                    nh = total // 60
                    nm = total % 60
                    return f"{nh:02d}:{nm:02d}"
                except Exception:
                    return ''

            def _add_row(initial=None):
                initial = initial or {}
                row_index = len(row_widgets) + 1

                start_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                start_entry.grid(row=row_index, column=8, sticky='e', padx=8, pady=4)
                start_entry.insert(0, initial.get('start_time', '07:30'))

                end_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                end_entry.grid(row=row_index, column=7, sticky='e', padx=8, pady=4)
                end_entry.insert(0, initial.get('end_time', '07:45'))

                pts_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                pts_entry.grid(row=row_index, column=5, sticky='e', padx=8, pady=4)
                pts_entry.insert(0, str(initial.get('bonus_points', 2)))

                general_var = tk.IntVar(value=int(initial.get('is_general', 1) or 0))
                general_cb = tk.Checkbutton(grid_frame, variable=general_var, bg='#ecf0f1')
                general_cb.grid(row=row_index, column=4, sticky='e', padx=8, pady=4)

                # Classes selection with button
                classes_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                classes_cell.grid(row=row_index, column=3, sticky='e', padx=8, pady=4)
                
                classes_entry = tk.Entry(classes_cell, font=('Arial', 11), width=12, justify='right', state='readonly')
                classes_entry.pack(side=tk.LEFT, padx=2)
                
                selected_classes_list = []
                initial_classes_str = str(initial.get('classes') or '').strip()
                if initial_classes_str:
                    selected_classes_list.extend([c.strip() for c in initial_classes_str.split(',') if c.strip()])
                    classes_entry.config(state='normal')
                    classes_entry.insert(0, ', '.join(selected_classes_list))
                    classes_entry.config(state='readonly')
                
                def open_classes_selector():
                    all_classes = set()
                    try:
                        students = self.db.get_all_students()
                        for s in students:
                            cn = (s.get('class_name') or '').strip()
                            if cn:
                                all_classes.add(cn)
                    except Exception:
                        pass
                    
                    if not all_classes:
                        messagebox.showinfo('מידע', 'אין כיתות במערכת')
                        return
                    
                    selector_dialog = tk.Toplevel(add_dialog)
                    selector_dialog.title('בחירת כיתות')
                    selector_dialog.geometry('400x500')
                    selector_dialog.configure(bg='#ecf0f1')
                    selector_dialog.transient(add_dialog)
                    selector_dialog.grab_set()
                    
                    tk.Label(
                        selector_dialog,
                        text=fix_rtl_text('בחר כיתות'),
                        font=('Arial', 14, 'bold'),
                        bg='#ecf0f1',
                        fg='#2c3e50'
                    ).pack(pady=(12, 6))
                    
                    cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                    
                    canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                    scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                    scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                    
                    scrollable_frame.bind(
                        "<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                    )
                    
                    canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                    canvas.configure(yscrollcommand=scrollbar.set)
                    
                    checkbox_vars = {}
                    
                    for cls in sorted(all_classes):
                        var = tk.BooleanVar(value=(cls in selected_classes_list))
                        checkbox_vars[cls] = var
                        
                        cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                        cb_row.pack(fill=tk.X, padx=20, pady=2)
                        
                        cb = tk.Checkbutton(
                            cb_row,
                            text=fix_rtl_text(cls),
                            variable=var,
                            font=('Arial', 11),
                            bg='#ecf0f1',
                            anchor='w'
                        )
                        cb.pack(side=tk.LEFT, fill=tk.X)
                    
                    canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                    
                    btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=(0, 12))
                    
                    def select_all():
                        for var in checkbox_vars.values():
                            var.set(True)
                    
                    def clear_all():
                        for var in checkbox_vars.values():
                            var.set(False)
                    
                    def apply_selection():
                        selected_classes_list.clear()
                        selected_classes_list.extend([cls for cls, var in checkbox_vars.items() if var.get()])
                        classes_entry.config(state='normal')
                        classes_entry.delete(0, tk.END)
                        classes_entry.insert(0, ', '.join(selected_classes_list) if selected_classes_list else '')
                        classes_entry.config(state='readonly')
                        if selected_classes_list:
                            general_var.set(0)
                        selector_dialog.destroy()
                    
                    tk.Button(
                        btn_frame,
                        text='בחר הכל',
                        command=select_all,
                        font=('Arial', 10),
                        bg='#3498db',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='נקה הכל',
                        command=clear_all,
                        font=('Arial', 10),
                        bg='#e74c3c',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='אישור',
                        command=apply_selection,
                        font=('Arial', 10, 'bold'),
                        bg='#27ae60',
                        fg='white',
                        padx=16,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                
                classes_btn = tk.Button(
                    classes_cell,
                    text='...',
                    command=open_classes_selector,
                    font=('Arial', 9),
                    bg='#3498db',
                    fg='white',
                    padx=4,
                    pady=0
                )
                classes_btn.pack(side=tk.LEFT)

                # Days selection with button
                days_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                days_cell.grid(row=row_index, column=2, sticky='e', padx=8, pady=4)

                days_entry = tk.Entry(days_cell, font=('Arial', 11), width=8, justify='right', state='readonly')
                days_entry.pack(side=tk.LEFT, padx=2)
                
                selected_days_list = []
                initial_days_str = str(initial.get('days_of_week') or '').strip()
                if initial_days_str:
                    selected_days_list.extend([day.strip() for day in initial_days_str.split(',') if day.strip()])

                all_days_var = tk.IntVar(value=1 if not selected_days_list else 0)
                days_entry.config(state='normal')
                if selected_days_list:
                    days_entry.insert(0, ','.join(selected_days_list))
                else:
                    days_entry.insert(0, 'כל')
                days_entry.config(state='readonly')
                
                def open_days_selector():
                    all_days = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ש']
                    
                    selector_dialog = tk.Toplevel(add_dialog)
                    selector_dialog.title('בחירת ימים')
                    selector_dialog.geometry('350x450')
                    selector_dialog.configure(bg='#ecf0f1')
                    selector_dialog.transient(add_dialog)
                    selector_dialog.grab_set()
                    
                    tk.Label(
                        selector_dialog,
                        text=fix_rtl_text('בחר ימים'),
                        font=('Arial', 14, 'bold'),
                        bg='#ecf0f1',
                        fg='#2c3e50'
                    ).pack(pady=(12, 6))
                    
                    cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                    
                    checkbox_vars = {}
                    
                    for day in all_days:
                        var = tk.BooleanVar(value=(day in selected_days_list if selected_days_list else True))
                        checkbox_vars[day] = var
                        
                        cb_row = tk.Frame(cb_frame, bg='#ecf0f1')
                        cb_row.pack(fill=tk.X, padx=20, pady=4)
                        
                        day_names = {'א': 'ראשון', 'ב': 'שני', 'ג': 'שלישי', 'ד': 'רביעי', 'ה': 'חמישי', 'ו': 'שישי', 'ש': 'שבת'}
                        
                        cb = tk.Checkbutton(
                            cb_row,
                            text=fix_rtl_text(f"{day_names.get(day, day)} ({day})"),
                            variable=var,
                            font=('Arial', 11),
                            bg='#ecf0f1',
                            anchor='w'
                        )
                        cb.pack(side=tk.LEFT, fill=tk.X)
                    
                    btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=(0, 12))
                    
                    def select_all():
                        for var in checkbox_vars.values():
                            var.set(True)
                    
                    def clear_all():
                        for var in checkbox_vars.values():
                            var.set(False)
                    
                    def apply_selection():
                        selected_days_list.clear()
                        selected_days_list.extend([day for day, var in checkbox_vars.items() if var.get()])
                        days_entry.config(state='normal')
                        days_entry.delete(0, tk.END)
                        try:
                            all_selected = (len(selected_days_list) == len(all_days))
                        except Exception:
                            all_selected = False
                        try:
                            all_days_var.set(1 if all_selected else 0)
                        except Exception:
                            pass
                        if all_selected or (not selected_days_list):
                            days_entry.insert(0, 'כל')
                        else:
                            days_entry.insert(0, ','.join(selected_days_list))
                        days_entry.config(state='readonly')
                        selector_dialog.destroy()
                    
                    tk.Button(
                        btn_frame,
                        text='בחר הכל',
                        command=select_all,
                        font=('Arial', 10),
                        bg='#3498db',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='נקה הכל',
                        command=clear_all,
                        font=('Arial', 10),
                        bg='#e74c3c',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='אישור',
                        command=apply_selection,
                        font=('Arial', 10, 'bold'),
                        bg='#27ae60',
                        fg='white',
                        padx=16,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                
                days_btn = tk.Button(
                    days_cell,
                    text='...',
                    command=open_days_selector,
                    font=('Arial', 9),
                    bg='#3498db',
                    fg='white',
                    padx=4,
                    pady=0
                )
                days_btn.pack(side=tk.LEFT)

                shown_var = tk.IntVar(value=int(initial.get('is_shown_public', 1) or 0))
                sound_var = tk.StringVar(value=str(initial.get('sound_key') or '').strip())

                sound_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                sound_cell.grid(row=row_index, column=1, sticky='e', padx=8, pady=4)

                shown_cb = tk.Checkbutton(sound_cell, variable=shown_var, bg='#ecf0f1')
                shown_cb.pack(side=tk.RIGHT)

                sound_values = [''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or [])
                sound_combo = ttk.Combobox(
                    sound_cell,
                    textvariable=sound_var,
                    values=sound_values,
                    state='readonly',
                    width=6,
                    justify='center'
                )
                sound_combo.pack(side=tk.RIGHT, padx=(0, 6))

                def _refresh_sound_values(_event=None, cb=sound_combo):
                    try:
                        cb.configure(values=[''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or []))
                    except Exception:
                        pass

                sound_combo.bind('<Button-1>', _refresh_sound_values)

                tk.Button(
                    sound_cell,
                    text="📁",
                    command=lambda v=sound_var: v.set(self._import_sound_file_to_folder(['לבונוס זמנים']) or v.get()),
                    font=('Arial', 10),
                    bg='#bdc3c7',
                    fg='black',
                    width=3,
                    cursor='hand2'
                ).pack(side=tk.RIGHT, padx=(0, 6))

                tk.Button(
                    sound_cell,
                    text="▶",
                    command=lambda v=sound_var: self._admin_play_sound_key(str(v.get() or '').strip()),
                    font=('Arial', 10, 'bold'),
                    bg='#bdc3c7',
                    fg='black',
                    width=3,
                    cursor='hand2'
                ).pack(side=tk.RIGHT, padx=(0, 6))

                del_btn = tk.Button(grid_frame, text="מחק", font=('Arial', 10), bg='#e74c3c', fg='white')
                del_btn.grid(row=row_index, column=0, sticky='e', padx=8, pady=4)

                def _sync_classes_state(*_):
                    try:
                        if int(general_var.get() or 0) == 1:
                            classes_entry.config(state='disabled')
                            classes_btn.config(state='disabled')
                        else:
                            classes_entry.config(state='normal')
                            classes_btn.config(state='normal')
                    except Exception:
                        pass

                try:
                    general_var.trace_add('write', _sync_classes_state)
                except Exception:
                    pass
                _sync_classes_state()

                def _reflow_rows():
                    for i, item in enumerate(row_widgets):
                        _, s_ent, e_ent, p_ent, g_var, g_cb_w, c_ent, d_ent, all_days_v, sh_var, sound_var, sound_cell_w, del_btn_w = item
                        rr = i + 1
                        try:
                            s_ent.grid_configure(row=rr, column=8)
                            e_ent.grid_configure(row=rr, column=7)
                            p_ent.grid_configure(row=rr, column=5)
                            g_cb_w.grid_configure(row=rr, column=4)
                            sound_cell_w.grid_configure(row=rr, column=1)
                            c_ent.grid_configure(row=rr, column=3)
                            d_ent.grid_configure(row=rr, column=2)
                            del_btn_w.grid_configure(row=rr, column=0)
                        except Exception:
                            pass

                def _on_delete_row():
                    try:
                        # find index by identity
                        idx_to_remove = None
                        for i, item in enumerate(row_widgets):
                            if item[1] is start_entry:
                                idx_to_remove = i
                                break
                        if idx_to_remove is None:
                            return
                        _, s_ent, e_ent, p_ent, g_var, g_cb_w, c_ent, d_ent, all_days_v, sh_var, sound_var, sound_cell_w, del_btn_w = row_widgets[idx_to_remove]
                        for w in [s_ent, e_ent, p_ent, c_ent, d_ent, sound_cell_w, del_btn_w]:
                            try:
                                w.destroy()
                            except Exception:
                                pass
                        try:
                            g_cb_w.destroy()
                        except Exception:
                            pass
                        del row_widgets[idx_to_remove]
                        _reflow_rows()
                        add_dialog.update_idletasks()
                        w = max(add_dialog.winfo_width(), add_dialog.winfo_reqwidth() + 40)
                        h = add_dialog.winfo_reqheight() + 40
                        add_dialog.geometry(f"{w}x{h}")
                        add_dialog.minsize(w, h)
                    except Exception:
                        pass

                del_btn.config(command=_on_delete_row)

                row_widgets.append((None, start_entry, end_entry, pts_entry, general_var, general_cb, classes_entry, days_entry, all_days_var, shown_var, sound_var, sound_cell, del_btn))
                try:
                    start_entry.focus_set()
                except Exception:
                    pass

            def _validate_rows() -> list:
                time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'

                out = []
                for _, s_ent, e_ent, p_ent, general_var, _general_cb, classes_ent, days_ent, all_days_var, shown_var, sound_var, _sound_cell_w, _del_btn in row_widgets:
                    start_raw = s_ent.get().strip()
                    end_raw = e_ent.get().strip()
                    start = _normalize_time_str(start_raw)
                    end = _normalize_time_str(end_raw)
                    try:
                        if start and start != start_raw:
                            s_ent.delete(0, tk.END)
                            s_ent.insert(0, start)
                    except Exception:
                        pass
                    try:
                        if end and end != end_raw:
                            e_ent.delete(0, tk.END)
                            e_ent.insert(0, end)
                    except Exception:
                        pass

                    if not re.match(time_pattern, start):
                        raise ValueError("פורמט 'משעה' לא תקין")
                    if not re.match(time_pattern, end):
                        raise ValueError("פורמט 'עד שעה' לא תקין")
                    if start >= end:
                        raise ValueError("'עד שעה' חייב להיות אחרי 'משעה'")
                    try:
                        pts = int(p_ent.get().strip())
                    except Exception:
                        raise ValueError("נקודות חייבות להיות מספר שלם")
                    if pts < 0:
                        raise ValueError("נקודות חייבות להיות 0 או יותר")
                    is_general = 1 if int(general_var.get() or 0) == 1 else 0
                    classes_text = str(classes_ent.get() or '').strip()
                    all_days = 1 if int(all_days_var.get() or 0) == 1 else 0
                    days_text = '' if all_days == 1 else str(days_ent.get() or '').strip()
                    is_shown = 1 if int(shown_var.get() or 0) == 1 else 0
                    try:
                        sound_key = str(sound_var.get() or '').strip()
                    except Exception:
                        sound_key = ''
                    if is_general == 1:
                        classes_text = ''
                    if all_days == 0 and not days_text:
                        raise ValueError("יש לסמן 'כל הימים' או להזין ימים")
                    out.append({'start_time': start, 'end_time': end, 'bonus_points': pts, 'is_general': is_general, 'classes': classes_text, 'days_of_week': days_text, 'is_shown_public': is_shown, 'sound_key': sound_key})
                if not out:
                    raise ValueError("יש להוסיף לפחות שורה אחת")
                out = sorted(out, key=lambda x: x.get('start_time') or '')

                # ולידציה פנימית: אין חפיפות לפי כללי כיתות בתוך הקבוצה שנבנית
                for i in range(len(out)):
                    for j in range(i + 1, len(out)):
                        a = out[i]
                        b = out[j]
                        # No conflict if weekday sets do not intersect
                        try:
                            a_days = _parse_days_list(a.get('days_of_week') or '')
                            b_days = _parse_days_list(b.get('days_of_week') or '')
                            if not a_days.intersection(b_days):
                                continue
                        except Exception:
                            pass
                        if not _ranges_overlap(a.get('start_time') or '', a.get('end_time') or '', b.get('start_time') or '', b.get('end_time') or ''):
                            continue
                        a_general = int(a.get('is_general', 0) or 0) == 1
                        b_general = int(b.get('is_general', 0) or 0) == 1
                        if a_general or b_general:
                            raise ValueError("אין אפשרות לחפיפה בשעות כאשר קיימת שורה 'כללי' באותה קבוצה")
                        a_classes = set(_parse_classes_list(a.get('classes') or ''))
                        b_classes = set(_parse_classes_list(b.get('classes') or ''))
                        if a_classes.intersection(b_classes):
                            raise ValueError("אין אפשרות לחפיפה בשעות עבור אותן כיתות באותה קבוצה")
                return out

            # שורה ראשונה
            _add_row({'start_time': '07:30', 'end_time': '07:45', 'bonus_points': 2, 'is_general': 1, 'classes': '', 'days_of_week': '', 'is_shown_public': 1, 'sound_key': ''})

            bottom = tk.Frame(container, bg='#ecf0f1')
            bottom.grid(row=2, column=0, columnspan=3, sticky='e', pady=(18, 0))

            def on_add_row():
                start_guess = '07:45'
                end_guess = '08:00'
                try:
                    if row_widgets:
                        last = row_widgets[-1]
                        last_end = ''
                        try:
                            last_end = str(last[2].get() or '').strip()
                        except Exception:
                            last_end = ''
                        last_end = _normalize_time_str(last_end)
                        next_start = _add_minutes(last_end, 1)
                        if next_start:
                            start_guess = next_start
                            # ברירת מחדל: טווח של 15 דקות קדימה (כדי ש"עד שעה" לא יישאר זהה)
                            next_end = _add_minutes(next_start, 15)
                            end_guess = next_end if next_end else _add_minutes(next_start, 1) or last_end
                except Exception:
                    pass

                _add_row({'start_time': start_guess, 'end_time': end_guess, 'bonus_points': 1, 'is_general': 1, 'classes': '', 'days_of_week': '', 'is_shown_public': 1, 'sound_key': ''})
                try:
                    add_dialog.update_idletasks()
                    w = max(add_dialog.winfo_width(), add_dialog.winfo_reqwidth() + 40)
                    h = add_dialog.winfo_reqheight() + 40
                    add_dialog.geometry(f"{w}x{h}")
                    add_dialog.minsize(w, h)
                except Exception:
                    pass

            def on_save():
                group_name = group_entry.get().strip()
                if not group_name:
                    messagebox.showerror("שגיאה", "יש להזין שם לבונוס")
                    return
                try:
                    rows = _validate_rows()
                except Exception as e:
                    messagebox.showerror("שגיאה", str(e))
                    return

                # ולידציה מול מסד נתונים (למקרה שיש כבר קבוצה באותו שם)
                try:
                    existing = self.db.get_all_time_bonuses()
                except Exception:
                    existing = []
                try:
                    for nr in rows:
                        _validate_time_bonus_conflicts(group_name, nr, existing)
                except Exception as e:
                    messagebox.showerror("שגיאה", str(e))
                    return
                try:
                    for r in rows:
                        self.db.add_time_bonus(
                            group_name,
                            r['start_time'],
                            r['end_time'],
                            int(r['bonus_points']),
                            group_name=group_name,
                            is_general=int(r.get('is_general', 1) or 0),
                            classes=(r.get('classes') or None),
                            days_of_week=(r.get('days_of_week') or None),
                            is_shown_public=int(r.get('is_shown_public', 1) or 0),
                            sound_key=(r.get('sound_key') or None),
                        )
                    load_bonuses()
                    add_dialog.destroy()
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בשמירת הבונוס: {e}")

            tk.Button(bottom, text="ביטול", command=add_dialog.destroy, font=('Arial', 11), bg='#95a5a6', fg='white', padx=14, pady=8).pack(side=tk.RIGHT)
            tk.Button(bottom, text="שמור", command=on_save, font=('Arial', 11, 'bold'), bg='#3498db', fg='white', padx=18, pady=8).pack(side=tk.RIGHT, padx=8)
            tk.Button(bottom, text="הוסף שורה", command=on_add_row, font=('Arial', 11), bg='#2ecc71', fg='white', padx=14, pady=8).pack(side=tk.RIGHT)

            group_entry.focus_set()

            try:
                add_dialog.update_idletasks()
                w = add_dialog.winfo_reqwidth() + 40
                h = add_dialog.winfo_reqheight() + 40
                add_dialog.geometry(f"{w}x{h}")
                add_dialog.minsize(w, h)
            except Exception:
                pass

        def _get_selected_group_name() -> str:
            sel = tree.selection()
            if not sel:
                return ""
            item_id = sel[0]
            tags = tree.item(item_id).get('tags') or ()
            for t in tags:
                if isinstance(t, str) and t.startswith('group:'):
                    return t.split(':', 1)[1]
            try:
                vals = tree.item(item_id).get('values') or ()
                if vals and len(vals) >= 1:
                    return str(vals[-1] or '').strip()
            except Exception:
                pass
            return ""

        def add_row_to_group():
            """הוספת שורה לבונוס קיים"""
            group_name = _get_selected_group_name()
            if not group_name:
                messagebox.showwarning("אזהרה", "בחר בונוס (קבוצה) או שורה מתוכו")
                return
            import re
            d = tk.Toplevel(dialog)
            d.title(f"הוסף שורה לבונוס: {group_name}")
            d.geometry("640x220")
            d.configure(bg='#ecf0f1')
            d.transient(dialog)
            d.grab_set()

            rf = tk.Frame(d, bg='#ecf0f1')
            rf.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)

            row_line = tk.Frame(rf, bg='#ecf0f1')
            row_line.pack(fill=tk.X, pady=8)

            shown_var = tk.IntVar(value=1)
            tk.Label(row_line, text=":מוצג", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)
            shown_cb = tk.Checkbutton(row_line, variable=shown_var, bg='#ecf0f1')
            shown_cb.pack(side=tk.RIGHT, padx=8)

            sound_var = tk.StringVar(value='')
            tk.Label(row_line, text=":צליל", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)
            sound_combo = ttk.Combobox(
                row_line,
                textvariable=sound_var,
                values=[''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or []),
                state='readonly',
                width=6,
                justify='center'
            )
            sound_combo.pack(side=tk.RIGHT, padx=8)

            def _refresh_sound_values(_event=None, cb=sound_combo):
                try:
                    cb.configure(values=[''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or []))
                except Exception:
                    pass

            sound_combo.bind('<Button-1>', _refresh_sound_values)

            tk.Button(
                row_line,
                text="📁",
                command=lambda v=sound_var: v.set(self._import_sound_file_to_folder(['לבונוס זמנים']) or v.get()),
                font=('Arial', 10),
                bg='#bdc3c7',
                fg='black',
                width=3,
                cursor='hand2'
            ).pack(side=tk.RIGHT, padx=4)

            tk.Button(
                row_line,
                text="▶",
                command=lambda v=sound_var: self._admin_play_sound_key(str(v.get() or '').strip()),
                font=('Arial', 10, 'bold'),
                bg='#bdc3c7',
                fg='black',
                width=3,
                cursor='hand2'
            ).pack(side=tk.RIGHT, padx=4)

            classes_entry = tk.Entry(row_line, font=('Arial', 12), width=18, justify='right')
            tk.Label(row_line, text=":כיתות (הפרד בפסיק)", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)
            classes_entry.pack(side=tk.RIGHT, padx=8)
            _apply_rtl_entry_behavior(classes_entry)

            days_entry = tk.Entry(row_line, font=('Arial', 12), width=14, justify='right')
            tk.Label(row_line, text=":ימים (א,ב,ג...)", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)
            days_entry.pack(side=tk.RIGHT, padx=6)
            _apply_rtl_entry_behavior(days_entry)

            all_days_var = tk.IntVar(value=1)
            all_days_cb = tk.Checkbutton(row_line, text=fix_rtl_text("כל הימים"), variable=all_days_var, bg='#ecf0f1')
            all_days_cb.pack(side=tk.RIGHT, padx=6)

            def _sync_days_state(*_):
                try:
                    if int(all_days_var.get() or 0) == 1:
                        days_entry.delete(0, tk.END)
                        days_entry.config(state='disabled')
                    else:
                        days_entry.config(state='normal')
                except Exception:
                    pass

            try:
                all_days_var.trace_add('write', _sync_days_state)
            except Exception:
                pass
            _sync_days_state()

            general_var = tk.IntVar(value=1)
            tk.Label(row_line, text=":כללי", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)
            general_cb = tk.Checkbutton(row_line, variable=general_var, bg='#ecf0f1')
            general_cb.pack(side=tk.RIGHT, padx=8)

            points_entry = tk.Entry(row_line, font=('Arial', 12), width=10, justify='right')
            points_entry.pack(side=tk.RIGHT, padx=8)
            points_entry.insert(0, '1')
            tk.Label(row_line, text=":נקודות בונוס", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)

            end_entry = tk.Entry(row_line, font=('Arial', 12), width=10, justify='right')
            end_entry.pack(side=tk.RIGHT, padx=8)
            end_entry.insert(0, '08:00')
            tk.Label(row_line, text=":עד שעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)

            start_entry = tk.Entry(row_line, font=('Arial', 12), width=10, justify='right')
            start_entry.pack(side=tk.RIGHT, padx=8)
            start_entry.insert(0, '07:45')
            tk.Label(row_line, text=":משעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').pack(side=tk.RIGHT, padx=8)

            def _sync_classes_state(*_):
                try:
                    if int(general_var.get() or 0) == 1:
                        classes_entry.config(state='disabled')
                    else:
                        classes_entry.config(state='normal')
                except Exception:
                    pass

            try:
                general_var.trace_add('write', _sync_classes_state)
            except Exception:
                pass

            try:
                _sync_classes_state()
            except Exception:
                pass

            # התאמת גודל החלון כך שלא ייפתח בגודל "ס"מ"
            try:
                d.update_idletasks()
                w2 = max(680, d.winfo_reqwidth() + 40)
                h2 = max(220, d.winfo_reqheight() + 40)
                d.geometry(f"{w2}x{h2}")
                d.minsize(w2, h2)
            except Exception:
                pass
            _sync_classes_state()

            btns = tk.Frame(container, bg='#ecf0f1')
            btns.pack(pady=10)

            def save_row():
                time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'
                start = start_entry.get().strip()
                end = end_entry.get().strip()
                if not re.match(time_pattern, start):
                    messagebox.showerror("שגיאה", "פורמט 'משעה' לא תקין")
                    return
                if not re.match(time_pattern, end):
                    messagebox.showerror("שגיאה", "פורמט 'עד שעה' לא תקין")
                    return
                if start >= end:
                    messagebox.showerror("שגיאה", "'עד שעה' חייב להיות אחרי 'משעה'")
                    return
                try:
                    pts = int(points_entry.get().strip())
                except Exception:
                    messagebox.showerror("שגיאה", "נקודות חייבות להיות מספר שלם")
                    return
                if pts < 0:
                    messagebox.showerror("שגיאה", "נקודות חייבות להיות 0 או יותר")
                    return

                is_general = 1 if int(general_var.get() or 0) == 1 else 0
                classes_text = str(classes_entry.get() or '').strip()
                all_days = 1 if int(all_days_var.get() or 0) == 1 else 0
                days_text = '' if all_days == 1 else str(days_entry.get() or '').strip()
                is_shown = 1 if int(shown_var.get() or 0) == 1 else 0
                try:
                    sound_key = str(sound_var.get() or '').strip()
                except Exception:
                    sound_key = ''
                if is_general == 1:
                    classes_text = ''
                if all_days == 0 and not days_text:
                    messagebox.showerror("שגיאה", "יש לסמן 'כל הימים' או להזין ימים")
                    return
                new_row = {
                    'start_time': start,
                    'end_time': end,
                    'bonus_points': pts,
                    'is_general': is_general,
                    'classes': classes_text,
                    'days_of_week': days_text,
                    'is_shown_public': is_shown,
                    'sound_key': sound_key,
                }
                try:
                    existing = self.db.get_all_time_bonuses()
                except Exception:
                    existing = []
                try:
                    _validate_time_bonus_conflicts(group_name, new_row, existing)
                except Exception as e:
                    messagebox.showerror("שגיאה", str(e))
                    return
                try:
                    self.db.add_time_bonus(
                        group_name,
                        start,
                        end,
                        pts,
                        group_name=group_name,
                        is_general=is_general,
                        classes=(classes_text or None),
                        days_of_week=(days_text or None),
                        is_shown_public=is_shown,
                        sound_key=(sound_key or None),
                    )
                    load_bonuses()
                    d.destroy()
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בהוספת שורה: {e}")

            tk.Button(btns, text="שמור", width=10, command=save_row, bg='#27ae60', fg='white').pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text="ביטול", width=10, command=d.destroy, bg='#95a5a6', fg='white').pack(side=tk.LEFT, padx=6)

            start_entry.focus_set()
        
        def toggle_active():
            """החלפת סטטוס פעיל/לא פעיל: שורה/שורות נבחרות, או כל הבונוס (לפי אישור)."""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור שורה")
                return

            bonuses = self.db.get_all_time_bonuses()

            row_ids = []
            groups = set()
            for item_id in selected:
                tags = tree.item(item_id).get('tags') or ()
                for t in tags:
                    if isinstance(t, str) and t.startswith('row:'):
                        try:
                            row_ids.append(int(t.split(':', 1)[1]))
                        except Exception:
                            pass
                    if isinstance(t, str) and t.startswith('group:'):
                        groups.add(t.split(':', 1)[1])

            if not row_ids:
                return

            # אם נבחרה שורה אחת בלבד – נציע להפעיל/לכבות את כל הבונוס (כל השורות של הקבוצה)
            if len(row_ids) == 1:
                group_name = _get_selected_group_name()
                if group_name:
                    try:
                        if messagebox.askyesno("פעיל/כבוי", "לשנות סטטוס לכל הבונוס (כל השורות)?"):
                            rows = [b for b in bonuses if (b.get('group_name') or b.get('name') or '').strip() == group_name]
                            if rows:
                                any_active = any(int(r.get('is_active', 0) or 0) == 1 for r in rows)
                                new_active = 0 if any_active else 1
                                for r in rows:
                                    self.db.update_time_bonus(
                                        int(r['id']),
                                        r.get('name') or group_name,
                                        r.get('start_time') or '',
                                        r.get('end_time') or '',
                                        int(r.get('bonus_points', 0) or 0),
                                        new_active,
                                        group_name=group_name,
                                        is_general=int(r.get('is_general', 1) or 0),
                                        classes=(r.get('classes') or None),
                                        days_of_week=(r.get('days_of_week') or None),
                                        is_shown_public=int(r.get('is_shown_public', 1) or 0),
                                    )
                                load_bonuses()
                                return
                    except Exception:
                        pass

            # מצב רגיל: החלפה רק לשורות הנבחרות
            for rid in row_ids:
                current = next((b for b in bonuses if int(b.get('id', 0) or 0) == rid), None)
                if not current:
                    continue
                group_name = (current.get('group_name') or current.get('name') or '').strip()
                new_active = 0 if int(current.get('is_active', 0) or 0) == 1 else 1
                self.db.update_time_bonus(
                    rid,
                    current.get('name') or group_name,
                    current.get('start_time') or '',
                    current.get('end_time') or '',
                    int(current.get('bonus_points', 0) or 0),
                    new_active,
                    group_name=group_name,
                    is_general=int(current.get('is_general', 1) or 0),
                    classes=(current.get('classes') or None),
                    days_of_week=(current.get('days_of_week') or None),
                    is_shown_public=int(current.get('is_shown_public', 1) or 0),
                )
            load_bonuses()
        
        def delete_bonus():
            """מחיקה: שורה/שורות נבחרות, או כל הבונוס (לפי אישור)."""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור שורה")
                return

            bonuses = self.db.get_all_time_bonuses()

            row_ids = []
            for item_id in selected:
                tags = tree.item(item_id).get('tags') or ()
                for t in tags:
                    if isinstance(t, str) and t.startswith('row:'):
                        try:
                            row_ids.append(int(t.split(':', 1)[1]))
                        except Exception:
                            pass

            if not row_ids:
                messagebox.showwarning("אזהרה", "יש לבחור שורה")
                return

            group_name = _get_selected_group_name()

            confirm = tk.Toplevel(dialog)
            confirm.title("מחיקה")
            confirm.geometry("420x180")
            try:
                confirm.minsize(420, 180)
            except Exception:
                pass
            confirm.configure(bg='#ecf0f1')
            confirm.transient(dialog)
            confirm.grab_set()

            msg = "למחוק?"
            if group_name:
                msg = f"למחוק?\n\nקבוצה: {group_name}"
            tk.Label(confirm, text=msg, font=('Arial', 12, 'bold'), bg='#ecf0f1', justify='right').pack(pady=18, padx=18, anchor='e')

            btns = tk.Frame(confirm, bg='#ecf0f1')
            btns.pack(pady=10)

            def _delete_selected_rows():
                try:
                    if not row_ids:
                        confirm.destroy()
                        return
                    for rid in row_ids:
                        self.db.delete_time_bonus(int(rid))
                    load_bonuses()
                finally:
                    try:
                        confirm.destroy()
                    except Exception:
                        pass

            def _delete_entire_group():
                try:
                    if not group_name:
                        confirm.destroy()
                        return
                    rows = [b for b in bonuses if (b.get('group_name') or b.get('name') or '').strip() == group_name]
                    for r in rows:
                        self.db.delete_time_bonus(int(r['id']))
                    load_bonuses()
                finally:
                    try:
                        confirm.destroy()
                    except Exception:
                        pass

            tk.Button(btns, text="מחק שורה", width=12, command=_delete_selected_rows, bg='#e74c3c', fg='white').pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text="מחק את כל הבונוס", width=16, command=_delete_entire_group, bg='#c0392b', fg='white').pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text="ביטול", width=10, command=confirm.destroy, bg='#95a5a6', fg='white').pack(side=tk.LEFT, padx=6)
        
        def edit_bonus():
            """עריכת שורה (משעה/עד שעה/נקודות)."""
            selected = tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור שורה לעריכה")
                return
            item_id = selected[0]
            tags = tree.item(item_id).get('tags') or ()
            row_id = None
            for t in tags:
                if isinstance(t, str) and t.startswith('row:'):
                    row_id = int(t.split(':', 1)[1])
                    break
            if not row_id:
                messagebox.showwarning("אזהרה", "יש לבחור שורה")
                return

            bonuses = self.db.get_all_time_bonuses()
            current = next((b for b in bonuses if int(b.get('id', 0) or 0) == row_id), None)
            if not current:
                messagebox.showerror("שגיאה", "לא נמצא שורה לעריכה")
                return

            import re
            group_name = (current.get('group_name') or current.get('name') or '').strip()

            d = tk.Toplevel(dialog)
            d.title(f"עריכת שורה בבונוס: {group_name}")
            d.geometry("1x1")
            d.configure(bg='#ecf0f1')
            d.transient(dialog)
            d.grab_set()

            container = tk.Frame(d, bg='#ecf0f1')
            container.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)

            # כמו חלון הוספה: כותרות למעלה
            grid_frame = tk.Frame(container, bg='#ecf0f1')
            grid_frame.pack(fill=tk.X)

            for c in range(7):
                grid_frame.grid_columnconfigure(c, weight=0)

            tk.Label(grid_frame, text=":משעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=6, sticky='e', padx=8)
            tk.Label(grid_frame, text=":עד שעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=5, sticky='e', padx=8)
            tk.Label(grid_frame, text=":נקודות בונוס", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=4, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כללי", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=3, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כיתות (הפרד בפסיק)", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=2, sticky='e', padx=8)
            tk.Label(grid_frame, text=":ימים (א,ב,ג...)", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=1, sticky='e', padx=8)
            tk.Label(grid_frame, text=":מוצג", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=0, sticky='e', padx=8)

            start_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
            start_entry.grid(row=1, column=6, sticky='e', padx=8, pady=6)
            start_entry.insert(0, current.get('start_time') or '')

            end_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
            end_entry.grid(row=1, column=5, sticky='e', padx=8, pady=6)
            end_entry.insert(0, current.get('end_time') or '')

            points_entry = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
            points_entry.grid(row=1, column=4, sticky='e', padx=8, pady=6)
            points_entry.insert(0, str(int(current.get('bonus_points', 0) or 0)))

            general_var = tk.IntVar(value=int(current.get('is_general', 1) or 0))
            general_cb = tk.Checkbutton(grid_frame, variable=general_var, bg='#ecf0f1')
            general_cb.grid(row=1, column=3, sticky='e', padx=8, pady=6)

            classes_entry = tk.Entry(grid_frame, font=('Arial', 12), width=18, justify='right')
            classes_entry.grid(row=1, column=2, sticky='e', padx=8, pady=6)
            classes_entry.insert(0, str(current.get('classes') or ''))
            _apply_rtl_entry_behavior(classes_entry)

            days_entry = tk.Entry(grid_frame, font=('Arial', 12), width=14, justify='right')
            days_entry.grid(row=1, column=1, sticky='e', padx=8, pady=6)
            days_entry.insert(0, str(current.get('days_of_week') or ''))
            _apply_rtl_entry_behavior(days_entry)

            all_days_var = tk.IntVar(value=1 if not str(current.get('days_of_week') or '').strip() else 0)
            all_days_cb = tk.Checkbutton(grid_frame, text=fix_rtl_text("כל הימים"), variable=all_days_var, bg='#ecf0f1')
            all_days_cb.grid(row=2, column=1, sticky='e', padx=8, pady=(0, 6))

            def _sync_days_state(*_):
                try:
                    if int(all_days_var.get() or 0) == 1:
                        days_entry.delete(0, tk.END)
                        days_entry.config(state='disabled')
                    else:
                        days_entry.config(state='normal')
                except Exception:
                    pass

            try:
                all_days_var.trace_add('write', _sync_days_state)
            except Exception:
                pass
            _sync_days_state()

            shown_var = tk.IntVar(value=int(current.get('is_shown_public', 1) or 0))
            shown_cb = tk.Checkbutton(grid_frame, variable=shown_var, bg='#ecf0f1')
            shown_cb.grid(row=1, column=0, sticky='e', padx=8, pady=6)

            def _sync_classes_state(*_):
                try:
                    if int(general_var.get() or 0) == 1:
                        classes_entry.config(state='disabled')
                    else:
                        classes_entry.config(state='normal')
                except Exception:
                    pass

            try:
                general_var.trace_add('write', _sync_classes_state)
            except Exception:
                pass
            _sync_classes_state()

            btns = tk.Frame(rf, bg='#ecf0f1')
            btns.pack(pady=10)

            def save_edit():
                time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'
                start = start_entry.get().strip()
                end = end_entry.get().strip()
                if not re.match(time_pattern, start):
                    messagebox.showerror("שגיאה", "פורמט 'משעה' לא תקין")
                    return
                if not re.match(time_pattern, end):
                    messagebox.showerror("שגיאה", "פורמט 'עד שעה' לא תקין")
                    return
                if start >= end:
                    messagebox.showerror("שגיאה", "'עד שעה' חייב להיות אחרי 'משעה'")
                    return
                try:
                    pts = int(points_entry.get().strip())
                except Exception:
                    messagebox.showerror("שגיאה", "נקודות חייבות להיות מספר שלם")
                    return
                if pts < 0:
                    messagebox.showerror("שגיאה", "נקודות חייבות להיות 0 או יותר")
                    return

                is_general = 1 if int(general_var.get() or 0) == 1 else 0
                classes_text = str(classes_entry.get() or '').strip()
                all_days = 1 if int(all_days_var.get() or 0) == 1 else 0
                days_text = '' if all_days == 1 else str(days_entry.get() or '').strip()
                is_shown = 1 if int(shown_var.get() or 0) == 1 else 0
                if is_general == 1:
                    classes_text = ''
                if all_days == 0 and not days_text:
                    messagebox.showerror("שגיאה", "יש לסמן 'כל הימים' או להזין ימים")
                    return

                new_row = {
                    'start_time': start,
                    'end_time': end,
                    'bonus_points': pts,
                    'is_general': is_general,
                    'classes': classes_text,
                    'days_of_week': days_text,
                    'is_shown_public': is_shown,
                }
                try:
                    existing = self.db.get_all_time_bonuses()
                except Exception:
                    existing = []
                try:
                    _validate_time_bonus_conflicts(group_name, new_row, existing, exclude_row_id=row_id)
                except Exception as e:
                    messagebox.showerror("שגיאה", str(e))
                    return
                try:
                    self.db.update_time_bonus(
                        row_id,
                        current.get('name') or group_name,
                        start,
                        end,
                        pts,
                        int(current.get('is_active', 1) or 0),
                        group_name=group_name,
                        is_general=is_general,
                        classes=(classes_text or None),
                        days_of_week=(days_text or None),
                        is_shown_public=is_shown,
                    )
                    load_bonuses()
                    d.destroy()
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בעדכון שורה: {e}")

            tk.Button(btns, text="שמור", width=10, command=save_edit, bg='#27ae60', fg='white').pack(side=tk.LEFT, padx=6)
            tk.Button(btns, text="ביטול", width=10, command=d.destroy, bg='#95a5a6', fg='white').pack(side=tk.LEFT, padx=6)

            start_entry.focus_set()

            try:
                d.update_idletasks()
                w = d.winfo_reqwidth() + 40
                h = d.winfo_reqheight() + 40
                d.geometry(f"{w}x{h}")
                d.minsize(w, h)
            except Exception:
                pass
        
        # כפתורים
        btn_frame = tk.Frame(dialog, bg='#ecf0f1')
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="➕ הוסף בונוס",
            command=add_group,
            font=('Arial', 11, 'bold'),
            bg='#27ae60',
            fg='white',
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)

        def edit_group():
            group_name = _get_selected_group_name()
            if not group_name:
                messagebox.showwarning("אזהרה", "בחר בונוס (קבוצה) או שורה מתוכו")
                return

            bonuses = self.db.get_all_time_bonuses()
            rows = [b for b in bonuses if (b.get('group_name') or b.get('name') or '').strip() == group_name]
            if not rows:
                messagebox.showerror("שגיאה", "לא נמצאה קבוצה לעריכה")
                return

            import re

            d = tk.Toplevel(dialog)
            d.title(f"עריכת בונוס: {group_name}")
            d.geometry("1x1")
            d.configure(bg='#ecf0f1')
            d.transient(dialog)
            d.grab_set()

            container = tk.Frame(d, bg='#ecf0f1')
            container.pack(fill=tk.BOTH, expand=True, padx=14, pady=14)

            container.grid_columnconfigure(0, weight=0)
            container.grid_columnconfigure(1, weight=1)
            container.grid_columnconfigure(2, weight=0)

            tk.Label(container, text=":בונוס", font=('Arial', 12, 'bold'), bg='#ecf0f1').grid(row=0, column=2, sticky='e', padx=(8, 0), pady=(0, 10))
            group_entry = tk.Entry(container, font=('Arial', 13), width=28, justify='right')
            group_entry.grid(row=0, column=1, sticky='e', pady=(0, 10))
            group_entry.insert(0, group_name)

            grid_frame = tk.Frame(container, bg='#ecf0f1')
            grid_frame.grid(row=1, column=0, columnspan=3, sticky='nsew')

            for c in range(8):
                grid_frame.grid_columnconfigure(c, weight=0)

            tk.Label(grid_frame, text=":משעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=7, sticky='e', padx=8)
            tk.Label(grid_frame, text=":עד שעה", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=6, sticky='e', padx=8)
            tk.Label(grid_frame, text=":נקודות בונוס", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=5, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כללי", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=4, sticky='e', padx=8)
            tk.Label(grid_frame, text=":כיתות", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=3, sticky='e', padx=8)
            tk.Label(grid_frame, text=":ימים", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=2, sticky='e', padx=8)
            tk.Label(grid_frame, text=":מוצג", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=1, sticky='e', padx=8)
            tk.Label(grid_frame, text=":מחק", font=('Arial', 11, 'bold'), bg='#ecf0f1').grid(row=0, column=0, sticky='e', padx=8)

            row_widgets = []

            def _reflow_rows():
                for i, item in enumerate(row_widgets):
                    s_ent, e_ent, p_ent, general_var, g_cb_w, c_ent, d_ent, all_days_var, shown_var, sound_var, sound_cell_w, del_btn_w = item
                    rr = i + 1
                    try:
                        s_ent.grid_configure(row=rr, column=7)
                        e_ent.grid_configure(row=rr, column=6)
                        p_ent.grid_configure(row=rr, column=5)
                        g_cb_w.grid_configure(row=rr, column=4)
                    except Exception:
                        pass
                    try:
                        c_ent.master.grid_configure(row=rr, column=3)
                    except Exception:
                        pass
                    try:
                        d_ent.master.grid_configure(row=rr, column=2)
                    except Exception:
                        pass
                    try:
                        sound_cell_w.grid_configure(row=rr, column=1)
                    except Exception:
                        pass
                    try:
                        del_btn_w.grid_configure(row=rr, column=0)
                    except Exception:
                        pass

            def _add_row(initial=None):
                initial = initial or {}
                rr = len(row_widgets) + 1
                s_ent = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                s_ent.grid(row=rr, column=7, sticky='e', padx=8, pady=4)
                s_ent.insert(0, initial.get('start_time') or '07:30')

                e_ent = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                e_ent.grid(row=rr, column=6, sticky='e', padx=8, pady=4)
                e_ent.insert(0, initial.get('end_time') or '07:45')

                p_ent = tk.Entry(grid_frame, font=('Arial', 12), width=10, justify='right')
                p_ent.grid(row=rr, column=5, sticky='e', padx=8, pady=4)
                p_ent.insert(0, str(initial.get('bonus_points', 1)))

                general_var = tk.IntVar(value=int(initial.get('is_general', 1) or 0))
                g_cb = tk.Checkbutton(grid_frame, variable=general_var, bg='#ecf0f1')
                g_cb.grid(row=rr, column=4, sticky='e', padx=8, pady=4)

                # Classes selection with button
                classes_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                classes_cell.grid(row=rr, column=3, sticky='e', padx=8, pady=4)
                
                c_ent = tk.Entry(classes_cell, font=('Arial', 11), width=10, justify='right', state='readonly')
                c_ent.pack(side=tk.LEFT, padx=2)
                
                selected_classes_list = []
                initial_classes_str = str(initial.get('classes') or '').strip()
                if initial_classes_str:
                    selected_classes_list.extend([c.strip() for c in initial_classes_str.split(',') if c.strip()])
                    c_ent.config(state='normal')
                    c_ent.insert(0, ', '.join(selected_classes_list))
                    c_ent.config(state='readonly')
                
                def open_classes_selector():
                    nonlocal d
                    all_classes = set()
                    try:
                        students = self.db.get_all_students()
                        for s in students:
                            cn = (s.get('class_name') or '').strip()
                            if cn:
                                all_classes.add(cn)
                    except Exception:
                        pass
                    
                    if not all_classes:
                        messagebox.showinfo('מידע', 'אין כיתות במערכת')
                        return
                    
                    selector_dialog = tk.Toplevel(d)
                    selector_dialog.title('בחירת כיתות')
                    selector_dialog.geometry('400x500')
                    selector_dialog.configure(bg='#ecf0f1')
                    selector_dialog.transient(d)
                    selector_dialog.grab_set()
                    
                    tk.Label(
                        selector_dialog,
                        text=fix_rtl_text('בחר כיתות'),
                        font=('Arial', 14, 'bold'),
                        bg='#ecf0f1',
                        fg='#2c3e50'
                    ).pack(pady=(12, 6))
                    
                    cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                    
                    canvas = tk.Canvas(cb_frame, bg='#ecf0f1', highlightthickness=0)
                    scrollbar = ttk.Scrollbar(cb_frame, orient=tk.VERTICAL, command=canvas.yview)
                    scrollable_frame = tk.Frame(canvas, bg='#ecf0f1')
                    
                    scrollable_frame.bind(
                        "<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
                    )
                    
                    canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
                    canvas.configure(yscrollcommand=scrollbar.set)
                    
                    checkbox_vars = {}
                    
                    for cls in sorted(all_classes):
                        var = tk.BooleanVar(value=(cls in selected_classes_list))
                        checkbox_vars[cls] = var
                        
                        cb_row = tk.Frame(scrollable_frame, bg='#ecf0f1')
                        cb_row.pack(fill=tk.X, padx=20, pady=2)
                        
                        cb = tk.Checkbutton(
                            cb_row,
                            text=fix_rtl_text(cls),
                            variable=var,
                            font=('Arial', 11),
                            bg='#ecf0f1',
                            anchor='w'
                        )
                        cb.pack(side=tk.LEFT, fill=tk.X)
                    
                    canvas.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                    scrollbar.pack(side=tk.LEFT, fill=tk.Y)
                    
                    btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=(0, 12))
                    
                    def select_all():
                        for var in checkbox_vars.values():
                            var.set(True)
                    
                    def clear_all():
                        for var in checkbox_vars.values():
                            var.set(False)
                    
                    def apply_selection():
                        selected = [cls for cls, var in checkbox_vars.items() if var.get()]
                        selected_classes_list.clear()
                        selected_classes_list.extend(selected)
                        c_ent.config(state='normal')
                        c_ent.delete(0, tk.END)
                        c_ent.insert(0, ', '.join(selected) if selected else '')
                        c_ent.config(state='readonly')
                        if selected:
                            general_var.set(0)
                        selector_dialog.destroy()
                    
                    tk.Button(
                        btn_frame,
                        text='בחר הכל',
                        command=select_all,
                        font=('Arial', 10),
                        bg='#3498db',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='נקה הכל',
                        command=clear_all,
                        font=('Arial', 10),
                        bg='#e74c3c',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='אישור',
                        command=apply_selection,
                        font=('Arial', 10, 'bold'),
                        bg='#27ae60',
                        fg='white',
                        padx=16,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                
                classes_btn = tk.Button(classes_cell, text='בחר...', command=open_classes_selector, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=8, pady=2)
                classes_btn.pack(side=tk.LEFT, padx=2)

                # Days selection with button
                days_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                days_cell.grid(row=rr, column=2, sticky='e', padx=8, pady=4)
                
                d_ent = tk.Entry(days_cell, font=('Arial', 11), width=8, justify='right', state='readonly')
                d_ent.pack(side=tk.LEFT, padx=2)
                
                selected_days_list = []
                initial_days_str = str(initial.get('days_of_week') or '').strip()
                if initial_days_str:
                    selected_days_list.extend([day.strip() for day in initial_days_str.split(',') if day.strip()])
                
                d_ent.config(state='normal')
                if selected_days_list:
                    d_ent.insert(0, ','.join(selected_days_list))
                else:
                    d_ent.insert(0, 'כל')
                d_ent.config(state='readonly')

                all_days_var = tk.IntVar(value=0 if selected_days_list else 1)
                
                def open_days_selector():
                    nonlocal d
                    all_days = ['א', 'ב', 'ג', 'ד', 'ה', 'ו', 'ש']
                    
                    selector_dialog = tk.Toplevel(d)
                    selector_dialog.title('בחירת ימים')
                    selector_dialog.geometry('350x400')
                    selector_dialog.configure(bg='#ecf0f1')
                    selector_dialog.transient(d)
                    selector_dialog.grab_set()
                    
                    tk.Label(
                        selector_dialog,
                        text=fix_rtl_text('בחר ימים'),
                        font=('Arial', 14, 'bold'),
                        bg='#ecf0f1',
                        fg='#2c3e50'
                    ).pack(pady=(12, 6))
                    
                    cb_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    cb_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                    
                    checkbox_vars = {}
                    day_names = {'א': 'ראשון', 'ב': 'שני', 'ג': 'שלישי', 'ד': 'רביעי', 'ה': 'חמישי', 'ו': 'שישי', 'ש': 'שבת'}
                    
                    for day in all_days:
                        var = tk.BooleanVar(value=(day in selected_days_list if selected_days_list else True))
                        checkbox_vars[day] = var
                        
                        cb_row = tk.Frame(cb_frame, bg='#ecf0f1')
                        cb_row.pack(fill=tk.X, padx=20, pady=4)
                        
                        cb = tk.Checkbutton(
                            cb_row,
                            text=fix_rtl_text(f"{day_names.get(day, day)} ({day})"),
                            variable=var,
                            font=('Arial', 11),
                            bg='#ecf0f1',
                            anchor='w'
                        )
                        cb.pack(side=tk.LEFT, fill=tk.X)
                    
                    btn_frame = tk.Frame(selector_dialog, bg='#ecf0f1')
                    btn_frame.pack(pady=(0, 12))
                    
                    def select_all():
                        for var in checkbox_vars.values():
                            var.set(True)
                    
                    def clear_all():
                        for var in checkbox_vars.values():
                            var.set(False)
                    
                    def apply_selection():
                        selected = [day for day, var in checkbox_vars.items() if var.get()]
                        selected_days_list.clear()
                        selected_days_list.extend(selected)
                        d_ent.config(state='normal')
                        d_ent.delete(0, tk.END)
                        if selected:
                            d_ent.insert(0, ','.join(selected))
                            try:
                                all_days_var.set(0)
                            except Exception:
                                pass
                        else:
                            d_ent.insert(0, 'כל')
                            try:
                                all_days_var.set(1)
                            except Exception:
                                pass
                        d_ent.config(state='readonly')
                        selector_dialog.destroy()
                    
                    tk.Button(
                        btn_frame,
                        text='בחר הכל',
                        command=select_all,
                        font=('Arial', 10),
                        bg='#3498db',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='נקה הכל',
                        command=clear_all,
                        font=('Arial', 10),
                        bg='#e74c3c',
                        fg='white',
                        padx=12,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                    
                    tk.Button(
                        btn_frame,
                        text='אישור',
                        command=apply_selection,
                        font=('Arial', 10, 'bold'),
                        bg='#27ae60',
                        fg='white',
                        padx=16,
                        pady=4
                    ).pack(side=tk.RIGHT, padx=4)
                
                days_btn = tk.Button(days_cell, text='בחר...', command=open_days_selector, font=('Arial', 9, 'bold'), bg='#3498db', fg='white', padx=8, pady=2)
                days_btn.pack(side=tk.LEFT, padx=2)

                shown_var = tk.IntVar(value=int(initial.get('is_shown_public', 1) or 0))
                sound_var = tk.StringVar(value=str(initial.get('sound_key') or '').strip())

                sound_cell = tk.Frame(grid_frame, bg='#ecf0f1')
                sound_cell.grid(row=rr, column=1, sticky='e', padx=8, pady=4)

                sh_cb = tk.Checkbutton(sound_cell, variable=shown_var, bg='#ecf0f1')
                sh_cb.pack(side=tk.RIGHT)

                sound_values = [''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or [])
                sound_combo = ttk.Combobox(
                    sound_cell,
                    textvariable=sound_var,
                    values=sound_values,
                    state='readonly',
                    width=6,
                    justify='center'
                )
                sound_combo.pack(side=tk.RIGHT, padx=(0, 6))

                def _refresh_sound_values(_event=None, cb=sound_combo):
                    try:
                        cb.configure(values=[''] + list(self._list_sound_keys_in_folder(['לבונוס זמנים']) or []))
                    except Exception:
                        pass

                sound_combo.bind('<Button-1>', _refresh_sound_values)

                tk.Button(
                    sound_cell,
                    text="📁",
                    command=lambda v=sound_var: v.set(self._import_sound_file_to_folder(['לבונוס זמנים']) or v.get()),
                    font=('Arial', 10),
                    bg='#bdc3c7',
                    fg='black',
                    width=3,
                    cursor='hand2'
                ).pack(side=tk.RIGHT, padx=(0, 6))

                tk.Button(
                    sound_cell,
                    text="▶",
                    command=lambda v=sound_var: self._admin_play_sound_key(str(v.get() or '').strip()),
                    font=('Arial', 10, 'bold'),
                    bg='#bdc3c7',
                    fg='black',
                    width=3,
                    cursor='hand2'
                ).pack(side=tk.RIGHT, padx=(0, 6))

                del_btn = tk.Button(grid_frame, text="מחק", font=('Arial', 10), bg='#e74c3c', fg='white')
                del_btn.grid(row=rr, column=0, sticky='e', padx=8, pady=4)

                def _sync_classes_state(*_):
                    try:
                        if int(general_var.get() or 0) == 1:
                            classes_btn.config(state='disabled')
                        else:
                            classes_btn.config(state='normal')
                    except Exception:
                        pass

                try:
                    general_var.trace_add('write', _sync_classes_state)
                except Exception:
                    pass
                _sync_classes_state()

                def _on_delete():
                    try:
                        idx_to_remove = None
                        for i2, it in enumerate(row_widgets):
                            if it[0] is s_ent:
                                idx_to_remove = i2
                                break
                        if idx_to_remove is None:
                            return
                        s2, e2, p2, _gv2, g2, c2, d2, _all_days2, _shown2, _sound2, sound_cell2, db2 = row_widgets[idx_to_remove]
                        for w in [s2, e2, p2, g2]:
                            try:
                                w.destroy()
                            except Exception:
                                pass
                        try:
                            c2.master.destroy()
                        except Exception:
                            pass
                        try:
                            d2.master.destroy()
                        except Exception:
                            pass
                        try:
                            sound_cell2.destroy()
                        except Exception:
                            pass
                        try:
                            db2.destroy()
                        except Exception:
                            pass
                        del row_widgets[idx_to_remove]
                        _reflow_rows()
                    except Exception:
                        pass

                del_btn.config(command=_on_delete)

                row_widgets.append((s_ent, e_ent, p_ent, general_var, g_cb, c_ent, d_ent, all_days_var, shown_var, sound_var, sound_cell, del_btn))

            # populate existing group rows
            for r in rows:
                _add_row({
                    'start_time': r.get('start_time') or '',
                    'end_time': r.get('end_time') or '',
                    'bonus_points': int(r.get('bonus_points', 0) or 0),
                    'is_general': int(r.get('is_general', 1) or 0),
                    'classes': r.get('classes') or '',
                    'days_of_week': r.get('days_of_week') or '',
                    'is_shown_public': int(r.get('is_shown_public', 1) or 0),
                    'sound_key': (r.get('sound_key') or ''),
                })

            bottom = tk.Frame(container, bg='#ecf0f1')
            bottom.grid(row=2, column=0, columnspan=3, sticky='e', pady=(18, 0))

            def on_add_row():
                _add_row({'start_time': '07:45', 'end_time': '08:00', 'bonus_points': 1, 'is_general': 1, 'classes': '', 'is_shown_public': 1, 'sound_key': ''})
                try:
                    d.update_idletasks()
                    w2 = max(d.winfo_width(), d.winfo_reqwidth() + 40)
                    h2 = d.winfo_reqheight() + 40
                    d.geometry(f"{w2}x{h2}")
                    d.minsize(w2, h2)
                except Exception:
                    pass

            def on_save_group():
                time_pattern = r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$'
                new_group_name = (group_entry.get() or '').strip() or group_name

                new_rows = []
                for s_ent, e_ent, p_ent, general_var, _g_cb_w, c_ent, d_ent, all_days_var, shown_var, sound_var, _sound_cell_w, _del_btn_w in row_widgets:
                    start = (s_ent.get() or '').strip()
                    end = (e_ent.get() or '').strip()
                    if not re.match(time_pattern, start):
                        messagebox.showerror("שגיאה", "פורמט 'משעה' לא תקין")
                        return
                    if not re.match(time_pattern, end):
                        messagebox.showerror("שגיאה", "פורמט 'עד שעה' לא תקין")
                        return
                    if start >= end:
                        messagebox.showerror("שגיאה", "'עד שעה' חייב להיות אחרי 'משעה'")
                        return
                    try:
                        pts = int((p_ent.get() or '').strip())
                    except Exception:
                        messagebox.showerror("שגיאה", "נקודות חייבות להיות מספר שלם")
                        return
                    if pts < 0:
                        messagebox.showerror("שגיאה", "נקודות חייבות להיות 0 או יותר")
                        return

                    is_general = 1 if int(general_var.get() or 0) == 1 else 0
                    is_shown = 1 if int(shown_var.get() or 0) == 1 else 0

                    try:
                        sound_key = str(sound_var.get() or '').strip()
                    except Exception:
                        sound_key = ''

                    classes_text = str(c_ent.get() or '').strip()
                    if is_general == 1:
                        classes_text = ''

                    all_days = 1 if int(all_days_var.get() or 0) == 1 else 0
                    days_text = '' if all_days == 1 else str(d_ent.get() or '').strip()
                    if all_days == 0 and not days_text:
                        messagebox.showerror("שגיאה", "יש לסמן 'כל הימים' או להזין ימים")
                        return

                    new_rows.append({
                        'start_time': start,
                        'end_time': end,
                        'bonus_points': pts,
                        'is_general': is_general,
                        'classes': classes_text,
                        'days_of_week': days_text,
                        'is_shown_public': is_shown,
                        'sound_key': sound_key,
                    })

                if not new_rows:
                    messagebox.showerror("שגיאה", "יש להוסיף לפחות שורה אחת")
                    return

                # internal overlap validation
                try:
                    new_rows_sorted = sorted(new_rows, key=lambda x: x.get('start_time') or '')
                    for i in range(len(new_rows_sorted)):
                        for j in range(i + 1, len(new_rows_sorted)):
                            a = new_rows_sorted[i]
                            b = new_rows_sorted[j]
                            # No conflict if weekday sets do not intersect
                            try:
                                a_days = _parse_days_list(a.get('days_of_week') or '')
                                b_days = _parse_days_list(b.get('days_of_week') or '')
                                if not a_days.intersection(b_days):
                                    continue
                            except Exception:
                                pass
                            if not _ranges_overlap(a.get('start_time') or '', a.get('end_time') or '', b.get('start_time') or '', b.get('end_time') or ''):
                                continue
                            a_general = int(a.get('is_general', 0) or 0) == 1
                            b_general = int(b.get('is_general', 0) or 0) == 1
                            if a_general or b_general:
                                raise ValueError("אין אפשרות לחפיפה בשעות כאשר קיימת שורה 'כללי' באותה קבוצה")
                            a_classes = set(_parse_classes_list(a.get('classes') or ''))
                            b_classes = set(_parse_classes_list(b.get('classes') or ''))
                            if a_classes.intersection(b_classes):
                                raise ValueError("אין אפשרות לחפיפה בשעות עבור אותן כיתות באותה קבוצה")
                except Exception as e:
                    messagebox.showerror("שגיאה", str(e))
                    return

                # persist: update existing ids where possible, add or delete extras
                try:
                    rows_existing = list(rows)
                    for idx, nr in enumerate(new_rows):
                        if idx < len(rows_existing):
                            r0 = rows_existing[idx]
                            self.db.update_time_bonus(
                                int(r0['id']),
                                new_group_name,
                                nr['start_time'],
                                nr['end_time'],
                                int(nr['bonus_points']),
                                int(r0.get('is_active', 1) or 0),
                                group_name=new_group_name,
                                is_general=int(nr.get('is_general', 1) or 0),
                                classes=(nr.get('classes') or None),
                                days_of_week=(nr.get('days_of_week') or None),
                                is_shown_public=int(nr.get('is_shown_public', 1) or 0),
                                sound_key=(nr.get('sound_key') or None),
                            )
                        else:
                            self.db.add_time_bonus(
                                new_group_name,
                                nr['start_time'],
                                nr['end_time'],
                                int(nr['bonus_points']),
                                group_name=new_group_name,
                                is_general=int(nr.get('is_general', 1) or 0),
                                classes=(nr.get('classes') or None),
                                days_of_week=(nr.get('days_of_week') or None),
                                is_shown_public=int(nr.get('is_shown_public', 1) or 0),
                                sound_key=(nr.get('sound_key') or None),
                            )
                    # delete extras
                    if len(rows_existing) > len(new_rows):
                        for r_ex in rows_existing[len(new_rows):]:
                            self.db.delete_time_bonus(int(r_ex['id']))

                    load_bonuses()
                    d.destroy()
                except Exception as e:
                    messagebox.showerror("שגיאה", f"שגיאה בעדכון קבוצה: {e}")

            tk.Button(bottom, text="ביטול", command=d.destroy, font=('Arial', 11), bg='#95a5a6', fg='white', padx=14, pady=8).pack(side=tk.RIGHT)
            tk.Button(bottom, text="שמור", command=on_save_group, font=('Arial', 11, 'bold'), bg='#3498db', fg='white', padx=18, pady=8).pack(side=tk.RIGHT, padx=8)
            tk.Button(bottom, text="הוסף שורה", command=on_add_row, font=('Arial', 11), bg='#2ecc71', fg='white', padx=14, pady=8).pack(side=tk.RIGHT)

            group_entry.focus_set()
            try:
                d.update_idletasks()
                w2 = d.winfo_reqwidth() + 40
                h2 = d.winfo_reqheight() + 40
                d.geometry(f"{w2}x{h2}")
                d.minsize(w2, h2)
            except Exception:
                pass

        tk.Button(
            btn_frame,
            text="🧩 עריכת קבוצה",
            command=edit_group,
            font=('Arial', 11),
            bg='#8e44ad',
            fg='white',
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)

        # עריכת שורה בודדת הוסרה – משתמשים בעריכת קבוצה בלבד.

        try:
            tree.bind('<Double-1>', lambda e: edit_group())
        except Exception:
            pass
        
        tk.Button(
            btn_frame,
            text="🔄 פעיל/כבוי",
            command=toggle_active,
            font=('Arial', 11),
            bg='#f39c12',
            fg='white',
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="🗑️ מחק",
            command=delete_bonus,
            font=('Arial', 11),
            bg='#e74c3c',
            fg='white',
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        tk.Button(
            btn_frame,
            text="✖ סגור",
            command=dialog.destroy,
            font=('Arial', 11),
            bg='#95a5a6',
            fg='white',
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)
        
        # טעינה ראשונית
        load_bonuses()

        try:
            dialog.update_idletasks()
            w = max(760, dialog.winfo_reqwidth() + 30)
            h = max(520, dialog.winfo_reqheight() + 30)
            dialog.geometry(f"{w}x{h}")
        except Exception:
            pass
    
    def show_login_screen(self):
        """מסך התחברות עם כרטיס"""
        try:
            self._login_active = True
        except Exception:
            pass
        try:
            # בזמן התחברות נצמצם את הצגת החלון הראשי כדי למנוע הבהוב קצר בפינה.
            # לא נשתמש ב-withdraw לפני יצירת ה-Toplevel, כי ב-Windows זה עלול למנוע ממנו להופיע.
            self.root.attributes('-alpha', 0.0)
        except Exception:
            pass
        # יצירת חלון התחברות
        login_window = tk.Toplevel(self.root)
        login_window.title("התחברות - עמדת ניהול")
        login_window.geometry("550x350")
        login_window.configure(bg='#ecf0f1')
        login_window.resizable(True, True)
        try:
            # כאשר root שקוף/מוסתר, transient לפעמים גורם לחלון ההתחברות לא להופיע.
            # נשאיר את החלון עצמאי אך עדיין נשתמש ב-grab_set.
            login_window.transient(None)
        except Exception:
            pass
        login_window.grab_set()
        
        # כעת, אחרי שחלון ההתחברות קיים, אפשר להסתיר לגמרי את ה-root
        try:
            self.root.withdraw()
        except Exception:
            pass
        
        # משתנה לשמירת הכרטיס
        card_buffer = {'text': ''}
        
        tk.Label(
            login_window,
            text="ברוכים הבאים למערכת הניהול",
            font=('Arial', 16, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50'
        ).pack(pady=20)
        
        tk.Label(
            login_window,
            text="🔐 העבר את כרטיס המורה שלך",
            font=('Arial', 14),
            bg='#ecf0f1',
            fg='#34495e'
        ).pack(pady=15)
        
        # שדה הצגת כרטיס
        card_display = tk.Label(
            login_window,
            text="ממתין לכרטיס...",
            font=('Arial', 12),
            bg='white',
            fg='#7f8c8d',
            width=30,
            height=2,
            relief=tk.SUNKEN
        )
        card_display.pack(pady=20)
        
        def on_key(event):
            """טיפול בלחיצת מקש"""
            if event.char and event.char not in ['\r', '\n']:
                card_buffer['text'] += event.char
                card_display.config(text=f"קורא כרטיס: {'*' * len(card_buffer['text'])}")
        
        def on_enter(event):
            """טיפול בלחיצת Enter - סיום קריאת כרטיס"""
            card_number = card_buffer['text'].strip()
            if not card_number:
                return
            
            if card_number == UNIVERSAL_MASTER_CODE:
                self.current_teacher = {
                    'name': 'מאסטר אוניברסלי',
                    'is_admin': 1,
                }
                login_window.destroy()
                self.continue_init()
                return
            
            # בדיקת כרטיס במסד הנתונים
            teacher = self.db.get_teacher_by_card(card_number)
            
            if teacher:
                # מורה נמצא - התחבר
                self.current_teacher = teacher
                login_window.destroy()
                try:
                    self._login_active = False
                except Exception:
                    pass
                self.continue_init()
            else:
                # כרטיס לא נמצא
                card_display.config(
                    text="❌ כרטיס לא מורשה! פנה למנהל",
                    fg='#e74c3c'
                )
                card_buffer['text'] = ''
                login_window.after(3000, lambda: card_display.config(
                    text="ממתין לכרטיס...",
                    fg='#7f8c8d'
                ))
        
        # טיפול בסגירת חלון ההתחברות
        def on_login_close():
            """כשסוגרים את חלון ההתחברות - סוגרים את כל התוכנה"""
            try:
                login_window.destroy()
            except Exception:
                pass
            try:
                self.root.quit()
            except Exception:
                pass
            try:
                self.root.destroy()
            except Exception:
                pass
        
        login_window.protocol("WM_DELETE_WINDOW", on_login_close)
        
        # קישור אירועי מקלדת
        login_window.bind('<Key>', on_key)
        login_window.bind('<Return>', on_enter)
        login_window.focus_force()

        # ב-Windows לפעמים קליק ראשון לא נותן פוקוס לחלון. נבצע "גניבת פוקוס" חד-פעמית.
        _one_click_focus_done = {'done': False}

        def _one_click_focus(_e=None):
            if _one_click_focus_done.get('done'):
                return
            _one_click_focus_done['done'] = True
            try:
                login_window.deiconify()
            except Exception:
                pass
            try:
                login_window.lift()
            except Exception:
                pass
            try:
                login_window.attributes('-topmost', True)
            except Exception:
                pass
            try:
                login_window.focus_force()
            except Exception:
                pass

            def _restore_topmost():
                try:
                    login_window.attributes('-topmost', False)
                except Exception:
                    pass

            try:
                login_window.after(250, _restore_topmost)
            except Exception:
                pass

        try:
            login_window.bind('<Button-1>', _one_click_focus)
        except Exception:
            pass
        try:
            def _on_login_map(_e=None):
                # בטעינת החלון (Map) נרצה רק להרים ולתת focus,
                # בלי "לשרוף" את הדגל שמאפשר לבצע קליק WinAPI מתוזמן.
                try:
                    login_window.deiconify()
                except Exception:
                    pass
                try:
                    login_window.lift()
                except Exception:
                    pass
                try:
                    login_window.focus_force()
                except Exception:
                    pass

            login_window.bind('<Map>', _on_login_map)
        except Exception:
            pass

        def _initial_focus_once():
            # העתקה של הגישה מעמדה ציבורית (ללא קליק אמיתי): topmost לזמן קצר + focus_force
            if _one_click_focus_done.get('done'):
                return
            _one_click_focus_done['done'] = True
            try:
                login_window.deiconify()
            except Exception:
                pass
            try:
                login_window.lift()
            except Exception:
                pass
            try:
                login_window.attributes('-topmost', True)
            except Exception:
                pass
            try:
                login_window.focus_force()
                login_window.update_idletasks()
            except Exception:
                pass

            # WinAPI fallback (Windows לעיתים חוסם focus_force): ננסה להביא ל-foreground בכוח
            try:
                import ctypes
                try:
                    hwnd = int(login_window.winfo_id())
                except Exception:
                    hwnd = 0
                if hwnd:
                    try:
                        user32 = ctypes.windll.user32
                        try:
                            user32.ShowWindow(hwnd, 5)  # SW_SHOW
                        except Exception:
                            pass
                        try:
                            user32.SetForegroundWindow(hwnd)
                        except Exception:
                            pass
                        try:
                            user32.SetFocus(hwnd)
                        except Exception:
                            pass
                    except Exception:
                        pass
            except Exception:
                pass

            # קליק עכבר אמיתי + הזזת הסמן (WinAPI) כדי ש-Windows יתייחס לחלון כחלון פעיל
            try:
                import ctypes
                try:
                    login_window.update_idletasks()
                    try:
                        x0 = int(login_window.winfo_rootx() or 0)
                        y0 = int(login_window.winfo_rooty() or 0)
                        ww = int(login_window.winfo_width() or login_window.winfo_reqwidth() or 500)
                        wh = int(login_window.winfo_height() or login_window.winfo_reqheight() or 320)
                    except Exception:
                        x0, y0, ww, wh = 0, 0, 500, 320
                    cx = int(x0) + (int(ww) // 2)
                    cy = int(y0) + (int(wh) // 2)

                    user32 = ctypes.windll.user32
                    user32.SetCursorPos(int(cx), int(cy))
                    MOUSEEVENTF_LEFTDOWN = 0x0002
                    MOUSEEVENTF_LEFTUP = 0x0004
                    try:
                        user32.mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                        user32.mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception:
                pass

            def _restore_topmost():
                try:
                    login_window.attributes('-topmost', False)
                except Exception:
                    pass

            try:
                login_window.after(1000, _restore_topmost)
            except Exception:
                pass

        # מרכז את חלון ההתחברות אחרי שה-UI נבנה (כדי שלא יזוז/ייצא מהמרכז)
        try:
            login_window.update_idletasks()
            w = max(450, int(login_window.winfo_reqwidth() or 450))
            h = max(300, int(login_window.winfo_reqheight() or 300))
            sw = int(login_window.winfo_screenwidth() or 1360)
            sh = int(login_window.winfo_screenheight() or 760)
            x = max(0, (sw // 2) - (w // 2))
            y = max(0, (sh // 2) - (h // 2))
            login_window.geometry(f'{w}x{h}+{x}+{y}')
            try:
                login_window.lift()
                login_window.focus_force()
            except Exception:
                pass
        except Exception:
            pass

        try:
            # פעם אחת בלבד אחרי שהכול כבר הוצג וממורכז
            login_window.after(200, _initial_focus_once)
        except Exception:
            pass
        
        # כפתור ניהול מורים (רק לראשונים - ייצור מנהל ראשוני)
        # המתן עד שהחלון ייסגר
        self.root.wait_window(login_window)

        # אם חלון ההתחברות נסגר בלי התחברות (למשל לחיצה על X) – נחזיר את החלון הראשי
        try:
            if not getattr(self, 'current_teacher', None):
                self.root.deiconify()
                try:
                    self.root.attributes('-alpha', 1.0)
                except Exception:
                    pass
                try:
                    self.root.state('zoomed')
                except Exception:
                    try:
                        sw = int(self.root.winfo_screenwidth() or 1360)
                        sh = int(self.root.winfo_screenheight() or 760)
                        self.root.geometry(f"{sw}x{sh}+0+0")
                    except Exception:
                        pass
        except Exception:
            pass
    
    def continue_init(self):
        """המשך אתחול אחרי התחברות מוצלחת"""
        try:
            self._login_active = False
        except Exception:
            pass
        try:
            self.root.deiconify()
            self.root.attributes('-alpha', 1.0)
            try:
                self.root.state('zoomed')
            except Exception:
                try:
                    sw = int(self.root.winfo_screenwidth() or 1360)
                    sh = int(self.root.winfo_screenheight() or 760)
                    self.root.geometry(f"{sw}x{sh}+0+0")
                except Exception:
                    pass
            self.root.lift()
            self.root.focus_force()
        except Exception:
            pass

        # איפוס caches שתלויים במשתמש, כדי למנוע מצב שבו מורה רואה רשימה ישנה עד רענון ידני
        try:
            self.teacher_classes_cache = []
        except Exception:
            pass
        try:
            self._swipe_stats_cache = None
        except Exception:
            pass
        try:
            # לאחר החלפת משתמש/התחברות נרצה טעינה מיידית של הטבלה
            self._suppress_auto_refresh_until = 0.0
        except Exception:
            pass
        # הצגת הודעת טעינה
        loading_label = tk.Label(
            self.root,
            text="⏳ טוען נתונים...",
            font=('Arial', 18, 'bold'),
            bg='#ecf0f1',
            fg='#3498db'
        )
        loading_label.place(relx=0.5, rely=0.5, anchor='center')
        self.root.update()
        
        # עדכון כותרת החלון
        teacher_name = self.current_teacher['name']
        is_admin = self.current_teacher['is_admin'] == 1
        role = "מנהל" if is_admin else "מורה"
        self.root.title(f"עמדת ניהול - {teacher_name} ({role})")
        
        # טעינת כיתות המורה ל-cache (פעם אחת בלבד!)
        if not is_admin:
            self.teacher_classes_cache = self.db.get_teacher_classes(self.current_teacher['id'])
            safe_print(f"📚 כיתות של {teacher_name}: {', '.join(self.teacher_classes_cache)}")
        
        # נתיב קובץ Excel - בדיקה במספר מיקומים
        parent_dir = os.path.dirname(self.base_dir)
        
        # 1. בדוק אם יש נתיב מוגדר בקובץ
        custom_path = None
        path_file = os.path.join(self.base_dir, 'excel_path.txt')
        if os.path.exists(path_file):
            try:
                with open(path_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            custom_path = line
                            break
            except:
                pass

        # קריאת הגדרות מקובץ ההגדרות החי (load_app_config יודע לעבוד עם מיקום כתיב)
        config_excel_path = None
        shared_folder = None
        try:
            cfg = self.load_app_config()
            if isinstance(cfg, dict):
                config_excel_path = cfg.get('excel_path')
                if not config_excel_path:
                    shared_folder = cfg.get('shared_folder') or cfg.get('network_root')
        except Exception as e:
            safe_print(f"שגיאה בקריאת config.json: {e}")
        
        # 2. רשימת מיקומים אפשריים
        possible_paths = []
        if config_excel_path:
            possible_paths.append(config_excel_path)
        if shared_folder:
            possible_paths.append(os.path.join(shared_folder, "טבלה למבצע אשראי.xlsx"))
        if custom_path:
            possible_paths.append(custom_path)
        possible_paths.append(os.path.join(parent_dir, "טבלה למבצע אשראי.xlsx"))
        possible_paths.append(os.path.join(self.base_dir, "טבלה למבצע אשראי.xlsx"))
        
        # 3. בחר את הראשון שקיים
        self.excel_path = None
        for path in possible_paths:
            if os.path.exists(path):
                self.excel_path = path
                try:
                    print(f"מיקום Excel: {path}")
                except:
                    pass  # אם יש בעיה בהדפסה, התעלם
                break
        
        # 4. אם לא נמצא, השתמש בברירת מחדל ויצור קובץ שבלונה
        if not self.excel_path:
            self.excel_path = possible_paths[1] if len(possible_paths) > 1 else possible_paths[0]
            safe_print(f"⚠️ קובץ Excel לא נמצא, ברירת מחדל: {self.excel_path}")
            # יצירת קובץ אקסל שבלונה עם כל העמודות הנכונות
            try:
                self.importer.export_to_excel(self.excel_path)
                safe_print("✅ נוצר קובץ אקסל שבלונה ראשוני לייבוא תלמידים")
            except Exception as e:
                safe_print(f"⚠️ לא ניתן היה ליצור קובץ אקסל שבלונה: {e}")
        
        # סינכרון אוטומטי כל 2 דקות
        self.sync_interval = 120000  # 2 דקות במילישניות
        self.last_excel_mod_time = None  # זמן שינוי אחרון של Excel
        
        # מצב עדכון מהיר
        self.quick_update_mode = False
        self.quick_update_points = 0
        
        # סימון אם היו שינויים
        self.has_changes = False

        # ייצוא אקסל אוטומטי (batch) כדי למנוע כתיבה על כל שינוי
        self._excel_export_job = None
        self._excel_export_pending = False
        self._excel_export_last_done_ts = 0.0
        self._excel_export_interval_sec = 300.0
        self._excel_export_min_gap_sec = 3.0
        self._excel_auto_export_enabled = True
        try:
            cfg_exp = self.load_app_config() or {}
            if isinstance(cfg_exp, dict):
                try:
                    self._excel_auto_export_enabled = bool(cfg_exp.get('excel_auto_export_enabled', True))
                except Exception:
                    self._excel_auto_export_enabled = True
                v = cfg_exp.get('excel_auto_export_interval_sec', None)
                if v is None:
                    v = cfg_exp.get('excel_auto_export_minutes', None)
                    if v is not None:
                        try:
                            v = float(v) * 60.0
                        except Exception:
                            v = None
                if v is not None:
                    try:
                        v = float(v)
                    except Exception:
                        v = None
                if v is not None and v >= 5.0:
                    self._excel_export_interval_sec = float(v)
        except Exception:
            pass
        
        # המשך אתחול רגיל...
        self.setup_ui()

        # טעינה מיידית של תלמידים (במיוחד אחרי החלפת משתמש)
        try:
            self.load_students(keep_selection=False)
        except Exception:
            try:
                self.root.after(50, lambda: self.load_students(keep_selection=False))
            except Exception:
                pass

        # בדיקת עדכון גרסה – רק מנהל ובעמדת ניהול
        try:
            if int(self.current_teacher.get('is_admin', 0) or 0) == 1:
                self._schedule_update_checks()
        except Exception:
            pass

        # הודעת מצב רישיון לאחר התחברות והטענת הממשק
        lm = getattr(self, "license_manager", None)
        if lm is not None:
            msg = lm.get_startup_message()
            if msg:
                try:
                    messagebox.showinfo("מצב רישיון", msg)
                except Exception:
                    pass

        # עדכון סטטוס כפתור בונוס
        self.update_bonus_button()

        # טעינה ראשונית - סינכרון מהיר + טעינת תלמידים
        self.initial_load()

        # ודא שהטבלה נטענת/מתרעננת מיידית לאחר התחברות (במיוחד למורים)
        try:
            self.load_students(keep_selection=False)
        except Exception:
            pass

        # הסרת הודעת טעינה
        loading_label.destroy()

        # רענון אוטומטי של הטבלה כל 10 שניות (לעדכונים מהעמדה הציבורית)
        self.auto_refresh_interval = 10000  # 10 שניות
        try:
            self.auto_refresh_job = self.root.after(self.auto_refresh_interval, self.auto_refresh_table)
        except Exception:
            self.auto_refresh_job = None

        # נעילה אוטומטית אחרי אי-פעולה
        try:
            self._start_idle_monitor()
        except Exception:
            pass

        # סינכרון ביציאה
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _get_downloads_dir(self) -> str:
        try:
            home = os.path.expanduser('~')
            p = os.path.join(home, 'Downloads')
            if os.path.isdir(p):
                return p
        except Exception:
            pass
        try:
            return os.path.expanduser('~')
        except Exception:
            return ''

    def _resolve_update_download_path(self, download_url: str, remote_version: str) -> str:
        try:
            url = str(download_url or '').strip()
        except Exception:
            url = ''
        if not url:
            return ''
        try:
            parsed = urllib.parse.urlparse(url)
            base = os.path.basename(parsed.path or '')
        except Exception:
            base = ''
        name, ext = os.path.splitext(base or '')
        if not name:
            name = 'SchoolPoints_Setup'
        if not ext:
            ext = '.exe'
        if remote_version:
            try:
                if str(remote_version) not in name:
                    name = f"{name}_v{remote_version}"
            except Exception:
                pass
        filename = f"{name}{ext}"
        downloads_dir = self._get_downloads_dir()
        if not downloads_dir:
            try:
                downloads_dir = os.path.dirname(os.path.abspath(__file__))
            except Exception:
                downloads_dir = ''
        if downloads_dir:
            return os.path.join(downloads_dir, filename)
        return filename

    def _open_update_download(self, download_path: str, download_url: str) -> None:
        try:
            path = str(download_path or '').strip()
        except Exception:
            path = ''
        if path and os.path.exists(path):
            try:
                os.startfile(path)
                return
            except Exception:
                try:
                    os.startfile(os.path.dirname(path))
                    return
                except Exception:
                    pass
        try:
            url = str(download_url or '').strip()
        except Exception:
            url = ''
        if url:
            try:
                webbrowser.open(url)
            except Exception:
                pass

    def _download_update_package_async(self, download_url: str, remote_version: str) -> str:
        try:
            url = str(download_url or '').strip()
        except Exception:
            url = ''
        if not url:
            return ''
        try:
            local_path = ''
            if url.lower().startswith('file://'):
                try:
                    local_path = urllib.request.url2pathname(urllib.parse.urlparse(url).path or '')
                except Exception:
                    local_path = ''
            elif os.path.exists(url):
                local_path = url
            if local_path and os.path.exists(local_path):
                return local_path
        except Exception:
            pass
        download_path = self._resolve_update_download_path(url, remote_version)
        if not download_path:
            return ''
        try:
            if os.path.exists(download_path) and os.path.getsize(download_path) > 0:
                return download_path
        except Exception:
            pass
        try:
            if getattr(self, '_update_download_version_in_progress', '') == remote_version:
                return download_path
        except Exception:
            pass
        self._update_download_version_in_progress = remote_version

        def _worker():
            ok = False
            tmp_path = download_path + '.part'
            try:
                try:
                    os.makedirs(os.path.dirname(download_path), exist_ok=True)
                except Exception:
                    pass
                with urllib.request.urlopen(url, timeout=20) as resp:
                    with open(tmp_path, 'wb') as f:
                        while True:
                            chunk = resp.read(65536)
                            if not chunk:
                                break
                            f.write(chunk)
                try:
                    os.replace(tmp_path, download_path)
                except Exception:
                    shutil.copy2(tmp_path, download_path)
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
                ok = True
            except Exception:
                try:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass

            def _on_ui():
                try:
                    self._update_download_version_in_progress = ''
                except Exception:
                    pass
                if not ok:
                    return
                try:
                    self._update_downloaded_version = remote_version
                except Exception:
                    pass
                try:
                    prompt = (
                        "העדכון ירד בהצלחה ונשמר ב:\n"
                        f"{download_path}\n\n"
                        "האם לפתוח את קובץ ההתקנה כעת?"
                    )
                    if messagebox.askyesno("עדכון הורד", prompt):
                        self._open_update_download(download_path, url)
                except Exception:
                    pass

            try:
                self.root.after(0, _on_ui)
            except Exception:
                pass

        try:
            threading.Thread(target=_worker, daemon=True).start()
        except Exception:
            try:
                self._update_download_version_in_progress = ''
            except Exception:
                pass
        return download_path

    def _schedule_update_checks(self):
        try:
            self._check_for_updates_async(show_no_update=False)
            try:
                self.update_check_job = self.root.after(60000, lambda: self._check_for_updates_async(show_no_update=False))
            except Exception:
                self.update_check_job = None
        except Exception:
            pass

    def _compare_versions(self, local_v: str, remote_v: str) -> int:
        """Return -1 if local<remote, 0 if equal, 1 if local>remote."""
        def _parts(v: str):
            try:
                return [int(p) for p in str(v).strip().split('.') if p.strip().isdigit()]
            except Exception:
                return []

        a = _parts(local_v)
        b = _parts(remote_v)
        n = max(len(a), len(b))
        a += [0] * (n - len(a))
        b += [0] * (n - len(b))
        for i in range(n):
            if a[i] < b[i]:
                return -1
            if a[i] > b[i]:
                return 1
        return 0

    def _check_for_updates_async(self, show_no_update: bool = False):
        try:
            cfg = self.load_app_config() or {}
            manifest_url = (cfg.get('update_manifest_url') or '').strip()
            if not manifest_url:
                return
        except Exception:
            return

        def _worker():
            payload = None
            err = None
            try:
                local_path = ''
                if manifest_url.lower().startswith('file://'):
                    try:
                        local_path = urllib.request.url2pathname(urllib.parse.urlparse(manifest_url).path or '')
                    except Exception:
                        local_path = ''
                elif os.path.exists(manifest_url):
                    local_path = manifest_url

                if local_path:
                    with open(local_path, 'r', encoding='utf-8') as f:
                        data = f.read()
                else:
                    req = urllib.request.Request(
                        manifest_url,
                        headers={'User-Agent': f'SchoolPoints/{APP_VERSION}'}
                    )
                    with urllib.request.urlopen(req, timeout=6) as resp:
                        data = resp.read().decode('utf-8', errors='replace')
                payload = json.loads(data)
            except Exception as e:
                err = str(e)

            def _on_ui():
                if err is not None:
                    return
                if not isinstance(payload, dict):
                    return

                remote_v = str(payload.get('version') or '').strip()
                if not remote_v:
                    return
                if self._compare_versions(APP_VERSION, remote_v) >= 0:
                    if show_no_update:
                        try:
                            messagebox.showinfo("עדכון", "אין עדכון חדש")
                        except Exception:
                            pass
                    return

                cfg2 = self.load_app_config() or {}
                download_url = (payload.get('download_url') or cfg2.get('update_download_url') or '').strip()
                msg = (payload.get('message') or cfg2.get('update_message') or '').strip()
                download_path = ''
                if download_url:
                    try:
                        download_path = self._download_update_package_async(download_url, remote_v)
                    except Exception:
                        download_path = ''
                self._show_update_dialog(
                    local_version=APP_VERSION,
                    remote_version=remote_v,
                    message=msg,
                    download_url=download_url,
                    download_path=download_path,
                )

            try:
                self.root.after(0, _on_ui)
            except Exception:
                pass

        try:
            threading.Thread(target=_worker, daemon=True).start()
        except Exception:
            pass

    def _show_update_dialog(self, local_version: str, remote_version: str, message: str, download_url: str, download_path: str = ''):
        try:
            if getattr(self, '_update_dialog_open', False):
                return
            self._update_dialog_open = True
        except Exception:
            pass

        d = tk.Toplevel(self.root)
        d.title("גרסה חדשה זמינה")
        d.geometry("520x260")
        d.configure(bg='#ecf0f1')
        d.transient(self.root)
        d.grab_set()

        def _close():
            try:
                self._update_dialog_open = False
            except Exception:
                pass
            try:
                d.destroy()
            except Exception:
                pass

        text = f"גרסה חדשה זמינה!\n\nגרסה חדשה: {remote_version}\nהגרסה שלך: {local_version}"
        if message:
            text += f"\n\n{message}"
        if download_path:
            try:
                if os.path.exists(download_path):
                    text += f"\n\nהקובץ כבר ירד: {download_path}"
                else:
                    text += "\n\nהעדכון יורד ברקע..."
            except Exception:
                text += "\n\nהעדכון יורד ברקע..."

        tk.Label(d, text=text, font=('Arial', 12, 'bold'), bg='#ecf0f1', justify='right').pack(pady=18, padx=18, anchor='e')

        btns = tk.Frame(d, bg='#ecf0f1')
        btns.pack(pady=10)

        def _open_download():
            if not download_url:
                try:
                    messagebox.showwarning("עדכון", "לא הוגדר קישור הורדה")
                except Exception:
                    pass
                return
            self._open_update_download(download_path, download_url)

        btn_label = "פתח עדכון" if (download_path and os.path.exists(download_path)) else "עדכון"
        tk.Button(btns, text=btn_label, width=12, command=_open_download, bg='#3498db', fg='white').pack(side=tk.LEFT, padx=6)
        tk.Button(btns, text="אחר כך", width=10, command=_close, bg='#95a5a6', fg='white').pack(side=tk.LEFT, padx=6)

        d.protocol("WM_DELETE_WINDOW", _close)

        try:
            d.update_idletasks()
            w = d.winfo_reqwidth() + 20
            h = d.winfo_reqheight() + 20
            d.geometry(f"{w}x{h}")
            d.minsize(w, h)
        except Exception:
            pass
    
    def on_closing(self):
        """טיפול בסגירת החלון"""
        try:
            if getattr(self, '_excel_export_job', None) is not None:
                try:
                    self.root.after_cancel(self._excel_export_job)
                except Exception:
                    pass
                self._excel_export_job = None
        except Exception:
            pass
        # ייצוא רק אם היו שינויים
        if self.has_changes:
            safe_print("📤 ייצוא שינויים לפני סגירה (רק G, H, I)...")
            try:
                self.importer.export_columns_only(self.excel_path)
                safe_print("✅ ייצוא הסתיים")
            except:
                pass
        else:
            safe_print("ℹ️ אין שינויים - דילוג על ייצוא")
        self.root.destroy()


def main():
    """הפעלה רגילה - ללא splash (גורם לבעיות)"""
    _set_windows_dpi_awareness()
    root = tk.Tk()
    try:
        import sys
        base_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        candidates = [
            os.path.join(base_dir, 'icons', 'admin.ico'),
            os.path.join(os.path.dirname(base_dir), 'icons', 'admin.ico'),
        ]
        for p in candidates:
            if p and os.path.exists(p):
                root.iconbitmap(p)
                break
    except Exception:
        pass
    _apply_tk_scaling(root)
    app = AdminStation(root)
    root.mainloop()


def main_no_splash():
    """הפעלה ללא splash screen (לבדיקות)"""
    _set_windows_dpi_awareness()
    root = tk.Tk()
    _apply_tk_scaling(root)
    app = AdminStation(root)
    root.mainloop()


if __name__ == "__main__":
    main()
