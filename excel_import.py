"""
מודול ייבוא נתונים מקובץ Excel
"""
try:
    import pandas as pd
    _PANDAS_AVAILABLE = True
except Exception:
    pd = None
    _PANDAS_AVAILABLE = False
_PANDAS_WARNED = False
_PANDAS_MISSING_MSG = "חסר רכיב pandas. פעולות ייבוא/ייצוא Excel לא זמינות עד להתקנה."
from database import Database
from typing import List, Dict
from collections import Counter
from datetime import date, timedelta, datetime, timezone


def _strip_asterisk_annotations(text: str) -> str:
    try:
        import re
        if not text:
            return text
        cleaned = re.sub(r'\*[^*]*\*', '', str(text))
        cleaned = re.sub(r'\s{2,}', ' ', cleaned).strip()
        return cleaned
    except Exception:
        return text


def _warn_missing_pandas() -> None:
    global _PANDAS_WARNED
    if _PANDAS_WARNED:
        return
    _PANDAS_WARNED = True
    msg = _PANDAS_MISSING_MSG
    try:
        from tkinter import messagebox
        messagebox.showerror("שגיאה", msg)
    except Exception:
        try:
            print(msg)
        except Exception:
            pass


def _require_pandas() -> bool:
    if _PANDAS_AVAILABLE:
        return True
    _warn_missing_pandas()
    return False


class ExcelImporter:
    def __init__(self, db: Database):
        self.db = db

    def export_points_log_excel(self, excel_path: str, student: dict, logs: list) -> bool:
        """ייצוא היסטוריית נקודות מפורטת לתלמיד לקובץ Excel.
        logs: רשימת רשומות מ-DB.points_log (dict).
        """
        if not _require_pandas():
            return False
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            student_name = f"{student.get('first_name', '')} {student.get('last_name', '')}".strip()
            class_name = str(student.get('class_name') or '').strip()

            def _split_dt(dt_val: str) -> tuple[str, str]:
                s = str(dt_val or '').strip()
                if not s:
                    return '', ''
                # SQLite default: YYYY-MM-DD HH:MM:SS
                if ' ' in s:
                    d, t = s.split(' ', 1)
                    return d.strip(), t.strip()
                return s, ''

            def _actor_kind(action_type: str, actor_name: str, reason: str) -> str:
                at = str(action_type or '').strip().lower()
                an = str(actor_name or '').strip()
                rs = str(reason or '').strip()
                if at in ('purchase', 'purchase_with_scheduled'):
                    return 'קניות'
                if at in ('refund',):
                    return 'החזר'
                if 'בונוס' in rs or at.startswith('בונוס'):
                    return 'בונוס'
                if an == 'מנהל':
                    return 'מנהל'
                if an == 'מערכת':
                    return 'מערכת'
                if an:
                    return 'מורה'
                return ''

            def _fmt_delta(n: int) -> str:
                try:
                    n = int(n)
                except Exception:
                    return ''
                return f"+{n}" if n > 0 else str(n)

            columns = [
                'שם תלמיד',
                'כיתה',
                'תאריך',
                'שעה',
                'שינוי',
                'לפני',
                'אחרי',
                'סוג פעולה',
                'מבצע פעולה',
                'בוצע ע"י',
                'סיבה',
            ]

            data = []
            for row in (logs or []):
                d, t = _split_dt(row.get('created_at'))
                action_type = row.get('action_type') or ''
                actor_name = row.get('actor_name') or ''
                reason = row.get('reason') or ''
                data.append({
                    'שם תלמיד': student_name,
                    'כיתה': class_name,
                    'תאריך': d,
                    'שעה': t,
                    'שינוי': _fmt_delta(row.get('delta')),
                    'לפני': row.get('old_points'),
                    'אחרי': row.get('new_points'),
                    'סוג פעולה': action_type,
                    'מבצע פעולה': _actor_kind(action_type, actor_name, reason),
                    'בוצע ע"י': actor_name,
                    'סיבה': reason,
                })

            df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)
            df.to_excel(excel_path, index=False, engine='openpyxl')

            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Apply RTL and alternating colors styling
            try:
                from excel_styling import apply_rtl_and_alternating_colors
                apply_rtl_and_alternating_colors(ws, has_header=True)
            except Exception:
                # Fallback to basic RTL if styling module not available
                ws.sheet_view.rightToLeft = True
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

            # הקפאת שורת כותרת + כותרת חוזרת בהדפסה
            try:
                ws.freeze_panes = 'A2'
                ws.print_title_rows = '1:1'
            except Exception:
                pass

            # שם גליון
            try:
                base = f"היסטוריה - {student.get('first_name','')}{student.get('last_name','')}".strip() or 'היסטוריה'
                ws.title = str(base)[:31]
            except Exception:
                pass

            # הוספת מעברי עמוד לכל הגליונות
            for ws in wb.worksheets:
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToHeight = 0
                ws.page_setup.fitToWidth = 1

            wb.save(excel_path)
            return True
        except Exception as e:
            print(f"שגיאה בייצוא היסטוריית נקודות: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def quick_update_from_excel(self, excel_path: str) -> tuple[int, List[str]]:
        """
        עדכון מהיר - רק עמודות G, H, I (כרטיס, נקודות, הודעה)
        משמש לפתיחת תוכנה
        
        Returns:
            tuple של (מספר תלמידים שעודכנו, רשימת שגיאות)
        """
        if not _require_pandas():
            return 0, [_PANDAS_MISSING_MSG]
        errors = []
        updated_count = 0
        
        try:
            # קריאת קובץ Excel
            df = pd.read_excel(excel_path, dtype={'מס\' כרטיס': str})
            
            # קבלת כל התלמידים פעם אחת
            all_students = self.db.get_all_students()
            
            # רק תלמידים עם שם
            for index, row in df.iterrows():
                try:
                    last_name = str(row.get('שם משפחה', '')).strip() if pd.notna(row.get('שם משפחה')) else ''
                    first_name = str(row.get('שם פרטי', '')).strip() if pd.notna(row.get('שם פרטי')) else ''
                    
                    if not last_name or not first_name:
                        continue
                    
                    # חיפוש תלמיד קיים - לפי שם מדויק
                    matching = [s for s in all_students if s['first_name'] == first_name and s['last_name'] == last_name]
                    
                    if not matching:
                        continue  # לא מעדכנים תלמידים שלא קיימים
                    
                    student = matching[0]
                    student_id = student['id']
                    
                    # עדכון כרטיס (עמודה G)
                    card_number = str(row.get('מס\' כרטיס', '')).strip() if pd.notna(row.get('מס\' כרטיס')) else ''
                    card_number = card_number.lstrip("'")
                    if card_number == 'nan' or card_number == '' or card_number == '0':
                        card_number = None
                    
                    if card_number != student.get('card_number'):
                        self.db.update_card_number(student_id, card_number)
                    
                    # עדכון נקודות (עמודה H)
                    points = 0
                    if 'מס\' נקודות' in df.columns and pd.notna(row.get('מס\' נקודות')):
                        try:
                            points = int(float(row.get('מס\' נקודות')))
                        except:
                            points = 0
                    
                    if points != student['points']:
                        self.db.update_student_points(student_id, points, "סנכרון מ-Excel", "מערכת")
                    
                    # עדכון הודעה פרטית (עמודה I)
                    if 'הודעה פרטית' in df.columns:
                        try:
                            msg_val = row['הודעה פרטית']
                            if pd.notna(msg_val) and str(msg_val).strip() and str(msg_val).strip().lower() != 'nan':
                                self.db.update_private_message(student_id, str(msg_val).strip())
                            else:
                                self.db.update_private_message(student_id, None)
                        except:
                            pass
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"שגיאה בשורה {index + 2}: {str(e)}")
            
            return updated_count, errors
            
        except Exception as e:
            errors.append(f"שגיאה כללית בקריאת הקובץ: {str(e)}")
            return 0, errors
    
    def import_from_excel(self, excel_path: str, clear_existing: bool = False, dtype_dict: dict = None) -> tuple[int, List[str]]:
        """
        ייבוא תלמידים מקובץ Excel
        
        Args:
            excel_path: נתיב לקובץ Excel
            clear_existing: האם למחוק תלמידים קיימים לפני הייבוא
            
        Returns:
            tuple של (מספר תלמידים שיובאו, רשימת שגיאות)
        """
        if not _require_pandas():
            return 0, [_PANDAS_MISSING_MSG]
        errors = []
        imported_count = 0
        # נתוני תיקופים לייבוא (רק כאשר clear_existing=True)
        swipe_totals_by_student: Dict[int, int] = {}
        approx_days_candidates: List[int] = []
        
        try:
            # קריאת קובץ Excel - עמודת כרטיס כטקסט!
            df = pd.read_excel(excel_path, dtype={'מס\' כרטיס': str})
            
            # וידוא שהעמודות הנדרשות קיימות
            required_columns = ['שם משפחה', 'שם פרטי']
            for col in required_columns:
                if col not in df.columns:
                    errors.append(f"חסרה עמודה נדרשת: {col}")
                    return 0, errors
            
            # מחיקת נתונים קיימים אם נדרש
            if clear_existing:
                self.db.clear_all_students()
            
            # ייבוא כל תלמיד
            # תמונת תלמיד: תומך גם בעמודה ישנה "מס' תמונה" וגם חדשה "נתיב תמונה"
            photo_col_name = "נתיב תמונה" if "נתיב תמונה" in df.columns else "מס' תמונה"

            for index, row in df.iterrows():
                try:
                    # המרת ערכים ריקים ל-string ריק
                    last_name = str(row.get('שם משפחה', '')).strip()
                    first_name = str(row.get('שם פרטי', '')).strip()
                    
                    # דילוג על שורות ריקות
                    if not last_name or not first_name or last_name == 'nan' or first_name == 'nan':
                        continue
                    
                    id_number = str(row.get('ת"ז', '')).strip() if pd.notna(row.get('ת"ז')) else ''
                    if id_number == 'nan':
                        id_number = ''
                    
                    class_name = str(row.get('כיתה', '')).strip() if pd.notna(row.get('כיתה')) else ''
                    if class_name == 'nan':
                        class_name = ''
                    
                    photo_number = str(row.get(photo_col_name, '')).strip() if pd.notna(row.get(photo_col_name)) else ''
                    if photo_number == 'nan':
                        photo_number = ''
                    
                    # כרטיס - אם ריק, השאר NULL במקום string ריק כדי להימנע מ-UNIQUE constraint
                    card_number = str(row.get('מס\' כרטיס', '')).strip() if pd.notna(row.get('מס\' כרטיס')) else ''
                    # הסר apostroph אם יש (מטקסט באקסל)
                    card_number = card_number.lstrip("'")
                    if card_number == 'nan' or card_number == '' or card_number == '0':
                        card_number = None  # NULL במסד הנתונים
                    
                    # נקודות - ברירת מחדל 0
                    points = 0
                    if 'מס\' נקודות' in df.columns and pd.notna(row.get('מס\' נקודות')):
                        try:
                            points = int(float(row.get('מס\' נקודות')))
                        except:
                            points = 0

                    # תיקופים (סיכום) - מס' תיקופים וממוצע תיקופים (אופציונלי)
                    total_swipes_val = None
                    avg_swipes_val = None
                    if "מס' תיקופים" in df.columns and pd.notna(row.get("מס' תיקופים")):
                        try:
                            total_swipes_val = int(float(row.get("מס' תיקופים")))
                            if total_swipes_val <= 0:
                                total_swipes_val = None
                        except Exception:
                            total_swipes_val = None

                    if "ממוצע תיקופים" in df.columns and pd.notna(row.get("ממוצע תיקופים")):
                        try:
                            # יכול להיות float עם עיגול
                            avg_swipes_val = float(row.get("ממוצע תיקופים"))
                            if avg_swipes_val <= 0:
                                avg_swipes_val = None
                        except Exception:
                            avg_swipes_val = None

                    # אם יש גם מס' תיקופים וגם ממוצע – ניתן להעריך מספר ימי לימודים גלובלי
                    if total_swipes_val is not None and avg_swipes_val is not None:
                        try:
                            approx_days = int(round(total_swipes_val / avg_swipes_val))
                            if approx_days >= 1:
                                approx_days_candidates.append(approx_days)
                        except Exception:
                            pass
                    
                    # הודעה פרטית
                    private_message = None
                    if 'הודעה פרטית' in df.columns:
                        try:
                            msg_val = row.get('הודעה פרטית')
                            if pd.notna(msg_val) and str(msg_val).strip() and str(msg_val).strip().lower() != 'nan':
                                private_message = str(msg_val).strip()
                        except:
                            pass

                    # מס' סידורי (אם קיים בעמודה באקסל - זיהוי גמיש לפי שם העמודה)
                    # אם אין עמודה כזו – ברירת מחדל: המספר לפי סדר השורה באקסל (1,2,3...)
                    serial_number = None
                    serial_col_name = None
                    for col in df.columns:
                        try:
                            if 'סידורי' in str(col):
                                serial_col_name = col
                                break
                        except Exception:
                            continue

                    if serial_col_name is not None:
                        try:
                            serial_val = row.get(serial_col_name)
                            if pd.notna(serial_val) and str(serial_val).strip() and str(serial_val).strip().lower() != 'nan':
                                serial_number = int(float(serial_val))
                        except Exception:
                            serial_number = None
                    else:
                        # אין עמודת "מס' סידורי" – השתמש בסדר השורות כמות שהוא
                        serial_number = int(index) + 1

                    # מגדר
                    import_gender = None
                    if 'מגדר' in df.columns:
                        try:
                            g = str(row.get('מגדר', '') or '').strip()
                            if g in ('בן', 'M', 'זכר'):
                                import_gender = 'M'
                            elif g in ('בת', 'F', 'נקבה'):
                                import_gender = 'F'
                        except Exception:
                            pass

                    # יום הולדת עברי
                    import_hb_day = None
                    import_hb_month = None
                    import_hb_year = None
                    _heb_month_map = {
                        'תשרי': 7, 'חשון': 8, 'כסלו': 9, 'טבת': 10, 'שבט': 11,
                        'אדר': 12, "אדר ב'": 13, "אדר ב": 13,
                        'ניסן': 1, 'אייר': 2, 'סיון': 3, 'תמוז': 4, 'אב': 5, 'אלול': 6,
                    }
                    if 'יום הולדת עברי - יום' in df.columns:
                        try:
                            dv = row.get('יום הולדת עברי - יום')
                            if pd.notna(dv) and str(dv).strip() and str(dv).strip().lower() != 'nan':
                                import_hb_day = int(float(dv))
                        except Exception:
                            pass
                    if 'יום הולדת עברי - חודש' in df.columns:
                        try:
                            mv = str(row.get('יום הולדת עברי - חודש') or '').strip()
                            if mv and mv.lower() != 'nan':
                                import_hb_month = _heb_month_map.get(mv)
                                if import_hb_month is None:
                                    try:
                                        import_hb_month = int(float(mv))
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                    if 'יום הולדת עברי - שנה' in df.columns:
                        try:
                            yv = row.get('יום הולדת עברי - שנה')
                            if pd.notna(yv) and str(yv).strip() and str(yv).strip().lower() != 'nan':
                                import_hb_year = int(float(yv))
                        except Exception:
                            pass

                    # חיפוש תלמיד קיים (לפי שם)
                    existing = self.db.search_students(f"{first_name} {last_name}")
                    
                    if existing:
                        # עדכון תלמיד קיים
                        student = existing[0]
                        student_id = student['id']

                        # עדכון מס' סידורי אם השתנה
                        if serial_number is not None and serial_number != student.get('serial_number'):
                            self.db.update_serial_number(student_id, serial_number)
                        
                        # עדכון כרטיס אם השתנה
                        if card_number and card_number != student['card_number']:
                            self.db.update_card_number(student_id, card_number)
                        
                        # עדכון נקודות אם השתנו
                        if points != student['points']:
                            self.db.update_student_points(student_id, points, "ייבוא מ-Excel", "מערכת")

                        # עדכון נתיב תמונה אם השתנה
                        if photo_number != (student.get('photo_number') or ''):
                            self.db.update_photo_number(student_id, photo_number)
                        
                        # עדכון הודעה פרטית - תמיד מעדכן (גם אם ריק)!
                        self.db.update_private_message(student_id, private_message)

                        # עדכון מגדר ויום הולדת עברי (אם הגיעו מהאקסל)
                        if import_gender is not None or import_hb_day is not None or import_hb_month is not None or import_hb_year is not None:
                            try:
                                self.db.update_student_basic(
                                    student_id=student_id,
                                    last_name=student.get('last_name', last_name),
                                    first_name=student.get('first_name', first_name),
                                    id_number=student.get('id_number') or id_number,
                                    class_name=student.get('class_name') or class_name,
                                    card_number=student.get('card_number') or (card_number or ''),
                                    photo_number=student.get('photo_number') or photo_number,
                                    serial_number=serial_number,
                                    hebrew_birth_day=import_hb_day if import_hb_day is not None else student.get('hebrew_birth_day'),
                                    hebrew_birth_month=import_hb_month if import_hb_month is not None else student.get('hebrew_birth_month'),
                                    hebrew_birth_year=import_hb_year if import_hb_year is not None else student.get('hebrew_birth_year'),
                                    gender=import_gender if import_gender is not None else (student.get('gender') or None),
                                )
                            except Exception:
                                pass

                        # שמירת סיכום תיקופים ליצירת swipe_log (רק אם clear_existing)
                        if clear_existing and total_swipes_val is not None:
                            swipe_totals_by_student[student_id] = total_swipes_val

                        imported_count += 1
                    elif last_name and first_name:
                        # הוספת תלמיד חדש
                        student_id = self.db.add_student(
                            last_name=last_name,
                            first_name=first_name,
                            id_number=id_number,
                            class_name=class_name,
                            photo_number=photo_number,
                            card_number=card_number,
                            points=points,
                            serial_number=serial_number,
                            hebrew_birth_day=import_hb_day,
                            hebrew_birth_month=import_hb_month,
                            hebrew_birth_year=import_hb_year,
                            gender=import_gender,
                        )
                        # אם קיימת הודעה פרטית בשורה – עדכן גם אותה לתלמיד החדש
                        if private_message is not None:
                            self.db.update_private_message(student_id, private_message)

                        # שמירת סיכום תיקופים ליצירת swipe_log (רק אם clear_existing)
                        if clear_existing and total_swipes_val is not None:
                            swipe_totals_by_student[student_id] = total_swipes_val

                        imported_count += 1
                        
                except Exception as e:
                    errors.append(f"שגיאה בשורה {index + 2}: {str(e)}")
            # לאחר ייבוא תלמידים – אם ביקשו איפוס נתונים קיימים ויש סיכומי תיקופים,
            # נשחזר היסטוריית swipe_log סינתטית כך שהסטטיסטיקות יתאימו ככל האפשר.
            if clear_existing and swipe_totals_by_student:
                try:
                    # קביעת מספר ימי הלימודים הכולל (גלובלי לכל התלמידים)
                    total_days = None
                    if approx_days_candidates:
                        counter = Counter(approx_days_candidates)
                        total_days = max(counter.items(), key=lambda kv: kv[1])[0]

                    if not total_days or total_days < 1:
                        total_days = 1

                    # בניית רשימת תאריכים (לא כולל שבת) – נסוגים מ-היום אחורה
                    days_list = []
                    current = date.today()
                    while len(days_list) < total_days:
                        # ב-SQLite שבת היא '%w' == '6'. ב-Python weekday(): שבת היא 5.
                        if current.weekday() != 5:  # הימנעות משבת
                            days_list.append(current)
                        current = current - timedelta(days=1)

                    # יצירת רשומות swipe_log
                    for student_id, total_swipes in swipe_totals_by_student.items():
                        try:
                            if total_swipes <= 0:
                                continue
                            base_per_day = total_swipes // total_days
                            remainder = total_swipes % total_days

                            # נפזר את השארית על הימים הראשונים כדי שהסכום הכולל יתאים
                            swipe_index = 0
                            for day_idx, day in enumerate(days_list):
                                count_for_day = base_per_day + (1 if day_idx < remainder else 0)
                                if count_for_day <= 0:
                                    continue
                                for i in range(count_for_day):
                                    # שעה קבועה + דקות שונות רק לצורך ייחוד חלקי
                                    minute = (swipe_index + i) % 60
                                    swiped_at = f"{day.isoformat()} 08:{minute:02d}:00"
                                    self.db.insert_swipe(student_id, swiped_at=swiped_at, card_number="", station_type="public")
                                swipe_index += count_for_day
                        except Exception as ex:
                            errors.append(f"שגיאה ביצירת תיקופים לתלמיד {student_id}: {str(ex)}")

                except Exception as e:
                    errors.append(f"שגיאה בשחזור היסטוריית תיקופים: {str(e)}")

            return imported_count, errors
            
        except Exception as e:
            errors.append(f"שגיאה כללית בקריאת הקובץ: {str(e)}")
            return 0, errors
    
    # ---------- ייבוא סלקטיבי ----------
    SELECTABLE_FIELDS = [
        ('כיתה', 'class'), ('ת"ז', 'id_number'), ('נתיב תמונה', 'photo'),
        ("מס' כרטיס", 'card'), ("מס' נקודות", 'points'), ('הודעה פרטית', 'private_message'),
        ('מגדר', 'gender'), ('יום הולדת עברי', 'hebrew_birthday'), ("מס' סידורי", 'serial'),
    ]

    def read_excel_info(self, excel_path: str) -> dict:
        """קריאת מידע על קובץ Excel: עמודות זמינות, מספר שורות, שדות לייבוא."""
        if not _require_pandas():
            return {'error': _PANDAS_MISSING_MSG}
        try:
            df = pd.read_excel(excel_path, dtype={"מס' כרטיס": str})
            cols = list(df.columns)
            if 'שם משפחה' not in cols or 'שם פרטי' not in cols:
                return {'error': 'חסרות עמודות חובה: שם משפחה, שם פרטי'}
            valid = df[df['שם משפחה'].notna() & df['שם פרטי'].notna()]
            valid = valid[valid['שם משפחה'].astype(str).str.strip().ne('')]
            valid = valid[valid['שם פרטי'].astype(str).str.strip().ne('')]
            valid = valid[~valid['שם משפחה'].astype(str).str.strip().str.lower().eq('nan')]
            valid = valid[~valid['שם פרטי'].astype(str).str.strip().str.lower().eq('nan')]
            available = []
            _col_detect = {
                'class': 'כיתה' in cols, 'id_number': 'ת"ז' in cols,
                'photo': ('נתיב תמונה' in cols or "מס' תמונה" in cols),
                'card': "מס' כרטיס" in cols, 'points': "מס' נקודות" in cols,
                'private_message': 'הודעה פרטית' in cols, 'gender': 'מגדר' in cols,
                'hebrew_birthday': any(c in cols for c in [
                    'יום הולדת עברי - יום', 'יום הולדת עברי - חודש', 'יום הולדת עברי - שנה']),
                'serial': any('סידורי' in str(c) for c in cols),
            }
            for display_name, field_key in self.SELECTABLE_FIELDS:
                if _col_detect.get(field_key):
                    available.append((display_name, field_key))
            return {'columns': cols, 'row_count': len(valid), 'available_fields': available}
        except Exception as e:
            return {'error': str(e)}

    def selective_import(self, excel_path: str, selected_keys: set,
                         add_new: bool = False) -> tuple:
        """ייבוא סלקטיבי – עדכון רק שדות נבחרים לתלמידים שמזוהים לפי שם.

        Returns: (updated_count, added_count, skipped_count, errors)
        """
        if not _require_pandas():
            return 0, 0, 0, [_PANDAS_MISSING_MSG]
        errors: List[str] = []
        updated = 0
        added = 0
        skipped = 0
        try:
            df = pd.read_excel(excel_path, dtype={"מס' כרטיס": str})
            cols = list(df.columns)
            if 'שם משפחה' not in cols or 'שם פרטי' not in cols:
                return 0, 0, 0, ['חסרות עמודות חובה: שם משפחה, שם פרטי']

            all_students = self.db.get_all_students()
            # מילון שם -> רשימת תלמידים (לטיפול בכפילויות שם)
            lookup: Dict[tuple, list] = {}
            serial_lookup: Dict[str, dict] = {}
            for s in all_students:
                k = (str(s.get('first_name') or '').strip(),
                     str(s.get('last_name') or '').strip())
                if k[0] and k[1]:
                    lookup.setdefault(k, []).append(s)
                sn = str(s.get('serial_number') or '').strip()
                if sn and sn != 'None' and sn != 'nan':
                    serial_lookup[sn] = s

            photo_col = "נתיב תמונה" if "נתיב תמונה" in cols else "מס' תמונה"
            serial_col = None
            for c in cols:
                if 'סידורי' in str(c):
                    serial_col = c
                    break
            _heb_month_map = {
                'תשרי': 7, 'חשון': 8, 'כסלו': 9, 'טבת': 10, 'שבט': 11,
                'אדר': 12, "אדר ב'": 13, "אדר ב": 13,
                'ניסן': 1, 'אייר': 2, 'סיון': 3, 'תמוז': 4, 'אב': 5, 'אלול': 6,
            }

            for index, row in df.iterrows():
                try:
                    last_name = str(row.get('שם משפחה', '')).strip()
                    first_name = str(row.get('שם פרטי', '')).strip()
                    if not last_name or not first_name or last_name == 'nan' or first_name == 'nan':
                        continue
                    candidates = lookup.get((first_name, last_name)) or []
                    existing = None
                    if len(candidates) == 1:
                        existing = candidates[0]
                    elif len(candidates) > 1:
                        # כפילות שם — ננסה לזהות לפי מס' סידורי
                        row_serial = ''
                        if serial_col:
                            try:
                                rv = row.get(serial_col)
                                if pd.notna(rv):
                                    row_serial = str(rv).strip().replace('.0', '')
                            except Exception:
                                pass
                        if row_serial and row_serial in serial_lookup:
                            existing = serial_lookup[row_serial]
                        else:
                            skipped += 1
                            errors.append(
                                f"שורה {index + 2}: {first_name} {last_name} — "
                                f"נמצאו {len(candidates)} תלמידים עם שם זהה, לא ניתן לזהות")
                            continue
                    if existing:
                        sid = existing['id']
                        self._selective_update_student(
                            sid, existing, row, cols, selected_keys,
                            photo_col, serial_col, _heb_month_map)
                        updated += 1
                    elif add_new:
                        self._selective_add_student(
                            row, cols, selected_keys, first_name, last_name,
                            photo_col, serial_col, _heb_month_map, index)
                        added += 1
                    else:
                        skipped += 1
                except Exception as e:
                    errors.append(f"שגיאה בשורה {index + 2}: {str(e)}")
            return updated, added, skipped, errors
        except Exception as e:
            errors.append(f"שגיאה כללית: {str(e)}")
            return updated, added, skipped, errors

    def _selective_update_student(self, sid, existing, row, cols, selected_keys,
                                  photo_col, serial_col, _heb_month_map):
        """עדכון שדות נבחרים לתלמיד קיים."""
        # שדות שמעדכנים דרך update_student_basic
        needs_basic = selected_keys & {'class', 'id_number', 'gender', 'hebrew_birthday'}
        basic_kwargs = dict(
            student_id=sid,
            last_name=existing.get('last_name', ''),
            first_name=existing.get('first_name', ''),
            id_number=existing.get('id_number') or '',
            class_name=existing.get('class_name') or '',
            card_number=existing.get('card_number') or '',
            photo_number=existing.get('photo_number') or '',
            serial_number=existing.get('serial_number'),
            hebrew_birth_day=existing.get('hebrew_birth_day'),
            hebrew_birth_month=existing.get('hebrew_birth_month'),
            hebrew_birth_year=existing.get('hebrew_birth_year'),
            gender=existing.get('gender'),
        )
        if 'class' in selected_keys and 'כיתה' in cols:
            val = str(row.get('כיתה', '')).strip() if pd.notna(row.get('כיתה')) else ''
            if val == 'nan': val = ''
            basic_kwargs['class_name'] = val
        if 'id_number' in selected_keys and 'ת"ז' in cols:
            val = str(row.get('ת"ז', '')).strip() if pd.notna(row.get('ת"ז')) else ''
            if val == 'nan': val = ''
            basic_kwargs['id_number'] = val
        if 'gender' in selected_keys and 'מגדר' in cols:
            g = str(row.get('מגדר', '') or '').strip()
            if g in ('בן', 'M', 'זכר'): basic_kwargs['gender'] = 'M'
            elif g in ('בת', 'F', 'נקבה'): basic_kwargs['gender'] = 'F'
        if 'hebrew_birthday' in selected_keys:
            if 'יום הולדת עברי - יום' in cols:
                try:
                    dv = row.get('יום הולדת עברי - יום')
                    if pd.notna(dv) and str(dv).strip() and str(dv).strip().lower() != 'nan':
                        basic_kwargs['hebrew_birth_day'] = int(float(dv))
                except Exception: pass
            if 'יום הולדת עברי - חודש' in cols:
                try:
                    mv = str(row.get('יום הולדת עברי - חודש') or '').strip()
                    if mv and mv.lower() != 'nan':
                        basic_kwargs['hebrew_birth_month'] = _heb_month_map.get(mv)
                        if basic_kwargs['hebrew_birth_month'] is None:
                            basic_kwargs['hebrew_birth_month'] = int(float(mv))
                except Exception: pass
            if 'יום הולדת עברי - שנה' in cols:
                try:
                    yv = row.get('יום הולדת עברי - שנה')
                    if pd.notna(yv) and str(yv).strip() and str(yv).strip().lower() != 'nan':
                        basic_kwargs['hebrew_birth_year'] = int(float(yv))
                except Exception: pass
        if needs_basic:
            self.db.update_student_basic(**basic_kwargs)

        if 'card' in selected_keys and "מס' כרטיס" in cols:
            val = str(row.get("מס' כרטיס", '')).strip() if pd.notna(row.get("מס' כרטיס")) else ''
            val = val.lstrip("'")
            if val in ('nan', '', '0'): val = None
            self.db.update_card_number(sid, val or '')
        if 'photo' in selected_keys and photo_col in cols:
            val = str(row.get(photo_col, '')).strip() if pd.notna(row.get(photo_col)) else ''
            if val == 'nan': val = ''
            self.db.update_photo_number(sid, val)
        if 'points' in selected_keys and "מס' נקודות" in cols:
            try:
                pts = int(float(row.get("מס' נקודות"))) if pd.notna(row.get("מס' נקודות")) else 0
            except Exception: pts = 0
            if pts != int(existing.get('points') or 0):
                self.db.update_student_points(sid, pts, "ייבוא סלקטיבי מ-Excel", "מערכת")
        if 'private_message' in selected_keys and 'הודעה פרטית' in cols:
            val = None
            try:
                msg_val = row.get('הודעה פרטית')
                if pd.notna(msg_val) and str(msg_val).strip() and str(msg_val).strip().lower() != 'nan':
                    val = str(msg_val).strip()
            except Exception: pass
            self.db.update_private_message(sid, val)
        if 'serial' in selected_keys and serial_col and serial_col in cols:
            try:
                sv = row.get(serial_col)
                if pd.notna(sv) and str(sv).strip() and str(sv).strip().lower() != 'nan':
                    self.db.update_serial_number(sid, int(float(sv)))
            except Exception: pass

    def _selective_add_student(self, row, cols, selected_keys, first_name, last_name,
                               photo_col, serial_col, _heb_month_map, index):
        """הוספת תלמיד חדש בייבוא סלקטיבי."""
        kwargs = dict(last_name=last_name, first_name=first_name)
        if 'class' in selected_keys and 'כיתה' in cols:
            val = str(row.get('כיתה', '')).strip() if pd.notna(row.get('כיתה')) else ''
            if val != 'nan': kwargs['class_name'] = val
        if 'id_number' in selected_keys and 'ת"ז' in cols:
            val = str(row.get('ת"ז', '')).strip() if pd.notna(row.get('ת"ז')) else ''
            if val != 'nan': kwargs['id_number'] = val
        if 'photo' in selected_keys and photo_col in cols:
            val = str(row.get(photo_col, '')).strip() if pd.notna(row.get(photo_col)) else ''
            if val != 'nan': kwargs['photo_number'] = val
        if 'card' in selected_keys and "מס' כרטיס" in cols:
            val = str(row.get("מס' כרטיס", '')).strip() if pd.notna(row.get("מס' כרטיס")) else ''
            val = val.lstrip("'")
            if val not in ('nan', '', '0'): kwargs['card_number'] = val
        if 'points' in selected_keys and "מס' נקודות" in cols:
            try:
                kwargs['points'] = int(float(row.get("מס' נקודות"))) if pd.notna(row.get("מס' נקודות")) else 0
            except Exception: pass
        if 'serial' in selected_keys and serial_col and serial_col in cols:
            try:
                sv = row.get(serial_col)
                if pd.notna(sv) and str(sv).strip().lower() != 'nan':
                    kwargs['serial_number'] = int(float(sv))
            except Exception: pass
        else:
            kwargs['serial_number'] = int(index) + 1
        if 'gender' in selected_keys and 'מגדר' in cols:
            g = str(row.get('מגדר', '') or '').strip()
            if g in ('בן', 'M', 'זכר'): kwargs['gender'] = 'M'
            elif g in ('בת', 'F', 'נקבה'): kwargs['gender'] = 'F'
        if 'hebrew_birthday' in selected_keys:
            if 'יום הולדת עברי - יום' in cols:
                try:
                    dv = row.get('יום הולדת עברי - יום')
                    if pd.notna(dv) and str(dv).strip().lower() != 'nan':
                        kwargs['hebrew_birth_day'] = int(float(dv))
                except Exception: pass
            if 'יום הולדת עברי - חודש' in cols:
                try:
                    mv = str(row.get('יום הולדת עברי - חודש') or '').strip()
                    if mv and mv.lower() != 'nan':
                        val = _heb_month_map.get(mv)
                        if val is None:
                            val = int(float(mv))
                        kwargs['hebrew_birth_month'] = val
                except Exception: pass
            if 'יום הולדת עברי - שנה' in cols:
                try:
                    yv = row.get('יום הולדת עברי - שנה')
                    if pd.notna(yv) and str(yv).strip().lower() != 'nan':
                        kwargs['hebrew_birth_year'] = int(float(yv))
                except Exception: pass
        new_id = self.db.add_student(**kwargs)
        if 'private_message' in selected_keys and 'הודעה פרטית' in cols:
            try:
                msg_val = row.get('הודעה פרטית')
                if pd.notna(msg_val) and str(msg_val).strip() and str(msg_val).strip().lower() != 'nan':
                    self.db.update_private_message(new_id, str(msg_val).strip())
            except Exception: pass

    def export_columns_only(self, excel_path: str) -> bool:
        """
        ייצוא חכם - עדכון רק עמודות G, H, I (כרטיס, נקודות, הודעה)
        לא דורס את שאר העמודות!
        """
        if not _require_pandas():
            return False
        try:
            from openpyxl import load_workbook
            import pandas as pd
            
            # קריאת האקסל הקיים
            df = pd.read_excel(excel_path, dtype={'מס\' כרטיס': str})
            
            # קבלת כל התלמידים מה-DB
            students = self.db.get_all_students()
            
            # עדכון רק עמודות G, H, I
            for index, row in df.iterrows():
                try:
                    last_name = str(row.get('שם משפחה', '')).strip() if pd.notna(row.get('שם משפחה')) else ''
                    first_name = str(row.get('שם פרטי', '')).strip() if pd.notna(row.get('שם פרטי')) else ''
                    
                    if not last_name or not first_name:
                        continue
                    
                    # חיפוש התלמיד ב-DB
                    matching = [s for s in students if s['last_name'] == last_name and s['first_name'] == first_name]
                    if not matching:
                        continue
                    
                    student = matching[0]
                    
                    # עדכון רק G, H, I
                    df.at[index, 'מס\' כרטיס'] = student['card_number'] if student['card_number'] else ''
                    df.at[index, 'מס\' נקודות'] = student['points']
                    df.at[index, 'הודעה פרטית'] = student.get('private_message', '') or ''
                    
                except Exception as e:
                    print(f"שגיאה בעדכון שורה {index}: {e}")
                    continue
            
            # שמירה חזרה
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            # פורמט עמודת כרטיס כטקסט
            wb = load_workbook(excel_path)
            ws = wb.active
            try:
                from excel_styling import apply_rtl_and_alternating_colors
                apply_rtl_and_alternating_colors(ws, has_header=True)
            except Exception:
                ws.sheet_view.rightToLeft = True
            
            # עמודת כרטיס כטקסט
            card_col = 7  # G
            for row_num in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=card_col)
                cell.number_format = '@'
            
            # הוספת מעברי עמוד לכל הגליונות
            for ws in wb.worksheets:
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToHeight = 0
                ws.page_setup.fitToWidth = 1

            wb.save(excel_path)
            return True
            
        except Exception as e:
            print(f"שגיאה בייצוא עמודות: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def export_to_excel(self, excel_path: str, allowed_classes: list = None) -> bool:
        """ייצוא נתונים לקובץ Excel עם פורמט טקסט למס' כרטיס וכיוון RTL.
        אם אין תלמידים במסד הנתונים – ייווצר קובץ שבלונה עם שורת כותרות בלבד.
        כולל גם מספר תיקופים וממוצע תיקופים לכל תלמיד (מבוסס swipe_log).
        """
        if not _require_pandas():
            return False
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            students = self.db.get_all_students()
            if allowed_classes:
                allowed = set([str(c).strip() for c in (allowed_classes or []) if str(c).strip()])
                if allowed:
                    students = [s for s in (students or []) if str(s.get('class_name') or '').strip() in allowed]

            # חישוב נתוני תיקופים לכל התלמידים
            student_ids = [s['id'] for s in students]
            swipe_totals = self.db.get_swipe_totals_for_students(student_ids, station_type="public")
            total_days = self.db.get_total_school_days(station_type="public")

            # עמודות לקובץ ה-Excel (ללא עמודת ת"ז)
            columns = [
                "מס' סידורי",
                'שם משפחה',
                'שם פרטי',
                'כיתה',
                'נתיב תמונה',
                "מס' כרטיס",
                "מס' נקודות",
                "מס' תיקופים",
                "ממוצע תיקופים",
                'הודעה פרטית',
                'מגדר',
                'יום הולדת עברי - יום',
                'יום הולדת עברי - חודש',
                'יום הולדת עברי - שנה',
            ]

            # המרה ל-DataFrame
            data = []
            for i, student in enumerate(students, 1):
                # שמירת מספר כרטיס כטקסט - בדיוק כמו שהוא!
                card_str = ''
                if student['card_number']:
                    card_str = str(student['card_number'])

                # מס' סידורי: קודם כל מהמסד, ואם חסר – לפי הסדר בקובץ
                serial_val = student.get('serial_number')
                serial_out = serial_val if serial_val not in (None, 0) else i

                # תיקופים
                total_swipes = int(swipe_totals.get(student['id'], 0) or 0)
                if total_days > 0 and total_swipes > 0:
                    avg_swipes = round(total_swipes / float(total_days), 2)
                else:
                    avg_swipes = 0

                # מגדר ויום הולדת עברי
                gender_val = str(student.get('gender') or '').strip()
                gender_display = {'M': 'בן', 'F': 'בת'}.get(gender_val, '')
                hb_day = student.get('hebrew_birth_day') or ''
                hb_month_num = student.get('hebrew_birth_month')
                hb_month_display = ''
                if hb_month_num:
                    _heb_m = {1:'ניסן',2:'אייר',3:'סיון',4:'תמוז',5:'אב',6:'אלול',7:'תשרי',8:'חשון',9:'כסלו',10:'טבת',11:'שבט',12:'אדר',13:"אדר ב'"}
                    hb_month_display = _heb_m.get(int(hb_month_num), str(hb_month_num))
                hb_year = student.get('hebrew_birth_year') or ''

                data.append({
                    "מס' סידורי": serial_out,
                    'שם משפחה': student['last_name'],
                    'שם פרטי': student['first_name'],
                    'כיתה': student['class_name'],
                    "נתיב תמונה": student['photo_number'],
                    "מס' כרטיס": card_str,
                    "מס' נקודות": student['points'],
                    "מס' תיקופים": total_swipes,
                    "ממוצע תיקופים": avg_swipes,
                    'הודעה פרטית': student.get('private_message', '') or '',
                    'מגדר': gender_display,
                    'יום הולדת עברי - יום': hb_day if hb_day else '',
                    'יום הולדת עברי - חודש': hb_month_display,
                    'יום הולדת עברי - שנה': hb_year if hb_year else '',
                })
            
            # גם אם data ריק – ניצור DataFrame עם העמודות כדי לקבל שורת כותרות בלבד
            if data:
                df = pd.DataFrame(data, columns=columns)
                # מיון לפי כיתה (stable sort — שומר סדר פנימי)
                try:
                    df = df.sort_values(by='כיתה', kind='mergesort', key=lambda x: x.astype(str).str.strip())
                except Exception:
                    pass
            else:
                df = pd.DataFrame(columns=columns)
            
            # שמירה ראשונית
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            # פתיחה מחדש לפורמט
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # פורמט עמודת כרטיס כטקסט - FORCE TEXT!
            # סדר העמודות: A סידורי, B שם משפחה, C שם פרטי, D כיתה, E נתיב תמונה,
            # F מס' כרטיס, G מס' נקודות, H מס' תיקופים, I ממוצע תיקופים, J הודעה פרטית
            card_col = 6  # עמודה F (מס' כרטיס)
            
            # הגדר את כותרת העמודה גם בטקסט
            header_cell = ws.cell(row=1, column=card_col)
            header_cell.number_format = '@'
            
            # הגדר כל תא בעמודת הכרטיס כטקסט (ללא הוספת גרש לתוכן)
            for row in range(2, ws.max_row + 1):  # מתחיל מ-2 (אחרי הכותרות)
                cell = ws.cell(row=row, column=card_col)
                if cell.value is not None:
                    cell.value = str(cell.value)
                cell.number_format = '@'
            
            # Apply RTL and alternating colors styling
            try:
                from excel_styling import apply_rtl_and_alternating_colors
                apply_rtl_and_alternating_colors(ws, has_header=True)
            except Exception:
                # Fallback to basic RTL if styling module not available
                ws.sheet_view.rightToLeft = True
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # מעברי עמוד בין כיתות + הקפאת כותרת
            try:
                from excel_styling import add_class_page_breaks_and_freeze
                add_class_page_breaks_and_freeze(ws, has_header=True)
            except Exception:
                pass
            
            # הגדרות הדפסה
            for _ws in wb.worksheets:
                _ws.page_setup.paperSize = _ws.PAPERSIZE_A4
                _ws.page_setup.orientation = _ws.ORIENTATION_LANDSCAPE
                _ws.page_setup.fitToPage = True
                _ws.page_setup.fitToHeight = 0
                _ws.page_setup.fitToWidth = 1

            wb.save(excel_path)
            
            return True
            
        except Exception as e:
            print(f"שגיאה בייצוא ל-Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_attendance_excel(self, excel_path: str, mode: str = "day", target_date_iso: str = None, bonus_id: int = None,
                               allowed_classes: list = None) -> bool:
        """ייצוא נוכחות מבוססת בונוס זמנים.

        mode:
            "day"         – נוכחות ליום מסוים (כל הבונוסים).
            "bonus"       – נוכחות לפי בונוס אחד, כל התאריכים (עמודה לכל תאריך).
            "all_bonuses" – כל הבונוסים וכל התאריכים, גליון נפרד לכל בונוס.
        """
        try:
            from datetime import date
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            students = self.db.get_all_students()
            if allowed_classes:
                allowed = set([str(c).strip() for c in (allowed_classes or []) if str(c).strip()])
                if allowed:
                    students = [s for s in (students or []) if str(s.get('class_name') or '').strip() in allowed]
            time_bonuses = self.db.get_all_time_bonuses()

            def _bonus_group_name(b: dict) -> str:
                return (b.get('group_name') or b.get('name') or '').strip()

            def _given_at_to_time_str(given_at_val: str) -> str:
                given_at = str(given_at_val or '')
                if len(given_at) < 16:
                    return ''
                try:
                    # given_at is historically stored as UTC (SQLite CURRENT_TIMESTAMP).
                    # Convert UTC -> local time (DST-aware) for display/export.
                    from datetime import timezone
                    dt_utc = datetime.strptime(given_at[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                    dt_local = dt_utc.astimezone()

                    return dt_local.strftime("%H.%M")
                except Exception:
                    try:
                        hhmm = given_at[11:16]
                        return hhmm.replace(':', '.')
                    except Exception:
                        return ''

            # כלי עזר לפורמט תאריך
            def _format_date(d_iso: str) -> str:
                if not d_iso or len(d_iso) < 10:
                    return d_iso
                y, m, d = d_iso[0:4], d_iso[5:7], d_iso[8:10]
                return f"{d}.{m}.{y}"

            def _style_and_setup_wb(wb_arg):
                """עיצוב RTL + מעברי עמוד בין כיתות + הקפאת כותרת + הגדרות הדפסה"""
                try:
                    from excel_styling import apply_rtl_and_alternating_colors, add_class_page_breaks_and_freeze
                    for _ws in wb_arg.worksheets:
                        apply_rtl_and_alternating_colors(_ws, has_header=True)
                        add_class_page_breaks_and_freeze(_ws, has_header=True)
                except Exception:
                    for _ws in wb_arg.worksheets:
                        _ws.sheet_view.rightToLeft = True
                for _ws in wb_arg.worksheets:
                    _ws.page_setup.paperSize = _ws.PAPERSIZE_A4
                    _ws.page_setup.orientation = _ws.ORIENTATION_LANDSCAPE
                    _ws.page_setup.fitToPage = True
                    _ws.page_setup.fitToHeight = 0
                    _ws.page_setup.fitToWidth = 1

            def _sort_by_class(df_arg):
                try:
                    return df_arg.sort_values(by='כיתה', kind='mergesort', key=lambda x: x.astype(str).str.strip())
                except Exception:
                    return df_arg

            # מצב 1: נוכחות ליום מסוים (כל הבונוסים)
            if mode == "day":
                if not target_date_iso:
                    target_date_iso = date.today().isoformat()

                bonuses = time_bonuses
                bonuses_by_id = {b['id']: b for b in bonuses}

                group_names = []
                group_name_by_id = {}
                for b in bonuses:
                    g = _bonus_group_name(b)
                    if not g:
                        continue
                    group_name_by_id[b['id']] = g
                    if g not in group_names:
                        group_names.append(g)

                attendance_rows = self.db.get_time_bonus_given_for_date(target_date_iso)

                # לכל תלמיד: נרשום שעה רק כאשר קיבל בפועל בונוס מהקבוצה.
                # כך נמנע ערבוב שעות מכיתות אחרות.
                by_student_group = {}
                for row in attendance_rows:
                    try:
                        sid = int(row.get('student_id'))
                        bid = int(row.get('bonus_schedule_id'))
                    except Exception:
                        continue
                    b = bonuses_by_id.get(bid)
                    if not b:
                        continue
                    g = group_name_by_id.get(bid) or _bonus_group_name(b)
                    if not g:
                        continue
                    t = _given_at_to_time_str(row.get('given_at'))
                    if not t:
                        continue
                    pts = int(b.get('bonus_points', 0) or 0)

                    cur = by_student_group.get(sid, {}).get(g)
                    if cur is None:
                        by_student_group.setdefault(sid, {})[g] = {'time': t, 'pts': pts}
                    else:
                        try:
                            # שעה מוקדמת יותר
                            if t < (cur.get('time') or t):
                                cur['time'] = t
                            # נקודות מקסימליות (לתצוגה)
                            if pts > int(cur.get('pts', 0) or 0):
                                cur['pts'] = pts
                        except Exception:
                            pass

                columns = [
                    "מס' סידורי",
                    'שם משפחה',
                    'שם פרטי',
                    'כיתה',
                ] + group_names

                data = []
                for i, student in enumerate(students, 1):
                    serial_val = student.get('serial_number')
                    serial_out = serial_val if serial_val not in (None, 0) else i
                    row_dict = {
                        "מס' סידורי": serial_out,
                        'שם משפחה': student['last_name'],
                        'שם פרטי': student['first_name'],
                        'כיתה': student['class_name'],
                    }

                    sid = int(student['id'])
                    for g in group_names:
                        info = by_student_group.get(sid, {}).get(g)
                        if not info:
                            row_dict[g] = ''
                            continue
                        t = str(info.get('time') or '').strip()
                        pts = int(info.get('pts', 0) or 0)
                        if not t:
                            row_dict[g] = ''
                        elif pts > 0:
                            row_dict[g] = f"{t} (+{pts})"
                        else:
                            row_dict[g] = t

                    data.append(row_dict)

                df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)
                df = _sort_by_class(df)
                df.to_excel(excel_path, index=False, engine='openpyxl')

                wb = load_workbook(excel_path)
                _style_and_setup_wb(wb)
                wb.save(excel_path)
                return True

            # מצב 2: לפי בונוס אחד – כל התאריכים, עמודה לכל תאריך
            if mode == "bonus":
                if not bonus_id:
                    return False

                chosen_bonus = next((b for b in time_bonuses if b['id'] == bonus_id), None)
                if not chosen_bonus:
                    return False

                chosen_group = _bonus_group_name(chosen_bonus)
                if not chosen_group:
                    chosen_group = str(chosen_bonus.get('name') or '').strip()

                group_bonus_ids = [int(b['id']) for b in time_bonuses if _bonus_group_name(b) == chosen_group]
                if not group_bonus_ids:
                    group_bonus_ids = [int(bonus_id)]

                rows = []
                for bid in group_bonus_ids:
                    try:
                        rows.extend(self.db.get_time_bonus_given_for_bonus(bid) or [])
                    except Exception:
                        pass

                attendance_map = {}
                dates_set = set()
                for row in rows:
                    sid = row['student_id']
                    d_iso = row['given_date']
                    dates_set.add(d_iso)
                    time_str = _given_at_to_time_str(row.get('given_at'))
                    if not time_str:
                        continue
                    prev = attendance_map.setdefault(sid, {}).get(d_iso, '')
                    if not prev or time_str < prev:
                        attendance_map[sid][d_iso] = time_str

                sorted_dates = sorted(dates_set)
                date_headers = [_format_date(d) for d in sorted_dates]

                columns = [
                    "מס' סידורי",
                    'שם משפחה',
                    'שם פרטי',
                    'כיתה',
                ] + date_headers

                data = []
                for i, student in enumerate(students, 1):
                    serial_val = student.get('serial_number')
                    serial_out = serial_val if serial_val not in (None, 0) else i
                    row_dict = {
                        "מס' סידורי": serial_out,
                        'שם משפחה': student['last_name'],
                        'שם פרטי': student['first_name'],
                        'כיתה': student['class_name'],
                    }

                    student_times = attendance_map.get(student['id'], {})
                    for d_iso, header in zip(sorted_dates, date_headers):
                        row_dict[header] = student_times.get(d_iso, '')

                    data.append(row_dict)

                df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)
                df = _sort_by_class(df)
                df.to_excel(excel_path, index=False, engine='openpyxl')

                wb = load_workbook(excel_path)
                ws = wb.active
                try:
                    sheet_name = str(chosen_group or chosen_bonus.get('name', 'נוכחות'))[:31]
                    ws.title = sheet_name
                except Exception:
                    pass
                _style_and_setup_wb(wb)
                wb.save(excel_path)
                return True

            # מצב 3: כל הבונוסים וכל התאריכים – גליון נפרד לכל בונוס
            if mode == "all_bonuses":
                # מיזוג לפי "קבוצה" כדי למנוע כפילויות ולוודא שמייצאים גם כאשר יש כמה שורות באותה קבוצה
                groups = {}
                for b in time_bonuses:
                    g = _bonus_group_name(b)
                    if not g:
                        continue
                    groups.setdefault(g, []).append(int(b['id']))

                wrote_any = False
                used_sheet_names = set()

                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    for group_name, bonus_ids in groups.items():
                        rows = []
                        for bid in bonus_ids:
                            try:
                                rows.extend(self.db.get_time_bonus_given_for_bonus(bid) or [])
                            except Exception:
                                pass
                        if not rows:
                            continue

                        attendance_map = {}
                        dates_set = set()
                        for row in rows:
                            sid = row.get('student_id')
                            d_iso = row.get('given_date')
                            if not sid or not d_iso:
                                continue
                            dates_set.add(d_iso)
                            time_str = _given_at_to_time_str(row.get('given_at'))
                            if not time_str:
                                continue
                            prev = attendance_map.setdefault(sid, {}).get(d_iso, '')
                            if not prev or time_str < prev:
                                attendance_map[sid][d_iso] = time_str

                        sorted_dates = sorted(dates_set)
                        date_headers = [_format_date(d) for d in sorted_dates]

                        columns = [
                            "מס' סידורי",
                            'שם משפחה',
                            'שם פרטי',
                            'כיתה',
                        ] + date_headers

                        data = []
                        for i, student in enumerate(students, 1):
                            serial_val = student.get('serial_number')
                            serial_out = serial_val if serial_val not in (None, 0) else i
                            row_dict = {
                                "מס' סידורי": serial_out,
                                'שם משפחה': student['last_name'],
                                'שם פרטי': student['first_name'],
                                'כיתה': student['class_name'],
                            }

                            student_times = attendance_map.get(student['id'], {})
                            for d_iso, header in zip(sorted_dates, date_headers):
                                row_dict[header] = student_times.get(d_iso, '')

                            data.append(row_dict)

                        df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)

                        base_name = str(group_name).strip() or 'נוכחות'
                        sheet_name = base_name[:31]
                        if sheet_name in used_sheet_names:
                            for n in range(2, 100):
                                suffix = f" ({n})"
                                cand = (base_name[: max(1, 31 - len(suffix))] + suffix)
                                if cand not in used_sheet_names:
                                    sheet_name = cand
                                    break
                        used_sheet_names.add(sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        wrote_any = True

                    if not wrote_any:
                        # אם אין נתונים בכלל – עדיין לייצר קובץ תקין עם כותרות בסיסיות
                        columns = ["מס' סידורי", 'שם משפחה', 'שם פרטי', 'כיתה']
                        pd.DataFrame([], columns=columns).to_excel(writer, sheet_name='נוכחות', index=False)

                wb = load_workbook(excel_path)
                _style_and_setup_wb(wb)
                wb.save(excel_path)
                return True

            # מצב לא מוכר
            return False

        except Exception as e:
            print(f"שגיאה בייצוא נוכחות ל-Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_daily_points_summary_excel(self, excel_path: str, *, allowed_classes: list = None) -> bool:
        if not _require_pandas():
            return False
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment

            rows, headers = self.db.get_daily_points_summary_matrix(allowed_classes=allowed_classes)
            base_cols = ["מס' סידורי", 'שם משפחה', 'שם פרטי', 'כיתה']
            cols = base_cols + list(headers or []) + ["סה\"כ נקודות"]

            try:
                df = pd.DataFrame(rows or [])
            except Exception:
                df = pd.DataFrame([])

            for c in cols:
                if c not in df.columns:
                    df[c] = ''
            df = df[cols]

            # מיון לפי כיתה
            try:
                df = df.sort_values(by='כיתה', kind='mergesort', key=lambda x: x.astype(str).str.strip())
            except Exception:
                pass

            df.to_excel(excel_path, index=False, engine='openpyxl')

            wb = load_workbook(excel_path)
            ws = wb.active
            try:
                ws.title = 'תשקיף יומי'[:31]
            except Exception:
                pass

            try:
                from excel_styling import apply_rtl_and_alternating_colors, add_class_page_breaks_and_freeze
                apply_rtl_and_alternating_colors(ws, has_header=True)
                add_class_page_breaks_and_freeze(ws, has_header=True)
            except Exception:
                ws.sheet_view.rightToLeft = True
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

            for _ws in wb.worksheets:
                _ws.page_setup.paperSize = _ws.PAPERSIZE_A4
                _ws.page_setup.orientation = _ws.ORIENTATION_LANDSCAPE
                _ws.page_setup.fitToPage = True
                _ws.page_setup.fitToHeight = 0
                _ws.page_setup.fitToWidth = 1

            wb.save(excel_path)
            return True
        except Exception as e:
            print(f"שגיאה בייצוא תשקיף יומי נקודות: {e}")
            import traceback
            traceback.print_exc()
            return False
